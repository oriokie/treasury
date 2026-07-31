"""Fixes from the Transactions page review:
- M-Pesa Reference column added to the Trust Fund Pending Receipts export.
- New quick-filter tabs (All / Needs review / Unallocated / Trust pending
  receipt) on the transactions page, layered on top of the existing,
  already-tested filter mechanism without changing it.
- Fixed a duplicate-button / over-restrictive-permission bug introduced while
  building the tabs: the trust-pending-receipt export is read-only and was
  already correctly available to any user with page access (Treasurer and
  Auditor alike) — caught before shipping by testing as an auditor, not just
  a treasurer."""
import datetime as dt
from decimal import Decimal
from django.test import TestCase, Client
from django.contrib.auth.models import User, Group
from departments.models import Department
from members.models import Member
from giving.models import Transaction


def _tr():
    u = User.objects.create_user("tr_txpage", password="x", is_superuser=True)
    u.groups.add(Group.objects.get_or_create(name="Treasurer")[0])
    return u


def _auditor():
    u = User.objects.create_user("au_txpage", password="x")
    u.groups.add(Group.objects.get_or_create(name="Auditor")[0])
    return u


class TrustPendingReceiptMpesaColumnTests(TestCase):
    def setUp(self):
        self.tr = _tr()
        self.trust = Department.objects.create(name="TxPageTrust", fund_type="TRUST",
            category="OFFERING")
        self.member = Member.objects.create(name="Tx Page Giver", phone="254711222333")
        self.c = Client(); self.c.force_login(self.tr)

    def _col(self, name):
        from giving.services.pending_receipt import HEADER
        return HEADER.index(name)

    def _rows(self, response):
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content))
        return list(wb.active.iter_rows(values_only=True))

    def test_header_includes_mpesa_reference(self):
        rows = self._rows(self.c.get("/transactions/?export=trust-pending-receipt"))
        header = next(r for r in rows if r and r[0] == "Date")
        from giving.services.pending_receipt import HEADER
        self.assertEqual(list(header), HEADER)
        self.assertIn("M-Pesa Reference", header)

    def test_mpesa_reference_populated(self):
        Transaction.objects.create(date=dt.date(2026, 6, 10), amount=Decimal("1500"),
            direction="CREDIT", confirmed=True, channel="BANK",
            allocation_status="MANUAL", department=self.trust, member=self.member,
            payer_phone="254711222333", reference="txpage-ref", mpesa_ref="QZZ1234AB")
        rows = self._rows(self.c.get("/transactions/?export=trust-pending-receipt"))
        row = next(r for r in rows if r and r[self._col("Reference")] == "txpage-ref")
        self.assertEqual(row[self._col("M-Pesa Reference")], "QZZ1234AB")

    def test_a_cash_gift_is_NOT_pending_receipt(self):
        """This used to assert the opposite — that a cash gift appears here with
        a blank M-Pesa reference. That was the bug: cash is receipted at the
        point of counting (it goes onto an envelope at the table), so it never
        arrives silently and waits to be chased. Listing it asked a treasurer to
        chase a receipt for money that was never going to have one."""
        Transaction.objects.create(date=dt.date(2026, 6, 11), amount=Decimal("500"),
            direction="CREDIT", confirmed=True, channel="CASH",
            allocation_status="MANUAL", department=self.trust, reference="cash-ref")
        rows = self._rows(self.c.get("/transactions/?export=pending-receipt"))
        self.assertFalse([r for r in rows if r and r[self._col("Reference")] == "cash-ref"],
                         "a cash gift should not be listed as pending receipt")

    def test_mpesa_reference_blank_when_the_bank_gift_has_none(self):
        """The case this test was really written for: a BANK credit that carries
        no M-Pesa reference (a direct transfer, say) still belongs on the list —
        with that column simply empty."""
        Transaction.objects.create(date=dt.date(2026, 6, 12), amount=Decimal("900"),
            direction="CREDIT", confirmed=True, channel="BANK",
            allocation_status="MANUAL", department=self.trust, reference="eft-ref")
        rows = self._rows(self.c.get("/transactions/?export=pending-receipt"))
        row = next(r for r in rows if r and r[self._col("Reference")] == "eft-ref")
        self.assertIn(row[self._col("M-Pesa Reference")], ("", None))


class TransactionsPageQuickTabsTests(TestCase):
    def setUp(self):
        self.tr = _tr()
        self.au = _auditor()

    def test_tabs_render_for_treasurer(self):
        c = Client(); c.force_login(self.tr)
        b = c.get("/transactions/").content.decode()
        self.assertIn("tx-quicktabs", b)
        self.assertIn("Needs review", b)
        self.assertIn("Unallocated", b)

    def test_trust_pending_link_appears_exactly_once(self):
        c = Client(); c.force_login(self.tr)
        b = c.get("/transactions/").content.decode()
        self.assertEqual(b.count("Pending receipt"), 1)

    def test_trust_pending_link_visible_to_auditor_too(self):
        # read-only export — an auditor legitimately needs this for oversight,
        # same access level as the page itself (ReadAccessMixin)
        c = Client(); c.force_login(self.au)
        b = c.get("/transactions/").content.decode()
        self.assertEqual(b.count("Pending receipt"), 1)
        r = c.get("/transactions/?export=trust-pending-receipt")
        self.assertEqual(r.status_code, 200)

    def test_needs_review_tab_filters_correctly(self):
        c = Client(); c.force_login(self.tr)
        b = c.get("/transactions/?status=REVIEW").content.decode()
        self.assertIn('tx-qtab active">Needs review', b)

    def test_cash_only_page_hides_trust_pending_tab(self):
        c = Client(); c.force_login(self.tr)
        b = c.get("/cash/").content.decode()
        self.assertNotIn("Pending receipt", b)
        self.assertIn("tx-quicktabs", b)

    def test_existing_filter_form_still_works(self):
        d = Department.objects.create(name="TabFilterFund", fund_type="LOCAL",
            category="MINISTRY")
        Transaction.objects.create(date=dt.date(2026, 6, 12), amount=Decimal("777"),
            direction="CREDIT", confirmed=True, channel="CASH",
            allocation_status="MANUAL", department=d, reference="findme777")
        c = Client(); c.force_login(self.tr)
        b = c.get("/transactions/?q=findme777").content.decode()
        self.assertIn("findme777", b)
