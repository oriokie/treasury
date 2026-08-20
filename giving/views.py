from decimal import Decimal, InvalidOperation
from django.db import transaction as db_tx

def _cash_duplicate(d, dept, amount, name=None):
    """True if a near-identical manual cash entry already exists. A duplicate must
    match the fund and amount (within +/-1 day) AND have a similar payer name
    (>50% match) — matching on amount alone wrongly flags different givers who
    happen to give the same amount on the same day."""
    if not (d and dept and amount):
        return False
    import datetime as _dt
    from difflib import SequenceMatcher
    from members.models import name_key
    from giving.models import Transaction
    cands = Transaction.objects.filter(
        channel=Transaction.Channel.CASH, department=dept, amount=amount,
        date__range=(d - _dt.timedelta(days=1), d + _dt.timedelta(days=1)))
    new_key = name_key(name or "")
    if not new_key:
        return False        # no name to compare -> don't guess it's a duplicate
    for t in cands:
        other = name_key(t.payer_name or "")
        if not other:
            continue
        if SequenceMatcher(None, new_key, other).ratio() > 0.5:
            return True
    return False


from core.utils import block_if_locked as _block_if_locked, PrefPaginationMixin
from django.contrib import messages
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import ListView, CreateView, View
from django import forms

from core.permissions import (DataEntryRequiredMixin, ReadAccessMixin, TreasurerRequiredMixin,
                              AllocateRequiredMixin, DebitClassifyRequiredMixin)
from core.utils import sabbath_week_of
from departments.models import Department
from .models import Transaction, AllocationRule
from .forms import CashEntryForm, QueueResolveForm, RuleForm
from .services.allocation import normalize_reference


def _group_split_siblings(transactions):
    """Group a list of Transaction rows so that siblings from the same
    original split contribution (Transaction.split_into — one lump sum
    divided across several funds/groups) are combined back into a single
    logical entry, the way it looked before the split. Returns an ordered
    list of groups (each a list of 1+ Transaction rows); a row with no
    siblings among those passed in comes back as its own group of one.

    Grouping happens only among the rows actually passed in — e.g. if an
    export is filtered to a single fund, only the sibling(s) that fund's
    filter matches are grouped; a sibling excluded by the filter is not
    silently pulled back in, so the export always aggregates exactly what's
    on screen, nothing more.

    What makes two rows siblings is deliberately confined to identifiers that
    CANNOT be shared by two different gifts: the `split_of` link the split
    itself recorded (resolved to the ROOT of the chain, so a split of a split
    is still one gift — see _root), a bank-assigned unique identifier (the
    core_ref base, or the M-Pesa reference for that day), or the issued
    receipt number an envelope's fund lines share, which is admitted on that
    one channel only and for the reasons set out below. A row with none of
    those groups with nothing but itself.

    That last rule is the point of this function's current shape, and it was
    learned the hard way. It used to end in a "same reference text + same date
    + same direction" fallback, which for a manually-entered CASH gift — no
    core_ref, no mpesa_ref, and a reference field that is optional on
    CashEntryForm and routinely left blank for a walk-in giver — meant any two
    unrelated cash gifts recorded on the same day collapsed into ONE exported
    row: the amounts summed, the funds joined into "Fund1 + Fund2", and only
    the FIRST giver's name printed. Alice's 100 and Bob's 250 left the building
    as a single 350 from ALICE, and Bob was not in the export at all. It is the
    same false positive Transaction.strict_split_siblings() already refuses to
    make ("a plain reference like tithe or offering is payer-entered free text,
    not a unique identifier"), and the same reason EnvelopeReceiptOneBankView
    keys its sibling gather on `core_ref base or mpesa_ref or str(id)`. Wrongly
    showing one gift as two rows is visible and merely untidy; wrongly merging
    two givers is invisible and loses a name — so the tie breaks toward not
    grouping.

    The one place a reference alone still groups is the ENVELOPE channel, and
    that is deliberate. `_save_envelope` posts one ENVELOPE-channel row per
    fund line of a receipt: no `split_of` link (nothing was split_into — each
    line was entered as its own row), no core_ref, no mpesa_ref, no bank
    receipt, and the only thing the lines of one receipt share is the
    reference it stamps on every one of them, "envelope R001". Refuse that and
    a two-fund envelope exports as two half-rows. It is safe here where free
    text is not because the text is not free: it is Envelope.receipt_no, which
    is unique=True, so no two envelopes can present the same reference and be
    mistaken for one receipt. The branch is also gated on channel == ENVELOPE,
    so no cash or bank row reaches it however its payer worded the reference,
    and it sits LAST before "solo", so a row that does carry a real identifier
    is already grouped by that instead — the legacy import's ENVELOPE-channel
    bank rows all share the constant reference "Processed via envelope", and
    they stay apart only because each one also carries its own core_ref.
    The channel alone cannot prove a row came from an envelope — CashEntryForm
    offers ENVELOPE as a channel with a free-text reference, so two hand-keyed
    rows both saying "tithe" would have combined, the original defect surviving
    on a narrower path. So the branch requires the reference to be one the
    system ISSUED, in `_save_envelope`'s own "envelope <receipt_no>" form,
    rather than any text that happens to sit on an envelope-channel row. A
    hand-keyed reference falls through to "solo" and exports as its own row."""
    # split_into() stamps `split_of` on the CHILDREN only — the original keeps
    # its own id and points at nothing — so a parent is recognised by being
    # pointed AT. Gathered once from the rows in hand instead of asking the
    # database per row (this runs over a whole export), which costs nothing in
    # correctness: grouping is confined to the rows passed in anyway, so a
    # parent that the filter excluded has no sibling here to join.
    parent_of = {t.pk: t.split_of_id for t in transactions if t.split_of_id}
    parent_ids = set(parent_of.values())

    def _root(pk):
        """Walk to the TRUE root of a split chain, not merely the immediate
        parent. A part of a split can be split again — split_into() sets
        `split_of` to whatever row it was called on, so X -> X-S1 ->
        X-S1-S1 is an ordinary two-level chain, not corruption — and keying
        on the immediate parent gave the grandchild a key of its own: ONE
        gift left the export as two rows with its amount torn between them,
        the fund label on each naming only part of where the money went.
        Nested splits are the whole reason the key is resolved rather than
        read off the row.

        The `seen` set guards against corrupt data only. `split_of` is a
        nullable self-FK and nothing in the database forbids a cycle; an
        unguarded walk would spin forever, and this function runs on the
        transactions page and the pending-receipt export as well as here, so
        one bad row would hang the whole surface instead of merely
        mis-grouping. A cycle terminates at the row we re-enter, which
        leaves its members in separate groups — the same direction every
        other tie here breaks: showing one gift as two rows is untidy and
        visible, merging two gifts is invisible and loses a name.
        """
        seen = set()
        while pk in parent_of and pk not in seen:
            seen.add(pk)
            pk = parent_of[pk]
        return pk

    def _key(t):
        # a reversal (or a reversed original) is a correction entry, never a
        # split sibling — it must never be silently combined with anything
        # else, even if it happens to share a core_ref/mpesa_ref/reference
        # with some unrelated row, so it always gets its own unique key
        if t.is_reversed or t.is_reversal:
            return ("solo", t.pk)
        # the `split_of` link FIRST, in the same order and for the same reason
        # as Transaction.split_siblings(): two rows are siblings because the
        # database says so, not because they look alike. It is also the only
        # thing that can recognise a split of a CASH entry, which has no
        # bank-assigned identifier of any kind to fall back on — so dropping
        # the loose reference match costs genuine cash splits nothing.
        base = t.split_of_id or (t.pk if t.pk in parent_ids else None)
        if base:
            # keyed on the ROOT of the chain so every generation of a nested
            # split lands in one group. A chain whose middle row the filter
            # excluded stops walking at that absent row rather than reaching
            # the true root, so the surviving ends group separately — the
            # same rule as everywhere else in here: only the rows actually
            # passed in may be joined, never one the filter took out.
            return ("split", _root(base))
        if t.core_ref:
            return ("core", t.core_ref.split("-S")[0], t.direction)
        if t.mpesa_ref:
            return ("mpesa", t.mpesa_ref, t.date, t.direction)
        ref = (t.reference or "").strip().lower()
        # The reference must be one the SYSTEM issued, not merely one that
        # arrived on an envelope-channel row. `_save_envelope` stamps exactly
        # "envelope <receipt_no>" on every line of a receipt, and receipt_no is
        # unique — that is the whole basis for trusting this text where free
        # text is never trusted. The channel alone does not carry that promise:
        # CashEntryForm offers ENVELOPE as a channel with a free-text
        # reference, so two hand-keyed rows both saying "tithe" would have
        # merged into one exported row and dropped a giver's name, which is the
        # very defect this function was fixed for, surviving on a narrower path.
        if (t.channel == Transaction.Channel.ENVELOPE
                and ref.startswith("envelope ")):
            return ("envelope", ref, t.direction)
        return ("solo", t.pk)

    groups, order = {}, []
    for t in transactions:
        key = _key(t)
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(t)
    # the first (lowest-id) sibling is always the representative row — it's
    # the original entry split_into() keeps as `self`, regardless of what
    # order the underlying queryset happens to list rows in (e.g. a "most
    # recent first" display ordering)
    return [sorted(groups[k], key=lambda t: t.id) for k in order]


def _combined_fund_label(members):
    """The fund label for a (possibly combined) group: the single department's
    name if there's only one, otherwise the split fund that caused the
    original split (found via the allocation rule for the group's shared
    reference), falling back to a plain join of the distinct fund names."""
    from .models import AllocationRule
    dept_ids = {t.department_id for t in members if t.department_id}
    if len(dept_ids) <= 1:
        first = members[0]
        if first.department_id:
            return first.department.name
        return ("Excluded (via envelope)" if first.excluded_from_income
                else "Unallocated")
    ref = (members[0].reference or "").strip().lower()
    if ref:
        rule = (AllocationRule.objects.filter(reference=ref, split_fund__isnull=False)
                .select_related("split_fund").first())
        if rule:
            return rule.split_fund.name
    names = sorted({t.department.name for t in members if t.department_id})
    return " + ".join(names)


class TransactionListView(PrefPaginationMixin, ReadAccessMixin, ListView):
    model = Transaction
    template_name = "giving/transaction_list.html"
    context_object_name = "transactions"
    paginate_by = 50

    def get(self, request, *args, **kwargs):
        export = request.GET.get("export")
        # The old "trust-" keys still work: a bookmark, and the Telegram bot's
        # /pending route, both point at them. Renaming a URL a user has saved is
        # not a rename, it is a breakage.
        if export in ("pending-receipt", "trust-pending-receipt"):
            return self._pending_receipt_export(request)
        if export in ("pending-receipt-pdf", "trust-pending-receipt-pdf"):
            return self._pending_receipt_pdf(request)
        if export in ("csv", "xlsx"):
            from reports.exports import csv_response, xlsx_response
            from core.models import SiteConfig
            qs = self.get_queryset()
            header = ["Date", "Sabbath", "Channel", "Direction", "Payer", "Member",
                      "Phone", "Fund", "Dev group", "Reference", "M-Pesa ref",
                      "Core ref", "Bank receipt", "Receipt status", "Status",
                      "Confirmed", "Entry status", "Amount"]

            def _receipt_status(t):
                if t.is_bank_memo:
                    return "Memo — receipted manually (cash on envelope entry)"
                if t.processed_via_envelope or t.channel == Transaction.Channel.ENVELOPE:
                    return "Receipted (envelope)"
                if t.manual_receipt:
                    return "Receipted (manual)"
                if t.excluded_from_income:
                    return "Non-income (loan / financing)"
                return "Not receipted"

            def _dev_group_label(members):
                labels = {t.dev_group.label for t in members if t.dev_group_id}
                return " + ".join(sorted(labels)) if labels else ""

            def _entry_status(t):
                if t.is_reversal:
                    return "Reversal"
                if t.is_reversed:
                    return "Reversed"
                return ""

            rows = []
            for members in _group_split_siblings(list(qs)):
                first = members[0]
                # The Amount column is the row's TRUE effect on cash, signed
                # the one canonical way (Transaction.signed_cash_amount): a
                # reversal nets its original to zero, and a manually-receipted
                # bank memo contributes zero (its cash lives on the envelope
                # entry — the Receipt status column says so) — so summing the
                # column equals reality instead of double-counting reversed or
                # manually-receipted money.
                total = sum((t.signed_cash_amount for t in members), Decimal(0))
                rows.append([first.date.isoformat(),
                     first.service_sabbath.isoformat() if first.service_sabbath else "",
                     first.get_channel_display(), first.get_direction_display(),
                     first.payer_name or (first.member.name if first.member else ""),
                     first.member.name if first.member_id else "",
                     first.payer_phone or (first.member.receipt_phone if first.member_id else "") or "",
                     _combined_fund_label(members),
                     _dev_group_label(members),
                     first.reference or "", first.mpesa_ref or "", first.core_ref or "",
                     first.bank_receipt or "", _receipt_status(first),
                     first.get_allocation_status_display(),
                     "Yes" if first.confirmed else "",
                     _entry_status(first),
                     float(total)])   # signed_cash_amount already carries direction
            if export == "xlsx":
                return xlsx_response("transactions.xlsx", header, rows,
                                     title="Transactions", church=SiteConfig.get().church_name)
            return csv_response("transactions.csv", header, rows)
        return super().get(request, *args, **kwargs)

    def _pending_receipt_export(self, request):
        """Credits in a RECEIPTABLE fund not yet formally receipted — Date,
        Phone, Member, Amount, Reference, M-Pesa Reference.

        "Receiptable" is Trust funds AND the Local Church Budget family (the LCB
        funds configured in Settings, plus their subgroups) — the same "Trust +
        LCB" the Sabbath-confirm scope names, and now the same code behind it
        (`departments.models.receiptable_fund_ids`). This list was Trust-only,
        so LCB money a church receipts exactly as it receipts trust money never
        appeared — which is why it was called "Trust pending receipt", a name
        that described the bug rather than the intent.

        A split contribution (one lump sum divided across several funds, e.g.
        Combined Offering = 50% trust + 50% local) is posted as several ledger
        rows sharing a payment reference; here they're recombined into one row
        with the FULL original amount and the split fund's own name. A giver who
        contributed 40 to "Combined Offering" gets one receipt for 40, not one
        for the 20 that happened to land in a trust account. A group qualifies
        whenever ANY of its parts landed in a receiptable fund — the whole gift
        is one receipt, so receipting half of it is not a thing.

        Sorted by name and a repeated name highlighted — same as the on-page
        view and the PDF (including the one the Telegram bot sends): one
        function decides the order and the duplicates, so none of these three
        can quietly disagree.

        The Fund column is deliberately absent from the downloads (it remains
        on the page, where you can also sort by it): this sheet is worked
        through name by name to issue receipts, and the fund is settled by the
        receipt itself."""
        from reports.exports import xlsx_response
        from core.models import SiteConfig
        from giving.services.pending_receipt import (HEADER, duplicate_name_flags,
                                                     export_rows,
                                                     pending_receipt_rows)

        pr_rows = pending_receipt_rows()
        dup_flags = duplicate_name_flags(pr_rows)
        # A repeated name is shown by the row highlight alone \u2014 no "repeats"
        # label appended to the name, which also kept the Member column clean
        # for anyone sorting or matching on it.
        rows = [[d.isoformat(), phone, name, float(amount), ref, mpesa]
               for (d, phone, name, amount, ref, mpesa) in export_rows(pr_rows)]
        return xlsx_response("pending_receipt.xlsx", HEADER, rows,
                             title="Items pending receipt",
                             church=SiteConfig.get().church_name,
                             row_highlight=dup_flags)

    def _pending_receipt_pdf(self, request):
        """The same data as a PDF — for printing, or for a copy someone
        wants outside a spreadsheet (e.g. via the Telegram bot's /pending
        pdf route)."""
        from django.http import HttpResponse
        from core.models import SiteConfig
        from giving.services.pending_receipt import pending_receipt_pdf_bytes
        pdf = pending_receipt_pdf_bytes(church=SiteConfig.get().church_name)
        resp = HttpResponse(pdf, content_type="application/pdf")
        resp["Content-Disposition"] = 'attachment; filename="pending_receipt.pdf"'
        return resp

    def get_queryset(self):
        order = ("date", "id") if self.request.GET.get("sort") == "oldest" else ("-date", "-id")
        qs = (Transaction.objects.select_related("department", "member", "dev_group",
                                                  "bank_account")
              .order_by(*order))
        q = self.request.GET.get("q")
        channel = self.request.GET.get("channel")
        status = self.request.GET.get("status")
        dept = self.request.GET.get("department")
        if q:
            cond = (Q(payer_name__icontains=q) | Q(reference__icontains=q) |
                    Q(core_ref__icontains=q) | Q(raw_narration__icontains=q) |
                    Q(mpesa_ref__icontains=q) | Q(bank_receipt__icontains=q))
            # let the same box find an entry by its amount, e.g. "250" or "1,250.50"
            try:
                qstr = q.replace(",", "").strip()
                amt = Decimal(qstr)
                if "." in qstr:
                    cond |= Q(amount=amt)            # decimals: exact
                else:
                    cond |= Q(amount__gte=amt) & Q(amount__lt=amt + 1)  # "1234" finds 1234.x
            except (InvalidOperation, AttributeError):
                pass
            qs = qs.filter(cond)
        if channel:
            qs = qs.filter(channel=channel)
        if status:
            qs = qs.filter(allocation_status=status)
        if dept == "none":
            qs = qs.filter(department__isnull=True)
        elif dept:
            qs = qs.filter(department_id=dept)
        date_from = self.request.GET.get("date_from")
        date_to = self.request.GET.get("date_to")
        # parse defensively: an invalid date string would otherwise raise and
        # break the page; ignore anything that isn't a real YYYY-MM-DD date.
        from core.utils import default_to_current_month
        df, dtv = default_to_current_month(self.request)
        if df:
            qs = qs.filter(date__gte=df)
        if dtv:
            qs = qs.filter(date__lte=dtv)

        direction = self.request.GET.get("direction")
        if direction:
            qs = qs.filter(direction=direction)
        amt_min = self.request.GET.get("amount_min")
        amt_max = self.request.GET.get("amount_max")
        try:
            if amt_min not in (None, ""):
                qs = qs.filter(amount__gte=Decimal(amt_min))
        except InvalidOperation:
            pass
        try:
            if amt_max not in (None, ""):
                qs = qs.filter(amount__lte=Decimal(amt_max))
        except InvalidOperation:
            pass
        member_q = self.request.GET.get("member")
        if member_q:
            qs = qs.filter(Q(member__name__icontains=member_q) |
                          Q(payer_name__icontains=member_q))
        bank_account = self.request.GET.get("bank_account")
        if bank_account:
            qs = qs.filter(bank_account_id=bank_account)
        imported_by = self.request.GET.get("imported_by")
        if imported_by:
            qs = qs.filter(statement_import__uploaded_by_id=imported_by)
        if self.request.GET.get("reversed_only"):
            qs = qs.filter(Q(is_reversed=True) | Q(is_reversal=True))
        if self.request.GET.get("receipted_only"):
            qs = qs.filter(Q(processed_via_envelope=True) | Q(manual_receipt=True))
        if self.request.GET.get("manual_receipt_only"):
            qs = qs.filter(manual_receipt=True)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["departments"] = Department.objects.filter(active=True, selectable=True)
        ctx["channels"] = Transaction.Channel.choices
        ctx["statuses"] = Transaction.Status.choices
        ctx["directions"] = Transaction.Direction.choices
        # The date inputs must show what's ACTUALLY being filtered on, including
        # the current-month default applied on a bare visit (see
        # default_to_current_month) — not stay blank while secretly filtering,
        # which would look like "all data" is showing when it isn't.
        from core.utils import default_to_current_month
        filters = self.request.GET.copy()
        df, dtv = default_to_current_month(self.request)
        filters["date_from"] = df.isoformat() if df else ""
        filters["date_to"] = dtv.isoformat() if dtv else ""
        ctx["filters"] = filters
        ctx["date_default_applied"] = not self.request.GET
        ctx["unallocated_count"] = Transaction.objects.active().filter(
            department__isnull=True, excluded_from_income=False).count()
        ctx["noflund_total"] = Transaction.objects.active().filter(
            department__isnull=True).count()
        from statements.models import BankAccount
        from django.contrib.auth.models import User
        ctx["bank_accounts"] = BankAccount.objects.filter(active=True).order_by("name")
        ctx["importers"] = (User.objects.filter(statementimport__isnull=False)
                           .distinct().order_by("username"))
        # summary stats for the filtered set (header cards)
        from django.db.models import Sum, Count, Q as _Q
        qs = self.get_queryset()
        agg = qs.aggregate(
            credits=Sum("amount", filter=_Q(direction="CREDIT")),
            debits=Sum("amount", filter=_Q(direction="DEBIT")),
            n=Count("id"),
            review=Count("id", filter=_Q(allocation_status="REVIEW")),
            # Held money is not "in review" — the rules allocated it
            # confidently and the confirmation setting kept it back. It was
            # counted in neither card, so it sat here looking settled.
            held=Count("id", filter=_Q(confirmed=False)),
            held_total=Sum("amount", filter=_Q(confirmed=False)))
        ctx["sum_credits"] = agg["credits"] or 0
        ctx["sum_debits"] = agg["debits"] or 0
        ctx["sum_net"] = (agg["credits"] or 0) - (agg["debits"] or 0)
        ctx["sum_count"] = agg["n"] or 0
        ctx["sum_review"] = agg["review"] or 0
        ctx["sum_held"] = agg["held"] or 0
        ctx["sum_held_total"] = agg["held_total"] or 0
        ctx["has_filters"] = any(self.request.GET.get(k) for k in
                                 ("q", "channel", "status", "department",
                                  "date_from", "date_to", "direction", "amount_min",
                                  "amount_max", "member", "bank_account", "imported_by",
                                  "reversed_only", "receipted_only", "manual_receipt_only"))
        ctx["sort"] = self.request.GET.get("sort") or "newest"
        ctx["running_balances"] = self._running_balances(ctx["transactions"], qs)
        return ctx

    def _running_balances(self, page_items, filtered_qs):
        """Balance after each transaction on the current page, computed
        chronologically (oldest to newest) and scoped to whatever filters
        are currently applied — so filtering to one fund shows that fund's
        own running balance, not the whole church's.

        Every row is signed by the CANONICAL Transaction.signed_cash_case /
        signed_cash_amount definition: a reversal is subtracted (an offsetting
        entry, not new income), and a manually-receipted bank row contributes
        ZERO — its cash lives on its envelope counterpart, and counting both
        halves of that pair was inflating the running balance by every
        manually-receipted amount (the memo row stays visible, badged, so the
        audit trail is intact).

        Only ever queries the current page's rows plus one aggregate for
        "everything before this page" — never the full, unbounded history —
        so this stays cheap regardless of how many transactions exist."""
        page_list = list(page_items)
        if not page_list:
            return {}
        chrono = sorted(page_list, key=lambda t: (t.date, t.id))
        earliest = chrono[0]

        opening = filtered_qs.filter(
            Q(date__lt=earliest.date)
            | Q(date=earliest.date, id__lt=earliest.id)).signed_cash_total()

        balances = {}
        running = opening
        for t in chrono:
            running += t.signed_cash_amount
            balances[t.id] = running
        return balances


class PendingReceiptView(ReadAccessMixin, View):
    """Credits pending receipt, on screen — not just as a download.

    Same data, same order and the same idea of "duplicate" as the Excel/PDF
    downloads and the PDF the Telegram bot sends (`pending_receipt_rows` /
    `duplicate_name_flags` in giving.services.pending_receipt — one function
    each, read by all four surfaces, so none of them can quietly disagree).
    Sorted by name by default so the same giver's entries sit together, with
    any name that appears more than once highlighted — someone paying twice
    for the same thing, or a name recorded two slightly different ways, is
    exactly the kind of thing that's obvious at a glance and easy to miss
    reading top to bottom by date. "Same name" is order-insensitive
    (`members.models.name_key`), the same test the system already uses for
    member matching.
    """
    template_name = "giving/pending_receipt.html"

    def get(self, request):
        from members.models import name_key
        from giving.services.pending_receipt import (duplicate_name_flags,
                                                     pending_receipt_rows)

        raw_rows = pending_receipt_rows()          # already name-sorted
        dup_flags = duplicate_name_flags(raw_rows)
        sort = request.GET.get("sort") or "name"

        rows = []
        for (date, phone, name, amount, fund, reference, mpesa_ref), is_dup \
                in zip(raw_rows, dup_flags):
            rows.append({
                "date": date, "phone": phone, "name": name, "amount": amount,
                "fund": fund, "reference": reference, "mpesa_ref": mpesa_ref,
                "name_key": name_key(name or ""), "is_duplicate_name": is_dup,
                # What a reader sorts by is the name in front of them, not the
                # order-insensitive key used to match two spellings together.
                "sort_name": (name or "").strip().upper() or "~",
            })

        if sort == "date":
            rows.sort(key=lambda r: (r["date"], r["sort_name"]))
        elif sort == "amount":
            rows.sort(key=lambda r: -r["amount"])
        elif sort == "fund":
            rows.sort(key=lambda r: (r["fund"] or "", r["sort_name"]))
        else:
            sort = "name"   # rows are already in this order; re-sort is a no-op,
                            # kept explicit so a stable sort always holds even if
                            # pending_receipt_rows' own order ever changes

        # When sorted by name the same giver's rows sit together, so mark where
        # each new giver starts and how many rows/how much they account for.
        # A flat list of 80+ highlighted rows is hard to work through; reading
        # it as one block per giver is what a treasurer is actually doing.
        if sort == "name":
            for i, r in enumerate(rows):
                prev = rows[i - 1]["name_key"] if i else None
                r["group_start"] = r["name_key"] != prev
            for r in rows:
                if r["group_start"]:
                    same = [x for x in rows if x["name_key"] == r["name_key"]]
                    r["group_count"] = len(same)
                    r["group_total"] = sum((x["amount"] for x in same), Decimal(0))
        else:
            for r in rows:
                r["group_start"] = False

        total = sum((r["amount"] for r in rows), Decimal(0))
        duplicate_names = len({r["name_key"] for r in rows if r["is_duplicate_name"]})

        return render(request, self.template_name, {
            "rows": rows, "sort": sort, "total": total,
            "count": len(rows), "duplicate_names": duplicate_names,
            "distinct_givers": len({r["name_key"] for r in rows if r["name_key"]}),
        })


class ReviewQueueView(ReadAccessMixin, ListView):
    template_name = "giving/queue.html"
    context_object_name = "items"
    paginate_by = 25

    def get_queryset(self):
        # Only credits (giving) awaiting allocation belong here; bank debits with a
        # REVIEW status are handled in the separate bank-debit queue. Anything
        # already receipted (system envelope) or marked as a manual paper receipt
        # is handled and must not appear here.
        return (Transaction.objects.filter(
                    allocation_status=Transaction.Status.REVIEW,
                    direction=Transaction.Direction.CREDIT,
                    processed_via_envelope=False, manual_receipt=False)
                .select_related("member").order_by("date", "id"))

    def get_context_data(self, **kwargs):
        from departments.models import DevelopmentGroup
        from giving.models import SplitFund
        ctx = super().get_context_data(**kwargs)
        ctx["departments"] = Department.objects.filter(active=True, selectable=True)
        ctx["split_funds"] = SplitFund.objects.filter(active=True).order_by("name")
        ctx["dev_groups"] = DevelopmentGroup.objects.filter(active=True).order_by("number")
        # item 5: gifts in the ledger that still need a fund but aren't in the queue
        ctx["unallocated_in_ledger"] = FetchUnallocatedView.pending_qs().count()
        return ctx


class RunRulesOnQueueView(AllocateRequiredMixin, View):
    """Re-run allocation rules over the items still in the review queue, so rules
    added after an import can clear matching items without re-importing the file."""

    def post(self, request):
        from giving.services.allocation import reallocate_pending
        r = reallocate_pending()
        if r["scanned"] == 0:
            messages.info(request, "The review queue is empty — nothing to allocate.")
        elif r["allocated"]:
            msg = (f"Allocated {r['allocated']} of {r['scanned']} item(s) using the "
                   f"current rules. {r['remaining']} still need attention.")
            extra = []
            if r["skipped_locked"]:
                extra.append(f"{r['skipped_locked']} in a locked period were skipped")
            if r["skipped_split"]:
                extra.append(f"{r['skipped_split']} matched a split fund (allocate manually)")
            if extra:
                msg += " (" + "; ".join(extra) + ")."
            messages.success(request, msg)
        else:
            note = ""
            if r["skipped_locked"]:
                note = f" {r['skipped_locked']} were in a locked period."
            messages.info(request, "No queued items matched the current rules — "
                                   "check the rule reference and match type." + note)
        return redirect("queue")


class BulkAllocateView(AllocateRequiredMixin, View):
    """Item 1: allocate several review-queue contributions to one fund in a single action,
    for faster clearing of the queue. Optionally sets a development group when the
    chosen fund is a development fund."""

    def post(self, request):
        from departments.models import DevelopmentGroup
        ids = request.POST.getlist("txn")
        raw_dept = request.POST.get("department", "")
        if not ids:
            messages.error(request, "Pick a fund and at least one contribution to allocate.")
            return redirect("queue")

        base_qs = Transaction.objects.filter(
            id__in=ids, allocation_status=Transaction.Status.REVIEW,
            direction=Transaction.Direction.CREDIT,
            processed_via_envelope=False, manual_receipt=False)

        # --- split fund (e.g. Combined Offering = Trust + Local) -------------
        if raw_dept.startswith("sf:"):
            from giving.models import SplitFund
            sf = SplitFund.objects.filter(pk=raw_dept[3:], active=True).first()
            if not sf:
                messages.error(request, "That split fund is no longer available.")
                return redirect("queue")
            n = 0
            for txn in base_qs:
                parts = [(d, amt, None) for d, amt in sf.split(txn.amount)]
                try:
                    txn.split_into(parts, user=request.user)
                except (ValueError, ArithmeticError):
                    continue
                txn.claimed_by = request.user
                txn.claimed_at = timezone.now()
                txn.save(update_fields=["claimed_by", "claimed_at"])
                n += 1
            if n:
                messages.success(request, f"Allocated {n} contribution(s) to {sf.name} — "
                                          "each split into its parts; the trust "
                                          "portion is queued for receipting.")
            else:
                messages.info(request, "No matching contributions to allocate.")
            return redirect("queue")

        # --- ordinary fund ----------------------------------------------------
        dept = Department.objects.filter(pk=raw_dept, active=True).first()
        if not dept:
            messages.error(request, "Pick a fund and at least one contribution to allocate.")
            return redirect("queue")
        grp = None
        if dept.category == Department.Category.DEVELOPMENT:
            grp = DevelopmentGroup.objects.filter(pk=request.POST.get("dev_group")).first()
        n = 0
        for txn in base_qs:
            txn.department = dept
            txn.dev_group = grp if grp else None
            txn.allocation_status = Transaction.Status.MANUAL
            txn.claimed_by = request.user
            txn.claimed_at = timezone.now()
            txn.save(update_fields=["department", "dev_group", "allocation_status",
                                    "claimed_by", "claimed_at"])
            n += 1
        if n:
            label = dept.name + (f" · {grp.label}" if grp else "")
            messages.success(request, f"Allocated {n} contribution(s) to {label}.")
        else:
            messages.info(request, "No matching contributions to allocate.")
        return redirect("queue")


class FetchUnallocatedView(AllocateRequiredMixin, View):
    """Item 5: pull credits that still need a fund — sitting in the ledger without
    a department but not currently in the review queue — into the queue so they
    can be allocated. A contribution can fall out of REVIEW (e.g. imported already
    'confirmed' but with no fund); this gathers them back for allocation."""

    @staticmethod
    def pending_qs():
        return Transaction.objects.filter(
            direction=Transaction.Direction.CREDIT, department__isnull=True,
            processed_via_envelope=False, manual_receipt=False,
            is_reversal=False, is_reversed=False,
            excluded_from_income=False).exclude(
            allocation_status=Transaction.Status.REVIEW)

    def post(self, request):
        n = self.pending_qs().update(allocation_status=Transaction.Status.REVIEW)
        if n:
            messages.success(request,
                f"Fetched {n} unallocated contribution(s) from the ledger into the queue.")
        else:
            messages.info(request, "No unallocated contributions found in the ledger.")
        return redirect("queue")


class ClaimResolveView(AllocateRequiredMixin, View):
    """Claim + resolve a review item; optionally remember the rule."""

    def post(self, request, pk):
        txn = get_object_or_404(Transaction, pk=pk,
                                allocation_status=Transaction.Status.REVIEW,
                                direction=Transaction.Direction.CREDIT)
        # --- split allocation: one bank gift meant for several funds ---
        if request.POST.get("split") == "1":
            from departments.models import Department as _D, DevelopmentGroup as _G
            from giving.models import SplitFund as _SF
            from decimal import Decimal as _Dec
            parts = []
            grps = request.POST.getlist("split_grp")
            for n, (d_id, amt) in enumerate(zip(request.POST.getlist("split_dept"),
                                                request.POST.getlist("split_amount"))):
                if not d_id or not str(amt).strip():
                    continue
                # a split-fund target sub-divides this part across its components
                if str(d_id).startswith("sf:"):
                    sf = _SF.objects.filter(pk=d_id[3:], active=True).first()
                    if not sf:
                        continue
                    try:
                        row_amt = _Dec(str(amt))
                    except (ArithmeticError, ValueError):
                        continue
                    for sub_d, sub_amt in sf.split(row_amt):
                        parts.append((sub_d, sub_amt, None))
                    continue
                d = _D.objects.filter(pk=d_id, active=True).first()
                if not d:
                    continue
                grp = None
                if d.category == Department.Category.DEVELOPMENT and n < len(grps) and grps[n]:
                    grp = _G.objects.filter(pk=grps[n]).first()
                parts.append((d, amt, grp))
            try:
                txn.split_into(parts, user=request.user)
            except (ValueError, ArithmeticError) as e:
                messages.error(request, f"Could not split: {e}")
                return redirect("queue")
            txn.claimed_by = request.user
            txn.claimed_at = timezone.now()
            txn.save(update_fields=["claimed_by", "claimed_at"])
            messages.success(request,
                f"Split across {len(parts)} funds: " +
                ", ".join(f"{d.name} {a}" for d, a, _ in parts) + ".")
            return redirect("queue")

        # --- split fund (e.g. Combined Offering = 50% Trust + 50% Local) ---
        raw_dept = request.POST.get("department", "")
        if raw_dept.startswith("sf:"):
            from giving.models import SplitFund
            sf = SplitFund.objects.filter(pk=raw_dept[3:], active=True).first()
            if not sf:
                messages.error(request, "That split fund is no longer available.")
                return redirect("queue")
            parts = [(d, amt, None) for d, amt in sf.split(txn.amount)]
            try:
                txn.split_into(parts, user=request.user)
            except (ValueError, ArithmeticError) as e:
                messages.error(request, f"Could not split: {e}")
                return redirect("queue")
            txn.claimed_by = request.user
            txn.claimed_at = timezone.now()
            txn.save(update_fields=["claimed_by", "claimed_at"])
            if request.POST.get("remember_rule") and txn.reference:
                ref = normalize_reference(txn.reference)
                AllocationRule.objects.update_or_create(
                    reference=ref,
                    defaults={"split_fund": sf, "department": None,
                              "source": AllocationRule.Source.LEARNED})
            messages.success(request,
                f"Allocated to {sf.name} — split into " +
                ", ".join(f"{d.name} {a}" for d, a in sf.split(txn.amount)) +
                ". The trust portion is queued for receipting.")
            return redirect("queue")

        form = QueueResolveForm(request.POST)
        if not form.is_valid():
            messages.error(request, "Pick a fund to allocate to.")
            return redirect("queue")

        dept = form.cleaned_data["department"]
        txn.department = dept
        if dept.category == Department.Category.DEVELOPMENT and request.POST.get("dev_group"):
            from departments.models import DevelopmentGroup as _G
            txn.dev_group = _G.objects.filter(pk=request.POST["dev_group"]).first()
        txn.allocation_status = Transaction.Status.MANUAL
        txn.claimed_by = request.user
        txn.claimed_at = timezone.now()
        txn.save()

        resolved_similar = 0
        if form.cleaned_data["remember_rule"] and txn.reference:
            ref = normalize_reference(txn.reference)
            AllocationRule.objects.update_or_create(
                reference=ref,
                defaults={"department": dept, "split_fund": None,
                          "source": AllocationRule.Source.LEARNED},
            )
            # apply to all other queued items with the same reference
            similar = Transaction.objects.filter(
                allocation_status=Transaction.Status.REVIEW,
                reference__iexact=txn.reference,
            ).exclude(pk=txn.pk)
            resolved_similar = similar.update(
                department=dept, allocation_status=Transaction.Status.LEARNED)

        msg = f"Allocated to {dept.name}."
        if resolved_similar:
            msg += f" Rule applied to {resolved_similar} similar item(s)."
        messages.success(request, msg)
        return redirect("queue")


class CashEntryListView(TransactionListView):
    """A focused, filterable table of cash collections (channel = CASH), so cash
    entries can be checked the same way envelopes are."""
    def get_queryset(self):
        return super().get_queryset().filter(channel="CASH",
                                             direction=Transaction.Direction.CREDIT)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["cash_only"] = True
        return ctx


class CashEntryCreate(DataEntryRequiredMixin, CreateView):
    model = Transaction
    form_class = CashEntryForm
    template_name = "giving/cash_form.html"
    success_url = reverse_lazy("transaction_list")

    def form_valid(self, form):
        if _block_if_locked(self.request, form.cleaned_data.get("date")):
            return redirect("cash_new")
        # H3: guard against double-entering the same collection (e.g. once by the
        # assistant and once by the treasurer). Require explicit confirmation.
        if not self.request.POST.get("confirm_duplicate") and _cash_duplicate(
                form.cleaned_data.get("date"), form.cleaned_data.get("department"),
                form.cleaned_data.get("amount"), form.cleaned_data.get("payer_name")):
            messages.warning(self.request,
                "A cash entry for this fund, date and amount already exists. If this is a "
                "separate collection, tick \u201cThis is not a duplicate\u201d and save again.")
            ctx = self.get_context_data(form=form)
            ctx["duplicate_warning"] = True
            return self.render_to_response(ctx)
        split = form.split_fund
        if split:
            base = form.save(commit=False)
            from core.utils import sabbath_of as _sof
            _svc = _sof(base.date) if base.date else None
            for dept, amt in split.split(base.amount):
                Transaction.objects.create(
                    date=base.date, channel=base.channel,
                    direction=Transaction.Direction.CREDIT,
                    allocation_status=Transaction.Status.MANUAL,
                    sabbath_week=sabbath_week_of(base.date),
                    service_sabbath=_svc,
                    amount=amt, department=dept, member=base.member,
                    reference=base.reference, payer_name=base.payer_name)
            messages.success(self.request, f"Entry recorded and split across {split.name}.")
            return redirect(self.success_url)
        txn = form.save(commit=False)
        txn.direction = Transaction.Direction.CREDIT
        txn.allocation_status = Transaction.Status.MANUAL
        txn.sabbath_week = sabbath_week_of(txn.date)
        # the treasurer dated this cash to a specific Sabbath; honour it directly
        # rather than rolling a "closed" Sabbath forward (that roll is for bank
        # gifts that physically arrive after a Sabbath, not counted cash).
        from core.utils import sabbath_of as _sof
        if txn.date and txn.service_sabbath is None:
            txn.service_sabbath = _sof(txn.date)
        txn.save()
        # offer/apply a pledge match if this giver has an active pledge
        try:
            from pledges.services.matching import handle_new_contribution
            note = handle_new_contribution(txn, user=self.request.user)
        except Exception:
            from core.utils import log_exception as _lx; _lx('giving/views.py')
            note = None
        if note:
            messages.success(self.request, f"Entry recorded — {note}.")
        else:
            messages.success(self.request, "Entry recorded.")
        return redirect(self.success_url)


class RuleListView(DataEntryRequiredMixin, ListView):
    model = AllocationRule
    template_name = "giving/rule_list.html"
    context_object_name = "rules"
    paginate_by = 50

    def get_queryset(self):
        view = self.request.GET.get("view", "active")
        qs = (AllocationRule.objects.select_related("department", "split_fund")
              .order_by("reference", "id"))
        if view == "archived":
            qs = qs.filter(archived=True)
        elif view == "expired":
            import datetime as _d
            qs = qs.filter(archived=False, valid_to__lt=_d.date.today())
        else:  # active = not archived
            qs = qs.filter(archived=False)
        return qs

    def get_context_data(self, **kwargs):
        import datetime as _d
        ctx = super().get_context_data(**kwargs)
        ctx["form"] = RuleForm()
        ctx["view"] = self.request.GET.get("view", "active")
        ctx["archived_count"] = AllocationRule.objects.filter(archived=True).count()
        ctx["expired_count"] = AllocationRule.objects.filter(
            archived=False, valid_to__lt=_d.date.today()).count()
        ctx["today"] = _d.date.today()
        return ctx


class RuleCreateView(DataEntryRequiredMixin, CreateView):
    model = AllocationRule
    form_class = RuleForm
    template_name = "giving/rule_form.html"
    success_url = reverse_lazy("rule_list")

    def form_valid(self, form):
        form.instance.reference = normalize_reference(form.instance.reference)
        messages.success(self.request, "Rule saved.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "That reference already has a rule.")
        return redirect("rule_list")


class RuleDeleteView(TreasurerRequiredMixin, View):
    def post(self, request, pk):
        from django.shortcuts import get_object_or_404
        rule = get_object_or_404(AllocationRule, pk=pk)
        ref = rule.reference
        rule.delete()
        messages.success(request, f"Deleted allocation rule “{ref}”.")
        return redirect("rule_list")

    def get(self, request, *args, **kwargs):
        return self.post(request, *args, **kwargs)


from django.views.generic import UpdateView
from .forms import TransactionEditForm


class TransactionUpdateView(DataEntryRequiredMixin, UpdateView):
    model = Transaction
    form_class = TransactionEditForm
    template_name = "giving/transaction_form.html"

    def get_success_url(self):
        return reverse_lazy("transaction_list")

    def form_valid(self, form):
        # M6: block if either the original or the new date falls in a locked period
        original = type(self.object).objects.filter(pk=self.object.pk).values_list(
            "date", flat=True).first()
        if _block_if_locked(self.request, form.instance.date) or \
           (original and _block_if_locked(self.request, original)):
            return redirect("transaction_list")
        # did the "manual receipt" box change on this save? (reversible both ways)
        changed_manual = "manual_receipt" in form.changed_data
        new_value = form.instance.manual_receipt
        response = super().form_valid(form)
        if changed_manual:
            # marking on: pull it (and split siblings) out of the queues so it
            # isn't receipted again. Un-marking: clear the flag on the whole gift
            # so it becomes eligible for a system receipt once more.
            n = self.object.mark_manual_receipt(value=new_value, cascade_split=True)
            if new_value and n > 1:
                messages.success(self.request,
                    f"Marked {n} split parts as manual receipts and cleared them "
                    f"from the queue.")
            elif new_value:
                messages.success(self.request, "Marked as a manual receipt.")
            else:
                messages.success(self.request,
                    "Manual-receipt mark removed — this contribution can be receipted again.")
        else:
            messages.success(self.request, "Entry updated.")
        from core.models import reconciled_period_warning
        warn = reconciled_period_warning(self.object.date)
        if warn:
            messages.warning(self.request, warn)
        return response


# ---- Bank-statement debit handling ----
from cashbook.models import Expense


def _float_fund():
    fund, _ = Department.objects.get_or_create(
        name="Float / Cash on hand",
        defaults=dict(fund_type=Department.FundType.LOCAL,
                      category=Department.Category.HOLDING))
    return fund


def _dept_from_post(request, field="department"):
    """Look up a Department from a POST field, tolerating a blank/missing value
    (an empty '— fund —' option, or a hidden field the JS never revealed) rather
    than letting Department.objects.filter(pk='') raise a ValueError."""
    raw = (request.POST.get(field) or "").strip()
    if not raw:
        return None
    return Department.objects.filter(pk=raw).first()


class DebitQueueView(DebitClassifyRequiredMixin, ListView):
    """Bank-statement debits awaiting classification."""
    template_name = "giving/debit_queue.html"
    context_object_name = "debits"
    paginate_by = 50

    def get_queryset(self):
        return (Transaction.objects.filter(
            direction=Transaction.Direction.DEBIT,
            channel=Transaction.Channel.BANK,
            allocation_status=Transaction.Status.REVIEW)
            .order_by("-date"))

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["funds"] = Department.objects.filter(active=True, selectable=True)
        ctx["categories"] = Expense.Category.choices
        # Payment Register suggestions: for each debit on this page, an
        # outstanding instrument matched by number-in-narration or unique
        # exact amount — one click marks it cleared on the debit's date
        from cashbook.services.payments import suggest_instrument_for_debit
        suggestions = {}
        for d in ctx["debits"]:
            inst, how = suggest_instrument_for_debit(d)
            if inst:
                suggestions[d.pk] = {"inst": inst, "how": how}
        ctx["instrument_suggestions"] = suggestions
        ctx["pending_expenses"] = Expense.objects.filter(
            status__in=[Expense.Status.PENDING, Expense.Status.APPROVED],
            bank_transaction__isnull=True).order_by("-date")[:200]
        # open remittance batches — a trust remittance debit usually settles a
        # whole batch (one payment covering several trust funds), so it should
        # be matched to the batch's per-fund lines rather than one fund
        from cashbook.models import RemittanceBatch
        ctx["open_batches"] = (RemittanceBatch.objects
            .filter(status__in=[RemittanceBatch.Status.DRAFT,
                                RemittanceBatch.Status.APPROVED])
            .order_by("-date", "-id")[:20])
        return ctx


class DebitResolveView(DebitClassifyRequiredMixin, View):
    """Classify one debit as a bank charge, a general expense, a float
    withdrawal, or a match to an existing expense."""

    def post(self, request, pk):
        txn = get_object_or_404(
            Transaction, pk=pk, direction=Transaction.Direction.DEBIT)
        # a debit carries a date; resolving it posts an expense/transfer on that
        # date, so it must honour the period lock just like any other entry.
        if _block_if_locked(request, txn.date):
            return redirect("debit_queue")
        kind = request.POST.get("kind")

        if kind == "bank_charge":
            dept = _dept_from_post(request) or _float_fund()
            Expense.objects.create(
                date=txn.date, sabbath_week=sabbath_week_of(txn.date), department=dept,
                description=(txn.raw_narration or "Bank charge")[:200], amount=txn.amount,
                category=Expense.Category.BANK_CHARGE, method=Expense.Method.BANK,
                status=Expense.Status.PAID, paid_date=txn.date,
                recorded_by=request.user, approved_by=request.user, bank_transaction=txn)
            txn.department = dept
            txn.allocation_status = Transaction.Status.MANUAL
            txn.save(update_fields=["department", "allocation_status"])
            messages.success(request, "Recorded as a bank charge.")

        elif kind == "expense":
            dept = _dept_from_post(request)
            if not dept:
                messages.error(request, "Choose a fund for the expense.")
                return redirect("debit_queue")
            cat = request.POST.get("category") or Expense.Category.OTHER
            Expense.objects.create(
                date=txn.date, sabbath_week=sabbath_week_of(txn.date), department=dept,
                description=(request.POST.get("description") or txn.raw_narration or "Expense")[:200],
                amount=txn.amount, category=cat, method=Expense.Method.BANK,
                status=Expense.Status.PAID, paid_date=txn.date,
                recorded_by=request.user, approved_by=request.user, bank_transaction=txn)
            txn.department = dept
            txn.allocation_status = Transaction.Status.MANUAL
            txn.save(update_fields=["department", "allocation_status"])
            messages.success(request, "Recorded as an expense.")

        elif kind == "remittance_batch":
            # a trust remittance usually settles a whole batch — one bank payment
            # covering several trust funds. Match the debit to the batch: every
            # per-fund expense line in it becomes PAID (each fund is charged its
            # own share), so the payment is never forced onto a single fund.
            from cashbook.models import RemittanceBatch
            raw = (request.POST.get("batch") or "").strip()
            batch = (RemittanceBatch.objects.filter(pk=raw).first()
                     if raw.isdigit() else None)
            if not batch:
                messages.error(request, "Choose the remittance batch this payment settles.")
                return redirect("debit_queue")
            if batch.status == RemittanceBatch.Status.REMITTED:
                messages.error(request, f"Batch {batch.batch_number} is already marked sent.")
                return redirect("debit_queue")
            if abs((batch.total_amount or Decimal(0)) - txn.amount) > Decimal("0.01"):
                messages.error(request,
                    f"This debit is {txn.amount:,.2f} but batch {batch.batch_number} "
                    f"totals {batch.total_amount:,.2f}. Match it to the right batch, "
                    "or adjust the batch first.")
                return redirect("debit_queue")
            from django.utils import timezone as _tz
            from django.db import transaction as _db_transaction
            from reports.views import _repost_to_ledger
            with _db_transaction.atomic():
                batch.status = RemittanceBatch.Status.REMITTED
                batch.remitted_at = _tz.now()
                batch.save(update_fields=["status", "remitted_at"])
                batch.expenses.update(status=Expense.Status.PAID, paid_date=txn.date)
                _repost_to_ledger(batch.expenses.all())
                txn.allocation_status = Transaction.Status.MANUAL
                txn.raw_narration = (txn.raw_narration or "") + \
                    f"\n[Settles remittance batch {batch.batch_number} — " \
                    f"{batch.expenses.count()} trust fund(s)]"
                txn.save(update_fields=["allocation_status", "raw_narration"])
            messages.success(request,
                f"Matched to batch {batch.batch_number}: {batch.expenses.count()} "
                "trust fund line(s) marked paid — each fund charged its own share.")

        elif kind == "remittance":
            dept = _dept_from_post(request)
            if not dept:
                messages.error(request, "Choose the trust fund being remitted.")
                return redirect("debit_queue")
            # if there's a recent DRAFT/APPROVED remittance batch that hasn't been
            # sent yet, link this expense to it — this is very likely the payment
            # that settles it, and keeps the batch's status accurate.
            from cashbook.models import RemittanceBatch
            recent_batch = (RemittanceBatch.objects
                            .filter(status__in=[RemittanceBatch.Status.DRAFT,
                                                RemittanceBatch.Status.APPROVED])
                            .order_by("-created_at").first())
            exp = Expense.objects.create(
                date=txn.date, sabbath_week=sabbath_week_of(txn.date), department=dept,
                description=(request.POST.get("description") or "Remittance to field")[:200],
                amount=txn.amount, category=Expense.Category.REMITTANCE,
                method=Expense.Method.BANK, status=Expense.Status.PAID,
                paid_date=txn.date, recorded_by=request.user, approved_by=request.user,
                bank_transaction=txn, remittance_batch=recent_batch)
            txn.department = dept
            txn.allocation_status = Transaction.Status.MANUAL
            txn.save(update_fields=["department", "allocation_status"])
            if recent_batch:
                messages.success(request, f"Recorded as a trust remittance for {dept.name} "
                    f"and linked to remittance batch {recent_batch.batch_number}.")
            else:
                messages.success(request, f"Recorded as a trust remittance for {dept.name}. "
                    "No open remittance batch was found to link it to.")

        elif kind == "match":
            ids = request.POST.getlist("expense")
            exps = list(Expense.objects.filter(pk__in=ids))
            if not exps:
                messages.error(request, "Choose at least one expense to match.")
                return redirect("debit_queue")
            total = sum((e.amount for e in exps), Decimal(0))
            if abs(total - txn.amount) > Decimal("0.01"):
                messages.error(
                    request, f"The selected expense(s) total {total:,.2f} but the "
                             f"debit is {txn.amount:,.2f} — they must match.")
                return redirect("debit_queue")
            for exp in exps:
                exp.bank_transaction = txn
                if exp.status != Expense.Status.PAID:
                    exp.status = Expense.Status.PAID
                    exp.paid_date = txn.date
                exp.save()
            # Payment Register integration: the matched expenses' outstanding
            # payment instruments are cleared with THIS DEBIT'S DATE as the
            # cleared date (the bank's clearance date) and linked to the debit
            # for the reconciliation trail — no duplicate records, and
            # historical reconciliations pick the right date automatically.
            from cashbook.services.payments import clear_for_bank_debit
            cleared_insts = clear_for_bank_debit(txn, request.user, exps)
            dept_ids = {e.department_id for e in exps}
            if len(dept_ids) == 1:
                txn.department = exps[0].department
            else:
                # one withdrawal spanning several funds: the per-fund split is held
                # by the linked expenses (each reduces its own fund); the debit line
                # itself is left unallocated and flagged as a multi-fund payment.
                txn.department = None
                names = ", ".join(sorted({e.department.name for e in exps}))
                txn.raw_narration = (txn.raw_narration or "") + \
                    f"\n[Split across {len(dept_ids)} funds: {names}]"
            txn.allocation_status = Transaction.Status.MANUAL
            txn.save()
            note = (f" across {len(dept_ids)} funds" if len(dept_ids) > 1 else "")
            pay_note = (f" {len(cleared_insts)} payment instrument(s) marked "
                        f"cleared on {txn.date:%d %b}." if cleared_insts else "")
            messages.success(
                request, f"Matched {len(exps)} expense(s) totalling {total:,.2f}{note} "
                         f"to this debit.{pay_note}")

        elif kind == "clear_instrument":
            # the debit IS the clearance of a known outstanding instrument:
            # mark it cleared on the debit's date, link both ways, and settle
            # the instrument's source expense(s) if still unpaid
            from cashbook.models import PaymentInstrument
            from cashbook.services.payments import apply_event
            inst = PaymentInstrument.objects.filter(
                pk=request.POST.get("instrument")).first()
            if not inst:
                messages.error(request, "Choose the payment instrument this debit clears.")
                return redirect("debit_queue")
            if inst.bank_transaction_id:
                messages.error(request, "That instrument is already cleared by another debit.")
                return redirect("debit_queue")
            if abs(inst.amount - txn.amount) > Decimal("0.01"):
                messages.error(request,
                    f"The debit is {txn.amount:,.2f} but the instrument is "
                    f"{inst.amount:,.2f} — they must match.")
                return redirect("debit_queue")
            try:
                apply_event(inst, "CLEAR", request.user, on=txn.date,
                            bank_transaction=txn,
                            comment="Cleared from the debit review queue")
            except Exception as exc:  # noqa: BLE001 — surface validation errors
                messages.error(request, str(exc))
                return redirect("debit_queue")
            depts = set()
            for exp in inst.all_expenses:
                if exp.bank_transaction_id is None:
                    exp.bank_transaction = txn
                if exp.status != Expense.Status.PAID:
                    exp.status = Expense.Status.PAID
                    exp.paid_date = txn.date
                exp.save()
                depts.add(exp.department_id)
            if len(depts) == 1 and inst.expense_id:
                txn.department = inst.expense.department
            txn.allocation_status = Transaction.Status.MANUAL
            txn.save()
            messages.success(request,
                f"{inst.get_method_display()} {inst.instrument_number or inst.pk} "
                f"marked cleared on {txn.date:%d %b %Y}.")

        elif kind == "float":
            fund = _float_fund()
            txn.department = fund
            txn.allocation_status = Transaction.Status.MANUAL
            txn.save(update_fields=["department", "allocation_status"])
            messages.success(
                request, "Marked as a float withdrawal (cash on hand). Record "
                         "expenses against it as they are paid.")

        elif kind == "petty_cash":
            # the bank withdrawal funded the petty-cash float: record a top-up so
            # the float reflects it. This moves money bank -> cash on hand; it
            # doesn't reduce total cash, so the debit is not booked as an expense.
            from cashbook.models import PettyCashTopUp
            PettyCashTopUp.objects.create(
                date=txn.date, amount=txn.amount,
                note=(request.POST.get("description") or txn.raw_narration
                      or "Bank withdrawal to petty cash")[:200],
                recorded_by=request.user)
            txn.department = _float_fund()
            txn.allocation_status = Transaction.Status.MANUAL
            txn.save(update_fields=["department", "allocation_status"])
            messages.success(
                request, f"Allocated to petty cash — the float has been topped up by "
                         f"{txn.amount:,.2f}.")
        elif kind == "already_accounted":
            # the treasurer confirms this payment was already recorded elsewhere
            # (e.g. entered manually, a duplicate, or out of scope) — resolve the
            # queue item without creating an expense or touching any fund balance.
            reason = (request.POST.get("description") or "").strip()
            if not reason:
                messages.error(request, "Add a short note on why this is already accounted for.")
                return redirect("debit_queue")
            txn.allocation_status = Transaction.Status.MANUAL
            txn.raw_narration = (txn.raw_narration or "") + \
                f"\n[Marked already accounted for — {reason} — by {request.user.get_username()}]"
            txn.save(update_fields=["allocation_status", "raw_narration"])
            messages.success(request, "Marked as already accounted for. It won't "
                             "affect any fund balance.")

        else:
            messages.error(request, "Choose how to treat this debit.")
        return redirect("debit_queue")


# ---- Review-queue export / import (offline matching) ----
import csv as _csv
import io as _io
from django.http import HttpResponse


class QueueExportView(ReadAccessMixin, View):
    """Download the giving review queue as CSV for offline allocation."""
    def get(self, request):
        resp = HttpResponse(content_type="text/csv")
        resp["Content-Disposition"] = 'attachment; filename="review_queue.csv"'
        w = _csv.writer(resp)
        w.writerow(["id", "date", "amount", "payer_name", "payer_phone",
                    "reference", "narration", "allocate_to_fund"])
        for t in (Transaction.objects.filter(
                allocation_status=Transaction.Status.REVIEW,
                direction=Transaction.Direction.CREDIT).order_by("date")):
            w.writerow([t.id, t.date, t.amount, t.payer_name, t.payer_phone,
                        t.reference, (t.raw_narration or "").replace("\n", " "), ""])
        return resp


class QueueImportView(DataEntryRequiredMixin, View):
    """Upload the filled queue CSV: each row's 'allocate_to_fund' (a fund name)
    allocates that transaction. Optionally remembers a rule for the reference."""
    template_name = "giving/queue_import.html"

    def get(self, request):
        return render(request, self.template_name, {})

    def post(self, request):
        f = request.FILES.get("file")
        remember = request.POST.get("remember") == "1"
        if not f:
            messages.error(request, "Choose the filled CSV.")
            return redirect("queue_import")
        try:
            reader = _csv.DictReader(_io.TextIOWrapper(f.file, encoding="utf-8-sig"))
        except Exception:
            from core.utils import log_exception as _lx; _lx('giving/views.py')
            messages.error(request, "Could not read that CSV.")
            return redirect("queue_import")
        funds = {d.name.lower(): d for d in Department.objects.filter(active=True)}
        done = skipped = 0
        for row in reader:
            fund_name = (row.get("allocate_to_fund") or "").strip().lower()
            tid = (row.get("id") or "").strip()
            if not fund_name or not tid.isdigit():
                continue
            dept = funds.get(fund_name)
            txn = Transaction.objects.filter(pk=tid,
                  allocation_status=Transaction.Status.REVIEW).first()
            if not dept or not txn:
                skipped += 1
                continue
            txn.department = dept
            txn.allocation_status = Transaction.Status.MANUAL
            txn.save(update_fields=["department", "allocation_status"])
            done += 1
            if remember and txn.reference:
                from .services.allocation import normalize_reference
                AllocationRule.objects.get_or_create(
                    reference=normalize_reference(txn.reference),
                    defaults=dict(department=dept, source=AllocationRule.Source.LEARNED))
        msg = f"Allocated {done} item(s)."
        if skipped:
            msg += f" Skipped {skipped} (unknown fund or already allocated)."
        messages.success(request, msg)
        return redirect("queue")


class MarkProcessedImportView(DataEntryRequiredMixin, View):
    """Bulk-mark bank entries as 'processed via envelope' — handled already, so
    kept out of the receipting/review flow, but NOT receipted (no envelope record
    is created). This is for contributions a member wrote on a physical envelope: the
    money is on the bank statement, but it must not be receipted again.

    Upload a small file with just a REFERENCE and an AMOUNT per row. The reference
    finds the bank transaction; the amount confirms it's the right record (a
    mismatch is reported, not applied). Accepts .csv or .xlsx.
    """
    template_name = "giving/mark_processed_import.html"

    def get(self, request):
        if request.GET.get("template"):
            return self._template()
        return render(request, self.template_name, {})

    def _template(self):
        from django.http import HttpResponse
        resp = HttpResponse(content_type="text/csv")
        resp["Content-Disposition"] = 'attachment; filename="mark_processed_template.csv"'
        w = _csv.writer(resp)
        w.writerow(["reference", "amount"])
        w.writerow(["UER2Q5NF2W", "1500"])
        w.writerow(["AC0C40FD2E26", "2000"])
        return resp

    def _rows_from_upload(self, f):
        """Yield (reference, amount_or_None) from a .csv or .xlsx upload, tolerant
        of header names and column order."""
        name = (getattr(f, "name", "") or "").lower()
        rows = []
        if name.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            data = list(ws.iter_rows(values_only=True))
            if not data:
                return rows
            header = [str(c).strip().lower() if c is not None else "" for c in data[0]]
            ref_i = next((i for i, h in enumerate(header)
                          if h in ("reference", "ref", "core ref", "receipt")), 0)
            amt_i = next((i for i, h in enumerate(header)
                          if h in ("amount", "amt", "value")), 1)
            for r in data[1:]:
                ref = str(r[ref_i]).strip() if ref_i < len(r) and r[ref_i] not in (None, "") else ""
                amt = r[amt_i] if amt_i < len(r) else None
                if ref:
                    rows.append((ref, amt))
        else:
            reader = _csv.DictReader(_io.TextIOWrapper(f.file, encoding="utf-8-sig"))
            # normalise header keys
            for raw in reader:
                row = { (k or "").strip().lower(): v for k, v in raw.items() }
                ref = (row.get("reference") or row.get("ref")
                       or row.get("core ref") or row.get("receipt") or "").strip()
                amt = row.get("amount") or row.get("amt") or row.get("value")
                if ref:
                    rows.append((ref, amt))
        return rows

    def post(self, request):
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "Choose a file with reference and amount columns.")
            return redirect("mark_processed_import")
        try:
            rows = self._rows_from_upload(f)
        except Exception:
            from core.utils import log_exception as _lx; _lx('giving/views.py')
            messages.error(request, "Could not read that file — upload the .csv or "
                                    ".xlsx from the template.")
            return redirect("mark_processed_import")
        if not rows:
            messages.warning(request, "No rows with a reference were found.")
            return redirect("mark_processed_import")

        marked = already = not_found = mismatched = ambiguous = 0
        problems = []

        def _mark(txn):
            """Mark one transaction as a manual (paper) receipt; return True if
            newly changed. Cascade is off because the importer has already matched
            the full split group by its total and marks each row explicitly."""
            return txn.mark_manual_receipt(value=True, cascade_split=False) > 0

        for ref, raw_amt in rows:
            # match a bank CREDIT by any of the reference-bearing fields
            qs = Transaction.objects.filter(
                channel=Transaction.Channel.BANK,
                direction=Transaction.Direction.CREDIT,
                is_reversal=False, is_reversed=False).filter(
                Q(reference__iexact=ref) | Q(core_ref__iexact=ref)
                | Q(bank_receipt__iexact=ref) | Q(mpesa_ref__iexact=ref))
            n = qs.count()
            if n == 0:
                not_found += 1
                if len(problems) < 12:
                    problems.append(f"“{ref}”: no matching bank entry")
                continue

            # parse the confirming amount, if supplied
            amt = None
            if raw_amt not in (None, ""):
                try:
                    amt = Decimal(str(raw_amt).replace(",", "").strip())
                except Exception:
                    from core.utils import log_exception as _lx; _lx('giving/views.py')
                    amt = None

            if n == 1:
                txn = qs.first()
                if amt is not None and txn.amount != amt:
                    mismatched += 1
                    if len(problems) < 12:
                        problems.append(f"“{ref}”: amount {amt} ≠ recorded {txn.amount}")
                    continue
                if _mark(txn):
                    marked += 1
                else:
                    already += 1
                continue

            # --- multiple matches: most often a SPLIT-FUND gift -----------------
            # A split gift (e.g. Combined Offering) is posted as several rows that
            # share the reference but divide the amount. The uploaded amount is the
            # original lump sum, so it equals the SUM of the group, not any one row.
            rows_qs = list(qs)
            total = sum((t.amount for t in rows_qs), Decimal(0))
            if amt is not None and total == amt:
                # whole split group confirmed by its total — mark every part
                newly = sum(1 for t in rows_qs if _mark(t))
                if newly:
                    marked += newly
                else:
                    already += 1
                continue
            # otherwise, an exact single-row amount match still disambiguates
            if amt is not None:
                exact = [t for t in rows_qs if t.amount == amt]
                if len(exact) == 1:
                    if _mark(exact[0]):
                        marked += 1
                    else:
                        already += 1
                    continue
            # genuinely ambiguous — report with both the count and the group total
            ambiguous += 1
            if len(problems) < 12:
                hint = (f"sum is {total}" if amt is None
                        else f"amount {amt} ≠ any row and ≠ split total {total}")
                problems.append(f"“{ref}”: matches {n} entries — {hint}")

        parts = [f"{marked} marked as manual receipt"]
        if already:
            parts.append(f"{already} already marked")
        if mismatched:
            parts.append(f"{mismatched} amount mismatch")
        if ambiguous:
            parts.append(f"{ambiguous} ambiguous")
        if not_found:
            parts.append(f"{not_found} not found")
        msg = ", ".join(parts) + "."
        if problems:
            msg += " Issues: " + "; ".join(problems)
        (messages.success if marked else messages.warning)(request, msg)
        return redirect("transaction_list")


class TransactionSendToReviewView(TreasurerRequiredMixin, View):
    """Undo a wrong allocation and send it back to the review queue for
    correct re-allocation — the answer to "this was wrongly auto-split
    across funds/groups, how do I put it back as one fund?"

    Reverses the entry (contra posting, same as TransactionReverseView) and,
    if it's part of a split contribution, every sibling too — then creates
    ONE new entry for the full combined original amount, in REVIEW status,
    ready to be correctly allocated as a single fund via the normal review
    queue. Nothing is ever deleted: the original split rows stay on the
    ledger, reversed, with contra postings — the replacement entry is a new,
    separate row, fully traceable back to what it replaced via its
    raw_narration."""
    def post(self, request, pk):
        from core.models import period_locked
        t = get_object_or_404(Transaction, pk=pk)
        lock = period_locked(t.date)
        if lock:
            messages.error(request, f"{lock} is locked. An administrator must unlock "
                                    "the period before this entry can be sent back to review.")
            return redirect(request.META.get("HTTP_REFERER") or "transaction_list")
        if t.is_reversed or t.is_reversal:
            messages.error(request, "This entry has already been reversed, or is itself "
                                    "a reversal, and can't be sent back to review.")
            return redirect(request.META.get("HTTP_REFERER") or "transaction_list")

        reason = (request.POST.get("reason") or "").strip()
        group = [t] + list(t.strict_split_siblings())
        total = Decimal(0)
        reversed_ids = []
        for member in group:
            if member.is_reversed or member.is_reversal or period_locked(member.date):
                continue
            total += member.amount
            member.reverse(request.user,
                          reason=reason or "Sent back to review for re-allocation")
            reversed_ids.append(member.pk)
            TransactionReverseView._delete_linked_envelope(member, request.user)

        if not reversed_ids:
            messages.error(request, "Nothing could be reversed (already reversed, "
                                    "or the period is locked).")
            return redirect(request.META.get("HTTP_REFERER") or "transaction_list")

        replacement = Transaction.objects.create(
            date=t.date, sabbath_week=t.sabbath_week, channel=t.channel,
            direction=t.direction, amount=total, member=t.member,
            reference=t.reference, payer_name=t.payer_name, payer_phone=t.payer_phone,
            mpesa_ref=t.mpesa_ref, confirmed=t.confirmed,
            allocation_status=Transaction.Status.REVIEW,
            raw_narration=(f"Replaces #{', #'.join(str(i) for i in reversed_ids)} — "
                          f"sent back to review for correct allocation"
                          + (f': "{reason}"' if reason else "")))

        n = len(reversed_ids)
        messages.success(request,
            f"{n} entr{'y' if n == 1 else 'ies'} reversed and combined into one new "
            f"entry of {total:,.2f} in the review queue — allocate it to the right "
            f"fund from there.")
        return redirect("queue")


class TransactionHistoryView(ReadAccessMixin, View):
    """The audit trail for one ledger entry — who created it and every
    change since, from django-simple-history's own tracking (already
    recorded on every save; this view is the first place it's actually
    surfaced to a user for a single transaction, rather than only in the
    all-models Audit Log report)."""
    def get(self, request, pk):
        from django.shortcuts import render
        t = get_object_or_404(Transaction, pk=pk)
        records = list(t.history.all().select_related("history_user").order_by("history_date"))
        FIELDS = ["amount", "department_id", "dev_group_id", "allocation_status",
                 "confirmed", "is_reversed", "is_reversal", "manual_receipt",
                 "processed_via_envelope", "reference", "payer_name", "payer_phone"]
        entries = []
        prev = None
        for rec in records:
            changes = []
            if prev is not None:
                for f in FIELDS:
                    old, new = getattr(prev, f, None), getattr(rec, f, None)
                    if old != new:
                        changes.append((f, old, new))
            entries.append({"record": rec, "changes": changes})
            prev = rec
        entries.reverse()   # most recent first
        return render(request, "giving/transaction_history.html",
                     {"txn": t, "entries": entries})


class TransactionReverseView(TreasurerRequiredMixin, View):
    """Reverse a ledger entry (treasury never deletes — it posts a contra entry).
    Blocked inside a locked period unless an admin overrides."""
    def post(self, request, pk):
        from core.models import period_locked
        t = get_object_or_404(Transaction, pk=pk)
        lock = period_locked(t.date)
        if lock:
            messages.error(request, f"{lock} is locked. An administrator must unlock "
                                    "the period before this entry can be reversed.")
            return redirect(request.META.get("HTTP_REFERER") or "transaction_list")
        try:
            t.reverse(request.user, reason=(request.POST.get("reason") or "").strip())
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect(request.META.get("HTTP_REFERER") or "transaction_list")

        # If this ledger entry came from an envelope, reverse the rest of that
        # envelope's entries too and DELETE the envelope receipt (rather than
        # leaving it struck through in the list).
        removed_envelope = self._delete_linked_envelope(t, request.user)

        if removed_envelope:
            messages.success(request, "Entry reversed and its envelope receipt removed "
                                      "— contra postings remain on the ledger for audit.")
        else:
            messages.success(request, "Entry reversed — a contra posting was created and "
                                      "both remain on the ledger for the audit trail.")
        return redirect(request.META.get("HTTP_REFERER") or "transaction_list")

    @staticmethod
    def _delete_linked_envelope(txn, user):
        """If `txn` belongs to an envelope, reverse the envelope's other ledger
        entries and delete the envelope (and its lines). Returns True if done."""
        from envelopes.models import Envelope
        env = (Envelope.objects.filter(lines__transaction=txn).first()
               or Envelope.objects.filter(bank_transaction=txn).first())
        if env is None:
            return False
        for st in env.linked_transactions:
            if st and st.pk != txn.pk and not st.is_reversed and not st.is_reversal:
                try:
                    st.reverse(user, reason=f"Envelope #{env.receipt_no} reversed")
                except ValueError:
                    pass
        env.delete()   # cascades EnvelopeLine rows; reversed ledger entries remain
        return True


class TransactionSplitView(TreasurerRequiredMixin, View):
    """Split one lump-sum entry across several funds and/or development groups
    (e.g. a single 2,000 bank deposit meant for two groups)."""
    template_name = "giving/transaction_split.html"

    def _blocked_reason(self, t):
        """Splitting must never be allowed once a receipt has already been
        issued for this entry (envelope or manual) — dividing the ledger row
        afterward would create a mismatch with what the issued receipt says,
        letting someone alter an already-allocated receipt without realising
        it. Also blocked for a reversed original or a reversal itself, which
        are correction entries, not something to further subdivide."""
        if t.is_reversed or t.is_reversal:
            return "This entry has been reversed and can't be split."
        if t.processed_via_envelope or t.manual_receipt:
            return "This entry has already been receipted and can't be split."
        return None

    def get(self, request, pk):
        from departments.models import Department, DevelopmentGroup
        t = get_object_or_404(Transaction, pk=pk)
        reason = self._blocked_reason(t)
        if reason:
            messages.error(request, reason)
            return redirect("transaction_list")
        return render(request, self.template_name, {
            "txn": t,
            "departments": Department.objects.filter(active=True, selectable=True).order_by("name"),
            "dev_groups": DevelopmentGroup.objects.filter(active=True).order_by("number"),
        })

    def post(self, request, pk):
        from departments.models import Department, DevelopmentGroup
        from core.models import period_locked
        t = get_object_or_404(Transaction, pk=pk)
        reason = self._blocked_reason(t)
        if reason:
            messages.error(request, reason)
            return redirect("transaction_list")
        lock = period_locked(t.date)
        if lock:
            messages.error(request, f"{lock} is locked. Unlock the period first.")
            return redirect("transaction_list")
        depts = request.POST.getlist("department")
        amounts = request.POST.getlist("amount")
        groups = request.POST.getlist("dev_group")
        parts = []
        for i, did in enumerate(depts):
            amt = amounts[i] if i < len(amounts) else None
            if not did or amt in (None, ""):
                continue
            dept = Department.objects.filter(pk=did).first()
            gid = groups[i] if i < len(groups) else ""
            grp = DevelopmentGroup.objects.filter(pk=gid).first() if gid else None
            parts.append((dept, amt, grp))
        try:
            out = t.split_into(parts, request.user)
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect("transaction_split", pk=pk)
        messages.success(request, f"Split into {len(out)} allocations.")
        return redirect("transaction_list")


class TransactionShiftSabbathView(TreasurerRequiredMixin, View):
    """Move a contribution to the next or previous Sabbath WITHOUT changing its real
    transaction date — used for late/after-cutoff items so a closed Sabbath is
    never altered. Audit-tracked via history."""
    def post(self, request, pk):
        from django.shortcuts import get_object_or_404
        from django.urls import reverse
        import datetime as _dt
        from core.utils import sabbath_of, sabbath_week_of
        t = get_object_or_404(Transaction, pk=pk)
        direction = request.POST.get("dir", "next")
        current = t.service_sabbath or sabbath_of(t.date)
        new = current + _dt.timedelta(days=7 if direction == "next" else -7)
        # don't allow shifting into a locked period
        try:
            from core.models import period_locked
            if period_locked(new) or period_locked(current):
                messages.error(request, "That Sabbath falls in a locked period — "
                                        "unlock it first.")
                return redirect(request.META.get("HTTP_REFERER", reverse("transaction_list")))
        except Exception:
            from core.utils import log_exception as _lx; _lx('giving/views.py')
            pass
        t.service_sabbath = new
        t.sabbath_week = sabbath_week_of(new)
        t.save(update_fields=["service_sabbath", "sabbath_week"])
        messages.success(request, f"Contribution moved to the Sabbath of {new:%d %b %Y} "
                                  f"(transaction date unchanged: {t.date:%d %b %Y}).")
        return redirect(request.META.get("HTTP_REFERER", reverse("transaction_list")))


class SabbathConfirmQueueView(DataEntryRequiredMixin, View):
    """Contributions imported after their service Sabbath had already passed — confirm
    whether each stays on that Sabbath or moves to the next one. Grouped per
    Sabbath so a whole import can be confirmed in one click."""
    template_name = "giving/sabbath_queue.html"

    def _qs(self):
        return (Transaction.objects.filter(sabbath_confirm_pending=True)
                .select_related("department", "member")
                .order_by("service_sabbath", "-date", "-id"))

    def get(self, request):
        from itertools import groupby
        qs = list(self._qs())
        groups = []
        for sab, items in groupby(qs, key=lambda t: t.service_sabbath):
            items = list(items)
            groups.append({"sabbath": sab, "items": items,
                           "count": len(items),
                           "total": sum((t.amount for t in items), Decimal(0))})
        return render(request, self.template_name, {"groups": groups,
                                                    "pending_total": len(qs)})

    def post(self, request):
        import datetime as _dt
        from core.models import next_open_sabbath, entry_blocked
        from core.utils import sabbath_week_of as _swk
        action = request.POST.get("action")
        ids = request.POST.getlist("txn")
        sab_raw = request.POST.get("sabbath")
        qs = Transaction.objects.filter(sabbath_confirm_pending=True)
        if sab_raw and not ids:        # whole-group action
            try:
                qs = qs.filter(service_sabbath=_dt.date.fromisoformat(sab_raw))
            except ValueError:
                qs = qs.none()
        elif ids:
            qs = qs.filter(id__in=ids)
        else:
            qs = qs.none()
        n = 0
        if action == "keep":
            n = qs.update(sabbath_confirm_pending=False)
            messages.success(request, f"Kept {n} contribution(s) on their original Sabbath.")
        elif action == "move":
            for t in qs:
                target = next_open_sabbath(t.service_sabbath + _dt.timedelta(days=7))
                why = entry_blocked(target)
                if why:
                    messages.error(request, f"Could not move a contribution: {why}")
                    continue
                t.service_sabbath = target
                t.sabbath_week = _swk(target)
                t.sabbath_confirm_pending = False
                t.save(update_fields=["service_sabbath", "sabbath_week",
                                      "sabbath_confirm_pending"])
                n += 1
            messages.success(request, f"Moved {n} contribution(s) to the next Sabbath.")
        return redirect("sabbath_queue")


# ===========================================================================
# Allocation-rules Excel import (item 1)
# ===========================================================================
class RuleImportView(TreasurerRequiredMixin, View):
    """Bulk-load allocation rules from a spreadsheet: reference, match type, the
    fund (or split fund) to allocate to, and optional valid-from/to dates."""
    template_name = "giving/rule_import.html"

    MATCH_LABELS = {
        "EXACT": "EXACT", "EXACTLY": "EXACT", "MATCHES EXACTLY": "EXACT", "IS": "EXACT",
        "STARTS": "STARTS", "STARTS WITH": "STARTS", "BEGINS": "STARTS", "PREFIX": "STARTS",
        "ENDS": "ENDS", "ENDS WITH": "ENDS", "SUFFIX": "ENDS",
        "CONTAINS": "CONTAINS", "INCLUDES": "CONTAINS", "HAS": "CONTAINS",
        "REGEX": "REGEX", "PATTERN": "REGEX", "MATCHES A PATTERN (REGEX)": "REGEX",
    }

    def get(self, request):
        if request.GET.get("download"):
            return self._download()
        return render(request, self.template_name, {"stage": "upload"})

    def post(self, request):
        if request.POST.get("apply"):
            return self._apply(request)
        return self._parse(request)

    def _download(self):
        import io, openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
        from django.http import HttpResponse
        from departments.models import Department
        from giving.models import SplitFund
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Rules"
        head = ["Reference", "Match type", "Fund", "Split fund", "Valid from", "Valid to"]
        ws.append(head)
        for c in range(1, len(head) + 1):
            ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
            ws.cell(1, c).fill = PatternFill("solid", fgColor="1F5F4F")
        ws.append(["tithe", "Exact", "TITHE", "", "", ""])
        ws.append(["grp12dev", "Exact", "DEVELOPMENT", "", "", ""])
        ws.append(["expense1", "Starts with", "", "Combined Offering", "", ""])
        ref = wb.create_sheet("Lists")
        ref["A1"] = "Funds"; ref["A1"].font = Font(bold=True)
        funds = list(Department.objects.filter(active=True, selectable=True).order_by("name"))
        for i, d in enumerate(funds, start=2):
            ref.cell(i, 1, d.name)
        ref["B1"] = "Match types"; ref["B1"].font = Font(bold=True)
        for i, m in enumerate(["Exact", "Starts with", "Ends with", "Contains"], start=2):
            ref.cell(i, 2, m)
        ref["C1"] = "Split funds"; ref["C1"].font = Font(bold=True)
        splits = list(SplitFund.objects.filter(active=True).order_by("name"))
        for i, s in enumerate(splits, start=2):
            ref.cell(i, 3, s.name)
        nrows = 400
        if funds:
            dv = DataValidation(type="list", formula1=f"=Lists!$A$2:$A${len(funds)+1}", allow_blank=True)
            ws.add_data_validation(dv); dv.add(f"C2:C{nrows}")
        dvm = DataValidation(type="list", formula1="=Lists!$B$2:$B$5", allow_blank=True)
        ws.add_data_validation(dvm); dvm.add(f"B2:B{nrows}")
        if splits:
            dvs = DataValidation(type="list", formula1=f"=Lists!$C$2:$C${len(splits)+1}", allow_blank=True)
            ws.add_data_validation(dvs); dvs.add(f"D2:D{nrows}")
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 20
        info = wb.create_sheet("How to fill this in")
        for i, line in enumerate([
            "Allocation rules import",
            "",
            "One row per rule. A rule sends a payment reference to a fund.",
            "  - Reference — the M-Pesa/bank reference text (e.g. tithe, grp12dev).",
            "      It is matched case- and space-insensitively.",
            "  - Match type — Exact / Starts with / Ends with / Contains. Use",
            "      'Contains' to catch variations (e.g. exp1, expense1 all contain 'exp1'?",
            "      pick the common fragment).",
            "  - Fund — the fund to allocate to (pick from the list). Leave blank if",
            "      you are using a Split fund instead.",
            "  - Split fund — if the reference should split across funds (e.g. a",
            "      combined offering), name it here and leave Fund blank.",
            "  - Valid from / Valid to — optional (YYYY-MM-DD). Leave both blank for a",
            "      permanent rule.",
            "",
            "Give either a Fund or a Split fund on each row, not both.",
            "Existing rules with the same reference are updated.",
        ], start=1):
            info.cell(i, 1, line)
        info.column_dimensions["A"].width = 76
        buf = io.BytesIO(); wb.save(buf)
        resp = HttpResponse(buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = 'attachment; filename="allocation_rules_template.xlsx"'
        return resp

    def _parse(self, request):
        import openpyxl, datetime as dt
        from departments.models import Department
        from giving.models import SplitFund
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "Choose a spreadsheet to upload.")
            return redirect("rule_import")
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception:
            from core.utils import log_exception as _lx; _lx('giving/views.py')
            messages.error(request, "Could not read that file — please upload a .xlsx.")
            return redirect("rule_import")
        ws = wb["Rules"] if "Rules" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            messages.error(request, "The sheet is empty.")
            return redirect("rule_import")
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

        def col(*names):
            for n in names:
                if n in header:
                    return header.index(n)
            return None
        c_ref = col("reference", "ref")
        c_match = col("match type", "match", "type")
        c_fund = col("fund", "department")
        c_split = col("split fund", "split")
        c_from = col("valid from", "from")
        c_to = col("valid to", "to")
        if c_ref is None:
            messages.error(request, "Couldn't find a Reference column — use the template.")
            return redirect("rule_import")

        funds = {d.name.strip().lower(): d for d in Department.objects.all()}
        splits = {s.name.strip().lower(): s for s in SplitFund.objects.all()}

        def cell(r, idx):
            if idx is None or idx >= len(r) or r[idx] in (None, ""):
                return ""
            return str(r[idx]).strip()

        def pdate(r, idx):
            if idx is None or idx >= len(r) or r[idx] in (None, ""):
                return None
            v = r[idx]
            if isinstance(v, dt.datetime):
                return v.date().isoformat()
            if isinstance(v, dt.date):
                return v.isoformat()
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
                try:
                    return dt.datetime.strptime(str(v).strip(), fmt).date().isoformat()
                except ValueError:
                    continue
            return None

        plan = []
        for r in rows[1:]:
            ref = normalize_reference(cell(r, c_ref))
            if not ref:
                continue
            match = self.MATCH_LABELS.get(cell(r, c_match).upper(), "EXACT")
            fund_raw = cell(r, c_fund)
            split_raw = cell(r, c_split)
            fund = funds.get(fund_raw.lower()) if fund_raw else None
            split = splits.get(split_raw.lower()) if split_raw else None
            plan.append({
                "reference": ref, "match_type": match,
                "fund_raw": fund_raw, "fund_id": fund.id if fund else None,
                "fund_name": fund.name if fund else None,
                "split_raw": split_raw, "split_id": split.id if split else None,
                "split_name": split.name if split else None,
                "valid_from": pdate(r, c_from), "valid_to": pdate(r, c_to),
                "ok": bool(fund or split) and not (fund and split),
            })
        if not plan:
            messages.error(request, "No rules with a reference were found.")
            return redirect("rule_import")
        request.session["rule_import_plan"] = plan
        return render(request, self.template_name, {
            "stage": "review", "plan": plan,
            "ready": sum(1 for p in plan if p["ok"]),
            "problems": sum(1 for p in plan if not p["ok"]),
        })

    @db_tx.atomic
    def _apply(self, request):
        from departments.models import Department
        from giving.models import SplitFund
        plan = request.session.get("rule_import_plan")
        if not plan:
            messages.error(request, "Your import session expired — please upload again.")
            return redirect("rule_import")
        created = updated = skipped = 0
        for p in plan:
            if not p["ok"]:
                skipped += 1
                continue
            fund = Department.objects.filter(pk=p["fund_id"]).first() if p["fund_id"] else None
            split = SplitFund.objects.filter(pk=p["split_id"]).first() if p["split_id"] else None
            if not fund and not split:
                skipped += 1
                continue
            obj, was_created = AllocationRule.objects.update_or_create(
                reference=p["reference"],
                defaults={"match_type": p["match_type"], "department": fund,
                          "split_fund": split, "source": AllocationRule.Source.SEED,
                          "valid_from": p["valid_from"], "valid_to": p["valid_to"]})
            created += 1 if was_created else 0
            updated += 0 if was_created else 1
        request.session.pop("rule_import_plan", None)
        parts = [f"{created} rule(s) created"]
        if updated:
            parts.append(f"{updated} updated")
        if skipped:
            parts.append(f"{skipped} skipped (no fund, or both fund and split set)")
        messages.success(request, ", ".join(parts) + ".")
        return redirect("rule_list")


class CashEntryDeleteView(DataEntryRequiredMixin, View):
    """Delete a loose cash entry. A cash entry IS its ledger row (one Transaction),
    so this removes the single record — there's no separate copy to fall out of
    sync. Split cash entries delete all their parts together. Edits still happen
    at the ledger. Guarded: only manual CASH credits that aren't reconciled,
    reversed, receipted via an envelope, or in a locked period."""

    def post(self, request, pk):
        txn = get_object_or_404(Transaction, pk=pk)
        # only loose cash credits may be deleted here
        if txn.channel != Transaction.Channel.CASH or \
           txn.direction != Transaction.Direction.CREDIT:
            messages.error(request, "Only cash collections can be deleted here. "
                           "Use the ledger for other entries.")
            return redirect("cash_list")
        if _block_if_locked(request, txn.date):
            return redirect("cash_list")
        # don't delete something tied to other records or already reversed
        blockers = []
        if txn.is_reversed or txn.is_reversal:
            blockers.append("it has already been reversed")
        if getattr(txn, "processed_via_envelope", False):
            blockers.append("it was receipted via an envelope")
        if txn.envelope_lines.exists():
            blockers.append("it is linked to an envelope")
        if blockers:
            messages.error(request, "Can't delete this cash entry because "
                           + " and ".join(blockers) + ". Reverse it at the ledger instead.")
            return redirect("cash_list")
        siblings = list(txn.split_siblings()) + [txn]
        n = len(siblings)
        with db_tx.atomic():
            for s in siblings:
                s.delete()
        messages.success(request,
            f"Cash entry deleted{f' ({n} split parts)' if n > 1 else ''}. "
            "The ledger row was removed with it.")
        return redirect("cash_list")


# --- Campaign fallback allocation -------------------------------------------
class CampaignListView(ReadAccessMixin, View):
    """Manage campaigns (e.g. Camp Meeting): their fund, trigger words and the
    member→group table used as a fallback when the normal rules miss."""
    template_name = "giving/campaign_list.html"

    def get(self, request):
        from giving.models import Campaign
        from departments.models import Department
        from django.db.models import Count
        camps = (Campaign.objects.select_related("department")
                 .annotate(n_members=Count("members", distinct=True),
                           n_txns=Count("transactions", distinct=True))
                 .order_by("-active", "name"))
        return render(request, self.template_name, {
            "campaigns": camps,
            "funds": Department.objects.filter(active=True, selectable=True).order_by("name"),
        })


class CampaignCreateView(DataEntryRequiredMixin, View):
    def post(self, request):
        from giving.models import Campaign
        from departments.models import Department
        name = (request.POST.get("name") or "").strip()
        dept = Department.objects.filter(pk=request.POST.get("department")).first()
        triggers = (request.POST.get("triggers") or "").strip()
        if not name or not dept:
            messages.error(request, "A campaign needs a name and a fund.")
            return redirect("campaign_list")
        camp, created = Campaign.objects.get_or_create(
            name=name, defaults={"department": dept, "triggers": triggers})
        if not created:
            camp.department = dept
            camp.triggers = triggers
            camp.active = True
            camp.save()
        messages.success(request, f"Campaign “{name}” saved. Now upload its members.")
        return redirect("campaign_list")


class CampaignMemberImportView(DataEntryRequiredMixin, View):
    """Upload the Name / Mobile / Group sheet for a campaign. Reads .xlsx or .csv
    tolerantly, skips unusable rows and reports what happened — a bad row never
    aborts the whole upload."""
    @staticmethod
    def _phone_cell(v):
        # a numeric phone cell arrives as int/float (e.g. 254791896792.0)
        if v is None:
            return ""
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    def _read(self, f):
        name = (getattr(f, "name", "") or "").lower()
        rows = []
        if name.endswith((".xlsx", ".xls")):
            import openpyxl
            ws = openpyxl.load_workbook(f, data_only=True).active
            data = list(ws.iter_rows(values_only=True))
            if not data:
                return rows
            hdr = [str(c).strip().lower() if c else "" for c in data[0]]
            ni = next((i for i, h in enumerate(hdr) if "name" in h), 0)
            pi = next((i for i, h in enumerate(hdr) if h in ("mobile", "phone", "msisdn")), 1)
            gi = next((i for i, h in enumerate(hdr) if "group" in h), 2)
            for r in data[1:]:
                nm = str(r[ni]).strip() if ni < len(r) and r[ni] not in (None, "") else ""
                rows.append((nm, self._phone_cell(r[pi] if pi < len(r) else None),
                             str(r[gi]).strip() if gi < len(r) and r[gi] else ""))
        else:
            import csv as _csv, io as _io
            for raw in _csv.DictReader(_io.TextIOWrapper(f.file, encoding="utf-8-sig")):
                row = {(k or "").strip().lower(): v for k, v in raw.items()}
                rows.append(((row.get("name") or "").strip(),
                             (row.get("mobile") or row.get("phone") or "").strip(),
                             (row.get("group") or "").strip()))
        return rows

    def post(self, request, pk):
        from giving.models import Campaign, CampaignMember
        camp = get_object_or_404(Campaign, pk=pk)
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "Choose a .xlsx or .csv with Name, Mobile, Group columns.")
            return redirect("campaign_list")
        try:
            rows = self._read(f)
        except Exception:
            from core.utils import log_exception as _lx; _lx('giving/views.py')
            messages.error(request, "Could not read that file — use the sample layout "
                                    "(Name, Mobile, Group). Try downloading the sample.")
            return redirect("campaign_list")
        if not rows:
            messages.warning(request, "That file had no data rows.")
            return redirect("campaign_list")
        camp.members.all().delete()          # replace the table
        made = skipped = 0
        for nm, ph, grp in rows:
            if not nm:
                skipped += 1
                continue
            try:
                CampaignMember.objects.create(campaign=camp, name=nm, phone=ph, group=grp)
                made += 1
            except Exception:
                from core.utils import log_exception as _lx; _lx('giving/views.py')
                skipped += 1
        msg = f"Loaded {made} members into “{camp.name}”."
        if skipped:
            msg += f" {skipped} row(s) skipped (no name or unreadable)."
        (messages.success if made else messages.warning)(request, msg)
        return redirect("campaign_list")


class CampaignTemplateView(ReadAccessMixin, View):
    """Download a sample member-upload file (CSV)."""
    def get(self, request):
        import csv as _csv
        from django.http import HttpResponse
        resp = HttpResponse(content_type="text/csv")
        resp["Content-Disposition"] = 'attachment; filename="campaign_members_sample.csv"'
        w = _csv.writer(resp)
        w.writerow(["Name", "Mobile", "Group"])
        w.writerow(["Amos Ndegwa", "254791896792", "CAMP_1"])
        w.writerow(["Calvince Ouma", "0726410608", "CAMP_1"])
        w.writerow(["Caroline Nyalick", "254705321239", "CAMP_2"])
        return resp


class CampaignDeleteView(TreasurerRequiredMixin, View):
    """Delete a finished campaign: its member table goes; rows it allocated keep
    their group tag (campaign link is set null)."""
    def post(self, request, pk):
        from giving.models import Campaign
        camp = get_object_or_404(Campaign, pk=pk)
        nm = camp.name
        camp.delete()
        messages.success(request, f"Campaign “{nm}” deleted. Past allocations keep their group tag.")
        return redirect("campaign_list")


class TransactionBulkReverseView(TreasurerRequiredMixin, View):
    """Bulk actions on several selected ledger entries at once: reverse them,
    or send them back to the review queue for re-allocation. Treasury never
    hard-deletes — each becomes a contra posting (and a linked envelope
    receipt is removed, its siblings reversed). Locked-period rows and ones
    already reversed (or that are themselves reversals) are skipped and
    counted."""
    def post(self, request):
        ids = request.POST.getlist("ids")
        if not ids:
            messages.info(request, "No entries were selected.")
            return redirect(request.META.get("HTTP_REFERER") or "transaction_list")
        if request.POST.get("action") == "send_to_review":
            return self._send_to_review(request, ids)
        return self._reverse(request, ids)

    def _reverse(self, request, ids):
        from core.models import period_locked
        reason = (request.POST.get("reason") or "").strip()
        done = skipped = 0
        for t in Transaction.objects.filter(pk__in=ids):
            if t.is_reversed or t.is_reversal or period_locked(t.date):
                skipped += 1
                continue
            try:
                t.reverse(request.user, reason=reason)
            except ValueError:
                skipped += 1
                continue
            TransactionReverseView._delete_linked_envelope(t, request.user)
            done += 1
        msg = f"{done} entr{'y' if done == 1 else 'ies'} reversed (contra postings kept for audit)."
        if skipped:
            msg += f" {skipped} skipped (locked period, or already reversed)."
        (messages.success if done else messages.info)(request, msg)
        return redirect(request.META.get("HTTP_REFERER") or "transaction_list")

    def _send_to_review(self, request, ids):
        """Groups selected entries by split family so ticking both siblings
        of the same split contribution still produces exactly one combined
        replacement, not two — the same underlying logic as
        TransactionSendToReviewView, applied to a multi-select batch."""
        from core.models import period_locked
        reason = (request.POST.get("reason") or "").strip()
        seen_ids = set()
        groups_done = entries_done = skipped = 0
        total_combined = Decimal(0)
        for t in Transaction.objects.filter(pk__in=ids).order_by("id"):
            if t.pk in seen_ids:
                continue
            if t.is_reversed or t.is_reversal or t.allocation_status == Transaction.Status.REVIEW \
                    or not t.department_id:
                skipped += 1
                seen_ids.add(t.pk)
                continue
            group = [t] + list(t.strict_split_siblings())
            reversible = [m for m in group if not m.is_reversed and not m.is_reversal
                         and not period_locked(m.date)]
            if not reversible:
                skipped += 1
                seen_ids.update(m.pk for m in group)
                continue
            group_total = Decimal(0)
            for member in reversible:
                group_total += member.amount
                member.reverse(request.user,
                              reason=reason or "Sent back to review for re-allocation")
                TransactionReverseView._delete_linked_envelope(member, request.user)
                entries_done += 1
            Transaction.objects.create(
                date=t.date, sabbath_week=t.sabbath_week, channel=t.channel,
                direction=t.direction, amount=group_total, member=t.member,
                reference=t.reference, payer_name=t.payer_name, payer_phone=t.payer_phone,
                mpesa_ref=t.mpesa_ref, confirmed=t.confirmed,
                allocation_status=Transaction.Status.REVIEW,
                raw_narration=(f"Replaces #{', #'.join(str(m.pk) for m in reversible)} — "
                              f"sent back to review for correct allocation"
                              + (f': "{reason}"' if reason else "")))
            total_combined += group_total
            groups_done += 1
            seen_ids.update(m.pk for m in group)

        if groups_done:
            messages.success(request,
                f"{entries_done} entr{'y' if entries_done == 1 else 'ies'} reversed and "
                f"combined into {groups_done} new entr{'y' if groups_done == 1 else 'ies'} "
                f"(total {total_combined:,.2f}) in the review queue."
                + (f" {skipped} skipped." if skipped else ""))
            return redirect("queue")
        messages.info(request, "Nothing could be sent back to review (already in "
                              "review, already reversed, or the period is locked).")
        return redirect(request.META.get("HTTP_REFERER") or "transaction_list")


class RuleEditView(DataEntryRequiredMixin, View):
    """Edit an allocation rule (reference / match type / target fund)."""
    template_name = "giving/rule_form.html"

    def get(self, request, pk):
        rule = get_object_or_404(AllocationRule, pk=pk)
        return render(request, self.template_name,
                      {"form": RuleForm(instance=rule), "edit_obj": rule})

    def post(self, request, pk):
        rule = get_object_or_404(AllocationRule, pk=pk)
        form = RuleForm(request.POST, instance=rule)
        if form.is_valid():
            form.instance.reference = normalize_reference(form.instance.reference)
            form.save()
            messages.success(request, "Rule updated.")
            return redirect("rule_list")
        return render(request, self.template_name, {"form": form, "edit_obj": rule})


# ===================== Development-group reference patterns (#8) =============
class DevPatternForm(forms.ModelForm):
    class Meta:
        from giving.models import DevGroupPattern
        model = DevGroupPattern
        fields = ["label", "pattern", "kind", "enabled", "sort_order", "note"]
        widgets = {"kind": forms.Select(attrs={"class": "field--select"})}


class DevPatternListView(TreasurerRequiredMixin, View):
    """Manage the regexes that recognise development-group references, with a
    live tester so you can see what a sample reference would match."""
    template_name = "giving/dev_patterns.html"

    def get(self, request):
        from giving.models import DevGroupPattern
        from giving.services.allocation import normalize_reference, detect_dev_group
        patterns = DevGroupPattern.objects.all()
        sample = request.GET.get("test", "")
        result = None
        if sample:
            norm = normalize_reference(sample)
            hit = detect_dev_group(norm)
            if hit and hit[0] == "NUMBER":
                result = {"ok": True, "msg": f"Matches → development group {hit[1]}.",
                          "norm": norm}
            elif hit and hit[0] == "WORD":
                result = {"ok": True, "msg": "Matches as development (no number) → "
                          "booked to development, awaiting a group.", "norm": norm}
            else:
                result = {"ok": False, "msg": "No development pattern matches this "
                          "reference.", "norm": norm}
        return render(request, self.template_name, {
            "patterns": patterns, "form": DevPatternForm(),
            "sample": sample, "result": result})

    def post(self, request):
        from giving.models import DevGroupPattern
        pk = request.POST.get("id")
        instance = DevGroupPattern.objects.filter(pk=pk).first() if pk else None
        form = DevPatternForm(request.POST, instance=instance)
        if form.is_valid():
            obj = form.save(commit=False)
            if not obj.created_by_id:
                obj.created_by = request.user
            try:
                obj.full_clean()
            except Exception as exc:  # noqa: BLE001
                from django.core.exceptions import ValidationError
                msg = "; ".join(m for v in exc.message_dict.values() for m in v) \
                    if isinstance(exc, ValidationError) else str(exc)
                messages.error(request, msg)
                return redirect("dev_patterns")
            obj.save()
            messages.success(request, "Pattern saved." if instance else "Pattern added.")
        else:
            messages.error(request, "; ".join(
                f"{f}: {', '.join(e)}" for f, e in form.errors.items()))
        return redirect("dev_patterns")


class DevPatternToggle(TreasurerRequiredMixin, View):
    def post(self, request, pk):
        from giving.models import DevGroupPattern
        p = get_object_or_404(DevGroupPattern, pk=pk)
        p.enabled = not p.enabled
        p.save()
        messages.success(request, f"{p.label} {'enabled' if p.enabled else 'disabled'}.")
        return redirect("dev_patterns")


class DevPatternDelete(TreasurerRequiredMixin, View):
    def post(self, request, pk):
        from giving.models import DevGroupPattern
        p = get_object_or_404(DevGroupPattern, pk=pk)
        p.delete()
        messages.success(request, "Pattern removed.")
        return redirect("dev_patterns")


# ===================== Allocation-rule lifecycle (#9) =======================
class RuleArchiveView(TreasurerRequiredMixin, View):
    """Archive a single rule (soft) or, with ?action=restore, bring it back."""
    def post(self, request, pk):
        rule = get_object_or_404(AllocationRule, pk=pk)
        if request.POST.get("action") == "restore":
            rule.restore()
            messages.success(request, "Rule restored — it can allocate giving again.")
        else:
            rule.archive()
            messages.success(request, "Rule archived — kept for the audit trail but "
                "no longer used for new giving.")
        return redirect(f"{reverse_lazy('rule_list')}?view="
                        f"{request.POST.get('view', 'active')}")


class RuleArchiveExpiredView(TreasurerRequiredMixin, View):
    """Bulk-archive every rule whose validity window has already ended."""
    def post(self, request):
        import datetime as _d
        qs = AllocationRule.objects.filter(archived=False, valid_to__lt=_d.date.today())
        n = 0
        for r in qs:
            r.archive(); n += 1
        messages.success(request, f"Archived {n} expired rule{'s' if n != 1 else ''}.")
        return redirect("rule_list")


class AllocationSettingsView(TreasurerRequiredMixin, View):
    """Allocation & categories — its own page.

    Moved out of Settings → Channels, where it sat as a card among bank
    accounts and opening balances that have nothing to do with it. Allocation is
    a *rules* concern: it belongs next to the allocation rules and the
    development-group patterns a treasurer manages, not buried in a settings tab
    they open once a year.

    The old dev-group "extra prefixes" field is gone from here. It built exactly
    the regex a DevGroupPattern of kind NUMBERED builds, but could not be
    labelled, ordered, disabled or audited — two places to configure one
    behaviour, neither able to see the other. Migration `giving.0025` turned
    whatever any church had configured into real patterns on the
    Development-group patterns page, which is where such things now live.
    """
    template_name = "giving/allocation_settings.html"

    def get(self, request):
        from core.forms import SiteConfigForm
        from core.models import SiteConfig
        return render(request, self.template_name,
                      {"form": SiteConfigForm(instance=SiteConfig.get())})

    def post(self, request):
        from core.forms import SiteConfigForm
        from core.models import SiteConfig
        cfg = SiteConfig.get()
        form = SiteConfigForm(request.POST, request.FILES, instance=cfg)
        if form.is_valid():
            form.save()
            messages.success(request, "Allocation settings saved.")
            return redirect("allocation_settings")
        messages.error(request, "Check the form.")
        return render(request, self.template_name, {"form": form})


def _sms_enabled():
    from core.models import SiteConfig
    return SiteConfig.get().sms_enabled


class CampaignDetailView(ReadAccessMixin, View):
    """One campaign: the sheet that was uploaded, and its groups.

    The list page could say how many members a campaign had but not who they
    were, so an uploaded sheet was effectively write-only — a treasurer could
    not check that the groups had come through correctly, which is the first
    thing anyone wants to do after an import.
    """
    template_name = "giving/campaign_detail.html"

    def get(self, request, pk):
        from django.db.models import Count, Sum
        from giving.models import Campaign
        from giving.services import campaign_sms

        campaign = get_object_or_404(
            Campaign.objects.select_related("department"), pk=pk)
        progress = campaign_sms.group_progress(campaign)
        groups = progress
        behind = [r for r in progress if r["behind"]]
        met = [r for r in progress if r["has_target"] and not r["behind"]]
        # What has already gone to each group, so a treasurer can see it before
        # composing another one.
        history = {g["name"]: campaign_sms.recent_sends(campaign, g["name"], limit=3)
                   for g in groups}
        for g in groups:
            g["history"] = history.get(g["name"], [])
        txns = campaign.transactions.all()
        return render(request, self.template_name, {
            "campaign": campaign,
            "groups": groups,
            "total_members": sum(g["count"] for g in groups),
            "total_reachable": sum(g["reachable"] for g in groups),
            "n_txns": txns.count(),
            "raised": txns.aggregate(t=Sum("amount"))["t"] or 0,
            "placeholders": campaign_sms.PLACEHOLDERS,
            "sms_enabled": _sms_enabled(),
            # Writing to the whole campaign at once. Worth its own box rather
            # than a checkbox on a group's: it is a different-sized action, and
            # the count beside the button is what makes that plain before the
            # press rather than after it.
            "all_groups_token": campaign_sms.ALL_GROUPS,
            "all_groups_history": campaign_sms.recent_sends(
                campaign, campaign_sms.ALL_GROUPS, limit=3),
            # Chasing the groups still short of the target set on the fund's
            # budget page. Computed here so the button can say how many groups
            # and how much — a "remind the ones behind" button that does not
            # say who is behind is asking for a blind press.
            "behind_token": campaign_sms.BEHIND_TARGET,
            "progress": progress,
            "behind": behind,
            "behind_short": sum((r["short"] for r in behind), 0),
            "behind_reachable": sum(
                (r["reachable"] for r in behind), 0),
            "any_targets": any(r["has_target"] for r in progress),
            "behind_history": campaign_sms.recent_sends(
                campaign, campaign_sms.BEHIND_TARGET, limit=3),
            # And the other half: the groups that got there. Same shape as the
            # chase above, because thanking people who finished is the same
            # sized action as chasing people who have not — and a church whose
            # members only hear from the treasurer when they are short is
            # teaching them what a message from the treasurer means.
            "met_token": campaign_sms.TARGET_MET,
            "met": met,
            "met_raised": sum((r["collected"] for r in met), 0),
            "met_reachable": sum((r["reachable"] for r in met), 0),
            "met_history": campaign_sms.recent_sends(
                campaign, campaign_sms.TARGET_MET, limit=3),
        })


class CampaignGroupSmsView(TreasurerRequiredMixin, View):
    """Write to one group.

    Treasurer-level on purpose. This is the only action in the application that
    costs money on every press and cannot be recalled, so it sits with the role
    that answers for spending rather than with general data entry.

    Two steps, always: compose then confirm. The confirmation screen is built
    from the same resolution the send uses, so the number on the button is the
    number of messages that will leave.
    """
    template_name = "giving/campaign_sms_confirm.html"

    def post(self, request, pk):
        from giving.models import Campaign
        from giving.services import campaign_sms

        campaign = get_object_or_404(Campaign, pk=pk)
        group = (request.POST.get("group") or "").strip()
        template = (request.POST.get("message") or "").strip()

        if not template:
            messages.error(request, "Write the message first.")
            return redirect("campaign_detail", pk=pk)

        if request.POST.get("confirm") != "yes":
            plan = campaign_sms.preview(campaign, group, template)
            if not plan["count"]:
                messages.error(
                    request,
                    "Nobody on this campaign has a usable phone number, so "
                    "there is nothing to send."
                    if group == campaign_sms.ALL_GROUPS else
                    "Nobody in that group has a usable phone number, so there "
                    "is nothing to send.")
                return redirect("campaign_detail", pk=pk)
            return render(request, self.template_name, {
                "campaign": campaign, "group": group, "message": template,
                "group_label": campaign_sms.group_label(group),
                # "spans more than one group", not "is the all-groups send" —
                # chasing the groups behind target is equally a multi-group
                # send, and the confirmation screen has the same job for both:
                # show which groups, and show each one's own figures.
                "all_groups": group in (campaign_sms.ALL_GROUPS,
                                        campaign_sms.BEHIND_TARGET),
                "behind_send": group == campaign_sms.BEHIND_TARGET,
                "group_breakdown": campaign_sms.breakdown(plan),
                "gap_members": campaign_sms.gap_warning(plan, template),
                "plan": plan,
                # Not a block — a church may legitimately repeat a reminder —
                # but the person about to press send should know they are
                # repeating it.
                "duplicate": campaign_sms.already_sent(campaign, group, template),
                "recent": campaign_sms.recent_sends(campaign, group, limit=3),
                "sms_enabled": _sms_enabled(),
            })

        result = campaign_sms.send(campaign, group, template, user=request.user)
        parts = [f"{result['sent']} message(s) sent to "
                 f"{campaign_sms.group_label(group)}"]
        if result["failed"]:
            parts.append(f"{result['failed']} failed")
        if result["skipped"]:
            parts.append(f"{result['skipped']} had no phone number")
        messages.success(request, ", ".join(parts) + ".")
        return redirect("campaign_detail", pk=pk)
