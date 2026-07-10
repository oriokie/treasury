from decimal import Decimal

from django.test import TestCase

from departments.models import Department, DevelopmentGroup
from giving.models import AllocationRule, Transaction
from giving.services.allocation import normalize_reference, allocate


class NormalizeReferenceTests(TestCase):
    def test_lowercases_and_strips_spaces(self):
        self.assertEqual(normalize_reference("  Sabbath School "), "sabbathschool")

    def test_handles_none(self):
        self.assertEqual(normalize_reference(None), "")


class AllocationEngineTests(TestCase):
    def setUp(self):
        self.tithe = Department.objects.create(
            name="Tithe", fund_type=Department.FundType.TRUST)
        AllocationRule.objects.create(
            reference="tithe", department=self.tithe,
            source=AllocationRule.Source.SEED)

    def test_fund_type_sets_is_trust(self):
        self.assertTrue(self.tithe.is_trust)

    def test_seeded_rule_resolves_auto(self):
        dept, status = allocate("tithe")
        self.assertEqual(dept, self.tithe)
        self.assertEqual(status, "AUTO")

    def test_unknown_reference_goes_to_review(self):
        resolver, status = allocate("something-unmapped")
        self.assertEqual(resolver, "UNALLOCATED")
        self.assertEqual(status, "REVIEW")

    def test_dev_group_reference_autodetected(self):
        resolver, status = allocate("Grp12dev")
        self.assertEqual(resolver, "DEV_GROUP_12")
        self.assertEqual(status, "AUTO")

    def test_messy_dev_references_resolve(self):
        # spellings seen in the real statement
        for ref, n in [("DEVGR7", 7), ("devg14", 14), ("DEVLOP GP14", 14),
                       ("dev grp5", 5), ("DEv Gp39", 39), ("DEVGRP3*", 3),
                       (" DEVGR26", 26), ("Devgrp11", 11)]:
            resolver, status = allocate(ref)
            self.assertEqual(resolver, f"DEV_GROUP_{n}", ref)
            self.assertEqual(status, "AUTO", ref)

    def test_dev_without_number_is_na(self):
        resolver, status = allocate("dev")
        self.assertEqual(resolver, "DEV_GROUP_NA")

    def test_resolver_token_maps_to_fund_and_group(self):
        from statements.services.importer import _resolve
        resolver, _ = allocate("Grp12dev")
        dept, grp = _resolve(resolver)
        self.assertEqual(dept.name.upper(), "DEVELOPMENT")
        self.assertEqual(grp.number, 12)
        self.assertTrue(DevelopmentGroup.objects.filter(number=12).exists())


class TransactionDedupTests(TestCase):
    def test_core_ref_is_unique(self):
        dept = Department.objects.create(name="Combined", fund_type=Department.FundType.LOCAL)
        Transaction.objects.create(
            date="2026-06-06", channel=Transaction.Channel.BANK,
            direction=Transaction.Direction.CREDIT, amount=Decimal("100"),
            department=dept, core_ref="UNIQUE123", allocation_status="AUTO")
        from django.db import IntegrityError, transaction
        with self.assertRaises(IntegrityError):
            with transaction.atomic():
                Transaction.objects.create(
                    date="2026-06-06", channel=Transaction.Channel.BANK,
                    direction=Transaction.Direction.CREDIT, amount=Decimal("100"),
                    department=dept, core_ref="UNIQUE123", allocation_status="AUTO")


class SplitFundTests(TestCase):
    def setUp(self):
        from giving.models import SplitFund, SplitComponent
        trust = Department.objects.create(name="Combined Trust",
                                          fund_type=Department.FundType.TRUST)
        local = Department.objects.create(name="Combined Local",
                                          fund_type=Department.FundType.LOCAL)
        self.sf = SplitFund.objects.create(name="Combined Offering")
        SplitComponent.objects.create(split_fund=self.sf, department=trust, percent=50)
        SplitComponent.objects.create(split_fund=self.sf, department=local, percent=50)

    def test_split_sums_exactly_with_rounding(self):
        parts = self.sf.split(Decimal("1001"))
        self.assertEqual(sum(a for _, a in parts), Decimal("1001"))
        self.assertEqual(parts[0][1], Decimal("500.50"))

    def test_allocate_returns_split_fund(self):
        from giving.models import AllocationRule
        AllocationRule.objects.create(reference="combined", split_fund=self.sf,
                                      source=AllocationRule.Source.SEED)
        resolver, status = allocate("combined")
        self.assertEqual(resolver, self.sf)
        self.assertEqual(status, "AUTO")


class DebitMultiMatchTests(TestCase):
    def setUp(self):
        import datetime as dt
        from decimal import Decimal
        from django.contrib.auth.models import User
        from departments.models import Department
        from giving.models import Transaction
        from cashbook.models import Expense
        self.u = User.objects.create_superuser("dm", password="x")
        self.dept = Department.objects.create(name="Tithe", fund_type=Department.FundType.TRUST)
        self.e1 = Expense.objects.create(date=dt.date(2026, 5, 10), department=self.dept,
            description="A", amount=Decimal("300"), status=Expense.Status.APPROVED, recorded_by=self.u)
        self.e2 = Expense.objects.create(date=dt.date(2026, 5, 10), department=self.dept,
            description="B", amount=Decimal("700"), status=Expense.Status.APPROVED, recorded_by=self.u)
        self.deb = Transaction.objects.create(date=dt.date(2026, 5, 10), channel="BANK",
            direction="DEBIT", amount=Decimal("1000"), allocation_status="REVIEW")
        self.client.login(username="dm", password="x")

    def test_multi_match_requires_equal_total(self):
        from django.urls import reverse
        # mismatch is rejected
        self.client.post(reverse("debit_resolve", args=[self.deb.id]),
                         {"kind": "match", "expense": [self.e1.id]})
        self.e1.refresh_from_db()
        self.assertIsNone(self.e1.bank_transaction_id)
        # matching total links both
        self.client.post(reverse("debit_resolve", args=[self.deb.id]),
                         {"kind": "match", "expense": [self.e1.id, self.e2.id]})
        self.e1.refresh_from_db(); self.e2.refresh_from_db()
        self.assertEqual(self.e1.bank_transaction_id, self.deb.id)
        self.assertEqual(self.e2.bank_transaction_id, self.deb.id)


class QueueImportTests(TestCase):
    def test_import_allocates(self):
        import datetime as dt
        from decimal import Decimal
        from django.contrib.auth.models import User
        from django.urls import reverse
        from django.core.files.uploadedfile import SimpleUploadedFile
        from departments.models import Department
        from giving.models import Transaction
        User.objects.create_superuser("qi", password="x")
        d = Department.objects.create(name="Tithe", fund_type=Department.FundType.TRUST)
        t = Transaction.objects.create(date=dt.date(2026, 5, 1), channel="BANK",
            direction="CREDIT", amount=Decimal("100"), allocation_status="REVIEW")
        self.client.login(username="qi", password="x")
        csv = f"id,date,amount,payer_name,payer_phone,reference,narration,allocate_to_fund\n{t.id},2026-05-01,100,X,,r,n,Tithe\n"
        self.client.post(reverse("queue_import"),
                         {"file": SimpleUploadedFile("q.csv", csv.encode())})
        t.refresh_from_db()
        self.assertEqual(t.allocation_status, "MANUAL")
        self.assertEqual(t.department_id, d.id)


class RuleMatchTypeTests(TestCase):
    def setUp(self):
        from departments.models import Department
        from giving.models import AllocationRule
        self.d = Department.objects.create(name="YOUTH", fund_type=Department.FundType.LOCAL)
        AllocationRule.objects.create(reference="potluck", department=self.d,
                                      match_type="CONTAINS", source="SEED")
        AllocationRule.objects.create(reference="yth", department=self.d,
                                      match_type="STARTS", source="SEED")

    def test_contains_and_starts(self):
        from giving.services.allocation import allocate
        self.assertEqual(allocate("abc potluck 99")[0], self.d)
        self.assertEqual(allocate("yth camp")[0], self.d)

    def test_exact_still_wins_review_otherwise(self):
        from giving.services.allocation import allocate
        target, status = allocate("totally-unknown-xyz")
        self.assertEqual(target, "UNALLOCATED")


class DeleteEntryTests(TestCase):
    def setUp(self):
        import datetime as dt
        from decimal import Decimal
        from django.contrib.auth.models import User, Group
        from departments.models import Department
        from giving.models import Transaction
        self.tre = User.objects.create_superuser("tre", password="x")
        self.asst = User.objects.create_user("asst", password="x")
        Group.objects.get_or_create(name="Assistant")[0].user_set.add(self.asst)
        d = Department.objects.create(name="Tithe", fund_type=Department.FundType.TRUST)
        self.t = Transaction.objects.create(date=dt.date(2026, 5, 1), channel="CASH",
            direction="CREDIT", amount=Decimal("100"), department=d, allocation_status="MANUAL")

    def test_treasurer_can_reverse(self):
        from django.urls import reverse
        from giving.models import Transaction
        self.client.login(username="tre", password="x")
        self.client.post(reverse("transaction_reverse", args=[self.t.id]))
        self.t.refresh_from_db()
        self.assertTrue(self.t.is_reversed)
        self.assertTrue(Transaction.objects.filter(id=self.t.id).exists())  # not deleted

    def test_assistant_cannot_reverse(self):
        from django.urls import reverse
        self.client.login(username="asst", password="x")
        self.client.post(reverse("transaction_reverse", args=[self.t.id]))
        self.t.refresh_from_db()
        self.assertFalse(self.t.is_reversed)


class ReversalTests(TestCase):
    def setUp(self):
        import datetime as dt
        from decimal import Decimal
        from django.contrib.auth.models import User
        from departments.models import Department
        from giving.models import Transaction
        self.u = User.objects.create_superuser("rv", password="x")
        d = Department.objects.create(name="Choir", fund_type=Department.FundType.LOCAL)
        self.t = Transaction.objects.create(date=dt.date(2026, 5, 1), channel="CASH",
            direction="CREDIT", amount=Decimal("300"), department=d, allocation_status="MANUAL")
        self.client.login(username="rv", password="x")

    def test_reverse_creates_contra_and_keeps_original(self):
        from django.urls import reverse
        from giving.models import Transaction, TransactionReversal
        self.client.post(reverse("transaction_reverse", args=[self.t.id]), {"reason": "dup"})
        self.t.refresh_from_db()
        self.assertTrue(self.t.is_reversed)
        self.assertTrue(Transaction.objects.filter(id=self.t.id).exists())  # never deleted
        contra = Transaction.objects.get(is_reversal=True)
        self.assertEqual(contra.amount, self.t.amount)   # positive (validator-safe)
        self.assertGreater(contra.amount, 0)
        self.assertTrue(TransactionReversal.objects.filter(original=self.t).exists())

    def test_cannot_reverse_twice(self):
        self.t.reverse(self.u)
        with self.assertRaises(ValueError):
            self.t.reverse(self.u)


class PeriodLockTests(TestCase):
    def test_lock_blocks_non_admin_create(self):
        import datetime as dt
        from django.contrib.auth.models import User, Group
        from django.urls import reverse
        from departments.models import Department
        from cashbook.models import Expense
        from core.models import PeriodLock
        admin = User.objects.create_superuser("adm", password="x")
        clerk = User.objects.create_user("clk", password="x")
        Group.objects.get_or_create(name="Assistant")[0].user_set.add(clerk)
        d = Department.objects.create(name="Choir", fund_type=Department.FundType.LOCAL)
        PeriodLock.objects.create(year=2026, month=5, locked_by=admin)
        self.client.login(username="clk", password="x")
        n = Expense.objects.count()
        self.client.post(reverse("expense_create"), {"date": "2026-05-10",
            "department": d.id, "description": "x", "amount": "20",
            "category": "OTHER", "claimant": "A", "method": "CASH", "voucher_no": ""})
        self.assertEqual(Expense.objects.count(), n)  # blocked


class TransactionDateFilterTests(TestCase):
    def setUp(self):
        import datetime as dt
        from django.contrib.auth.models import User
        from departments.models import Department
        from giving.models import Transaction
        self.u = User.objects.create_superuser("tf", password="x")
        self.fund = Department.objects.create(name="LCB", fund_type=Department.FundType.LOCAL)
        for d in [dt.date(2026, 1, 5), dt.date(2026, 3, 5), dt.date(2026, 6, 5)]:
            Transaction.objects.create(date=d, channel="CASH", direction="CREDIT",
                amount=100, department=self.fund, allocation_status="AUTO")
        self.client.login(username="tf", password="x")

    def test_date_range_filters(self):
        from django.urls import reverse
        url = reverse("transaction_list")
        r = self.client.get(url, {"date_from": "2026-02-01", "date_to": "2026-04-30"})
        self.assertEqual(r.status_code, 200)
        self.assertEqual(len(r.context["transactions"]), 1)  # only the March one
        r2 = self.client.get(url, {"date_from": "2026-01-01"})
        self.assertEqual(len(r2.context["transactions"]), 3)


class CashDuplicateTests(TestCase):
    def setUp(self):
        from django.contrib.auth.models import User
        from departments.models import Department
        self.u = User.objects.create_superuser("cd", password="x")
        self.fund = Department.objects.create(name="LCB", fund_type=Department.FundType.LOCAL)
        self.client.login(username="cd", password="x")

    def test_duplicate_requires_confirmation(self):
        import datetime as dt
        from decimal import Decimal
        from django.urls import reverse
        from giving.models import Transaction
        Transaction.objects.create(date=dt.date(2026, 5, 2), channel="CASH",
            direction="CREDIT", amount=Decimal("2000"), department=self.fund,
            allocation_status="MANUAL", payer_name="Mary Wanjiku")
        # second entry with same amount AND a similar name -> blocked without confirm
        self.client.post(reverse("cash_new"), {
            "date": "2026-05-02", "department": self.fund.id, "amount": "2000",
            "channel": "CASH", "reference": "", "payer_name": "Wanjiku Mary"})
        self.assertEqual(Transaction.objects.filter(amount=Decimal("2000")).count(), 1)
        # with confirmation -> saved
        self.client.post(reverse("cash_new"), {
            "date": "2026-05-02", "department": self.fund.id, "amount": "2000",
            "channel": "CASH", "reference": "", "payer_name": "Wanjiku Mary",
            "confirm_duplicate": "1"})
        self.assertEqual(Transaction.objects.filter(amount=Decimal("2000")).count(), 2)

    def test_negative_amount_rejected_by_model(self):
        import datetime as dt
        from decimal import Decimal
        from django.core.exceptions import ValidationError
        from giving.models import Transaction
        t = Transaction(date=dt.date(2026, 5, 2), channel="CASH", direction="CREDIT",
                        amount=Decimal("-5"), department=self.fund, allocation_status="MANUAL")
        with self.assertRaises(ValidationError):
            t.full_clean()


class ReviewVsDebitQueueTests(TestCase):
    def setUp(self):
        import datetime as dt
        from decimal import Decimal
        from django.contrib.auth.models import User
        from departments.models import Department
        from giving.models import Transaction
        self.u = User.objects.create_superuser("rq", password="x")
        self.fund = Department.objects.create(name="LCB", fund_type=Department.FundType.LOCAL)
        # a credit awaiting allocation -> review queue
        self.credit = Transaction.objects.create(date=dt.date(2026, 5, 2),
            channel="BANK", direction="CREDIT", amount=Decimal("500"),
            allocation_status="REVIEW", raw_narration="unmatched giving")
        # a bank debit awaiting classification -> debit queue, NOT review queue
        self.debit = Transaction.objects.create(date=dt.date(2026, 5, 3),
            channel="BANK", direction="DEBIT", amount=Decimal("900"),
            allocation_status="REVIEW", raw_narration="LEDGER FEE")
        self.client.login(username="rq", password="x")

    def test_debit_not_in_review_queue(self):
        from django.urls import reverse
        ids = [t.id for t in self.client.get(reverse("queue")).context["items"]]
        self.assertIn(self.credit.id, ids)
        self.assertNotIn(self.debit.id, ids)

    def test_debit_in_debit_queue(self):
        from django.urls import reverse
        ids = [t.id for t in self.client.get(reverse("debit_queue")).context["debits"]]
        self.assertIn(self.debit.id, ids)
        self.assertNotIn(self.credit.id, ids)


class BulkReceiptsTests(TestCase):
    def setUp(self):
        import datetime as dt
        from decimal import Decimal
        from django.contrib.auth.models import User
        from departments.models import Department
        from envelopes.models import Envelope, EnvelopeLine
        self.u = User.objects.create_superuser("br", password="x")
        d = Department.objects.create(name="Tithe", fund_type=Department.FundType.TRUST)
        sab = dt.date(2026, 5, 2)  # a Saturday
        for i in range(3):
            e = Envelope.objects.create(receipt_no=f"R-{i}", contributor_name=f"Giver {i}",
                date=sab, total=Decimal("1000"), recorded_by=self.u)
            EnvelopeLine.objects.create(envelope=e, department=d, amount=Decimal("1000"))
        self.sab = sab
        self.client.login(username="br", password="x")

    def test_bulk_a4_lists_all_receipts(self):
        from django.urls import reverse
        r = self.client.get(reverse("envelope_receipts_bulk") + f"?date={self.sab.isoformat()}")
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertEqual(body.count('class="rcpt"'), 3)
        self.assertIn("a4-grid", body)

    def test_bulk_etr_format(self):
        from django.urls import reverse
        r = self.client.get(reverse("envelope_receipts_bulk")
                            + f"?date={self.sab.isoformat()}&format=etr")
        self.assertEqual(r.status_code, 200)
        self.assertIn('class="etr"', r.content.decode())


class AllocationPeriodTests(TestCase):
    def setUp(self):
        from departments.models import Department
        from giving.models import AllocationRule
        self.permanent = Department.objects.create(name="General", fund_type="LOCAL")
        self.camp = Department.objects.create(name="Camp Meeting", fund_type="LOCAL")
        import datetime as dt
        AllocationRule.objects.create(reference="campoffer", department=self.permanent, source="SEED")
        AllocationRule.objects.create(reference="campoffer", department=self.camp, source="SEED",
            valid_from=dt.date(2026, 12, 1), valid_to=dt.date(2026, 12, 31))

    def test_period_rule_supersedes_permanent_in_range(self):
        import datetime as dt
        from giving.services.allocation import allocate
        r, _ = allocate("campoffer", dt.date(2026, 12, 15))
        self.assertEqual(r, self.camp)

    def test_permanent_used_outside_period(self):
        import datetime as dt
        from giving.services.allocation import allocate
        r, _ = allocate("campoffer", dt.date(2026, 6, 1))
        self.assertEqual(r, self.permanent)


class CashDuplicateNameTests(TestCase):
    def setUp(self):
        import datetime as dt
        from decimal import Decimal
        from departments.models import Department
        from giving.models import Transaction
        self.fund = Department.objects.create(name="LCB", fund_type="LOCAL")
        Transaction.objects.create(date=dt.date(2026, 5, 2), channel="CASH",
            direction="CREDIT", amount=Decimal("500"), department=self.fund,
            allocation_status="MANUAL", payer_name="Ruth Momanyi")

    def test_same_amount_different_name_not_duplicate(self):
        import datetime as dt
        from decimal import Decimal
        from giving.views import _cash_duplicate
        self.assertFalse(_cash_duplicate(dt.date(2026, 5, 2), self.fund,
                                         Decimal("500"), "John Otieno"))

    def test_same_amount_similar_name_is_duplicate(self):
        import datetime as dt
        from decimal import Decimal
        from giving.views import _cash_duplicate
        self.assertTrue(_cash_duplicate(dt.date(2026, 5, 2), self.fund,
                                        Decimal("500"), "Momanyi Ruth"))


class ImportConfirmationTests(TestCase):
    def setUp(self):
        import datetime as dt
        from decimal import Decimal
        from django.contrib.auth.models import User
        from departments.models import Department
        from giving.models import Transaction
        from statements.models import StatementImport
        self.u = User.objects.create_superuser("ic", password="x")
        self.fund = Department.objects.create(name="LCB", fund_type="LOCAL")
        self.imp = StatementImport.objects.create(uploaded_by=self.u, filename="x.csv")
        self.t = Transaction.objects.create(date=dt.date(2026, 5, 2), channel="BANK",
            direction="CREDIT", amount=Decimal("1000"), department=self.fund,
            allocation_status="AUTO", confirmed=False, statement_import=self.imp)
        self.client.login(username="ic", password="x")

    def test_unconfirmed_excluded_from_balances(self):
        from decimal import Decimal
        from reports.services import balances
        rows = {r["department"].id: r["closing"]
                for r in balances.department_summary(None, None, consolidated=False)}
        self.assertEqual(rows.get(self.fund.id, Decimal(0)), Decimal(0))

    def test_confirm_makes_it_count(self):
        from decimal import Decimal
        from django.urls import reverse
        from reports.services import balances
        self.client.post(reverse("statement_auto_review", args=[self.imp.id]),
                         {"confirm_all": "1", f"dept_{self.t.id}": str(self.fund.id)})
        self.t.refresh_from_db()
        self.assertTrue(self.t.confirmed)
        rows = {r["department"].id: r["closing"]
                for r in balances.department_summary(None, None, consolidated=False)}
        self.assertEqual(rows.get(self.fund.id, Decimal(0)), Decimal("1000"))

    def test_excel_export(self):
        from django.urls import reverse
        r = self.client.get(reverse("statement_auto_excel", args=[self.imp.id]))
        self.assertEqual(r.status_code, 200)
        self.assertIn("spreadsheet", r["Content-Type"])


class RuleDeleteAndConfigTests(TestCase):
    def setUp(self):
        from django.contrib.auth.models import User, Group
        from core.roles import TREASURER
        from departments.models import Department
        from giving.models import AllocationRule
        u = User.objects.create_user("rd", password="x")
        u.groups.add(Group.objects.get_or_create(name=TREASURER)[0])
        self.client.login(username="rd", password="x")
        self.fund = Department.objects.create(name="LCB", fund_type="LOCAL")
        self.rule = AllocationRule.objects.create(reference="tithe", department=self.fund,
                                                  source="SEED")

    def test_rule_delete_redirects_and_removes(self):
        from django.urls import reverse
        from giving.models import AllocationRule
        r = self.client.post(reverse("rule_delete", args=[self.rule.id]))
        self.assertEqual(r.status_code, 302)
        self.assertFalse(AllocationRule.objects.filter(id=self.rule.id).exists())

    def test_configurable_dev_prefix(self):
        import datetime as dt
        from core.models import SiteConfig
        from giving.services.allocation import allocate
        cfg = SiteConfig.get(); cfg.dev_group_extra_prefixes = "project, phase"; cfg.save()
        r, status = allocate("project7", dt.date(2026, 5, 2))
        self.assertEqual(r, "DEV_GROUP_7")


class ExpectedCashTests(TestCase):
    def test_expected_cash_includes_envelopes_excludes_bank_less_disbursed(self):
        import datetime as dt
        from decimal import Decimal
        from django.contrib.auth.models import User
        from departments.models import Department
        from giving.models import Transaction
        from cashbook.models import Expense
        from core.utils import sabbath_of
        from envelopes.views import CountSessionCreate
        u = User.objects.create_superuser("ec", password="x")
        fund = Department.objects.create(name="Loose", fund_type="LOCAL")
        sab = sabbath_of(dt.date(2026, 5, 2))
        Transaction.objects.create(date=sab, channel="CASH", direction="CREDIT",
            amount=Decimal("1000"), department=fund, allocation_status="MANUAL", confirmed=True)
        Transaction.objects.create(date=sab, channel="ENVELOPE", direction="CREDIT",
            amount=Decimal("2000"), department=fund, allocation_status="MANUAL", confirmed=True)
        Transaction.objects.create(date=sab, channel="BANK", direction="CREDIT",
            amount=Decimal("5000"), department=fund, allocation_status="MANUAL", confirmed=True)
        Expense.objects.create(date=sab, department=fund, description="Cash out",
            amount=Decimal("300"), category="OTHER", method="CASH", status="PAID",
            recorded_by=u, approved_by=u)
        self.assertEqual(CountSessionCreate()._expected(sab), Decimal("2700"))


class TransactionExportDetailTests(TestCase):
    """The transactions Excel export includes the richer columns (M-Pesa ref etc)."""

    def test_export_has_mpesa_ref_column(self):
        import io, openpyxl
        from django.contrib.auth.models import User, Group
        from django.test import Client
        from giving.models import Transaction
        import datetime as dt
        from decimal import Decimal
        u = User.objects.create_user("tx", password="x")
        g, _ = Group.objects.get_or_create(name="Treasurer")
        u.groups.add(g)
        Transaction.objects.create(date=dt.date(2026, 6, 6), channel="BANK",
            direction="CREDIT", amount=Decimal("500"), allocation_status="AUTO",
            confirmed=True, mpesa_ref="UF6EXP01", core_ref="UF6EXP01",
            payer_name="Exporter")
        c = Client(); c.force_login(u)
        r = c.get("/transactions/?export=xlsx")
        self.assertEqual(r.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # the workbook has church/title rows before the header; find the header row
        header = []
        for row in ws.iter_rows(values_only=True):
            if row and "Date" in row:
                header = list(row)
                break
        self.assertIn("M-Pesa ref", header)
        self.assertIn("Core ref", header)
        self.assertIn("Sabbath", header)


class MarkProcessedImportTests(TestCase):
    """Bulk 'mark processed (via envelope)': matches a bank credit by reference,
    confirms with the amount, sets manual_receipt, and reports problems
    without applying bad rows. These entries are handled — not receipted."""

    def setUp(self):
        import datetime as dt
        from decimal import Decimal
        from django.contrib.auth.models import User
        from departments.models import Department
        from giving.models import Transaction
        User.objects.create_superuser("mp", password="x")
        self.client.login(username="mp", password="x")
        self.d = Department.objects.create(name="Combined Offering",
                                           fund_type=Department.FundType.LOCAL)
        mk = lambda ref, amt: Transaction.objects.create(
            date=dt.date(2026, 6, 1), channel="BANK", direction="CREDIT",
            amount=Decimal(amt), department=self.d, reference=ref,
            allocation_status="AUTO", confirmed=True)
        self.t1 = mk("REFAAA111", "1500")
        self.t2 = mk("REFBBB222", "2000")

    def _post(self, csv_text):
        from django.urls import reverse
        from django.core.files.uploadedfile import SimpleUploadedFile
        return self.client.post(reverse("mark_processed_import"),
            {"file": SimpleUploadedFile("p.csv", csv_text.encode())})

    def test_match_marks_processed(self):
        self._post("reference,amount\nREFAAA111,1500\n")
        self.t1.refresh_from_db()
        self.assertTrue(self.t1.manual_receipt)

    def test_amount_mismatch_is_rejected(self):
        self._post("reference,amount\nREFBBB222,9999\n")
        self.t2.refresh_from_db()
        self.assertFalse(self.t2.manual_receipt)   # wrong amount → skipped

    def test_unknown_reference_is_skipped(self):
        # should not raise, and nothing gets marked
        self._post("reference,amount\nNOSUCHREF,100\n")
        self.t1.refresh_from_db(); self.t2.refresh_from_db()
        self.assertFalse(self.t1.manual_receipt)
        self.assertFalse(self.t2.manual_receipt)

    def test_amount_disambiguates_two_matches(self):
        import datetime as dt
        from decimal import Decimal
        from giving.models import Transaction
        # a second entry sharing the reference but a different amount
        dup = Transaction.objects.create(date=dt.date(2026, 6, 2), channel="BANK",
            direction="CREDIT", amount=Decimal("7777"), department=self.d,
            reference="REFAAA111", allocation_status="AUTO", confirmed=True)
        self._post("reference,amount\nREFAAA111,7777\n")
        self.t1.refresh_from_db(); dup.refresh_from_db()
        self.assertFalse(self.t1.manual_receipt)   # 1500 one untouched
        self.assertTrue(dup.manual_receipt)        # 7777 one marked

    def test_xlsx_upload_path(self):
        import io
        import openpyxl
        from django.urls import reverse
        from django.core.files.uploadedfile import SimpleUploadedFile
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["reference", "amount"]); ws.append(["REFAAA111", 1500])
        buf = io.BytesIO(); wb.save(buf)
        self.client.post(reverse("mark_processed_import"),
            {"file": SimpleUploadedFile("p.xlsx", buf.getvalue())})
        self.t1.refresh_from_db()
        self.assertTrue(self.t1.manual_receipt)

    def test_template_download(self):
        from django.urls import reverse
        r = self.client.get(reverse("mark_processed_import") + "?template=1")
        self.assertEqual(r.status_code, 200)
        self.assertIn("text/csv", r["Content-Type"])
        self.assertIn(b"reference", r.content)


class MarkProcessedSplitFundTests(TestCase):
    """A split-fund contribution posts as several rows sharing the reference with divided
    amounts. The importer must confirm the whole group by its TOTAL and mark
    every part — not look for a single row equal to the lump sum."""

    def setUp(self):
        import datetime as dt
        from decimal import Decimal
        from django.contrib.auth.models import User
        from departments.models import Department
        from giving.models import Transaction
        User.objects.create_superuser("mps", password="x")
        self.client.login(username="mps", password="x")
        self.trust = Department.objects.create(name="Combined Trust half",
            fund_type=Department.FundType.TRUST, selectable=False)
        self.local = Department.objects.create(name="Combined Local half",
            fund_type=Department.FundType.LOCAL, selectable=False)
        common = dict(date=dt.date(2026, 6, 1), channel="BANK", direction="CREDIT",
                      reference="SPLITREF", confirmed=True, allocation_status="MANUAL")
        # 2000 split 50/50 -> two 1000 rows sharing the reference
        self.h1 = Transaction.objects.create(amount=Decimal("1000"),
            department=self.trust, core_ref="SPCORE", bank_receipt="SPRC", **common)
        self.h2 = Transaction.objects.create(amount=Decimal("1000"),
            department=self.local, core_ref="SPCORE-S1", **common)

    def _post(self, csv_text):
        from django.urls import reverse
        from django.core.files.uploadedfile import SimpleUploadedFile
        return self.client.post(reverse("mark_processed_import"),
            {"file": SimpleUploadedFile("p.csv", csv_text.encode())})

    def test_total_marks_all_split_parts(self):
        self._post("reference,amount\nSPLITREF,2000\n")     # the lump sum
        self.h1.refresh_from_db(); self.h2.refresh_from_db()
        self.assertTrue(self.h1.manual_receipt)
        self.assertTrue(self.h2.manual_receipt)

    def test_wrong_total_marks_nothing(self):
        self._post("reference,amount\nSPLITREF,5000\n")
        self.h1.refresh_from_db(); self.h2.refresh_from_db()
        self.assertFalse(self.h1.manual_receipt)
        self.assertFalse(self.h2.manual_receipt)

    def test_three_way_split_by_total(self):
        import datetime as dt
        from decimal import Decimal
        from departments.models import Department
        from giving.models import Transaction
        third = Department.objects.create(name="Combined Third",
            fund_type=Department.FundType.LOCAL, selectable=False)
        # add a third part so the group is 1000+1000+500 = 2500, ref TRIO
        common = dict(date=dt.date(2026, 6, 3), channel="BANK", direction="CREDIT",
                      reference="TRIO", confirmed=True, allocation_status="MANUAL")
        a = Transaction.objects.create(amount=Decimal("1000"), department=self.trust,
            core_ref="TRIOC", **common)
        b = Transaction.objects.create(amount=Decimal("1000"), department=self.local,
            core_ref="TRIOC-S1", **common)
        c = Transaction.objects.create(amount=Decimal("500"), department=third,
            core_ref="TRIOC-S2", **common)
        self._post("reference,amount\nTRIO,2500\n")
        for t in (a, b, c):
            t.refresh_from_db()
            self.assertTrue(t.manual_receipt)


class MarkProcessedClearsQueueTests(TestCase):
    """Marking an entry processed must also remove it from the review queue, and
    cascade to all parts of a split contribution."""

    def setUp(self):
        import datetime as dt
        from decimal import Decimal
        from django.contrib.auth.models import User
        from departments.models import Department
        from giving.models import Transaction
        User.objects.create_superuser("mpq", password="x")
        self.client.login(username="mpq", password="x")
        self.trust = Department.objects.create(name="MPQ Trust",
            fund_type=Department.FundType.TRUST, selectable=False)
        self.local = Department.objects.create(name="MPQ Local",
            fund_type=Department.FundType.LOCAL, selectable=False)

    def _review_queue_ids(self):
        from giving.models import Transaction
        return set(Transaction.objects.filter(
            allocation_status="REVIEW", direction="CREDIT"
        ).values_list("id", flat=True))

    def test_model_method_clears_review_status(self):
        import datetime as dt
        from decimal import Decimal
        from giving.models import Transaction
        t = Transaction.objects.create(date=dt.date(2026, 6, 1), channel="BANK",
            direction="CREDIT", amount=Decimal("500"), reference="R1",
            core_ref="C1", allocation_status="REVIEW", confirmed=True)
        self.assertIn(t.id, self._review_queue_ids())
        t.mark_manual_receipt()
        t.refresh_from_db()
        self.assertTrue(t.manual_receipt)
        self.assertNotEqual(t.allocation_status, "REVIEW")
        self.assertNotIn(t.id, self._review_queue_ids())

    def test_cascade_clears_split_siblings_from_queue(self):
        import datetime as dt
        from decimal import Decimal
        from giving.models import Transaction
        common = dict(date=dt.date(2026, 6, 2), channel="BANK", direction="CREDIT",
                      reference="SPLIT", confirmed=True, allocation_status="REVIEW")
        h1 = Transaction.objects.create(amount=Decimal("1000"),
            department=self.trust, core_ref="SC", **common)
        h2 = Transaction.objects.create(amount=Decimal("1000"),
            department=self.local, core_ref="SC-S1", **common)
        # mark just one half via the model method with cascade
        n = h1.mark_manual_receipt(cascade_split=True)
        self.assertEqual(n, 2)                      # both parts marked
        h1.refresh_from_db(); h2.refresh_from_db()
        self.assertNotIn(h1.id, self._review_queue_ids())
        self.assertNotIn(h2.id, self._review_queue_ids())

    def test_bulk_import_removes_from_queue(self):
        import datetime as dt
        from decimal import Decimal
        from django.urls import reverse
        from django.core.files.uploadedfile import SimpleUploadedFile
        from giving.models import Transaction
        t = Transaction.objects.create(date=dt.date(2026, 6, 3), channel="BANK",
            direction="CREDIT", amount=Decimal("750"), reference="BR1",
            core_ref="BC1", allocation_status="REVIEW", confirmed=True)
        self.assertIn(t.id, self._review_queue_ids())
        self.client.post(reverse("mark_processed_import"),
            {"file": SimpleUploadedFile("p.csv", b"reference,amount\nBR1,750\n")})
        t.refresh_from_db()
        self.assertNotIn(t.id, self._review_queue_ids())

    def test_edit_checkbox_clears_queue_and_cascades(self):
        import datetime as dt
        from decimal import Decimal
        from giving.models import Transaction
        common = dict(date=dt.date(2026, 6, 4), channel="BANK", direction="CREDIT",
                      reference="ESPLIT", confirmed=True, allocation_status="REVIEW")
        h1 = Transaction.objects.create(amount=Decimal("1200"),
            department=self.trust, core_ref="EC", **common)
        h2 = Transaction.objects.create(amount=Decimal("1200"),
            department=self.local, core_ref="EC-S1", **common)
        self.client.post(f"/transactions/{h1.id}/edit/", {
            "date": "2026-06-04", "channel": "BANK", "direction": "CREDIT",
            "department": self.trust.id, "amount": "1200", "reference": "ESPLIT",
            "allocation_status": "REVIEW", "manual_receipt": "on"})
        h1.refresh_from_db(); h2.refresh_from_db()
        self.assertNotIn(h1.id, self._review_queue_ids())
        self.assertNotIn(h2.id, self._review_queue_ids())   # sibling cascaded


class CashEntrySabbathTests(TestCase):
    """Bug: loose cash dated to a Sabbath that has been closed was rolling forward
    to the next Sabbath in the cash count. Manually-dated cash must count for the
    Sabbath the treasurer assigned it to."""

    def setUp(self):
        import datetime as dt
        from django.contrib.auth.models import User
        from departments.models import Department
        from core.models import SabbathClose
        self.u = User.objects.create_superuser("cs", password="x")
        self.client.login(username="cs", password="x")
        self.d = Department.objects.create(name="Cash Sabbath Fund",
            fund_type="LOCAL", selectable=True)
        # the 13th is a Saturday; close it (as on the live system)
        SabbathClose.objects.create(sabbath=dt.date(2026, 6, 13), closed_by=self.u)

    def test_cash_dated_closed_sabbath_counts_for_that_sabbath(self):
        import datetime as dt
        from giving.models import Transaction
        self.client.post("/cash/new/", {
            "date": "2026-06-13", "department": str(self.d.id),
            "channel": "CASH", "amount": "450", "payer_name": "Loose 13th",
            "confirm_duplicate": "1"})
        t = Transaction.objects.filter(payer_name="LOOSE 13TH",
                                       channel="CASH").first()
        self.assertIsNotNone(t)
        self.assertEqual(t.service_sabbath, dt.date(2026, 6, 13))


class BulkAllocateSplitFundTests(TestCase):
    """Item 3: split funds must be selectable in the bulk-allocate dropdown and
    split each selected contribution into its parts."""

    def setUp(self):
        import datetime as dt
        from decimal import Decimal
        from django.contrib.auth.models import User
        from departments.models import Department
        from giving.models import SplitFund, SplitComponent
        self.u = User.objects.create_superuser("bs", password="x")
        self.client.login(username="bs", password="x")
        trust = Department.objects.create(name="BS Trust", fund_type="TRUST",
                                          selectable=True, category="TRUST")
        local = Department.objects.create(name="BS Local", fund_type="LOCAL",
                                          selectable=True, category="OFFERING")
        self.sf = SplitFund.objects.create(name="BS Combined", active=True)
        SplitComponent.objects.create(split_fund=self.sf, department=trust,
                                          percent=Decimal("50"))
        SplitComponent.objects.create(split_fund=self.sf, department=local,
                                          percent=Decimal("50"))

    def test_split_fund_in_queue_dropdown(self):
        r = self.client.get("/queue/")
        self.assertContains(r, f"sf:{self.sf.id}")

    def test_bulk_allocate_splits_each_gift(self):
        import datetime as dt
        from decimal import Decimal
        from giving.models import Transaction
        t = Transaction.objects.create(date=dt.date(2026, 3, 7), channel="BANK",
            direction="CREDIT", amount=Decimal("1000"), payer_name="BSX",
            reference="BSX1", core_ref="BSX1", allocation_status="REVIEW",
            confirmed=True)
        self.client.post("/queue/bulk-allocate/",
            {"txn": [str(t.id)], "department": f"sf:{self.sf.id}"})
        self.assertGreaterEqual(
            Transaction.objects.filter(core_ref__startswith="BSX1").count(), 2)


class RuleImportTests(TestCase):
    """Item 1: bulk import of allocation rules from a spreadsheet."""

    def setUp(self):
        from django.contrib.auth.models import User, Group
        u = User.objects.create_user("ri", password="x")
        g, _ = Group.objects.get_or_create(name="Treasurer")
        u.groups.add(g)
        self.client.login(username="ri", password="x")
        from departments.models import Department
        self.fund = Department.objects.create(name="RI Tithe", fund_type="LOCAL",
                                              selectable=True, category="OFFERING")

    def _file(self, rows):
        import io, openpyxl
        from django.core.files.uploadedfile import SimpleUploadedFile
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Rules"
        ws.append(["Reference", "Match type", "Fund", "Split fund", "Valid from", "Valid to"])
        for r in rows:
            ws.append(r)
        buf = io.BytesIO(); wb.save(buf)
        return SimpleUploadedFile("r.xlsx", buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def test_template_downloads(self):
        r = self.client.get("/rules/import/?download=1")
        self.assertEqual(r.status_code, 200)
        self.assertIn("spreadsheet", r["Content-Type"])

    def test_import_creates_rules_with_match_type(self):
        from giving.models import AllocationRule
        f = self._file([["camprule", "Contains", self.fund.name, "", "", ""],
                        ["norfund", "Exact", "", "", "", ""]])
        self.client.post("/rules/import/", {"file": f})
        self.client.post("/rules/import/", {"apply": "1"})
        rule = AllocationRule.objects.filter(reference="camprule").first()
        self.assertIsNotNone(rule)
        self.assertEqual(rule.match_type, "CONTAINS")
        self.assertEqual(rule.department_id, self.fund.id)
        # row with no fund is skipped
        self.assertFalse(AllocationRule.objects.filter(reference="norfund").exists())

    def test_existing_rule_updated(self):
        from giving.models import AllocationRule
        AllocationRule.objects.create(reference="dup", department=self.fund,
                                      match_type="EXACT")
        f = self._file([["dup", "Starts with", self.fund.name, "", "", ""]])
        self.client.post("/rules/import/", {"file": f})
        self.client.post("/rules/import/", {"apply": "1"})
        self.assertEqual(AllocationRule.objects.filter(reference="dup").count(), 1)
        self.assertEqual(AllocationRule.objects.get(reference="dup").match_type, "STARTS")


class RegexAllocationRuleTests(TestCase):
    """Item 2: a REGEX match type lets one rule cover many narration variations
    (EXPENSE_1, exp1, expe1, expense1) for a camp/expense group."""

    def setUp(self):
        from departments.models import Department
        self.d = Department.objects.create(name="Camp Group 1", fund_type="LOCAL",
                                           selectable=True, category="DEVELOPMENT")

    def test_one_regex_matches_all_variations(self):
        from giving.models import AllocationRule
        from giving.services.allocation import allocate
        AllocationRule.objects.create(reference=r"exp(e|ense)?_?0*1\b",
            match_type="REGEX", department=self.d, source="SEED")
        for ref in ["EXPENSE_1", "exp1", "expe1", "expense1"]:
            res, status = allocate(ref)
            self.assertEqual(getattr(res, "name", res), self.d.name,
                             f"{ref} should map to the camp group")
        # must not catch group 10
        res, _ = allocate("expense10")
        self.assertNotEqual(getattr(res, "name", res), self.d.name)

    def test_invalid_regex_never_crashes(self):
        from giving.models import AllocationRule
        from giving.services.allocation import allocate
        # a deliberately broken pattern is stored but simply never matches
        AllocationRule.objects.create(reference="exp[1", match_type="REGEX",
            department=self.d, source="SEED")
        res, status = allocate("anything")
        self.assertEqual(status, "REVIEW")   # no crash, falls through to review

    def test_form_rejects_bad_regex(self):
        from giving.forms import RuleForm
        f = RuleForm(data={"reference": "exp[1", "match_type": "REGEX",
                           "department": str(self.d.id), "source": "SEED"})
        self.assertFalse(f.is_valid())
        self.assertIn("reference", f.errors)


class CashEntryDeleteTests(TestCase):
    """Item 3: a cash entry is its ledger row; deleting it removes the one record.
    Bank, reversed, or envelope-receipted rows are protected here."""

    def setUp(self):
        import datetime as dt
        from decimal import Decimal
        from django.contrib.auth.models import User, Group
        from departments.models import Department
        from giving.models import Transaction
        u = User.objects.create_user("cashdel", password="x")
        g, _ = Group.objects.get_or_create(name="Treasurer")
        u.groups.add(g)
        self.client.login(username="cashdel", password="x")
        self.fund = Department.objects.create(name="CD Fund", fund_type="LOCAL",
                                              category="OFFERING", selectable=True)
        self.cash = Transaction.objects.create(date=dt.date(2026, 6, 6), channel="CASH",
            direction="CREDIT", amount=Decimal("250"), department=self.fund,
            confirmed=True, allocation_status="MANUAL")

    def test_delete_cash_entry_removes_ledger_row(self):
        from giving.models import Transaction
        r = self.client.post(f"/cash/{self.cash.id}/delete/")
        self.assertEqual(r.status_code, 302)
        self.assertFalse(Transaction.objects.filter(id=self.cash.id).exists())

    def test_bank_row_not_deletable_here(self):
        import datetime as dt
        from decimal import Decimal
        from giving.models import Transaction
        b = Transaction.objects.create(date=dt.date(2026, 6, 6), channel="BANK",
            direction="CREDIT", amount=Decimal("500"), department=self.fund,
            confirmed=True, allocation_status="MANUAL")
        self.client.post(f"/cash/{b.id}/delete/")
        self.assertTrue(Transaction.objects.filter(id=b.id).exists())

    def test_split_cash_parts_deleted_together(self):
        import datetime as dt
        from decimal import Decimal
        from giving.models import Transaction
        base = Transaction.objects.create(date=dt.date(2026, 6, 6), channel="CASH",
            direction="CREDIT", amount=Decimal("600"), department=self.fund,
            confirmed=True, allocation_status="MANUAL", core_ref="CD1")
        part = Transaction.objects.create(date=dt.date(2026, 6, 6), channel="CASH",
            direction="CREDIT", amount=Decimal("400"), department=self.fund,
            confirmed=True, allocation_status="MANUAL", core_ref="CD1-S1")
        self.client.post(f"/cash/{base.id}/delete/")
        self.assertFalse(Transaction.objects.filter(id__in=[base.id, part.id]).exists())


class NumberedFundFamilyTests(TestCase):
    """A single 'numbered fund family' config routes EXPENSE<n> -> fund CAMP_<n>
    for every group, handling narration variations, without a rule per group."""

    def setUp(self):
        from core.models import SiteConfig
        from departments.models import Department
        cfg = SiteConfig.get()
        cfg.numbered_fund_families = "expense, exp, expe = CAMP_{n}"
        cfg.dev_group_extra_prefixes = ""
        cfg.save()
        for n in (1, 2, 10, 30):
            Department.objects.create(name=f"CAMP_{n}", fund_type="LOCAL",
                                      category="DEVELOPMENT", selectable=True)

    def test_variations_route_to_group_1(self):
        from giving.services.allocation import allocate
        for ref in ["EXPENSE1", "exp1", "expe1", "expense1", "EXPENSE 1", "EXPENSE_1"]:
            res, st = allocate(ref)
            self.assertEqual(getattr(res, "name", res), "CAMP_1", ref)
            self.assertEqual(st, "AUTO")

    def test_distinguishes_one_from_ten(self):
        from giving.services.allocation import allocate
        self.assertEqual(getattr(allocate("EXPENSE10")[0], "name", None), "CAMP_10")
        self.assertEqual(getattr(allocate("EXPENSE30")[0], "name", None), "CAMP_30")
        self.assertEqual(getattr(allocate("expense2")[0], "name", None), "CAMP_2")

    def test_missing_fund_falls_through(self):
        from giving.services.allocation import allocate
        res, st = allocate("expense99")          # no CAMP_99 fund exists
        self.assertEqual(res, "UNALLOCATED")
        self.assertEqual(st, "REVIEW")

    def test_no_config_means_no_routing(self):
        from core.models import SiteConfig
        from giving.services.allocation import allocate
        cfg = SiteConfig.get(); cfg.numbered_fund_families = ""; cfg.save()
        res, st = allocate("expense1")
        self.assertEqual(res, "UNALLOCATED")


class CampaignFallbackTests(TestCase):
    """The campaign table allocates only after the normal rules miss, gated by
    trigger words, matching the payer by phone or a unique name."""
    def setUp(self):
        from departments.models import Department
        from giving.models import Campaign, CampaignMember
        self.d = Department.objects.create(name="Camp Fund", fund_type="LOCAL", category="OFFERING")
        self.camp = Campaign.objects.create(name="Camp", department=self.d,
                                            triggers="expense, campexpense", active=True)
        CampaignMember.objects.create(campaign=self.camp, name="Amos Ndegwa",
                                      phone="254791896792", group="CAMP_1")

    def test_trigger_and_phone_allocates_to_campaign(self):
        from giving.services.allocation import campaign_allocate
        camp, grp, dept, status = campaign_allocate("441211#campexpense", "X", "254791896792")
        # the matched member's contribution splits to their subgroup fund (CAMP_1),
        # which is created under the campaign's parent fund and inherits its type.
        self.assertEqual(dept.name, "CAMP_1")
        self.assertEqual(dept.parent, self.d)
        self.assertEqual(dept.fund_type, self.d.fund_type)
        self.assertEqual(grp, "CAMP_1")
        self.assertEqual(status, "AUTO")

    def test_trigger_but_no_member_routes_to_parent_for_review(self):
        from giving.services.allocation import campaign_allocate
        camp, grp, dept, status = campaign_allocate("441211#campexpense", "Nobody Here", "")
        self.assertEqual(dept, self.d)        # parent fund, flagged for review
        self.assertEqual(status, "REVIEW")

    def test_no_trigger_falls_through_to_normal_rules(self):
        from giving.services.allocation import campaign_allocate
        self.assertIsNone(campaign_allocate("tithe", "Amos Ndegwa", "254791896792")[0])

    def test_trigger_without_member_routes_to_fund_for_review(self):
        from giving.services.allocation import campaign_allocate
        camp, grp, dept, status = campaign_allocate("expense", "Stranger", "254700000000")
        self.assertEqual(dept, self.d)
        self.assertEqual(status, "REVIEW")

    def test_inactive_campaign_is_ignored(self):
        from giving.services.allocation import campaign_allocate
        self.camp.active = False
        self.camp.save()
        self.assertIsNone(campaign_allocate("campexpense", "Amos Ndegwa", "254791896792")[0])
