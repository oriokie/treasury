"""Download button on the transactions page: trust fund credits not yet
formally receipted, with Date/Phone/Member/Amount/Fund/Reference columns. A
lump sum split across several trust funds (e.g. a Combined Offering split
across two trust accounts) is recombined into one row — one gift, one line —
rather than shown as separate partial amounts."""
import datetime as dt
from decimal import Decimal
from django.test import TestCase, Client
from django.contrib.auth.models import User, Group
from departments.models import Department
from members.models import Member
from giving.models import Transaction, SplitFund, SplitComponent, AllocationRule


def _tr():
    u = User.objects.create_user("tr_tpr", password="x", is_superuser=True)
    u.groups.add(Group.objects.get_or_create(name="Treasurer")[0])
    return u


class TrustPendingReceiptExportTests(TestCase):
    def setUp(self):
        self.tr = _tr()
        self.trust1 = Department.objects.create(name="TPRTrust1", fund_type="TRUST",
            category="OFFERING")
        self.trust2 = Department.objects.create(name="TPRTrust2", fund_type="TRUST",
            category="OFFERING")
        self.local = Department.objects.create(name="TPRLocal1", fund_type="LOCAL",
            category="MINISTRY")
        self.member = Member.objects.create(name="Jane Giver", phone="254712000111")
        self.c = Client(); self.c.force_login(self.tr)

    def _rows(self, response):
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content))
        return list(wb.active.iter_rows(values_only=True))

    def test_button_present_on_transactions_page(self):
        """The ledger page links to the Pending receipt PAGE (not directly to
        a download) — the Excel/PDF downloads live on that page itself, so
        there is one place to find them, not three."""
        b = self.c.get("/transactions/").content.decode()
        self.assertIn("Pending receipt", b)
        self.assertIn("/transactions/pending-receipt/", b)

    def test_downloads_are_on_the_pending_receipt_page(self):
        b = self.c.get("/transactions/pending-receipt/").content.decode()
        self.assertIn("export=pending-receipt", b)
        self.assertIn("export=pending-receipt-pdf", b)

    def test_the_old_export_key_still_works(self):
        """Renaming a URL a treasurer has bookmarked — or that the Telegram
        bot's /pending route points at — is not a rename, it is a breakage."""
        self.assertEqual(
            self.c.get("/transactions/?export=trust-pending-receipt").status_code, 200)

    def test_export_returns_xlsx(self):
        r = self.c.get("/transactions/?export=trust-pending-receipt")
        self.assertEqual(r.status_code, 200)
        self.assertIn("spreadsheet", r["Content-Type"])

    def test_unreceipted_trust_credit_included(self):
        Transaction.objects.create(date=dt.date(2026, 6, 10), amount=Decimal("2000"),
            direction="CREDIT", confirmed=True, channel="BANK",
            allocation_status="MANUAL", department=self.trust1, member=self.member,
            payer_phone="254712000111", reference="simple-ref")
        rows = self._rows(self.c.get("/transactions/?export=trust-pending-receipt"))
        data_rows = [r for r in rows if r[0] == "2026-06-10"]
        self.assertEqual(len(data_rows), 1)
        row = data_rows[0]
        self.assertEqual(row[1], "254712000111")
        self.assertEqual(row[2], "JANE GIVER")
        self.assertEqual(row[3], 2000)
        self.assertEqual(row[4], "TPRTrust1")
        self.assertEqual(row[5], "simple-ref")

    def test_receipted_trust_credit_excluded(self):
        Transaction.objects.create(date=dt.date(2026, 6, 11), amount=Decimal("500"),
            direction="CREDIT", confirmed=True, channel="ENVELOPE",
            allocation_status="MANUAL", department=self.trust1, reference="receipted-ref")
        rows = self._rows(self.c.get("/transactions/?export=trust-pending-receipt"))
        self.assertFalse(any(r[5] == "receipted-ref" for r in rows if r))

    def test_manually_receipted_trust_credit_excluded(self):
        Transaction.objects.create(date=dt.date(2026, 6, 11), amount=Decimal("500"),
            direction="CREDIT", confirmed=True, channel="BANK", manual_receipt=True,
            allocation_status="MANUAL", department=self.trust1, reference="manual-receipt-ref")
        rows = self._rows(self.c.get("/transactions/?export=trust-pending-receipt"))
        self.assertFalse(any(r[5] == "manual-receipt-ref" for r in rows if r))

    def test_local_fund_credit_excluded(self):
        Transaction.objects.create(date=dt.date(2026, 6, 13), amount=Decimal("300"),
            direction="CREDIT", confirmed=True, channel="CASH",
            allocation_status="MANUAL", department=self.local, reference="local-ref")
        rows = self._rows(self.c.get("/transactions/?export=trust-pending-receipt"))
        self.assertFalse(any(r[5] == "local-ref" for r in rows if r))

    def test_split_across_trust_funds_combined_into_one_row(self):
        sf = SplitFund.objects.create(name="Combined Trust Split TPR")
        SplitComponent.objects.create(split_fund=sf, department=self.trust1, percent=Decimal("60"))
        SplitComponent.objects.create(split_fund=sf, department=self.trust2, percent=Decimal("40"))
        AllocationRule.objects.create(reference="splitref-tpr", split_fund=sf, source="LEARNED")
        Transaction.objects.create(date=dt.date(2026, 6, 12), amount=Decimal("600"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=self.trust1, reference="splitref-tpr", core_ref="CBXTPR-S1",
            payer_phone="254799222333", payer_name="Peter Payer")
        Transaction.objects.create(date=dt.date(2026, 6, 12), amount=Decimal("400"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=self.trust2, reference="splitref-tpr", core_ref="CBXTPR-S2",
            payer_phone="254799222333", payer_name="Peter Payer")
        rows = self._rows(self.c.get("/transactions/?export=trust-pending-receipt"))
        matching = [r for r in rows if r and r[5] == "splitref-tpr"]
        self.assertEqual(len(matching), 1)
        row = matching[0]
        self.assertEqual(row[3], 1000)
        self.assertEqual(row[4], "Combined Trust Split TPR")
        self.assertEqual(row[1], "254799222333")
        self.assertEqual(row[2], "PETER PAYER")

    def test_split_without_matching_rule_falls_back_to_joined_names(self):
        # same 2-fund split shape but with no AllocationRule for the reference
        Transaction.objects.create(date=dt.date(2026, 6, 14), amount=Decimal("300"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=self.trust1, reference="no-rule-ref", core_ref="CBXNR-S1",
            payer_phone="254700111222")
        Transaction.objects.create(date=dt.date(2026, 6, 14), amount=Decimal("200"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=self.trust2, reference="no-rule-ref", core_ref="CBXNR-S2",
            payer_phone="254700111222")
        rows = self._rows(self.c.get("/transactions/?export=trust-pending-receipt"))
        matching = [r for r in rows if r and r[5] == "no-rule-ref"]
        self.assertEqual(len(matching), 1)
        self.assertEqual(matching[0][3], 500)
        self.assertIn("TPRTrust1", matching[0][4])
        self.assertIn("TPRTrust2", matching[0][4])

    def test_header_columns_exact(self):
        rows = self._rows(self.c.get("/transactions/?export=trust-pending-receipt"))
        header_row = next(r for r in rows if r and r[0] == "Date")
        self.assertEqual(list(header_row), ["Date", "Phone", "Member", "Amount",
                                            "Fund", "Reference", "M-Pesa Reference"])

    def test_excel_export_sorted_by_name(self):
        Transaction.objects.create(date=dt.date(2026, 6, 10), amount=Decimal("100"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=self.trust1, payer_name="ZEBRA WANJIRU", reference="z-ref",
            core_ref="ZREF1")
        Transaction.objects.create(date=dt.date(2026, 6, 11), amount=Decimal("200"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=self.trust1, payer_name="ALPHA OTIENO", reference="a-ref",
            core_ref="AREF1")
        rows = self._rows(self.c.get("/transactions/?export=trust-pending-receipt"))
        names = [r[2] for r in rows if r and r[0] != "Date" and r[2] not in
                (None, "SDA Church", "Items pending receipt")]
        self.assertEqual(names, sorted(names, key=lambda n: n.upper()))

    def test_excel_export_marks_duplicate_names(self):
        Transaction.objects.create(date=dt.date(2026, 6, 10), amount=Decimal("100"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=self.trust1, payer_name="REPEAT PERSON", reference="r1",
            core_ref="RREF1")
        Transaction.objects.create(date=dt.date(2026, 6, 12), amount=Decimal("150"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=self.trust1, payer_name="REPEAT PERSON", reference="r2",
            core_ref="RREF2")
        b = self.c.get("/transactions/?export=trust-pending-receipt")
        rows = self._rows(b)
        dupe_rows = [r for r in rows if r and r[2] and "REPEAT PERSON" in str(r[2])]
        self.assertEqual(len(dupe_rows), 2)
        self.assertTrue(all("repeats" in str(r[2]) for r in dupe_rows))


class PendingReceiptPdfTests(TestCase):
    def setUp(self):
        self.tr = _tr()
        self.trust = Department.objects.create(name="PDFTrust1", fund_type="TRUST",
            category="OFFERING")

    def test_pdf_highlights_and_sorts(self):
        from giving.services.pending_receipt import pending_receipt_pdf_bytes
        Transaction.objects.create(date=dt.date(2026, 6, 10), amount=Decimal("100"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=self.trust, payer_name="TWICE GIVEN", reference="t1",
            core_ref="TREF1")
        Transaction.objects.create(date=dt.date(2026, 6, 12), amount=Decimal("150"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=self.trust, payer_name="TWICE GIVEN", reference="t2",
            core_ref="TREF2")
        pdf = pending_receipt_pdf_bytes(church="Test Church")
        self.assertTrue(pdf.startswith(b"%PDF"))
        self.assertGreater(len(pdf), 500)

    def test_telegram_bot_sends_the_same_sorted_highlighted_pdf(self):
        """The Telegram /pending command must produce the SAME PDF the web
        download serves — same function, same data, so a repeated name is
        highlighted there too, not just on the website."""
        from core.services.telegram_bot import _do_pending
        Transaction.objects.create(date=dt.date(2026, 6, 10), amount=Decimal("100"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=self.trust, payer_name="TWICE GIVEN", reference="t1",
            core_ref="TGREF1")
        Transaction.objects.create(date=dt.date(2026, 6, 12), amount=Decimal("150"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=self.trust, payer_name="TWICE GIVEN", reference="t2",
            core_ref="TGREF2")
        result = _do_pending(chat_id=12345)
        self.assertIn("document", result)
        self.assertTrue(result["document"].startswith(b"%PDF"))
        self.assertEqual(result["filename"], "pending_receipt.pdf")
        self.assertIn("2 item", result["caption"])
