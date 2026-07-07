"""Critical fix: the Transactions Excel/CSV export summed a reversal's
amount as if it were ordinary income, instead of the offsetting entry it
actually is. A reversal (contra) keeps the same direction and a positive
amount as its original by design (the ledger nets it to zero by not
posting either side, not by inverting the stored sign) - so summing the
Amount column double-counted a reversed transaction (original + reversal
both positive) instead of netting to zero, exactly the bug reported: the
export's totals were wrong wherever a reversal existed.

Also hardened _group_split_siblings() so a reversed/reversal transaction
can never be grouped with anything else (a correction entry, not a split
sibling) - it always appears as its own row, even for a manually-entered
cash transaction with no core_ref/mpesa_ref that would otherwise fall
through to the loose reference+date+direction matching."""
import datetime as dt
from decimal import Decimal
from django.test import TestCase, Client
from django.contrib.auth.models import User, Group
from departments.models import Department
from giving.models import Transaction


def _tr():
    u = User.objects.create_user("tr_exportreversal", password="x", is_superuser=True)
    u.groups.add(Group.objects.get_or_create(name="Treasurer")[0])
    return u


def _rows(response):
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(response.content))
    return list(wb.active.iter_rows(values_only=True))


class ExportReversalNettingTests(TestCase):
    def setUp(self):
        self.tr = _tr()
        self.c = Client(); self.c.force_login(self.tr)
        self.d = Department.objects.create(name="ExportRevFund", fund_type="LOCAL",
            category="MINISTRY")

    def test_reversal_and_original_net_to_zero(self):
        t = Transaction.objects.create(date=dt.date(2026, 6, 10), amount=Decimal("500"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="AUTO",
            department=self.d, reference="nettotest1", core_ref="NETTOTEST1")
        t.reverse(self.tr, reason="test")
        rows = _rows(self.c.get("/transactions/?export=xlsx"))
        matching = [r for r in rows if r and r[9] == "nettotest1"]
        self.assertEqual(sum(r[-1] for r in matching), 0)

    def test_reversal_shown_as_negative_amount(self):
        t = Transaction.objects.create(date=dt.date(2026, 6, 11), amount=Decimal("400"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="AUTO",
            department=self.d, reference="nettotest2", core_ref="NETTOTEST2")
        t.reverse(self.tr, reason="test")
        rows = _rows(self.c.get("/transactions/?export=xlsx"))
        matching = [r for r in rows if r and r[9] == "nettotest2"]
        amounts = sorted(r[-1] for r in matching)
        self.assertEqual(amounts, [-400, 400])

    def test_entry_status_column_identifies_reversal_and_reversed(self):
        t = Transaction.objects.create(date=dt.date(2026, 6, 12), amount=Decimal("250"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="AUTO",
            department=self.d, reference="nettotest3", core_ref="NETTOTEST3")
        t.reverse(self.tr, reason="test")
        rows = _rows(self.c.get("/transactions/?export=xlsx"))
        matching = {r[-2]: r for r in rows if r and r[9] == "nettotest3"}
        self.assertIn("Reversal", matching)
        self.assertIn("Reversed", matching)
        self.assertEqual(matching["Reversal"][-1], -250)
        self.assertEqual(matching["Reversed"][-1], 250)

    def test_unrelated_entry_sharing_reference_never_wrongly_grouped(self):
        """The core scenario: a manually-entered cash transaction (no
        core_ref/mpesa_ref) that gets reversed, alongside a completely
        unrelated cash gift sharing the same reference/date/direction —
        must never be combined."""
        t = Transaction.objects.create(date=dt.date(2026, 6, 13), amount=Decimal("300"),
            direction="CREDIT", confirmed=True, channel="CASH", allocation_status="MANUAL",
            department=self.d, reference="cashnettotest")
        t.reverse(self.tr, reason="test")
        Transaction.objects.create(date=dt.date(2026, 6, 13), amount=Decimal("777"),
            direction="CREDIT", confirmed=True, channel="CASH", allocation_status="MANUAL",
            department=self.d, reference="cashnettotest")
        rows = _rows(self.c.get("/transactions/?export=xlsx"))
        matching = [r for r in rows if r and r[9] == "cashnettotest"]
        self.assertEqual(len(matching), 3)
        amounts = sorted(r[-1] for r in matching)
        self.assertEqual(amounts, [-300, 300, 777])

    def test_genuine_split_still_combines_correctly_unaffected(self):
        from giving.models import SplitFund, SplitComponent, AllocationRule
        d2 = Department.objects.create(name="ExportRevFund2", fund_type="LOCAL",
            category="MINISTRY")
        sf = SplitFund.objects.create(name="ExportRevSplit")
        SplitComponent.objects.create(split_fund=sf, department=self.d, percent=Decimal("60"))
        SplitComponent.objects.create(split_fund=sf, department=d2, percent=Decimal("40"))
        AllocationRule.objects.create(reference="splitunaffected", split_fund=sf, source="LEARNED")
        Transaction.objects.create(date=dt.date(2026, 6, 14), amount=Decimal("600"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="AUTO",
            department=self.d, reference="splitunaffected", core_ref="SPLITUNAFFECTED")
        Transaction.objects.create(date=dt.date(2026, 6, 14), amount=Decimal("400"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="AUTO",
            department=d2, reference="splitunaffected", core_ref="SPLITUNAFFECTED-S1")
        rows = _rows(self.c.get("/transactions/?export=xlsx"))
        matching = [r for r in rows if r and r[9] == "splitunaffected"]
        self.assertEqual(len(matching), 1)
        self.assertEqual(matching[0][-1], 1000)

    def test_csv_export_also_fixed(self):
        t = Transaction.objects.create(date=dt.date(2026, 6, 15), amount=Decimal("100"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="AUTO",
            department=self.d, reference="csvnettotest", core_ref="CSVNETTOTEST")
        t.reverse(self.tr, reason="test")
        b = self.c.get("/transactions/?export=csv").content.decode()
        self.assertIn("-100", b)
        self.assertIn("Reversal", b)
        self.assertIn("Reversed", b)
