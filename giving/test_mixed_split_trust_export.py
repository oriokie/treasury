"""Bug fix: the Trust Fund Pending Receipts export pre-filtered to
fund_type=TRUST before grouping split siblings, so a mixed split (e.g.
"Combined Offering" = 50% trust ENF + 50% local LCB) only ever showed the
trust-side partial amount — a giver's 40 showed as 20. The department-type
filter must apply at the GROUP level (include if any sibling is a trust
credit) so the full original gift amount is shown, since the giver's receipt
covers the whole gift regardless of the internal accounting split."""
import datetime as dt
from decimal import Decimal
from django.test import TestCase, Client
from django.contrib.auth.models import User, Group
from departments.models import Department
from giving.models import Transaction, SplitFund, SplitComponent, AllocationRule


def _tr():
    u = User.objects.create_user("tr_mixedsplit", password="x", is_superuser=True)
    u.groups.add(Group.objects.get_or_create(name="Treasurer")[0])
    return u


class MixedSplitTrustExportTests(TestCase):
    def setUp(self):
        self.tr = _tr()
        self.c = Client(); self.c.force_login(self.tr)

    def _rows(self, response):
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content))
        return list(wb.active.iter_rows(values_only=True))

    def _make_split(self, ref, name, enf_type, lcb_type, enf_amt, lcb_amt,
                     enf_receipted=False, lcb_receipted=False):
        enf = Department.objects.create(name=f"{name}ENF", fund_type=enf_type,
            category="TRUST" if enf_type == "TRUST" else "OFFERING")
        lcb = Department.objects.create(name=f"{name}LCB", fund_type=lcb_type,
            category="TRUST" if lcb_type == "TRUST" else "OFFERING")
        sf = SplitFund.objects.create(name=name)
        SplitComponent.objects.create(split_fund=sf, department=enf, percent=Decimal("50"))
        SplitComponent.objects.create(split_fund=sf, department=lcb, percent=Decimal("50"))
        AllocationRule.objects.create(reference=ref, split_fund=sf, source="LEARNED")
        Transaction.objects.create(date=dt.date(2026, 6, 10), amount=enf_amt,
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=enf, reference=ref, core_ref=ref.upper(), mpesa_ref=ref.upper(),
            manual_receipt=enf_receipted)
        Transaction.objects.create(date=dt.date(2026, 6, 10), amount=lcb_amt,
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=lcb, reference=ref, core_ref=f"{ref.upper()}-S1",
            mpesa_ref=ref.upper(), manual_receipt=lcb_receipted)
        return enf, lcb

    def test_mixed_trust_local_split_shows_full_original_amount(self):
        self._make_split("ug44gartjd", "COMBINED", "TRUST", "LOCAL",
                         Decimal("20"), Decimal("20"))
        rows = self._rows(self.c.get("/transactions/?export=pending-receipt"))
        match = next(r for r in rows if r and r[5] == "ug44gartjd")
        self.assertEqual(match[3], 40)
        self.assertEqual(match[4], "COMBINED")

    def test_a_split_whose_local_half_is_LCB_is_now_INCLUDED(self):
        """Behaviour deliberately changed. This split's halves are both LOCAL
        by fund_type — but one of them is the Local Church Budget (that is what
        the "LCB" half of an ENF/LCB combined offering IS), and a church
        receipts LCB money exactly as it receipts trust money.

        The pending-receipt list used to be Trust-ONLY, so this was excluded —
        which is precisely the gap reported: LCB gifts a church does receipt
        never appeared on the list of things awaiting a receipt. Receiptable is
        now Trust + the LCB family.
        """
        self._make_split("purelocalref", "PureLocal", "LOCAL", "LOCAL",
                         Decimal("10"), Decimal("10"))
        rows = self._rows(self.c.get("/transactions/?export=pending-receipt"))
        self.assertTrue(any(r and r[5] == "purelocalref" for r in rows),
                        "an unreceipted LCB gift must appear on the pending-receipt list")

    def test_a_gift_to_a_plain_local_fund_is_still_excluded(self):
        """The scope widened, but it did not become everything: an ordinary
        local fund — no trust, no LCB — is still not a receipting concern.
        (_make_split cannot be used here: it names one half "…LCB", which under
        the fallback IS the Local Church Budget.)"""
        d = Department.objects.create(name="PlainYouthFund", fund_type="LOCAL",
                                      category="OFFERING")
        Transaction.objects.create(
            date=dt.date(2026, 6, 10), amount=Decimal("40"), direction="CREDIT",
            confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=d, reference="plainlocalref")
        rows = self._rows(self.c.get("/transactions/?export=pending-receipt"))
        self.assertFalse(any(r and r[5] == "plainlocalref" for r in rows))

    def test_pure_trust_split_unaffected(self):
        self._make_split("puretrustref", "PureTrust", "TRUST", "TRUST",
                         Decimal("300"), Decimal("200"))
        rows = self._rows(self.c.get("/transactions/?export=pending-receipt"))
        match = next(r for r in rows if r and r[5] == "puretrustref")
        self.assertEqual(match[3], 500)

    def test_standalone_trust_transaction_unaffected(self):
        d = Department.objects.create(name="StandaloneTrustX", fund_type="TRUST",
            category="OFFERING")
        Transaction.objects.create(date=dt.date(2026, 6, 11), amount=Decimal("777"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=d, reference="standaloneref")
        rows = self._rows(self.c.get("/transactions/?export=pending-receipt"))
        match = next(r for r in rows if r and r[5] == "standaloneref")
        self.assertEqual(match[3], 777)

    def test_standalone_local_transaction_excluded(self):
        d = Department.objects.create(name="StandaloneLocalX", fund_type="LOCAL",
            category="MINISTRY")
        Transaction.objects.create(date=dt.date(2026, 6, 12), amount=Decimal("444"),
            direction="CREDIT", confirmed=True, channel="CASH", allocation_status="MANUAL",
            department=d, reference="standalonelocalref2")
        rows = self._rows(self.c.get("/transactions/?export=pending-receipt"))
        self.assertFalse(any(r and r[5] == "standalonelocalref2" for r in rows))

    def test_fully_receipted_mixed_split_excluded(self):
        self._make_split("fullyreceiptedref", "FullyReceipted", "TRUST", "LOCAL",
                         Decimal("20"), Decimal("20"),
                         enf_receipted=True, lcb_receipted=True)
        rows = self._rows(self.c.get("/transactions/?export=pending-receipt"))
        self.assertFalse(any(r and r[5] == "fullyreceiptedref" for r in rows))

    def test_partially_receipted_mixed_split_still_shows_full_total(self):
        self._make_split("partialreceiptref", "PartialReceipt", "TRUST", "LOCAL",
                         Decimal("15"), Decimal("15"), enf_receipted=True)
        rows = self._rows(self.c.get("/transactions/?export=pending-receipt"))
        match = next(r for r in rows if r and r[5] == "partialreceiptref")
        self.assertEqual(match[3], 30)
