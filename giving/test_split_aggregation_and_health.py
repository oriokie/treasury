"""Follow-up fixes:
3. The main Transactions Excel/CSV export now aggregates split-fund siblings
   into one combined row (matching the trust-pending-receipt export's
   existing behaviour), instead of showing each partial line separately.
4. The General Ledger Health Check's "shared M-Pesa/bank reference" check
   previously compared core_ref values directly, but split_into() gives
   every sibling its OWN distinct core_ref (base + "-S1", "-S2", ...) — so a
   legitimate split was never actually recognised as one, always flagged as
   "worth checking". Fixed to compare the base reference (stripping the
   "-S<n>" suffix) instead.
6. Missing postings persisted after every rebuild for reversed/reversal
   transactions, which post_transaction() (and therefore rebuild()) both
   correctly decline to post by design — a false positive, not a real gap.
   Fixed by excluding them from the check. Each remaining genuinely-missing
   item now links to its record."""
import datetime as dt
from decimal import Decimal
from django.test import TestCase, Client
from django.contrib.auth.models import User, Group
from departments.models import Department
from giving.models import Transaction, SplitFund, SplitComponent, AllocationRule
from ledger.services.posting import ensure_chart, rebuild
from ledger.services.health import missing_source_documents, duplicate_references
from ledger.models import JournalEntry


def _tr():
    u = User.objects.create_user("tr_splitagg", password="x", is_superuser=True)
    u.groups.add(Group.objects.get_or_create(name="Treasurer")[0])
    return u


class TransactionExportSplitAggregationTests(TestCase):
    def setUp(self):
        self.tr = _tr()
        self.d1 = Department.objects.create(name="SplitAggFund1", fund_type="LOCAL",
            category="MINISTRY")
        self.d2 = Department.objects.create(name="SplitAggFund2", fund_type="LOCAL",
            category="MINISTRY")
        self.c = Client(); self.c.force_login(self.tr)

    def _rows(self, response):
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content))
        return list(wb.active.iter_rows(values_only=True))

    def test_split_siblings_combined_into_one_row(self):
        sf = SplitFund.objects.create(name="Export Split Fund")
        SplitComponent.objects.create(split_fund=sf, department=self.d1, percent=Decimal("60"))
        SplitComponent.objects.create(split_fund=sf, department=self.d2, percent=Decimal("40"))
        AllocationRule.objects.create(reference="exportsplitref", split_fund=sf, source="LEARNED")
        Transaction.objects.create(date=dt.date(2026, 6, 10), amount=Decimal("600"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=self.d1, reference="exportsplitref", core_ref="EXPSPLIT-S1")
        Transaction.objects.create(date=dt.date(2026, 6, 10), amount=Decimal("400"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=self.d2, reference="exportsplitref", core_ref="EXPSPLIT-S2")
        rows = self._rows(self.c.get("/transactions/?export=xlsx"))
        matching = [r for r in rows if r and r[9] == "exportsplitref"]
        self.assertEqual(len(matching), 1)
        self.assertEqual(matching[0][-1], 1000)
        self.assertEqual(matching[0][7], "Export Split Fund")

    def test_non_split_transaction_unaffected(self):
        Transaction.objects.create(date=dt.date(2026, 6, 11), amount=Decimal("250"),
            direction="CREDIT", confirmed=True, channel="CASH", allocation_status="MANUAL",
            department=self.d1, reference="plainref")
        rows = self._rows(self.c.get("/transactions/?export=xlsx"))
        matching = [r for r in rows if r and r[9] == "plainref"]
        self.assertEqual(len(matching), 1)
        self.assertEqual(matching[0][-1], 250)

    def test_csv_export_also_aggregates(self):
        sf = SplitFund.objects.create(name="CSV Split Fund")
        SplitComponent.objects.create(split_fund=sf, department=self.d1, percent=Decimal("50"))
        SplitComponent.objects.create(split_fund=sf, department=self.d2, percent=Decimal("50"))
        AllocationRule.objects.create(reference="csvsplitref", split_fund=sf, source="LEARNED")
        Transaction.objects.create(date=dt.date(2026, 6, 12), amount=Decimal("300"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=self.d1, reference="csvsplitref", core_ref="CSVSPLIT-S1")
        Transaction.objects.create(date=dt.date(2026, 6, 12), amount=Decimal("300"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=self.d2, reference="csvsplitref", core_ref="CSVSPLIT-S2")
        b = self.c.get("/transactions/?export=csv").content.decode()
        self.assertEqual(b.count("csvsplitref"), 1)
        self.assertIn("600", b)


class LedgerHealthSplitReferenceTests(TestCase):
    def test_split_siblings_recognised_via_base_reference(self):
        d1 = Department.objects.create(name="HealthSplitFund1", fund_type="LOCAL",
            category="MINISTRY")
        d2 = Department.objects.create(name="HealthSplitFund2", fund_type="LOCAL",
            category="MINISTRY")
        Transaction.objects.create(date=dt.date(2026, 6, 10), amount=Decimal("600"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=d1, mpesa_ref="HEALTHSHARED", core_ref="HEALTHCORE-S1")
        Transaction.objects.create(date=dt.date(2026, 6, 10), amount=Decimal("400"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=d2, mpesa_ref="HEALTHSHARED", core_ref="HEALTHCORE-S2")
        refs = duplicate_references()
        match = next(r for r in refs if r["mpesa_ref"] == "HEALTHSHARED")
        self.assertTrue(match["likely_split"])

    def test_genuinely_different_transactions_still_flagged(self):
        d1 = Department.objects.create(name="HealthGenuineFund1", fund_type="LOCAL",
            category="MINISTRY")
        d2 = Department.objects.create(name="HealthGenuineFund2", fund_type="LOCAL",
            category="MINISTRY")
        Transaction.objects.create(date=dt.date(2026, 6, 10), amount=Decimal("600"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=d1, mpesa_ref="GENUINESHARED", core_ref="TOTALLYDIFFERENT1")
        Transaction.objects.create(date=dt.date(2026, 6, 11), amount=Decimal("400"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=d2, mpesa_ref="GENUINESHARED", core_ref="TOTALLYDIFFERENT2")
        refs = duplicate_references()
        match = next(r for r in refs if r["mpesa_ref"] == "GENUINESHARED")
        self.assertFalse(match["likely_split"])

    def test_health_page_shows_split_count_separately(self):
        tr = _tr()
        d1 = Department.objects.create(name="HealthPageFund1", fund_type="LOCAL",
            category="MINISTRY")
        d2 = Department.objects.create(name="HealthPageFund2", fund_type="LOCAL",
            category="MINISTRY")
        Transaction.objects.create(date=dt.date(2026, 6, 10), amount=Decimal("600"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=d1, mpesa_ref="PAGESHARED", core_ref="PAGECORE-S1")
        Transaction.objects.create(date=dt.date(2026, 6, 10), amount=Decimal("400"),
            direction="CREDIT", confirmed=True, channel="BANK", allocation_status="MANUAL",
            department=d2, mpesa_ref="PAGESHARED", core_ref="PAGECORE-S2")
        c = Client(); c.force_login(tr)
        b = c.get("/ledger/health/").content.decode()
        self.assertIn("legitimate split gift", b)


class MissingPostingsReversalExclusionTests(TestCase):
    def setUp(self):
        ensure_chart()
        self.tr = _tr()
        self.d = Department.objects.create(name="MissingRevFund", fund_type="LOCAL",
            category="MINISTRY")

    def test_reversed_transaction_not_flagged_as_missing(self):
        t = Transaction.objects.create(date=dt.date(2026, 6, 10), amount=Decimal("500"),
            direction="CREDIT", confirmed=True, channel="CASH", allocation_status="MANUAL",
            department=self.d, is_reversed=True)
        missing = missing_source_documents()
        self.assertNotIn(t.id, [x.id for x in missing["transactions"]])

    def test_reversed_transaction_not_flagged_even_after_rebuild(self):
        t = Transaction.objects.create(date=dt.date(2026, 6, 11), amount=Decimal("500"),
            direction="CREDIT", confirmed=True, channel="CASH", allocation_status="MANUAL",
            department=self.d, is_reversed=True)
        rebuild()
        missing = missing_source_documents()
        self.assertNotIn(t.id, [x.id for x in missing["transactions"]])

    def test_genuinely_missing_transaction_still_flagged(self):
        t = Transaction.objects.create(date=dt.date(2026, 6, 12), amount=Decimal("777"),
            direction="CREDIT", confirmed=True, channel="CASH", allocation_status="MANUAL",
            department=self.d)
        JournalEntry.objects.filter(source_type="transaction", source_id=t.id).delete()
        missing = missing_source_documents()
        self.assertIn(t.id, [x.id for x in missing["transactions"]])

    def test_missing_item_has_a_working_link(self):
        t = Transaction.objects.create(date=dt.date(2026, 6, 13), amount=Decimal("888"),
            direction="CREDIT", confirmed=True, channel="CASH", allocation_status="MANUAL",
            department=self.d)
        JournalEntry.objects.filter(source_type="transaction", source_id=t.id).delete()
        c = Client(); c.force_login(self.tr)
        b = c.get("/ledger/health/").content.decode()
        self.assertIn(f"/transactions/{t.id}/edit/", b)
