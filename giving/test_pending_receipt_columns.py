"""Pending receipt: what the downloads carry, and how a repeat is shown.

Two deliberate changes, both requested:

  * A repeated name is indicated by the row highlight ALONE. The "⚠ repeats"
    label that used to be appended to the name is gone from all three surfaces
    (page, Excel, PDF) — appending it also dirtied the Member column for
    anyone sorting or matching on it.
  * The Fund column is gone from the DOWNLOADS (Excel and PDF, and therefore
    the PDF the Telegram /pending command sends, which is the same function).
    It stays on the page, where you can also sort by it.

The service still derives the fund label — the page uses it — so dropping the
column is a presentation change, not a loss of data.
"""
import datetime as dt
import io
from decimal import Decimal

from django.contrib.auth.models import Group, User
from django.test import Client, TestCase

from core.roles import TREASURER
from departments.models import Department
from giving.models import Transaction
from giving.services.pending_receipt import (HEADER, duplicate_name_flags,
                                             export_rows,
                                             pending_receipt_pdf_bytes,
                                             pending_receipt_rows)


def _treasurer(username="prc_tr"):
    u = User.objects.create_user(username, password="x", is_superuser=True)
    u.groups.add(Group.objects.get_or_create(name=TREASURER)[0])
    return u


class _Seed(TestCase):
    def setUp(self):
        self.tr = _treasurer()
        self.c = Client()
        self.c.force_login(self.tr)
        self.trust = Department.objects.create(name="PrcTrust", fund_type="TRUST",
                                               category="OFFERING")

    def _credit(self, name, amount, ref, when=dt.date(2026, 6, 10), core=None):
        return Transaction.objects.create(
            date=when, amount=Decimal(amount), direction="CREDIT", confirmed=True,
            channel="BANK", allocation_status="MANUAL", department=self.trust,
            payer_name=name, reference=ref, core_ref=core or ref.upper())

    def _xlsx(self):
        from openpyxl import load_workbook
        r = self.c.get("/transactions/?export=pending-receipt")
        return list(load_workbook(io.BytesIO(r.content)).active.iter_rows(values_only=True))

    def _pdf_text(self):
        from pypdf import PdfReader
        data = pending_receipt_pdf_bytes(church="Test Church")
        return "\n".join(p.extract_text() or "" for p in PdfReader(io.BytesIO(data)).pages)


class DownloadColumnsTests(_Seed):
    def test_header_has_no_fund_column(self):
        self.assertNotIn("Fund", HEADER)
        self.assertEqual(HEADER, ["Date", "Phone", "Member", "Amount",
                                  "Reference", "M-Pesa Reference"])

    def test_export_rows_match_the_header_width(self):
        self._credit("ALPHA ONE", "100", "aref")
        rows = pending_receipt_rows()
        er = export_rows(rows)
        self.assertTrue(er)
        for row in er:
            self.assertEqual(len(row), len(HEADER))

    def test_export_rows_drop_only_the_fund(self):
        """Everything else survives, in order — the fund is the one column
        removed, not an arbitrary reshuffle."""
        self._credit("ALPHA ONE", "100", "aref")
        (date, phone, name, amount, fund, ref, mpesa) = pending_receipt_rows()[0]
        self.assertEqual(export_rows(pending_receipt_rows())[0],
                         [date, phone, name, amount, ref, mpesa])

    def test_the_service_still_derives_the_fund_label(self):
        """The page still shows Fund and sorts by it, so the data must remain."""
        self._credit("ALPHA ONE", "100", "aref")
        self.assertEqual(pending_receipt_rows()[0][4], "PrcTrust")

    def test_excel_has_no_fund_column(self):
        self._credit("ALPHA ONE", "100", "aref")
        rows = self._xlsx()
        header = next(r for r in rows if r and r[0] == "Date")
        self.assertEqual(list(header), HEADER)
        self.assertNotIn("Fund", [str(c) for c in header])
        self.assertFalse(any(r and "PrcTrust" in str(r) for r in rows),
                         "the fund name still appears somewhere in the Excel")

    def test_pdf_has_no_fund_column(self):
        self._credit("ALPHA ONE", "100", "aref")
        text = self._pdf_text()
        self.assertIn("Reference", text)
        self.assertNotIn("PrcTrust", text)

    def test_the_page_keeps_the_fund_column(self):
        self._credit("ALPHA ONE", "100", "aref")
        body = self.c.get("/transactions/pending-receipt/").content.decode()
        self.assertIn("<th>Fund</th>", body)
        self.assertIn("PrcTrust", body)


class RepeatWordingRemovedTests(_Seed):
    def setUp(self):
        super().setUp()
        self._credit("REPEAT PERSON", "100", "r1", core="RREF1")
        self._credit("REPEAT PERSON", "150", "r2",
                     when=dt.date(2026, 6, 12), core="RREF2")

    def test_the_duplicate_is_still_detected(self):
        rows = pending_receipt_rows()
        self.assertTrue(any(duplicate_name_flags(rows)))

    def test_excel_name_column_is_clean(self):
        rows = self._xlsx()
        names = [str(r[2]) for r in rows if r and r[2] and "REPEAT" in str(r[2])]
        self.assertEqual(len(names), 2)
        self.assertTrue(all(n == "REPEAT PERSON" for n in names), names)
        self.assertFalse(any("repeat" in n.lower().replace("repeat person", "")
                             for n in names))

    def test_excel_still_highlights_the_duplicate_rows(self):
        """The highlight is the whole signal now, so it must actually be
        applied — losing it would make duplicates invisible."""
        from openpyxl import load_workbook
        r = self.c.get("/transactions/?export=pending-receipt")
        ws = load_workbook(io.BytesIO(r.content)).active
        filled = [row for row in ws.iter_rows()
                  if any(c.fill and c.fill.fgColor and c.fill.fgColor.rgb
                         not in (None, "00000000") for c in row)]
        self.assertTrue(filled, "no row in the Excel carries a highlight fill")

    def test_pdf_has_no_repeats_wording(self):
        text = self._pdf_text()
        self.assertIn("REPEAT PERSON", text)
        self.assertNotIn("repeats", text.lower())

    def test_page_has_no_repeats_badge(self):
        body = self.c.get("/transactions/pending-receipt/").content.decode()
        self.assertIn("prv-dupe-row", body)          # the highlight is applied
        self.assertNotIn("prv-dupe-badge", body)     # the label is gone
        self.assertNotIn("⚠ repeats", body)


class PageGroupingTests(_Seed):
    def test_rows_are_grouped_by_giver_when_name_sorted(self):
        self._credit("ALPHA ONE", "100", "a1", core="A1")
        self._credit("ALPHA ONE", "200", "a2", core="A2")
        self._credit("BETA TWO", "300", "b1", core="B1")
        r = self.c.get("/transactions/pending-receipt/")
        rows = r.context["rows"]
        starts = [x["group_start"] for x in rows]
        self.assertEqual(sum(starts), 2, "expected one group start per giver")
        self.assertTrue(starts[0], "the first row starts a group")

    def test_grouping_is_off_for_other_sorts(self):
        self._credit("ALPHA ONE", "100", "a1", core="A1")
        r = self.c.get("/transactions/pending-receipt/?sort=amount")
        self.assertFalse(any(x["group_start"] for x in r.context["rows"]))

    def test_distinct_giver_count_is_reported(self):
        self._credit("ALPHA ONE", "100", "a1", core="A1")
        self._credit("ALPHA ONE", "200", "a2", core="A2")
        self._credit("BETA TWO", "300", "b1", core="B1")
        r = self.c.get("/transactions/pending-receipt/")
        self.assertEqual(r.context["distinct_givers"], 2)
        self.assertEqual(r.context["count"], 3)


class TelegramSharesTheSamePdfTests(_Seed):
    def test_the_bot_sends_the_same_columns(self):
        """/pending sends a PDF built by pending_receipt_pdf_bytes — the very
        function the web download uses — so the column change reaches Telegram
        without the bot needing to know about it."""
        self._credit("ALPHA ONE", "100", "aref")
        from core.services.telegram_bot import _do_pending
        msg = _do_pending(chat_id=1)
        self.assertEqual(msg["filename"], "pending_receipt.pdf")
        from pypdf import PdfReader
        text = "\n".join(p.extract_text() or ""
                         for p in PdfReader(io.BytesIO(msg["document"])).pages)
        self.assertNotIn("PrcTrust", text)      # no Fund column
        self.assertNotIn("repeats", text.lower())
