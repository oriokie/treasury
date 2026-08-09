"""Critical fix: the Transactions Excel/CSV export merged two DIFFERENT
givers into one row.

`_group_split_siblings` ended in a "same reference text + same date + same
direction" fallback. A manually-entered CASH gift has no core_ref and no
mpesa_ref, and its reference is optional on CashEntryForm — routinely blank
for a walk-in giver — so two unrelated cash gifts taken the same day matched
that fallback on an EMPTY reference and were combined: the amounts summed,
the funds joined into "Fund1 + Fund2", and only the FIRST giver's name
printed. Alice's 100 and Bob's 250 left the building as a single 350 from
ALICE, and Bob appeared nowhere in the export.

The same trap with a non-empty reference is no better: "tithe" and
"offering" are the most common things a giver writes, which is precisely why
Transaction.strict_split_siblings() refuses to treat reference text as an
identifier at all. So sibling-hood is now decided only by things that cannot
belong to two gifts — the `split_of` link, the core_ref base, the M-Pesa
reference — and anything else groups with itself alone.

giving/test_export_reversal_fix.py's
test_unrelated_entry_sharing_reference_never_wrongly_grouped covers the
neighbouring case (a reversal must not be absorbed into an unrelated row);
these tests cover the plain two-givers case it never reached, plus the thing
that must NOT regress: a genuine split of a cash entry, which has no bank
identifier of any kind and is held together purely by `split_of`.

Leaning that hard on `split_of` then exposed the mirror-image fault, covered
by NestedSplitIsStillOneGiftTests: the key was the row's IMMEDIATE parent, so
a split of a split (X -> X-S1 -> X-S1-S1 — an ordinary second cut, since
split_into() stamps `split_of` on whichever row it was called on) put the
grandchild in a group of its own and one gift left as two rows with its
amount torn between them. GenuineSplitsStillGroupTests holds the other
direction down across both export formats, because under-grouping is the
quieter failure: an export that shows every gift in pieces looks plausible.
"""
import contextlib
import datetime as dt
import signal
import threading
from decimal import Decimal
from django.test import TestCase, Client
from django.contrib.auth.models import User, Group
from departments.models import Department
from giving.models import Transaction
from giving.views import _group_split_siblings

DAY = dt.date(2026, 7, 4)

# column positions in the export header, by name (see TransactionListView.get)
PAYER, FUND, REFERENCE, AMOUNT = 4, 7, 9, -1


def _tr():
    u = User.objects.create_user("tr_blankref", password="x", is_superuser=True)
    u.groups.add(Group.objects.get_or_create(name="Treasurer")[0])
    return u


def _rows(response):
    """The export's data rows — the workbook carries church/title rows above
    the header, so everything up to and including the header is dropped."""
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(response.content))
    all_rows = list(wb.active.iter_rows(values_only=True))
    for i, row in enumerate(all_rows):
        if row and "Date" in row:
            return all_rows[i + 1:]
    return all_rows


@contextlib.contextmanager
def _hard_deadline(seconds, what):
    """Turn a hang into an ordinary test failure.

    The cycle test below cannot be written as a plain assertion: without the
    `seen` guard in _root() the walk never returns, so the assertion is never
    reached and the whole suite simply stops with no message. SIGALRM
    interrupts the loop and reports it instead. Only the main thread can take
    a signal, so anywhere else the body still runs, just unbounded.
    """
    if threading.current_thread() is not threading.main_thread():
        yield
        return

    def _fire(signum, frame):
        raise AssertionError(
            f"{what} did not finish within {seconds}s — a `split_of` cycle is "
            "being walked without a visited-set guard.")

    previous = signal.signal(signal.SIGALRM, _fire)
    signal.setitimer(signal.ITIMER_REAL, seconds)
    try:
        yield
    finally:
        signal.setitimer(signal.ITIMER_REAL, 0)
        signal.signal(signal.SIGALRM, previous)


class UnrelatedCashGiftsAreNeverMergedTests(TestCase):
    def setUp(self):
        self.tr = _tr()
        self.c = Client(); self.c.force_login(self.tr)
        self.f1 = Department.objects.create(name="BlankRefFund1", fund_type="LOCAL",
            category="MINISTRY")
        self.f2 = Department.objects.create(name="BlankRefFund2", fund_type="LOCAL",
            category="MINISTRY")

    def _cash(self, payer, amount, dept, reference=""):
        return Transaction.objects.create(
            date=DAY, amount=Decimal(amount), direction="CREDIT", confirmed=True,
            channel="CASH", allocation_status="MANUAL", department=dept,
            payer_name=payer, reference=reference)

    def test_two_walk_in_cash_gifts_with_no_reference_stay_two_rows(self):
        """Alice's 100 and Bob's 250, same day, no reference on either (the
        normal state of a walk-in cash gift) — two givers, two funds, so two
        rows. Before the fix this was ONE row reading ALICE / "BlankRefFund1 +
        BlankRefFund2" / 350, and Bob was not in the export at all."""
        self._cash("Alice Wanjiru", "100", self.f1)
        self._cash("Bob Otieno", "250", self.f2)
        rows = [r for r in _rows(self.c.get("/transactions/?export=xlsx")) if r and r[PAYER]]
        by_payer = {r[PAYER]: r for r in rows}
        self.assertEqual(sorted(by_payer), ["ALICE WANJIRU", "BOB OTIENO"])
        self.assertEqual(by_payer["ALICE WANJIRU"][AMOUNT], 100)
        self.assertEqual(by_payer["BOB OTIENO"][AMOUNT], 250)
        self.assertEqual(by_payer["ALICE WANJIRU"][FUND], "BlankRefFund1")
        self.assertEqual(by_payer["BOB OTIENO"][FUND], "BlankRefFund2")

    def test_two_cash_gifts_sharing_the_word_tithe_stay_two_rows(self):
        """A non-empty reference is no safer: "tithe" is free text two
        unrelated people write on the same Sabbath, not an identifier. Same
        invariant as strict_split_siblings()' docstring states."""
        self._cash("Clara Njeri", "300", self.f1, reference="tithe")
        self._cash("Daniel Kip", "700", self.f2, reference="tithe")
        rows = [r for r in _rows(self.c.get("/transactions/?export=xlsx"))
                if r and r[REFERENCE] == "tithe"]
        self.assertEqual(len(rows), 2)
        self.assertEqual(sorted(r[PAYER] for r in rows), ["CLARA NJERI", "DANIEL KIP"])
        self.assertEqual(sorted(r[AMOUNT] for r in rows), [300, 700])

    def test_the_csv_export_keeps_both_givers_too(self):
        """The CSV goes through the same grouping, so it must not be the one
        surface where a giver still disappears."""
        self._cash("Esther Auma", "120", self.f1)
        self._cash("Felix Mwangi", "480", self.f2)
        body = self.c.get("/transactions/?export=csv").content.decode()
        self.assertIn("ESTHER AUMA", body)
        self.assertIn("FELIX MWANGI", body)
        self.assertNotIn("BlankRefFund1 + BlankRefFund2", body)

    def test_a_real_cash_split_still_comes_back_as_one_row(self):
        """The other half of the invariant, and the reason the loose reference
        match could be dropped safely: a cash entry split across two funds has
        no core_ref and no mpesa_ref, but split_into() records `split_of` on
        the part it creates — the authoritative link — so the export still
        recombines the gift into the single 400 it was, under the two funds it
        went to, even with a blank reference."""
        t = self._cash("Grace Kamau", "400", self.f1)
        t.split_into([(self.f1, Decimal("250"), None), (self.f2, Decimal("150"), None)])
        rows = [r for r in _rows(self.c.get("/transactions/?export=xlsx"))
                if r and r[PAYER] == "GRACE KAMAU"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][AMOUNT], 400)
        self.assertEqual(rows[0][FUND], "BlankRefFund1 + BlankRefFund2")

    def test_a_split_part_is_not_joined_by_an_unrelated_blank_reference_gift(self):
        """Both rules at once: the two halves of Grace's split combine, and
        Henry's unrelated same-day cash gift — which shares their blank
        reference, date and direction — stays his own row."""
        t = self._cash("Grace Kamau", "400", self.f1)
        t.split_into([(self.f1, Decimal("250"), None), (self.f2, Decimal("150"), None)])
        self._cash("Henry Owino", "60", self.f2)
        rows = [r for r in _rows(self.c.get("/transactions/?export=xlsx")) if r and r[PAYER]]
        self.assertEqual(sorted((r[PAYER], r[AMOUNT]) for r in rows),
                         [("GRACE KAMAU", 400), ("HENRY OWINO", 60)])


class NestedSplitIsStillOneGiftTests(TestCase):
    """A part of a split can be split again — the treasurer divides a 400
    deposit into two funds, then realises the 150 half belongs to two funds of
    its own. split_into() stamps `split_of` on whatever row it was called on,
    so the second cut hangs off the FIRST CUT'S CHILD, not off the original:
    the chain is X -> X-S1 -> X-S1-S1.

    Keying the export on the immediate `split_of` therefore gave the
    grandchild a group of its own, and one gift came out as two rows with its
    amount torn between them — each row's fund label naming only part of where
    the money actually went, and neither row's total being the gift. The key
    now resolves the chain to its root.
    """

    def setUp(self):
        self.tr = _tr()
        self.c = Client(); self.c.force_login(self.tr)
        self.f1 = Department.objects.create(name="NestFundA", fund_type="LOCAL",
            category="MINISTRY")
        self.f2 = Department.objects.create(name="NestFundB", fund_type="LOCAL",
            category="MINISTRY")
        self.f3 = Department.objects.create(name="NestFundC", fund_type="LOCAL",
            category="MINISTRY")

    def _cash(self, payer, amount, dept, reference=""):
        return Transaction.objects.create(
            date=DAY, amount=Decimal(amount), direction="CREDIT", confirmed=True,
            channel="CASH", allocation_status="MANUAL", department=dept,
            payer_name=payer, reference=reference)

    def _split_of_a_split(self, payer):
        """400 -> (250 A, 150 B), then that 150 B -> (100 B, 50 C).
        Returns (original, middle, grandchild)."""
        root = self._cash(payer, "400", self.f1)
        middle = root.split_into([(self.f1, Decimal("250"), None),
                                  (self.f2, Decimal("150"), None)])[1]
        grandchild = middle.split_into([(self.f2, Decimal("100"), None),
                                        (self.f3, Decimal("50"), None)])[1]
        return root, middle, grandchild

    def test_a_split_of_a_split_exports_as_one_row_of_the_whole_gift(self):
        self._split_of_a_split("Ivy Nyambura")
        rows = [r for r in _rows(self.c.get("/transactions/?export=xlsx"))
                if r and r[PAYER] == "IVY NYAMBURA"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][AMOUNT], 400)
        # every fund the gift reached, including the one only the second cut
        # touched — the grandchild used to be exported as a separate 50
        self.assertEqual(rows[0][FUND], "NestFundA + NestFundB + NestFundC")

    def test_a_bank_nested_split_is_one_row_too(self):
        """The `split_of` key is consulted BEFORE core_ref, so a bank gift is
        not saved by having identifiers: X-S1-S1's core_ref would have shared
        the root's base, but the split branch returned first and split it off
        anyway. Same chain, same single row."""
        root = Transaction.objects.create(
            date=DAY, amount=Decimal("900"), direction="CREDIT", confirmed=True,
            channel="BANK", allocation_status="AUTO", department=self.f1,
            payer_name="Joseph Kariuki", reference="nestbank", core_ref="NESTBANK")
        middle = root.split_into([(self.f1, Decimal("600"), None),
                                  (self.f2, Decimal("300"), None)])[1]
        middle.split_into([(self.f2, Decimal("200"), None),
                           (self.f3, Decimal("100"), None)])
        rows = [r for r in _rows(self.c.get("/transactions/?export=xlsx"))
                if r and r[REFERENCE] == "nestbank"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][AMOUNT], 900)

    def test_the_csv_export_combines_the_nested_split_as_well(self):
        """Both export formats build their rows from one pass of
        _group_split_siblings and differ only in how they are serialised, so
        neither may be the surface where a gift arrives in pieces."""
        self._split_of_a_split("Kevin Ochieng")
        body = self.c.get("/transactions/?export=csv").content.decode()
        gift_rows = [ln for ln in body.splitlines() if "KEVIN OCHIENG" in ln]
        self.assertEqual(len(gift_rows), 1)
        self.assertIn("400", gift_rows[0])
        self.assertIn("NestFundA + NestFundB + NestFundC", gift_rows[0])

    def test_a_missing_middle_row_does_not_glue_the_ends_together(self):
        """Grouping never reaches for a row the caller did not pass in — an
        export filtered to one fund must aggregate exactly what is on screen.
        With the middle of the chain filtered out, the walk stops at the
        absent row, so the original and the grandchild stay separate rather
        than being silently recombined from rows the filter excluded."""
        root, middle, grandchild = self._split_of_a_split("Lydia Chebet")
        groups = _group_split_siblings([root, grandchild])
        self.assertEqual([[t.pk for t in g] for g in groups],
                         [[root.pk], [grandchild.pk]])

    def test_a_corrupt_split_of_cycle_terminates_instead_of_hanging(self):
        """`split_of` is a nullable self-FK with nothing in the database
        forbidding A -> B -> A. Walking to the root must survive that: this
        function also runs the transactions page and the pending-receipt
        export, so a cycle would hang those surfaces, not just this one."""
        a = self._cash("Cycle One", "10", self.f1)
        b = self._cash("Cycle Two", "20", self.f2)
        # .update() so the cycle is written exactly as corrupt data would be
        Transaction.objects.filter(pk=a.pk).update(split_of=b)
        Transaction.objects.filter(pk=b.pk).update(split_of=a)
        a.refresh_from_db(); b.refresh_from_db()
        with _hard_deadline(10, "_group_split_siblings"):
            groups = _group_split_siblings([a, b])
        # no row may be lost or duplicated, whatever the cycle does to grouping
        self.assertEqual(sorted(t.pk for g in groups for t in g),
                         sorted([a.pk, b.pk]))


class GenuineSplitsStillGroupTests(TestCase):
    """The other half of the invariant, and the more dangerous direction to
    regress: tightening what counts as a sibling must not stop REAL siblings
    combining. An over-grouped export is loudly wrong; an under-grouped one
    just quietly shows every gift in pieces and no one notices.
    """

    def setUp(self):
        self.tr = _tr()
        self.c = Client(); self.c.force_login(self.tr)
        self.f1 = Department.objects.create(name="KeepGroupedA", fund_type="LOCAL",
            category="MINISTRY")
        self.f2 = Department.objects.create(name="KeepGroupedB", fund_type="LOCAL",
            category="MINISTRY")

    def test_the_csv_export_combines_a_real_cash_split_into_one_row(self):
        """The CSV is the export the treasurer actually reconciles in a
        spreadsheet. It shares the grouping pass with the workbook, and this
        pins that: a cash split (held together by nothing but `split_of`)
        must be one line there too."""
        t = Transaction.objects.create(
            date=DAY, amount=Decimal("500"), direction="CREDIT", confirmed=True,
            channel="CASH", allocation_status="MANUAL", department=self.f1,
            payer_name="Mercy Wambui")
        t.split_into([(self.f1, Decimal("320"), None), (self.f2, Decimal("180"), None)])
        body = self.c.get("/transactions/?export=csv").content.decode()
        gift_rows = [ln for ln in body.splitlines() if "MERCY WAMBUI" in ln]
        self.assertEqual(len(gift_rows), 1)
        self.assertIn("KeepGroupedA + KeepGroupedB", gift_rows[0])
        self.assertIn("500", gift_rows[0])

    def test_the_fund_lines_of_one_envelope_still_come_back_as_one_receipt(self):
        """Written the way envelopes/services/posting.py._save_envelope writes
        them: one ENVELOPE-channel row per fund line, no `split_of`, no bank
        identifier of any kind, sharing only the "envelope <receipt_no>"
        reference. That reference is all there is to rebuild the receipt from,
        which is why the ENVELOPE channel keeps its reference-based branch."""
        for dept, amt in ((self.f1, "700"), (self.f2, "300")):
            Transaction.objects.create(
                date=DAY, amount=Decimal(amt), direction="CREDIT", confirmed=True,
                channel="ENVELOPE", allocation_status="MANUAL", department=dept,
                payer_name="NAOMI ACHIENG", reference="envelope KG901",
                raw_narration="ENVELOPE KG901")
        rows = [r for r in _rows(self.c.get("/transactions/?export=xlsx"))
                if r and r[REFERENCE] == "envelope KG901"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][AMOUNT], 1000)
        self.assertEqual(rows[0][FUND], "KeepGroupedA + KeepGroupedB")

    def test_two_hand_keyed_envelope_channel_gifts_are_not_merged_by_their_words(self):
        """The original defect, surviving on a narrower path.

        The reference branch is safe because `_save_envelope` issues the text
        and `Envelope.receipt_no` is unique — but the CHANNEL does not prove the
        text was issued. `CashEntryForm` offers ENVELOPE alongside CASH with a
        free-text, optional reference, so a treasurer keying two members' gifts
        and typing "tithe" on both produced exactly the merge this whole
        function was fixed to stop: one row, the amounts summed, the second
        giver's name gone from the export handed to the auditors.
        """
        for name, amt in (("HANNAH W", "400"), ("PETER O", "600")):
            Transaction.objects.create(
                date=DAY, amount=Decimal(amt), direction="CREDIT", confirmed=True,
                channel="ENVELOPE", allocation_status="MANUAL", department=self.f1,
                payer_name=name, reference="tithe")
        rows = [r for r in _rows(self.c.get("/transactions/?export=xlsx"))
                if r and r[REFERENCE] == "tithe"]
        self.assertEqual(len(rows), 2,
                         "two hand-keyed envelope-channel gifts merged on the "
                         "word 'tithe' — a giver's name is missing from the export")
        self.assertEqual({r[PAYER] for r in rows}, {"HANNAH W", "PETER O"})

    def test_a_second_envelope_is_never_absorbed_into_the_first(self):
        """Envelope.receipt_no is unique, so the reference branch can only
        ever gather one receipt — the property the whole branch rests on."""
        Transaction.objects.create(
            date=DAY, amount=Decimal("700"), direction="CREDIT", confirmed=True,
            channel="ENVELOPE", allocation_status="MANUAL", department=self.f1,
            payer_name="OLIVE ADHIAMBO", reference="envelope KG902")
        Transaction.objects.create(
            date=DAY, amount=Decimal("300"), direction="CREDIT", confirmed=True,
            channel="ENVELOPE", allocation_status="MANUAL", department=self.f1,
            payer_name="PETER MUTISO", reference="envelope KG903")
        rows = [r for r in _rows(self.c.get("/transactions/?export=xlsx"))
                if r and (r[REFERENCE] or "").startswith("envelope KG90")]
        self.assertEqual(sorted((r[PAYER], r[AMOUNT]) for r in rows),
                         [("OLIVE ADHIAMBO", 700), ("PETER MUTISO", 300)])

    def test_two_statement_lines_of_one_mpesa_payment_still_combine(self):
        """The bank-side sibling rule: same M-Pesa receipt, same day — an
        identifier the bank issued, not text a payer typed."""
        for dept, amt in ((self.f1, "150"), (self.f2, "250")):
            Transaction.objects.create(
                date=DAY, amount=Decimal(amt), direction="CREDIT", confirmed=True,
                channel="BANK", allocation_status="AUTO", department=dept,
                payer_name="Quinter Atieno", reference="kgmpesa",
                mpesa_ref="KG9MPESA01")
        rows = [r for r in _rows(self.c.get("/transactions/?export=xlsx"))
                if r and r[REFERENCE] == "kgmpesa"]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][AMOUNT], 400)
