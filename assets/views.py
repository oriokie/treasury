import datetime as dt
import re
from decimal import Decimal

from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import ListView, CreateView, UpdateView

from core.permissions import ReadAccessMixin, TreasurerRequiredMixin, DataEntryRequiredMixin
from .models import FixedAsset, DepreciationRule
from .forms import FixedAssetForm, DepreciationRuleForm


def _as_of(request):
    try:
        return dt.date.fromisoformat(request.GET.get("as_of", ""))
    except ValueError:
        return dt.date.today()


class AssetListView(ReadAccessMixin, ListView):
    model = FixedAsset
    template_name = "assets/list.html"
    context_object_name = "assets"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        as_of = _as_of(self.request)
        rows, t_cost, t_acc, t_nbv = [], Decimal(0), Decimal(0), Decimal(0)
        for a in self.object_list:
            acc = a.accumulated_depreciation(as_of)
            nbv = a.net_book_value(as_of)
            rows.append({"a": a, "acc": acc, "nbv": nbv})
            t_cost += a.cost
            t_acc += acc
            t_nbv += nbv
        ctx.update({"rows": rows, "as_of": as_of, "t_cost": t_cost,
                    "t_acc": t_acc, "t_nbv": t_nbv})
        return ctx


class AssetCreate(TreasurerRequiredMixin, CreateView):
    model = FixedAsset
    form_class = FixedAssetForm
    template_name = "assets/form.html"
    success_url = reverse_lazy("asset_list")

    def form_valid(self, form):
        from .models import Acquisition
        resp = super().form_valid(form)
        asset = self.object
        source = form.cleaned_data.get("acq_source") or Acquisition.Source.PURCHASE
        Acquisition.objects.create(
            asset=asset, source=source,
            date=asset.acquired_on or dt.date.today(),
            amount=asset.cost or 0, fund=asset.department,
            donor_name=form.cleaned_data.get("donor_name", "") or "",
            recorded_by=self.request.user if self.request.user.is_authenticated else None)
        if source == Acquisition.Source.DONATION:
            messages.success(self.request, f"{asset.name} added to the register as a "
                                           f"donated asset — recognised at fair value.")
        else:
            messages.success(self.request, "Asset added to the register.")
        return resp


class AssetUpdate(TreasurerRequiredMixin, UpdateView):
    model = FixedAsset
    form_class = FixedAssetForm
    template_name = "assets/form.html"
    success_url = reverse_lazy("asset_list")

    def form_valid(self, form):
        messages.success(self.request, "Asset updated.")
        return super().form_valid(form)


class AssetDisposeView(TreasurerRequiredMixin, View):
    def post(self, request, pk):
        from django.db import transaction as db_tx
        from giving.models import Transaction
        from departments.models import Department
        from ledger.services import posting
        from .services import depreciation as dep
        from .services import lifecycle
        from . import signals as asset_signals
        a = get_object_or_404(FixedAsset, pk=pk)
        if a.disposed:
            messages.info(request, f"{a.name} is already disposed.")
            return redirect("asset_detail", pk=pk)
        # The same guard the lifecycle service applies to "held for disposal",
        # asked here because a disposal never goes through transition(): without
        # it the register could write off an asset that is still checked out,
        # and leave its assignment open for ever with nobody able to return it.
        try:
            lifecycle.check_not_issued(a, "recording the disposal")
        except lifecycle.TransitionError as exc:
            messages.error(request, str(exc))
            return redirect("asset_detail", pk=pk)
        try:
            on = dt.date.fromisoformat(request.POST.get("disposed_on", ""))
        except ValueError:
            on = dt.date.today()
        try:
            proceeds = Decimal(request.POST.get("proceeds") or "0")
        except Exception:
            from core.utils import log_exception as _lx; _lx('assets/views.py')
            proceeds = Decimal("0")
        method = request.POST.get("method") or FixedAsset.DisposalMethod.SOLD
        fund_id = request.POST.get("fund")
        # The fund is required: it receives any proceeds AND carries the gain or
        # loss, so every disposal is attributable to a fund. Without it the
        # proceeds could not be reclassified in the ledger.
        fund = Department.objects.filter(pk=fund_id).first() if fund_id else None
        if fund is None:
            messages.error(request, "Choose the fund that receives the proceeds "
                                    "and carries the gain or loss.")
            return redirect("asset_detail", pk=pk)
        # What the disposal is worth is decided HERE, once, and stored. The
        # proceeds and the gain/(loss) between them pin down the carrying value
        # it was struck against, and depreciation.accumulated_at_disposal reads
        # that back for everyone who asks later — the ledger's disposal journal,
        # the disposals report, the movement in fixed assets. Nothing downstream
        # re-runs the depreciation engine at the disposal date any more, so a
        # rate changed next week cannot restate a disposal recorded today.
        #
        # Both figures come from the depreciation service rather than from
        # net_book_value(), which reports zero from the disposal date onwards —
        # see carrying_value_at_disposal for what that costs the unwary.
        nbv = dep.carrying_value_at_disposal(a, on)
        gain_loss = dep.gain_or_loss_on_disposal(a, on, proceeds)
        a.disposed = True
        a.disposed_on = on
        a.disposal_proceeds = proceeds
        a.disposal_method = method
        a.disposal_gain_loss = gain_loss
        a.disposal_fund = fund
        # The cash proceeds are recorded as a receipt into the nominated fund so
        # fund balances and cash reports reflect the money received. The general
        # ledger entry then reclassifies that receipt via post_disposal —
        # removing the asset's cost and accumulated depreciation and recognising
        # the balancing gain/(loss) — so the proceeds are not double-counted and
        # the disposal is a proper journal.
        #
        # All three writes are one atomic act. A disposal used to reach the
        # ledger only when someone next rebuilt it; now it posts here, and if
        # that posting cannot be written the register is not left claiming a
        # disposal the ledger has never heard of — the whole thing is rolled
        # back and the treasurer told, rather than half-recorded in silence.
        try:
            with db_tx.atomic():
                a.save()
                if proceeds and fund:
                    Transaction.objects.create(
                        date=on, channel=Transaction.Channel.BANK,
                        direction=Transaction.Direction.CREDIT,
                        amount=proceeds, department=fund,
                        allocation_status=Transaction.Status.MANUAL,
                        excluded_from_income=True,
                        reference=f"Disposal of {a.name}"[:60],
                        payer_name=f"Asset disposal ({a.get_disposal_method_display()})"[:120])
                asset_signals.post_to_ledger(posting.post_disposal, a)
        except Exception:  # noqa: BLE001 — the detail is for the log, not the treasurer
            from core.utils import log_exception as _lx; _lx('assets/views.py')
            messages.error(request, f"{a.name} was NOT disposed of — the disposal could "
                                    f"not be posted to the general ledger, so nothing was "
                                    f"recorded. Please try again, or tell whoever supports "
                                    f"the system.")
            return redirect("asset_detail", pk=pk)
        verb = {"SOLD": "sold", "SCRAPPED": "scrapped", "DONATED": "donated",
                "LOST": "written off"}.get(method, "disposed")
        gl = (f"gain of {gain_loss}" if gain_loss > 0 else
              (f"loss of {-gain_loss}" if gain_loss < 0 else "no gain or loss"))
        messages.success(request, f"{a.name} {verb}. Net book value {nbv}, proceeds {proceeds} "
                                  f"— {gl}.")
        return redirect("asset_detail", pk=pk)


class DepreciationRulesView(TreasurerRequiredMixin, View):
    """Manage per-category depreciation rules (linked from settings)."""
    template_name = "assets/depreciation_rules.html"

    def get(self, request):
        from core.models import SiteConfig
        rules = {r.category: r for r in DepreciationRule.objects.all()}
        cats = [{"value": v, "label": l, "rule": rules.get(v)}
                for v, l in FixedAsset.Category.choices]
        return render(request, self.template_name,
                      {"cats": cats, "methods": DepreciationRule.Method.choices,
                       "cfg": SiteConfig.get()})

    def post(self, request):
        for v, _ in FixedAsset.Category.choices:
            method = request.POST.get(f"method_{v}")
            rate = request.POST.get(f"rate_{v}")
            if method is None:
                continue
            try:
                rate_val = Decimal(rate) if rate else Decimal(0)
            except Exception:
                from core.utils import log_exception as _lx; _lx('assets/views.py')
                rate_val = Decimal(0)
            DepreciationRule.objects.update_or_create(
                category=v, defaults={"method": method, "rate": rate_val})
        messages.success(request, "Depreciation rules saved.")
        return redirect("depreciation_rules")


from django.views.generic import DetailView
from .models import AssetAttachment


class AssetDetailView(ReadAccessMixin, DetailView):
    model = FixedAsset
    template_name = "assets/asset_detail.html"
    context_object_name = "asset"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        a = self.object
        ctx["nbv"] = a.net_book_value()
        ctx["accum"] = a.accumulated_depreciation()
        from departments.models import Department
        ctx["funds"] = Department.objects.filter(active=True, is_trust=False).order_by("name")
        ctx["today"] = dt.date.today()
        le = a.source_expenses.select_related("department").order_by("date")
        ctx["linked_expenses"] = le
        from decimal import Decimal as _D
        from django.db.models import Sum as _Sum
        ctx["linked_total"] = le.aggregate(t=_Sum("amount"))["t"] or _D(0)
        # --- Asset 360: lifecycle, custody, movement, and the journals it made ---
        from django.contrib.auth.models import User
        from .models import Location, AssetTransfer
        from .services import lifecycle
        ctx["allowed_transitions"] = [
            {"value": s, "label": FixedAsset.Status(s).label}
            for s in lifecycle.allowed_transitions(a)]
        ctx["open_assignment"] = lifecycle.open_assignment(a)
        ctx["assignments"] = a.assignments.select_related(
            "custodian", "location", "issued_by", "received_by")[:20]
        ctx["transfers"] = a.transfers.select_related(
            "from_location", "to_location", "from_fund", "to_fund", "approved_by")[:20]
        ctx["pending_transfers"] = [t for t in ctx["transfers"]
                                    if t.status == AssetTransfer.Status.PENDING]
        ctx["events"] = a.events.select_related("actor")[:40]
        ctx["locations"] = Location.objects.filter(active=True).order_by("name")
        ctx["users"] = User.objects.filter(is_active=True).order_by("username")
        ctx["all_funds"] = Department.objects.filter(active=True).order_by("name")
        ctx["acquisition"] = getattr(a, "acquisition", None)
        from ledger.models import JournalLine
        ctx["journal_lines"] = (JournalLine.objects.filter(
            entry__source_type__in=["asset_acq", "asset_disposal", "asset_transfer"],
            entry__source_id=a.pk)
            .select_related("entry", "account").order_by("entry__date", "id"))
        ctx["depreciation_lines"] = (a.depreciation_lines
                                     .select_related("run").order_by("-run__year", "-run__month")[:12])
        return ctx


class AssetAttachmentUpload(DataEntryRequiredMixin, View):
    def post(self, request, pk):
        a = get_object_or_404(FixedAsset, pk=pk)
        f = request.FILES.get("file")
        if f:
            AssetAttachment.objects.create(asset=a, file=f,
                label=request.POST.get("label", "")[:120], uploaded_by=request.user)
            messages.success(request, "Document attached.")
        else:
            messages.error(request, "Please choose a file to upload.")
        return redirect("asset_detail", pk=pk)


class AssetAttachmentDelete(DataEntryRequiredMixin, View):
    def post(self, request, pk, att):
        x = get_object_or_404(AssetAttachment, pk=att, asset_id=pk)
        x.file.delete(save=False)
        x.delete()
        messages.success(request, "Attachment removed.")
        return redirect("asset_detail", pk=pk)


class AssetAccumulateView(TreasurerRequiredMixin, View):
    """Accumulate a construction/building asset's cost from CAPITAL expenses.

    Only capital expenses NOT yet attached to any asset are picked up, then each
    is LINKED to this asset (capitalized_asset) and its amount added to the cost.
    Re-running is therefore safe — already-included expenses are skipped, so the
    cost can't be double-counted. The asset detail page lists exactly which
    expenses make up the cost. Manual editing of the cost stays available."""
    def post(self, request, pk):
        import datetime as dt
        from decimal import Decimal
        from django.db.models import Sum
        from django.shortcuts import get_object_or_404, redirect
        from django.contrib import messages
        from cashbook.models import Expense
        from departments.models import Department
        a = get_object_or_404(FixedAsset, pk=pk)
        fund_id = request.POST.get("fund")
        fund = Department.objects.filter(pk=fund_id).first() if fund_id else None
        if not fund:
            messages.error(request, "Choose the fund the construction is paid from.")
            return redirect("asset_detail", pk=a.pk)

        def _date(key):
            try:
                return dt.date.fromisoformat(request.POST.get(key))
            except (TypeError, ValueError):
                return None
        start, end = _date("start"), _date("end")
        qs = Expense.objects.filter(
            department=fund,
            expenditure_type=Expense.ExpenditureType.CAPITAL,
            status__in=[Expense.Status.APPROVED, Expense.Status.PAID],
            capitalized_asset__isnull=True)        # never re-add an already-linked expense
        if start:
            qs = qs.filter(date__gte=start)
        if end:
            qs = qs.filter(date__lte=end)
        total = qs.aggregate(t=Sum("amount"))["t"] or Decimal(0)
        n = qs.count()
        if not n:
            messages.info(request, "No new capital expenses to add — everything matching is "
                                   "already included in this asset's cost.")
            return redirect("asset_detail", pk=a.pk)
        qs.update(capitalized_asset=a)             # link them so they're exempt next time
        a.cost = (a.cost or Decimal(0)) + total
        a.save()
        rng = f" ({start or '...'} to {end or '...'})" if (start or end) else ""
        messages.success(request, f"Added {n} capital expense(s) totalling {total:,.2f} "
                                  f"from {fund.name}{rng}. Cost is now {a.cost:,.2f}.")
        return redirect("asset_detail", pk=a.pk)


from core.permissions import ReadAccessMixin as _ReadMixin  # noqa: E402
from core import roles as _roles  # noqa: E402


class DepreciationRunsView(_ReadMixin, View):
    """Monthly depreciation runs: review, generate, and post to the ledger.

    Viewing is read-access; generating and posting are treasurer actions. The
    register↔ledger reconciliation is shown at the top so the treasurer can see
    at a glance whether the control accounts agree with the register."""
    template_name = "assets/depreciation_runs.html"

    def get(self, request):
        from .models import DepreciationRun
        from core.metrics import metrics
        import datetime as _d
        today = _d.date.today()
        runs = list(DepreciationRun.objects.all()[:36])
        posted_months = {(r.year, r.month) for r in runs
                         if r.status != DepreciationRun.Status.DRAFT}
        # the next month that still needs a run (previous calendar month, if unposted)
        prev = today.replace(day=1) - _d.timedelta(days=1)
        rec = metrics.register_vs_ledger(today)
        reconciled = all(v["diff"] == 0 for v in rec.values())
        return render(request, self.template_name, {
            "runs": runs, "rec": rec, "reconciled": reconciled,
            "suggest_year": prev.year, "suggest_month": prev.month,
            "suggest_pending": (prev.year, prev.month) not in posted_months,
            "can_manage": _roles.can_approve(request.user),
        })

    def post(self, request):
        from .services import runs as run_svc
        if not _roles.can_approve(request.user):
            messages.error(request, "Posting depreciation requires the Treasurer role.")
            return redirect("depreciation_runs")
        action = request.POST.get("action")
        try:
            year = int(request.POST.get("year"))
            month = int(request.POST.get("month"))
        except (TypeError, ValueError):
            messages.error(request, "Choose a valid month.")
            return redirect("depreciation_runs")
        try:
            if action == "generate":
                run = run_svc.generate_run(year, month, user=request.user)
                messages.success(request, f"Generated draft depreciation for "
                                          f"{year}-{month:02d}: {run.lines.count()} assets, "
                                          f"{run.total_charge:,.2f}. Review, then post.")
            elif action in ("post", "generate_post"):
                run = run_svc.generate_run(year, month, user=request.user) \
                    if action == "generate_post" else \
                    __import__("assets.models", fromlist=["DepreciationRun"]) \
                    .DepreciationRun.objects.get(year=year, month=month)
                run_svc.post_run(run)
                messages.success(request, f"Posted depreciation for {year}-{month:02d} "
                                          f"to the ledger ({run.total_charge:,.2f}).")
        except Exception as e:  # noqa: BLE001 — surface the business-rule message
            messages.error(request, str(e))
        return redirect("depreciation_runs")


class ExpenseCapitaliseView(TreasurerRequiredMixin, View):
    """Convert a capital expense into a fixed asset (EAM §9.3).

    Capital purchases belong on the register, but they are entered day to day as
    payments. This turns such a payment into an asset: it creates the register
    record at the amount paid, links the expense to it (so the expense's own
    posting moves from capital work-in-progress to the fixed-asset account —
    no second entry, no double count), and records the acquisition.

    An existing asset can be nominated instead, in which case the payment is
    added to that asset's cost — the same "accumulate" behaviour used for
    construction, kept in one place.
    """
    def get(self, request, pk):
        from cashbook.models import Expense
        exp = get_object_or_404(Expense, pk=pk)
        return render(request, "assets/capitalise.html", {
            "expense": exp,
            "assets": FixedAsset.objects.filter(disposed=False).order_by("name"),
            "categories": FixedAsset.Category.choices,
            "already": exp.capitalized_asset,
        })

    def post(self, request, pk):
        from cashbook.models import Expense
        from .models import Acquisition
        exp = get_object_or_404(Expense, pk=pk)
        if exp.expenditure_type != Expense.ExpenditureType.CAPITAL:
            messages.error(request, "Only capital expenditure can be capitalised. "
                                    "Change the expense to capital first.")
            return redirect("expense_detail", pk=exp.pk)
        if exp.capitalized_asset_id:
            messages.info(request, f"That payment is already part of "
                                   f"{exp.capitalized_asset.name}.")
            return redirect("asset_detail", pk=exp.capitalized_asset_id)

        existing_id = request.POST.get("existing")
        if existing_id:
            asset = FixedAsset.objects.filter(pk=existing_id).first()
            if not asset:
                messages.error(request, "Choose an asset to add this payment to.")
                return redirect("expense_capitalise", pk=exp.pk)
            asset.cost = (asset.cost or Decimal(0)) + exp.amount
            asset.save(update_fields=["cost"])
            exp.capitalized_asset = asset
            exp.save(update_fields=["capitalized_asset"])
            messages.success(request, f"{exp.amount:,.2f} added to {asset.name}; "
                                      f"its cost is now {asset.cost:,.2f}.")
            return redirect("asset_detail", pk=asset.pk)

        name = (request.POST.get("name") or exp.description or "").strip()
        if not name:
            messages.error(request, "Give the asset a name.")
            return redirect("expense_capitalise", pk=exp.pk)
        from core.models import SiteConfig
        floor = SiteConfig.get().capitalisation_threshold or Decimal(0)
        if floor and exp.amount < floor:
            messages.error(request, f"{exp.amount:,.2f} is below the capitalisation "
                                    f"threshold of {floor:,.2f} — it belongs in running "
                                    f"costs, not on the register.")
            return redirect("expense_detail", pk=exp.pk)

        def _date(key, fallback):
            try:
                return dt.date.fromisoformat(request.POST.get(key) or "")
            except ValueError:
                return fallback

        asset = FixedAsset.objects.create(
            name=name[:120],
            category=request.POST.get("category") or FixedAsset.Category.EQUIPMENT,
            cost=exp.amount, salvage_value=Decimal(0),
            acquired_on=_date("acquired_on", exp.date),
            in_service_on=_date("in_service_on", exp.date),
            department=exp.department,
            reference=(exp.voucher_no or "")[:60],
            notes=f"Capitalised from payment {exp.pk}"[:250])
        exp.capitalized_asset = asset
        exp.save(update_fields=["capitalized_asset"])
        Acquisition.objects.create(
            asset=asset, source=Acquisition.Source.PURCHASE, date=asset.acquired_on,
            amount=exp.amount, expense=exp, fund=exp.department,
            reference=(exp.voucher_no or "")[:60],
            recorded_by=request.user if request.user.is_authenticated else None)
        messages.success(request, f"{name} is now on the asset register at "
                                  f"{exp.amount:,.2f}, and depreciates from "
                                  f"{asset.in_service_on:%d %b %Y}.")
        return redirect("asset_detail", pk=asset.pk)


class AssetTransitionView(TreasurerRequiredMixin, View):
    """Move an asset along its lifecycle. Every rule lives in the lifecycle
    service, so this view only reports what it decides."""
    def post(self, request, pk):
        from .services import lifecycle
        a = get_object_or_404(FixedAsset, pk=pk)
        target = request.POST.get("status") or ""
        try:
            lifecycle.transition(a, target, user=request.user,
                                 note=request.POST.get("note", "")[:120])
            messages.success(request, f"{a.name} is now {a.get_status_display().lower()}.")
        except lifecycle.TransitionError as exc:
            messages.error(request, str(exc))
        if request.POST.get("next") == "board":
            return redirect("asset_board")
        return redirect("asset_detail", pk=a.pk)


class AssetAssignView(TreasurerRequiredMixin, View):
    """Issue an asset to a custodian. An asset can only be in one pair of hands
    at a time, so an open assignment must be closed before another opens."""
    def post(self, request, pk):
        from django.contrib.auth.models import User
        from .models import AssetAssignment, AssetEvent, Location
        from .services import lifecycle
        a = get_object_or_404(FixedAsset, pk=pk)
        if a.disposed:
            messages.error(request, f"{a.name} has been disposed of.")
            return redirect("asset_detail", pk=a.pk)
        if lifecycle.open_assignment(a):
            messages.error(request, f"{a.name} is already issued. Check it in first.")
            return redirect("asset_detail", pk=a.pk)
        custodian = User.objects.filter(pk=request.POST.get("custodian")).first()
        holder_name = (request.POST.get("holder_name") or "").strip()
        if not custodian and not holder_name:
            messages.error(request, "Say who is taking the asset.")
            return redirect("asset_detail", pk=a.pk)
        try:
            frm = dt.date.fromisoformat(request.POST.get("from_date") or "")
        except ValueError:
            frm = dt.date.today()
        loc = Location.objects.filter(pk=request.POST.get("location")).first()
        asn = AssetAssignment.objects.create(
            asset=a, custodian=custodian, holder_name=holder_name[:120], location=loc,
            from_date=frm, condition_out=request.POST.get("condition_out", "")[:120],
            note=request.POST.get("note", "")[:250],
            issued_by=request.user if request.user.is_authenticated else None)
        if loc:
            a.location_fk = loc
        a.custodian = custodian
        a.save(update_fields=["custodian", "location_fk"])
        lifecycle.log(a, AssetEvent.Kind.ASSIGNED, f"Issued to {asn.holder}", request.user)
        messages.success(request, f"{a.name} issued to {asn.holder}.")
        return redirect("asset_detail", pk=a.pk)


class AssetCheckInView(TreasurerRequiredMixin, View):
    """Take an asset back into the church's own hands."""
    def post(self, request, pk):
        from .models import AssetEvent
        from .services import lifecycle
        a = get_object_or_404(FixedAsset, pk=pk)
        asn = lifecycle.open_assignment(a)
        if not asn:
            messages.info(request, f"{a.name} is not currently issued to anyone.")
            return redirect("asset_detail", pk=a.pk)
        try:
            asn.to_date = dt.date.fromisoformat(request.POST.get("to_date") or "")
        except ValueError:
            asn.to_date = dt.date.today()
        asn.condition_in = request.POST.get("condition_in", "")[:120]
        asn.received_by = request.user if request.user.is_authenticated else None
        asn.save()
        holder = asn.holder
        a.custodian = None
        a.save(update_fields=["custodian"])
        lifecycle.log(a, AssetEvent.Kind.RETURNED, f"Returned by {holder}", request.user)
        messages.success(request, f"{a.name} checked back in from {holder}.")
        return redirect("asset_detail", pk=a.pk)


class AssetTransferCreateView(TreasurerRequiredMixin, View):
    """Request a move of location and/or owning fund. A fund change is an
    accounting event, so it waits for approval before it posts."""
    def post(self, request, pk):
        from departments.models import Department
        from .models import AssetTransfer, Location
        a = get_object_or_404(FixedAsset, pk=pk)
        if a.disposed:
            messages.error(request, f"{a.name} has been disposed of.")
            return redirect("asset_detail", pk=a.pk)
        to_loc = Location.objects.filter(pk=request.POST.get("to_location")).first()
        to_fund = Department.objects.filter(pk=request.POST.get("to_fund")).first()
        if not to_loc and not to_fund:
            messages.error(request, "Choose a new location or a new fund.")
            return redirect("asset_detail", pk=a.pk)
        try:
            on = dt.date.fromisoformat(request.POST.get("date") or "")
        except ValueError:
            on = dt.date.today()
        tr = AssetTransfer.objects.create(
            asset=a, date=on, from_location=a.location_fk, to_location=to_loc or a.location_fk,
            from_fund=a.department, to_fund=to_fund or a.department,
            reason=request.POST.get("reason", "")[:250],
            requested_by=request.user if request.user.is_authenticated else None)
        messages.success(request, "Transfer requested — it takes effect once approved."
                         if tr.changes_fund else "Transfer requested.")
        return redirect("asset_detail", pk=a.pk)


class AssetTransferDecideView(TreasurerRequiredMixin, View):
    """Approve or reject a transfer. Approving is what actually moves the asset,
    and posts the inter-fund equity move when the owning fund changes."""
    def post(self, request, pk):
        from django.db import transaction as db_tx
        from django.utils import timezone
        from core import roles
        from ledger.services import posting
        from .models import AssetTransfer, AssetEvent
        from .services import lifecycle
        from . import signals as asset_signals
        tr = get_object_or_404(AssetTransfer, pk=pk)
        if not roles.can_approve(request.user):
            messages.error(request, "Approving a transfer is restricted to Treasurers.")
            return redirect("asset_detail", pk=tr.asset_id)
        if tr.status != AssetTransfer.Status.PENDING:
            messages.info(request, "That transfer has already been decided.")
            return redirect("asset_detail", pk=tr.asset_id)
        if request.POST.get("decision") == "reject":
            tr.status = AssetTransfer.Status.REJECTED
            tr.approved_by = request.user
            tr.approved_at = timezone.now()
            tr.save()
            messages.info(request, "Transfer rejected.")
            return redirect("asset_detail", pk=tr.asset_id)
        if tr.requested_by_id and tr.requested_by_id == request.user.pk:
            messages.error(request, "A transfer must be approved by someone other "
                                    "than the person who requested it.")
            return redirect("asset_detail", pk=tr.asset_id)
        a = tr.asset
        where = tr.to_location.name if tr.to_location else ""
        fund = tr.to_fund.name if tr.to_fund else ""
        # Approving is the accounting event: it moves the asset on the register
        # AND, when the owning fund changes, moves its carrying value between the
        # two funds' equity. Those used to be separated by however long it took
        # somebody to rebuild the ledger — and nothing reported the gap, because
        # an equity-only entry never disturbs the control accounts the
        # register↔ledger reconciliation watches. They are now one act: if the
        # journal cannot be written, the approval does not happen either.
        try:
            with db_tx.atomic():
                tr.status = AssetTransfer.Status.APPROVED
                tr.approved_by = request.user
                tr.approved_at = timezone.now()
                tr.save()
                a.location_fk = tr.to_location
                a.department = tr.to_fund
                a.save(update_fields=["location_fk", "department"])
                lifecycle.log(a, AssetEvent.Kind.TRANSFERRED,
                              f"Transferred to {' / '.join(x for x in (where, fund) if x)}",
                              request.user)
                asset_signals.post_to_ledger(posting.post_asset_transfer, tr)
        except Exception:  # noqa: BLE001 — the detail is for the log, not the treasurer
            from core.utils import log_exception as _lx; _lx('assets/views.py')
            messages.error(request, f"{a.name} was NOT transferred — the move could not be "
                                    f"posted to the general ledger, so the transfer is "
                                    f"still awaiting approval. Please try again, or tell "
                                    f"whoever supports the system.")
            return redirect("asset_detail", pk=tr.asset_id)
        messages.success(request, f"{a.name} transferred.")
        return redirect("asset_detail", pk=tr.asset_id)


class AssetBoardView(ReadAccessMixin, View):
    """The register as a board — every asset in the column for the stage of life
    it is at, so gaps (nothing commissioned, things stuck in maintenance) are
    visible at a glance."""
    def get(self, request):
        from .services import lifecycle
        order = [FixedAsset.Status.PLANNED, FixedAsset.Status.ON_ORDER,
                 FixedAsset.Status.IN_CWIP, FixedAsset.Status.IN_SERVICE,
                 FixedAsset.Status.IDLE, FixedAsset.Status.MAINTENANCE,
                 FixedAsset.Status.IMPAIRED, FixedAsset.Status.HELD_SALE,
                 FixedAsset.Status.DISPOSED]
        as_of = _as_of(request)
        assets = (FixedAsset.objects.select_related("department", "location_fk")
                  .exclude(status=FixedAsset.Status.ARCHIVED).order_by("name"))
        buckets = {s: [] for s in order}
        for a in assets:
            st = a.status if a.status in buckets else FixedAsset.Status.IN_SERVICE
            buckets[st].append({
                "asset": a,
                "nbv": a.net_book_value(as_of),
                "targets": [{"value": t, "label": FixedAsset.Status(t).label}
                            for t in lifecycle.allowed_transitions(a)],
            })
        columns = [{"status": s, "label": FixedAsset.Status(s).label,
                    "cards": buckets[s], "count": len(buckets[s]),
                    "total": sum((c["nbv"] for c in buckets[s]), Decimal(0))}
                   for s in order]
        return render(request, "assets/board.html",
                      {"columns": columns, "as_of": as_of})


class AssetPreflightView(ReadAccessMixin, View):
    """Read-only: which assets would break the reconciliation if register cost
    were made temporal on the acquisition date, and by how much."""
    def get(self, request):
        from core.metrics import metrics
        report = metrics.acquisition_coverage(_as_of(request))
        return render(request, "assets/preflight.html", {"r": report,
                                                         "as_of": report["as_of"]})


class AssetImportView(TreasurerRequiredMixin, View):
    """Bring an existing asset register in from a spreadsheet.

    Deliberately forgiving about the sheet and strict about the accounting:
    column headings are matched by name so a treasurer's own spreadsheet works
    without being re-typed, but nothing is written until the whole file has been
    checked and shown back for confirmation. Every asset imported gets an
    acquisition record, so the register still knows where each one came from.
    """
    HEADERS = {
        "name": ["name", "asset", "description", "item", "particulars"],
        "category": ["category", "class", "type"],
        "acquired_on": ["acquired", "acquired on", "date acquired", "purchase date",
                        "date", "date purchased"],
        "cost": ["cost", "amount", "value", "purchase price", "price"],
        "salvage_value": ["salvage", "salvage value", "residual", "residual value"],
        "department": ["fund", "department", "owning fund"],
        "location": ["location", "where", "room", "building"],
        "serial_no": ["serial", "serial no", "serial number"],
        "tag": ["tag", "asset tag", "code", "asset no", "asset number"],
        "reference": ["reference", "ref", "voucher", "invoice"],
        "method": ["method", "depreciation method", "basis"],
        "rate": ["rate", "depreciation rate", "rate %", "%"],
        "in_service_on": ["in service", "in service on", "commissioned",
                          "date in service"],
        "notes": ["notes", "note", "remarks", "comment", "comments"],
    }

    SAMPLE_ROWS = [
        ["Church building", "BUILDING", "2015-06-01", "8000000", "Development",
         "Main compound", "", "STRAIGHT", "2", "Owned since the church was built"],
        ["Church van", "VEHICLE", "2020-01-05", "1200000", "Development",
         "Car park", "KBX 123A", "REDUCING", "25", ""],
        ["PA / sound system", "EQUIPMENT", "2023-06-01", "420000", "Church Budget",
         "Sanctuary", "", "STRAIGHT", "10", ""],
        ["Plastic chairs (200)", "FURNITURE", "2022-03-15", "300000", "Church Budget",
         "Hall", "", "STRAIGHT", "10", "Counted as one item"],
    ]
    SAMPLE_HEADERS = ["Name", "Category", "Date acquired", "Cost", "Fund",
                      "Location", "Serial no", "Method", "Rate", "Notes"]

    def get(self, request):
        from departments.models import Department
        if request.GET.get("sample"):
            return self._sample()
        return render(request, "assets/import.html", {
            "categories": FixedAsset.Category.choices,
            "funds": Department.objects.filter(active=True).order_by("name"),
            "headers": self.HEADERS,
        })

    def _sample(self):
        """A filled-in example to start from — real column names, real dates, and
        one row per kind of asset a church usually holds."""
        import csv
        import io
        from django.http import HttpResponse
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(self.SAMPLE_HEADERS)
        for row in self.SAMPLE_ROWS:
            writer.writerow(row)
        # a BOM so Excel opens it with the accents and columns intact
        response = HttpResponse("\ufeff" + buf.getvalue(),
                                content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="asset-register-sample.csv"'
        return response

    # -- reading -------------------------------------------------------------

    def _map_columns(self, header_row):
        """Match the sheet's headings to fields, by name, case/space-insensitive."""
        found = {}
        for idx, raw in enumerate(header_row):
            label = str(raw or "").strip().lower().replace("_", " ")
            if not label:
                continue
            for field, aliases in self.HEADERS.items():
                if field in found:
                    continue
                if label in aliases or label.rstrip("s") in aliases:
                    found[field] = idx
                    break
        return found

    @staticmethod
    def _as_date(value):
        if value in (None, ""):
            return None
        if isinstance(value, dt.datetime):
            return value.date()
        if isinstance(value, dt.date):
            return value
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y", "%d %B %Y",
                    "%m/%d/%Y", "%Y/%m/%d"):
            try:
                return dt.datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _as_amount(value):
        if value in (None, ""):
            return None
        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value))
        text = re.sub(r"[^0-9.\-]", "", str(value))
        if not text or text in ("-", "."):
            return None
        try:
            return Decimal(text)
        except Exception:  # noqa: BLE001
            return None

    def _read_csv(self, upload):
        """CSV, read tolerantly: any common delimiter, and a UTF-8 BOM (which
        Excel writes) stripped so the first heading is still recognised."""
        import csv
        import io
        raw = upload.read()
        if isinstance(raw, bytes):
            for encoding in ("utf-8-sig", "utf-8", "latin-1"):
                try:
                    text = raw.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise ValueError("the text in this file is in an encoding I cannot read")
        else:
            text = raw
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        return [tuple(r) for r in csv.reader(io.StringIO(text), dialect)]

    def _read_workbook(self, upload):
        import openpyxl
        name = (getattr(upload, "name", "") or "").lower()
        if name.endswith(".xls"):
            raise ValueError("this is the older .xls format — open it and choose "
                             "Save As .xlsx, or save it as CSV")
        try:
            wb = openpyxl.load_workbook(upload, data_only=True, read_only=True)
        except Exception:
            # read-only mode is fussy about files written by other programs;
            # the ordinary reader copes with more of them
            upload.seek(0)
            wb = openpyxl.load_workbook(upload, data_only=True)
        if not wb.sheetnames:
            raise ValueError("I could not find any sheets in this file")
        ws = wb[wb.sheetnames[0]]
        try:
            rows = list(ws.iter_rows(values_only=True))
        except Exception:
            # some files declare the wrong size; re-reading without read-only
            # mode recalculates it
            upload.seek(0)
            wb2 = openpyxl.load_workbook(upload, data_only=True)
            rows = list(wb2[wb2.sheetnames[0]].iter_rows(values_only=True))
            wb2.close()
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass
        return rows

    def _read(self, upload):
        name = (getattr(upload, "name", "") or "").lower()
        upload.seek(0)
        if name.endswith((".csv", ".txt")):
            rows = self._read_csv(upload)
        else:
            rows = self._read_workbook(upload)
        rows = [r for r in rows if r is not None]
        if not rows:
            return {}, []
        # the header is the first row with at least two recognisable headings
        mapping, start = {}, 0
        for i, row in enumerate(rows[:10]):
            candidate = self._map_columns(row)
            if len(candidate) >= 2:
                mapping, start = candidate, i + 1
                break
        return mapping, rows[start:]

    def _examine(self, mapping, rows, defaults):
        """Check every row and describe what would happen. Writes nothing."""
        from departments.models import Department
        funds = {d.name.strip().lower(): d for d in Department.objects.filter(active=True)}
        categories = {label.lower(): value for value, label in FixedAsset.Category.choices}
        categories.update({value.lower(): value for value, _ in FixedAsset.Category.choices})
        existing = {n.strip().lower() for n in
                    FixedAsset.objects.values_list("name", flat=True)}
        from core.models import SiteConfig
        floor = SiteConfig.get().capitalisation_threshold or Decimal(0)

        ready, problems, seen = [], [], set()
        for n, row in enumerate(rows, start=1):
            def cell(field):
                idx = mapping.get(field)
                return row[idx] if idx is not None and idx < len(row) else None

            name = str(cell("name") or "").strip()
            if not name and not any(str(c or "").strip() for c in row):
                continue                      # a blank spacer row
            cost = self._as_amount(cell("cost"))
            acquired = self._as_date(cell("acquired_on"))
            issues = []
            if not name:
                issues.append("no name")
            if cost is None:
                issues.append("no cost")
            elif cost < 0:
                issues.append("cost is negative")
            elif floor and cost < floor:
                issues.append(f"below the {floor:,.0f} capitalisation threshold")
            if acquired is None:
                issues.append("no acquisition date I could read")
            elif acquired > dt.date.today():
                issues.append("acquired in the future")
            key = name.strip().lower()
            if key and key in existing:
                issues.append("already on the register")
            if key and key in seen:
                issues.append("appears twice in this file")
            seen.add(key)

            cat_raw = str(cell("category") or "").strip().lower()
            category = categories.get(cat_raw) or defaults["category"]
            fund_raw = str(cell("department") or "").strip().lower()
            fund = funds.get(fund_raw) or defaults["fund"]
            in_service = self._as_date(cell("in_service_on")) or acquired
            record = {
                "row": n, "name": name[:120], "category": category,
                "cost": cost, "acquired_on": acquired, "in_service_on": in_service,
                "salvage_value": self._as_amount(cell("salvage_value")) or Decimal(0),
                "fund": fund,
                "location": str(cell("location") or "").strip()[:120],
                "serial_no": str(cell("serial_no") or "").strip()[:80],
                "tag": str(cell("tag") or "").strip()[:40],
                "reference": str(cell("reference") or "").strip()[:60],
                "notes": str(cell("notes") or "").strip()[:250],
                "method": str(cell("method") or "").strip().upper()[:10],
                "rate": self._as_amount(cell("rate")),
                "issues": issues,
            }
            (problems if issues else ready).append(record)
        return ready, problems

    def post(self, request):
        from departments.models import Department
        from .models import Acquisition, AssetEvent, Location
        from .services import lifecycle

        upload = request.FILES.get("file")
        if not upload:
            messages.error(request, "Choose a spreadsheet to bring in.")
            return redirect("asset_import")
        defaults = {
            "category": request.POST.get("default_category") or FixedAsset.Category.EQUIPMENT,
            "fund": Department.objects.filter(pk=request.POST.get("default_fund")).first(),
        }
        try:
            mapping, rows = self._read(upload)
        except ValueError as exc:
            # a cause we recognised and can explain
            messages.error(request, f"I could not read that file — {exc}.")
            return redirect("asset_import")
        except Exception as exc:  # noqa: BLE001
            # anything else: the technical detail means nothing to the reader, so
            # say what to try instead, and log the detail for whoever is asked.
            import logging
            logging.getLogger("treasury").exception("Asset import could not read %s",
                                                    getattr(upload, "name", "?"))
            messages.error(
                request,
                "I could not read that file. If it came from another program, open "
                "it and save it again as .xlsx, or save it as CSV — the sample on "
                "this page shows the shape I expect. "
                f"(Technical detail, for support: {type(exc).__name__}: {exc})")
            return redirect("asset_import")
        if "name" not in mapping or "cost" not in mapping:
            messages.error(request, "I could not find a name column and a cost column. "
                                    "Name the headings so they can be recognised — "
                                    "'Name' and 'Cost' work.")
            return redirect("asset_import")

        ready, problems = self._examine(mapping, rows, defaults)

        # First pass: show what would happen. Nothing is written.
        if request.POST.get("confirm") != "yes":
            return render(request, "assets/import.html", {
                "categories": FixedAsset.Category.choices,
                "funds": Department.objects.filter(active=True).order_by("name"),
                "headers": self.HEADERS,
                "checked": True, "ready": ready, "problems": problems,
                "matched": sorted(mapping.keys()),
                "defaults": defaults,
                "filename": upload.name,
            })

        # Second pass: the treasurer has seen it and confirmed.
        created = 0
        for r in ready:
            loc = None
            if r["location"]:
                loc, _ = Location.objects.get_or_create(name=r["location"])
            asset = FixedAsset.objects.create(
                name=r["name"], category=r["category"], cost=r["cost"],
                salvage_value=r["salvage_value"], acquired_on=r["acquired_on"],
                in_service_on=r["in_service_on"], department=r["fund"],
                location_fk=loc, serial_no=r["serial_no"], tag=r["tag"] or None,
                reference=r["reference"], notes=r["notes"],
                method=r["method"] if r["method"] in dict(DepreciationRule.Method.choices) else "",
                rate=r["rate"])
            Acquisition.objects.create(
                asset=asset, source=Acquisition.Source.OPENING,
                date=r["acquired_on"], amount=r["cost"], fund=r["fund"],
                reference=r["reference"],
                notes="Imported from a spreadsheet"[:250],
                recorded_by=request.user if request.user.is_authenticated else None)
            lifecycle.log(asset, AssetEvent.Kind.CREATED,
                          "Imported from a spreadsheet", request.user)
            created += 1
        if created:
            messages.success(request, f"{created} asset{'s' if created != 1 else ''} "
                                      f"added to the register.")
        if problems:
            messages.info(request, f"{len(problems)} row(s) were left out — see the "
                                   f"reasons listed before you re-import them.")
        return redirect("asset_list")
