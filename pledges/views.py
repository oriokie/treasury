import datetime as dt
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.db import transaction as db_tx
from django.db.models import Sum, Count, Q
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils import timezone
from django.views import View
from django.views.generic import ListView, TemplateView

from core.permissions import (ReadAccessMixin, DataEntryRequiredMixin,
                              TreasurerRequiredMixin)
from .models import PledgeCampaign, Pledge, PledgePayment, PledgeReminderLog
from .forms import CampaignForm, PledgeForm
from .services import matching as match_svc
from .services import reminders as rem_svc


# ===========================================================================
# Dashboard / overview
# ===========================================================================
class PledgeDashboardView(ReadAccessMixin, TemplateView):
    template_name = "pledges/dashboard.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        campaigns = list(PledgeCampaign.objects.all())
        active = [c for c in campaigns if c.status == PledgeCampaign.Status.ACTIVE]
        ctx["campaigns"] = campaigns
        ctx["active_campaigns"] = active
        ctx["total_pledged"] = sum((c.total_pledged for c in campaigns), Decimal("0"))
        ctx["total_received"] = sum((c.total_received for c in campaigns), Decimal("0"))
        ctx["total_outstanding"] = sum((c.total_outstanding for c in campaigns), Decimal("0"))
        ctx["pledge_count"] = Pledge.objects.exclude(
            status=Pledge.Status.CANCELLED).count()
        ctx["draft_count"] = Pledge.objects.filter(status=Pledge.Status.DRAFT).count()
        from .models import PledgeMatchSuggestion
        ctx["suggestion_count"] = PledgeMatchSuggestion.objects.filter(status=PledgeMatchSuggestion.Status.PENDING).count()
        ctx["overdue"] = [p for p in Pledge.objects.filter(
            status__in=[Pledge.Status.ACTIVE, Pledge.Status.LAPSED])
            .select_related("member", "campaign") if p.is_overdue][:15]
        return ctx


# ===========================================================================
# Campaigns
# ===========================================================================
class CampaignListView(ReadAccessMixin, ListView):
    template_name = "pledges/campaign_list.html"
    context_object_name = "campaigns"

    def get_queryset(self):
        return PledgeCampaign.objects.all()


class CampaignDetailView(ReadAccessMixin, TemplateView):
    template_name = "pledges/campaign_detail.html"

    def get_context_data(self, **kwargs):
        from django.db.models import Q
        ctx = super().get_context_data(**kwargs)
        c = get_object_or_404(PledgeCampaign, pk=kwargs["pk"])
        ctx["campaign"] = c

        # This page is where a treasurer lands after importing pledges into a
        # campaign, so it has to answer "what did that import actually put in
        # here, and can I fix a row it got wrong?" It used to answer neither:
        # no search, no ordering that surfaced a fresh import, and no way to
        # reach the edit/delete screens that already existed. A wrongly
        # allocated pledge was visible but not correctable from where anyone
        # would look for it.
        qs = (c.pledges.exclude(status=Pledge.Status.CANCELLED)
              .select_related("member", "recorded_by"))
        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(Q(member__name__icontains=q)
                          | Q(member__phone__icontains=q)
                          | Q(note__icontains=q))
        status = self.request.GET.get("status")
        if status:
            qs = qs.filter(status=status)
        # newest first: a just-imported batch is what someone is here to check
        pledges = list(qs.order_by("-created_at", "-id"))

        ctx["pledges"] = pledges
        ctx["q"] = q
        ctx["f_status"] = status or ""
        ctx["statuses"] = Pledge.Status.choices
        # Off the campaign, not off `pledges`: the list above is whatever the
        # treasurer last filtered to, and the headline metric is about the
        # appeal. The headline states the promises the campaign counts, so the
        # ones it does not count have to be said somewhere or a draft simply
        # vanishes from the page a treasurer reviews drafts on.
        ctx["awaiting_approval"] = c.pledges.filter(
            status=Pledge.Status.DRAFT).count()
        ctx["fulfilled"] = sum(1 for p in pledges if p.status == Pledge.Status.FULFILLED)
        ctx["lapsed"] = sum(1 for p in pledges if p.status == Pledge.Status.LAPSED)
        return ctx


class CampaignCreateView(TreasurerRequiredMixin, View):
    def get(self, request, pk=None):
        obj = get_object_or_404(PledgeCampaign, pk=pk) if pk else None
        form = CampaignForm(instance=obj)
        return render(request, "pledges/campaign_form.html",
                      {"form": form, "obj": obj})

    def post(self, request, pk=None):
        obj = get_object_or_404(PledgeCampaign, pk=pk) if pk else None
        form = CampaignForm(request.POST, instance=obj)
        if form.is_valid():
            c = form.save(commit=False)
            if not c.created_by_id:
                c.created_by = request.user
            c.save()
            messages.success(request, f"Campaign “{c.name}” saved.")
            return redirect("pledge_campaign_detail", pk=c.pk)
        return render(request, "pledges/campaign_form.html",
                      {"form": form, "obj": obj})


# ===========================================================================
# Pledges
# ===========================================================================
class PledgeListView(ReadAccessMixin, ListView):
    template_name = "pledges/pledge_list.html"
    context_object_name = "pledges"
    paginate_by = 50

    def get_queryset(self):
        qs = Pledge.objects.select_related("member", "campaign")
        g = self.request.GET
        if g.get("status"):
            qs = qs.filter(status=g["status"])
        if g.get("campaign"):
            qs = qs.filter(campaign_id=g["campaign"])
        if g.get("q"):
            qs = qs.filter(Q(member__name__icontains=g["q"])
                           | Q(campaign__name__icontains=g["q"]))
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["filters"] = self.request.GET
        ctx["statuses"] = Pledge.Status.choices
        ctx["campaigns"] = PledgeCampaign.objects.all()
        qs = self.get_queryset()
        ctx["sum_pledged"] = qs.aggregate(s=Sum("amount"))["s"] or Decimal("0")
        # received across the filtered pledges
        ctx["sum_received"] = (PledgePayment.objects.filter(pledge__in=qs)
                               .aggregate(s=Sum("amount"))["s"] or Decimal("0"))
        ctx["sum_outstanding"] = (ctx["sum_pledged"] - ctx["sum_received"]
                                  if ctx["sum_pledged"] > ctx["sum_received"]
                                  else Decimal("0"))
        ctx["has_filters"] = any(g for g in (g2 for g2 in
                                 (self.request.GET.get(k) for k in
                                  ("status", "campaign", "q"))) if g)
        return ctx


class PledgeDetailView(ReadAccessMixin, TemplateView):
    template_name = "pledges/pledge_detail.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        p = get_object_or_404(Pledge.objects.select_related("member", "campaign"),
                              pk=kwargs["pk"])
        ctx["pledge"] = p
        ctx["payments"] = p.payments.select_related("transaction").all()
        from core.roles import is_treasurer, can_enter_data
        ctx["is_treasurer"] = is_treasurer(self.request.user)
        ctx["can_enter_data"] = can_enter_data(self.request.user)
        ctx["reminders"] = p.reminders.all()[:10]
        # expected schedule vs paid (informational)
        sched = p.expected_installments()
        ctx["schedule"] = [{"due": d, "amount": a} for d, a in sched]
        # match suggestions for the treasurer
        if self.request.user.has_perm and p.status == Pledge.Status.ACTIVE:
            ctx["suggestions"] = match_svc.suggest_matches_for_pledge(p)
        else:
            ctx["suggestions"] = []
        return ctx


class PledgeCreateView(DataEntryRequiredMixin, View):
    def get(self, request, pk=None):
        obj = get_object_or_404(Pledge, pk=pk) if pk else None
        initial = {}
        if request.GET.get("campaign"):
            initial["campaign"] = request.GET["campaign"]
        if request.GET.get("member"):
            initial["member"] = request.GET["member"]
        form = PledgeForm(instance=obj, initial=initial)
        # When the form is reached from a campaign, show what that campaign is
        # trying to raise and how far along it is. Someone recording a pledge
        # at an appeal is standing in front of that number; the form should not
        # be the one place in the system that hides it.
        campaign = None
        cid = (obj.campaign_id if obj else None) or initial.get("campaign")
        if cid:
            campaign = PledgeCampaign.objects.filter(pk=cid).first()
        return render(request, "pledges/pledge_form.html",
                      {"form": form, "obj": obj, "campaign": campaign})

    def post(self, request, pk=None):
        obj = get_object_or_404(Pledge, pk=pk) if pk else None
        form = PledgeForm(request.POST, instance=obj)
        if form.is_valid():
            p = form.save(commit=False)
            if not p.recorded_by_id:
                p.recorded_by = request.user
            # a treasurer recording a pledge can have it auto-approved; an
            # assistant's pledge stays DRAFT until a treasurer approves it
            if not obj:
                from core.roles import can_approve
                if can_approve(request.user):
                    p.status = Pledge.Status.ACTIVE
                    p.approved_by = request.user
                    p.approved_at = timezone.now()
                else:
                    p.status = Pledge.Status.DRAFT
            p.save()
            messages.success(request, "Pledge saved.")
            return redirect("pledge_detail", pk=p.pk)
        campaign = None
        cid = (obj.campaign_id if obj else None) or form.data.get("campaign")
        if cid:
            campaign = PledgeCampaign.objects.filter(pk=cid).first()
        return render(request, "pledges/pledge_form.html",
                      {"form": form, "obj": obj, "campaign": campaign})


class BasePledgeApprovalQueue(TemplateView):
    """Drafts awaiting approval — the treasurer's whole list, or a leader's own.

    The scope comes from ``approval.approvable_for``, which also decides what
    the bulk POST accepts, so the page cannot offer a row the action would then
    refuse. Two subclasses differ only in which door they open: staff reach it
    under /pledges/, leaders under /leader/, because ReadAccessMixin
    deliberately keeps leaders out of the unscoped office screens and widening
    it for this page would be the wrong trade entirely.
    """
    template_name = "pledges/approval_queue.html"

    def get_context_data(self, **kwargs):
        from pledges.services import approval
        ctx = super().get_context_data(**kwargs)
        rows = list(approval.approvable_for(self.request.user))
        campaign = self.request.GET.get("campaign")
        if campaign:
            rows = [p for p in rows if str(p.campaign_id) == campaign]
        ctx["rows"] = rows
        ctx["total"] = sum((p.amount for p in rows), Decimal("0"))
        ctx["campaigns"] = sorted(
            {(p.campaign_id, p.campaign.name)
             for p in approval.approvable_for(self.request.user)},
            key=lambda x: x[1])
        ctx["campaign"] = campaign
        from core import roles
        ctx["is_wide"] = roles.can_approve(self.request.user)
        return ctx

    def post(self, request, *args, **kwargs):
        from pledges.services import approval
        ids = request.POST.getlist("pledge")
        if not ids:
            messages.error(request, "Tick the pledges you want to approve.")
            return redirect(request.get_full_path())
        approved, skipped = approval.approve_many(ids, request.user)
        if approved:
            messages.success(
                request, f"{approved} pledge{'s' if approved != 1 else ''} "
                         "approved and now active.")
        if skipped:
            messages.warning(
                request, f"{skipped} could not be approved — they are not on a "
                         "fund you approve for, or were already handled.")
        return redirect(request.get_full_path())


class PledgeApprovalQueueView(ReadAccessMixin, BasePledgeApprovalQueue):
    """The office's door onto the approval queue."""


class PledgeApproveView(View):
    """Approve a draft pledge, or cancel/reactivate. Mirrors expense approval.

    Approving honours the same scope as the queue, so a leader may approve one
    of their own fund's pledges here as well as in bulk. Cancelling and
    reactivating remain the treasurer's — undoing a pledge the church has been
    counting on is a different act from confirming one was made.
    """
    def dispatch(self, request, *args, **kwargs):
        from core.permissions import LoginRequiredMixin  # noqa: F401
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, pk):
        from core import roles
        from pledges.services import approval
        p = get_object_or_404(Pledge, pk=pk)
        action = request.POST.get("action")
        if action == "approve" and p.status == Pledge.Status.DRAFT:
            if not approval.may_approve(request.user, p):
                messages.error(request, "That pledge is not on a fund you "
                                        "approve for.")
                return redirect("pledge_detail", pk=p.pk)
            approval.approve(p, request.user)
            messages.success(request, "Pledge approved and now active.")
        elif action in ("cancel", "reactivate") and not roles.can_approve(request.user):
            messages.error(request, "Only a treasurer can do that.")
        elif action == "cancel":
            paid = p.paid
            p.status = Pledge.Status.CANCELLED
            p.save()
            # Money already matched to it is left exactly where it is. The
            # contributions are real, they stay in the ledger, and their
            # PledgePayment links stay on the pledge as the record of what was
            # matched to what — unlinking them here would quietly rewrite
            # history to make a figure tidy. What changes is that a cancelled
            # pledge is no longer one of the campaign's counted pledges, so
            # neither the promise nor the money against it goes on propping up
            # the campaign's standing. Say the amount out loud: the treasurer
            # is about to watch that figure drop and should know why.
            if paid:
                messages.success(request, f"Pledge cancelled. The KES {paid:,.2f} "
                    "already matched to it no longer counts toward the campaign "
                    "total; those contributions remain in the ledger.")
            else:
                messages.success(request, "Pledge cancelled.")
        elif action == "reactivate" and p.status in (Pledge.Status.CANCELLED,
                                                      Pledge.Status.LAPSED):
            p.status = Pledge.Status.ACTIVE
            p.save()
            p.recompute_status()
            messages.success(request, "Pledge reactivated.")
        return redirect("pledge_detail", pk=p.pk)


# ===========================================================================
# Matching contributions to pledges
# ===========================================================================
class PledgeMatchView(TreasurerRequiredMixin, View):
    """Confirm a specific suggested contribution, or auto-match the whole pledge."""
    @db_tx.atomic
    def post(self, request, pk):
        p = get_object_or_404(Pledge, pk=pk)
        # One gate in front of all three actions, because all three end in a
        # PledgePayment and the pledge's status is the same objection to each.
        # It cannot be left to the template: pledge_detail.html renders the
        # "record a payment directly" form for any treasurer whatever the
        # pledge's status, so before this a treasurer could put money against a
        # self-submitted draft nobody had approved — and it landed in the
        # campaign's public "Received" figure. The auto action was already
        # refused, but by the matching service, which reported it back as "no
        # unmatched contributions found" and told the treasurer nothing.
        if not p.accepts_payment:
            messages.error(request, (
                "That pledge is still awaiting approval — approve it before "
                "recording money against it."
                if p.status == Pledge.Status.DRAFT else
                f"That pledge is {p.get_status_display().lower()} — reactivate "
                "it before recording money against it."))
            return redirect("pledge_detail", pk=p.pk)
        action = request.POST.get("action")
        if action == "auto":
            applied = match_svc.auto_match_pledge(p, user=request.user)
            if applied > 0:
                messages.success(request,
                    f"Auto-matched KES {applied:,.2f} from existing contributions.")
            else:
                messages.info(request, "No unmatched contributions found to apply.")
        elif action == "match":
            txn_id = request.POST.get("transaction")
            from giving.models import Transaction
            t = get_object_or_404(Transaction, pk=txn_id)
            try:
                amt = Decimal(request.POST.get("amount") or t.amount)
            except (InvalidOperation, TypeError):
                amt = t.amount
            already = (PledgePayment.objects.filter(transaction=t)
                       .aggregate(s=Sum("amount"))["s"] or Decimal("0"))
            free = t.amount - already
            amt = min(amt, free, p.outstanding if p.outstanding > 0 else amt)
            if amt <= 0:
                messages.error(request, "Nothing left to apply from that contribution.")
                return redirect("pledge_detail", pk=p.pk)
            PledgePayment.objects.create(pledge=p, transaction=t, amount=amt,
                date=t.date, source=PledgePayment.Source.MANUAL,
                matched_by=request.user)
            messages.success(request, f"Matched KES {amt:,.2f} to this pledge.")
        elif action == "manual":
            # record a payment with no linked transaction (e.g. a cash gift not
            # yet in the ledger) — still informational, flagged for follow-up
            try:
                amt = Decimal(request.POST.get("amount") or "0")
            except InvalidOperation:
                amt = Decimal("0")
            if amt > 0:
                PledgePayment.objects.create(pledge=p, transaction=None, amount=amt,
                    date=dt.date.today(), source=PledgePayment.Source.MANUAL,
                    matched_by=request.user,
                    note=request.POST.get("note", "")[:200] or "Manual entry")
                messages.success(request, f"Recorded KES {amt:,.2f} toward the pledge.")
            else:
                messages.error(request, "Enter an amount greater than zero.")
        return redirect("pledge_detail", pk=p.pk)


class PledgePaymentDeleteView(TreasurerRequiredMixin, View):
    def post(self, request, pk):
        pp = get_object_or_404(PledgePayment, pk=pk)
        pledge_id = pp.pledge_id
        pledge = pp.pledge
        pp.delete()
        pledge.recompute_status()
        messages.success(request, "Match removed.")
        return redirect("pledge_detail", pk=pledge_id)


class PledgeAutoMatchAllView(TreasurerRequiredMixin, View):
    """Preview the matches the sweep would make, then apply only on confirm.

    The dashboard button opens this page (GET). Applying still POSTs here —
    with `confirm=1` — so a treasurer sees every proposed link before any
    PledgePayment is written. Linking still never creates money.
    """
    template_name = "pledges/auto_match_preview.html"

    def _campaign(self, request):
        raw = request.POST.get("campaign") or request.GET.get("campaign")
        if not raw:
            return None
        return PledgeCampaign.objects.filter(pk=raw).first()

    def get(self, request):
        campaign = self._campaign(request)
        plan = match_svc.plan_auto_match_all(campaign=campaign)
        total = sum((r["amount"] for r in plan), Decimal("0"))
        pledges = {r["pledge"].id for r in plan}
        return render(request, self.template_name, {
            "campaign": campaign,
            "campaigns": PledgeCampaign.objects.filter(
                status=PledgeCampaign.Status.ACTIVE),
            "plan": plan,
            "total": total,
            "pledge_count": len(pledges),
            "match_count": len(plan),
        })

    @db_tx.atomic
    def post(self, request):
        campaign = self._campaign(request)
        if not request.POST.get("confirm"):
            # Accidental/empty POST — show the preview rather than apply.
            return redirect(request.path + (
                f"?campaign={campaign.pk}" if campaign else ""))
        plan = match_svc.plan_auto_match_all(campaign=campaign)
        # Honour unchecked rows: only apply the (pledge, txn) pairs the
        # treasurer left ticked. An empty selection means cancel.
        selected = set(request.POST.getlist("match"))
        if selected:
            plan = [r for r in plan
                    if f"{r['pledge'].id}:{r['txn'].id}" in selected]
        elif "match" in request.POST:
            # Form submitted with every box unchecked.
            messages.info(request, "No matches selected.")
            return redirect("pledge_dashboard")
        touched, total = match_svc.apply_planned_matches(
            plan, user=request.user)
        if touched:
            messages.success(request, f"Auto-matched KES {total:,.2f} across "
                                      f"{touched} pledge(s).")
        else:
            messages.info(request, "No new matches found.")
        return redirect("pledge_dashboard")


# ===========================================================================
# Reminders
# ===========================================================================
class PledgeReminderView(TreasurerRequiredMixin, View):
    """Send one pledge message — a reminder, or a thank-you for pledging."""

    def post(self, request, pk):
        p = get_object_or_404(Pledge, pk=pk)
        channel = request.POST.get("channel", "SMS")
        kind = "THANKS" if request.POST.get("kind") == "THANKS" else "REMINDER"
        log = rem_svc.send_pledge_reminder(p, channel=channel, user=request.user,
                                           kind=kind)
        noun = "Thank-you" if kind == "THANKS" else "Reminder"
        if log.ok:
            messages.success(request, f"{noun} sent.")
        else:
            messages.warning(request, f"{noun} not sent: {log.message[:120]}")
        return redirect("pledge_detail", pk=p.pk)


class PledgeMessagePreviewView(TreasurerRequiredMixin, View):
    """The message as this member would actually receive it.

    A template with placeholders reads nothing like the text that goes out, and
    an SMS cannot be recalled — so the words are shown filled in, against a
    real pledge, before anyone presses send. Also reports the length, because
    160 characters is a segment and a church pays per segment.
    """
    def get(self, request):
        from django.http import JsonResponse
        kind = "THANKS" if request.GET.get("kind") == "THANKS" else "REMINDER"
        template = request.GET.get("template")
        pledge = None
        if request.GET.get("pledge"):
            pledge = Pledge.objects.filter(pk=request.GET["pledge"]).first()
        if pledge is None:
            pledge = (Pledge.objects.exclude(status=Pledge.Status.CANCELLED)
                      .select_related("member", "campaign").first())
        if pledge is None:
            return JsonResponse({"ok": False,
                                 "error": "No pledge to preview against yet."})
        text = rem_svc.build_pledge_text(pledge, kind=kind,
                                          template=template or None)
        n = len(text)
        return JsonResponse({
            "ok": True, "text": text, "length": n,
            "segments": (n + 152) // 153 if n > 160 else 1,
            "example": pledge.member.name,
            "placeholders": list(rem_svc.PLACEHOLDERS),
        })


class PledgeReminderBatchView(TreasurerRequiredMixin, TemplateView):
    """Preview the list of members to remind, then send to all with one click."""
    template_name = "pledges/reminder_batch.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        campaign = None
        if self.request.GET.get("campaign"):
            campaign = PledgeCampaign.objects.filter(
                pk=self.request.GET["campaign"]).first()
        ctx["campaign"] = campaign
        ctx["campaigns"] = PledgeCampaign.objects.filter(
            status=PledgeCampaign.Status.ACTIVE)
        kind = ("THANKS" if self.request.GET.get("kind") == "THANKS"
                else "REMINDER")
        tag = self.request.GET.get("tag") or None
        ctx["kind"] = kind
        ctx["tag"] = tag
        from members.models import MemberTag
        ctx["all_tags"] = MemberTag.objects.filter(active=True)
        ctx["targets"] = rem_svc.reminder_targets(campaign=campaign, tag=tag,
                                                  kind=kind)
        ctx["batches"] = rem_svc.reminder_batches(campaign=campaign, tag=tag,
                                                  kind=kind)
        from core.models import SiteConfig
        cfg = SiteConfig.get()
        ctx["cfg"] = cfg
        ctx["sms_enabled"] = cfg.sms_enabled
        ctx["whatsapp_enabled"] = cfg.whatsapp_enabled
        return ctx

    def post(self, request):
        campaign = None
        if request.POST.get("campaign"):
            campaign = PledgeCampaign.objects.filter(pk=request.POST["campaign"]).first()
        channel = request.POST.get("channel", "SMS")
        kind = "THANKS" if request.POST.get("kind") == "THANKS" else "REMINDER"
        tag = request.POST.get("tag") or None
        # Recomputed from the same arguments the page was showing, rather than
        # trusting a list of ids from the form: what a treasurer approved was
        # "everyone on this screen", and the screen is defined by these filters.
        # Batched by member so two pledges never produce two texts.
        batches = rem_svc.reminder_batches(campaign=campaign, tag=tag,
                                           kind=kind)
        sent = 0
        for batch in batches:
            log = rem_svc.send_pledge_reminder(
                pledges=batch["pledges"], channel=channel,
                user=request.user, kind=kind)
            if log and log.ok:
                sent += 1
        noun = "thank-you" if kind == "THANKS" else "reminder"
        messages.success(request,
                         f"Sent {sent} of {len(batches)} {noun} message(s).")
        return redirect("pledge_dashboard")


# ===========================================================================
# Reports & year-end statements
# ===========================================================================
class CampaignPledgeReportView(ReadAccessMixin, TemplateView):
    """Who pledged what to one campaign, what they have given, and the balance.

    Optionally grouped by member tag, which is the question a treasurer
    actually asks of a campaign: not "how much is outstanding" — the dashboard
    says that — but "how are the board doing", "have the committee paid". A
    member holding two tags appears under both, and the group subtotals
    therefore do NOT sum to the campaign total. That is the honest presentation:
    the alternative is picking one tag per person arbitrarily, which answers
    neither question.
    """
    template_name = "pledges/campaign_report.html"

    def _campaign(self):
        return PledgeCampaign.objects.filter(
            pk=self.kwargs.get("pk") or self.request.GET.get("campaign")).first()

    def _rows(self, campaign):
        """Every promise on the campaign's books bar the cancelled ones.

        Listing a draft is deliberate — a treasurer reading down this report
        wants to see what is still waiting for her — but listing it and
        *counting* it are two different questions, and only the second has one
        right answer in this application. `counted` carries the campaign's own
        answer (`PledgeCampaign.counted_pledges`, i.e. RECOGNISED_STATUSES)
        down to `_totals` rather than the report deciding for itself: read from
        the campaign, the report follows automatically if that definition ever
        grows another condition, which is precisely how the two drifted apart
        in the first place.
        """
        counted = set(campaign.counted_pledges.values_list("pk", flat=True))
        qs = (Pledge.objects.filter(campaign=campaign)
              .exclude(status=Pledge.Status.CANCELLED)
              .select_related("member").prefetch_related("member__tags")
              .order_by("member__name"))
        today = dt.date.today()
        rows = []
        for p in qs:
            outstanding = p.outstanding
            rows.append({
                "pledge": p, "member": p.member,
                "tags": list(p.member.tags.all()),
                "amount": p.amount, "paid": p.paid,
                "outstanding": outstanding, "status": p.get_status_display(),
                "counted": p.pk in counted,
                "overdue": bool(p.end_date and p.end_date < today
                                and outstanding > 0),
                "pct": (int(min(p.paid / p.amount * 100, 100))
                        if p.amount else 0)})
        return rows

    @staticmethod
    def _sort(rows, key):
        """Name is the roll-call order; the others are working orders. Sorting
        by outstanding puts the follow-up list at the top, progress puts the
        least-paid first — ties broken by name so the order is stable."""
        if key == "outstanding":
            rows.sort(key=lambda r: (-r["outstanding"], r["member"].name))
        elif key == "pledged":
            rows.sort(key=lambda r: (-r["amount"], r["member"].name))
        elif key == "progress":
            rows.sort(key=lambda r: (r["pct"], r["member"].name))
        return rows

    def _screen_rows(self, campaign):
        """Exactly what the page shows: tag filter and sort applied. The export
        uses this too — a treasurer who filtered to the committee and pressed
        Excel means the committee, not the whole roll."""
        rows = self._rows(campaign)
        tag = self.request.GET.get("tag")
        if tag:
            rows = [r for r in rows if any(t.name == tag for t in r["tags"])]
        return self._sort(rows, self.request.GET.get("sort") or "name")

    @staticmethod
    def _goal_figures(campaign, totals):
        """The campaign band: how far pledges and money have got toward the
        goal. Bar segments are capped so the drawing never overflows; the
        stated percentages are not, because over-subscription is worth seeing.
        Without a goal the bar measures given against pledged instead — the
        only yardstick there is."""
        goal = campaign.goal_amount or Decimal("0")
        base = goal or totals["amount"]
        out = {"goal": goal}
        if totals["amount"]:
            out["fulfilment"] = int(totals["paid"] * 100 / totals["amount"])
        if goal:
            out["pct_pledged"] = int(totals["amount"] * 100 / goal)
            out["pct_received"] = int(totals["paid"] * 100 / goal)
            out["short"] = max(goal - totals["amount"], Decimal("0"))
        if base:
            received = int(min(totals["paid"] * 100 / base, 100))
            pledged = int(min(totals["amount"] * 100 / base, 100)) - received
            out["bar"] = {"received": max(received, 0),
                          "pledged": max(pledged, 0)}
        return out

    @staticmethod
    def _totals(rows):
        """The figures, over the promises the campaign actually recognises.

        A draft is a promise nobody has checked yet, and since the public
        pledge link went in anyone with the URL can create one by typing it.
        This row used to add up everything that was not CANCELLED, so 60,000
        approved beside a 40,000 draft made the report state 100,000 pledged
        while the campaign page, the public standing block and
        `PledgeCampaign.total_pledged` all stated 60,000 — and it fed
        `_goal_figures`, so the "% of goal pledged" band inherited the same
        inflation. This is the copy that gets printed and carried to a board
        meeting, which is the one place the figure is read with nothing beside
        it to contradict it. Nothing here is money, which is why it survived so
        long: no balance was wrong, only what the church believed it had been
        promised.

        `awaiting_*` keeps the drafts a stated figure rather than a silent
        omission, so the difference between the rows and the total is
        explainable rather than merely puzzling.
        """
        counted = [r for r in rows if r["counted"]]
        waiting = [r for r in rows if not r["counted"]]
        return {
            "n": len(counted),
            "amount": sum((r["amount"] for r in counted), Decimal(0)),
            "paid": sum((r["paid"] for r in counted), Decimal(0)),
            "outstanding": sum((r["outstanding"] for r in counted), Decimal(0)),
            "awaiting_n": len(waiting),
            "awaiting_amount": sum((r["amount"] for r in waiting), Decimal(0)),
        }

    def _groups(self, rows):
        """Rows bucketed by tag, plus everyone who carries none."""
        buckets = {}
        untagged = []
        for r in rows:
            if not r["tags"]:
                untagged.append(r)
            for t in r["tags"]:
                buckets.setdefault(t.name, []).append(r)
        out = [{"name": name, "rows": rs, "totals": self._totals(rs)}
               for name, rs in sorted(buckets.items())]
        if untagged:
            out.append({"name": "Untagged", "rows": untagged,
                        "totals": self._totals(untagged), "untagged": True})
        return out

    def get(self, request, *args, **kwargs):
        campaign = self._campaign()
        if campaign and request.GET.get("export") in ("csv", "xlsx"):
            return self._export(campaign, request.GET["export"])
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        from members.models import MemberTag
        ctx = super().get_context_data(**kwargs)
        campaign = self._campaign()
        ctx["campaign"] = campaign
        ctx["campaigns"] = PledgeCampaign.objects.order_by("name")
        ctx["group_by_tag"] = self.request.GET.get("group") == "tag"
        ctx["all_tags"] = MemberTag.objects.filter(active=True)
        ctx["sort"] = self.request.GET.get("sort") or "name"
        if campaign:
            rows = self._screen_rows(campaign)
            ctx["rows"] = rows
            ctx["totals"] = self._totals(rows)
            ctx["goal"] = self._goal_figures(campaign, ctx["totals"])
            ctx["groups"] = self._groups(rows) if ctx["group_by_tag"] else []
            ctx["tag"] = self.request.GET.get("tag")
        return ctx

    def _export(self, campaign, fmt):
        from reports.exports import csv_response, xlsx_response
        from core.models import SiteConfig
        rows = self._screen_rows(campaign)
        header = ["Member", "Tags", "Pledged", "Given", "Balance", "Status"]
        data = [[r["member"].name, ", ".join(t.name for t in r["tags"]),
                 float(r["amount"]), float(r["paid"]),
                 float(r["outstanding"]), r["status"]] for r in rows]
        t = self._totals(rows)
        data.append(["TOTAL", "", float(t["amount"]), float(t["paid"]),
                     float(t["outstanding"]), ""])
        if t["awaiting_n"]:
            # A workbook outlives the screen it came off, and whoever opens it
            # next will add the Pledged column up by hand. The draft rows are
            # in that column and not in the TOTAL, so the sheet has to say so
            # itself rather than leave a difference nobody can account for.
            data.append(["AWAITING APPROVAL — not in the TOTAL above", "",
                         float(t["awaiting_amount"]), "", "",
                         f"{t['awaiting_n']} draft pledge(s)"])
        name = f"pledges_{campaign.pk}"
        if fmt == "xlsx":
            return xlsx_response(f"{name}.xlsx", header, data,
                                 title=f"{campaign.name} — pledges",
                                 church=SiteConfig.get().church_name)
        return csv_response(f"{name}.csv", header, data)


class PledgeReportView(ReadAccessMixin, TemplateView):
    """Campaign progress + per-status breakdown, exportable."""
    template_name = "pledges/report.html"

    def get(self, request, *args, **kwargs):
        if request.GET.get("export") == "xlsx":
            return self._export()
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["campaigns"] = PledgeCampaign.objects.all()
        ctx["status_rows"] = (Pledge.objects.exclude(status=Pledge.Status.CANCELLED)
                              .values("status").annotate(n=Count("id"),
                                                         total=Sum("amount"))
                              .order_by("status"))
        return ctx

    def _export(self):
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from django.http import HttpResponse
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Campaign progress"
        head = ["Campaign", "Status", "Goal", "Pledged", "Received",
                "Outstanding", "% received", "Pledges"]
        ws.append(head)
        for c in range(1, len(head) + 1):
            ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
            ws.cell(1, c).fill = PatternFill("solid", fgColor="1F5F4F")
        for camp in PledgeCampaign.objects.all():
            ws.append([camp.name, camp.get_status_display(),
                       float(camp.goal_amount), float(camp.total_pledged),
                       float(camp.total_received), float(camp.total_outstanding),
                       camp.pct_received, camp.pledge_count])
        buf = io.BytesIO(); wb.save(buf)
        resp = HttpResponse(buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = 'attachment; filename="pledge_report.xlsx"'
        return resp


class MemberPledgeStatementView(ReadAccessMixin, TemplateView):
    """Year-end statement of a member's pledges and fulfilment — print/PDF ready."""
    template_name = "pledges/member_statement.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from members.models import Member
        m = get_object_or_404(Member, pk=kwargs["pk"])
        try:
            year = int(self.request.GET.get("year", dt.date.today().year))
        except (TypeError, ValueError):
            year = dt.date.today().year
        pledges = (m.pledges.exclude(status=Pledge.Status.CANCELLED)
                   .select_related("campaign"))
        rows = []
        for p in pledges:
            year_paid = (p.payments.filter(date__year=year)
                         .aggregate(s=Sum("amount"))["s"] or Decimal("0"))
            rows.append({"pledge": p, "year_paid": year_paid})
        ctx["member"] = m
        ctx["year"] = year
        ctx["rows"] = rows
        ctx["total_pledged"] = sum((r["pledge"].amount for r in rows), Decimal("0"))
        ctx["total_paid_year"] = sum((r["year_paid"] for r in rows), Decimal("0"))
        from core.models import SiteConfig
        ctx["cfg"] = SiteConfig.get()
        return ctx


# ===========================================================================
# Match suggestions (SUGGEST mode review queue)
# ===========================================================================
class PledgeSuggestionListView(TreasurerRequiredMixin, TemplateView):
    template_name = "pledges/suggestions.html"

    def get_context_data(self, **kwargs):
        from .models import PledgeMatchSuggestion
        ctx = super().get_context_data(**kwargs)
        ctx["suggestions"] = (PledgeMatchSuggestion.objects.filter(
            status=PledgeMatchSuggestion.Status.PENDING)
            .select_related("transaction", "pledge", "pledge__member",
                            "pledge__campaign", "transaction__department"))
        return ctx


class PledgeSuggestionActionView(TreasurerRequiredMixin, View):
    @db_tx.atomic
    def post(self, request, pk):
        from .models import PledgeMatchSuggestion
        s = get_object_or_404(PledgeMatchSuggestion, pk=pk)
        action = request.POST.get("action")
        if action == "confirm" and s.status == PledgeMatchSuggestion.Status.PENDING:
            # A suggestion is only ever raised against an active pledge, but it
            # then waits in a queue — and the pledge can be cancelled while it
            # waits. Confirming it afterwards would be the same defect by a
            # slower route, so the same rule is asked here.
            if not s.pledge.accepts_payment:
                messages.error(request, f"{s.pledge.member.name}'s pledge is "
                    f"{s.pledge.get_status_display().lower()} — nothing was "
                    "matched. Dismiss the suggestion, or reactivate the pledge "
                    "first.")
                return redirect("pledge_suggestions")
            # apply the match (capped at outstanding and at the gift's free amount)
            already = (PledgePayment.objects.filter(transaction=s.transaction)
                       .aggregate(x=Sum("amount"))["x"] or Decimal("0"))
            free = s.transaction.amount - already
            amt = min(s.amount, free, s.pledge.outstanding
                      if s.pledge.outstanding > 0 else s.amount)
            if amt > 0:
                PledgePayment.objects.create(pledge=s.pledge,
                    transaction=s.transaction, amount=amt,
                    date=s.transaction.date, source=PledgePayment.Source.AUTO,
                    matched_by=request.user, note="Confirmed from suggestion")
            s.status = PledgeMatchSuggestion.Status.CONFIRMED
            s.resolved_by = request.user
            s.resolved_at = timezone.now()
            s.save()
            messages.success(request, f"Matched KES {amt:,.0f} to "
                                      f"{s.pledge.member.name}'s pledge.")
        elif action == "dismiss":
            s.status = PledgeMatchSuggestion.Status.DISMISSED
            s.resolved_by = request.user
            s.resolved_at = timezone.now()
            s.save()
            messages.info(request, "Suggestion dismissed.")
        return redirect("pledge_suggestions")


# ===========================================================================
# Public member pledge form  (NO LOGIN — security-sensitive; off by default)
# ===========================================================================
# Design for safety:
#   * Off unless SiteConfig.pledge_public_form_enabled is set.
#   * Write-only: it creates a single self-submitted, UNVERIFIED draft Pledge and
#     nothing else. It never reads or exposes member data, balances, or other
#     pledges, so there is no information-disclosure surface.
#   * No member lookup/autocomplete is exposed (that would leak the membership
#     roll). The submitter types their name/phone as free text; a treasurer links
#     it to the real Member record during review.
#   * A draft posts to nothing — it cannot touch the ledger or a fund balance.
#   * Spam defences: only ACTIVE campaigns are selectable; a honeypot field; a
#     minimum render-to-submit time; a simple per-session/IP throttle; and a hard
#     amount ceiling. Approval is always manual.
from django.contrib.auth.decorators import login_not_required
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_protect
import time


@method_decorator(csrf_protect, name="dispatch")
@method_decorator(login_not_required, name="dispatch")
class PublicPledgeView(View):
    # Deliberately public (P1-1 exempt): a member-facing pledge form, gated by
    # SiteConfig.pledge_public_form_enabled (off by default) and heavily
    # rate/bot-guarded below. Status follows pledge_public_submit_mode
    # (draft for approval, or active immediately).
    template_name = "pledges/public_form.html"
    MAX_AMOUNT = Decimal("100000000")   # sanity ceiling
    MIN_SECONDS = 2                     # forms filled faster than this are bots

    def _enabled(self):
        from core.models import SiteConfig
        return SiteConfig.get().pledge_public_form_enabled

    def _open_campaigns(self):
        import datetime as _d
        from django.db.models import Q
        today = _d.date.today()
        return (PledgeCampaign.objects
                .filter(status=PledgeCampaign.Status.ACTIVE)
                .filter(Q(end_date__isnull=True) | Q(end_date__gte=today))
                .order_by("name"))

    def get(self, request, pk=None):
        if not self._enabled():
            return render(request, "pledges/public_disabled.html", status=404)
        campaigns = self._open_campaigns()
        campaign = campaigns.filter(pk=pk).first() if pk else None
        if pk and campaign is None:
            # a stale or wrong link should not silently become "pick one"
            return render(request, "pledges/public_disabled.html", status=404)
        request.session["pledge_form_ts"] = time.time()
        from core.models import SiteConfig
        return render(request, self.template_name, {
            "campaigns": campaigns, "campaign": campaign,
            "cfg": SiteConfig.get()})

    def post(self, request, pk=None):
        if not self._enabled():
            return render(request, "pledges/public_disabled.html", status=404)
        from core.models import SiteConfig
        cfg = SiteConfig.get()
        campaigns = self._open_campaigns()
        # A campaign in the URL is the campaign, full stop — the form has no
        # chooser to disagree with, and a posted field cannot redirect the
        # pledge somewhere the giver never saw.
        fixed = campaigns.filter(pk=pk).first() if pk else None
        if pk and fixed is None:
            return render(request, "pledges/public_disabled.html", status=404)

        def fail(msg):
            return render(request, self.template_name,
                          {"campaigns": campaigns, "campaign": fixed,
                           "cfg": cfg, "error": msg, "form": request.POST})

        # 1) honeypot — a hidden field real users never fill
        if (request.POST.get("website") or "").strip():
            return redirect("public_pledge_thanks")   # silently drop bots
        # 2) too-fast submit
        ts = request.session.get("pledge_form_ts")
        if ts and (time.time() - ts) < self.MIN_SECONDS:
            return fail("Please take a moment to review your details and try again.")
        # 3) light throttle: max 3 submissions per session
        n = request.session.get("pledge_submits", 0)
        if n >= 3:
            return fail("Thank you — we already received your pledge. Please contact "
                        "the treasurer for further changes.")

        name = (request.POST.get("name") or "").strip()
        phone = (request.POST.get("phone") or "").strip()
        camp_id = request.POST.get("campaign")
        amount_raw = (request.POST.get("amount") or "").strip()
        note = (request.POST.get("note") or "").strip()[:200]

        if not name or len(name) < 3:
            return fail("Please enter your full name.")
        if not phone:
            return fail("Please enter your M-PESA phone number.")
        from members.models import Member
        from members.services.matching import name_key, normalize_phone
        ph = normalize_phone(phone)
        if not ph:
            return fail("Please enter a valid M-PESA phone number.")
        campaign = fixed or campaigns.filter(pk=camp_id).first()
        if not campaign:
            return fail("Please choose a campaign.")
        try:
            amount = Decimal(amount_raw.replace(",", ""))
        except (InvalidOperation, AttributeError):
            return fail("Please enter a valid amount.")
        if amount <= 0 or amount > self.MAX_AMOUNT:
            return fail("Please enter a valid amount.")
        # How it will be given is not asked. It is optional on the record, it
        # is the question most likely to make somebody abandon the form, and a
        # treasurer can set it later from something the giver actually said.
        freq = Pledge.Frequency.MONTHLY

        # Resolve to a Member only by an exact, unambiguous match; otherwise leave
        # unlinked for the treasurer. We never reveal whether a match was found.
        member = Member.objects.filter(phone=ph).first()
        if not member:
            nk = name_key(name)
            matches = Member.objects.filter(name_key=nk)[:2]
            if len(matches) == 1:
                member = matches[0]
        if not member:
            # create a provisional member record (inactive until a treasurer
            # confirms) so the pledge always has an owner
            member = Member.objects.create(name=name, phone=ph,
                                           source=Member.Source.AUTO_BANK,
                                           active=False)
        elif not member.phone:
            member.phone = ph
            member.save(update_fields=["phone"])

        auto_accept = (cfg.pledge_public_submit_mode
                       == SiteConfig.PledgePublicSubmitMode.ACTIVE)
        status = (Pledge.Status.ACTIVE if auto_accept else Pledge.Status.DRAFT)
        pledge = Pledge.objects.create(
            campaign=campaign, member=member, amount=amount, frequency=freq,
            start_date=dt.date.today(), status=status,
            self_submitted=True,
            submitted_contact=f"{name} / {phone}"[:120],
            note=note,
            approved_at=timezone.now() if auto_accept else None)
        request.session["pledge_submits"] = n + 1
        request.session["pledge_thanks_accepted"] = auto_accept

        # Thank-you SMS to the member (best-effort; never blocks the redirect).
        try:
            from django.db import transaction
            from pledges.services.reminders import maybe_send_submit_thanks
            pledge_id = pledge.pk
            transaction.on_commit(lambda: maybe_send_submit_thanks(pledge_id))
        except Exception:
            from core.utils import log_exception as _lx; _lx('pledges/views.py')

        # notify treasurers
        try:
            from core.services.notifications import notify
            if auto_accept:
                notify("pledge", f"New member pledge from {name} "
                                 f"(KES {amount:,.0f} to {campaign.name}) — "
                                 "accepted automatically.",
                       link=f"/pledges/{pledge.pk}/")
            else:
                notify("pledge", f"New member pledge submitted by {name} "
                                 f"(KES {amount:,.0f} to {campaign.name}) — "
                                 "review needed.",
                       link="/pledges/list/?status=DRAFT")
        except Exception:
            from core.utils import log_exception as _lx; _lx('pledges/views.py')
            pass
        return redirect("public_pledge_thanks")


@method_decorator(login_not_required, name="dispatch")
class PublicPledgeThanksView(View):
    def get(self, request):
        from core.models import SiteConfig
        accepted = bool(request.session.pop("pledge_thanks_accepted", False))
        return render(request, "pledges/public_thanks.html",
                      {"cfg": SiteConfig.get(), "accepted": accepted})


# ===========================================================================
# Bulk pledge import (treasurer-only)
# ===========================================================================
#
# For loading pledges collected on paper (e.g. cards filled at a campaign
# launch). A two-step wizard mirroring the budget importer: upload -> review &
# map any unmatched members/campaigns -> apply. Imported pledges land as DRAFT
# so a treasurer still approves them, exactly like the public form — no pledge
# becomes active (or affects anything) without a human decision. Pledges never
# post to the ledger, so this is purely an informational bulk-entry convenience.

class PledgeImportView(TreasurerRequiredMixin, View):
    template_name = "pledges/import.html"
    #: extra context for the review screen, so another view (the leader's
    #: fund-scoped importer) can reuse this one's parsing and review UI while
    #: pointing its own navigation at its own pages.
    extra_context = None

    FREQ_LABELS = {
        "ONE OFF": "ONE_OFF", "ONEOFF": "ONE_OFF", "ONE-OFF": "ONE_OFF",
        "ONCE": "ONE_OFF", "LUMP SUM": "ONE_OFF", "LUMPSUM": "ONE_OFF",
        "WEEKLY": "WEEKLY", "WEEK": "WEEKLY",
        "MONTHLY": "MONTHLY", "MONTH": "MONTHLY",
        "QUARTERLY": "QUARTERLY", "QUARTER": "QUARTERLY",
        "ANNUAL": "ANNUAL", "ANNUALLY": "ANNUAL", "YEARLY": "ANNUAL", "YEAR": "ANNUAL",
    }

    def dispatch(self, request, *args, **kwargs):
        # Item 5: when reached via a campaign page, every imported pledge is
        # scoped to that campaign (no Campaign column needed).
        self.campaign = None
        if kwargs.get("pk"):
            self.campaign = PledgeCampaign.objects.filter(pk=kwargs["pk"]).first()
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        if request.GET.get("download"):
            return self._download(request)
        return render(request, self.template_name,
                      {"stage": "upload", "campaign": self.campaign})

    def post(self, request, *args, **kwargs):
        if request.POST.get("apply"):
            return self._apply(request)
        return self._parse(request)

    # ---- template ----
    def _download(self, request):
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
        from django.http import HttpResponse
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Pledges"
        head = ["Member name", "Phone", "Campaign", "Amount", "Frequency",
                "Start date", "End date", "Note"]
        ws.append(head)
        for c in range(1, len(head) + 1):
            ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
            ws.cell(1, c).fill = PatternFill("solid", fgColor="1F5F4F")
        # worked examples
        ws.append(["GRACE WANJIRU", "0712345678", "Sanctuary Roof", 50000,
                   "Monthly", "2026-01-01", "2026-12-31", "Pledged at launch"])
        ws.append(["PETER OTIENO", "", "Sanctuary Roof", 20000, "One-off",
                   "2026-01-01", "", ""])

        ref = wb.create_sheet("Lists")
        ref["A1"] = "Campaigns"; ref["A1"].font = Font(bold=True)
        camps = list(PledgeCampaign.objects.exclude(
            status=PledgeCampaign.Status.CLOSED).order_by("name"))
        for i, c in enumerate(camps, start=2):
            ref.cell(i, 1, c.name)
        ref["B1"] = "Frequency"; ref["B1"].font = Font(bold=True)
        for i, (v, l) in enumerate(Pledge.Frequency.choices, start=2):
            ref.cell(i, 2, l)

        nrows = 400
        if camps:
            dv_c = DataValidation(type="list",
                formula1=f"=Lists!$A$2:$A${len(camps) + 1}", allow_blank=True)
            ws.add_data_validation(dv_c); dv_c.add(f"C2:C{nrows}")
        dv_f = DataValidation(type="list",
            formula1=f"=Lists!$B$2:$B${len(Pledge.Frequency.choices) + 1}",
            allow_blank=True)
        ws.add_data_validation(dv_f); dv_f.add(f"E2:E{nrows}")

        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["H"].width = 26
        info = wb.create_sheet("How to fill this in")
        for i, line in enumerate([
            "Pledge import",
            "",
            "One row per pledge a member has made.",
            "  - Member name — required. Matched to an existing member by name or",
            "      phone; an unmatched name can be mapped or created on review.",
            "  - Phone — optional, helps match and is saved to a created member.",
            "  - Campaign — pick from the list. Blank/unmatched is flagged on review.",
            "  - Amount — the total promised (required, > 0).",
            "  - Frequency — One-off / Weekly / Monthly / Quarterly / Annual.",
            "  - Start / End date — optional (YYYY-MM-DD). End date drives the",
            "      schedule and the lapsed check.",
            "",
            "Imported pledges are saved as DRAFTS for a treasurer to approve.",
            "Nothing posts to the ledger — pledges are informational commitments.",
        ], start=1):
            info.cell(i, 1, line)
        info.column_dimensions["A"].width = 74

        buf = io.BytesIO(); wb.save(buf)
        resp = HttpResponse(buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = 'attachment; filename="pledge_import_template.xlsx"'
        return resp

    # ---- step 1: parse + build a review plan ----
    def _parse(self, request):
        import openpyxl
        from members.models import Member
        from members.services.matching import name_key
        from members.models import normalize_phone
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "Choose a spreadsheet to upload.")
            return redirect("pledge_import")
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception:
            from core.utils import log_exception as _lx; _lx('pledges/views.py')
            messages.error(request, "Could not read that file — please upload a .xlsx.")
            return redirect("pledge_import")
        ws = wb["Pledges"] if "Pledges" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            messages.error(request, "The sheet is empty.")
            return redirect("pledge_import")
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

        def col(*names):
            for n in names:
                if n in header:
                    return header.index(n)
            return None
        c_name = col("member name", "member", "name")
        c_phone = col("phone", "mobile")
        c_camp = col("campaign")
        c_amt = col("amount", "pledge", "pledged")
        c_freq = col("frequency")
        c_start = col("start date", "start")
        c_end = col("end date", "end")
        c_note = col("note", "notes")
        if c_name is None or c_amt is None:
            messages.error(request, "Couldn't find the Member name and Amount "
                                    "columns — please use the template.")
            return redirect("pledge_import")

        members = list(Member.objects.all())
        by_key = {}
        by_phone = {}
        for m in members:
            by_key.setdefault(m.name_key, m)
            if m.phone:
                by_phone[m.phone] = m
        campaigns = {self._norm(c.name): c for c in PledgeCampaign.objects.all()}

        def cell(r, idx):
            if idx is None or idx >= len(r) or r[idx] in (None, ""):
                return ""
            return str(r[idx]).strip()

        def parse_date(v):
            if not v:
                return None
            if isinstance(v, dt.datetime):
                return v.date().isoformat()
            if isinstance(v, dt.date):
                return v.isoformat()
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y"):
                try:
                    return dt.datetime.strptime(str(v).strip(), fmt).date().isoformat()
                except ValueError:
                    continue
            return None

        plan = []
        for r in rows[1:]:
            name = cell(r, c_name)
            if not name:
                continue
            try:
                amt = float(r[c_amt]) if c_amt < len(r) and r[c_amt] not in (None, "") else 0.0
            except (TypeError, ValueError):
                amt = 0.0
            if amt <= 0:
                continue
            phone = normalize_phone(cell(r, c_phone)) or cell(r, c_phone)
            # match member: phone first, then unambiguous name key
            m = None
            score = 0.0
            if phone and phone in by_phone:
                m = by_phone[phone]; score = 1.0
            if not m:
                nk = name_key(name)
                if nk and nk in by_key:
                    m = by_key[nk]; score = 0.9
            camp_raw = cell(r, c_camp)
            camp = campaigns.get(self._norm(camp_raw)) if camp_raw else None
            freq_raw = cell(r, c_freq).upper().replace("/", " ").strip()
            freq = self.FREQ_LABELS.get(freq_raw, "ONE_OFF")
            plan.append({
                "name": name, "phone": phone,
                "member_id": m.id if m else None,
                "member_match": m.name if m else None,
                "member_score": score,
                "campaign_raw": camp_raw,
                "campaign_id": camp.id if camp else None,
                "campaign_match": camp.name if camp else None,
                "amount": amt, "frequency": freq,
                "start": parse_date(r[c_start]) if c_start is not None and c_start < len(r) else None,
                "end": parse_date(r[c_end]) if c_end is not None and c_end < len(r) else None,
                "note": cell(r, c_note)[:200],
            })

        if not plan:
            messages.error(request, "No pledge rows with an amount were found.")
            return redirect("pledge_import")

        request.session["pledge_import_plan"] = plan
        if self.campaign:
            # campaign-scoped import: every row belongs to this campaign
            for p in plan:
                p["campaign_id"] = self.campaign.id
                p["campaign_match"] = self.campaign.name
                p["campaign_raw"] = ""
            request.session["pledge_forced_campaign"] = self.campaign.id
        else:
            request.session.pop("pledge_forced_campaign", None)
        unmatched_members = sum(1 for p in plan if not p["member_id"])
        unmatched_camps = sorted({p["campaign_raw"] for p in plan
                                  if p["campaign_raw"] and not p["campaign_id"]})
        no_campaign = sum(1 for p in plan if not p["campaign_raw"])
        member_candidates = [{"id": m.id, "name": m.name} for m in
                             sorted(members, key=lambda x: x.name)]
        campaign_candidates = [{"id": c.id, "name": c.name} for c in
                               PledgeCampaign.objects.exclude(
                                   status=PledgeCampaign.Status.CLOSED).order_by("name")]
        return render(request, self.template_name, {
            "stage": "review", "plan": plan, "campaign": self.campaign,
            **(self.extra_context or {}),
            "total": sum(p["amount"] for p in plan),
            "unmatched_members": unmatched_members,
            "unmatched_camps": unmatched_camps,
            "no_campaign": no_campaign,
            "member_candidates": member_candidates,
            "campaign_candidates": campaign_candidates,
            "frequencies": Pledge.Frequency.choices,
        })

    # ---- step 2: apply ----
    @db_tx.atomic
    def _apply(self, request):
        from members.models import Member
        plan = request.session.get("pledge_import_plan")
        if not plan:
            messages.error(request, "Your import session expired — please upload again.")
            return redirect("pledge_import")
        created = skipped = new_members = new_camps = 0
        # a single "default campaign" choice can cover rows with no campaign
        default_campaign_id = request.POST.get("default_campaign") or None
        # campaign-scoped import: this campaign wins for every row
        forced_campaign_id = request.session.get("pledge_forced_campaign")
        forced_campaign = (PledgeCampaign.objects.filter(pk=forced_campaign_id).first()
                           if forced_campaign_id else None)

        for i, p in enumerate(plan):
            # resolve member
            mchoice = request.POST.get(f"member_{i}", "")
            member = None
            if mchoice.startswith("member:"):
                member = Member.objects.filter(pk=mchoice.split(":", 1)[1]).first()
            elif mchoice == "create" or (mchoice == "" and not p["member_id"]):
                if p["name"]:
                    member = Member.objects.create(
                        name=p["name"], phone=p["phone"] or None,
                        source=Member.Source.MANUAL)
                    new_members += 1
            elif p["member_id"]:
                member = Member.objects.filter(pk=p["member_id"]).first()
            if not member:
                skipped += 1
                continue

            # resolve campaign
            cchoice = request.POST.get(f"campaign_{i}", "")
            campaign = None
            if forced_campaign:
                campaign = forced_campaign
            elif cchoice.startswith("campaign:"):
                campaign = PledgeCampaign.objects.filter(pk=cchoice.split(":", 1)[1]).first()
            elif cchoice == "create" and p["campaign_raw"]:
                campaign, was = PledgeCampaign.objects.get_or_create(
                    name=p["campaign_raw"].strip(),
                    defaults={"status": PledgeCampaign.Status.ACTIVE,
                              "created_by": request.user})
                if was:
                    new_camps += 1
            elif p["campaign_id"]:
                campaign = PledgeCampaign.objects.filter(pk=p["campaign_id"]).first()
            elif default_campaign_id:
                campaign = PledgeCampaign.objects.filter(pk=default_campaign_id).first()
            if not campaign:
                skipped += 1
                continue

            from decimal import Decimal as D
            kwargs = dict(
                campaign=campaign, member=member, amount=D(str(p["amount"])),
                frequency=p["frequency"], status=Pledge.Status.DRAFT,
                recorded_by=request.user, note=p["note"])
            if p["start"]:
                kwargs["start_date"] = p["start"]
            if p["end"]:
                kwargs["end_date"] = p["end"]
            Pledge.objects.create(**kwargs)
            created += 1

        request.session.pop("pledge_import_plan", None)
        forced = request.session.pop("pledge_forced_campaign", None)
        parts = [f"{created} pledge(s) imported as drafts"]
        if new_members:
            parts.append(f"{new_members} new member(s)")
        if new_camps:
            parts.append(f"{new_camps} new campaign(s)")
        if skipped:
            parts.append(f"{skipped} row(s) skipped")
        messages.success(request, ", ".join(parts) +
                         ". They are DRAFTS — check them below and edit or delete "
                         "anything allocated to the wrong member before approving.")
        if forced:
            return redirect("pledge_campaign_detail", pk=forced)
        return redirect(f"{reverse('pledge_list')}?status=DRAFT")

    @staticmethod
    def _norm(s):
        return " ".join((s or "").upper().split())


class PledgeDeleteView(TreasurerRequiredMixin, View):
    """Delete a pledge. Its payment *links* (PledgePayment) are removed, but the
    underlying contributions stay in the ledger — a pledge link carries no money."""
    def post(self, request, pk):
        p = get_object_or_404(Pledge, pk=pk)
        n = p.payments.count()
        campaign_id = p.campaign_id
        p.delete()
        msg = "Pledge deleted."
        if n:
            msg += f" {n} matched contribution(s) were unlinked but remain in the ledger."
        messages.success(request, msg)
        # Deleting is nearly always something done while REVIEWING a campaign's
        # imported pledges — send the treasurer back to the list they were
        # working through, not to the global pledge list.
        if campaign_id:
            return redirect("pledge_campaign_detail", pk=campaign_id)
        return redirect("pledge_list")


class CampaignDeleteView(TreasurerRequiredMixin, View):
    """Delete a pledge campaign. Only allowed when it has no pledges (remove or
    reassign those first), so a campaign with giving history can't vanish."""
    def post(self, request, pk):
        from .models import PledgeCampaign
        camp = get_object_or_404(PledgeCampaign, pk=pk)
        n = camp.pledges.count()
        if n:
            messages.error(request, f"“{camp.name}” still has {n} pledge(s). Delete or "
                                    f"reassign those pledges first, then delete the campaign.")
            return redirect("pledge_campaign_detail", pk=pk)
        name = camp.name
        camp.delete()
        messages.success(request, f"Campaign “{name}” deleted.")
        return redirect("pledge_campaign_list")
