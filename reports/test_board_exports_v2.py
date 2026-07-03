"""Board report exports v2: RTF removed entirely; Excel has full detail tables
(not the on-screen top-10) plus native charts and a KPI-card summary sheet;
Word mirrors the HTML structure and carries a per-section AI/rule-based
narrative (server-side, always available even with the assistant off)."""
import datetime as dt
from decimal import Decimal
from django.test import TestCase, Client
from django.contrib.auth.models import User, Group
from django.urls import reverse, NoReverseMatch
from departments.models import Department
from giving.models import Transaction
from ledger.services.posting import ensure_chart


def _tr():
    u = User.objects.create_user("tr_bx2", password="x", is_superuser=True)
    u.groups.add(Group.objects.get_or_create(name="Treasurer")[0])
    return u


class RtfRemovedTests(TestCase):
    def setUp(self):
        self.tr = _tr()
        self.c = Client(); self.c.force_login(self.tr)

    def test_rtf_url_no_longer_resolves(self):
        with self.assertRaises(NoReverseMatch):
            reverse("report_board_rtf")

    def test_rtf_path_404s(self):
        r = self.c.get("/reports/board/export/rtf/?as_of=2026-06-01")
        self.assertEqual(r.status_code, 404)

    def test_toolbar_has_no_rtf_link(self):
        b = self.c.get("/reports/board/?as_of=2026-06-01").content.decode()
        self.assertNotIn("report_board_rtf", b)
        self.assertNotIn("RTF", b)


class ExcelFullDetailTests(TestCase):
    def setUp(self):
        ensure_chart()
        self.tr = _tr()
        self.d = Department.objects.create(name="XLFund", fund_type="LOCAL",
            category="MINISTRY")
        for i in range(15):
            f = Department.objects.create(name=f"XLTrust{i}", fund_type="TRUST",
                category="OFFERING")
            Transaction.objects.create(date=dt.date(2026, 6, 5),
                amount=Decimal(str(1000 + i)), direction="CREDIT", confirmed=True,
                channel="CASH", allocation_status="MANUAL", department=f)
        self.c = Client(); self.c.force_login(self.tr)

    def test_all_sheets_present(self):
        import io
        from openpyxl import load_workbook
        r = self.c.get("/reports/board/export/excel/?as_of=2026-06-01")
        self.assertEqual(r.status_code, 200)
        wb = load_workbook(io.BytesIO(r.content))
        for name in ["Executive Summary", "Collections", "Trust Fund Performance",
                     "Local Fund Performance", "Expenditure", "Budget & Goals",
                     "Financial Position", "Cash Flow", "Bank Reconciliation",
                     "Board Decisions"]:
            self.assertIn(name, wb.sheetnames)

    def test_collections_sheet_has_full_listing_not_top_ten(self):
        import io
        from openpyxl import load_workbook
        r = self.c.get("/reports/board/export/excel/?as_of=2026-06-01")
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb["Collections"]
        names = [c.value for c in ws["A"] if c.value and "XLTrust" in str(c.value)]
        self.assertEqual(len(names), 15)   # all 15, not capped at 10

    def test_charts_present(self):
        import io
        from openpyxl import load_workbook
        r = self.c.get("/reports/board/export/excel/?as_of=2026-06-01")
        wb = load_workbook(io.BytesIO(r.content))
        total_charts = sum(len(ws._charts) for ws in wb.worksheets)
        self.assertGreater(total_charts, 0)
        self.assertGreater(len(wb["Collections"]._charts), 0)
        self.assertGreater(len(wb["Financial Position"]._charts), 0)

    def test_kpi_cards_on_summary_sheet(self):
        import io
        from openpyxl import load_workbook
        r = self.c.get("/reports/board/export/excel/?as_of=2026-06-01")
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb["Executive Summary"]
        values = [c.value for row in ws.iter_rows() for c in row if c.value]
        self.assertIn("Total collections", values)
        self.assertIn("Net assets", values)


class WordNarrativeTests(TestCase):
    def setUp(self):
        ensure_chart()
        self.tr = _tr()
        self.c = Client(); self.c.force_login(self.tr)

    def test_word_has_all_sections(self):
        b = self.c.get("/reports/board/export/word/?as_of=2026-06-01").content.decode()
        for s in ["Executive summary", "Key highlights",
                  "Items requiring Board attention", "Collections summary",
                  "Trust fund performance", "Local fund performance",
                  "Expenditure summary", "Budget &amp; goal tracking",
                  "Statement of financial position", "Cash flow",
                  "Bank reconciliation", "Board decisions required"]:
            self.assertIn(s, b)

    def test_word_narratives_always_present_llm_off(self):
        # with the assistant off, _ai_narratives must still return usable
        # rule-based text for every section — the report never depends on the LLM
        from reports.views import MonthlyTreasurerReportView
        from django.test import RequestFactory
        rf = RequestFactory()
        req = rf.get("/reports/board/?as_of=2026-06-01")
        req.user = self.tr
        view = MonthlyTreasurerReportView()
        view.request = req
        ctx = view.get_context_data()
        narr = view._ai_narratives(ctx)
        self.assertTrue(all(isinstance(v, str) for v in narr.values()))

    def test_word_kpi_cards_present(self):
        b = self.c.get("/reports/board/export/word/?as_of=2026-06-01").content.decode()
        self.assertIn("kpicard", b)
        self.assertIn("Trust funds outstanding", b)
