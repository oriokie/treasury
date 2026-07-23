"""Split from reports/views.py (P1-2). Behaviour identical; the
package __init__ reproduces the original module namespace."""
from decimal import Decimal
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.shortcuts import get_object_or_404, redirect, render
from django.views.generic import TemplateView
from core.permissions import (ReportAccessMixin, TreasurerRequiredMixin,
                              RightRequiredMixin, ReportAccessMixin)
from core.utils import parse_period, safe_json
from cashbook.models import Expense
from departments.models import Department
from giving.models import Transaction
from members.models import Member
from ..services import balances
from ..exports import csv_response
import datetime as dt
from ._shared import PeriodMixin


class ReportIndexView(ReportAccessMixin, TemplateView):
    template_name = "reports/index.html"

class MonthlyReportView(PeriodMixin, TemplateView):
    template_name = "reports/monthly.html"

    def get(self, request, *args, **kwargs):
        start, end = self.period()
        rows = balances.department_summary(start, end)
        if request.GET.get("export") == "csv":
            data = [(r["department"].name,
                     "Trust" if r["is_trust"] else "Local",
                     r["opening"], r["receipts"], r["expenses"], r["closing"])
                    for r in rows]
            return csv_response(
                f"monthly_{start}_{end}.csv",
                ["Fund", "Type", "Opening", "Receipts", "Expenses", "Closing"], data)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        rows = balances.department_summary(ctx["start"], ctx["end"])
        ctx["rows"] = rows
        ctx["totals"] = balances.totals(rows)
        ctx["trust_total"] = sum(r["receipts"] for r in rows if r["is_trust"])
        ctx["local_total"] = sum(r["receipts"] for r in rows if not r["is_trust"])
        return ctx

class OfferingSummaryView(PeriodMixin, TemplateView):
    template_name = "reports/offering_summary.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        data = balances.offering_summary(ctx["start"], ctx["end"])
        ctx["sabbaths"] = data["sabbaths"]
        ctx["rows"] = data["rows"]
        return ctx

class TitheReportView(PeriodMixin, TemplateView):
    template_name = "reports/tithe.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        s, e = ctx["start"], ctx["end"]
        ctx["tithe"] = balances.tithe_total(s, e)
        ctx["count"] = Transaction.objects.filter(
            date__gte=s, date__lte=e, direction=Transaction.Direction.CREDIT,
            department__name__icontains="tithe").count()
        return ctx

class GroupGivingView(PeriodMixin, TemplateView):
    template_name = "reports/by_group.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        data = balances.giving_by_group(ctx["start"], ctx["end"])
        labels = dict(Member.Group.choices)
        ctx["rows"] = [{"group": labels.get(k, k), "total": v}
                       for k, v in sorted(data.items(), key=lambda x: -x[1])]
        ctx["grand_total"] = sum(data.values())
        return ctx

class ExpenseReportView(PeriodMixin, TemplateView):
    template_name = "reports/expenses.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        s, e = ctx["start"], ctx["end"]
        eff = Q(status__in=[Expense.Status.APPROVED, Expense.Status.PAID])
        base = (Expense.objects.filter(eff, date__gte=s, date__lte=e)
                .exclude(doc_class=Expense.DocClass.LIABILITY))
        # consolidate by the top-level fund (sub-account spend rolls into its parent)
        from collections import defaultdict
        from departments.models import Department
        from ..services.budget import budget_amounts_bulk
        all_depts = list(Department.objects.select_related("parent"))
        parent_of = {}
        tops = {}
        for d in all_depts:
            top = d.parent or d
            parent_of[d.id] = top.name
            tops.setdefault(top.name, top)
        top_budgets = budget_amounts_bulk(e.year, tops.values())
        budget_of = {name: top_budgets.get(top.id) for name, top in tops.items()}
        agg = defaultdict(Decimal)
        for r in base.values("department_id").annotate(total=Sum("amount")):
            agg[parent_of.get(r["department_id"], "Unallocated")] += r["total"] or Decimal(0)
        ctx["by_dept"] = [{"name": k, "total": v, "budget": budget_of.get(k)}
                          for k, v in sorted(agg.items(), key=lambda x: -x[1])]
        ctx["by_category"] = (base.values("category")
                              .annotate(total=Sum("amount")).order_by("-total"))
        ctx["by_claimant"] = (base.exclude(claimant="").values("claimant")
                              .annotate(total=Sum("amount")).order_by("-total"))
        ctx["outstanding"] = Expense.objects.filter(
            status__in=[Expense.Status.PENDING, Expense.Status.APPROVED]).order_by("date")
        ctx["cat_labels"] = dict(Expense.Category.choices)
        return ctx

class IncomeExpenditureView(PeriodMixin, TemplateView):
    template_name = "reports/income_expenditure.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        s, e = ctx["start"], ctx["end"]
        # Income = the church's own (local) revenue only. Trust funds (tithe,
        # the remitted share of combined/thanksgiving offerings, etc.) are
        # collected on behalf of the field — a liability, not revenue — so they
        # are excluded here and shown separately as a memo.
        income = (Transaction.objects.confirmed_credits().filter(
            date__gte=s, date__lte=e, department__is_trust=False,
            excluded_from_income=False)
            .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        trust_collected = (Transaction.objects.confirmed_credits().filter(
            date__gte=s, date__lte=e, department__is_trust=True,
            excluded_from_income=False)
            .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        # Expenditure excludes remittances (they settle the trust liability) AND
        # capital purchases (an asset, not consumed in the period — only any
        # depreciation belongs in an income & expenditure account). Capital
        # additions are shown separately as a memo.
        paid = [Expense.Status.APPROVED, Expense.Status.PAID]
        expense = (Expense.objects.filter(
            date__gte=s, date__lte=e, status__in=paid)
            .exclude(doc_class=Expense.DocClass.LIABILITY)
            .exclude(expenditure_type=Expense.ExpenditureType.CAPITAL)
            .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        capital = (Expense.objects.filter(
            date__gte=s, date__lte=e, status__in=paid,
            expenditure_type=Expense.ExpenditureType.CAPITAL)
            .exclude(doc_class=Expense.DocClass.LIABILITY)
            .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        remittances = (Expense.objects.filter(
            date__gte=s, date__lte=e, status__in=paid,
            category=Expense.Category.REMITTANCE)
            .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        ctx["income"] = income
        ctx["expense"] = expense
        # Gain/(loss) on asset disposals in the period — the only part of a disposal
        # that belongs in the income result (the proceeds themselves are a capital
        # receipt, excluded from income above). Sourced from the metrics registry.
        from core.metrics import metrics
        non_cash = metrics.non_cash_items(s, e)
        disposal_gl = non_cash["disposal_gain_loss"]
        ctx["disposal_gain_loss"] = disposal_gl
        ctx["net"] = income - expense + disposal_gl
        # Non-cash contributions: assets donated in kind. There is no Transaction
        # for these — they are credited to net assets, not income (EAM §9.4) — so
        # they are reported as their own section, outside the cash income total,
        # and deliberately excluded from `net`.
        ctx["non_cash"] = non_cash
        ctx["donated_assets"] = non_cash["donated_assets"]
        ctx["donated_assets_detail"] = metrics.donated_assets_detail(s, e)
        # Depreciation is a real, posted expense but not a payment, so it does not
        # belong in the cash expenditure figure above. It is reported with the
        # other non-cash items, and the surplus after them is shown alongside the
        # cash result so the statement tells the whole story without either
        # figure being mistaken for the other.
        ctx["depreciation"] = non_cash["depreciation"]
        ctx["depreciation_charge"] = -ctx["depreciation"]   # a deduction, for display
        ctx["net_after_noncash"] = (disposal_gl + income - expense
                                    + ctx["donated_assets"] - ctx["depreciation"])
        ctx["has_noncash"] = bool(ctx["donated_assets"] or ctx["depreciation"])
        ctx["capital"] = capital
        ctx["trust_collected"] = trust_collected
        ctx["remittances"] = remittances
        return ctx

class CashBookView(PeriodMixin, TemplateView):
    template_name = "reports/cashbook.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        s, e = ctx["start"], ctx["end"]
        entries = []
        # Cash-effective receipts only, per the one canonical definition:
        # confirmed, not reversed, not a reversal (confirmed_credits), and not
        # a bank-memo row (a manually-receipted bank line whose cash is counted
        # on its envelope entry). The cash book previously summed EVERY credit —
        # unconfirmed, reversed originals, their reversal contras AND memo
        # rows — overstating receipts on all four counts.
        for t in (Transaction.objects.confirmed_credits()
                  .filter(date__gte=s, date__lte=e)
                  .exclude(channel=Transaction.Channel.BANK,
                           manual_receipt=True)):
            entries.append({"date": t.date, "desc": t.payer_name or t.reference or "Receipt",
                            "credit": t.amount, "debit": None})
        for x in Expense.objects.filter(date__gte=s, date__lte=e,
                                        status__in=[Expense.Status.APPROVED, Expense.Status.PAID]):
            entries.append({"date": x.date, "desc": x.description,
                            "credit": None, "debit": x.amount})
        entries.sort(key=lambda r: r["date"])
        running = Decimal(0)
        for en in entries:
            running += (en["credit"] or 0) - (en["debit"] or 0)
            en["balance"] = running
        ctx["entries"] = entries
        ctx["closing"] = running
        return ctx

class ReconciliationView(PeriodMixin, TemplateView):
    template_name = "reports/reconciliation.html"

    def get_context_data(self, **kwargs):
        from statements.models import BankReconciliation
        ctx = super().get_context_data(**kwargs)
        s, e = ctx["start"], ctx["end"]
        bank = Transaction.objects.active().filter(channel=Transaction.Channel.BANK,
                                          date__gte=s, date__lte=e)
        ctx["ledger_credits"] = bank.filter(direction=Transaction.Direction.CREDIT).aggregate(
            t=Sum("amount"))["t"] or Decimal(0)
        ctx["ledger_debits"] = bank.filter(direction=Transaction.Direction.DEBIT).aggregate(
            t=Sum("amount"))["t"] or Decimal(0)
        ctx["unreconciled"] = bank.filter(
            allocation_status=Transaction.Status.REVIEW).order_by("date")

        # Proper bank reconciliation: the bank STATEMENT balance reconciles to the
        # BOOK (cashbook) balance through timing/at-hand adjustments — NOT to the
        # raw bank-credit total, which is why a naive "credits vs bank balance"
        # comparison looks wildly off. Show the latest statement reconciliation.
        rec = (BankReconciliation.objects.order_by("-statement_date")
               .prefetch_related("items").first())
        ctx["rec"] = rec
        if rec:
            adds = [i for i in rec.items.all() if i.effect == "ADD"]
            subs = [i for i in rec.items.all() if i.effect == "SUBTRACT"]
            ctx["rec_adds"] = adds
            ctx["rec_subs"] = subs
            ctx["rec_add_total"] = sum((i.amount for i in adds), Decimal(0))
            ctx["rec_sub_total"] = sum((i.amount for i in subs), Decimal(0))
            ctx["rec_computed_book"] = (rec.bank_balance + ctx["rec_add_total"]
                                        - ctx["rec_sub_total"])
            ctx["rec_variance"] = ctx["rec_computed_book"] - (rec.book_balance or Decimal(0))
        return ctx

class AnnualSummaryView(ReportAccessMixin, TemplateView):
    template_name = "reports/annual.html"

    def get_context_data(self, **kwargs):
        from django.db.models.functions import ExtractYear
        ctx = super().get_context_data(**kwargs)
        income = (Transaction.objects.confirmed_credits()
                  .filter(excluded_from_income=False)
                  .annotate(yr=ExtractYear("date"))
                  .values("yr").annotate(total=Sum("amount")).order_by("yr"))
        expense = (Expense.objects.filter(
            status__in=[Expense.Status.APPROVED, Expense.Status.PAID])
            .exclude(doc_class=Expense.DocClass.LIABILITY)
            .annotate(yr=ExtractYear("date"))
            .values("yr").annotate(total=Sum("amount")).order_by("yr"))
        inc = {r["yr"]: r["total"] for r in income}
        exp = {r["yr"]: r["total"] for r in expense}
        years = sorted(set(inc) | set(exp))
        ctx["rows"] = [{"year": y, "income": inc.get(y, 0), "expense": exp.get(y, 0),
                        "net": (inc.get(y, 0) or 0) - (exp.get(y, 0) or 0)} for y in years]
        # historical reference years (collection / trust fund / expenditure)
        from core.models import HistoricalYear, HistoricalMonth
        ctx["historical"] = list(HistoricalYear.objects.all())
        # seasonality: average collection / trust / expenditure by calendar month
        MN = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep",
              "Oct", "Nov", "Dec"]
        agg = {m: {"c": 0.0, "t": 0.0, "e": 0.0, "n": 0} for m in range(1, 13)}
        for hm in HistoricalMonth.objects.all():
            a = agg[hm.month]
            a["c"] += float(hm.collection); a["t"] += float(hm.trust_fund)
            a["e"] += float(hm.expenditure); a["n"] += 1
        season = {"labels": MN,
                  "collection": [round(agg[m]["c"] / agg[m]["n"], 2) if agg[m]["n"] else 0 for m in range(1, 13)],
                  "trust": [round(agg[m]["t"] / agg[m]["n"], 2) if agg[m]["n"] else 0 for m in range(1, 13)],
                  "expenditure": [round(agg[m]["e"] / agg[m]["n"], 2) if agg[m]["n"] else 0 for m in range(1, 13)]}
        ctx["season_json"] = safe_json(season)
        ctx["has_season"] = HistoricalMonth.objects.exists()
        return ctx

class HistoricalYearManageView(TreasurerRequiredMixin, TemplateView):
    """Add, edit, or remove prior-year comparison figures, now with per-month
    detail. When a year has monthly rows, its yearly totals are computed from
    them (so the two always agree); a year with no months keeps the figure typed
    by hand. Monthly data can be imported from Excel for fast back-filling and
    enables month-on-month trend analysis."""
    template_name = "reports/historical_manage.html"

    MONTHS = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    @staticmethod
    def _recompute_year(year):
        """Set a HistoricalYear's totals from its months, if any exist."""
        from decimal import Decimal
        from django.db.models import Sum
        from core.models import HistoricalYear, HistoricalMonth
        months = HistoricalMonth.objects.filter(year=year)
        if not months.exists():
            return
        agg = months.aggregate(c=Sum("collection"), t=Sum("trust_fund"),
                               e=Sum("expenditure"))
        HistoricalYear.objects.update_or_create(year=year, defaults=dict(
            collection=agg["c"] or Decimal(0), trust_fund=agg["t"] or Decimal(0),
            expenditure=agg["e"] or Decimal(0), note="Computed from monthly records"))

    def get(self, request, *args, **kwargs):
        if request.GET.get("sample"):
            return self._sample_xlsx()
        return super().get(request, *args, **kwargs)

    def _sample_xlsx(self):
        import io
        import datetime as _d
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Monthly history"
        head = ["Year", "Month (1-12)", "Collection", "Trust fund", "Expenditure"]
        ws.append(head)
        for c in range(1, len(head) + 1):
            cell = ws.cell(1, c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F5F4F")
            cell.alignment = Alignment(horizontal="center")
        # a couple of illustrative rows for last year
        ly = _d.date.today().year - 1
        for m, coll, tr, ex in [(1, 120000, 70000, 45000), (2, 98000, 60000, 52000)]:
            ws.append([ly, m, coll, tr, ex])
        for col, w in zip("ABCDE", (8, 14, 14, 14, 14)):
            ws.column_dimensions[col].width = w
        info = wb.create_sheet("How to use")
        for line in [
            "One row per month. Year and Month (1-12) identify the period.",
            "Collection = total receipts that month (all funds).",
            "Trust fund = the portion that is trust/remittable.",
            "Expenditure = total spending that month.",
            "Re-importing a month overwrites that month. Yearly totals are computed automatically.",
        ]:
            info.append([line])
        info.column_dimensions["A"].width = 90
        buf = io.BytesIO(); wb.save(buf)
        resp = HttpResponse(buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = 'attachment; filename="monthly_history_template.xlsx"'
        return resp

    def get_context_data(self, **kwargs):
        from collections import defaultdict
        from core.models import HistoricalYear, HistoricalMonth
        ctx = super().get_context_data(**kwargs)
        ctx["years"] = HistoricalYear.objects.all()
        by_year = defaultdict(lambda: [None] * 13)
        for hm in HistoricalMonth.objects.all():
            by_year[hm.year][hm.month] = hm
        month_rows = []
        for yr in sorted(by_year, reverse=True):
            cells = by_year[yr]
            present = [c for c in cells[1:] if c]
            month_rows.append({
                "year": yr, "cells": cells[1:],
                "collection": sum((c.collection for c in present), 0),
                "trust_fund": sum((c.trust_fund for c in present), 0),
                "expenditure": sum((c.expenditure for c in present), 0),
                "count": len(present)})
        ctx["month_rows"] = month_rows
        ctx["months"] = self.MONTHS[1:]
        return ctx

    def post(self, request, *args, **kwargs):
        from decimal import Decimal, InvalidOperation
        from core.models import HistoricalYear, HistoricalMonth
        action = request.POST.get("action")

        def dec(k):
            return Decimal(str(request.POST.get(k) or "0").replace(",", ""))

        if action == "delete":
            HistoricalYear.objects.filter(pk=request.POST.get("pk")).delete()
            messages.success(request, "Historical year removed.")
            return redirect("historical_manage")

        if action == "delete_year_all":
            try:
                yr = int(request.POST.get("year"))
            except (TypeError, ValueError):
                return redirect("historical_manage")
            HistoricalMonth.objects.filter(year=yr).delete()
            HistoricalYear.objects.filter(year=yr).delete()
            messages.success(request, f"Deleted all historical data for {yr}.")
            return redirect("historical_manage")

        if action == "save_month":
            try:
                year = int(request.POST.get("year"))
                month = int(request.POST.get("month"))
                assert 1 <= month <= 12
            except (TypeError, ValueError, AssertionError):
                messages.error(request, "Enter a valid year and month (1–12).")
                return redirect("historical_manage")
            HistoricalMonth.objects.update_or_create(
                year=year, month=month, defaults=dict(
                    collection=dec("collection"), trust_fund=dec("trust_fund"),
                    expenditure=dec("expenditure")))
            self._recompute_year(year)
            messages.success(request, f"Saved {self.MONTHS[month]} {year}.")
            return redirect("historical_manage")

        if action == "delete_month":
            hm = HistoricalMonth.objects.filter(pk=request.POST.get("pk")).first()
            if hm:
                yr = hm.year; hm.delete(); self._recompute_year(yr)
                messages.success(request, "Month removed.")
            return redirect("historical_manage")

        if action == "import":
            return self._import(request)

        # save a whole-year figure by hand (only when there are no months for it)
        try:
            year = int(request.POST.get("year"))
            HistoricalYear.objects.update_or_create(
                year=year, defaults=dict(collection=dec("collection"),
                    trust_fund=dec("trust_fund"), expenditure=dec("expenditure"),
                    note=(request.POST.get("note") or "Entered manually")[:200]))
            messages.success(request, f"Saved historical figures for {year}.")
        except (TypeError, ValueError, InvalidOperation):
            messages.error(request, "Enter a valid year and numeric amounts.")
        return redirect("historical_manage")

    def _import(self, request):
        from decimal import Decimal, InvalidOperation
        import openpyxl
        from core.models import HistoricalMonth
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "Choose an Excel file to import.")
            return redirect("historical_manage")
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx; _lx("historical import")
            messages.error(request, "That file couldn't be read as an Excel workbook.")
            return redirect("historical_manage")
        ws = wb["Monthly history"] if "Monthly history" in wb.sheetnames else wb.active
        n = 0; years = set()
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or row is None:
                continue
            try:
                year = int(row[0]); month = int(row[1])
                if not (1 <= month <= 12):
                    continue
                def d(v):
                    return Decimal(str(v or "0").replace(",", ""))
                HistoricalMonth.objects.update_or_create(
                    year=year, month=month, defaults=dict(
                        collection=d(row[2]), trust_fund=d(row[3]),
                        expenditure=d(row[4])))
                years.add(year); n += 1
            except (TypeError, ValueError, InvalidOperation, IndexError):
                continue
        for yr in years:
            self._recompute_year(yr)
        messages.success(request, f"Imported {n} monthly record(s) across "
                                  f"{len(years)} year(s). Yearly totals updated.")
        return redirect("historical_manage")

class AuditLogView(ReportAccessMixin, TemplateView):
    template_name = "reports/audit.html"

    def _models(self):
        from giving.models import Transaction as T, AllocationRule as AR
        from cashbook.models import Expense as X
        from members.models import Member as M
        from envelopes.models import EnvelopeBatch as EB
        candidates = {"Transaction": T, "Expense": X, "Member": M,
                     "Allocation rule": AR, "Envelope batch": EB}
        return {n: m for n, m in candidates.items() if hasattr(m, "history")}

    def _collect(self, request, cap=1500):
        import datetime as dt
        models = self._models()
        model_f = request.GET.get("model", "")
        user_f = request.GET.get("user", "")
        type_f = request.GET.get("type", "")          # +, ~, -
        q = (request.GET.get("q", "") or "").strip().lower()

        def _date(name):
            raw = request.GET.get(name)
            try:
                return dt.date.fromisoformat(raw) if raw else None
            except ValueError:
                return None
        start, end = _date("start"), _date("end")

        records, users = [], set()
        # Department/SplitFund names for AllocationRule, prefetched once: its
        # __str__() touches self.split_fund or self.department, and
        # h.instance (a full model reconstructed from the historical row) has
        # no select_related — calling str() on it for every historical
        # AllocationRule row triggered a fresh FK query each time (up to
        # ~1500 extra queries for a church with a large rule history).
        from departments.models import Department
        from giving.models import SplitFund
        dept_names = dict(Department.objects.values_list("id", "name"))
        split_names = dict(SplitFund.objects.values_list("id", "name"))

        for name, model in models.items():
            hq = model.history.all().select_related("history_user")
            if start:
                hq = hq.filter(history_date__date__gte=start)
            if end:
                hq = hq.filter(history_date__date__lte=end)
            if user_f:
                hq = hq.filter(history_user__username=user_f)
            if type_f in ("+", "~", "-"):
                hq = hq.filter(history_type=type_f)
            # collect the user list for the filter dropdown (cheap, distinct)
            for u in (model.history.exclude(history_user__isnull=True)
                      .values_list("history_user__username", flat=True).distinct()[:200]):
                users.add(u)
            if model_f and model_f != name:
                continue
            for h in hq.order_by("-history_date")[:cap]:
                try:
                    if name == "Allocation rule":
                        # build the display string from the historical row's
                        # own FK id columns + the prefetched name maps,
                        # instead of str(h.instance) (see note above)
                        target = (split_names.get(h.split_fund_id)
                                  or dept_names.get(h.department_id) or "—")
                        obj = f"{h.reference} -> {target}"
                    else:
                        obj = str(h.instance)
                except Exception:
                    from core.utils import log_exception as _lx; _lx('reports/views.py')
                    obj = f"{name} #{h.id}"
                uname = getattr(h.history_user, "username", "") or "system"
                if q and q not in (obj + " " + uname + " " + name).lower():
                    continue
                records.append({
                    "model": name, "when": h.history_date, "user": uname,
                    "type": h.get_history_type_display(), "obj": obj})
        records.sort(key=lambda r: r["when"], reverse=True)
        return records, sorted(users)

    def get(self, request, *args, **kwargs):
        records, users = self._collect(request)
        if request.GET.get("export") == "csv":
            from reports.exports import csv_response
            header = ["When", "Record type", "Change", "By", "Detail"]
            rows = [[r["when"].strftime("%Y-%m-%d %H:%M:%S"), r["model"],
                     r["type"], r["user"], r["obj"]] for r in records]
            return csv_response("audit_log.csv", header, rows)

        from django.core.paginator import Paginator
        paginator = Paginator(records, 50)
        page = paginator.get_page(request.GET.get("page"))
        ctx = self.get_context_data(**kwargs)
        ctx.update({
            "page_obj": page, "records": page.object_list,
            "total": paginator.count,
            "models": list(self._models().keys()), "users": users,
            "f": {"model": request.GET.get("model", ""),
                  "user": request.GET.get("user", ""),
                  "type": request.GET.get("type", ""),
                  "q": request.GET.get("q", ""),
                  "start": request.GET.get("start", ""),
                  "end": request.GET.get("end", "")},
            "querystring": _qs_without(request, "page"),
        })
        return self.render_to_response(ctx)

def _qs_without(request, *drop):
    from urllib.parse import urlencode
    items = [(k, v) for k, v in request.GET.items() if k not in drop and v]
    return urlencode(items)

class MemberStatementView(PeriodMixin, TemplateView):
    template_name = "reports/member_statement.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        member = get_object_or_404(Member, pk=kwargs["pk"])
        s, e = ctx["start"], ctx["end"]
        txns = (Transaction.objects.confirmed_credits().filter(
            member=member, date__gte=s, date__lte=e,
            excluded_from_income=False).values("department__name")
            .annotate(total=Sum("amount")).order_by("-total"))
        ctx["member"] = member
        ctx["rows"] = txns
        ctx["total"] = sum(r["total"] for r in txns)
        return ctx
