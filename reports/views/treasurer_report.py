"""Split from reports/views.py (P1-2). Behaviour identical; the
package __init__ reproduces the original module namespace."""
from decimal import Decimal
from django.db.models import Sum, Count, Q
from django.shortcuts import get_object_or_404, redirect, render
from django.views import View
from django.views.generic import TemplateView
from core.permissions import (ReportAccessMixin, TreasurerRequiredMixin,
                              RightRequiredMixin, ReportAccessMixin)
from core.utils import parse_period, safe_json
from cashbook.models import Expense
from giving.models import Transaction
from ..services import balances
import datetime as _dt
from core.models import SiteConfig
from ..services.goals import (sentence_fund_name as _sfund,      # noqa: E402
                             camp_goal_records as _camp_goal_records)


class MonthlyTreasurerReportView(ReportAccessMixin, TemplateView):
    """Comprehensive monthly Treasurer's Report: collections, trust & LCB trends,
    a multi-year trend, expense and local-fund breakdowns, the income statement,
    statement of financial position, cash-flow statements, and the latest bank
    reconciliation — each with a short plain-language note. Compact, board-ready."""
    template_name = "reports/monthly_treasurer.html"

    def get_context_data(self, **kwargs):
        import datetime as _dt
        from decimal import Decimal
        from django.db.models import Sum
        from reports.services import treasurer as T
        from reports.services import budget as T_budget
        ctx = super().get_context_data(**kwargs)

        # month selection (default current month). An <input type="month">
        # submits "YYYY-MM"; a date picker may submit "YYYY-MM-DD" — accept both.
        raw = (self.request.GET.get("as_of") or "").strip()
        as_of = None
        for fmt in ("%Y-%m-%d", "%Y-%m"):
            try:
                as_of = _dt.datetime.strptime(raw, fmt).date()
                break
            except ValueError:
                continue
        if as_of is None:
            as_of = _dt.date.today()
        s, e = T.month_bounds(as_of)
        ctx["as_of"] = as_of; ctx["start"] = s; ctx["end"] = e
        ctx["today"] = _dt.date.today()
        ctx["church"] = SiteConfig.get().church_name

        # 1) collections summary
        csum = T.collections_summary(s, e)
        rows = csum["rows"]
        ctx["collections"] = csum

        # 2) trust receipted trend (current + previous 2 months)
        trust_trend = T.trust_receipted_trend(as_of, months=3)
        ctx["trust_trend"] = trust_trend
        ctx["trust_trend_json"] = safe_json({
            "labels": [c["label"] for c in trust_trend.get("columns", [])],
            "totals": [float(v or 0) for v in trust_trend.get("col_totals", [])],
        })
        # 3) LCB sub-account trend (all LCB accounts, current + previous 2 months)
        lcb_trend = T.lcb_subaccount_trend(as_of, months=3)
        ctx["lcb_trend"] = lcb_trend
        _lcb_latest = sorted(
            [r for r in lcb_trend.get("rows", []) if r.get("cells")],
            key=lambda r: r["cells"][-1], reverse=True)[:6]
        ctx["lcb_trend_json"] = safe_json({
            "labels": [r["dept"].name for r in _lcb_latest],
            "amounts": [float(r["cells"][-1] or 0) for r in _lcb_latest],
        })
        # 4) 5-year YTD trend (+ JSON for a chart)
        yearly = T.yearly_trend(as_of, years=5)
        ctx["yearly"] = yearly
        ctx["yearly_json"] = safe_json([{
            "year": str(y["year"]), "collection": float(y["collection"] or 0),
            "trust": float(y["trust"] or 0), "expense": float(y["expense"] or 0),
        } for y in yearly])
        # 5) LCB expenditure statement (fixed: matches all LCB departments)
        ctx["lcb_expenditure"] = T.lcb_expenditure(s, e)
        # 6) local funds movement statement: opening, receipts, expenses, closing
        ctx["local_statement"] = T.local_funds_statement(s, e)
        ctx["local_statement_more"] = max(
            0, len(ctx["local_statement"].get("rows", [])) - 10)

        # 7) income statement (recurrent basis)
        paid = T.PAID
        income = (Transaction.objects.confirmed_credits().filter(
            date__gte=s, date__lte=e, department__is_trust=False,
            excluded_from_income=False).aggregate(t=Sum("amount"))["t"] or Decimal(0))
        op_exp = (Expense.objects.filter(date__gte=s, date__lte=e, status__in=paid)
                  .exclude(doc_class=Expense.DocClass.LIABILITY)
                  .exclude(expenditure_type=Expense.ExpenditureType.CAPITAL)
                  .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        capital = (Expense.objects.filter(date__gte=s, date__lte=e, status__in=paid,
                   expenditure_type=Expense.ExpenditureType.CAPITAL)
                   .exclude(doc_class=Expense.DocClass.LIABILITY)
                   .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        from core.metrics import metrics as _metrics
        _nc = _metrics.non_cash_items(s, e)
        ctx["non_cash"] = _nc
        ctx["income_stmt"] = {"income": income, "expense": op_exp,
                              "surplus": income - op_exp, "capital": capital,
                              # the same result once assets consumed, gifts in
                              # kind and disposals are taken into account
                              "depreciation": _nc["depreciation"],
                              "donated_assets": _nc["donated_assets"],
                              "disposal_gain_loss": _nc["disposal_gain_loss"],
                              "surplus_accrual": income - op_exp + _nc["net"]}
        # detailed line items for a report-form income statement
        rev_rows = (Transaction.objects.confirmed_credits().filter(
            date__gte=s, date__lte=e, department__is_trust=False,
            excluded_from_income=False).values("department__name")
            .annotate(t=Sum("amount")).order_by("-t"))
        exp_cat = (Expense.objects.filter(date__gte=s, date__lte=e, status__in=paid)
                   .exclude(doc_class=Expense.DocClass.LIABILITY)
                   .exclude(expenditure_type=Expense.ExpenditureType.CAPITAL)
                   .values("category").annotate(t=Sum("amount")).order_by("-t"))
        _cat = dict(Expense.Category.choices)
        ctx["income_detail"] = {
            "revenue": [{"name": r["department__name"] or "Unallocated",
                         "amount": r["t"]} for r in rev_rows if r["t"]],
            "expenses": [{"name": _cat.get(r["category"], r["category"]),
                          "amount": r["t"]} for r in exp_cat if r["t"]]}
        # per-fund collections detail (trust then local, with amounts)
        ctx["collection_detail"] = {
            "trust": sorted([r for r in rows if r["is_trust"] and r["receipts"]],
                            key=lambda r: r["receipts"], reverse=True),
            "local": sorted([r for r in rows if not r["is_trust"] and r["receipts"]],
                            key=lambda r: r["receipts"], reverse=True)}
        ctx["collection_detail_more"] = {
            "trust": max(0, len(ctx["collection_detail"]["trust"]) - 10),
            "local": max(0, len(ctx["collection_detail"]["local"]) - 10)}

        # 8) statement of financial position (summary, period end)
        sofp_rows = balances.department_summary(None, e)
        cash = sum((r["closing"] for r in sofp_rows), Decimal(0))
        trust_payable = sum((r["closing"] for r in sofp_rows if r["is_trust"]), Decimal(0))
        local_funds_total = cash - trust_payable
        from cashbook.views import open_payables_total, open_accruals_total
        payables = open_payables_total(e); accruals = open_accruals_total(e)
        pending = balances.pending_receipts_total(e)
        nbv = Decimal(0)
        try:
            from assets.models import nbv_total
            nbv = nbv_total(e)
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx; _lx("monthly treasurer sofp")
        # full statement of financial position, matching the main report: trust
        # payable split into receipted vs not-yet-receipted, prepayments/advances,
        # and net assets classified into unallocated / allocated / property.
        _tsum = balances.trust_summary(None, e)
        trust_receipted = sum((r["to_remit"] for r in _tsum), Decimal(0))
        trust_unreceipted = trust_payable - trust_receipted
        try:
            from cashbook.views import (unexpired_prepayments_total,
                                        outstanding_advances_total)
            prepaid = unexpired_prepayments_total(e)
            advances = outstanding_advances_total(e)
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx; _lx("monthly treasurer sofp2")
            prepaid = advances = Decimal(0)
        allocated = sum((r["closing"] for r in rows if not r["is_trust"]
                         and r["department"].category == "DEVELOPMENT"), Decimal(0))
        unallocated = local_funds_total - allocated
        accrual_adj = prepaid - payables - accruals
        ctx["sofp"] = {
            "cash": cash, "cash_on_hand": cash - advances, "advances": advances,
            "nbv": nbv, "prepaid": prepaid, "pending": pending,
            "trust_payable": trust_payable, "trust_receipted": trust_receipted,
            "trust_unreceipted": trust_unreceipted,
            "payables": payables, "accruals": accruals,
            "total_assets": (cash - advances) + advances + pending + nbv + prepaid,
            "total_liabilities": trust_payable + payables + accruals + pending,
            "local_funds": local_funds_total,
            "unallocated": unallocated, "allocated": allocated,
            "accrual_adj": accrual_adj,
            "net_assets": unallocated + allocated + nbv + accrual_adj}

        # Statement of changes in net assets: opening + surplus/(deficit) = closing.
        try:
            s_prev = s - _dt.timedelta(days=1)
            _closing_na = unallocated + allocated + nbv + accrual_adj
            prev_rows = balances.department_summary(None, s_prev)
            local_prev = [r for r in prev_rows if not r["is_trust"]]
            lf_prev = sum((r["closing"] for r in local_prev), Decimal(0))
            alloc_prev = sum((r["closing"] for r in local_prev
                              if r["department"].category == "DEVELOPMENT"), Decimal(0))
            nbv_prev = nbv
            try:
                from assets.models import nbv_total as _nbvt
                nbv_prev = _nbvt(s_prev)
            except Exception:  # noqa: BLE001
                pass
            accr_prev = (unexpired_prepayments_total(s_prev)
                         - open_payables_total(s_prev) - open_accruals_total(s_prev))
            opening_na = lf_prev + nbv_prev + accr_prev
            ctx["net_asset_changes"] = {
                "opening": opening_na,
                "surplus": _closing_na - opening_na,
                "closing": _closing_na}
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx3; _lx3("MT net-asset changes")
            ctx["net_asset_changes"] = None

        # 9 & 10) cash-flow statements (operating / investing / financing) for the month
        local_receipts = sum((r["receipts"] for r in rows if not r["is_trust"]), Decimal(0))
        trust_receipts = sum((r["receipts"] for r in rows if r["is_trust"]), Decimal(0))
        remittances = (Expense.objects.filter(date__gte=s, date__lte=e, status__in=paid,
                       category=Expense.Category.REMITTANCE)
                       .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        net_operating = local_receipts + trust_receipts - op_exp - remittances
        cash_open = sum((r["opening"] for r in rows), Decimal(0))
        ctx["cashflow"] = {
            "local_receipts": local_receipts, "trust_receipts": trust_receipts,
            "operating_exp": op_exp, "remittances": remittances,
            "net_operating": net_operating, "capital": capital,
            "net_investing": -capital, "net_change": net_operating - capital,
            "cash_open": cash_open, "cash_close": cash_open + net_operating - capital}

        # 12) most recent reconciliation
        ctx["recon"] = T.recent_reconciliation(e)

        # 13) Camp Meeting goal records (expense + offering, never group goals)
        ctx["camp_goals"] = _camp_goal_records(as_of.year)

        # short notes per section (AI narrative if available, else concise text)
        ctx["notes"] = self._notes(ctx)
        ctx["ai_summary"] = self._ai_summary(ctx)
        ctx["insights"] = self._section_insights(ctx)
        # Board-ready executive summary: budget tracking, highlights, items
        # needing attention, and decisions the Board is asked to make.
        try:
            ctx["budget_summary"] = T_budget.budget_vs_actual(
                as_of.year, period="MONTH", month=as_of.month)
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx4; _lx4("MT budget summary")
            ctx["budget_summary"] = None
        self._board_focus(ctx)
        return ctx

    def _section_insights(self, ctx):
        """A line or two of analysis per section — trend direction and a takeaway.
        Rule-based so it always works; enriched by the LLM when it is enabled."""
        from decimal import Decimal
        ins = {}

        def pct_change(now, prev):
            now = Decimal(str(now or 0)); prev = Decimal(str(prev or 0))
            if prev == 0:
                return None
            return float((now - prev) / prev * 100)

        def phrase(delta):
            if delta is None:
                return "no comparable prior period"
            if delta > 1:
                return f"up {delta:.0f}% on the prior period"
            if delta < -1:
                return f"down {abs(delta):.0f}% on the prior period"
            return "broadly flat versus the prior period"

        c = ctx.get("collections") or {}
        yearly = ctx.get("yearly") or []
        # collections: compare YTD to last year's YTD
        try:
            if len(yearly) >= 2:
                d = pct_change(yearly[-1]["collection"], yearly[-2]["collection"])
                ins["collections"] = (
                    f"Year-to-date collections are {phrase(d)}. Trust makes up "
                    f"{(c.get('trust',0)/c['total']*100):.0f}% of this month's receipts."
                    if c.get("total") else f"Collections are {phrase(d)}.")
        except Exception:  # noqa: BLE001 — an optional narrative must never break the report
            from core.utils import log_exception as _lx; _lx("monthly treasurer insight: collections")
        # trust trend
        tt = ctx.get("trust_trend") or {}
        try:
            rows = tt.get("rows") or []
            if rows:
                cols = tt.get("cells_key") or []
                last_two = [sum(r["cells"][-1] for r in rows if r["cells"]),
                            sum(r["cells"][-2] for r in rows if len(r.get("cells", [])) > 1)]
                d = pct_change(last_two[0], last_two[1])
                ins["trust_trend"] = (f"Receipted trust giving is {phrase(d)} month-on-month. "
                                      "Only receipted trust is a firm remittance liability.")
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx; _lx("monthly treasurer insight: trust_trend")
        # income & expenditure
        isr = ctx.get("income_stmt") or {}
        try:
            surplus = isr.get("surplus", 0) or 0
            if surplus >= 0:
                ins["income"] = (f"Operations ran a surplus of {float(surplus):,.0f} this "
                                 "month — income covered operating costs.")
            else:
                ins["income"] = (f"Operations ran a deficit of {abs(float(surplus)):,.0f} — "
                                 "spending outpaced local income this month; watch reserves.")
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx; _lx("monthly treasurer insight: income")
        # cash-flow
        cf = ctx.get("cashflow") or {}
        try:
            nc = cf.get("net_change", 0) or 0
            direction = "rose" if nc > 0 else ("fell" if nc < 0 else "held steady")
            ins["cashflow"] = (f"Cash {direction} by {abs(float(nc)):,.0f} over the month, "
                               f"ending at {float(cf.get('cash_close', 0)):,.0f}.")
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx; _lx("monthly treasurer insight: cashflow")
        # SoFP / net assets
        sofp = ctx.get("sofp") or {}
        try:
            na = float(sofp.get("net_assets", 0) or 0)
            tp = float(sofp.get("trust_payable", sofp.get("trust_liab", 0)) or 0)
            ins["sofp"] = (f"Net assets stand at {na:,.0f}. Trust funds of {tp:,.0f} are a "
                           "liability owed to the field, not the church's own reserves.")
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx; _lx("monthly treasurer insight: sofp")
        # camp goals
        try:
            goals = ctx.get("camp_goals") or []
            if goals:
                best = max(goals, key=lambda g: g["pct"])
                ins["camp"] = (f"{best['name']} is {best['pct']}% funded. "
                               + ("On track." if best["pct"] >= 60 else
                                  "Momentum needed to reach the target."))
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx; _lx("monthly treasurer insight: camp")

        # optional single LLM enrichment pass (kept cheap; falls back silently)
        try:
            from core.services.assistant import _llm_call
            cfg = SiteConfig.get()
            if getattr(cfg, "llm_enabled", False):
                import json as _json
                facts = {
                    "month": f"{ctx['end']:%B %Y}",
                    "collections_total": float(c.get("total", 0) or 0),
                    "surplus": float(isr.get("surplus", 0) or 0),
                    "net_assets": float(sofp.get("net_assets", 0) or 0),
                    "cash_change": float(cf.get("net_change", 0) or 0),
                }
                prompt = (
                    "You are a church treasurer analyst. Given these monthly figures, "
                    "return ONLY a JSON object with keys collections, income, cashflow, "
                    "sofp — each a single insightful sentence (max 22 words) about the "
                    "trend or what to watch. No preamble.\n" + _json.dumps(facts))
                txt, err = _llm_call(prompt, cfg, context="(section insights)")
                if txt and not err:
                    obj = _json.loads(txt[txt.find("{"):txt.rfind("}") + 1])
                    for k, v in obj.items():
                        if v:
                            ins[k] = str(v)
        except Exception:
            pass
        return ins

    def _board_focus(self, ctx):
        """Populate ctx['highlights'], ctx['attention'] and ctx['decisions'] — the
        board-oriented executive summary. Rule-based and defensive: any single
        computation failing never breaks the report, it's just omitted."""
        from decimal import Decimal
        highlights, attention, decisions = [], [], []
        ins = ctx.get("insights") or {}
        c = ctx.get("collections") or {}
        sofp = ctx.get("sofp") or {}
        cf = ctx.get("cashflow") or {}

        # --- Highlights (4-6 short, factual lines) ---
        try:
            trust_local = (ctx.get("collection_detail") or {})
            top_fund = None
            pool = (trust_local.get("trust") or []) + (trust_local.get("local") or [])
            if pool:
                top_fund = max(pool, key=lambda r: r["receipts"])
            if top_fund:
                highlights.append(
                    f"Largest receiving fund this month: {_sfund(top_fund['department'].name)} "
                    f"({float(top_fund['receipts']):,.0f}).")
        except Exception:  # noqa: BLE001
            pass
        for key in ("collections", "trust_trend", "income", "cashflow", "camp", "sofp"):
            if ins.get(key) and len(highlights) < 6:
                highlights.append(ins[key])
        ctx["highlights"] = highlights[:6]

        # --- Items requiring Board attention ---
        recon = ctx.get("recon")
        if not recon:
            attention.append({"severity": "medium", "title": "No bank reconciliation on file",
                              "detail": "No reconciliation has been recorded for this period. "
                                       "The bank and cash-book balances have not been checked "
                                       "against each other this month."})
        elif not recon.is_reconciled:
            diff = recon.difference
            attention.append({"severity": "high", "title": "Bank reconciliation not balanced",
                              "detail": f"The latest reconciliation ({recon.statement_date:%d %b %Y}) "
                                       f"leaves a difference of {float(diff or 0):,.0f} between "
                                       "the bank statement and the cash book."})

        try:
            neg = [r for r in (ctx.get("local_statement") or {}).get("rows", [])
                   if (r.get("closing") or 0) < 0]
            if neg:
                names = ", ".join(f"{_sfund(r['department'].name)} ({float(r['closing']):,.0f})"
                                  for r in neg[:5])
                attention.append({"severity": "high", "title": "Negative fund balance(s)",
                                  "detail": f"{len(neg)} local fund(s) are overdrawn: {names}"
                                           + ("…" if len(neg) > 5 else "")})
        except Exception:  # noqa: BLE001
            pass

        try:
            trust_unrec = sofp.get("trust_unreceipted") or Decimal(0)
            if trust_unrec > 0:
                attention.append({"severity": "medium",
                                  "title": "Trust funds collected but not yet receipted",
                                  "detail": f"{float(trust_unrec):,.0f} in trust-fund bank "
                                           "credits have not yet been formally receipted, so "
                                           "they aren't yet reflected as a firm remittance "
                                           "liability."})
        except Exception:  # noqa: BLE001
            pass

        try:
            bs = ctx.get("budget_summary")
            over_rows = [r for r in (bs or {}).get("rows", []) if r["over"]] if bs else []
            if over_rows:
                names = ", ".join(f"{_sfund(r['department'].name)} (over by "
                                  f"{abs(float(r['variance'])):,.0f})" for r in over_rows[:5])
                attention.append({"severity": "medium", "title": "Budget overrun this month",
                                  "detail": f"{len(over_rows)} fund(s) exceeded their prorated "
                                           f"monthly budget: {names}"
                                           + ("…" if len(over_rows) > 5 else "")})
        except Exception:  # noqa: BLE001
            pass
        ctx["attention"] = attention

        # --- Board decisions required ---
        m = ctx["end"].strftime("%B %Y")
        decisions.append({"title": "Approve the financial statements",
                          "detail": f"Adopt the Monthly Treasurer's Report for {m} as "
                                   "presented, including the Statement of Financial "
                                   "Position and Income & Expenditure Statement."})
        try:
            outstanding = sofp.get("trust_payable") or Decimal(0)
            field = SiteConfig.get().field_name or "the field"
            if outstanding > 0:
                decisions.append({"title": "Approve trust-fund remittance",
                                  "detail": f"Authorise remittance of {float(outstanding):,.0f} "
                                           f"in outstanding trust funds to {field}."})
        except Exception:  # noqa: BLE001
            pass
        for item in attention:
            if item["severity"] == "high":
                decisions.append({"title": f"Resolve: {item['title']}",
                                  "detail": item["detail"]})
        ctx["decisions"] = decisions

    def _notes(self, ctx):
        """Concise, accurate one-liners describing each section."""
        f = ctx["end"].strftime("%B %Y")
        return {
            "collections": f"Everything received in {f}, split between trust funds "
                           "(remitted to the field) and local funds (kept by the church).",
            "trust_trend": "Receipted trust collections this month and the previous "
                           "two — the trend in what is owed onward to the field.",
            "lcb_trend": "Every Local Church Budget account this month and the previous "
                         "two, so you can see which areas are growing or slowing.",
            "yearly": "Year-to-date totals for the same point in each of the last five "
                      "years, for a like-for-like long-term comparison.",
            "lcb_expenditure": f"How the Local Church Budget was spent in {f}, by category.",
            "local_statement": "Each local fund's opening balance, receipts, expenses "
                               "and closing balance for the month.",
            "sofp": "What the church owns and owes at month-end; trust funds are a "
                    "liability owed to the field, not the church's own money.",
            "cashflow": "How cash actually moved this month — from operations, into "
                        "property (investing), and the net change in cash held.",
            "recon": "The latest check that the cash book agrees with the bank statement.",
        }

    def _ai_summary(self, ctx):
        """An AI-written headline for the month, with a rule-based fallback."""
        try:
            from core.services.assistant import _llm_call
            cfg = SiteConfig.get()
            if not getattr(cfg, "llm_enabled", False):
                raise RuntimeError("assistant off")
            c = ctx["collections"]; isr = ctx["income_stmt"]
            prompt = (
                "Write 2 short sentences (max 45 words) summarising a church's monthly "
                "treasury figures for a board. Be factual and encouraging, no preamble.\n"
                f"Month: {ctx['end']:%B %Y}. Total collections: {c['total']:,.0f}. "
                f"Trust: {c['trust']:,.0f}. Local: {c['local']:,.0f}. "
                f"Operating surplus/(deficit): {isr['surplus']:,.0f}.")
            txt, err = _llm_call(prompt, cfg, context="(monthly treasurer summary)")
            if txt and not err:
                return txt.strip()
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx; _lx("monthly treasurer ai")
        c = ctx["collections"]; isr = ctx["income_stmt"]
        verdict = "a surplus" if isr["surplus"] >= 0 else "a deficit"
        return (f"In {ctx['end']:%B %Y} the church received {c['total']:,.0f} "
                f"({c['trust']:,.0f} trust, {c['local']:,.0f} local) and recorded "
                f"{verdict} of {abs(isr['surplus']):,.0f} on operations.")

    def _ai_narratives(self, ctx):
        """One short analysis paragraph per section, for the Word export (and
        anywhere else that wants more than the one-line `insights`). Runs
        server-side for three reasons: (1) the LLM key and prompt logic never
        reach the browser, (2) one server call can cover every section at once
        instead of the client firing several, and (3) the same figures Django
        already computed are passed straight into the prompt with no risk of
        a client-side mismatch. Every section always has rule-based text —
        `ctx['notes']` / `ctx['insights']`, already computed above — so the
        report never depends on the LLM being configured; when it is, this
        rewrites each into a fuller paragraph in one batched call."""
        notes = ctx.get("notes") or {}
        insights = ctx.get("insights") or {}
        # rule-based baseline: always available, always correct
        base = {k: " ".join(filter(None, [notes.get(k), insights.get(k)]))
                for k in notes}
        try:
            from core.services.assistant import _llm_call
            cfg = SiteConfig.get()
            if not getattr(cfg, "llm_enabled", False):
                raise RuntimeError("assistant off")
            c = ctx["collections"]; isr = ctx["income_stmt"]; sof = ctx["sofp"]
            facts = (
                f"Month {ctx['end']:%B %Y}. Collections {c['total']:,.0f} "
                f"(trust {c['trust']:,.0f}, local {c['local']:,.0f}). "
                f"Surplus/(deficit) {isr['surplus']:,.0f}. Cash & bank "
                f"{sof.get('cash', 0):,.0f}. Net assets {sof.get('net_assets', 0):,.0f}. "
                f"Trust outstanding {sof.get('trust_payable', 0):,.0f}.")
            sections = ", ".join(base.keys())
            prompt = (
                "You are writing analysis paragraphs for a church board's monthly "
                "treasury report. For EACH of these sections, write exactly one "
                f"factual, board-appropriate paragraph (25-45 words): {sections}. "
                f"Church figures this month: {facts}\n"
                "Respond ONLY as JSON: {\"section_key\": \"paragraph\", ...} using "
                "the exact section keys given, no other text.")
            txt, err = _llm_call(prompt, cfg, context="(monthly treasurer narratives)")
            if txt and not err:
                import json, re
                cleaned = re.sub(r"^```(json)?|```$", "", txt.strip(), flags=re.MULTILINE).strip()
                data = json.loads(cleaned)
                if isinstance(data, dict):
                    return {k: (data.get(k) or base.get(k, "")) for k in base}
        except Exception:  # noqa: BLE001 — narration is a nice-to-have, never blocking
            from core.utils import log_exception as _lx; _lx("monthly treasurer narratives")
        return base

def _monthly_report_context(request):
    """Reuse the full MonthlyTreasurerReportView context for exports."""
    view = MonthlyTreasurerReportView()
    view.request = request
    view.kwargs = {}
    view.args = ()
    return view.get_context_data()

class MonthlyReportExcelView(ReportAccessMixin, View):
    """Download the Monthly Treasurer's Report as a multi-sheet Excel workbook —
    full detail tables (not the on-screen top-10), a KPI summary sheet styled as
    cards, and native Excel charts for the figures that are charted on screen."""
    def get(self, request):
        import io
        from decimal import Decimal
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.chart import BarChart, PieChart, LineChart, Reference
        from django.http import HttpResponse

        ctx = _monthly_report_context(request)
        wb = Workbook()
        forest = "1F5F4F"; brass = "B07D2C"; red = "B3261E"
        head_fill = PatternFill("solid", fgColor=forest)
        card_fill = PatternFill("solid", fgColor="EAF1EE")
        warn_fill = PatternFill("solid", fgColor="FBEDEA")
        white = Font(color="FFFFFF", bold=True, size=12)
        bold = Font(bold=True)
        big = Font(bold=True, size=14, color=forest)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(bottom=thin)
        money = '#,##0.00'

        def sheet(ws, title, subtitle=None):
            ws.column_dimensions["A"].width = 40
            for col in "BCDEF":
                ws.column_dimensions[col].width = 16
            ws["A1"] = ctx.get("church") or "Church Treasury"
            ws["A1"].font = Font(bold=True, size=14, color=forest)
            ws["A2"] = title
            ws["A2"].font = Font(bold=True, size=12)
            ws["A3"] = subtitle or f"Period ending {ctx['end']:%d %B %Y}"
            ws["A3"].font = Font(italic=True, color="666666")
            return 5

        def hrow(ws, r, cells):
            for i, c in enumerate(cells):
                cell = ws.cell(row=r, column=1 + i, value=c)
                cell.fill = head_fill; cell.font = white
                cell.alignment = Alignment(horizontal="left" if i == 0 else "right")
            return r + 1

        def drow(ws, r, label, *vals, bold_row=False, flag=False):
            cell = ws.cell(row=r, column=1, value=label)
            if bold_row:
                cell.font = bold
            if flag:
                cell.font = Font(bold=True, color=red)
            for i, v in enumerate(vals):
                vc = ws.cell(row=r, column=2 + i, value=float(v) if isinstance(v, Decimal) else v)
                vc.number_format = money if isinstance(v, Decimal) else "General"
                vc.alignment = Alignment(horizontal="right")
                if bold_row:
                    vc.font = bold
                if flag:
                    vc.font = Font(bold=True, color=red)
            return r + 1

        # ---------------- Executive Summary (KPI "cards" + highlights + attention) ----------------
        ws0 = wb.active; ws0.title = "Executive Summary"
        r = sheet(ws0, "Executive Summary", f"For the month of {ctx['end']:%B %Y}")
        isr = ctx["income_stmt"]; sof = ctx["sofp"]; c = ctx["collections"]
        kpis = [
            ("Total collections", c.get("total", 0)), ("Local fund receipts", c.get("local", 0)),
            ("Trust fund receipts", c.get("trust", 0)), ("Total expenses", isr.get("expense", 0)),
            ("Monthly surplus / (deficit)", isr.get("surplus", 0)),
            ("Cash & bank balance", sof.get("cash", 0)), ("Net assets", sof.get("net_assets", 0)),
            ("Trust funds outstanding", sof.get("trust_payable", 0)),
        ]
        card_row = r
        for i, (label, val) in enumerate(kpis):
            col = 1 + (i % 2) * 2
            row = card_row + (i // 2) * 3
            lc = ws0.cell(row=row, column=col, value=label)
            lc.font = Font(size=9, color="666677"); lc.fill = card_fill
            vc = ws0.cell(row=row + 1, column=col, value=float(val))
            vc.font = big; vc.number_format = money; vc.fill = card_fill
            ws0.cell(row=row, column=col + 1).fill = card_fill
            ws0.cell(row=row + 1, column=col + 1).fill = card_fill
        r = card_row + ((len(kpis) + 1) // 2) * 3 + 1

        if ctx.get("highlights"):
            ws0.cell(row=r, column=1, value="Key highlights").font = Font(bold=True, size=12, color=forest)
            r += 1
            for h in ctx["highlights"]:
                ws0.cell(row=r, column=1, value=f"\u2022 {h}")
                r += 1
            r += 1
        if ctx.get("attention"):
            ws0.cell(row=r, column=1, value="Items requiring Board attention").font = Font(bold=True, size=12, color=red)
            r += 1
            for a in ctx["attention"]:
                cell = ws0.cell(row=r, column=1, value=f"{a['title']} — {a['detail']}")
                cell.fill = warn_fill; cell.font = Font(color=red)
                r += 1

        # ---------------- Collections (FULL listing, both trust and local) ----------------
        ws1 = wb.create_sheet("Collections")
        r = sheet(ws1, "Collections Summary — full listing")
        r = hrow(ws1, r, ["Fund", "Type", "Amount", "% of total"])
        tot = c.get("total") or Decimal(1)
        cd = ctx.get("collection_detail") or {}
        first_data_row = r
        for kind, rows in (("Trust", cd.get("trust", [])), ("Local", cd.get("local", []))):
            for row in rows:
                pct = float(row["receipts"] / tot * 100) if tot else 0
                r = drow(ws1, r, row["department"].name, kind, row["receipts"], f"{pct:.1f}%")
        last_data_row = r - 1
        r = drow(ws1, r, "Total collections", "", c.get("total", 0), "100%", bold_row=True)
        if last_data_row >= first_data_row:
            pie = PieChart(); pie.title = "Collections by fund"
            data = Reference(ws1, min_col=3, min_row=first_data_row - 1, max_row=min(last_data_row, first_data_row + 14))
            cats = Reference(ws1, min_col=1, min_row=first_data_row, max_row=min(last_data_row, first_data_row + 14))
            pie.add_data(data, titles_from_data=True); pie.set_categories(cats)
            pie.height = 9; pie.width = 14
            ws1.add_chart(pie, f"F{first_data_row}")

        # ---------------- Trust Fund Performance ----------------
        ws2 = wb.create_sheet("Trust Fund Performance")
        r = sheet(ws2, "Trust Fund Performance — 3-month trend")
        tt = ctx.get("trust_trend") or {}
        cols = tt.get("columns", [])
        r = hrow(ws2, r, ["Trust fund"] + [col["label"] for col in cols])
        first_data_row = r
        for row in tt.get("rows", []):
            r = drow(ws2, r, row["dept"].name, *row["cells"])
        last_data_row = r - 1
        if tt.get("rows"):
            r = drow(ws2, r, "Total", *tt.get("col_totals", []), bold_row=True)
            line = LineChart(); line.title = "Receipted trust — 3-month trend"
            data = Reference(ws2, min_col=2, max_col=1 + len(cols), min_row=first_data_row - 1, max_row=last_data_row)
            cats = Reference(ws2, min_col=1, min_row=first_data_row, max_row=last_data_row)
            line.add_data(data, titles_from_data=True); line.set_categories(cats)
            line.height = 8; line.width = 16
            ws2.add_chart(line, f"A{r + 2}")

        # ---------------- Local Fund Performance (FULL listing) ----------------
        ws3 = wb.create_sheet("Local Fund Performance")
        r = sheet(ws3, "Local Fund Performance — full listing")
        r = hrow(ws3, r, ["Fund", "Opening", "Receipts", "Expenses", "Closing"])
        first_data_row = r
        for row in ctx.get("local_statement", {}).get("rows", []):
            nm = getattr(row.get("department"), "name", "") or row.get("name", "")
            r = drow(ws3, r, nm, row.get("opening", Decimal(0)), row.get("receipts", Decimal(0)),
                     row.get("expenses", Decimal(0)), row.get("closing", Decimal(0)),
                     flag=(row.get("closing", 0) or 0) < 0)
        last_data_row = r - 1
        totloc = ctx.get("local_statement", {}).get("totals", {})
        if totloc:
            r = drow(ws3, r, "Total", totloc.get("opening", Decimal(0)), totloc.get("receipts", Decimal(0)),
                     totloc.get("expenses", Decimal(0)), totloc.get("closing", Decimal(0)), bold_row=True)
        if last_data_row >= first_data_row:
            bar = BarChart(); bar.title = "Local fund closing balances"; bar.type = "col"
            data = Reference(ws3, min_col=5, min_row=first_data_row - 1, max_row=min(last_data_row, first_data_row + 19))
            cats = Reference(ws3, min_col=1, min_row=first_data_row, max_row=min(last_data_row, first_data_row + 19))
            bar.add_data(data, titles_from_data=True); bar.set_categories(cats)
            bar.height = 9; bar.width = 18
            ws3.add_chart(bar, f"A{r + 2}")

        # ---------------- Expenditure Summary ----------------
        ws4 = wb.create_sheet("Expenditure")
        r = sheet(ws4, "Expenditure Summary — by category")
        r = hrow(ws4, r, ["Category", "Amount"])
        first_data_row = r
        for row in ctx.get("lcb_expenditure", {}).get("rows", []):
            r = drow(ws4, r, row["label"], row["total"])
        last_data_row = r - 1
        if ctx.get("lcb_expenditure", {}).get("rows"):
            r = drow(ws4, r, "Total LCB expenditure", ctx["lcb_expenditure"]["total"], bold_row=True)
            pie2 = PieChart(); pie2.title = "Expenditure by category"
            data = Reference(ws4, min_col=2, min_row=first_data_row - 1, max_row=last_data_row)
            cats = Reference(ws4, min_col=1, min_row=first_data_row, max_row=last_data_row)
            pie2.add_data(data, titles_from_data=True); pie2.set_categories(cats)
            pie2.height = 9; pie2.width = 14
            ws4.add_chart(pie2, f"D{first_data_row}")

        # ---------------- Budget & Goal Tracking ----------------
        ws5 = wb.create_sheet("Budget & Goals")
        r = sheet(ws5, "Budget & Goal Tracking")
        bs = ctx.get("budget_summary")
        if bs and bs.get("rows"):
            r = hrow(ws5, r, ["Fund", "Budget", "Actual", "Variance", "Variance %"])
            for row in bs["rows"]:
                pct = row.get("variance_pct")
                r = drow(ws5, r, row["department"].name, row["budget"], row["actual"],
                         row["variance"], f"{float(pct):.1f}%" if pct is not None else "",
                         flag=row.get("over"))
            t = bs["totals"]
            r = drow(ws5, r, "Total", t["budget"], t["actual"], t["variance"],
                     f"{float(t['variance_pct']):.1f}%" if t.get("variance_pct") is not None else "",
                     bold_row=True)
            r += 2
        if ctx.get("camp_goals"):
            r = hrow(ws5, r, ["Goal", "Target", "Collected", "Variance", "% Complete"])
            for g in ctx["camp_goals"]:
                r = drow(ws5, r, g["name"], g["goal"], g["collected"], g["variance"], f"{g['pct']}%")

        # ---------------- Statement of Financial Position ----------------
        ws6 = wb.create_sheet("Financial Position")
        r = sheet(ws6, "Statement of Financial Position")
        r = hrow(ws6, r, ["Item", "Amount"])
        r = drow(ws6, r, "Cash & bank", sof.get("cash_on_hand", Decimal(0)))
        r = drow(ws6, r, "Property (net book value)", sof.get("nbv", Decimal(0)))
        r = drow(ws6, r, "Total assets", sof.get("total_assets", Decimal(0)), bold_row=True)
        r = drow(ws6, r, "Trust payable — receipted", sof.get("trust_receipted", Decimal(0)))
        r = drow(ws6, r, "Trust payable — not yet receipted", sof.get("trust_unreceipted", Decimal(0)))
        r = drow(ws6, r, "Total liabilities", sof.get("total_liabilities", Decimal(0)), bold_row=True)
        r = drow(ws6, r, "General net assets", sof.get("unallocated", Decimal(0)))
        r = drow(ws6, r, "Designated development funds", sof.get("allocated", Decimal(0)))
        r = drow(ws6, r, "Invested in property", sof.get("nbv", Decimal(0)))
        fund_mix_row = r
        r = drow(ws6, r, "Total net assets", sof.get("net_assets", Decimal(0)), bold_row=True)
        pie3 = PieChart(); pie3.title = "Fund composition"
        ws6.cell(row=r + 2, column=1, value="General")
        ws6.cell(row=r + 2, column=2, value=float(sof.get("unallocated", 0) or 0))
        ws6.cell(row=r + 3, column=1, value="Designated dev")
        ws6.cell(row=r + 3, column=2, value=float(sof.get("allocated", 0) or 0))
        ws6.cell(row=r + 4, column=1, value="Property")
        ws6.cell(row=r + 4, column=2, value=float(sof.get("nbv", 0) or 0))
        ws6.cell(row=r + 5, column=1, value="Trust to remit")
        ws6.cell(row=r + 5, column=2, value=float(sof.get("trust_payable", 0) or 0))
        data = Reference(ws6, min_col=2, min_row=r + 2, max_row=r + 5)
        cats = Reference(ws6, min_col=1, min_row=r + 2, max_row=r + 5)
        pie3.add_data(data); pie3.set_categories(cats)
        pie3.height = 9; pie3.width = 14
        ws6.add_chart(pie3, f"D{fund_mix_row}")

        if ctx.get("net_asset_changes"):
            nac = ctx["net_asset_changes"]
            r += 8
            r = hrow(ws6, r, ["Changes in net assets", "Amount"])
            r = drow(ws6, r, "Net assets at start of period", nac["opening"])
            r = drow(ws6, r, "Surplus / (deficit) for the period", nac["surplus"])
            r = drow(ws6, r, "Net assets at end of period", nac["closing"], bold_row=True)

        # ---------------- Cash Flow ----------------
        ws7 = wb.create_sheet("Cash Flow")
        r = sheet(ws7, "Cash Flow Statement")
        cf = ctx.get("cashflow") or {}
        r = hrow(ws7, r, ["Item", "Amount"])
        r = drow(ws7, r, "Local receipts", cf.get("local_receipts", Decimal(0)))
        r = drow(ws7, r, "Trust receipts", cf.get("trust_receipts", Decimal(0)))
        r = drow(ws7, r, "Operating expenses paid", -(cf.get("operating_exp", Decimal(0))))
        r = drow(ws7, r, "Remittances to field", -(cf.get("remittances", Decimal(0))))
        r = drow(ws7, r, "Net operating cash", cf.get("net_operating", Decimal(0)), bold_row=True)
        r = drow(ws7, r, "Capital expenditure", -(cf.get("capital", Decimal(0))))
        r = drow(ws7, r, "Net change in cash", cf.get("net_change", Decimal(0)), bold_row=True)
        r = drow(ws7, r, "Cash at start of period", cf.get("cash_open", Decimal(0)))
        r = drow(ws7, r, "Cash at end of period", cf.get("cash_close", Decimal(0)))

        # ---------------- Bank Reconciliation ----------------
        ws8 = wb.create_sheet("Bank Reconciliation")
        r = sheet(ws8, "Bank Reconciliation")
        rec = ctx.get("recon")
        if rec:
            r = hrow(ws8, r, ["Item", "Amount"])
            r = drow(ws8, r, "Statement date", str(rec.statement_date))
            r = drow(ws8, r, "Balance per bank statement", rec.bank_balance)
            r = drow(ws8, r, "Adjusted bank balance", rec.adjusted_balance)
            r = drow(ws8, r, "Balance per cash book", rec.book_balance or Decimal(0))
            r = drow(ws8, r, "Difference", rec.difference or Decimal(0), bold_row=True,
                     flag=not rec.is_reconciled)
            r = drow(ws8, r, "Status", "Reconciled" if rec.is_reconciled else "Not yet balanced")
        else:
            ws8.cell(row=r, column=1, value="No bank reconciliation recorded yet for this period.")

        # ---------------- Board Decisions Required ----------------
        ws9 = wb.create_sheet("Board Decisions")
        r = sheet(ws9, "Board Decisions Required")
        for i, d in enumerate(ctx.get("decisions", []), 1):
            ws9.cell(row=r, column=1, value=f"{i}. {d['title']}").font = bold
            r += 1
            ws9.cell(row=r, column=1, value=d["detail"])
            r += 2

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f"treasurer_report_{ctx['end']:%Y_%m}.xlsx"
        resp = HttpResponse(buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp

class MonthlyReportWordView(ReportAccessMixin, View):
    """Download the Monthly Treasurer's Report as a Word document. Rendered as a
    Word-compatible HTML document (opens natively in Microsoft Word) so it needs
    no extra library on the server. Mirrors the on-screen report's structure —
    executive summary, then each management section with its own narrative —
    and adds a fuller per-section analysis paragraph (server-side, LLM-enriched
    with a rule-based fallback; see MonthlyTreasurerReportView._ai_narratives)
    plus a few chart images — Word can't run the on-screen report's JS charts,
    so these are rendered server-side with Pillow and embedded as base64."""
    def get(self, request):
        from django.http import HttpResponse
        from decimal import Decimal
        from reports.services.chart_image import bar_chart, donut_or_split
        view = MonthlyTreasurerReportView()
        view.request = request
        view.kwargs = {}
        view.args = ()
        ctx = view.get_context_data()
        ctx["narratives"] = view._ai_narratives(ctx)

        isr = ctx["income_stmt"]; c = ctx["collections"]
        try:
            ctx["chart_income_exp"] = bar_chart(
                "Income vs expenditure",
                [("Income", isr.get("income", 0), (31, 95, 79)),
                 ("Expenditure", isr.get("expense", 0), (179, 38, 30)),
                 ("Surplus / (deficit)", isr.get("surplus", 0), (176, 125, 44))])
        except Exception:  # noqa: BLE001 — a chart failing must never break the export
            ctx["chart_income_exp"] = None
        try:
            ctx["chart_collections_mix"] = donut_or_split(
                "Collections — local vs trust",
                [("Local", c.get("local", 0) or Decimal(0), (31, 95, 79)),
                 ("Trust", c.get("trust", 0) or Decimal(0), (176, 125, 44))])
        except Exception:  # noqa: BLE001
            ctx["chart_collections_mix"] = None
        try:
            camp = next((g for g in (ctx.get("camp_goals") or [])
                        if "Expense" in g["name"]), None)
            if camp and camp["goal"]:
                ctx["chart_camp_goal"] = bar_chart(
                    "Camp Meeting Expense Goal — progress",
                    [("Collected", camp["collected"], (31, 95, 79)),
                     ("Target", camp["goal"], (176, 125, 44))])
            else:
                ctx["chart_camp_goal"] = None
        except Exception:  # noqa: BLE001
            ctx["chart_camp_goal"] = None

        html = render(request, "reports/monthly_treasurer_word.html", ctx).content
        fname = f"treasurer_report_{ctx['end']:%Y_%m}.doc"
        resp = HttpResponse(html, content_type="application/msword")
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp
