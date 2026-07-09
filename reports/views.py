from decimal import Decimal

from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views import View
from django.views.generic import TemplateView

from core.permissions import (ReportAccessMixin, TreasurerRequiredMixin,
                              RightRequiredMixin, ReportAccessMixin)
from core.utils import parse_period, safe_json
from cashbook.models import Expense
from departments.models import Department
from giving.models import Transaction
from members.models import Member
from .services import balances
from .exports import csv_response


class PeriodMixin(ReportAccessMixin):
    def period(self):
        return parse_period(self.request)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["start"], ctx["end"] = self.period()
        ctx["filters"] = self.request.GET
        return ctx


class ReportIndexView(ReportAccessMixin, TemplateView):
    template_name = "reports/index.html"


class MonthlyReportView(PeriodMixin, TemplateView):
    template_name = "reports/monthly.html"

    def get(self, request, *args, **kwargs):
        start, end = self.period()
        rows = balances.department_summary(start, end)
        if request.GET.get("export") == "csv":
            data = [(r["department"].name,
                     "Trust" if r["is_trust"] else "Local",
                     r["opening"], r["receipts"], r["expenses"], r["closing"])
                    for r in rows]
            return csv_response(
                f"monthly_{start}_{end}.csv",
                ["Fund", "Type", "Opening", "Receipts", "Expenses", "Closing"], data)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        rows = balances.department_summary(ctx["start"], ctx["end"])
        ctx["rows"] = rows
        ctx["totals"] = balances.totals(rows)
        ctx["trust_total"] = sum(r["receipts"] for r in rows if r["is_trust"])
        ctx["local_total"] = sum(r["receipts"] for r in rows if not r["is_trust"])
        return ctx


class OfferingSummaryView(PeriodMixin, TemplateView):
    template_name = "reports/offering_summary.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        data = balances.offering_summary(ctx["start"], ctx["end"])
        ctx["sabbaths"] = data["sabbaths"]
        ctx["rows"] = data["rows"]
        return ctx


class TitheReportView(PeriodMixin, TemplateView):
    template_name = "reports/tithe.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        s, e = ctx["start"], ctx["end"]
        ctx["tithe"] = balances.tithe_total(s, e)
        ctx["count"] = Transaction.objects.filter(
            date__gte=s, date__lte=e, direction=Transaction.Direction.CREDIT,
            department__name__icontains="tithe").count()
        return ctx


class GroupGivingView(PeriodMixin, TemplateView):
    template_name = "reports/by_group.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        data = balances.giving_by_group(ctx["start"], ctx["end"])
        labels = dict(Member.Group.choices)
        ctx["rows"] = [{"group": labels.get(k, k), "total": v}
                       for k, v in sorted(data.items(), key=lambda x: -x[1])]
        ctx["grand_total"] = sum(data.values())
        return ctx


class DevGroupUnassignedView(RightRequiredMixin, TemplateView):
    """Development contributions sitting on the parent Development fund without a specific
    group — list them and reassign to the correct group for accurate per-group
    totals. Open to anyone with the development-offering allocation right."""
    required_right = "allocate_dev_offering"
    template_name = "reports/dev_unassigned.html"

    def _qs(self):
        from departments.models import Department
        return (Transaction.objects.active()
                .filter(direction=Transaction.Direction.CREDIT, confirmed=True,
                        department__category=Department.Category.DEVELOPMENT,
                        dev_group__isnull=True,
                        # loan financing on a Development fund is NOT a member
                        # development contribution — keep it out of the
                        # unassigned queue and every dev-group figure
                        excluded_from_income=False)
                .select_related("department", "member").order_by("-date"))

    def get_context_data(self, **kwargs):
        from departments.models import DevelopmentGroup
        ctx = super().get_context_data(**kwargs)
        qs = self._qs()
        ctx["items"] = qs[:500]
        ctx["count"] = qs.count()
        ctx["total"] = qs.aggregate(t=Sum("amount"))["t"] or Decimal(0)
        ctx["groups"] = DevelopmentGroup.objects.filter(active=True).order_by("number")
        return ctx

    def post(self, request):
        from departments.models import DevelopmentGroup
        gid = request.POST.get("dev_group")
        ids = request.POST.getlist("txn")
        if request.POST.get("all"):
            ids = list(self._qs().values_list("id", flat=True))
        grp = DevelopmentGroup.objects.filter(pk=gid).first()
        if not grp or not ids:
            messages.error(request, "Choose a group and at least one contribution to assign.")
            return redirect("dev_unassigned")
        n = (Transaction.objects.filter(id__in=ids, dev_group__isnull=True)
             .update(dev_group=grp))
        # keep the linked envelope lines in step (guarded: tolerate older schemas
        # where EnvelopeLine has no dev_group column)
        try:
            from envelopes.models import EnvelopeLine
            EnvelopeLine.objects.filter(transaction_id__in=ids,
                                        dev_group__isnull=True).update(dev_group=grp)
        except Exception:
            from core.utils import log_exception as _lx; _lx('reports/views.py')
            pass
        messages.success(request, f"Assigned {n} development contribution(s) to {grp.label}.")
        return redirect("dev_unassigned")


class DevGroupProgressView(PeriodMixin, TemplateView):
    template_name = "reports/dev_groups.html"

    def get(self, request, *args, **kwargs):
        export = request.GET.get("export")
        if export in ("csv", "xlsx"):
            from .exports import csv_response, xlsx_response
            from core.models import SiteConfig
            s, e = self.period()
            rows = balances.dev_group_progress(s, e)
            header = ["Group", "Collected", "Target", "Balance to target", "% complete"]
            out = [[(r["group"].label if hasattr(r["group"], "label") else r["group"].name),
                    float(r.get("collected") or 0), float(r.get("target") or 0),
                    float(r.get("balance") or 0), float(r.get("pct") or 0)] for r in rows]
            title = f"Development groups {s:%d %b %Y}–{e:%d %b %Y}"
            if export == "xlsx":
                return xlsx_response("dev_groups.xlsx", header, out, title=title,
                                     church=SiteConfig.get().church_name)
            return csv_response("dev_groups.csv", header, out)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["rows"] = balances.dev_group_progress(ctx["start"], ctx["end"])
        from departments.models import Department
        un = (Transaction.objects.active().filter(
                direction=Transaction.Direction.CREDIT, confirmed=True,
                department__category=Department.Category.DEVELOPMENT,
                dev_group__isnull=True, excluded_from_income=False))
        ctx["unassigned_count"] = un.count()
        ctx["unassigned_total"] = un.aggregate(t=Sum("amount"))["t"] or Decimal(0)
        return ctx


class ExpenseReportView(PeriodMixin, TemplateView):
    template_name = "reports/expenses.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        s, e = ctx["start"], ctx["end"]
        eff = Q(status__in=[Expense.Status.APPROVED, Expense.Status.PAID])
        base = (Expense.objects.filter(eff, date__gte=s, date__lte=e)
                .exclude(category__in=[Expense.Category.REMITTANCE, Expense.Category.LOAN_REPAYMENT]))
        # consolidate by the top-level fund (sub-account spend rolls into its parent)
        from collections import defaultdict
        from departments.models import Department
        from .services.budget import budget_amounts_bulk
        all_depts = list(Department.objects.select_related("parent"))
        parent_of = {}
        tops = {}
        for d in all_depts:
            top = d.parent or d
            parent_of[d.id] = top.name
            tops.setdefault(top.name, top)
        top_budgets = budget_amounts_bulk(e.year, tops.values())
        budget_of = {name: top_budgets.get(top.id) for name, top in tops.items()}
        agg = defaultdict(Decimal)
        for r in base.values("department_id").annotate(total=Sum("amount")):
            agg[parent_of.get(r["department_id"], "Unallocated")] += r["total"] or Decimal(0)
        ctx["by_dept"] = [{"name": k, "total": v, "budget": budget_of.get(k)}
                          for k, v in sorted(agg.items(), key=lambda x: -x[1])]
        ctx["by_category"] = (base.values("category")
                              .annotate(total=Sum("amount")).order_by("-total"))
        ctx["by_claimant"] = (base.exclude(claimant="").values("claimant")
                              .annotate(total=Sum("amount")).order_by("-total"))
        ctx["outstanding"] = Expense.objects.filter(
            status__in=[Expense.Status.PENDING, Expense.Status.APPROVED]).order_by("date")
        ctx["cat_labels"] = dict(Expense.Category.choices)
        return ctx


class IncomeExpenditureView(PeriodMixin, TemplateView):
    template_name = "reports/income_expenditure.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        s, e = ctx["start"], ctx["end"]
        # Income = the church's own (local) revenue only. Trust funds (tithe,
        # the remitted share of combined/thanksgiving offerings, etc.) are
        # collected on behalf of the field — a liability, not revenue — so they
        # are excluded here and shown separately as a memo.
        income = (Transaction.objects.confirmed_credits().filter(
            date__gte=s, date__lte=e, department__is_trust=False,
            excluded_from_income=False)
            .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        trust_collected = (Transaction.objects.confirmed_credits().filter(
            date__gte=s, date__lte=e, department__is_trust=True,
            excluded_from_income=False)
            .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        # Expenditure excludes remittances (they settle the trust liability) AND
        # capital purchases (an asset, not consumed in the period — only any
        # depreciation belongs in an income & expenditure account). Capital
        # additions are shown separately as a memo.
        paid = [Expense.Status.APPROVED, Expense.Status.PAID]
        expense = (Expense.objects.filter(
            date__gte=s, date__lte=e, status__in=paid)
            .exclude(category__in=[Expense.Category.REMITTANCE, Expense.Category.LOAN_REPAYMENT])
            .exclude(expenditure_type=Expense.ExpenditureType.CAPITAL)
            .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        capital = (Expense.objects.filter(
            date__gte=s, date__lte=e, status__in=paid,
            expenditure_type=Expense.ExpenditureType.CAPITAL)
            .exclude(category__in=[Expense.Category.REMITTANCE, Expense.Category.LOAN_REPAYMENT])
            .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        remittances = (Expense.objects.filter(
            date__gte=s, date__lte=e, status__in=paid,
            category=Expense.Category.REMITTANCE)
            .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        ctx["income"] = income
        ctx["expense"] = expense
        # Gain/(loss) on asset disposals in the period — the only part of a disposal
        # that belongs in the income result (the proceeds themselves are a capital
        # receipt, excluded from income above).
        from assets.models import FixedAsset
        disposal_gl = (FixedAsset.objects.filter(
            disposed=True, disposed_on__gte=s, disposed_on__lte=e)
            .aggregate(t=Sum("disposal_gain_loss"))["t"] or Decimal(0))
        ctx["disposal_gain_loss"] = disposal_gl
        ctx["net"] = income - expense + disposal_gl
        ctx["capital"] = capital
        ctx["trust_collected"] = trust_collected
        ctx["remittances"] = remittances
        return ctx


class FundLedgerView(PeriodMixin, TemplateView):
    template_name = "reports/fund_ledger.html"

    def get(self, request, *args, **kwargs):
        if request.GET.get("export") in ("xlsx", "csv", "subgroups", "subgroups-csv"):
            return self._export(request, *args, **kwargs)
        return super().get(request, *args, **kwargs)

    def _export(self, request, *args, **kwargs):
        from reports.exports import csv_response, xlsx_response
        from core.models import SiteConfig
        ctx = self.get_context_data(**kwargs)
        dept = ctx["department"]
        mode = request.GET.get("export")

        # subgroup breakdown export (sub-accounts beneath this fund)
        if mode in ("subgroups", "subgroups-csv"):
            show_pay = ctx.get("sub_show_payments", True)
            header = (["ID", "Subgroup", "Type", "Opening", "Receipts", "Payments", "Closing balance"]
                      if show_pay else
                      ["ID", "Subgroup", "Type", "Opening", "Receipts", "Closing balance"])
            rows = []
            for r in ctx["subgroups"]:
                sub = r["sub"]
                row = [sub.id, sub.name, "Trust" if sub.is_trust else "Local",
                       float(r["opening"]), float(r["receipts"])]
                if show_pay:
                    row.append(float(r["payments"]))
                row.append(float(r["closing"]))
                rows.append(row)
            for r in ctx.get("dev_rows", []):
                g = r["group"]
                row = [getattr(g, "id", ""), g.name, "Local", "", float(r["receipts"])]
                if show_pay:
                    row.append("")
                row.append(float(r["receipts"]))
                rows.append(row)
            total_row = ["", "TOTAL", "", float(ctx["combined_opening"]), ""]
            if show_pay:
                total_row.append("")
            total_row.append(float(ctx["subgroup_total"]))
            rows.append(total_row)
            fname = f"fund-{dept.slug or dept.id}-subgroups-{ctx['start']}-{ctx['end']}"
            if mode == "subgroups-csv":
                return csv_response(fname + ".csv", header, rows)
            return xlsx_response(fname + ".xlsx", header, rows,
                                 title=f"{dept.name} — sub-accounts ({ctx['start']} to {ctx['end']})",
                                 church=SiteConfig.get().church_name)

        header = ["ID", "Type", "Date", "Description", "Debit", "Credit", "Balance"]
        rows = [["", "", "", "Opening balance", "", "", float(ctx["opening"])]]
        for en in ctx["entries"]:
            rows.append([en.get("ref_id", ""), en.get("src", ""),
                         en["date"].isoformat(), en["desc"],
                         float(en["debit"]) if en["debit"] else "",
                         float(en["credit"]) if en["credit"] else "",
                         float(en["balance"])])
        rows.append(["", "", "", "Closing balance", "", "", float(ctx["closing"])])
        fname = f"fund-{dept.slug or dept.id}-{ctx['start']}-{ctx['end']}"
        if request.GET["export"] == "csv":
            return csv_response(fname + ".csv", header, rows)
        return xlsx_response(fname + ".xlsx", header, rows,
                             title=f"{dept.name} ledger ({ctx['start']} to {ctx['end']})",
                             church=SiteConfig.get().church_name)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        dept = get_object_or_404(Department, pk=kwargs["pk"])
        s, e = ctx["start"], ctx["end"]
        receipts = (Transaction.objects.confirmed_credits().filter(
            department=dept, date__gte=s, date__lte=e,
            excluded_from_income=False).order_by("date"))
        payments = Expense.objects.filter(
            department=dept, date__gte=s, date__lte=e,
            status__in=[Expense.Status.APPROVED, Expense.Status.PAID]).order_by("date")
        entries = []
        for t in receipts:
            entries.append({"date": t.date, "desc": t.payer_name or t.reference or "Receipt",
                            "credit": t.amount, "debit": None, "src": "Receipt", "ref_id": t.id})
        for x in payments:
            entries.append({"date": x.date, "desc": x.description,
                            "credit": None, "debit": x.amount, "src": "Expense", "ref_id": x.id})
        from cashbook.models import FundTransfer
        for tr in FundTransfer.objects.filter(destination=dept, date__gte=s, date__lte=e):
            entries.append({"date": tr.date, "desc": f"Transfer from {tr.source.name}"
                            + (f" — {tr.reason}" if tr.reason else ""),
                            "credit": tr.amount, "debit": None, "src": "Transfer", "ref_id": tr.id})
        for tr in FundTransfer.objects.filter(source=dept, date__gte=s, date__lte=e):
            entries.append({"date": tr.date, "desc": f"Transfer to {tr.destination.name}"
                            + (f" — {tr.reason}" if tr.reason else ""),
                            "credit": None, "debit": tr.amount, "src": "Transfer", "ref_id": tr.id})
        entries.sort(key=lambda r: r["date"])
        # opening = founding opening_balance + all net movement before `s` (not
        # just the raw founding field), so a fund with real prior-period activity
        # shows its true brought-forward balance rather than zero.
        from reports.services.balances import brought_forward, brought_forward_map
        opening_bf = brought_forward(dept, s)
        running = opening_bf
        for en in entries:
            running += (en["credit"] or 0) - (en["debit"] or 0)
            en["balance"] = running
        ctx["department"] = dept
        ctx["entries"] = entries
        ctx["opening"] = opening_bf
        ctx["closing"] = running

        # roll up any sub-accounts beneath this fund (two grouped queries, not 2/sub)
        subs = list(dept.subgroups.all())
        sub_ids = [x.id for x in subs]
        sub_rec = {r["department"]: (r["t"] or Decimal(0)) for r in
                   Transaction.objects.filter(
                       department_id__in=sub_ids,
                       direction=Transaction.Direction.CREDIT,
                       date__gte=s, date__lte=e)
                   .values("department").annotate(t=Sum("amount"))}
        sub_pay = {r["department"]: (r["t"] or Decimal(0)) for r in
                   Expense.objects.filter(
                       department_id__in=sub_ids, date__gte=s, date__lte=e,
                       status__in=[Expense.Status.APPROVED, Expense.Status.PAID])
                   .values("department").annotate(t=Sum("amount"))}
        subs_rows = []
        sub_total = Decimal(0)
        sub_bf = brought_forward_map(sub_ids, s) if sub_ids else {}
        for sub in subs:
            r = sub_rec.get(sub.id, Decimal(0))
            p = sub_pay.get(sub.id, Decimal(0))
            opening = sub_bf.get(sub.id, Decimal(0))
            closing = opening + r - p
            subs_rows.append({"sub": sub, "opening": opening, "receipts": r,
                              "payments": p, "closing": closing})
            sub_total += closing
        subs_rows.sort(key=lambda x: x["closing"], reverse=True)   # largest balance first
        subs = subs_rows

        # development groups are sub-accounts of the Development fund (one query)
        from departments.models import DevelopmentGroup
        dev_rows = []
        if dept.name.lower() == "development":
            dev_map = {r["dev_group"]: (r["t"] or Decimal(0)) for r in
                       Transaction.objects.filter(
                           dev_group__isnull=False,
                           direction=Transaction.Direction.CREDIT,
                           date__gte=s, date__lte=e)
                       .values("dev_group").annotate(t=Sum("amount"))}
            for grp in DevelopmentGroup.objects.filter(active=True):
                r = dev_map.get(grp.id, Decimal(0))
                dev_rows.append({"group": grp, "receipts": r})
                sub_total += r
            dev_rows.sort(key=lambda x: x["receipts"], reverse=True)
        ctx["dev_rows"] = dev_rows
        ctx["subgroups"] = subs
        ctx["subgroup_total"] = sub_total
        ctx["combined_closing"] = running + sub_total

        # combined (parent + sub-accounts) figures for the top cards, since the
        # sub-accounts are part of the parent fund.
        parent_receipts = sum((en["credit"] or Decimal(0)) for en in entries)
        parent_payments = sum((en["debit"] or Decimal(0)) for en in entries)
        sub_receipts_total = sum(sub_rec.values(), Decimal(0)) + sum(
            (r["receipts"] for r in dev_rows), Decimal(0))
        sub_payments_total = sum(sub_pay.values(), Decimal(0))
        sub_opening_total = sum((r["opening"] for r in subs_rows), Decimal(0))
        ctx["has_subaccounts"] = bool(subs_rows or dev_rows)
        ctx["combined_opening"] = opening_bf + sub_opening_total
        ctx["combined_receipts"] = parent_receipts + sub_receipts_total
        ctx["combined_payments"] = parent_payments + sub_payments_total
        ctx["parent"] = dept.parent
        # collection-only funds never take expenses/payments; hide that column in
        # the sub-accounts table when this fund and every sub-account shown are
        # collection-only, so the summary reads opening/receipts/closing only.
        _cols = [dept] + [r["sub"] for r in subs_rows]
        ctx["sub_show_payments"] = not all(
            getattr(d, "collection_only", False) for d in _cols) if _cols else True
        return ctx


class FundMembersView(PeriodMixin, TemplateView):
    """Aggregated giving for a fund and all its sub-accounts, grouped by member,
    so leaders/treasurers can see how much each person has given. Complements the
    chronological fund ledger (which this links to and from)."""
    template_name = "reports/fund_members.html"

    def get(self, request, *args, **kwargs):
        if request.GET.get("export") in ("xlsx", "csv"):
            return self._export(request, *args, **kwargs)
        return super().get(request, *args, **kwargs)

    def _fund_ids(self, dept):
        # the fund itself plus every descendant sub-account
        ids, frontier = {dept.id}, [dept]
        while frontier:
            nxt = []
            for d in frontier:
                for sub in d.subgroups.all():
                    if sub.id not in ids:
                        ids.add(sub.id); nxt.append(sub)
            frontier = nxt
        return ids

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        dept = get_object_or_404(Department, pk=kwargs["pk"])
        s, e = ctx["start"], ctx["end"]
        fund_ids = self._fund_ids(dept)
        qs = (Transaction.objects.confirmed_credits().filter(
            department_id__in=fund_ids, date__gte=s, date__lte=e,
            excluded_from_income=False))
        # group by member (named) and, separately, anonymous/loose giving
        rows = {}
        anon_total, anon_count = Decimal(0), 0
        for r in (qs.values("member", "member__name", "payer_name")
                    .annotate(total=Sum("amount"), n=Count("id"))):
            mid = r["member"]
            if mid:
                key, name = mid, r["member__name"]
            elif (r["payer_name"] or "").strip():
                key, name = f"p:{r['payer_name'].strip().lower()}", r["payer_name"].strip()
            else:
                anon_total += r["total"] or Decimal(0); anon_count += r["n"]; continue
            if key in rows:
                rows[key]["total"] += r["total"] or Decimal(0)
                rows[key]["n"] += r["n"]
            else:
                rows[key] = {"member_id": mid, "name": name,
                             "total": r["total"] or Decimal(0), "n": r["n"]}
        members = sorted(rows.values(), key=lambda x: x["total"], reverse=True)
        named_total = sum((m["total"] for m in members), Decimal(0))
        ctx.update({
            "department": dept, "members": members,
            "named_total": named_total, "anon_total": anon_total,
            "anon_count": anon_count, "grand_total": named_total + anon_total,
            "giver_count": len(members), "subaccount_count": len(fund_ids) - 1,
        })
        return ctx

    def _export(self, request, *args, **kwargs):
        from reports.exports import csv_response, xlsx_response
        from core.models import SiteConfig
        ctx = self.get_context_data(**kwargs)
        dept = ctx["department"]
        header = ["Member", "Gifts", "Total"]
        rows = [[m["name"], m["n"], float(m["total"])] for m in ctx["members"]]
        if ctx["anon_total"]:
            rows.append(["(unnamed / loose)", ctx["anon_count"], float(ctx["anon_total"])])
        rows.append(["TOTAL", "", float(ctx["grand_total"])])
        fname = f"fund-{dept.slug or dept.id}-by-member-{ctx['start']}-{ctx['end']}"
        if request.GET["export"] == "csv":
            return csv_response(fname + ".csv", header, rows)
        return xlsx_response(fname + ".xlsx", header, rows,
                             title=f"{dept.name} — giving by member ({ctx['start']} to {ctx['end']})",
                             church=SiteConfig.get().church_name)


def _balanced_partition(items, n, balance_size=True):
    """Partition members into n development groups balanced by giving capability.

    Two phases:

    1. **Greedy seed** - heaviest giver first into the currently lightest group,
       with a soft size cap (ceil(members / n)) so the groups also end up with an
       even *number* of members, not just even totals. (The previous version
       balanced totals only, which could leave one big giver alone in a group
       while everyone else clustered elsewhere.)

    2. **Local-search refinement** - repeatedly swap a member of the richest
       group with a lighter member of the poorest group whenever doing so shrinks
       the gap between them. Sizes are preserved (a swap is one-for-one), so this
       only redistributes capability to cut the variance between groups.

    Inherent limit: when giving is highly skewed - e.g. one member contributes
    ~90% of all development giving - *no* partition can equalise the totals; that
    member's group is unavoidably heavier. The algorithm still spreads everyone
    else as evenly as possible and keeps group sizes within one member of each
    other. `items` = [(member_id, name, phone, weight), ...].
    """
    items = sorted(items, key=lambda x: x[3], reverse=True)
    m = len(items)
    cap = (-(-m // n)) if (balance_size and n) else None  # ceil(m / n)
    buckets = [{"members": [], "total": Decimal(0)} for _ in range(n)]

    def lightest(respect_cap=True):
        cands = buckets
        if respect_cap and cap is not None:
            open_b = [b for b in buckets if len(b["members"]) < cap]
            if open_b:
                cands = open_b
        return min(cands, key=lambda x: (x["total"], len(x["members"])))

    for mid, name, phone, w in items:
        b = lightest()
        b["members"].append({"id": mid, "name": name, "phone": phone, "weight": w})
        b["total"] += w

    # Phase 2: variance-reducing swaps between the richest and poorest groups.
    for _ in range(2000):
        hi = max(buckets, key=lambda b: b["total"])
        lo = min(buckets, key=lambda b: b["total"])
        if hi is lo or hi["total"] == lo["total"]:
            break
        gap = hi["total"] - lo["total"]
        best = None  # (improvement, hi_member, lo_member)
        for hm in hi["members"]:
            for lm in lo["members"]:
                delta = hm["weight"] - lm["weight"]   # hm to lo, lm to hi
                if delta <= 0:
                    continue
                new_gap = abs(gap - 2 * delta)
                improve = gap - new_gap
                if improve > 0 and (best is None or improve > best[0]):
                    best = (improve, hm, lm)
        if not best:
            break
        _, hm, lm = best
        hi["members"].remove(hm); hi["total"] -= hm["weight"]
        lo["members"].remove(lm); lo["total"] -= lm["weight"]
        hi["members"].append(lm); hi["total"] += lm["weight"]
        lo["members"].append(hm); lo["total"] += hm["weight"]
    return buckets


class DevGroupBuilderView(RightRequiredMixin, TemplateView):
    """Propose N development groups balanced by members' historical development
    giving (capability), so each group carries a comparable giving capacity. By
    default this only produces a downloadable Excel/CSV proposal and changes no
    data; a treasurer can enable the live "apply" action in settings."""
    required_right = "build_dev_groups"
    template_name = "reports/dev_group_builder.html"

    def _member_totals(self):
        from departments.models import Department
        from members.models import Member
        rows = {r["member"]: r["t"] for r in (Transaction.objects.filter(
            member__isnull=False, direction=Transaction.Direction.CREDIT,
            confirmed=True, department__category=Department.Category.DEVELOPMENT)
            .values("member").annotate(t=Sum("amount")))}
        items = []
        for m in Member.objects.filter(active=True):
            items.append((m.id, m.name, m.phone or "", rows.get(m.id, Decimal(0))))
        return items

    def get(self, request, *args, **kwargs):
        if request.GET.get("export") in ("xlsx", "csv"):
            return self._export(request)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        from core.models import SiteConfig
        ctx = super().get_context_data(**kwargs)
        try:
            n = int(self.request.GET.get("n") or 0)
        except ValueError:
            n = 0
        items = self._member_totals()
        ctx["member_count"] = len(items)
        ctx["total_capability"] = sum((w for _, _, _, w in items), Decimal(0))
        ctx["n"] = n
        ctx["apply_enabled"] = SiteConfig.get().dev_group_builder_apply
        if 2 <= n <= 50 and items:
            buckets = _balanced_partition(items, n)
            for i, b in enumerate(buckets, 1):
                b["number"] = i
            ctx["preview"] = buckets
            totals = [b["total"] for b in buckets]
            ctx["spread"] = (max(totals) - min(totals)) if totals else Decimal(0)
        return ctx

    def _export(self, request):
        from reports.exports import csv_response, xlsx_response
        from core.models import SiteConfig
        try:
            n = int(request.GET.get("n") or 0)
        except ValueError:
            n = 0
        items = self._member_totals()
        header = ["Group", "Member", "Phone", "Development giving (capability)"]
        rows = []
        if 2 <= n <= 50 and items:
            buckets = _balanced_partition(items, n)
            for i, b in enumerate(buckets, 1):
                for m in sorted(b["members"], key=lambda x: x["weight"], reverse=True):
                    rows.append([f"Group {i}", m["name"], m["phone"], float(m["weight"])])
        else:
            # no group count given — just the member list with their capability
            header = ["Member", "Phone", "Development giving (capability)"]
            for mid, name, phone, w in sorted(items, key=lambda x: x[3], reverse=True):
                rows.append([name, phone, float(w)])
        fname = f"development-groups-{n}" if n else "development-members"
        if request.GET["export"] == "csv":
            return csv_response(fname + ".csv", header, rows)
        return xlsx_response(fname + ".xlsx", header, rows,
            title=(f"Proposed {n} balanced development groups" if n
                   else "Development giving by member"),
            church=SiteConfig.get().church_name)

    def post(self, request):
        from django.db import transaction
        from departments.models import DevelopmentGroup
        from members.models import Member
        from core.models import SiteConfig
        if not SiteConfig.get().dev_group_builder_apply:
            messages.error(request, "Creating groups is turned off — download the "
                "Excel proposal instead, or enable it in Settings → Channels.")
            return redirect(f"{request.path}?n={request.POST.get('n') or ''}")
        try:
            n = int(request.POST.get("n") or 0)
        except ValueError:
            n = 0
        if not (2 <= n <= 50):
            messages.error(request, "Choose between 2 and 50 groups.")
            return redirect(f"{request.path}?n={n}")
        prefix = (request.POST.get("prefix") or "Group").strip()[:40]
        items = self._member_totals()
        if not items:
            messages.error(request, "There are no active members to group.")
            return redirect(request.path)
        buckets = _balanced_partition(items, n)
        with transaction.atomic():
            # reuse groups 1..n, deactivate any beyond n
            DevelopmentGroup.objects.filter(number__gt=n).update(active=False)
            for i, b in enumerate(buckets, 1):
                grp, _ = DevelopmentGroup.objects.get_or_create(number=i)
                grp.name = f"{prefix} {i}"
                grp.active = True
                grp.save()
                ids = [m["id"] for m in b["members"]]
                Member.objects.filter(id__in=ids).update(dev_group=grp)
        messages.success(request, f"Built {n} balanced development groups from "
            f"{len(items)} members' giving history.")
        return redirect("report_dev_groups")


class TrustFundView(PeriodMixin, TemplateView):
    template_name = "reports/trust.html"

    def get_context_data(self, **kwargs):
        from cashbook.models import RemittanceBatch
        ctx = super().get_context_data(**kwargs)
        ctx["rows"] = balances.trust_summary(ctx["start"], ctx["end"])
        ctx["total_to_remit"] = sum(r["to_remit"] for r in ctx["rows"])
        ctx["total_unreceipted"] = sum((r["unreceipted"] for r in ctx["rows"]), Decimal(0))
        ctx["total_liability"] = sum((r["total_liability"] for r in ctx["rows"]), Decimal(0))
        ctx["batches"] = (RemittanceBatch.objects
                          .order_by("-date", "-id")[:25])
        ctx["remitted_total"] = sum(r["remitted"] for r in ctx["rows"])
        return ctx


class RemittanceView(PeriodMixin, TemplateView):
    template_name = "reports/remittance.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["rows"] = balances.trust_summary(ctx["start"], ctx["end"])
        ctx["total"] = sum(r["to_remit"] for r in ctx["rows"])
        ctx["total_unreceipted"] = sum((r["unreceipted"] for r in ctx["rows"]), Decimal(0))
        from statements.models import BankAccount
        ctx["bank_accounts"] = BankAccount.objects.all()
        return ctx


class RemitTrustView(TreasurerRequiredMixin, View):
    """Remit the amount still outstanding for each trust fund in the period as a
    single field payment. This now uses the same payment architecture as the
    batch workflow: it creates a remittance batch, raises the per-fund expenses
    against it, and settles the whole batch with one PaymentInstrument (cheque,
    EFT, RTGS, M-Pesa, etc.). The instrument is the settlement record; it posts
    no separate accounting entries."""

    def post(self, request):
        import datetime as _dt
        from cashbook.models import (Expense, RemittanceBatch, PaymentInstrument)
        from core.models import SiteConfig
        from core.utils import sabbath_week_of
        try:
            s = _dt.date.fromisoformat(request.POST["start"])
            e = _dt.date.fromisoformat(request.POST["end"])
        except (KeyError, ValueError):
            messages.error(request, "Pick a valid period to remit.")
            return redirect("report_remittance")
        field = SiteConfig.get().field_name or "the field"

        method = (request.POST.get("method") or "CHEQUE").upper()
        valid_methods = dict(PaymentInstrument.Method.choices)
        if method not in valid_methods:
            method = "CHEQUE"
        reference = (request.POST.get("instrument_number") or "").strip()
        try:
            paid = (_dt.date.fromisoformat(request.POST.get("date_issued"))
                    if request.POST.get("date_issued") else e)
        except ValueError:
            paid = e
        bank_id = request.POST.get("bank_account") or ""

        rows = balances.trust_summary(s, e)
        outstanding = [(r["department"], r["to_remit"]) for r in rows
                       if r["to_remit"] and r["to_remit"] > 0]
        if not outstanding:
            messages.info(request, "Nothing outstanding to remit for this period.")
            return redirect(f"{reverse('report_remittance')}?start={s}&end={e}")

        total = sum((amt for _, amt in outstanding), Decimal(0))
        batch = RemittanceBatch.create_batch(
            total_amount=total, status=RemittanceBatch.Status.REMITTED,
            period_start=s, period_end=e, created_by=request.user,
            approved_by=request.user, remitted_at=_tz.now())

        # one settlement instrument for the whole field payment
        inst = PaymentInstrument(
            method=method, instrument_number=reference[:40],
            payee=field, amount=total, date_issued=paid,
            status=PaymentInstrument.Status.ISSUED,
            source_kind=PaymentInstrument.SourceKind.REMITTANCE,
            remittance_batch=batch, recorded_by=request.user)
        if bank_id.isdigit():
            from statements.models import BankAccount
            inst.bank_account = BankAccount.objects.filter(pk=bank_id).first()
        inst.save()
        batch.payment = inst
        if method == "CHEQUE":          # keep legacy fields in step for old reports
            batch.cheque_no = reference[:30]
            batch.cheque_date = paid
        batch.save(update_fields=["payment", "cheque_no", "cheque_date"])

        for dept, amt in outstanding:
            Expense.objects.create(
                date=paid, sabbath_week=sabbath_week_of(paid), department=dept,
                description=f"Remittance to {field} ({s:%d %b}-{e:%d %b %Y})",
                amount=amt, category=Expense.Category.REMITTANCE,
                claimant=field, method=Expense.Method.CHEQUE,
                voucher_no=reference[:30], status=Expense.Status.PAID,
                paid_date=paid, remittance_batch=batch,
                recorded_by=request.user, approved_by=request.user)

        messages.success(
            request, f"Remitted {len(outstanding)} trust fund(s) totalling "
                     f"KES {total:,.2f} to {field}, settled by "
                     f"{inst.get_method_display()} {reference}." if reference else
                     f"Remitted {len(outstanding)} trust fund(s) totalling "
                     f"KES {total:,.2f} to {field} by {inst.get_method_display()}.")
        return redirect(f"{reverse('report_remittance')}?start={s}&end={e}")


class MemberStatementView(PeriodMixin, TemplateView):
    template_name = "reports/member_statement.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        member = get_object_or_404(Member, pk=kwargs["pk"])
        s, e = ctx["start"], ctx["end"]
        txns = (Transaction.objects.confirmed_credits().filter(
            member=member, date__gte=s, date__lte=e,
            excluded_from_income=False).values("department__name")
            .annotate(total=Sum("amount")).order_by("-total"))
        ctx["member"] = member
        ctx["rows"] = txns
        ctx["total"] = sum(r["total"] for r in txns)
        return ctx


class CashBookView(PeriodMixin, TemplateView):
    template_name = "reports/cashbook.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        s, e = ctx["start"], ctx["end"]
        entries = []
        for t in Transaction.objects.filter(date__gte=s, date__lte=e,
                                             direction=Transaction.Direction.CREDIT):
            entries.append({"date": t.date, "desc": t.payer_name or t.reference or "Receipt",
                            "credit": t.amount, "debit": None})
        for x in Expense.objects.filter(date__gte=s, date__lte=e,
                                        status__in=[Expense.Status.APPROVED, Expense.Status.PAID]):
            entries.append({"date": x.date, "desc": x.description,
                            "credit": None, "debit": x.amount})
        entries.sort(key=lambda r: r["date"])
        running = Decimal(0)
        for en in entries:
            running += (en["credit"] or 0) - (en["debit"] or 0)
            en["balance"] = running
        ctx["entries"] = entries
        ctx["closing"] = running
        return ctx


class ReconciliationView(PeriodMixin, TemplateView):
    template_name = "reports/reconciliation.html"

    def get_context_data(self, **kwargs):
        from statements.models import BankReconciliation
        ctx = super().get_context_data(**kwargs)
        s, e = ctx["start"], ctx["end"]
        bank = Transaction.objects.active().filter(channel=Transaction.Channel.BANK,
                                          date__gte=s, date__lte=e)
        ctx["ledger_credits"] = bank.filter(direction=Transaction.Direction.CREDIT).aggregate(
            t=Sum("amount"))["t"] or Decimal(0)
        ctx["ledger_debits"] = bank.filter(direction=Transaction.Direction.DEBIT).aggregate(
            t=Sum("amount"))["t"] or Decimal(0)
        ctx["unreconciled"] = bank.filter(
            allocation_status=Transaction.Status.REVIEW).order_by("date")

        # Proper bank reconciliation: the bank STATEMENT balance reconciles to the
        # BOOK (cashbook) balance through timing/at-hand adjustments — NOT to the
        # raw bank-credit total, which is why a naive "credits vs bank balance"
        # comparison looks wildly off. Show the latest statement reconciliation.
        rec = (BankReconciliation.objects.order_by("-statement_date")
               .prefetch_related("items").first())
        ctx["rec"] = rec
        if rec:
            adds = [i for i in rec.items.all() if i.effect == "ADD"]
            subs = [i for i in rec.items.all() if i.effect == "SUBTRACT"]
            ctx["rec_adds"] = adds
            ctx["rec_subs"] = subs
            ctx["rec_add_total"] = sum((i.amount for i in adds), Decimal(0))
            ctx["rec_sub_total"] = sum((i.amount for i in subs), Decimal(0))
            ctx["rec_computed_book"] = (rec.bank_balance + ctx["rec_add_total"]
                                        - ctx["rec_sub_total"])
            ctx["rec_variance"] = ctx["rec_computed_book"] - (rec.book_balance or Decimal(0))
        return ctx


class AnnualSummaryView(ReportAccessMixin, TemplateView):
    template_name = "reports/annual.html"

    def get_context_data(self, **kwargs):
        from django.db.models.functions import ExtractYear
        ctx = super().get_context_data(**kwargs)
        income = (Transaction.objects.confirmed_credits()
                  .filter(excluded_from_income=False)
                  .annotate(yr=ExtractYear("date"))
                  .values("yr").annotate(total=Sum("amount")).order_by("yr"))
        expense = (Expense.objects.filter(
            status__in=[Expense.Status.APPROVED, Expense.Status.PAID])
            .exclude(category__in=[Expense.Category.REMITTANCE, Expense.Category.LOAN_REPAYMENT])
            .annotate(yr=ExtractYear("date"))
            .values("yr").annotate(total=Sum("amount")).order_by("yr"))
        inc = {r["yr"]: r["total"] for r in income}
        exp = {r["yr"]: r["total"] for r in expense}
        years = sorted(set(inc) | set(exp))
        ctx["rows"] = [{"year": y, "income": inc.get(y, 0), "expense": exp.get(y, 0),
                        "net": (inc.get(y, 0) or 0) - (exp.get(y, 0) or 0)} for y in years]
        # historical reference years (collection / trust fund / expenditure)
        from core.models import HistoricalYear, HistoricalMonth
        ctx["historical"] = list(HistoricalYear.objects.all())
        # seasonality: average collection / trust / expenditure by calendar month
        MN = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep",
              "Oct", "Nov", "Dec"]
        agg = {m: {"c": 0.0, "t": 0.0, "e": 0.0, "n": 0} for m in range(1, 13)}
        for hm in HistoricalMonth.objects.all():
            a = agg[hm.month]
            a["c"] += float(hm.collection); a["t"] += float(hm.trust_fund)
            a["e"] += float(hm.expenditure); a["n"] += 1
        season = {"labels": MN,
                  "collection": [round(agg[m]["c"] / agg[m]["n"], 2) if agg[m]["n"] else 0 for m in range(1, 13)],
                  "trust": [round(agg[m]["t"] / agg[m]["n"], 2) if agg[m]["n"] else 0 for m in range(1, 13)],
                  "expenditure": [round(agg[m]["e"] / agg[m]["n"], 2) if agg[m]["n"] else 0 for m in range(1, 13)]}
        ctx["season_json"] = safe_json(season)
        ctx["has_season"] = HistoricalMonth.objects.exists()
        return ctx


class HistoricalYearManageView(TreasurerRequiredMixin, TemplateView):
    """Add, edit, or remove prior-year comparison figures, now with per-month
    detail. When a year has monthly rows, its yearly totals are computed from
    them (so the two always agree); a year with no months keeps the figure typed
    by hand. Monthly data can be imported from Excel for fast back-filling and
    enables month-on-month trend analysis."""
    template_name = "reports/historical_manage.html"

    MONTHS = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    @staticmethod
    def _recompute_year(year):
        """Set a HistoricalYear's totals from its months, if any exist."""
        from decimal import Decimal
        from django.db.models import Sum
        from core.models import HistoricalYear, HistoricalMonth
        months = HistoricalMonth.objects.filter(year=year)
        if not months.exists():
            return
        agg = months.aggregate(c=Sum("collection"), t=Sum("trust_fund"),
                               e=Sum("expenditure"))
        HistoricalYear.objects.update_or_create(year=year, defaults=dict(
            collection=agg["c"] or Decimal(0), trust_fund=agg["t"] or Decimal(0),
            expenditure=agg["e"] or Decimal(0), note="Computed from monthly records"))

    def get(self, request, *args, **kwargs):
        if request.GET.get("sample"):
            return self._sample_xlsx()
        return super().get(request, *args, **kwargs)

    def _sample_xlsx(self):
        import io
        import datetime as _d
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Monthly history"
        head = ["Year", "Month (1-12)", "Collection", "Trust fund", "Expenditure"]
        ws.append(head)
        for c in range(1, len(head) + 1):
            cell = ws.cell(1, c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F5F4F")
            cell.alignment = Alignment(horizontal="center")
        # a couple of illustrative rows for last year
        ly = _d.date.today().year - 1
        for m, coll, tr, ex in [(1, 120000, 70000, 45000), (2, 98000, 60000, 52000)]:
            ws.append([ly, m, coll, tr, ex])
        for col, w in zip("ABCDE", (8, 14, 14, 14, 14)):
            ws.column_dimensions[col].width = w
        info = wb.create_sheet("How to use")
        for line in [
            "One row per month. Year and Month (1-12) identify the period.",
            "Collection = total receipts that month (all funds).",
            "Trust fund = the portion that is trust/remittable.",
            "Expenditure = total spending that month.",
            "Re-importing a month overwrites that month. Yearly totals are computed automatically.",
        ]:
            info.append([line])
        info.column_dimensions["A"].width = 90
        buf = io.BytesIO(); wb.save(buf)
        resp = HttpResponse(buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = 'attachment; filename="monthly_history_template.xlsx"'
        return resp

    def get_context_data(self, **kwargs):
        from collections import defaultdict
        from core.models import HistoricalYear, HistoricalMonth
        ctx = super().get_context_data(**kwargs)
        ctx["years"] = HistoricalYear.objects.all()
        by_year = defaultdict(lambda: [None] * 13)
        for hm in HistoricalMonth.objects.all():
            by_year[hm.year][hm.month] = hm
        month_rows = []
        for yr in sorted(by_year, reverse=True):
            cells = by_year[yr]
            present = [c for c in cells[1:] if c]
            month_rows.append({
                "year": yr, "cells": cells[1:],
                "collection": sum((c.collection for c in present), 0),
                "trust_fund": sum((c.trust_fund for c in present), 0),
                "expenditure": sum((c.expenditure for c in present), 0),
                "count": len(present)})
        ctx["month_rows"] = month_rows
        ctx["months"] = self.MONTHS[1:]
        return ctx

    def post(self, request, *args, **kwargs):
        from decimal import Decimal, InvalidOperation
        from core.models import HistoricalYear, HistoricalMonth
        action = request.POST.get("action")

        def dec(k):
            return Decimal(str(request.POST.get(k) or "0").replace(",", ""))

        if action == "delete":
            HistoricalYear.objects.filter(pk=request.POST.get("pk")).delete()
            messages.success(request, "Historical year removed.")
            return redirect("historical_manage")

        if action == "delete_year_all":
            try:
                yr = int(request.POST.get("year"))
            except (TypeError, ValueError):
                return redirect("historical_manage")
            HistoricalMonth.objects.filter(year=yr).delete()
            HistoricalYear.objects.filter(year=yr).delete()
            messages.success(request, f"Deleted all historical data for {yr}.")
            return redirect("historical_manage")

        if action == "save_month":
            try:
                year = int(request.POST.get("year"))
                month = int(request.POST.get("month"))
                assert 1 <= month <= 12
            except (TypeError, ValueError, AssertionError):
                messages.error(request, "Enter a valid year and month (1–12).")
                return redirect("historical_manage")
            HistoricalMonth.objects.update_or_create(
                year=year, month=month, defaults=dict(
                    collection=dec("collection"), trust_fund=dec("trust_fund"),
                    expenditure=dec("expenditure")))
            self._recompute_year(year)
            messages.success(request, f"Saved {self.MONTHS[month]} {year}.")
            return redirect("historical_manage")

        if action == "delete_month":
            hm = HistoricalMonth.objects.filter(pk=request.POST.get("pk")).first()
            if hm:
                yr = hm.year; hm.delete(); self._recompute_year(yr)
                messages.success(request, "Month removed.")
            return redirect("historical_manage")

        if action == "import":
            return self._import(request)

        # save a whole-year figure by hand (only when there are no months for it)
        try:
            year = int(request.POST.get("year"))
            HistoricalYear.objects.update_or_create(
                year=year, defaults=dict(collection=dec("collection"),
                    trust_fund=dec("trust_fund"), expenditure=dec("expenditure"),
                    note=(request.POST.get("note") or "Entered manually")[:200]))
            messages.success(request, f"Saved historical figures for {year}.")
        except (TypeError, ValueError, InvalidOperation):
            messages.error(request, "Enter a valid year and numeric amounts.")
        return redirect("historical_manage")

    def _import(self, request):
        from decimal import Decimal, InvalidOperation
        import openpyxl
        from core.models import HistoricalMonth
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "Choose an Excel file to import.")
            return redirect("historical_manage")
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx; _lx("historical import")
            messages.error(request, "That file couldn't be read as an Excel workbook.")
            return redirect("historical_manage")
        ws = wb["Monthly history"] if "Monthly history" in wb.sheetnames else wb.active
        n = 0; years = set()
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0 or row is None:
                continue
            try:
                year = int(row[0]); month = int(row[1])
                if not (1 <= month <= 12):
                    continue
                def d(v):
                    return Decimal(str(v or "0").replace(",", ""))
                HistoricalMonth.objects.update_or_create(
                    year=year, month=month, defaults=dict(
                        collection=d(row[2]), trust_fund=d(row[3]),
                        expenditure=d(row[4])))
                years.add(year); n += 1
            except (TypeError, ValueError, InvalidOperation, IndexError):
                continue
        for yr in years:
            self._recompute_year(yr)
        messages.success(request, f"Imported {n} monthly record(s) across "
                                  f"{len(years)} year(s). Yearly totals updated.")
        return redirect("historical_manage")


class AuditLogView(ReportAccessMixin, TemplateView):
    template_name = "reports/audit.html"

    def _models(self):
        from giving.models import Transaction as T, AllocationRule as AR
        from cashbook.models import Expense as X
        from members.models import Member as M
        candidates = {"Transaction": T, "Expense": X, "Member": M, "Allocation rule": AR}
        return {n: m for n, m in candidates.items() if hasattr(m, "history")}

    def _collect(self, request, cap=1500):
        import datetime as dt
        models = self._models()
        model_f = request.GET.get("model", "")
        user_f = request.GET.get("user", "")
        type_f = request.GET.get("type", "")          # +, ~, -
        q = (request.GET.get("q", "") or "").strip().lower()

        def _date(name):
            raw = request.GET.get(name)
            try:
                return dt.date.fromisoformat(raw) if raw else None
            except ValueError:
                return None
        start, end = _date("start"), _date("end")

        records, users = [], set()
        # Department/SplitFund names for AllocationRule, prefetched once: its
        # __str__() touches self.split_fund or self.department, and
        # h.instance (a full model reconstructed from the historical row) has
        # no select_related — calling str() on it for every historical
        # AllocationRule row triggered a fresh FK query each time (up to
        # ~1500 extra queries for a church with a large rule history).
        from departments.models import Department
        from giving.models import SplitFund
        dept_names = dict(Department.objects.values_list("id", "name"))
        split_names = dict(SplitFund.objects.values_list("id", "name"))

        for name, model in models.items():
            hq = model.history.all().select_related("history_user")
            if start:
                hq = hq.filter(history_date__date__gte=start)
            if end:
                hq = hq.filter(history_date__date__lte=end)
            if user_f:
                hq = hq.filter(history_user__username=user_f)
            if type_f in ("+", "~", "-"):
                hq = hq.filter(history_type=type_f)
            # collect the user list for the filter dropdown (cheap, distinct)
            for u in (model.history.exclude(history_user__isnull=True)
                      .values_list("history_user__username", flat=True).distinct()[:200]):
                users.add(u)
            if model_f and model_f != name:
                continue
            for h in hq.order_by("-history_date")[:cap]:
                try:
                    if name == "Allocation rule":
                        # build the display string from the historical row's
                        # own FK id columns + the prefetched name maps,
                        # instead of str(h.instance) (see note above)
                        target = (split_names.get(h.split_fund_id)
                                  or dept_names.get(h.department_id) or "—")
                        obj = f"{h.reference} -> {target}"
                    else:
                        obj = str(h.instance)
                except Exception:
                    from core.utils import log_exception as _lx; _lx('reports/views.py')
                    obj = f"{name} #{h.id}"
                uname = getattr(h.history_user, "username", "") or "system"
                if q and q not in (obj + " " + uname + " " + name).lower():
                    continue
                records.append({
                    "model": name, "when": h.history_date, "user": uname,
                    "type": h.get_history_type_display(), "obj": obj})
        records.sort(key=lambda r: r["when"], reverse=True)
        return records, sorted(users)

    def get(self, request, *args, **kwargs):
        records, users = self._collect(request)
        if request.GET.get("export") == "csv":
            from reports.exports import csv_response
            header = ["When", "Record type", "Change", "By", "Detail"]
            rows = [[r["when"].strftime("%Y-%m-%d %H:%M:%S"), r["model"],
                     r["type"], r["user"], r["obj"]] for r in records]
            return csv_response("audit_log.csv", header, rows)

        from django.core.paginator import Paginator
        paginator = Paginator(records, 50)
        page = paginator.get_page(request.GET.get("page"))
        ctx = self.get_context_data(**kwargs)
        ctx.update({
            "page_obj": page, "records": page.object_list,
            "total": paginator.count,
            "models": list(self._models().keys()), "users": users,
            "f": {"model": request.GET.get("model", ""),
                  "user": request.GET.get("user", ""),
                  "type": request.GET.get("type", ""),
                  "q": request.GET.get("q", ""),
                  "start": request.GET.get("start", ""),
                  "end": request.GET.get("end", "")},
            "querystring": _qs_without(request, "page"),
        })
        return self.render_to_response(ctx)


def _qs_without(request, *drop):
    from urllib.parse import urlencode
    items = [(k, v) for k, v in request.GET.items() if k not in drop and v]
    return urlencode(items)


# ---- Envelope reports ----
import datetime as dt
from .services import envelope_reports


from core.utils import last_saturday as _last_saturday


class EnvelopeSabbathView(ReportAccessMixin, TemplateView):
    template_name = "reports/envelope_sabbath.html"

    def _date(self, request):
        raw = request.GET.get("date")
        try:
            return dt.date.fromisoformat(raw) if raw else _last_saturday()
        except ValueError:
            return _last_saturday()

    def get(self, request, *args, **kwargs):
        date = self._date(request)
        data = envelope_reports.sabbath_statement(date)
        if request.GET.get("export") == "csv":
            header = ["Receipt", "Contributor"] + [f.name for f in data["funds"]] + ["Total"]
            rows = []
            for r in data["rows"]:
                rows.append([r["envelope"].receipt_no, r["envelope"].contributor_name]
                            + [r["cells"].get(f.id, "") for f in data["funds"]]
                            + [r["total"]])
            rows.append(["", "TOTAL"] + [data["fund_totals"][f.id] for f in data["funds"]]
                        + [data["grand_total"]])
            return csv_response(f"envelopes_{date}.csv", header, rows)
        ctx = self.get_context_data(**kwargs)
        ctx["d"] = data
        ctx["date"] = date
        return self.render_to_response(ctx)


class EnvelopeSummaryView(ReportAccessMixin, TemplateView):
    template_name = "reports/envelope_summary.html"

    def get(self, request, *args, **kwargs):
        today = dt.date.today()
        try:
            year = int(request.GET.get("year", today.year))
            month = int(request.GET.get("month", today.month))
        except ValueError:
            year, month = today.year, today.month
        data = envelope_reports.monthly_summary(year, month)
        if request.GET.get("export") == "csv":
            header = ["Fund"] + [s.strftime("%d %b") for s in data["saturdays"]] + ["Total"]
            rows = [["— TRUST FUNDS —"]]
            for r in data["trust_rows"]:
                rows.append([r["fund"].name] + r["cols"] + [r["total"]])
            rows.append(["TOTAL TRUST FUNDS"] + data["trust_col_totals"] + [data["trust_total"]])
            rows.append(["— LOCAL FUNDS —"])
            for r in data["local_rows"]:
                rows.append([r["fund"].name] + r["cols"] + [r["total"]])
            rows.append(["TOTAL LOCAL FUNDS"] + data["local_col_totals"] + [data["local_total"]])
            return csv_response(f"offering_summary_{year}_{month:02d}.csv", header, rows)
        ctx = self.get_context_data(**kwargs)
        ctx["d"] = data
        ctx["month_label"] = dt.date(year, month, 1).strftime("%B %Y")
        ctx["year"], ctx["month"] = year, month
        return self.render_to_response(ctx)


# ---- Monthly account reports ----
from .services import monthly


def _year_from(request):
    try:
        return int(request.GET.get("year", dt.date.today().year))
    except (ValueError, TypeError):
        return dt.date.today().year


class MonthlyAccountsView(ReportAccessMixin, TemplateView):
    template_name = "reports/monthly_accounts.html"

    def get(self, request, *args, **kwargs):
        year = _year_from(request)
        coll = monthly.collections_by_account(year)
        exp = monthly.expenses_by_account(year)
        if request.GET.get("export") == "csv":
            header = ["Account"] + [lbl for _, lbl in coll["months"]] + ["Total"]
            rows = [["— COLLECTIONS —"]]
            for r in coll["rows"]:
                rows.append([str(r["dept"])] + r["cells"] + [r["total"]])
            rows.append(["Total collections"] + coll["col_totals"] + [coll["grand"]])
            rows.append(["— EXPENSES —"])
            for r in exp["rows"]:
                rows.append([str(r["dept"])] + r["cells"] + [r["total"]])
            rows.append(["Total expenses"] + exp["col_totals"] + [exp["grand"]])
            return csv_response(f"accounts_{year}.csv", header, rows)
        ctx = self.get_context_data(**kwargs)
        ctx.update(year=year, coll=coll, exp=exp,
                   years=range(dt.date.today().year, dt.date.today().year - 6, -1))
        return self.render_to_response(ctx)


class TrustMonthlyView(ReportAccessMixin, TemplateView):
    template_name = "reports/trust_monthly.html"

    def get(self, request, *args, **kwargs):
        year = _year_from(request)
        data = monthly.trust_monthly(year)
        if request.GET.get("export") == "csv":
            header = ["Trust account"] + [lbl for _, lbl in data["months"]] + ["Total"]
            rows = [[str(r["dept"])] + r["cells"] + [r["total"]] for r in data["rows"]]
            rows.append(["TOTAL TRUST FUNDS"] + data["col_totals"] + [data["grand"]])
            return csv_response(f"trust_monthly_{year}.csv", header, rows)
        ctx = self.get_context_data(**kwargs)
        ctx.update(year=year, d=data,
                   years=range(dt.date.today().year, dt.date.today().year - 6, -1))
        return self.render_to_response(ctx)


class CollectionsSummaryView(ReportAccessMixin, TemplateView):
    template_name = "reports/collections_summary.html"

    def get(self, request, *args, **kwargs):
        year = _year_from(request)
        data = monthly.collections_summary(year)
        if request.GET.get("export") == "csv":
            header = ["Month", "Collections", "Trust funds", "Local funds",
                      "Expenditure", "Net"]
            rows = [[r["month"], r["collections"], r["trust"], r["local"],
                     r["expenditure"], r["net"]] for r in data["rows"]]
            rows.append(["TOTAL", data["tot_collections"], data["tot_trust"],
                         data["tot_local"], data["tot_expenditure"], data["tot_net"]])
            return csv_response(f"collections_summary_{year}.csv", header, rows)
        ctx = self.get_context_data(**kwargs)
        ctx.update(year=year, d=data,
                   years=range(dt.date.today().year, dt.date.today().year - 6, -1))
        return self.render_to_response(ctx)


class CollectionsDetailView(PeriodMixin, TemplateView):
    """Detailed collections for any chosen period, broken down by fund. The grand
    total reconciles to the Collections figure on the Collections Summary for the
    same dates. Exports to Excel (.xlsx) and CSV."""
    template_name = "reports/collections_detail.html"

    def get(self, request, *args, **kwargs):
        s, e = self.period()
        data = monthly.collections_detail(s, e)
        export = request.GET.get("export")
        if export in ("xlsx", "csv"):
            header = ["Fund", "Type", "Receipts", "Collected"]
            rows = [[r["fund"], r["type"], r["n"], float(r["amount"])] for r in data["rows"]]
            rows.append(["Trust funds — subtotal", "", "", float(data["tot_trust"])])
            rows.append(["Local funds — subtotal", "", "", float(data["tot_local"])])
            rows.append(["TOTAL COLLECTIONS", "", data["n_receipts"], float(data["tot_collections"])])
            fname = f"collections_detail_{s}_{e}"
            if export == "csv":
                return csv_response(fname + ".csv", header, rows)
            from reports.exports import xlsx_response
            from core.models import SiteConfig
            return xlsx_response(fname + ".xlsx", header, rows,
                                 title=f"Collections detail ({s} to {e})",
                                 church=SiteConfig.get().church_name)
        ctx = self.get_context_data(**kwargs)
        ctx.update(d=data)
        return self.render_to_response(ctx)


# ===================== Trust Fund Remittance subsystem =====================
import datetime as _dt
from django.utils import timezone as _tz
from cashbook.models import RemittanceBatch
from core.models import SiteConfig
from core.utils import sabbath_week_of


def _days_outstanding(dept):
    """Days since the oldest *unremitted* trust receipt for a fund. Receipts up to
    the last remittance's period are already settled, so we count only from the
    first receipt after it — not the first contribution ever (which made everything look
    months overdue even right after remitting)."""
    from cashbook.models import RemittanceBatch
    last = (RemittanceBatch.objects.filter(
        status=RemittanceBatch.Status.REMITTED, period_end__isnull=False)
        .order_by("-period_end").first())
    q = Transaction.objects.filter(
        department=dept, direction=Transaction.Direction.CREDIT,
        confirmed=True, is_reversed=False, is_reversal=False)
    if last:
        q = q.filter(date__gt=last.period_end)
    first = q.order_by("date").first()
    if not first:
        return 0
    return (_dt.date.today() - first.date).days


def _repost_to_ledger(expenses=None):
    """After a bulk .update() (which bypasses post_save signals), rebuild the
    general ledger so it always reflects batch approve/remit and stays reconciled."""
    try:
        from ledger.services import posting
        if posting.chart_ready():
            posting.rebuild()
    except Exception:
        from core.utils import log_exception as _lx; _lx('reports/views.py')
        pass


def remittance_dashboard_rows(start=None, end=None):
    rows = []
    for r in balances.trust_summary(start, end):   # period (or lifetime) collected vs remitted
        out = r["to_remit"]
        rows.append({
            "department": r["department"], "collected": r["collected"],
            "remitted": r["remitted"], "outstanding": out,
            "unreceipted": r["unreceipted"],
            "days": _days_outstanding(r["department"]) if out > 0 else 0,
        })
    return rows


def _remit_period(request):
    """Resolve a remittance period from the request: ?month=YYYY-MM, or
    ?start=&end=, else None (lifetime)."""
    import datetime as _dt, calendar as _cal
    month = request.GET.get("month") or request.POST.get("month")
    if month:
        try:
            y, m = (int(x) for x in month.split("-"))
            last = _cal.monthrange(y, m)[1]
            return _dt.date(y, m, 1), _dt.date(y, m, last), month
        except (ValueError, TypeError):
            pass
    s = request.GET.get("start") or request.POST.get("start")
    e = request.GET.get("end") or request.POST.get("end")
    def pd(v):
        try:
            return _dt.datetime.strptime(v, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return None
    return pd(s), pd(e), None


class RemittanceDashboardView(ReportAccessMixin, TemplateView):
    template_name = "reports/remittance_dashboard.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        start, end, month = _remit_period(self.request)
        rows = remittance_dashboard_rows(start, end)
        ctx["rows"] = rows
        ctx["period_start"], ctx["period_end"], ctx["period_month"] = start, end, month
        ctx["period_active"] = bool(start or end)
        ctx["total_outstanding"] = sum((r["outstanding"] for r in rows), Decimal(0))
        ctx["total_collected"] = sum((r["collected"] for r in rows), Decimal(0))
        ctx["total_remitted"] = sum((r["remitted"] for r in rows), Decimal(0))
        ctx["total_unreceipted"] = sum((r["unreceipted"] for r in rows), Decimal(0))
        ctx["max_days"] = max([r["days"] for r in rows], default=0)
        # Next remittance: use the configured per-month deadlines if any exist
        # (their dates are set freely per month, not on a fixed day). We count
        # down to the *reporting Sabbath* — the Saturday whose count must be in
        # the remittance. Fall back to the configured due-day only if no
        # deadlines have been entered yet.
        import datetime as _dt, calendar as _cal
        from cashbook.models import RemittanceDeadline
        today = _dt.date.today()
        nxt = (RemittanceDeadline.objects.filter(remitted=False, deadline__gte=today)
               .order_by("deadline").first())
        overdue_dl = (RemittanceDeadline.objects.filter(remitted=False, deadline__lt=today)
                      .order_by("-deadline").first())
        active = nxt or overdue_dl
        if active:
            ctx["next_deadline"] = active
            ctx["due_date"] = active.deadline
            ctx["reporting_sabbath"] = active.reporting_sabbath
            ctx["days_to_deadline"] = (active.deadline - today).days
            ctx["days_to_sabbath"] = (active.reporting_sabbath - today).days
            ctx["deadline_period"] = active.get_period_display()
            ctx["has_deadlines"] = True
        else:
            # legacy fallback: configured day of the following month
            cfg = SiteConfig.get()
            ny, nm = (today.year + 1, 1) if today.month == 12 else (today.year, today.month + 1)
            due_day = min(cfg.trust_remit_due_day or 15, _cal.monthrange(ny, nm)[1])
            ctx["due_date"] = _dt.date(ny, nm, due_day)
            ctx["days_to_deadline"] = (ctx["due_date"] - today).days
            ctx["has_deadlines"] = False
        ctx["due_overdue"] = ctx["total_outstanding"] > 0 and ctx.get("days_to_deadline", 99) < 0
        ctx["batches"] = RemittanceBatch.objects.all()[:10]
        ctx["field_name"] = SiteConfig.get().field_name or "the field"
        return ctx


class RemittanceBatchCreateView(TreasurerRequiredMixin, View):
    """Generate a DRAFT batch. POST 'all'=1 to include every outstanding trust
    fund, or one or more 'fund' ids for a per-fund wizard."""
    def post(self, request):
        start, end, month = _remit_period(request)
        rows = {r["department"].id: r for r in remittance_dashboard_rows(start, end)}
        if request.POST.get("all"):
            chosen = [r for r in rows.values() if r["outstanding"] > 0]
        else:
            ids = [int(i) for i in request.POST.getlist("fund") if i.isdigit()]
            chosen = [rows[i] for i in ids if i in rows and rows[i]["outstanding"] > 0]
        if not chosen:
            messages.error(request, "No outstanding trust funds to remit for that period.")
            return redirect("remittance_dashboard")
        field = SiteConfig.get().field_name or "the field"
        batch = RemittanceBatch.create_batch(
            created_by=request.user, status=RemittanceBatch.Status.DRAFT,
            period_start=start, period_end=end)
        today = _dt.date.today()
        # date the remittance expense within the period it covers, so it ties to
        # the right month's trust collection
        rdate = end or today
        plabel = (f" for {month}" if month else
                  f" for {start:%d %b %Y}–{end:%d %b %Y}" if (start and end) else "")
        for r in chosen:
            Expense.objects.create(
                date=rdate, sabbath_week=sabbath_week_of(rdate), department=r["department"],
                description=f"Trust remittance to {field} — {batch.batch_number}{plabel}",
                amount=r["outstanding"], category=Expense.Category.REMITTANCE,
                claimant=field, method=Expense.Method.CHEQUE,
                status=Expense.Status.PENDING, recorded_by=request.user,
                remittance_batch=batch)
        batch.recompute_total()
        batch.save(update_fields=["total_amount"])
        messages.success(request, f"Created remittance batch {batch.batch_number} "
                                  f"covering {len(chosen)} fund(s){plabel}.")
        return redirect("remittance_batch_detail", pk=batch.pk)


class RemittanceBatchDetailView(ReportAccessMixin, TemplateView):
    template_name = "reports/remittance_batch.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        batch = get_object_or_404(RemittanceBatch, pk=kwargs["pk"])
        ctx["batch"] = batch
        ctx["lines"] = batch.expenses.select_related("department").all()
        ctx["field_name"] = SiteConfig.get().field_name or "the field"
        from statements.models import BankAccount
        ctx["bank_accounts"] = BankAccount.objects.all()
        return ctx


class RemittanceBatchApproveView(TreasurerRequiredMixin, View):
    def post(self, request, pk):
        batch = get_object_or_404(RemittanceBatch, pk=pk)
        if batch.status != RemittanceBatch.Status.DRAFT:
            messages.error(request, "Only a draft batch can be approved.")
            return redirect("remittance_batch_detail", pk=pk)
        batch.status = RemittanceBatch.Status.APPROVED
        batch.approved_by = request.user
        batch.save(update_fields=["status", "approved_by"])
        batch.expenses.update(status=Expense.Status.APPROVED, approved_by=request.user)
        _repost_to_ledger(batch.expenses.all())
        messages.success(request, f"Batch {batch.batch_number} approved.")
        return redirect("remittance_batch_detail", pk=pk)


class RemittanceBatchRemitView(TreasurerRequiredMixin, View):
    """Mark a batch as sent — only once a payment instrument has been issued and
    linked. The instrument is the settlement record for the trust liability;
    bank reconciliation later only flips it to Cleared (no extra journals)."""
    def post(self, request, pk):
        batch = get_object_or_404(RemittanceBatch, pk=pk)
        if batch.status != RemittanceBatch.Status.APPROVED:
            messages.error(request, "Approve the batch before marking it sent.")
            return redirect("remittance_batch_detail", pk=pk)
        if not batch.is_settled:
            messages.error(request, "Issue and link a payment instrument (cheque, "
                "EFT, M-Pesa, etc.) before marking this batch as sent.")
            return redirect("remittance_batch_detail", pk=pk)
        inst = batch.payment
        paid_date = inst.date_issued or _dt.date.today()
        batch.status = RemittanceBatch.Status.REMITTED
        batch.remitted_at = _tz.now()
        # keep the legacy fields in step for any old reports still reading them
        if inst.method == "CHEQUE":
            batch.cheque_no = inst.instrument_number[:30]
            batch.cheque_date = inst.date_issued
        batch.save(update_fields=["status", "remitted_at", "cheque_no", "cheque_date"])
        batch.expenses.update(status=Expense.Status.PAID, paid_date=paid_date,
                              voucher_no=inst.instrument_number[:30])
        _repost_to_ledger(batch.expenses.all())
        messages.success(request, f"Batch {batch.batch_number} marked sent, settled by "
                                  f"{inst.get_method_display()} {inst.instrument_number}.")
        return redirect("remittance_batch_detail", pk=pk)


class RemittanceBatchIssuePaymentView(TreasurerRequiredMixin, View):
    """Issue a payment instrument that settles this remittance batch and link it.
    Posts no journal entries — the batch's remittance expenses already account
    for the liability; this only records how it is being paid."""
    def post(self, request, pk):
        from cashbook.models import PaymentInstrument
        batch = get_object_or_404(RemittanceBatch, pk=pk)
        if batch.status not in (RemittanceBatch.Status.APPROVED,
                                RemittanceBatch.Status.DRAFT):
            messages.error(request, "A payment can only be issued for a draft or "
                                    "approved batch.")
            return redirect("remittance_batch_detail", pk=pk)
        method = request.POST.get("method") or "CHEQUE"
        number = (request.POST.get("instrument_number") or "").strip()[:40]
        try:
            issued = _dt.date.fromisoformat(request.POST.get("date_issued")) \
                if request.POST.get("date_issued") else _dt.date.today()
        except ValueError:
            issued = _dt.date.today()
        bank_id = request.POST.get("bank_account") or ""
        inst = PaymentInstrument(
            method=method, instrument_number=number,
            payee="Conference remittance", amount=batch.total_amount,
            date_issued=issued, status=PaymentInstrument.Status.ISSUED,
            source_kind=PaymentInstrument.SourceKind.REMITTANCE,
            remittance_batch=batch, recorded_by=request.user)
        if bank_id.isdigit():
            from statements.models import BankAccount
            inst.bank_account = BankAccount.objects.filter(pk=bank_id).first()
        inst.save()
        batch.payment = inst
        batch.save(update_fields=["payment"])
        messages.success(request, f"Issued {inst.get_method_display()} "
                                  f"{inst.instrument_number} for this remittance.")
        return redirect("remittance_batch_detail", pk=pk)


class RemittanceBatchListView(ReportAccessMixin, TemplateView):
    template_name = "reports/remittance_batches.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["batches"] = RemittanceBatch.objects.all()
        return ctx


class RemittanceCalendarView(ReportAccessMixin, TemplateView):
    """Per-year remittance calendar: each period's deadline and the reporting
    Sabbath it maps to (the most recent Saturday on/before the deadline). When a
    deadline doesn't fall on a Sabbath, the previous Sabbath is shown as the
    reporting Sabbath, and due-soon / overdue items are highlighted."""
    template_name = "reports/remittance_calendar.html"

    def get_context_data(self, **kwargs):
        from cashbook.models import RemittanceDeadline, RemittanceBatch
        import calendar as _cal
        ctx = super().get_context_data(**kwargs)
        today = _dt.date.today()
        try:
            year = int(self.request.GET.get("year", today.year))
        except (TypeError, ValueError):
            year = today.year
        deadlines = list(RemittanceDeadline.objects.filter(year=year))

        # auto-mark a period as remitted when a remitted batch covers it. A batch
        # covers a month if its period overlaps that month, or (if it has no
        # period set) if it was remitted within the month. This keeps the
        # calendar honest without the treasurer ticking each box by hand.
        remitted_batches = list(RemittanceBatch.objects.filter(
            status=RemittanceBatch.Status.REMITTED))
        auto_marked = 0
        for d in deadlines:
            if d.remitted:
                continue
            m_start = _dt.date(year, d.period_month, 1)
            m_end = _dt.date(year, d.period_month,
                             _cal.monthrange(year, d.period_month)[1])
            covered = False
            for b in remitted_batches:
                ps, pe = b.period_start, b.period_end
                if ps and pe:
                    if ps <= m_end and pe >= m_start:
                        covered = True; break
                elif b.remitted_at and m_start <= b.remitted_at.date() <= m_end:
                    covered = True; break
            if covered:
                d.remitted = True
                d.save(update_fields=["remitted"])
                auto_marked += 1
        if auto_marked:
            messages.info(self.request,
                f"{auto_marked} period(s) marked remitted automatically from "
                f"completed remittance batches.")

        ctx["year"] = year
        ctx["prev_year"] = year - 1
        ctx["next_year"] = year + 1
        ctx["deadlines"] = deadlines
        ctx["has_any"] = bool(deadlines)
        ctx["due_soon"] = [d for d in deadlines if d.is_due_soon]
        ctx["overdue"] = [d for d in deadlines if d.is_overdue]
        return ctx


class RemittanceCalendarGenerateView(TreasurerRequiredMixin, View):
    """Auto-generate a year's monthly remittance deadlines. The default deadline
    is the 1st of the following month (adjustable afterwards). Existing periods
    are left untouched."""

    def post(self, request):
        from cashbook.models import RemittanceDeadline
        import calendar
        today = _dt.date.today()
        try:
            year = int(request.POST.get("year", today.year))
        except (TypeError, ValueError):
            year = today.year
        try:
            due_day = int(request.POST.get("due_day", 1))
        except (TypeError, ValueError):
            due_day = 1
        due_day = min(max(due_day, 1), 28)
        created = 0
        for m in range(1, 13):
            # deadline falls in the FOLLOWING month by default
            dyear, dmonth = (year + 1, 1) if m == 12 else (year, m + 1)
            _, last_day = calendar.monthrange(dyear, dmonth)
            deadline = _dt.date(dyear, dmonth, min(due_day, last_day))
            _, was_created = RemittanceDeadline.objects.get_or_create(
                year=year, period_month=m,
                defaults={"deadline": deadline,
                          "label": f"{calendar.month_name[m]} remittance"})
            if was_created:
                created += 1
        messages.success(request, f"Generated {created} remittance deadline(s) for "
                                  f"{year}. You can adjust individual dates below.")
        return redirect(f"{reverse('remittance_calendar')}?year={year}")


class RemittanceDeadlineUpdateView(TreasurerRequiredMixin, View):
    """Edit one deadline's date/label/notes, or toggle its remitted flag."""

    def post(self, request, pk):
        from cashbook.models import RemittanceDeadline
        d = get_object_or_404(RemittanceDeadline, pk=pk)
        if request.POST.get("toggle_remitted"):
            d.remitted = not d.remitted
            d.save(update_fields=["remitted"])
            messages.success(request, f"{d.get_period_display()} marked "
                                      f"{'remitted' if d.remitted else 'not remitted'}.")
            return redirect(f"{reverse('remittance_calendar')}?year={d.year}")
        new_date = request.POST.get("deadline")
        if new_date:
            try:
                d.deadline = _dt.date.fromisoformat(new_date)
            except ValueError:
                messages.error(request, "Invalid date.")
                return redirect(f"{reverse('remittance_calendar')}?year={d.year}")
        d.label = (request.POST.get("label") or d.label)[:60]
        d.notes = (request.POST.get("notes") or "")[:200]
        d.save()
        messages.success(request, f"Updated {d.get_period_display()} deadline.")
        return redirect(f"{reverse('remittance_calendar')}?year={d.year}")


# ===================== Budget vs Actual report =====================
from .services import budget as budget_svc


class BudgetVsActualView(ReportAccessMixin, TemplateView):
    template_name = "reports/budget_vs_actual.html"

    def get(self, request, *args, **kwargs):
        today = _dt.date.today()
        try:
            year = int(request.GET.get("year", today.year))
        except (TypeError, ValueError):
            year = today.year
        period = (request.GET.get("period") or "ANNUAL").upper()
        month = request.GET.get("month")
        quarter = request.GET.get("quarter")
        data = budget_svc.budget_vs_actual(year, period, month, quarter)
        if request.GET.get("export") in ("csv", "xlsx"):
            header = ["Fund", "Budget", "Actual", "Variance", "Variance %"]
            rows = [[r["department"].name, r["budget"], r["actual"], r["variance"],
                     (round(float(r['variance_pct']), 1) if r["variance_pct"] is not None else "")]
                    for r in data["rows"]]
            t = data["totals"]
            rows.append(["TOTAL", t["budget"], t["actual"], t["variance"],
                         (round(float(t['variance_pct']), 1) if t["variance_pct"] is not None else "")])
            fname = f"budget_vs_actual_{data['label']}".replace(" ", "_")
            if request.GET.get("export") == "xlsx":
                return xlsx_response(fname + ".xlsx", header, rows,
                                     title=f"Budget vs Actual — {data['label']}",
                                     church=SiteConfig.get().church_name)
            return csv_response(fname + ".csv", header, rows)
        ctx = self.get_context_data(**kwargs)
        ctx.update({"d": data, "year": year, "period": period,
                    "month": int(month) if month else today.month,
                    "quarter": int(quarter) if quarter else ((today.month - 1) // 3 + 1),
                    "years": range(today.year + 1, today.year - 5, -1),
                    "months": [(m, _dt.date(2000, m, 1).strftime("%B")) for m in range(1, 13)]})
        return self.render_to_response(ctx)


# ===================== Reporting suite (new reports) =====================
from .exports import xlsx_response
from core.utils import sabbath_of
from core.models import SiteConfig


def _export(request, filename, header, rows, title):
    fmt = request.GET.get("export")
    if fmt == "csv":
        return csv_response(filename + ".csv", header, rows)
    if fmt == "xlsx":
        return xlsx_response(filename + ".xlsx", header, rows, title=title,
                             church=SiteConfig.get().church_name)
    return None


def _day_income_expense(start, end):
    inc = (Transaction.objects.filter(direction=Transaction.Direction.CREDIT, is_reversal=False,
           date__gte=start, date__lte=end).values("department__name")
           .annotate(t=Sum("amount")).order_by("-t"))
    eff = Q(status__in=[Expense.Status.APPROVED, Expense.Status.PAID])
    exp = (Expense.objects.filter(eff, date__gte=start, date__lte=end)
           .values("department__name").annotate(t=Sum("amount")).order_by("-t"))
    return inc, exp


class DailySummaryView(ReportAccessMixin, TemplateView):
    template_name = "reports/daily_summary.html"

    def get(self, request, *args, **kwargs):
        try:
            day = _dt.date.fromisoformat(request.GET.get("date", ""))
        except ValueError:
            day = _dt.date.today()
        inc, exp = _day_income_expense(day, day)
        inc = list(inc); exp = list(exp)
        ti = sum((r["t"] or Decimal(0) for r in inc), Decimal(0))
        te = sum((r["t"] or Decimal(0) for r in exp), Decimal(0))
        ex = _export(request, f"daily_{day}",
                     ["Type", "Fund", "Amount"],
                     [["Income", r["department__name"] or "—", r["t"]] for r in inc]
                     + [["Expense", r["department__name"] or "—", r["t"]] for r in exp],
                     f"Daily summary {day:%d %b %Y}")
        if ex:
            return ex
        ctx = self.get_context_data(**kwargs)
        ctx.update({"day": day, "income": inc, "expense": exp,
                    "total_income": ti, "total_expense": te, "net": ti - te})
        return self.render_to_response(ctx)


class WeeklySummaryView(ReportAccessMixin, TemplateView):
    template_name = "reports/weekly_summary.html"

    def get(self, request, *args, **kwargs):
        try:
            anchor = _dt.date.fromisoformat(request.GET.get("date", ""))
        except ValueError:
            anchor = _dt.date.today()
        sab = sabbath_of(anchor)               # the Sabbath of that week
        start = sab - _dt.timedelta(days=6)     # Sun..Sat window
        inc, exp = _day_income_expense(start, sab)
        inc = list(inc); exp = list(exp)
        ti = sum((r["t"] or Decimal(0) for r in inc), Decimal(0))
        te = sum((r["t"] or Decimal(0) for r in exp), Decimal(0))
        ex = _export(request, f"week_to_{sab}",
                     ["Type", "Fund", "Amount"],
                     [["Income", r["department__name"] or "—", r["t"]] for r in inc]
                     + [["Expense", r["department__name"] or "—", r["t"]] for r in exp],
                     f"Week to Sabbath {sab:%d %b %Y}")
        if ex:
            return ex
        ctx = self.get_context_data(**kwargs)
        ctx.update({"sabbath": sab, "start": start, "income": inc, "expense": exp,
                    "total_income": ti, "total_expense": te, "net": ti - te})
        return self.render_to_response(ctx)


class CashFlowView(ReportAccessMixin, TemplateView):
    template_name = "reports/cash_flow.html"

    def get(self, request, *args, **kwargs):
        try:
            year = int(request.GET.get("year", _dt.date.today().year))
        except (TypeError, ValueError):
            year = _dt.date.today().year
        eff = Q(status__in=[Expense.Status.APPROVED, Expense.Status.PAID])
        rows, running = [], Decimal(0)
        for m in range(1, 13):
            import calendar as _cal
            last = _cal.monthrange(year, m)[1]
            s, e = _dt.date(year, m, 1), _dt.date(year, m, last)
            inflow = (Transaction.objects.confirmed_credits()
                      .filter(date__gte=s, date__lte=e, excluded_from_income=False)
                      .aggregate(t=Sum("amount"))["t"] or Decimal(0))
            outflow = (Expense.objects.filter(eff, date__gte=s, date__lte=e)
                       .aggregate(t=Sum("amount"))["t"] or Decimal(0))
            net = inflow - outflow
            running += net
            rows.append({"month": _dt.date(year, m, 1).strftime("%B"),
                         "inflow": inflow, "outflow": outflow, "net": net, "running": running})
        ex = _export(request, f"cash_flow_{year}",
                     ["Month", "Inflow", "Outflow", "Net", "Running"],
                     [[r["month"], r["inflow"], r["outflow"], r["net"], r["running"]] for r in rows],
                     f"Cash flow {year}")
        if ex:
            return ex
        ctx = self.get_context_data(**kwargs)
        ctx.update({"year": year, "rows": rows,
                    "years": range(_dt.date.today().year, _dt.date.today().year - 6, -1),
                    "tot_in": sum((r["inflow"] for r in rows), Decimal(0)),
                    "tot_out": sum((r["outflow"] for r in rows), Decimal(0))})
        return self.render_to_response(ctx)


class BoardReportSettingsView(TreasurerRequiredMixin, View):
    """Configure which board-report sections appear, their order, and the notes
    shown on the report."""
    template_name = "reports/board_settings.html"

    def get(self, request):
        cfg = SiteConfig.get()
        return render(request, self.template_name, {
            "sections": cfg.board_settings()["sections"],
            "notes": cfg.board_settings()["notes"]})

    def post(self, request):
        cfg = SiteConfig.get()
        # ordering arrives as a list of keys; visibility as checkboxes
        order = request.POST.getlist("order")
        valid = dict(SiteConfig.BOARD_SECTIONS)
        sections = []
        for key in order:
            if key in valid:
                sections.append({"key": key,
                                 "visible": bool(request.POST.get(f"visible_{key}"))})
        cfg.board_config = {"sections": sections,
                            "notes": (request.POST.get("notes") or "")[:4000]}
        cfg.save(update_fields=["board_config"])
        messages.success(request, "Board report settings saved.")
        return redirect("board_settings")


class BoardReportView(PeriodMixin, TemplateView):
    """One-page board summary: position, fund groups, trust, budget, KPIs, plus an
    AI-written narrative (insights, trends, recommendations) with a rule-based
    fallback when the assistant is off or unavailable."""
    template_name = "reports/board_report.html"

    def get_context_data(self, **kwargs):
        from core.services.assistant import board_report_narrative
        from cashbook.views import open_payables_total, open_accruals_total
        from django.db.models.functions import ExtractYear
        ctx = super().get_context_data(**kwargs)
        s, e = ctx["start"], ctx["end"]
        cfg = SiteConfig.get()
        paid = [Expense.Status.APPROVED, Expense.Status.PAID]
        rows = balances.department_summary(s, e)
        ctx["totals"] = balances.totals(rows)
        ctx["trust_rows"] = [r for r in rows if r["is_trust"]]
        ctx["local_rows"] = [r for r in rows if not r["is_trust"]]
        ctx["trust_summary"] = balances.trust_summary(s, e)
        ctx["trust_outstanding"] = sum((r["to_remit"] for r in ctx["trust_summary"]), Decimal(0))
        ctx["trust_unreceipted"] = sum((r["unreceipted"] for r in ctx["trust_summary"]), Decimal(0))
        ctx["by_channel"] = balances.income_by_channel(s, e)
        ctx["church"] = cfg.church_name

        # ---- Income & Expenditure statement ----
        income = (Transaction.objects.confirmed_credits().filter(
            date__gte=s, date__lte=e, department__is_trust=False,
            excluded_from_income=False).aggregate(t=Sum("amount"))["t"] or Decimal(0))
        expense = (Expense.objects.filter(date__gte=s, date__lte=e, status__in=paid)
                   .exclude(category__in=[Expense.Category.REMITTANCE, Expense.Category.LOAN_REPAYMENT])
                   .exclude(expenditure_type=Expense.ExpenditureType.CAPITAL)
                   .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        capital = (Expense.objects.filter(date__gte=s, date__lte=e, status__in=paid,
                   expenditure_type=Expense.ExpenditureType.CAPITAL)
                   .exclude(category__in=[Expense.Category.REMITTANCE, Expense.Category.LOAN_REPAYMENT])
                   .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        ie_cats = (Expense.objects.filter(date__gte=s, date__lte=e, status__in=paid)
                   .exclude(category__in=[Expense.Category.REMITTANCE, Expense.Category.LOAN_REPAYMENT])
                   .exclude(expenditure_type=Expense.ExpenditureType.CAPITAL)
                   .values("category").annotate(t=Sum("amount")).order_by("-t"))
        cat_label = dict(Expense.Category.choices)
        ctx["ie_income"] = income
        ctx["ie_expense"] = expense
        ctx["ie_surplus"] = income - expense
        ctx["ie_capital"] = capital
        ctx["ie_expense_by_cat"] = [{"label": cat_label.get(r["category"], r["category"]),
                                     "total": r["t"]} for r in ie_cats]

        # ---- Statement of financial position (period end) ----
        income_all = (Transaction.objects.confirmed_credits()
                      .filter(excluded_from_income=False, date__lte=e)
                      .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        payments_all = (Expense.objects.filter(status__in=paid, date__lte=e)
                        .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        opening = (cfg.opening_bank_balance + cfg.opening_cash_on_hand
                   - cfg.opening_unremitted_trust)
        cash_bank = opening + income_all - payments_all
        payables = open_payables_total()
        accruals = open_accruals_total()
        asset_nbv = Decimal(0)
        try:
            from assets.models import FixedAsset
            asset_nbv = sum((a.net_book_value(e) for a in FixedAsset.objects.filter(disposed=False)),
                            Decimal(0))
        except Exception:
            from core.utils import log_exception as _lx; _lx('reports/views.py')
            asset_nbv = Decimal(0)
        trust_liab = ctx["trust_outstanding"]
        total_assets = cash_bank + asset_nbv
        total_liab = trust_liab + payables + accruals
        ctx["sofp"] = {
            "cash_bank": cash_bank, "asset_nbv": asset_nbv, "total_assets": total_assets,
            "trust_liab": trust_liab, "payables": payables, "accruals": accruals,
            "total_liab": total_liab, "net_assets": total_assets - total_liab,
        }

        # ---- Multi-year trend (like-for-like, year-to-date) ----
        # Compare each year only up to the SAME point in the year as the current
        # report period reaches, so a partial current year isn't unfairly shown
        # against full prior years. We cap each year's figures at the month/day
        # the current period ends on.
        cutoff_month, cutoff_day = e.month, e.day
        ytd_label = f" (Jan–{e:%b})" if (cutoff_month, cutoff_day) != (12, 31) else ""

        def _ytd_filter(qs):
            # keep rows whose (month, day) falls on/before the current cutoff
            from django.db.models.functions import ExtractMonth, ExtractDay
            return (qs.annotate(_m=ExtractMonth("date"), _dd=ExtractDay("date"))
                    .filter(Q(_m__lt=cutoff_month)
                            | Q(_m=cutoff_month, _dd__lte=cutoff_day)))

        inc_y = {r["yr"]: r["total"] for r in (_ytd_filter(
                 Transaction.objects.confirmed_credits()
                 .filter(excluded_from_income=False))
                 .annotate(yr=ExtractYear("date"))
                 .values("yr").annotate(total=Sum("amount")))}
        exp_y = {r["yr"]: r["total"] for r in (_ytd_filter(
                 Expense.objects.filter(status__in=paid)
                 .exclude(category__in=[Expense.Category.REMITTANCE, Expense.Category.LOAN_REPAYMENT]))
                 .annotate(yr=ExtractYear("date")).values("yr")
                 .annotate(total=Sum("amount")))}
        from core.models import HistoricalYear
        hist = {h.year: h for h in HistoricalYear.objects.all()}
        years = sorted(set(inc_y) | set(exp_y) | set(hist))
        ctx["trend_ytd_label"] = ytd_label
        trend = []
        for y in years:
            coll = inc_y.get(y) or (hist[y].collection if y in hist else 0)
            ex = exp_y.get(y) or (hist[y].expenditure if y in hist else 0)
            trend.append({"year": y, "income": coll, "expense": ex,
                          "net": (coll or 0) - (ex or 0)})
        ctx["trend"] = trend

        # ---- narrative (AI with deterministic fallback) ----
        label = f"{s:%d %b %Y} – {e:%d %b %Y}"
        ctx["board_income"] = income
        ctx["board_expenditure"] = expense
        ctx["board_surplus"] = income - expense
        context_str = self._context_str(ctx, label, income, expense)
        narrative, source, err = None, "fallback", None
        if cfg.llm_enabled and self.request.GET.get("ai") != "0":
            narrative, err = board_report_narrative(context_str, label, cfg)
            if narrative:
                source = "ai"
        if not narrative:
            narrative = self._fallback(ctx, label, income, expense)
        ctx["narrative"] = narrative
        ctx["narrative_source"] = source
        ctx["ai_enabled"] = cfg.llm_enabled
        ctx["ai_error"] = err

        # ---- Goals & targets (#3): expense, offering, group contribution ----
        from departments.models import Department as _Dept
        gyear = e.year

        def _fund_ids(d):
            ids = [d.id]
            for sub in d.subgroups.all():
                ids.extend(_fund_ids(sub))
            return ids

        def _collected(fund):
            return (Transaction.objects.confirmed_credits().filter(
                department_id__in=_fund_ids(fund), excluded_from_income=False,
                date__year=gyear).aggregate(t=Sum("amount"))["t"] or Decimal(0))

        def _goal_row(name, kind, goal, collected):
            goal = goal or Decimal(0)
            return {"name": name, "kind": kind, "goal": goal, "collected": collected,
                    "variance": collected - goal,
                    "pct": int(min(collected / goal * 100, 999)) if goal else 0,
                    "short": max(goal - collected, Decimal(0))}

        goals = []
        # church-wide Camp Meeting goals come from Settings → Goals, not from
        # any individual fund, so fund rows below stay purely per-fund
        goals.extend(_camp_goal_records(gyear))
        for d in _Dept.objects.filter(active=True).prefetch_related("subgroups"):
            if d.goal_type == "CAMP_EXPENSE":
                continue  # migrated to SiteConfig; avoid double rows
            if d.year_goal:
                goals.append(_goal_row(f"{d.name} — annual goal", "Expense (local)",
                                       d.year_goal, _collected(d)))
            if d.offering_goal and d.offering_fund:
                goals.append(_goal_row(f"{d.offering_fund.name} — offering goal",
                                       "Offering (trust)",
                                       d.offering_goal, _collected(d.offering_fund)))
            if d.contribution_goal:
                grp_total = sum((_collected(s) for s in d.subgroups.all()), Decimal(0))
                goals.append(_goal_row(f"{d.name} — group contribution goal",
                                       "Contribution", d.contribution_goal, grp_total))
        ctx["goals"] = goals
        ctx["goals_year"] = gyear

        ctx["board_sections"] = cfg.board_settings()["sections"]
        ctx["board_notes"] = cfg.board_settings()["notes"]
        ctx["bvis"] = {s["key"]: s["visible"] for s in ctx["board_sections"]}
        ctx["border_order"] = [s["key"] for s in ctx["board_sections"] if s["visible"]]
        return ctx

    def _context_str(self, ctx, label, income, expenditure):
        def f(v):
            return f"{float(v or 0):,.0f}"
        top = sorted(ctx["local_rows"], key=lambda r: r["receipts"], reverse=True)[:6]
        lines = [
            f"Period: {label}",
            f"Local income (I&E): {f(income)}",
            f"Local expenditure (I&E): {f(expenditure)}",
            f"Net surplus/(deficit): {f(income - expenditure)}",
            f"Cash & bank at period end: {f(ctx['sofp']['cash_bank'])}",
            f"Net assets: {f(ctx['sofp']['net_assets'])}",
            f"Trust outstanding to remit: {f(ctx['trust_outstanding'])}",
            "Top funds by receipts: " + "; ".join(
                f"{_sfund(r['department'].name)} {f(r['receipts'])}" for r in top if r["receipts"]),
        ]
        if len(ctx["trend"]) > 1:
            lines.append("Year trend (income): " + "; ".join(
                f"{t['year']}: {f(t['income'])}" for t in ctx["trend"][-4:]))
        deficits = [r for r in ctx["local_rows"] if r["closing"] < 0]
        if deficits:
            lines.append("Funds in deficit: " + "; ".join(
                f"{_sfund(r['department'].name)} {f(r['closing'])}" for r in deficits[:6]))
        return "\n".join(lines)

    def _fallback(self, ctx, label, income, expenditure):
        def f(v):
            return f"{float(v or 0):,.0f}"
        surplus = income - expenditure
        top = [r for r in sorted(ctx["local_rows"], key=lambda r: r["receipts"],
                                 reverse=True) if r["receipts"]][:3]
        deficits = [r for r in ctx["local_rows"] if r["closing"] < 0]
        p = ["Executive summary:",
             f"For {label}, collections were {f(income)} against expenditure of "
             f"{f(expenditure)}, a {'surplus' if surplus >= 0 else 'deficit'} of "
             f"{f(abs(surplus))}.",
             "\nKey insights:"]
        if top:
            p.append("- Largest funds: " + ", ".join(
                f"{_sfund(r['department'].name)} ({f(r['receipts'])})" for r in top) + ".")
        p.append(f"- Trust outstanding to remit: {f(ctx['trust_outstanding'])}.")
        if deficits:
            p.append("- Funds in deficit: " + ", ".join(
                r["department"].name for r in deficits[:5]) + ".")
        p.append("\nRecommendations:")
        if ctx["trust_outstanding"] > 0:
            p.append(f"- Remit the outstanding trust of {f(ctx['trust_outstanding'])}.")
        if deficits:
            p.append("- Review and rebalance funds carrying a negative balance.")
        p.append("- Confirm the bank reconciliation and file supporting vouchers.")
        return "\n".join(p)


class PastorReportView(PeriodMixin, TemplateView):
    """Pastoral summary: tithe, offerings, giving by group, participation."""
    template_name = "reports/pastor_report.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        s, e = ctx["start"], ctx["end"]
        ctx["tithe"] = balances.tithe_total(s, e)
        ctx["by_group"] = balances.giving_by_group(s, e)
        ctx["by_channel"] = balances.income_by_channel(s, e)
        rows = balances.department_summary(s, e)
        ctx["total_income"] = balances.totals(rows)
        from members.models import Member
        ctx["active_members"] = Member.objects.filter(active=True).count()
        givers = (Transaction.objects.confirmed_credits()
                  .filter(date__gte=s, date__lte=e, member__isnull=False)
                  .values("member").distinct().count())
        ctx["givers"] = givers
        ctx["church"] = SiteConfig.get().church_name
        return ctx


class ConferenceSubmissionView(PeriodMixin, TemplateView):
    """Conference submission: trust collected / remitted / to-remit + batches."""
    template_name = "reports/conference_submission.html"

    def get(self, request, *args, **kwargs):
        ctx = self.get_context_data(**kwargs)
        s, e = ctx["start"], ctx["end"]
        rows = balances.trust_summary(s, e)
        cfg = SiteConfig.get()
        if request.GET.get("export") in ("csv", "xlsx"):
            header = ["Trust fund", "Collected", "Remitted",
                      "Outstanding to remit (receipted)", "Unreceipted (pending)",
                      "Total trust liability"]
            data = [[r["department"].name, r["collected"], r["remitted"], r["to_remit"],
                     r["unreceipted"], r["total_liability"]] for r in rows]
            data.append(["TOTAL",
                         sum((r["collected"] for r in rows), Decimal(0)),
                         sum((r["remitted"] for r in rows), Decimal(0)),
                         sum((r["to_remit"] for r in rows), Decimal(0)),
                         sum((r["unreceipted"] for r in rows), Decimal(0)),
                         sum((r["total_liability"] for r in rows), Decimal(0))])
            ex = _export(request, f"conference_{s}_{e}", header, data, "Conference submission")
            if ex:
                return ex
        ctx["rows"] = rows
        ctx["totals"] = {
            "collected": sum((r["collected"] for r in rows), Decimal(0)),
            "remitted": sum((r["remitted"] for r in rows), Decimal(0)),
            "to_remit": sum((r["to_remit"] for r in rows), Decimal(0)),
            "unreceipted": sum((r["unreceipted"] for r in rows), Decimal(0)),
            "total_liability": sum((r["total_liability"] for r in rows), Decimal(0)),
        }
        ctx["field_name"] = cfg.field_name
        ctx["church"] = cfg.church_name
        return self.render_to_response(ctx)


# ===================== Financial statements =====================
class IncomeStatementView(PeriodMixin, TemplateView):
    """Statement of income & expenditure on a LOCAL (operating) basis: trust
    collections and their remittances are excluded, since trust money is held on
    behalf of the field rather than being the church's own income."""
    template_name = "reports/income_statement.html"

    def get(self, request, *args, **kwargs):
        ctx = self.get_context_data(**kwargs)
        s, e = ctx["start"], ctx["end"]
        rows = balances.department_summary(s, e)
        income = [{"name": r["department"].name, "amount": r["receipts"]}
                  for r in rows if not r["is_trust"] and r["receipts"]]
        income.sort(key=lambda x: -x["amount"])
        total_income = sum((r["amount"] for r in income), Decimal(0))
        trust_collected = sum((r["receipts"] for r in rows if r["is_trust"]), Decimal(0))
        # expenditure by category, excluding trust remittances, split recurrent/capital
        eff = Q(status__in=[Expense.Status.APPROVED, Expense.Status.PAID],
                date__gte=s, date__lte=e)
        cats = dict(Expense.Category.choices)
        base = Expense.objects.filter(eff).exclude(category__in=[Expense.Category.REMITTANCE, Expense.Category.LOAN_REPAYMENT])

        def _by_cat(qs):
            rows = [{"name": cats.get(r["category"], r["category"]), "amount": r["t"]}
                    for r in qs.values("category").annotate(t=Sum("amount")).order_by("-t")]
            return rows, sum((r["amount"] for r in rows), Decimal(0))

        recurrent, total_recurrent = _by_cat(
            base.filter(expenditure_type=Expense.ExpenditureType.RECURRENT))
        capital, total_capital = _by_cat(
            base.filter(expenditure_type=Expense.ExpenditureType.CAPITAL))
        total_exp = total_recurrent + total_capital
        operating = total_income - total_recurrent
        surplus = operating - total_capital
        # Change in net assets (fund basis) — ties the result back to the funds and
        # to the Statement of Financial Position.
        na_open = sum((r["opening"] for r in rows if not r["is_trust"]), Decimal(0))
        net_transfers = sum((r["net_transfer"] for r in rows if not r["is_trust"]), Decimal(0))
        na_close = na_open + surplus + net_transfers
        if request.GET.get("export") in ("csv", "xlsx"):
            header = ["Section", "Line", "Amount"]
            data = [["Revenue", r["name"], r["amount"]] for r in income]
            data.append(["Revenue", "TOTAL REVENUE", total_income])
            data += [["Operating (recurrent) expenditure", r["name"], r["amount"]] for r in recurrent]
            data.append(["Operating (recurrent) expenditure", "TOTAL RECURRENT", total_recurrent])
            data.append(["Result", "OPERATING SURPLUS/(DEFICIT)", operating])
            data += [["Capital expenditure", r["name"], r["amount"]] for r in capital]
            data.append(["Capital expenditure", "TOTAL CAPITAL", total_capital])
            data.append(["Result", "NET SURPLUS/(DEFICIT)", surplus])
            data += [
                ["Net assets", "Net assets brought forward", na_open],
                ["Net assets", "Net surplus/(deficit) for the period", surplus],
                ["Net assets", "Net inter-fund transfers", net_transfers],
                ["Net assets", "NET ASSETS CARRIED FORWARD", na_close],
            ]
            ex = _export(request, f"income_statement_{s}_{e}", header, data,
                         "Statement of Financial Activity")
            if ex:
                return ex
        ctx.update({"income": income, "total_income": total_income,
                    "recurrent": recurrent, "total_recurrent": total_recurrent,
                    "capital": capital, "total_capital": total_capital,
                    "operating": operating, "total_exp": total_exp,
                    "surplus": surplus, "trust_collected": trust_collected,
                    "na_open": na_open, "net_transfers": net_transfers,
                    "na_close": na_close,
                    "church": SiteConfig.get().church_name})
        return self.render_to_response(ctx)


class FinancialPositionView(ReportAccessMixin, TemplateView):
    """Statement of Financial Position (balance sheet) on a fund-accounting basis,
    as at a date. Assets = cash/bank (fund balances) + fixed assets (NBV);
    financed by trust funds payable, accumulated local funds, and a capital fund
    matching the carrying value of fixed assets."""
    template_name = "reports/financial_position.html"

    def get(self, request, *args, **kwargs):
        try:
            as_of = _dt.date.fromisoformat(request.GET.get("as_of", ""))
        except ValueError:
            as_of = _dt.date.today()
        rows = balances.department_summary(None, as_of)
        cash = sum((r["closing"] for r in rows), Decimal(0))
        trust_payable = sum((r["closing"] for r in rows if r["is_trust"]), Decimal(0))
        # Split the trust liability into RECEIPTED (firmly due to remit) vs
        # not-yet-receipted (trust money allocated to a trust fund but without a
        # formal receipt). The receipted figure comes from the trust summary
        # (opening + receipted − remitted); the remainder of the closing balance
        # is the unreceipted portion, so the two always sum to the trust payable
        # that ties the balance sheet. Bank money not yet allocated to any fund is
        # a DIFFERENT thing (suspense) and is shown on its own line below.
        _tsum = balances.trust_summary(None, as_of)
        trust_receipted = sum((r["to_remit"] for r in _tsum), Decimal(0))
        trust_unreceipted = trust_payable - trust_receipted
        local_rows = [r for r in rows if not r["is_trust"]]
        local_funds = sum((r["closing"] for r in local_rows), Decimal(0))
        from assets.models import nbv_total
        nbv = nbv_total(as_of)
        # Net assets, classified per the SDA framework: Board-designated funds
        # (development/projects) are "Allocated"; the rest are "Unallocated"; the
        # carrying value of property is held as "Invested in property".
        allocated = sum((r["closing"] for r in local_rows
                         if r["department"].category == "DEVELOPMENT"), Decimal(0))
        unallocated = local_funds - allocated
        # Accrual overlay (memoranda): credit purchases owed, expenses accrued, and
        # amounts prepaid. These adjust the cash-basis position to an accrual view.
        from cashbook.views import (open_payables_total, open_accruals_total,
                                     unexpired_prepayments_total,
                                     outstanding_advances_total)
        payables = open_payables_total(as_of)
        accruals = open_accruals_total(as_of)
        prepaid = unexpired_prepayments_total(as_of)
        # An unspent staff advance is cash that has physically left but not yet been
        # expensed — a receivable. Reclassify it out of cash so each is shown
        # correctly; totals are unchanged (it is still inside the cash figure).
        advances = outstanding_advances_total(as_of)
        cash_on_hand = cash - advances
        # Bank money received but not yet receipted/allocated to a fund — shown as
        # cash held in suspense with a matching "pending allocation" liability, so
        # the statement ties to the bank and the money is never invisible.
        pending = balances.pending_receipts_total(as_of)
        # Loans payable: the outstanding loan principal is a real liability, split
        # into current (≤12 months / on demand) and long-term. The loan cash is
        # already inside the `cash` asset figure (a loan receipt raises the fund's
        # balance), so recognising the matching liability here is exactly what
        # keeps the statement in balance once loans exist. This total ties to the
        # LOANS_PAYABLE ledger account by construction.
        from loans.services import reporting as loan_rep
        loan_liab = loan_rep.outstanding_liability(as_of)
        loans_current = loan_liab["current"]
        loans_long_term = loan_liab["long_term"]
        loans_payable = loan_liab["total"]
        accrual_adj = prepaid - payables - accruals
        # Loans payable is a liability the church must settle from its own cash,
        # so it reduces net assets (unlike trust payable, which sits against
        # trust cash held on the field's behalf). Deducting it here keeps
        # Assets = Liabilities + Net assets true.
        net_assets = unallocated + allocated + nbv + accrual_adj - loans_payable
        total_assets = cash_on_hand + advances + pending + nbv + prepaid
        total_liabilities = (trust_payable + payables + accruals + pending
                             + loans_payable)
        total_liab_and_na = total_liabilities + net_assets
        # committed-but-unpaid vouchers (memorandum)
        unpaid = (Expense.objects.filter(status__in=[Expense.Status.PENDING,
                  Expense.Status.APPROVED], date__lte=as_of)
                  .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        ctx = self.get_context_data(**kwargs)
        _cfg = SiteConfig.get()
        if request.GET.get("export") in ("csv", "xlsx"):
            header = ["Section", "Line", "Amount"]
            data = [
                ["Assets", "Cash & bank (current)", cash_on_hand],
                ["Assets", "Staff advances (receivable)", advances],
                ["Assets", "Bank receipts pending allocation", pending],
                ["Assets", "Prepayments", prepaid],
                ["Assets", "Property, plant & equipment (non-current)", nbv],
                ["Assets", "TOTAL ASSETS", total_assets],
                ["Liabilities", "Trust funds payable to the field", trust_payable],
                ["Liabilities", "Payables", payables],
                ["Liabilities", "Accruals", accruals],
                ["Liabilities", "Receipts pending allocation", pending],
                ["Liabilities", "Loans payable — current", loans_current],
                ["Liabilities", "Loans payable — long-term", loans_long_term],
                ["Liabilities", "TOTAL LIABILITIES", total_liabilities],
                ["Net assets", "General net assets", unallocated],
                ["Net assets", "Designated development funds", allocated],
                ["Net assets", "Invested in property", nbv],
                ["Net assets", "Accrual adjustment", accrual_adj],
                ["Net assets", "TOTAL NET ASSETS", net_assets],
                ["", "TOTAL LIABILITIES & NET ASSETS", total_liab_and_na],
            ]
            ex = _export(request, f"financial_position_{as_of}", header, data,
                         "Statement of Financial Position")
            if ex:
                return ex
        ctx.update({"as_of": as_of, "cash": cash, "nbv": nbv,
                    "cash_on_hand": cash_on_hand, "advances": advances,
                    "pending": pending,
                    "trust_payable": trust_payable, "local_funds": local_funds,
                    "trust_receipted": trust_receipted,
                    "trust_unreceipted": trust_unreceipted,
                    "trust_total_payable": trust_payable,
                    "unallocated": unallocated, "allocated": allocated,
                    "net_assets": net_assets, "total_assets": total_assets,
                    "total_liab_and_na": total_liab_and_na,
                    "loans_payable": loans_payable, "loans_current": loans_current,
                    "loans_long_term": loans_long_term,
                    "payables": payables, "accruals": accruals, "prepaid": prepaid,
                    "accrual_adj": accrual_adj, "total_liabilities": total_liabilities,
                    "balanced": total_assets == total_liab_and_na,
                    "unpaid": unpaid, "trust_rows": [r for r in rows if r["is_trust"]],
                    "local_rows": [r for r in local_rows if r["closing"]],
                    "opening_bank": _cfg.opening_bank_balance,
                    "opening_cash_on_hand": _cfg.opening_cash_on_hand,
                    "opening_unremitted_trust": _cfg.opening_unremitted_trust,
                    "opening_total": (_cfg.opening_bank_balance
                                      + _cfg.opening_cash_on_hand
                                      - _cfg.opening_unremitted_trust),
                    "church": SiteConfig.get().church_name})
        return self.render_to_response(ctx)

class ChangesInNetAssetsView(PeriodMixin, TemplateView):
    """Statement of Changes in Net Assets — how each class of net assets moved over
    the period: opening + surplus/(deficit) +/- capital reclassification +/-
    transfers = closing. Classified into Unallocated, Allocated (Board-designated)
    and Invested in property, and ties to the Statement of Financial Position."""
    template_name = "reports/changes_in_net_assets.html"

    def get(self, request, *args, **kwargs):
        ctx = self.get_context_data(**kwargs)
        s, e = ctx["start"], ctx["end"]
        rows = balances.department_summary(s, e)
        local = [r for r in rows if not r["is_trust"]]

        def is_alloc(r):
            return r["department"].category == "DEVELOPMENT"

        eff = Q(status__in=[Expense.Status.APPROVED, Expense.Status.PAID],
                date__gte=s, date__lte=e)

        def cap_for(alloc):
            qs = (Expense.objects.filter(eff, expenditure_type=Expense.ExpenditureType.CAPITAL,
                  department__fund_type="LOCAL").exclude(category__in=[Expense.Category.REMITTANCE, Expense.Category.LOAN_REPAYMENT]))
            tot = Decimal(0)
            for x in qs.select_related("department"):
                if (x.department.category == "DEVELOPMENT") == alloc:
                    tot += x.amount
            return tot

        def col(alloc):
            sub = [r for r in local if is_alloc(r) == alloc]
            opening = sum((r["opening"] for r in sub), Decimal(0))
            receipts = sum((r["receipts"] for r in sub), Decimal(0))
            expenses = sum((r["expenses"] for r in sub), Decimal(0))
            transfers = sum((r["net_transfer"] for r in sub), Decimal(0))
            closing = sum((r["closing"] for r in sub), Decimal(0))
            capital = cap_for(alloc)
            op_surplus = receipts - (expenses - capital)   # surplus before capital
            return {"opening": opening, "op_surplus": op_surplus, "capital": capital,
                    "transfers": transfers, "closing": closing}

        un, al = col(False), col(True)
        cap_total = un["capital"] + al["capital"]
        day_before = s - _dt.timedelta(days=1)
        from assets.models import nbv_total
        nbv_open = nbv_total(day_before)
        nbv_close = nbv_total(e)
        depr = nbv_close - nbv_open - cap_total      # balancing: depreciation + disposals
        prop = {"opening": nbv_open, "additions": cap_total, "depr": depr, "closing": nbv_close}

        t_open = un["opening"] + al["opening"] + nbv_open
        t_opsurplus = un["op_surplus"] + al["op_surplus"]
        t_transfers = un["transfers"] + al["transfers"]
        t_close = un["closing"] + al["closing"] + nbv_close

        if request.GET.get("export") in ("csv", "xlsx"):
            header = ["Line", "General net assets", "Designated development funds", "Invested in property", "Total"]
            data = [
                ["Net assets, beginning", un["opening"], al["opening"], nbv_open, t_open],
                ["Surplus/(deficit) from operations", un["op_surplus"], al["op_surplus"], 0, t_opsurplus],
                ["Capital expenditure (property acquired)", -un["capital"], -al["capital"], cap_total, 0],
                ["Depreciation & disposals", 0, 0, depr, depr],
                ["Net inter-fund transfers", un["transfers"], al["transfers"], 0, t_transfers],
                ["Net assets, end", un["closing"], al["closing"], nbv_close, t_close],
            ]
            ex = _export(request, f"changes_in_net_assets_{s}_{e}", header, data,
                         "Statement of Changes in Net Assets")
            if ex:
                return ex
        ctx.update({"un": un, "al": al, "prop": prop, "cap_total": cap_total,
                    "t_open": t_open, "t_opsurplus": t_opsurplus, "depr": depr,
                    "t_transfers": t_transfers, "t_close": t_close,
                    "church": SiteConfig.get().church_name})
        return self.render_to_response(ctx)


class StatementOfCashFlowsView(PeriodMixin, TemplateView):
    """Statement of Cash Flows on the SDA three-category basis (operating, investing,
    financing). Reconciles the movement in total cash & bank over the period."""
    template_name = "reports/cash_flows.html"

    def get(self, request, *args, **kwargs):
        ctx = self.get_context_data(**kwargs)
        s, e = ctx["start"], ctx["end"]
        rows = balances.department_summary(s, e)
        cash_open = sum((r["opening"] for r in rows), Decimal(0))
        cash_close = sum((r["closing"] for r in rows), Decimal(0))
        local_receipts = sum((r["receipts"] for r in rows if not r["is_trust"]), Decimal(0))
        trust_receipts = sum((r["receipts"] for r in rows if r["is_trust"]), Decimal(0))
        eff = Q(status__in=[Expense.Status.APPROVED, Expense.Status.PAID],
                date__gte=s, date__lte=e)

        def _sum(qs):
            return qs.aggregate(t=Sum("amount"))["t"] or Decimal(0)

        remittances = _sum(Expense.objects.filter(eff, category=Expense.Category.REMITTANCE))
        nonremit = Expense.objects.filter(eff).exclude(category__in=[Expense.Category.REMITTANCE, Expense.Category.LOAN_REPAYMENT])
        total_nonremit = _sum(nonremit)
        capital = _sum(nonremit.filter(expenditure_type=Expense.ExpenditureType.CAPITAL))
        # everything non-remittance that isn't explicitly capital is operating —
        # this way the three buckets always sum to total expenses, so the
        # statement reconciles even if some rows have no expenditure type set.
        operating_exp = total_nonremit - capital

        # Financing activities: loan receipts (cash in) and principal repayments
        # (cash out) belong here, never in operating.
        #  * `local_receipts` (from department_summary/receipts_by_department)
        #    INCLUDES loan receipts as fund cash, so they must be SUBTRACTED
        #    out of operating and shown in financing instead (not added twice).
        #  * principal repayments were excluded from `nonremit`, so they never
        #    hit operating expenses; they reduce cash only here in financing.
        #  * interest paid stays inside operating expenses per system policy
        #    (an ordinary voucher on the fund) — no adjustment needed.
        from loans.services import reporting as loan_rep
        fin = loan_rep.financing_activity(s, e)
        loan_receipts = fin["receipts"]
        loan_repayments = fin["repayments"]
        # Loan conversions / write-offs recognise income with NO cash movement
        # (a liability is reclassified to income). That income leg is a normal,
        # non-excluded contribution credit, so it is inside `local_receipts` —
        # remove it here so operating cash receipts reflect only real cash in.
        # (Its contra LOAN_REPAYMENT leg was already excluded from operating
        # expenses, so removing this keeps the statement reconciling.)
        loan_noncash_income = loan_rep.retirement_income(s, e)
        local_operating_receipts = local_receipts - loan_receipts - loan_noncash_income

        net_operating = local_operating_receipts + trust_receipts - operating_exp - remittances
        net_investing = -capital
        net_financing = loan_receipts - loan_repayments
        net_change = net_operating + net_investing + net_financing

        if request.GET.get("export") in ("csv", "xlsx"):
            header = ["Section", "Line", "Amount"]
            data = [
                ["Operating", "Local offerings & income received", local_operating_receipts],
                ["Operating", "Tithe & trust offerings received (held for the field)", trust_receipts],
                ["Operating", "Operating (recurrent) expenses paid", -operating_exp],
                ["Operating", "Remittances to the field paid", -remittances],
                ["Operating", "Net cash from operating activities", net_operating],
                ["Investing", "Purchase of property & equipment", -capital],
                ["Investing", "Net cash used in investing activities", net_investing],
                ["Financing", "Loan receipts (borrowings)", loan_receipts],
                ["Financing", "Loan principal repayments", -loan_repayments],
                ["Financing", "Net cash from financing activities", net_financing],
                ["Summary", "Net increase/(decrease) in cash", net_change],
                ["Summary", "Cash & bank at beginning of period", cash_open],
                ["Summary", "Cash & bank at end of period", cash_open + net_change],
            ]
            ex = _export(request, f"cash_flows_{s}_{e}", header, data,
                         "Statement of Cash Flows")
            if ex:
                return ex
        ctx.update({"local_receipts": local_operating_receipts, "trust_receipts": trust_receipts,
                    "operating_exp": operating_exp, "remittances": remittances,
                    "capital": capital, "net_operating": net_operating,
                    "loan_receipts": loan_receipts, "loan_repayments": loan_repayments,
                    "net_investing": net_investing, "net_financing": net_financing,
                    "net_change": net_change, "cash_open": cash_open,
                    "cash_close": cash_close, "cash_end_calc": cash_open + net_change,
                    "ties": (cash_open + net_change) == cash_close,
                    "church": SiteConfig.get().church_name})
        return self.render_to_response(ctx)


class BudgetBoardReportView(ReportAccessMixin, TemplateView):
    """Board-facing budget summary: per-department budget by source of funds, with
    Local Church Budget exposure (departmental allocations) and prior-year pegging."""
    template_name = "reports/budget_board.html"

    def get(self, request, *args, **kwargs):
        today = _dt.date.today()
        try:
            year = int(request.GET.get("year", today.year))
        except (TypeError, ValueError):
            year = today.year
        data = budget_svc.board_budget(year)
        if request.GET.get("export") in ("csv", "xlsx"):
            header = ["Department", "Trust?", "Own funds", "Local Church Budget",
                      "Other funds", "Total budget", f"{year - 1} total"]
            rows = [[r["dept"].name, "Yes" if r["is_trust"] else "No", r["own"],
                     r["lcb"], r["other"], r["total"], r["prior"]] for r in data["rows"]]
            t = data["totals"]
            rows.append(["TOTAL", "", t["own"], t["lcb"], t["other"], t["budget"], t["prior"]])
            return _export(request, f"board_budget_{year}", header, rows,
                           f"Board Budget Summary {year}")
        ctx = {"year": year, "data": data, "totals": data["totals"],
               "years": range(today.year + 1, today.year - 5, -1)}
        return self.render_to_response(ctx)


class DevGroupMembersView(PeriodMixin, TemplateView):
    """Members and their contributions to one development group, for the group
    leader's reconciliation. Can be emailed to the leader if an email is set."""
    template_name = "reports/dev_group_members.html"

    def get_context_data(self, **kwargs):
        from django.shortcuts import get_object_or_404
        from departments.models import DevelopmentGroup
        ctx = super().get_context_data(**kwargs)
        group = get_object_or_404(DevelopmentGroup, pk=kwargs["pk"])
        data = balances.dev_group_members(group, ctx["start"], ctx["end"])
        ctx.update({"group": group, "rows": data["rows"], "grand_total": data["total"]})
        return ctx

    def get(self, request, *args, **kwargs):
        ctx = self.get_context_data(**kwargs)
        if request.GET.get("export") in ("csv", "xlsx"):
            header = ["Member", "Phone", "Contributions", "Total"]
            rows = [[r["name"], r["phone"], r["count"], r["total"]] for r in ctx["rows"]]
            rows.append(["TOTAL", "", "", ctx["grand_total"]])
            return _export(request, f"devgroup_{ctx['group'].number}_members",
                           header, rows, f"{ctx['group'].label} — member contributions")
        return self.render_to_response(ctx)

    def post(self, request, *args, **kwargs):
        """Email the report to the group leader."""
        ctx = self.get_context_data(**kwargs)
        group = ctx["group"]
        if not group.leader_email:
            messages.error(request, "This group has no leader email set. Add one in the "
                                    "development group settings.")
            return redirect(request.get_full_path())
        from core.services.email import send_email, is_configured
        if not is_configured():
            messages.error(request, "Email isn't configured. Set it up in Settings → Email.")
            return redirect(request.get_full_path())
        lines = [f"{r['name']}: {r['total']:,.2f} ({r['count']} donation(s))" for r in ctx["rows"]]
        body = (f"{group.label} — member contributions\n"
                f"Period: {ctx['start']:%d %b %Y} to {ctx['end']:%d %b %Y}\n\n"
                + ("\n".join(lines) if lines else "No contributions in this period.")
                + f"\n\nTOTAL: {ctx['grand_total']:,.2f}")
        ok, detail = send_email(f"{group.label} contribution report",
                                body, group.leader_email)
        (messages.success if ok else messages.error)(
            request, f"Report to {group.leader_email}: {detail}")
        return redirect(request.get_full_path())


class DevGroupAllExcelView(ReportAccessMixin, View):
    """One workbook for all development groups: a summary sheet plus a per-group
    sheet of member contributions, for detailed offline analysis."""
    def get(self, request):
        import io
        import openpyxl
        from django.http import HttpResponse
        from departments.models import DevelopmentGroup
        start, end = parse_period(request)[:2] if request.GET.get("start") else (None, None)
        wb = openpyxl.Workbook()
        summary = wb.active
        summary.title = "Summary"
        summary.append(["Group", "Leader", "Email", "Collected", "Target",
                        "Balance", "% complete"])
        progress = {p["group"].id: p for p in balances.dev_group_progress()}
        for g in DevelopmentGroup.objects.filter(active=True):
            pr = progress.get(g.id, {})
            summary.append([g.label, g.leader_name, g.leader_email,
                            float(pr.get("collected", 0)),
                            float(pr.get("target", 0) or 0),
                            float(pr.get("balance", 0)), pr.get("pct", 0)])
            data = balances.dev_group_members(g, start, end)
            title = ("G%d" % g.number)[:28]
            ws = wb.create_sheet(title=title)
            ws.append([g.label + (" — " + g.leader_name if g.leader_name else "")])
            ws.append(["Member", "Phone", "Contributions", "Total"])
            for r in data["rows"]:
                ws.append([r["name"], r["phone"], r["count"], float(r["total"])])
            ws.append(["TOTAL", "", "", float(data["total"])])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = 'attachment; filename="development_groups_detailed.xlsx"'
        return resp


class DevGroupEmailAllView(TreasurerRequiredMixin, View):
    """Email each group's member report to its leader, spacing sends 30s apart so
    the mail provider doesn't flag a burst as spam. Runs in the background."""
    def post(self, request):
        import threading
        from departments.models import DevelopmentGroup
        from core.services.email import is_configured
        if not is_configured():
            messages.error(request, "Email isn't configured. Set it up in Settings → Email.")
            return redirect("report_dev_groups")
        start, end = (parse_period(request)[:2] if request.GET.get("start")
                      else (None, None))
        groups = list(DevelopmentGroup.objects.filter(active=True).exclude(leader_email=""))
        if not groups:
            messages.info(request, "No groups have a leader email set.")
            return redirect("report_dev_groups")
        ids = [g.id for g in groups]
        threading.Thread(target=self._send_all, args=(ids, start, end), daemon=True).start()
        messages.success(request, f"Queued reports for {len(ids)} group leader(s). "
                                  f"They'll be sent about 30 seconds apart to avoid "
                                  f"spam filters.")
        return redirect("report_dev_groups")

    @staticmethod
    def _send_all(group_ids, start, end):
        import time
        from django.db import connection
        from departments.models import DevelopmentGroup
        from core.services.email import send_email
        from core.models import Notification
        try:
            for i, gid in enumerate(group_ids):
                g = DevelopmentGroup.objects.filter(pk=gid).first()
                if not g or not g.leader_email:
                    continue
                data = balances.dev_group_members(g, start, end)
                lines = [f"{r['name']}: {r['total']:,.2f} ({r['count']} donation(s))"
                         for r in data["rows"]]
                body = (f"{g.label} — member contributions\n\n"
                        + ("\n".join(lines) if lines else "No contributions in the period.")
                        + f"\n\nTOTAL: {data['total']:,.2f}")
                ok, detail = send_email(f"{g.label} contribution report",
                                        body, g.leader_email)
                Notification.objects.create(
                    kind=Notification.Kind.GENERAL,
                    message=f"{g.label} report to {g.leader_email}: "
                            f"{'sent' if ok else 'failed — ' + detail[:80]}")
                if i < len(group_ids) - 1:
                    time.sleep(30)
        finally:
            connection.close()


class BankPositionView(ReportAccessMixin, TemplateView):
    """Bank reconciliation: does the system's bank balance agree with the bank?

    The system's bank position = opening bank balance + every confirmed BANK
    credit − every confirmed BANK debit (expenses paid from the bank appear as
    debit rows). The bank's own figure is the closing running balance of the most
    recent imported statement. If the two differ, an entry is on the statement but
    not in the app (or vice versa) — exactly the un-entered-entry case. We show the
    gap and list the most likely culprits so the treasurer can chase them.
    """
    template_name = "reports/bank_position.html"

    def get_context_data(self, **kwargs):
        from decimal import Decimal
        from statements.models import StatementImport
        ctx = super().get_context_data(**kwargs)
        cfg = SiteConfig.get()
        opening = cfg.opening_bank_balance or Decimal(0)

        # most recent statement with a captured closing balance
        stmt = (StatementImport.objects.exclude(status="PURGED")
                .exclude(stmt_closing_balance__isnull=True)
                .order_by("-stmt_last_date", "-uploaded_at").first())
        ctx["stmt"] = stmt

        # system bank movements up to the statement's last date (or all, if none)
        cutoff = stmt.stmt_last_date if stmt else None
        bank = Transaction.objects.filter(channel=Transaction.Channel.BANK,
                                          confirmed=True, is_reversal=False,
                                          is_reversed=False)
        if cutoff:
            bank = bank.filter(date__lte=cutoff)
        credits = bank.filter(direction=Transaction.Direction.CREDIT).aggregate(
            s=Sum("amount"))["s"] or Decimal(0)
        debits = bank.filter(direction=Transaction.Direction.DEBIT).aggregate(
            s=Sum("amount"))["s"] or Decimal(0)
        # Bank-paid expenses that AREN'T already represented by a bank DEBIT row
        # (i.e. entered directly with method=Bank, not resolved from the debit
        # queue). These are real outflows from the bank account and must reduce the
        # system bank balance, otherwise it overstates the cash at bank. Expenses
        # linked to a bank_transaction are excluded — they're already in `debits`.
        from cashbook.models import Expense
        bank_exp_qs = Expense.objects.filter(
            method=Expense.Method.BANK, status=Expense.Status.PAID,
            bank_transaction__isnull=True)
        if cutoff:
            bank_exp_qs = bank_exp_qs.filter(date__lte=cutoff)
        bank_expenses = bank_exp_qs.aggregate(s=Sum("amount"))["s"] or Decimal(0)
        system_balance = opening + credits - debits - bank_expenses

        ctx["opening"] = opening
        ctx["bank_credits"] = credits
        ctx["bank_debits"] = debits
        ctx["bank_expenses"] = bank_expenses
        ctx["system_balance"] = system_balance
        ctx["statement_balance"] = stmt.stmt_closing_balance if stmt else None
        ctx["difference"] = ((stmt.stmt_closing_balance - system_balance)
                             if stmt else None)

        # real-time cleared balance from the CBS feed (independent of an imported
        # statement) — often more current than the last uploaded statement.
        from statements.services.importer import latest_cleared_balance
        live = latest_cleared_balance()
        ctx["live_balance"] = live
        ctx["live_difference"] = ((live["balance"] - system_balance)
                                  if live else None)

        # if there is a gap, surface candidates: recent bank rows that look
        # suspicious (unallocated, in review, or unconfirmed) which often explain
        # a difference, plus a note about the statement's own integrity check.
        ctx["suspects"] = []
        if stmt and ctx["difference"] and abs(ctx["difference"]) > Decimal("0.01"):
            suspects = (Transaction.objects.filter(
                channel=Transaction.Channel.BANK)
                .filter(Q(confirmed=False) | Q(allocation_status="REVIEW")
                        | Q(department__isnull=True))
                .order_by("-date")[:50])
            ctx["suspects"] = [{
                "date": t.date, "payer": t.payer_name or "—",
                "amount": t.amount if t.direction == "CREDIT" else -t.amount,
                "ref": t.mpesa_ref or t.core_ref or t.reference or "",
                "why": ("not confirmed" if not t.confirmed
                        else "in review" if t.allocation_status == "REVIEW"
                        else "no fund"),
                "id": t.id} for t in suspects]
            ctx["stmt_integrity"] = stmt.balance_check
            ctx["stmt_integrity_detail"] = stmt.balance_detail
        return ctx


class CashFlowForecastView(ReportAccessMixin, TemplateView):
    """Forward-looking cash projection over 30 days / quarter / year."""
    template_name = "reports/cashflow_forecast.html"

    def get_context_data(self, **kwargs):
        from core.services import forecast
        ctx = super().get_context_data(**kwargs)
        h = forecast.horizons()
        ctx["horizons"] = h
        # a small bar/line dataset: projected position at each horizon
        ctx["chart_json"] = safe_json({
            "labels": ["Now", "30 days", "Quarter", "Year"],
            "values": [float(forecast.cash_now()),
                       float(h["30 days"]["projected"]),
                       float(h["Quarter"]["projected"]),
                       float(h["Year"]["projected"])],
        })
        return ctx


class FundThankSmsView(ReportAccessMixin, TemplateView):
    """Thank contributors to a fund (and its sub-accounts) for a period by SMS.

    Lumps each member's total giving across the fund and its sub-accounts within
    the selected period; the message is a customizable template. Treasurer only
    for sending; the preview is read-access."""
    template_name = "reports/fund_thank_sms.html"

    DEFAULT_TEMPLATE = ("Dear {name}, thank you for your contribution of KES {amount} "
                        "to {fund} ({period}). May God bless you. - {church}")

    def _period(self, request):
        import datetime as dt
        def _d(name, default):
            raw = request.GET.get(name) or request.POST.get(name)
            try:
                return dt.date.fromisoformat(raw) if raw else default
            except ValueError:
                return default
        today = dt.date.today()
        start = _d("start", today.replace(day=1))
        end = _d("end", today)
        return start, end

    def _recipients(self, dept, start, end):
        """[(member, total)] for members who gave to this fund or its sub-accounts
        in the period and have a phone on file."""
        from django.db.models import Sum
        from members.models import Member
        ids = [dept.id] + list(dept.subgroups.values_list("id", flat=True))
        rows = (Transaction.objects.filter(
                    department_id__in=ids, direction=Transaction.Direction.CREDIT,
                    confirmed=True, is_reversal=False, is_reversed=False,
                    excluded_from_income=False, member__isnull=False,
                    date__gte=start, date__lte=end)
                .values("member").annotate(t=Sum("amount")))
        totals = {r["member"]: r["t"] or Decimal(0) for r in rows}
        members = {m.id: m for m in Member.objects.filter(id__in=totals)}
        out = []
        for mid, total in totals.items():
            m = members.get(mid)
            if m and m.phone:
                out.append((m, total))
        out.sort(key=lambda x: x[1], reverse=True)
        return out

    def get_context_data(self, **kwargs):
        from core.models import SiteConfig
        from core.roles import is_treasurer
        ctx = super().get_context_data(**kwargs)
        dept = get_object_or_404(Department, pk=kwargs["pk"])
        start, end = self._period(self.request)
        recips = self._recipients(dept, start, end)
        ctx.update({
            "department": dept, "start": start, "end": end,
            "recipients": recips, "recipient_count": len(recips),
            "total": sum((t for _, t in recips), Decimal(0)),
            "template": self.DEFAULT_TEMPLATE,
            "church": SiteConfig.get().church_name or "",
            "can_send": is_treasurer(self.request.user) and SiteConfig.get().sms_enabled,
            "sms_enabled": SiteConfig.get().sms_enabled,
        })
        return ctx

    def post(self, request, *args, **kwargs):
        from core.roles import is_treasurer
        from core.models import SiteConfig
        from core.services.sms import send_sms, _format
        if not is_treasurer(request.user):
            messages.error(request, "Only a treasurer can send the thank-you messages.")
            return redirect("report_fund", pk=kwargs["pk"])
        dept = get_object_or_404(Department, pk=kwargs["pk"])
        start, end = self._period(request)
        template = request.POST.get("template") or self.DEFAULT_TEMPLATE
        church = SiteConfig.get().church_name or ""
        period_str = f"{start:%d %b %Y} – {end:%d %b %Y}"
        sent = failed = 0
        for member, total in self._recipients(dept, start, end):
            msg = _format(template, name=member.name.split()[0] if member.name else "member",
                          amount=f"{total:,.0f}", fund=dept.name,
                          period=period_str, church=church)
            log = send_sms(member.phone, msg)
            if getattr(log, "status", "") == "SENT":
                sent += 1
            else:
                failed += 1
        if sent:
            messages.success(request, f"Thank-you SMS sent to {sent} contributor(s)"
                                      + (f"; {failed} failed." if failed else "."))
        else:
            messages.error(request, "No messages were sent. "
                                    + ("Check SMS settings." if failed else "No recipients with a phone."))
        return redirect(f"{reverse('report_fund', kwargs={'pk': dept.id})}?start={start}&end={end}")


def _sfund(name):
    """Sentence-case a fund/department name for narrative text (many are
    stored in ALL CAPS). Same rule as the `sentence_fund` template filter —
    kept in sync so Python-built narrative strings match table cells."""
    from core.templatetags.treasury_extras import sentence_fund
    return sentence_fund(name)


def _camp_goal_records(year):
    """Camp Meeting expense goal (Local fund flagged CAMP_EXPENSE, aggregated over
    its sub-accounts, unchanged) paired with the Camp Meeting Offering goal — a
    single church-wide Trust-fund target now configured in Settings → Goals
    rather than on any individual fund."""
    from decimal import Decimal
    from django.db.models import Sum
    from departments.models import Department as _D

    def _ids(d):
        out = [d.id]
        for sub in d.subgroups.all():
            out.extend(_ids(sub))
        return out

    def _collected(fund):
        if fund is None:
            return Decimal(0)
        return (Transaction.objects.confirmed_credits().filter(
            department_id__in=_ids(fund), excluded_from_income=False,
            date__year=year).aggregate(t=Sum("amount"))["t"] or Decimal(0))

    def _row(name, kind, goal, fund):
        goal = goal or Decimal(0)
        col = _collected(fund)
        return {"name": name, "kind": kind, "goal": goal, "collected": col,
                "variance": col - goal,
                "pct": int(min(col / goal * 100, 999)) if goal else 0,
                "short": max(goal - col, Decimal(0))}

    rows = []
    # deterministic + defensive: if more than one fund is (mis)flagged
    # CAMP_EXPENSE, prefer the one that actually has a goal set rather than
    # an arbitrary DB-order pick (an unordered .first() is not guaranteed
    # stable across databases, and picking one with no year_goal would make
    # the goal silently vanish from every report even though it's really set
    # on a different fund).
    camp = (_D.objects.filter(active=True, goal_type="CAMP_EXPENSE")
            .prefetch_related("subgroups")
            .order_by("-year_goal", "id").first())
    if camp and camp.year_goal:
        rows.append(_row("Camp Meeting Expense Goal", "Expense (local)",
                         camp.year_goal, camp))
    cfg = SiteConfig.get()
    if cfg.camp_offering_goal and cfg.camp_offering_fund_id:
        rows.append(_row("Camp Meeting Offering Goal", "Offering (trust)",
                         cfg.camp_offering_goal, cfg.camp_offering_fund))
    return rows


class MonthlyTreasurerReportView(ReportAccessMixin, TemplateView):
    """Comprehensive monthly Treasurer's Report: collections, trust & LCB trends,
    a multi-year trend, expense and local-fund breakdowns, the income statement,
    statement of financial position, cash-flow statements, and the latest bank
    reconciliation — each with a short plain-language note. Compact, board-ready."""
    template_name = "reports/monthly_treasurer.html"

    def get_context_data(self, **kwargs):
        import datetime as _dt
        from decimal import Decimal
        from django.db.models import Sum
        from reports.services import treasurer as T
        from reports.services import budget as T_budget
        ctx = super().get_context_data(**kwargs)

        # month selection (default current month). An <input type="month">
        # submits "YYYY-MM"; a date picker may submit "YYYY-MM-DD" — accept both.
        raw = (self.request.GET.get("as_of") or "").strip()
        as_of = None
        for fmt in ("%Y-%m-%d", "%Y-%m"):
            try:
                as_of = _dt.datetime.strptime(raw, fmt).date()
                break
            except ValueError:
                continue
        if as_of is None:
            as_of = _dt.date.today()
        s, e = T.month_bounds(as_of)
        ctx["as_of"] = as_of; ctx["start"] = s; ctx["end"] = e
        ctx["today"] = _dt.date.today()
        ctx["church"] = SiteConfig.get().church_name

        # 1) collections summary
        csum = T.collections_summary(s, e)
        rows = csum["rows"]
        ctx["collections"] = csum

        # 2) trust receipted trend (current + previous 2 months)
        trust_trend = T.trust_receipted_trend(as_of, months=3)
        ctx["trust_trend"] = trust_trend
        ctx["trust_trend_json"] = safe_json({
            "labels": [c["label"] for c in trust_trend.get("columns", [])],
            "totals": [float(v or 0) for v in trust_trend.get("col_totals", [])],
        })
        # 3) LCB sub-account trend (all LCB accounts, current + previous 2 months)
        lcb_trend = T.lcb_subaccount_trend(as_of, months=3)
        ctx["lcb_trend"] = lcb_trend
        _lcb_latest = sorted(
            [r for r in lcb_trend.get("rows", []) if r.get("cells")],
            key=lambda r: r["cells"][-1], reverse=True)[:6]
        ctx["lcb_trend_json"] = safe_json({
            "labels": [r["dept"].name for r in _lcb_latest],
            "amounts": [float(r["cells"][-1] or 0) for r in _lcb_latest],
        })
        # 4) 5-year YTD trend (+ JSON for a chart)
        yearly = T.yearly_trend(as_of, years=5)
        ctx["yearly"] = yearly
        ctx["yearly_json"] = safe_json([{
            "year": str(y["year"]), "collection": float(y["collection"] or 0),
            "trust": float(y["trust"] or 0), "expense": float(y["expense"] or 0),
        } for y in yearly])
        # 5) LCB expenditure statement (fixed: matches all LCB departments)
        ctx["lcb_expenditure"] = T.lcb_expenditure(s, e)
        # 6) local funds movement statement: opening, receipts, expenses, closing
        ctx["local_statement"] = T.local_funds_statement(s, e)
        ctx["local_statement_more"] = max(
            0, len(ctx["local_statement"].get("rows", [])) - 10)

        # 7) income statement (recurrent basis)
        paid = T.PAID
        income = (Transaction.objects.confirmed_credits().filter(
            date__gte=s, date__lte=e, department__is_trust=False,
            excluded_from_income=False).aggregate(t=Sum("amount"))["t"] or Decimal(0))
        op_exp = (Expense.objects.filter(date__gte=s, date__lte=e, status__in=paid)
                  .exclude(category__in=[Expense.Category.REMITTANCE, Expense.Category.LOAN_REPAYMENT])
                  .exclude(expenditure_type=Expense.ExpenditureType.CAPITAL)
                  .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        capital = (Expense.objects.filter(date__gte=s, date__lte=e, status__in=paid,
                   expenditure_type=Expense.ExpenditureType.CAPITAL)
                   .exclude(category__in=[Expense.Category.REMITTANCE, Expense.Category.LOAN_REPAYMENT])
                   .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        ctx["income_stmt"] = {"income": income, "expense": op_exp,
                              "surplus": income - op_exp, "capital": capital}
        # detailed line items for a report-form income statement
        rev_rows = (Transaction.objects.confirmed_credits().filter(
            date__gte=s, date__lte=e, department__is_trust=False,
            excluded_from_income=False).values("department__name")
            .annotate(t=Sum("amount")).order_by("-t"))
        exp_cat = (Expense.objects.filter(date__gte=s, date__lte=e, status__in=paid)
                   .exclude(category__in=[Expense.Category.REMITTANCE, Expense.Category.LOAN_REPAYMENT])
                   .exclude(expenditure_type=Expense.ExpenditureType.CAPITAL)
                   .values("category").annotate(t=Sum("amount")).order_by("-t"))
        _cat = dict(Expense.Category.choices)
        ctx["income_detail"] = {
            "revenue": [{"name": r["department__name"] or "Unallocated",
                         "amount": r["t"]} for r in rev_rows if r["t"]],
            "expenses": [{"name": _cat.get(r["category"], r["category"]),
                          "amount": r["t"]} for r in exp_cat if r["t"]]}
        # per-fund collections detail (trust then local, with amounts)
        ctx["collection_detail"] = {
            "trust": sorted([r for r in rows if r["is_trust"] and r["receipts"]],
                            key=lambda r: r["receipts"], reverse=True),
            "local": sorted([r for r in rows if not r["is_trust"] and r["receipts"]],
                            key=lambda r: r["receipts"], reverse=True)}
        ctx["collection_detail_more"] = {
            "trust": max(0, len(ctx["collection_detail"]["trust"]) - 10),
            "local": max(0, len(ctx["collection_detail"]["local"]) - 10)}

        # 8) statement of financial position (summary, period end)
        sofp_rows = balances.department_summary(None, e)
        cash = sum((r["closing"] for r in sofp_rows), Decimal(0))
        trust_payable = sum((r["closing"] for r in sofp_rows if r["is_trust"]), Decimal(0))
        local_funds_total = cash - trust_payable
        from cashbook.views import open_payables_total, open_accruals_total
        payables = open_payables_total(e); accruals = open_accruals_total(e)
        pending = balances.pending_receipts_total(e)
        nbv = Decimal(0)
        try:
            from assets.models import nbv_total
            nbv = nbv_total(e)
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx; _lx("monthly treasurer sofp")
        # full statement of financial position, matching the main report: trust
        # payable split into receipted vs not-yet-receipted, prepayments/advances,
        # and net assets classified into unallocated / allocated / property.
        _tsum = balances.trust_summary(None, e)
        trust_receipted = sum((r["to_remit"] for r in _tsum), Decimal(0))
        trust_unreceipted = trust_payable - trust_receipted
        try:
            from cashbook.views import (unexpired_prepayments_total,
                                        outstanding_advances_total)
            prepaid = unexpired_prepayments_total(e)
            advances = outstanding_advances_total(e)
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx; _lx("monthly treasurer sofp2")
            prepaid = advances = Decimal(0)
        allocated = sum((r["closing"] for r in rows if not r["is_trust"]
                         and r["department"].category == "DEVELOPMENT"), Decimal(0))
        unallocated = local_funds_total - allocated
        accrual_adj = prepaid - payables - accruals
        ctx["sofp"] = {
            "cash": cash, "cash_on_hand": cash - advances, "advances": advances,
            "nbv": nbv, "prepaid": prepaid, "pending": pending,
            "trust_payable": trust_payable, "trust_receipted": trust_receipted,
            "trust_unreceipted": trust_unreceipted,
            "payables": payables, "accruals": accruals,
            "total_assets": (cash - advances) + advances + pending + nbv + prepaid,
            "total_liabilities": trust_payable + payables + accruals + pending,
            "local_funds": local_funds_total,
            "unallocated": unallocated, "allocated": allocated,
            "accrual_adj": accrual_adj,
            "net_assets": unallocated + allocated + nbv + accrual_adj}

        # Statement of changes in net assets: opening + surplus/(deficit) = closing.
        try:
            s_prev = s - _dt.timedelta(days=1)
            _closing_na = unallocated + allocated + nbv + accrual_adj
            prev_rows = balances.department_summary(None, s_prev)
            local_prev = [r for r in prev_rows if not r["is_trust"]]
            lf_prev = sum((r["closing"] for r in local_prev), Decimal(0))
            alloc_prev = sum((r["closing"] for r in local_prev
                              if r["department"].category == "DEVELOPMENT"), Decimal(0))
            nbv_prev = nbv
            try:
                from assets.models import nbv_total as _nbvt
                nbv_prev = _nbvt(s_prev)
            except Exception:  # noqa: BLE001
                pass
            accr_prev = (unexpired_prepayments_total(s_prev)
                         - open_payables_total(s_prev) - open_accruals_total(s_prev))
            opening_na = lf_prev + nbv_prev + accr_prev
            ctx["net_asset_changes"] = {
                "opening": opening_na,
                "surplus": _closing_na - opening_na,
                "closing": _closing_na}
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx3; _lx3("MT net-asset changes")
            ctx["net_asset_changes"] = None

        # 9 & 10) cash-flow statements (operating / investing / financing) for the month
        local_receipts = sum((r["receipts"] for r in rows if not r["is_trust"]), Decimal(0))
        trust_receipts = sum((r["receipts"] for r in rows if r["is_trust"]), Decimal(0))
        remittances = (Expense.objects.filter(date__gte=s, date__lte=e, status__in=paid,
                       category=Expense.Category.REMITTANCE)
                       .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        net_operating = local_receipts + trust_receipts - op_exp - remittances
        cash_open = sum((r["opening"] for r in rows), Decimal(0))
        ctx["cashflow"] = {
            "local_receipts": local_receipts, "trust_receipts": trust_receipts,
            "operating_exp": op_exp, "remittances": remittances,
            "net_operating": net_operating, "capital": capital,
            "net_investing": -capital, "net_change": net_operating - capital,
            "cash_open": cash_open, "cash_close": cash_open + net_operating - capital}

        # 12) most recent reconciliation
        ctx["recon"] = T.recent_reconciliation(e)

        # 13) Camp Meeting goal records (expense + offering, never group goals)
        ctx["camp_goals"] = _camp_goal_records(as_of.year)

        # short notes per section (AI narrative if available, else concise text)
        ctx["notes"] = self._notes(ctx)
        ctx["ai_summary"] = self._ai_summary(ctx)
        ctx["insights"] = self._section_insights(ctx)
        # Board-ready executive summary: budget tracking, highlights, items
        # needing attention, and decisions the Board is asked to make.
        try:
            ctx["budget_summary"] = T_budget.budget_vs_actual(
                as_of.year, period="MONTH", month=as_of.month)
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx4; _lx4("MT budget summary")
            ctx["budget_summary"] = None
        self._board_focus(ctx)
        return ctx

    def _section_insights(self, ctx):
        """A line or two of analysis per section — trend direction and a takeaway.
        Rule-based so it always works; enriched by the LLM when it is enabled."""
        from decimal import Decimal
        ins = {}

        def pct_change(now, prev):
            now = Decimal(str(now or 0)); prev = Decimal(str(prev or 0))
            if prev == 0:
                return None
            return float((now - prev) / prev * 100)

        def phrase(delta):
            if delta is None:
                return "no comparable prior period"
            if delta > 1:
                return f"up {delta:.0f}% on the prior period"
            if delta < -1:
                return f"down {abs(delta):.0f}% on the prior period"
            return "broadly flat versus the prior period"

        c = ctx.get("collections") or {}
        yearly = ctx.get("yearly") or []
        # collections: compare YTD to last year's YTD
        try:
            if len(yearly) >= 2:
                d = pct_change(yearly[-1]["collection"], yearly[-2]["collection"])
                ins["collections"] = (
                    f"Year-to-date collections are {phrase(d)}. Trust makes up "
                    f"{(c.get('trust',0)/c['total']*100):.0f}% of this month's receipts."
                    if c.get("total") else f"Collections are {phrase(d)}.")
        except Exception:  # noqa: BLE001 — an optional narrative must never break the report
            from core.utils import log_exception as _lx; _lx("monthly treasurer insight: collections")
        # trust trend
        tt = ctx.get("trust_trend") or {}
        try:
            rows = tt.get("rows") or []
            if rows:
                cols = tt.get("cells_key") or []
                last_two = [sum(r["cells"][-1] for r in rows if r["cells"]),
                            sum(r["cells"][-2] for r in rows if len(r.get("cells", [])) > 1)]
                d = pct_change(last_two[0], last_two[1])
                ins["trust_trend"] = (f"Receipted trust giving is {phrase(d)} month-on-month. "
                                      "Only receipted trust is a firm remittance liability.")
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx; _lx("monthly treasurer insight: trust_trend")
        # income & expenditure
        isr = ctx.get("income_stmt") or {}
        try:
            surplus = isr.get("surplus", 0) or 0
            if surplus >= 0:
                ins["income"] = (f"Operations ran a surplus of {float(surplus):,.0f} this "
                                 "month — income covered operating costs.")
            else:
                ins["income"] = (f"Operations ran a deficit of {abs(float(surplus)):,.0f} — "
                                 "spending outpaced local income this month; watch reserves.")
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx; _lx("monthly treasurer insight: income")
        # cash-flow
        cf = ctx.get("cashflow") or {}
        try:
            nc = cf.get("net_change", 0) or 0
            direction = "rose" if nc > 0 else ("fell" if nc < 0 else "held steady")
            ins["cashflow"] = (f"Cash {direction} by {abs(float(nc)):,.0f} over the month, "
                               f"ending at {float(cf.get('cash_close', 0)):,.0f}.")
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx; _lx("monthly treasurer insight: cashflow")
        # SoFP / net assets
        sofp = ctx.get("sofp") or {}
        try:
            na = float(sofp.get("net_assets", 0) or 0)
            tp = float(sofp.get("trust_payable", sofp.get("trust_liab", 0)) or 0)
            ins["sofp"] = (f"Net assets stand at {na:,.0f}. Trust funds of {tp:,.0f} are a "
                           "liability owed to the field, not the church's own reserves.")
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx; _lx("monthly treasurer insight: sofp")
        # camp goals
        try:
            goals = ctx.get("camp_goals") or []
            if goals:
                best = max(goals, key=lambda g: g["pct"])
                ins["camp"] = (f"{best['name']} is {best['pct']}% funded. "
                               + ("On track." if best["pct"] >= 60 else
                                  "Momentum needed to reach the target."))
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx; _lx("monthly treasurer insight: camp")

        # optional single LLM enrichment pass (kept cheap; falls back silently)
        try:
            from core.services.assistant import _llm_call
            cfg = SiteConfig.get()
            if getattr(cfg, "llm_enabled", False):
                import json as _json
                facts = {
                    "month": f"{ctx['end']:%B %Y}",
                    "collections_total": float(c.get("total", 0) or 0),
                    "surplus": float(isr.get("surplus", 0) or 0),
                    "net_assets": float(sofp.get("net_assets", 0) or 0),
                    "cash_change": float(cf.get("net_change", 0) or 0),
                }
                prompt = (
                    "You are a church treasurer analyst. Given these monthly figures, "
                    "return ONLY a JSON object with keys collections, income, cashflow, "
                    "sofp — each a single insightful sentence (max 22 words) about the "
                    "trend or what to watch. No preamble.\n" + _json.dumps(facts))
                txt, err = _llm_call(prompt, cfg, context="(section insights)")
                if txt and not err:
                    obj = _json.loads(txt[txt.find("{"):txt.rfind("}") + 1])
                    for k, v in obj.items():
                        if v:
                            ins[k] = str(v)
        except Exception:
            pass
        return ins

    def _board_focus(self, ctx):
        """Populate ctx['highlights'], ctx['attention'] and ctx['decisions'] — the
        board-oriented executive summary. Rule-based and defensive: any single
        computation failing never breaks the report, it's just omitted."""
        from decimal import Decimal
        highlights, attention, decisions = [], [], []
        ins = ctx.get("insights") or {}
        c = ctx.get("collections") or {}
        sofp = ctx.get("sofp") or {}
        cf = ctx.get("cashflow") or {}

        # --- Highlights (4-6 short, factual lines) ---
        try:
            trust_local = (ctx.get("collection_detail") or {})
            top_fund = None
            pool = (trust_local.get("trust") or []) + (trust_local.get("local") or [])
            if pool:
                top_fund = max(pool, key=lambda r: r["receipts"])
            if top_fund:
                highlights.append(
                    f"Largest receiving fund this month: {_sfund(top_fund['department'].name)} "
                    f"({float(top_fund['receipts']):,.0f}).")
        except Exception:  # noqa: BLE001
            pass
        for key in ("collections", "trust_trend", "income", "cashflow", "camp", "sofp"):
            if ins.get(key) and len(highlights) < 6:
                highlights.append(ins[key])
        ctx["highlights"] = highlights[:6]

        # --- Items requiring Board attention ---
        recon = ctx.get("recon")
        if not recon:
            attention.append({"severity": "medium", "title": "No bank reconciliation on file",
                              "detail": "No reconciliation has been recorded for this period. "
                                       "The bank and cash-book balances have not been checked "
                                       "against each other this month."})
        elif not recon.is_reconciled:
            diff = recon.difference
            attention.append({"severity": "high", "title": "Bank reconciliation not balanced",
                              "detail": f"The latest reconciliation ({recon.statement_date:%d %b %Y}) "
                                       f"leaves a difference of {float(diff or 0):,.0f} between "
                                       "the bank statement and the cash book."})

        try:
            neg = [r for r in (ctx.get("local_statement") or {}).get("rows", [])
                   if (r.get("closing") or 0) < 0]
            if neg:
                names = ", ".join(f"{_sfund(r['department'].name)} ({float(r['closing']):,.0f})"
                                  for r in neg[:5])
                attention.append({"severity": "high", "title": "Negative fund balance(s)",
                                  "detail": f"{len(neg)} local fund(s) are overdrawn: {names}"
                                           + ("…" if len(neg) > 5 else "")})
        except Exception:  # noqa: BLE001
            pass

        try:
            trust_unrec = sofp.get("trust_unreceipted") or Decimal(0)
            if trust_unrec > 0:
                attention.append({"severity": "medium",
                                  "title": "Trust funds collected but not yet receipted",
                                  "detail": f"{float(trust_unrec):,.0f} in trust-fund bank "
                                           "credits have not yet been formally receipted, so "
                                           "they aren't yet reflected as a firm remittance "
                                           "liability."})
        except Exception:  # noqa: BLE001
            pass

        try:
            bs = ctx.get("budget_summary")
            over_rows = [r for r in (bs or {}).get("rows", []) if r["over"]] if bs else []
            if over_rows:
                names = ", ".join(f"{_sfund(r['department'].name)} (over by "
                                  f"{abs(float(r['variance'])):,.0f})" for r in over_rows[:5])
                attention.append({"severity": "medium", "title": "Budget overrun this month",
                                  "detail": f"{len(over_rows)} fund(s) exceeded their prorated "
                                           f"monthly budget: {names}"
                                           + ("…" if len(over_rows) > 5 else "")})
        except Exception:  # noqa: BLE001
            pass
        ctx["attention"] = attention

        # --- Board decisions required ---
        m = ctx["end"].strftime("%B %Y")
        decisions.append({"title": "Approve the financial statements",
                          "detail": f"Adopt the Monthly Treasurer's Report for {m} as "
                                   "presented, including the Statement of Financial "
                                   "Position and Income & Expenditure Statement."})
        try:
            outstanding = sofp.get("trust_payable") or Decimal(0)
            field = SiteConfig.get().field_name or "the field"
            if outstanding > 0:
                decisions.append({"title": "Approve trust-fund remittance",
                                  "detail": f"Authorise remittance of {float(outstanding):,.0f} "
                                           f"in outstanding trust funds to {field}."})
        except Exception:  # noqa: BLE001
            pass
        for item in attention:
            if item["severity"] == "high":
                decisions.append({"title": f"Resolve: {item['title']}",
                                  "detail": item["detail"]})
        ctx["decisions"] = decisions

    def _notes(self, ctx):
        """Concise, accurate one-liners describing each section."""
        f = ctx["end"].strftime("%B %Y")
        return {
            "collections": f"Everything received in {f}, split between trust funds "
                           "(remitted to the field) and local funds (kept by the church).",
            "trust_trend": "Receipted trust collections this month and the previous "
                           "two — the trend in what is owed onward to the field.",
            "lcb_trend": "Every Local Church Budget account this month and the previous "
                         "two, so you can see which areas are growing or slowing.",
            "yearly": "Year-to-date totals for the same point in each of the last five "
                      "years, for a like-for-like long-term comparison.",
            "lcb_expenditure": f"How the Local Church Budget was spent in {f}, by category.",
            "local_statement": "Each local fund's opening balance, receipts, expenses "
                               "and closing balance for the month.",
            "sofp": "What the church owns and owes at month-end; trust funds are a "
                    "liability owed to the field, not the church's own money.",
            "cashflow": "How cash actually moved this month — from operations, into "
                        "property (investing), and the net change in cash held.",
            "recon": "The latest check that the cash book agrees with the bank statement.",
        }

    def _ai_summary(self, ctx):
        """An AI-written headline for the month, with a rule-based fallback."""
        try:
            from core.services.assistant import _llm_call
            cfg = SiteConfig.get()
            if not getattr(cfg, "llm_enabled", False):
                raise RuntimeError("assistant off")
            c = ctx["collections"]; isr = ctx["income_stmt"]
            prompt = (
                "Write 2 short sentences (max 45 words) summarising a church's monthly "
                "treasury figures for a board. Be factual and encouraging, no preamble.\n"
                f"Month: {ctx['end']:%B %Y}. Total collections: {c['total']:,.0f}. "
                f"Trust: {c['trust']:,.0f}. Local: {c['local']:,.0f}. "
                f"Operating surplus/(deficit): {isr['surplus']:,.0f}.")
            txt, err = _llm_call(prompt, cfg, context="(monthly treasurer summary)")
            if txt and not err:
                return txt.strip()
        except Exception:  # noqa: BLE001
            from core.utils import log_exception as _lx; _lx("monthly treasurer ai")
        c = ctx["collections"]; isr = ctx["income_stmt"]
        verdict = "a surplus" if isr["surplus"] >= 0 else "a deficit"
        return (f"In {ctx['end']:%B %Y} the church received {c['total']:,.0f} "
                f"({c['trust']:,.0f} trust, {c['local']:,.0f} local) and recorded "
                f"{verdict} of {abs(isr['surplus']):,.0f} on operations.")

    def _ai_narratives(self, ctx):
        """One short analysis paragraph per section, for the Word export (and
        anywhere else that wants more than the one-line `insights`). Runs
        server-side for three reasons: (1) the LLM key and prompt logic never
        reach the browser, (2) one server call can cover every section at once
        instead of the client firing several, and (3) the same figures Django
        already computed are passed straight into the prompt with no risk of
        a client-side mismatch. Every section always has rule-based text —
        `ctx['notes']` / `ctx['insights']`, already computed above — so the
        report never depends on the LLM being configured; when it is, this
        rewrites each into a fuller paragraph in one batched call."""
        notes = ctx.get("notes") or {}
        insights = ctx.get("insights") or {}
        # rule-based baseline: always available, always correct
        base = {k: " ".join(filter(None, [notes.get(k), insights.get(k)]))
                for k in notes}
        try:
            from core.services.assistant import _llm_call
            cfg = SiteConfig.get()
            if not getattr(cfg, "llm_enabled", False):
                raise RuntimeError("assistant off")
            c = ctx["collections"]; isr = ctx["income_stmt"]; sof = ctx["sofp"]
            facts = (
                f"Month {ctx['end']:%B %Y}. Collections {c['total']:,.0f} "
                f"(trust {c['trust']:,.0f}, local {c['local']:,.0f}). "
                f"Surplus/(deficit) {isr['surplus']:,.0f}. Cash & bank "
                f"{sof.get('cash', 0):,.0f}. Net assets {sof.get('net_assets', 0):,.0f}. "
                f"Trust outstanding {sof.get('trust_payable', 0):,.0f}.")
            sections = ", ".join(base.keys())
            prompt = (
                "You are writing analysis paragraphs for a church board's monthly "
                "treasury report. For EACH of these sections, write exactly one "
                f"factual, board-appropriate paragraph (25-45 words): {sections}. "
                f"Church figures this month: {facts}\n"
                "Respond ONLY as JSON: {\"section_key\": \"paragraph\", ...} using "
                "the exact section keys given, no other text.")
            txt, err = _llm_call(prompt, cfg, context="(monthly treasurer narratives)")
            if txt and not err:
                import json, re
                cleaned = re.sub(r"^```(json)?|```$", "", txt.strip(), flags=re.MULTILINE).strip()
                data = json.loads(cleaned)
                if isinstance(data, dict):
                    return {k: (data.get(k) or base.get(k, "")) for k in base}
        except Exception:  # noqa: BLE001 — narration is a nice-to-have, never blocking
            from core.utils import log_exception as _lx; _lx("monthly treasurer narratives")
        return base


# ===================== Monthly Treasurer's Report exports ===================
def _monthly_report_context(request):
    """Reuse the full MonthlyTreasurerReportView context for exports."""
    view = MonthlyTreasurerReportView()
    view.request = request
    view.kwargs = {}
    view.args = ()
    return view.get_context_data()


class MonthlyReportExcelView(ReportAccessMixin, View):
    """Download the Monthly Treasurer's Report as a multi-sheet Excel workbook —
    full detail tables (not the on-screen top-10), a KPI summary sheet styled as
    cards, and native Excel charts for the figures that are charted on screen."""
    def get(self, request):
        import io
        from decimal import Decimal
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.chart import BarChart, PieChart, LineChart, Reference
        from django.http import HttpResponse

        ctx = _monthly_report_context(request)
        wb = Workbook()
        forest = "1F5F4F"; brass = "B07D2C"; red = "B3261E"
        head_fill = PatternFill("solid", fgColor=forest)
        card_fill = PatternFill("solid", fgColor="EAF1EE")
        warn_fill = PatternFill("solid", fgColor="FBEDEA")
        white = Font(color="FFFFFF", bold=True, size=12)
        bold = Font(bold=True)
        big = Font(bold=True, size=14, color=forest)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(bottom=thin)
        money = '#,##0.00'

        def sheet(ws, title, subtitle=None):
            ws.column_dimensions["A"].width = 40
            for col in "BCDEF":
                ws.column_dimensions[col].width = 16
            ws["A1"] = ctx.get("church") or "Church Treasury"
            ws["A1"].font = Font(bold=True, size=14, color=forest)
            ws["A2"] = title
            ws["A2"].font = Font(bold=True, size=12)
            ws["A3"] = subtitle or f"Period ending {ctx['end']:%d %B %Y}"
            ws["A3"].font = Font(italic=True, color="666666")
            return 5

        def hrow(ws, r, cells):
            for i, c in enumerate(cells):
                cell = ws.cell(row=r, column=1 + i, value=c)
                cell.fill = head_fill; cell.font = white
                cell.alignment = Alignment(horizontal="left" if i == 0 else "right")
            return r + 1

        def drow(ws, r, label, *vals, bold_row=False, flag=False):
            cell = ws.cell(row=r, column=1, value=label)
            if bold_row:
                cell.font = bold
            if flag:
                cell.font = Font(bold=True, color=red)
            for i, v in enumerate(vals):
                vc = ws.cell(row=r, column=2 + i, value=float(v) if isinstance(v, Decimal) else v)
                vc.number_format = money if isinstance(v, Decimal) else "General"
                vc.alignment = Alignment(horizontal="right")
                if bold_row:
                    vc.font = bold
                if flag:
                    vc.font = Font(bold=True, color=red)
            return r + 1

        # ---------------- Executive Summary (KPI "cards" + highlights + attention) ----------------
        ws0 = wb.active; ws0.title = "Executive Summary"
        r = sheet(ws0, "Executive Summary", f"For the month of {ctx['end']:%B %Y}")
        isr = ctx["income_stmt"]; sof = ctx["sofp"]; c = ctx["collections"]
        kpis = [
            ("Total collections", c.get("total", 0)), ("Local fund receipts", c.get("local", 0)),
            ("Trust fund receipts", c.get("trust", 0)), ("Total expenses", isr.get("expense", 0)),
            ("Monthly surplus / (deficit)", isr.get("surplus", 0)),
            ("Cash & bank balance", sof.get("cash", 0)), ("Net assets", sof.get("net_assets", 0)),
            ("Trust funds outstanding", sof.get("trust_payable", 0)),
        ]
        card_row = r
        for i, (label, val) in enumerate(kpis):
            col = 1 + (i % 2) * 2
            row = card_row + (i // 2) * 3
            lc = ws0.cell(row=row, column=col, value=label)
            lc.font = Font(size=9, color="666677"); lc.fill = card_fill
            vc = ws0.cell(row=row + 1, column=col, value=float(val))
            vc.font = big; vc.number_format = money; vc.fill = card_fill
            ws0.cell(row=row, column=col + 1).fill = card_fill
            ws0.cell(row=row + 1, column=col + 1).fill = card_fill
        r = card_row + ((len(kpis) + 1) // 2) * 3 + 1

        if ctx.get("highlights"):
            ws0.cell(row=r, column=1, value="Key highlights").font = Font(bold=True, size=12, color=forest)
            r += 1
            for h in ctx["highlights"]:
                ws0.cell(row=r, column=1, value=f"\u2022 {h}")
                r += 1
            r += 1
        if ctx.get("attention"):
            ws0.cell(row=r, column=1, value="Items requiring Board attention").font = Font(bold=True, size=12, color=red)
            r += 1
            for a in ctx["attention"]:
                cell = ws0.cell(row=r, column=1, value=f"{a['title']} — {a['detail']}")
                cell.fill = warn_fill; cell.font = Font(color=red)
                r += 1

        # ---------------- Collections (FULL listing, both trust and local) ----------------
        ws1 = wb.create_sheet("Collections")
        r = sheet(ws1, "Collections Summary — full listing")
        r = hrow(ws1, r, ["Fund", "Type", "Amount", "% of total"])
        tot = c.get("total") or Decimal(1)
        cd = ctx.get("collection_detail") or {}
        first_data_row = r
        for kind, rows in (("Trust", cd.get("trust", [])), ("Local", cd.get("local", []))):
            for row in rows:
                pct = float(row["receipts"] / tot * 100) if tot else 0
                r = drow(ws1, r, row["department"].name, kind, row["receipts"], f"{pct:.1f}%")
        last_data_row = r - 1
        r = drow(ws1, r, "Total collections", "", c.get("total", 0), "100%", bold_row=True)
        if last_data_row >= first_data_row:
            pie = PieChart(); pie.title = "Collections by fund"
            data = Reference(ws1, min_col=3, min_row=first_data_row - 1, max_row=min(last_data_row, first_data_row + 14))
            cats = Reference(ws1, min_col=1, min_row=first_data_row, max_row=min(last_data_row, first_data_row + 14))
            pie.add_data(data, titles_from_data=True); pie.set_categories(cats)
            pie.height = 9; pie.width = 14
            ws1.add_chart(pie, f"F{first_data_row}")

        # ---------------- Trust Fund Performance ----------------
        ws2 = wb.create_sheet("Trust Fund Performance")
        r = sheet(ws2, "Trust Fund Performance — 3-month trend")
        tt = ctx.get("trust_trend") or {}
        cols = tt.get("columns", [])
        r = hrow(ws2, r, ["Trust fund"] + [col["label"] for col in cols])
        first_data_row = r
        for row in tt.get("rows", []):
            r = drow(ws2, r, row["dept"].name, *row["cells"])
        last_data_row = r - 1
        if tt.get("rows"):
            r = drow(ws2, r, "Total", *tt.get("col_totals", []), bold_row=True)
            line = LineChart(); line.title = "Receipted trust — 3-month trend"
            data = Reference(ws2, min_col=2, max_col=1 + len(cols), min_row=first_data_row - 1, max_row=last_data_row)
            cats = Reference(ws2, min_col=1, min_row=first_data_row, max_row=last_data_row)
            line.add_data(data, titles_from_data=True); line.set_categories(cats)
            line.height = 8; line.width = 16
            ws2.add_chart(line, f"A{r + 2}")

        # ---------------- Local Fund Performance (FULL listing) ----------------
        ws3 = wb.create_sheet("Local Fund Performance")
        r = sheet(ws3, "Local Fund Performance — full listing")
        r = hrow(ws3, r, ["Fund", "Opening", "Receipts", "Expenses", "Closing"])
        first_data_row = r
        for row in ctx.get("local_statement", {}).get("rows", []):
            nm = getattr(row.get("department"), "name", "") or row.get("name", "")
            r = drow(ws3, r, nm, row.get("opening", Decimal(0)), row.get("receipts", Decimal(0)),
                     row.get("expenses", Decimal(0)), row.get("closing", Decimal(0)),
                     flag=(row.get("closing", 0) or 0) < 0)
        last_data_row = r - 1
        totloc = ctx.get("local_statement", {}).get("totals", {})
        if totloc:
            r = drow(ws3, r, "Total", totloc.get("opening", Decimal(0)), totloc.get("receipts", Decimal(0)),
                     totloc.get("expenses", Decimal(0)), totloc.get("closing", Decimal(0)), bold_row=True)
        if last_data_row >= first_data_row:
            bar = BarChart(); bar.title = "Local fund closing balances"; bar.type = "col"
            data = Reference(ws3, min_col=5, min_row=first_data_row - 1, max_row=min(last_data_row, first_data_row + 19))
            cats = Reference(ws3, min_col=1, min_row=first_data_row, max_row=min(last_data_row, first_data_row + 19))
            bar.add_data(data, titles_from_data=True); bar.set_categories(cats)
            bar.height = 9; bar.width = 18
            ws3.add_chart(bar, f"A{r + 2}")

        # ---------------- Expenditure Summary ----------------
        ws4 = wb.create_sheet("Expenditure")
        r = sheet(ws4, "Expenditure Summary — by category")
        r = hrow(ws4, r, ["Category", "Amount"])
        first_data_row = r
        for row in ctx.get("lcb_expenditure", {}).get("rows", []):
            r = drow(ws4, r, row["label"], row["total"])
        last_data_row = r - 1
        if ctx.get("lcb_expenditure", {}).get("rows"):
            r = drow(ws4, r, "Total LCB expenditure", ctx["lcb_expenditure"]["total"], bold_row=True)
            pie2 = PieChart(); pie2.title = "Expenditure by category"
            data = Reference(ws4, min_col=2, min_row=first_data_row - 1, max_row=last_data_row)
            cats = Reference(ws4, min_col=1, min_row=first_data_row, max_row=last_data_row)
            pie2.add_data(data, titles_from_data=True); pie2.set_categories(cats)
            pie2.height = 9; pie2.width = 14
            ws4.add_chart(pie2, f"D{first_data_row}")

        # ---------------- Budget & Goal Tracking ----------------
        ws5 = wb.create_sheet("Budget & Goals")
        r = sheet(ws5, "Budget & Goal Tracking")
        bs = ctx.get("budget_summary")
        if bs and bs.get("rows"):
            r = hrow(ws5, r, ["Fund", "Budget", "Actual", "Variance", "Variance %"])
            for row in bs["rows"]:
                pct = row.get("variance_pct")
                r = drow(ws5, r, row["department"].name, row["budget"], row["actual"],
                         row["variance"], f"{float(pct):.1f}%" if pct is not None else "",
                         flag=row.get("over"))
            t = bs["totals"]
            r = drow(ws5, r, "Total", t["budget"], t["actual"], t["variance"],
                     f"{float(t['variance_pct']):.1f}%" if t.get("variance_pct") is not None else "",
                     bold_row=True)
            r += 2
        if ctx.get("camp_goals"):
            r = hrow(ws5, r, ["Goal", "Target", "Collected", "Variance", "% Complete"])
            for g in ctx["camp_goals"]:
                r = drow(ws5, r, g["name"], g["goal"], g["collected"], g["variance"], f"{g['pct']}%")

        # ---------------- Statement of Financial Position ----------------
        ws6 = wb.create_sheet("Financial Position")
        r = sheet(ws6, "Statement of Financial Position")
        r = hrow(ws6, r, ["Item", "Amount"])
        r = drow(ws6, r, "Cash & bank", sof.get("cash_on_hand", Decimal(0)))
        r = drow(ws6, r, "Property (net book value)", sof.get("nbv", Decimal(0)))
        r = drow(ws6, r, "Total assets", sof.get("total_assets", Decimal(0)), bold_row=True)
        r = drow(ws6, r, "Trust payable — receipted", sof.get("trust_receipted", Decimal(0)))
        r = drow(ws6, r, "Trust payable — not yet receipted", sof.get("trust_unreceipted", Decimal(0)))
        r = drow(ws6, r, "Total liabilities", sof.get("total_liabilities", Decimal(0)), bold_row=True)
        r = drow(ws6, r, "General net assets", sof.get("unallocated", Decimal(0)))
        r = drow(ws6, r, "Designated development funds", sof.get("allocated", Decimal(0)))
        r = drow(ws6, r, "Invested in property", sof.get("nbv", Decimal(0)))
        fund_mix_row = r
        r = drow(ws6, r, "Total net assets", sof.get("net_assets", Decimal(0)), bold_row=True)
        pie3 = PieChart(); pie3.title = "Fund composition"
        ws6.cell(row=r + 2, column=1, value="General")
        ws6.cell(row=r + 2, column=2, value=float(sof.get("unallocated", 0) or 0))
        ws6.cell(row=r + 3, column=1, value="Designated dev")
        ws6.cell(row=r + 3, column=2, value=float(sof.get("allocated", 0) or 0))
        ws6.cell(row=r + 4, column=1, value="Property")
        ws6.cell(row=r + 4, column=2, value=float(sof.get("nbv", 0) or 0))
        ws6.cell(row=r + 5, column=1, value="Trust to remit")
        ws6.cell(row=r + 5, column=2, value=float(sof.get("trust_payable", 0) or 0))
        data = Reference(ws6, min_col=2, min_row=r + 2, max_row=r + 5)
        cats = Reference(ws6, min_col=1, min_row=r + 2, max_row=r + 5)
        pie3.add_data(data); pie3.set_categories(cats)
        pie3.height = 9; pie3.width = 14
        ws6.add_chart(pie3, f"D{fund_mix_row}")

        if ctx.get("net_asset_changes"):
            nac = ctx["net_asset_changes"]
            r += 8
            r = hrow(ws6, r, ["Changes in net assets", "Amount"])
            r = drow(ws6, r, "Net assets at start of period", nac["opening"])
            r = drow(ws6, r, "Surplus / (deficit) for the period", nac["surplus"])
            r = drow(ws6, r, "Net assets at end of period", nac["closing"], bold_row=True)

        # ---------------- Cash Flow ----------------
        ws7 = wb.create_sheet("Cash Flow")
        r = sheet(ws7, "Cash Flow Statement")
        cf = ctx.get("cashflow") or {}
        r = hrow(ws7, r, ["Item", "Amount"])
        r = drow(ws7, r, "Local receipts", cf.get("local_receipts", Decimal(0)))
        r = drow(ws7, r, "Trust receipts", cf.get("trust_receipts", Decimal(0)))
        r = drow(ws7, r, "Operating expenses paid", -(cf.get("operating_exp", Decimal(0))))
        r = drow(ws7, r, "Remittances to field", -(cf.get("remittances", Decimal(0))))
        r = drow(ws7, r, "Net operating cash", cf.get("net_operating", Decimal(0)), bold_row=True)
        r = drow(ws7, r, "Capital expenditure", -(cf.get("capital", Decimal(0))))
        r = drow(ws7, r, "Net change in cash", cf.get("net_change", Decimal(0)), bold_row=True)
        r = drow(ws7, r, "Cash at start of period", cf.get("cash_open", Decimal(0)))
        r = drow(ws7, r, "Cash at end of period", cf.get("cash_close", Decimal(0)))

        # ---------------- Bank Reconciliation ----------------
        ws8 = wb.create_sheet("Bank Reconciliation")
        r = sheet(ws8, "Bank Reconciliation")
        rec = ctx.get("recon")
        if rec:
            r = hrow(ws8, r, ["Item", "Amount"])
            r = drow(ws8, r, "Statement date", str(rec.statement_date))
            r = drow(ws8, r, "Balance per bank statement", rec.bank_balance)
            r = drow(ws8, r, "Adjusted bank balance", rec.adjusted_balance)
            r = drow(ws8, r, "Balance per cash book", rec.book_balance or Decimal(0))
            r = drow(ws8, r, "Difference", rec.difference or Decimal(0), bold_row=True,
                     flag=not rec.is_reconciled)
            r = drow(ws8, r, "Status", "Reconciled" if rec.is_reconciled else "Not yet balanced")
        else:
            ws8.cell(row=r, column=1, value="No bank reconciliation recorded yet for this period.")

        # ---------------- Board Decisions Required ----------------
        ws9 = wb.create_sheet("Board Decisions")
        r = sheet(ws9, "Board Decisions Required")
        for i, d in enumerate(ctx.get("decisions", []), 1):
            ws9.cell(row=r, column=1, value=f"{i}. {d['title']}").font = bold
            r += 1
            ws9.cell(row=r, column=1, value=d["detail"])
            r += 2

        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f"treasurer_report_{ctx['end']:%Y_%m}.xlsx"
        resp = HttpResponse(buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp


class MonthlyReportWordView(ReportAccessMixin, View):
    """Download the Monthly Treasurer's Report as a Word document. Rendered as a
    Word-compatible HTML document (opens natively in Microsoft Word) so it needs
    no extra library on the server. Mirrors the on-screen report's structure —
    executive summary, then each management section with its own narrative —
    and adds a fuller per-section analysis paragraph (server-side, LLM-enriched
    with a rule-based fallback; see MonthlyTreasurerReportView._ai_narratives)
    plus a few chart images — Word can't run the on-screen report's JS charts,
    so these are rendered server-side with Pillow and embedded as base64."""
    def get(self, request):
        from django.http import HttpResponse
        from decimal import Decimal
        from reports.services.chart_image import bar_chart, donut_or_split
        view = MonthlyTreasurerReportView()
        view.request = request
        view.kwargs = {}
        view.args = ()
        ctx = view.get_context_data()
        ctx["narratives"] = view._ai_narratives(ctx)

        isr = ctx["income_stmt"]; c = ctx["collections"]
        try:
            ctx["chart_income_exp"] = bar_chart(
                "Income vs expenditure",
                [("Income", isr.get("income", 0), (31, 95, 79)),
                 ("Expenditure", isr.get("expense", 0), (179, 38, 30)),
                 ("Surplus / (deficit)", isr.get("surplus", 0), (176, 125, 44))])
        except Exception:  # noqa: BLE001 — a chart failing must never break the export
            ctx["chart_income_exp"] = None
        try:
            ctx["chart_collections_mix"] = donut_or_split(
                "Collections — local vs trust",
                [("Local", c.get("local", 0) or Decimal(0), (31, 95, 79)),
                 ("Trust", c.get("trust", 0) or Decimal(0), (176, 125, 44))])
        except Exception:  # noqa: BLE001
            ctx["chart_collections_mix"] = None
        try:
            camp = next((g for g in (ctx.get("camp_goals") or [])
                        if "Expense" in g["name"]), None)
            if camp and camp["goal"]:
                ctx["chart_camp_goal"] = bar_chart(
                    "Camp Meeting Expense Goal — progress",
                    [("Collected", camp["collected"], (31, 95, 79)),
                     ("Target", camp["goal"], (176, 125, 44))])
            else:
                ctx["chart_camp_goal"] = None
        except Exception:  # noqa: BLE001
            ctx["chart_camp_goal"] = None

        html = render(request, "reports/monthly_treasurer_word.html", ctx).content
        fname = f"treasurer_report_{ctx['end']:%Y_%m}.doc"
        resp = HttpResponse(html, content_type="application/msword")
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp


