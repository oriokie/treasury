import csv
import io
from django.http import HttpResponse


def csv_response(filename, header, rows):
    resp = HttpResponse(content_type="text/csv")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(resp)
    writer.writerow(header)
    for r in rows:
        writer.writerow(r)
    return resp


def xlsx_response(filename, header, rows, title=None, church=None):
    """A styled .xlsx with an optional title/church header, bold column headers
    and a frozen header row."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (title or "Report")[:31]
    r = 1
    if church:
        ws.cell(row=r, column=1, value=church).font = Font(bold=True, size=14); r += 1
    if title:
        ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=12, color="1F5F4F"); r += 1
    if church or title:
        r += 1
    head_row = r
    for c, h in enumerate(header, start=1):
        cell = ws.cell(row=head_row, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F5F4F")
        cell.alignment = Alignment(horizontal="center")
    rr = head_row
    for row in rows:
        rr += 1
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=rr, column=c, value=v)
            if isinstance(v, (int, float)) and c > 1:
                cell.number_format = "#,##0.00"
    for c, h in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(c)].width = max(12, len(str(h)) + 2)
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)
    buf = io.BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
