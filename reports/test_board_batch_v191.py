"""v1.91 board report batch: month filter fix, camp Type column removed,
statement of changes in net assets, exports, fonts."""
import io
from decimal import Decimal
from django.test import TestCase, Client
from django.contrib.auth.models import User, Group
from openpyxl import load_workbook
from ledger.services.posting import ensure_chart


def _tr():
    u = User.objects.create_user("tr_b191", password="x", is_superuser=True)
    u.groups.add(Group.objects.get_or_create(name="Treasurer")[0])
    return u


class BoardBatchTests(TestCase):
    def setUp(self):
        ensure_chart()
        self.tr = _tr()
        self.c = Client(); self.c.force_login(self.tr)

    def test_month_filter_accepts_yyyymm(self):
        # an <input type="month"> submits "YYYY-MM"; must not fall back to today
        b = self.c.get("/reports/board/?as_of=2026-05").content.decode()
        self.assertIn("May 2026", b)

    def test_month_filter_accepts_full_date(self):
        b = self.c.get("/reports/board/?as_of=2026-05-15").content.decode()
        self.assertIn("May 2026", b)

    def test_camp_goals_no_type_column(self):
        b = self.c.get("/reports/board/").content.decode()
        self.assertNotIn("<th>Type</th>", b)

    def test_changes_in_net_assets_section(self):
        b = self.c.get("/reports/board/?as_of=2026-06").content.decode()
        self.assertIn("Changes in net assets", b)
        self.assertIn("Net assets at start of period", b)
        self.assertIn("Net assets at end of period", b)

    def test_unallocated_note(self):
        b = self.c.get("/reports/board/").content.decode()
        self.assertIn("not earmarked for a specific project", b)

    def test_sofp_full_line_items(self):
        b = self.c.get("/reports/board/?as_of=2026-06").content.decode()
        self.assertIn("General net assets", b)
        self.assertIn("Designated development funds", b)
        self.assertIn("Total net assets", b)

    def test_excel_export_has_changes_sheet(self):
        xl = self.c.get("/reports/board/export/excel/?as_of=2026-06")
        wb = load_workbook(io.BytesIO(xl.content))
        self.assertIn("Changes in Net Assets", wb.sheetnames)

    def test_word_export_no_type_has_changes(self):
        wd = self.c.get("/reports/board/export/word/?as_of=2026-06").content.decode()
        self.assertNotIn("<th>Type</th>", wd)
        self.assertIn("changes in net assets", wd)

    def test_dashboard_card_font_uses_variable(self):
        css = open("static/css/app.css").read()
        # card header + report-card titles use the display-font variable, not a
        # hardcoded family, so the font preference applies to them
        self.assertIn('var(--font-display,"Fraunces",serif)', css)
        # no bare hardcoded Fraunces on the card header line
        import re
        bad = re.findall(r'font-family:"Fraunces",serif\}', css)
        self.assertEqual(bad, [])
