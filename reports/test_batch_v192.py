"""v1.92 batch: expense ID sort+export, edit-charge lifecycle, net-asset rename,
RTF export, dashboard bank-debit pill, per-section insights."""
import io, datetime as dt
from decimal import Decimal
from django.test import TestCase, Client
from django.contrib.auth.models import User, Group
from openpyxl import load_workbook
from core.models import SiteConfig
from departments.models import Department
from cashbook.models import Expense
from giving.models import Transaction
from ledger.services.posting import ensure_chart


def _tr(name="tr_192"):
    u = User.objects.create_user(name, password="x", is_superuser=True)
    u.groups.add(Group.objects.get_or_create(name="Treasurer")[0])
    return u


class ExpenseIdTests(TestCase):
    def setUp(self):
        ensure_chart()
        cfg = SiteConfig.get(); cfg.require_expense_approval = False
        cfg.enforce_fund_balance = False; cfg.save()
        self.tr = _tr(); self.c = Client(); self.c.force_login(self.tr)
        self.d = Department.objects.create(name="Fund 192", fund_type="LOCAL",
            category="OFFERING", show_in_expenses=True)

    def test_list_ordered_by_id_desc(self):
        a = Expense.objects.create(date=dt.date(2026, 6, 20), department=self.d,
            description="ZZZ-older-date-lower-id", amount=Decimal("10"),
            category="OTHER", status="APPROVED", recorded_by=self.tr)
        b = Expense.objects.create(date=dt.date(2026, 1, 1), department=self.d,
            description="YYY-newer-id-earlier-date", amount=Decimal("10"),
            category="OTHER", status="APPROVED", recorded_by=self.tr)
        self.assertGreater(b.id, a.id)
        body = self.c.get("/expenses/").content.decode()
        # b has the higher id, so it appears before a despite the earlier date
        self.assertLess(body.index("YYY-newer-id-earlier-date"),
                        body.index("ZZZ-older-date-lower-id"))

    def test_excel_export_includes_id(self):
        e = Expense.objects.create(date=dt.date(2026, 6, 1), department=self.d,
            description="x", amount=Decimal("10"), category="OTHER",
            status="APPROVED", recorded_by=self.tr)
        xl = self.c.get("/expenses/?export=xlsx")
        wb = load_workbook(io.BytesIO(xl.content))
        ws = wb.active
        header = [c.value for c in ws[1]] if ws.max_row else []
        # find the header row (title may occupy row 1)
        header_found = any("ID" == (c.value) for r in ws.iter_rows() for c in r)
        self.assertTrue(header_found)


class EditChargeTests(TestCase):
    def setUp(self):
        ensure_chart()
        cfg = SiteConfig.get(); cfg.require_expense_approval = False
        cfg.enforce_fund_balance = False; cfg.save()
        self.tr = _tr(); self.c = Client(); self.c.force_login(self.tr)
        self.d = Department.objects.create(name="Fund C", fund_type="LOCAL",
            category="OFFERING", show_in_expenses=True)
        self.e = Expense.objects.create(date=dt.date(2026, 6, 10), department=self.d,
            description="charge me", amount=Decimal("1000"), category="MATERIALS",
            status="APPROVED", recorded_by=self.tr, approved_by=self.tr, method="MPESA")

    def _edit(self, charge):
        return self.c.post(f"/expenses/{self.e.id}/edit/", {
            "date": "2026-06-10", "department": str(self.d.id),
            "description": "charge me", "amount": "1000", "category": "MATERIALS",
            "method": "MPESA", "charge": charge})

    def test_charge_created_on_edit(self):
        self._edit("35")
        self.assertEqual(self.e.charges.count(), 1)
        self.assertEqual(self.e.charges.first().amount, Decimal("35"))
        self.assertEqual(self.e.charges.first().category, "BANK_CHARGE")

    def test_charge_not_duplicated(self):
        self._edit("35"); self._edit("35")
        self.assertEqual(self.e.charges.count(), 1)

    def test_charge_updated_in_place(self):
        self._edit("35"); self._edit("60")
        self.assertEqual(self.e.charges.count(), 1)
        self.assertEqual(self.e.charges.first().amount, Decimal("60"))

    def test_charge_deleted_when_cleared(self):
        self._edit("35"); self._edit("")
        self.assertEqual(self.e.charges.count(), 0)

    def test_charge_prefilled_on_edit_form(self):
        self._edit("42")
        body = self.c.get(f"/expenses/{self.e.id}/edit/").content.decode()
        self.assertIn("42", body)


class NetAssetRenameTests(TestCase):
    def setUp(self):
        ensure_chart()
        self.c = Client(); self.c.force_login(_tr())

    def test_labels_renamed(self):
        b = self.c.get("/reports/board/?as_of=2026-06").content.decode()
        self.assertIn("General net assets", b)
        self.assertIn("Designated development funds", b)
        self.assertNotIn("Unallocated funds (general)", b)


class RtfExportTests(TestCase):
    def setUp(self):
        ensure_chart()
        self.c = Client(); self.c.force_login(_tr())

    def test_rtf_export(self):
        r = self.c.get("/reports/board/export/rtf/?as_of=2026-06")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r["Content-Type"], "application/rtf")
        body = r.content.decode()
        self.assertTrue(body.startswith(r"{\rtf1"))
        self.assertTrue(body.rstrip().endswith("}"))
        self.assertIn("General net assets", body)


class DashboardDebitPillTests(TestCase):
    def setUp(self):
        ensure_chart()
        self.c = Client(); self.c.force_login(_tr())

    def test_bank_debit_pill_links_to_debit_queue(self):
        Transaction.objects.create(date=dt.date(2026, 6, 1), amount=Decimal("500"),
            direction="DEBIT", channel="BANK", allocation_status="REVIEW",
            confirmed=True, reference="DBTX1")
        body = self.c.get("/").content.decode()
        self.assertIn("bank debit(s) to classify", body)
        import re
        m = re.search(r'href="([^"]+)"[^>]*>\u2691 \d+ bank debit', body)
        self.assertIsNotNone(m)
        self.assertIn("/debits/", m.group(1))


class SectionInsightsTests(TestCase):
    def setUp(self):
        ensure_chart()
        self.c = Client(); self.c.force_login(_tr())

    def test_insights_present(self):
        b = self.c.get("/reports/board/?as_of=2026-06").content.decode()
        self.assertIn("\U0001F4C8", b)  # at least one insight line rendered
