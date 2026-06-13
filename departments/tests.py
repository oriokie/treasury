from decimal import Decimal
from django.test import TestCase
from departments.models import Department


class SubgroupTests(TestCase):
    def test_subaccount_inherits_parent_fund_type(self):
        youth = Department.objects.create(name="Youth", fund_type=Department.FundType.LOCAL)
        choir = Department.objects.create(name="Youth Choir", parent=youth,
                                          fund_type=Department.FundType.TRUST)  # should be overridden
        self.assertEqual(choir.fund_type, Department.FundType.LOCAL)
        self.assertFalse(choir.is_trust)
        self.assertTrue(choir.is_subgroup)

    def test_subgroups_reverse_relation(self):
        youth = Department.objects.create(name="Youth", fund_type=Department.FundType.LOCAL)
        Department.objects.create(name="Potluck", parent=youth)
        Department.objects.create(name="Mission", parent=youth)
        self.assertEqual(youth.subgroups.count(), 2)

    def test_str_shows_parent_path(self):
        youth = Department.objects.create(name="Youth", fund_type=Department.FundType.LOCAL)
        potluck = Department.objects.create(name="Potluck", parent=youth)
        self.assertEqual(str(potluck), "Youth / Potluck")

    def test_deleting_parent_keeps_subgroup(self):
        youth = Department.objects.create(name="Youth", fund_type=Department.FundType.LOCAL)
        potluck = Department.objects.create(name="Potluck", parent=youth)
        youth.delete()
        potluck.refresh_from_db()
        self.assertIsNone(potluck.parent)


class BudgetSourceAndBoardTests(TestCase):
    def setUp(self):
        from django.contrib.auth.models import User
        from departments.models import Department, Budget, BudgetLine
        from decimal import Decimal
        self.u = User.objects.create_superuser("bs", password="x")
        self.lcb = Department.objects.create(name="LCB – Local Church Budget",
                                             fund_type=Department.FundType.LOCAL)
        self.youth = Department.objects.create(name="Youth", fund_type=Department.FundType.LOCAL)
        self.year = 2026
        b = Budget.objects.create(year=self.year, department=self.youth, amount=Decimal("0"))
        BudgetLine.objects.create(budget=b, name="Camp", amount=Decimal("70000"))          # own
        BudgetLine.objects.create(budget=b, name="PA hire", amount=Decimal("30000"),
                                  source_fund=self.lcb)                                     # LCB
        b.amount = b.lines_total
        b.save()
        self.budget = b
        self.client.login(username="bs", password="x")

    def test_source_kind(self):
        lines = {l.name: l.source_kind for l in self.budget.lines.all()}
        self.assertEqual(lines["Camp"], "OWN")
        self.assertEqual(lines["PA hire"], "LCB")

    def test_board_budget_splits_and_lcb_exposure(self):
        from reports.services.budget import board_budget
        from decimal import Decimal
        d = board_budget(self.year)
        self.assertEqual(d["totals"]["budget"], Decimal("100000"))
        self.assertEqual(d["totals"]["lcb"], Decimal("30000"))
        self.assertEqual(d["totals"]["own"], Decimal("70000"))
        alloc = {a["dept"].id: a["amount"] for a in d["lcb_alloc"]}
        self.assertEqual(alloc[self.youth.id], Decimal("30000"))

    def test_board_report_page_renders(self):
        from django.urls import reverse
        r = self.client.get(reverse("report_budget_board") + f"?year={self.year}")
        self.assertEqual(r.status_code, 200)

    def test_copy_prior_year_breakdown(self):
        from django.urls import reverse
        from departments.models import Budget, BudgetLine
        from decimal import Decimal
        # this budget's lines are year 2026; make 2027 copy from 2026
        cur, _ = Budget.objects.get_or_create(year=2027, department=self.youth,
                                              defaults={"amount": Decimal("0")})
        self.client.post(reverse("budget_lines", args=[self.youth.id]),
                         {"year": 2027, "action": "copy_prior"})
        cur.refresh_from_db()
        self.assertEqual(cur.lines.count(), 2)
        self.assertEqual(cur.lines_total, Decimal("100000"))
        # source_fund carried over
        self.assertTrue(cur.lines.filter(source_fund=self.lcb).exists())

    def test_blank_source_defaults_to_own(self):
        from departments.models import BudgetLine
        ln = self.budget.lines.get(name="Camp")
        self.assertIsNone(ln.source_fund)
        self.assertIn("own funds", ln.source_label)


class LcbFundLookupTests(TestCase):
    """lcb_fund() must not crash on its name-based fallback branches. Regression
    for a FieldError ('children' instead of 'subgroups') that 500'd the budget
    breakdown page when an LCB fund was matched by name rather than prefix."""

    def test_lcb_fund_by_full_name(self):
        from departments.models import Department, lcb_fund
        Department.objects.create(name="Local Church Budget", fund_type="LOCAL",
                                  category="OFFERING")
        # must resolve without raising
        result = lcb_fund()
        self.assertIsNotNone(result)

    def test_budget_lines_page_loads(self):
        from django.contrib.auth.models import User, Group
        from django.test import Client
        from departments.models import Department
        u = User.objects.create_user("bl", password="x")
        g, _ = Group.objects.get_or_create(name="Treasurer")
        u.groups.add(g)
        Department.objects.create(name="Local Church Budget", fund_type="LOCAL",
                                  category="OFFERING")
        d = Department.objects.create(name="Youth Fund", fund_type="LOCAL",
                                      category="MINISTRY")
        c = Client(); c.force_login(u)
        r = c.get(f"/budget/{d.id}/lines/?year=2026")
        self.assertEqual(r.status_code, 200)


class FundStructureImportTests(TestCase):
    """Dedicated fund/sub-account importer: template lists existing funds, and
    apply creates parents then sub-accounts (which inherit the parent's type)."""

    def setUp(self):
        from django.contrib.auth.models import User
        self.u = User.objects.create_superuser("fsi", password="x")
        self.client.login(username="fsi", password="x")

    def _file(self, rows):
        import io
        import openpyxl
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Funds"
        ws.append(["Fund name", "Parent fund (blank = top level)", "Fund type",
                   "Category", "Opening balance"])
        for r in rows:
            ws.append(r)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return buf.getvalue()

    def test_template_downloads_with_existing_funds(self):
        import io
        import openpyxl
        from departments.models import Department
        Department.objects.create(name="EXISTING FUND",
                                  fund_type=Department.FundType.LOCAL)
        r = self.client.get("/funds/structure-import/?template=1")
        self.assertEqual(r.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        self.assertIn("Funds", wb.sheetnames)
        # existing fund should be listed on the Lists sheet
        lists = list(wb["Lists"].iter_rows(values_only=True))
        names = {row[0] for row in lists if row and row[0]}
        self.assertIn("EXISTING FUND", names)

    def test_creates_parent_and_subaccounts(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from departments.models import Department
        data = self._file([
            ["YOUTH MINISTRY", "", "Local", "Ministry", 0],
            ["YOUTH POTLUCK", "YOUTH MINISTRY", "Local", "Ministry", 0],
            ["YOUTH MISSION", "YOUTH MINISTRY", "Trust", "Development", 0],
        ])
        self.client.post("/funds/structure-import/",
                         {"file": SimpleUploadedFile("f.xlsx", data)})
        self.client.post("/funds/structure-import/", {"apply": "1"})
        parent = Department.objects.get(name="YOUTH MINISTRY")
        sub1 = Department.objects.get(name="YOUTH POTLUCK")
        sub2 = Department.objects.get(name="YOUTH MISSION")
        self.assertIsNone(parent.parent_id)
        self.assertEqual(sub1.parent_id, parent.id)
        self.assertEqual(sub2.parent_id, parent.id)
        # sub-account inherits the parent's fund type (parent is Local)
        self.assertEqual(sub2.fund_type, parent.fund_type)

    def test_existing_fund_is_skipped(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from departments.models import Department
        Department.objects.create(name="ALREADY HERE",
                                  fund_type=Department.FundType.LOCAL)
        before = Department.objects.count()
        data = self._file([["ALREADY HERE", "", "Local", "Ministry", 0]])
        self.client.post("/funds/structure-import/",
                         {"file": SimpleUploadedFile("f.xlsx", data)})
        self.client.post("/funds/structure-import/", {"apply": "1"})
        self.assertEqual(Department.objects.count(), before)   # nothing created

    def test_order_independent_parent_after_child(self):
        # child row appears BEFORE its parent — should still link correctly
        from django.core.files.uploadedfile import SimpleUploadedFile
        from departments.models import Department
        data = self._file([
            ["CHILD FIRST", "PARENT LATER", "Local", "Ministry", 0],
            ["PARENT LATER", "", "Local", "Ministry", 0],
        ])
        self.client.post("/funds/structure-import/",
                         {"file": SimpleUploadedFile("f.xlsx", data)})
        self.client.post("/funds/structure-import/", {"apply": "1"})
        parent = Department.objects.get(name="PARENT LATER")
        child = Department.objects.get(name="CHILD FIRST")
        self.assertEqual(child.parent_id, parent.id)


class BudgetTemplatePrefilledTests(TestCase):
    """The budget template is pre-filled with one row per existing fund."""

    def test_template_lists_existing_funds_as_rows(self):
        import io
        import openpyxl
        from django.contrib.auth.models import User
        from departments.models import Department
        User.objects.create_superuser("bt", password="x")
        self.client.login(username="bt", password="x")
        Department.objects.create(name="FUND ALPHA",
            fund_type=Department.FundType.LOCAL, selectable=True)
        Department.objects.create(name="FUND BETA",
            fund_type=Department.FundType.LOCAL, selectable=True)
        r = self.client.get("/budget/template/?year=2026")
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb["Budget lines"]
        names = {row[0] for row in ws.iter_rows(values_only=True) if row[0]}
        self.assertIn("FUND ALPHA", names)
        self.assertIn("FUND BETA", names)
