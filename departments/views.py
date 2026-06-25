from django.contrib import messages
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView

from core.permissions import ReadAccessMixin, TreasurerRequiredMixin
from .forms import DepartmentForm, DevelopmentGroupForm
from .models import Department, DevelopmentGroup


class DepartmentListView(ReadAccessMixin, ListView):
    model = Department
    template_name = "departments/list.html"
    context_object_name = "departments"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        tops = (Department.objects.filter(parent__isnull=True)
                .prefetch_related("subgroups"))
        dev_groups = DevelopmentGroup.objects.all()
        tree = []
        for f in tops:
            node = {"fund": f, "subs": list(f.subgroups.all()), "dev_groups": []}
            if f.name.lower() == "development":
                node["dev_groups"] = list(dev_groups)
            tree.append(node)
        ctx["fund_tree"] = tree
        ctx["dev_groups"] = dev_groups
        from giving.models import SplitFund
        ctx["split_funds"] = SplitFund.objects.prefetch_related("components__department")
        return ctx


class DepartmentCreateView(TreasurerRequiredMixin, CreateView):
    model = Department
    form_class = DepartmentForm
    template_name = "departments/form.html"
    success_url = reverse_lazy("department_list")

    def form_valid(self, form):
        messages.success(self.request, "Fund created.")
        return super().form_valid(form)


class DepartmentUpdateView(TreasurerRequiredMixin, UpdateView):
    model = Department
    form_class = DepartmentForm
    template_name = "departments/form.html"
    success_url = reverse_lazy("department_list")

    def form_valid(self, form):
        messages.success(self.request, "Fund updated.")
        return super().form_valid(form)


# ---- Development groups (CRUD) ----
class DevGroupCreateView(TreasurerRequiredMixin, CreateView):
    model = DevelopmentGroup
    form_class = DevelopmentGroupForm
    template_name = "departments/dev_group_form.html"
    success_url = reverse_lazy("department_list")

    def form_valid(self, form):
        messages.success(self.request, "Development group created.")
        return super().form_valid(form)


class DevGroupUpdateView(TreasurerRequiredMixin, UpdateView):
    model = DevelopmentGroup
    form_class = DevelopmentGroupForm
    template_name = "departments/dev_group_form.html"
    success_url = reverse_lazy("department_list")

    def form_valid(self, form):
        messages.success(self.request, "Development group updated.")
        return super().form_valid(form)


class DevGroupDeleteView(TreasurerRequiredMixin, DeleteView):
    model = DevelopmentGroup
    template_name = "departments/dev_group_confirm_delete.html"
    success_url = reverse_lazy("department_list")

    def form_valid(self, form):
        messages.success(self.request, "Development group deleted.")
        return super().form_valid(form)


import datetime as _dt
from decimal import Decimal, InvalidOperation
from django.db import transaction as db_tx
from django.shortcuts import render, redirect
from django.views import View


class BudgetView(TreasurerRequiredMixin, View):
    """Annual budgeting: set each fund's annual budget and brought-forward
    opening balance in one place (kept off the structural Funds page)."""
    template_name = "departments/budget.html"

    def _ordered(self):
        from collections import defaultdict
        kids = defaultdict(list)
        tops = []
        for d in Department.objects.filter(active=True).select_related("parent"):
            (kids[d.parent_id].append(d) if d.parent_id else tops.append(d))
        rows = []
        for t in sorted(tops, key=lambda x: x.name):
            rows.append({"d": t, "child": False})
            for c in sorted(kids.get(t.id, []), key=lambda x: x.name):
                rows.append({"d": c, "child": True})
        return rows

    def get(self, request):
        from .models import Budget
        year = int(request.GET.get("year") or _dt.date.today().year)
        budget_objs = {b.department_id: b for b in
                       Budget.objects.filter(year=year).prefetch_related("lines")}
        rows = self._ordered()
        for r in rows:
            b = budget_objs.get(r["d"].id)
            if b is not None:
                lt = b.lines_total
                r["budget"] = lt if lt else b.amount
                r["from_lines"] = bool(lt)
            else:
                r["budget"] = None
                r["from_lines"] = False
        return render(request, self.template_name,
                      {"rows": rows, "year": year,
                       "years": range(_dt.date.today().year + 1, _dt.date.today().year - 5, -1)})

    def post(self, request):
        from .models import Budget
        year = int(request.POST.get("year") or _dt.date.today().year)
        n = 0
        for d in Department.objects.filter(active=True):
            b = request.POST.get(f"budget_{d.id}")
            o = request.POST.get(f"opening_{d.id}")
            changed = False
            if b is not None:
                try:
                    amt = Decimal(b) if b.strip() else None
                    if amt is None:
                        Budget.objects.filter(year=year, department=d).delete()
                    else:
                        Budget.objects.update_or_create(
                            year=year, department=d, defaults={"amount": amt})
                    n += 1
                except (InvalidOperation, AttributeError):
                    pass
            if o is not None:
                try:
                    d.opening_balance = Decimal(o) if o.strip() else Decimal(0)
                    changed = True
                except (InvalidOperation, AttributeError):
                    pass
            if changed:
                d.save(update_fields=["opening_balance"])
        messages.success(request, f"Saved budgets and opening balances for {year}.")
        return redirect(f"{reverse_lazy('budget')}?year={year}")


class BudgetLinesView(TreasurerRequiredMixin, View):
    """Manage the breakdown lines for one fund's annual budget."""
    template_name = "departments/budget_lines.html"

    def get(self, request, pk):
        from .models import Budget, BudgetLine, lcb_fund
        from cashbook.models import Expense
        from django.shortcuts import get_object_or_404
        dept = get_object_or_404(Department, pk=pk)
        year = int(request.GET.get("year") or _dt.date.today().year)
        budget, _ = Budget.objects.get_or_create(year=year, department=dept)
        lcb = lcb_fund()
        other_funds = Department.objects.filter(active=True).exclude(pk=dept.id)
        if lcb:
            other_funds = other_funds.exclude(pk=lcb.id)
        prior = Budget.objects.filter(year=year - 1, department=dept).first()
        return render(request, self.template_name, {
            "dept": dept, "year": year, "budget": budget,
            "lines": budget.lines.select_related("source_fund").all(),
            "categories": Expense.Category.choices,
            "quarters": BudgetLine._meta.get_field("quarter").choices,
            "lcb": lcb, "other_funds": other_funds.order_by("name"),
            "prior_year": year - 1,
            "prior_total": (prior.lines_total if prior else None),
            "prior_has_lines": bool(prior and prior.lines.exists()),
            "years": range(_dt.date.today().year + 1, _dt.date.today().year - 5, -1),
        })

    def post(self, request, pk):
        from .models import Budget, BudgetLine
        from django.shortcuts import get_object_or_404
        dept = get_object_or_404(Department, pk=pk)
        year = int(request.POST.get("year") or _dt.date.today().year)
        budget, _ = Budget.objects.get_or_create(year=year, department=dept)
        action = request.POST.get("action")
        if action == "add":
            name = (request.POST.get("name") or "").strip()
            if name:
                try:
                    amt = Decimal(request.POST.get("amount") or "0")
                except InvalidOperation:
                    amt = Decimal(0)
                src_id = request.POST.get("source_fund") or None
                source = Department.objects.filter(pk=src_id).first() if src_id else None
                BudgetLine.objects.create(budget=budget, name=name,
                                          category=request.POST.get("category", ""),
                                          quarter=request.POST.get("quarter", ""),
                                          amount=amt, source_fund=source)
                messages.success(request, f"Added budget line \u201c{name}\u201d.")
        elif action == "delete":
            BudgetLine.objects.filter(pk=request.POST.get("line"), budget=budget).delete()
            messages.success(request, "Budget line removed.")
        elif action == "copy_prior":
            prior = Budget.objects.filter(year=year - 1, department=dept).first()
            if prior and prior.lines.exists():
                n = 0
                for ln in prior.lines.all():
                    BudgetLine.objects.create(budget=budget, name=ln.name,
                        category=ln.category, amount=ln.amount, source_fund=ln.source_fund,
                        quarter=ln.quarter)
                    n += 1
                messages.success(request, f"Copied {n} line(s) from {year - 1}. "
                                          f"Adjust the amounts as needed.")
            else:
                messages.info(request, f"No {year - 1} breakdown to copy from.")
        # keep the headline Budget amount in step with the breakdown
        budget.amount = budget.lines_total
        budget.save(update_fields=["amount"])
        return redirect(f"{reverse_lazy('budget_lines', args=[dept.id])}?year={year}")


# ---------------------------------------------------------------------------
# Bulk fund + budget import (items 6 & 7)
# ---------------------------------------------------------------------------
import difflib
from .models import Budget, BudgetLine


def _norm_fund(s):
    """Normalise a fund name for matching: uppercase, spaces for separators,
    collapse whitespace."""
    s = (s or "").upper().replace("_", " ").replace("-", " ").replace("/", " ")
    return " ".join(s.split())


def _match_department(name, depts):
    """Return (department or None, score). Tries exact normalised match first,
    then a couple of known synonyms, then fuzzy."""
    n = _norm_fund(name)
    by_norm = {_norm_fund(d.name): d for d in depts}
    if n in by_norm:
        return by_norm[n], 1.0
    # common singular/plural and abbreviation synonyms seen in this church's data
    synonyms = {
        "PATHFINDER": "PATHFINDERS", "CHILDREN": "CHILDREN MINISTRY",
        "MASTER GUIDE": "MASTER GUIDE", "TRUST FUND": "TRUST FUND",
        "LOCAL CHURCH BUDGET": "LCB LOCAL CHURCH BUDGET",
        "ASAM": "ADVENTIST SINGLE ADULT MINISTRIES",
        "APM": "ADVENTIST POSSIBILITY MINISTRY",
    }
    if n in synonyms and _norm_fund(synonyms[n]) in by_norm:
        return by_norm[_norm_fund(synonyms[n])], 0.95
    close = difflib.get_close_matches(n, list(by_norm), n=1, cutoff=0.84)
    if close:
        return by_norm[close[0]], 0.85
    return None, 0.0


class BulkFundImportView(TreasurerRequiredMixin, View):
    """Two-step wizard: upload a fund/budget spreadsheet, review how each row maps
    to an existing department (creating new funds or sub-groups where needed),
    then apply — writing the per-year Budget and optional monthly breakdown.

    Step 1 (GET): show the upload form.
    Step 1 (POST file): parse, match each fund, stash a plan in the session, show
                        the review/mapping table with a prompt for unmatched funds.
    Step 2 (POST apply): create departments as chosen and write budgets.
    """
    template_name = "departments/bulk_import.html"

    MONTHS = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN",
              "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]

    def get(self, request):
        return render(request, self.template_name, {"stage": "upload"})

    def post(self, request):
        if request.POST.get("apply_lines"):
            return _bulk_apply_line_items(self, request)
        if request.POST.get("apply"):
            return self._apply(request)
        return self._parse(request)

    # ---- step 1: parse the file and build a mapping plan ----
    def _parse(self, request):
        import openpyxl
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "Choose a spreadsheet to upload.")
            return redirect("bulk_fund_import")
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception:
            messages.error(request, "Could not read that file — please upload a .xlsx.")
            return redirect("bulk_fund_import")
        try:
            year = int(request.POST.get("year") or _dt.date.today().year)
        except (TypeError, ValueError):
            year = _dt.date.today().year

        # the line-item template (downloaded from the budget page) has a
        # "Budget lines" sheet with Department / Line item / Amount columns —
        # handle that shape directly, writing a BudgetLine per row.
        if "Budget lines" in wb.sheetnames:
            return _bulk_parse_line_items(self, request, wb["Budget lines"], year)

        ws = wb["DEPARTMENTS"] if "DEPARTMENTS" in wb.sheetnames else wb.active
        # find the header row + the columns we care about (NAME, B/F, projected,
        # and the monthly income block). The budget sheet has two header rows.
        rows = list(ws.iter_rows(values_only=True))
        name_col = bf_col = pinc_col = pexp_col = None
        header_idx = None
        for i, r in enumerate(rows[:5]):
            cells = [str(c).strip().upper() if c is not None else "" for c in r]
            if "NAME" in cells:
                header_idx = i
                name_col = cells.index("NAME")
                for label, attr in (("B/F", "bf"), ("PROJECTED INCOME", "pinc"),
                                    ("PROJECTED EXPENSES", "pexp")):
                    if label in cells:
                        if attr == "bf": bf_col = cells.index(label)
                        elif attr == "pinc": pinc_col = cells.index(label)
                        elif attr == "pexp": pexp_col = cells.index(label)
                break
        if name_col is None:
            messages.error(request, "Couldn't find a NAME column — is this the budget "
                                    "workbook? Expected a DEPARTMENTS sheet with NAME, "
                                    "B/F and PROJECTED INCOME columns.")
            return redirect("bulk_fund_import")

        # locate the monthly columns. The workbook has two JAN…DEC blocks: an
        # INCOME block then an EXPENSES block. A budget is planned *spend*, so we
        # use the EXPENSES block (the second JAN) to stay consistent with the
        # PROJECTED EXPENSES headline. Fall back to the first block if there's
        # only one.
        month_start = None
        for r in rows[:5]:
            cells = [str(c).strip().upper() if c is not None else "" for c in r]
            jans = [j for j, v in enumerate(cells) if v == "JAN"]
            if jans:
                month_start = jans[1] if len(jans) > 1 else jans[0]
                break

        depts = list(Department.objects.all())
        plan = []
        for r in rows[header_idx + 1:]:
            if name_col >= len(r):
                continue
            raw_name = r[name_col]
            if not raw_name or not str(raw_name).strip():
                continue
            name = str(raw_name).strip()
            def num(col):
                if col is None or col >= len(r) or r[col] in (None, ""):
                    return 0.0
                try:
                    return float(r[col])
                except (TypeError, ValueError):
                    return 0.0
            pinc = num(pinc_col)
            pexp = num(pexp_col)
            bf = num(bf_col)
            if pinc == 0 and pexp == 0 and bf == 0:
                continue  # skip empty filler rows
            months = []
            if month_start is not None:
                for k in range(12):
                    months.append(num(month_start + k))
            dept, score = _match_department(name, depts)
            plan.append({
                "name": name,
                "bf": bf, "pinc": pinc, "pexp": pexp,
                "months": months,
                "match_id": dept.id if dept else None,
                "match_name": dept.name if dept else None,
                "score": score,
            })

        request.session["bulk_fund_plan"] = {"year": year, "rows": plan}
        # candidate departments for the manual-map dropdowns
        candidates = [{"id": d.id, "name": str(d)} for d in
                      sorted(depts, key=lambda x: str(x))]
        matched = [p for p in plan if p["match_id"]]
        unmatched = [p for p in plan if not p["match_id"]]
        return render(request, self.template_name, {
            "stage": "review", "year": year, "plan": plan,
            "candidates": candidates, "matched_count": len(matched),
            "unmatched_count": len(unmatched),
            "fund_types": Department.FundType.choices,
            "categories": Department.Category.choices,
        })

    # ---- step 2: apply the reviewed plan ----
    @db_tx.atomic
    def _apply(self, request):
        data = request.session.get("bulk_fund_plan")
        if not data:
            messages.error(request, "Your import session expired — please upload again.")
            return redirect("bulk_fund_import")
        year = data["year"]
        plan = data["rows"]
        created_funds = budgets_written = lines_written = skipped = 0

        for i, p in enumerate(plan):
            choice = request.POST.get(f"map_{i}", "")
            dept = None
            if choice.startswith("dept:"):
                dept = Department.objects.filter(pk=choice.split(":", 1)[1]).first()
            elif choice == "create":
                ftype = request.POST.get(f"ftype_{i}") or Department.FundType.LOCAL
                cat = request.POST.get(f"cat_{i}") or Department.Category.MINISTRY
                parent_id = request.POST.get(f"parent_{i}") or None
                parent = Department.objects.filter(pk=parent_id).first() if parent_id else None
                dept, _was = Department.objects.get_or_create(
                    name=p["name"].upper().strip(),
                    defaults={"fund_type": ftype, "category": cat, "parent": parent})
                if _was:
                    created_funds += 1
            elif choice == "skip" or (choice == "" and not p["match_id"]):
                skipped += 1
                continue
            else:
                # default: use the auto-matched department if present
                if p["match_id"]:
                    dept = Department.objects.filter(pk=p["match_id"]).first()
            if not dept:
                skipped += 1
                continue

            # write the year budget (projected expenses is the planned spend)
            amount = Decimal(str(p["pexp"] or p["pinc"] or 0))
            budget, _ = Budget.objects.update_or_create(
                year=year, department=dept,
                defaults={"amount": amount, "note": "Imported from budget workbook"})
            budgets_written += 1

            # optional monthly breakdown lines (only if the user asked for it)
            if request.POST.get("with_months") and p.get("months"):
                budget.lines.filter(name__startswith="Budget — ").delete()
                for mi, mval in enumerate(p["months"]):
                    if mval:
                        BudgetLine.objects.create(
                            budget=budget,
                            name=f"Budget — {self.MONTHS[mi]}",
                            amount=Decimal(str(mval)))
                        lines_written += 1
                budget.amount = budget.lines_total
                budget.save(update_fields=["amount"])

        request.session.pop("bulk_fund_plan", None)
        parts = [f"{budgets_written} budget(s) written for {year}"]
        if created_funds:
            parts.append(f"{created_funds} new fund(s) created")
        if lines_written:
            parts.append(f"{lines_written} monthly line(s)")
        if skipped:
            parts.append(f"{skipped} row(s) skipped")
        messages.success(request, ", ".join(parts) + ".")
        return redirect("budget")


# ---------------------------------------------------------------------------
# Budget template download + line-item budget import (item 1)
# ---------------------------------------------------------------------------
class BudgetTemplateDownloadView(TreasurerRequiredMixin, View):
    """Download a ready-to-fill budget template (one sheet of line items). The
    treasurer fills in a row per planned expense line for each department, says
    where the money comes from (the funding department — blank means the
    department's own funds), then re-imports it on the bulk-import screen.

    The per-department budget total is the sum of its line items, and a line's
    funding source is recorded so reports can show how much of each fund's plan
    is financed by the LCB, by another fund, or from its own balance.
    """

    def get(self, request):
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation
        from cashbook.models import Expense
        try:
            year = int(request.GET.get("year") or _dt.date.today().year)
        except (TypeError, ValueError):
            year = _dt.date.today().year

        depts = list(Department.objects.filter(active=True, selectable=True)
                     .order_by("name"))
        cats = [lbl for _, lbl in Expense.Category.choices]

        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Budget lines"
        head = ["Department", "Line item", "Category", "Amount", "Funded by"]
        ws.append(head)
        for c in range(1, len(head) + 1):
            cell = ws.cell(1, c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F5F4F")
            cell.alignment = Alignment(horizontal="center")
        # Pre-fill one row per existing fund so the treasurer just enters amounts
        # against funds already in the system (rather than typing fund names).
        # If a fund already has a budget for the year, show its current total as a
        # starting point. Funds with no budget get a blank amount to fill in.
        from .models import Budget
        existing_budgets = {b.department_id: b.amount for b in
                            Budget.objects.filter(year=year)}
        if depts:
            for d in depts:
                amt = existing_budgets.get(d.id)
                ws.append([d.name, "", "", (float(amt) if amt else None), ""])
        else:
            # no funds yet — fall back to worked examples
            ws.append(["YOUTH", "Youth camp transport", "Transport", 20000, ""])
            ws.append(["YOUTH", "PA system hire", "Materials / supplies", 15000,
                       "LCB – Local Church Budget"])
            ws.append(["CHILDREN MINISTRY", "VBS materials", "Materials / supplies",
                       8000, "YOUTH"])

        # reference lists for the dropdowns
        ref = wb.create_sheet("Lists")
        ref["A1"] = "Departments"; ref["A1"].font = Font(bold=True)
        for i, d in enumerate(depts, start=2):
            ref.cell(i, 1, d.name)
        ref["B1"] = "Funded by (blank = own funds)"; ref["B1"].font = Font(bold=True)
        ref.cell(2, 2, "")           # own funds (blank)
        for i, d in enumerate(depts, start=3):
            ref.cell(i, 2, d.name)
        ref["C1"] = "Categories"; ref["C1"].font = Font(bold=True)
        for i, c in enumerate(cats, start=2):
            ref.cell(i, 3, c)

        nrows = max(len(depts) + 5, 200)
        dv_dept = DataValidation(type="list",
            formula1=f"=Lists!$A$2:$A${len(depts) + 1}", allow_blank=False)
        dv_cat = DataValidation(type="list",
            formula1=f"=Lists!$C$2:$C${len(cats) + 1}", allow_blank=True)
        dv_src = DataValidation(type="list",
            formula1=f"=Lists!$B$2:$B${len(depts) + 2}", allow_blank=True)
        ws.add_data_validation(dv_dept); ws.add_data_validation(dv_cat)
        ws.add_data_validation(dv_src)
        dv_dept.add(f"A2:A{nrows}"); dv_cat.add(f"C2:C{nrows}")
        dv_src.add(f"E2:E{nrows}")

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 28
        # a short instructions block below the lists
        ref["A1"].comment = None
        info = wb.create_sheet("How to fill this in")
        for i, line in enumerate([
            f"Budget template for {year}",
            "",
            "One row per planned spending line.",
            "  • Department — the fund the line belongs to (pick from the list).",
            "  • Line item — a short description, e.g. 'Camp transport'.",
            "  • Category — optional; lets reports compare plan vs actual by category.",
            "  • Amount — the planned amount for the year.",
            "  • Funded by — where the money comes from. Leave BLANK if the",
            "      department pays from its own funds; otherwise pick the fund that",
            "      finances it (often the Local Church Budget, or another department).",
            "",
            "The department's budget total is the sum of its line items.",
            "A line funded by another department still counts in this department's",
            "plan, and is also recorded against the funding department so reports",
            "can show who finances whom.",
            "",
            "When done, go to Budgets → Bulk import and upload this file.",
        ], start=1):
            info.cell(i, 1, line)
        info.column_dimensions["A"].width = 76

        buf = io.BytesIO(); wb.save(buf)
        from django.http import HttpResponse
        resp = HttpResponse(buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="budget_template_{year}.xlsx"'
        return resp


# Line-item budget template handling (appended to BulkFundImportView via monkey
# methods would be ugly; instead these live as module functions the view calls).
def _li_match(name, depts):
    return _match_department(name, depts)


def _bulk_parse_line_items(view, request, ws, year):
    """Parse the line-item budget template. Each row is a planned budget line for
    a department, with an optional funding source. Build a review plan grouped by
    department, flagging any department or funding source we can't match."""
    from cashbook.models import Expense
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        messages.error(request, "The 'Budget lines' sheet is empty.")
        return redirect("bulk_fund_import")
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    def col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None
    c_dept = col("department")
    c_item = col("line item", "line", "description")
    c_cat = col("category")
    c_amt = col("amount")
    c_src = col("funded by", "source", "funding source")
    if c_dept is None or c_amt is None:
        messages.error(request, "Couldn't find the Department and Amount columns — "
                                "please use the downloaded template.")
        return redirect("bulk_fund_import")

    depts = list(Department.objects.all())
    cat_by_label = {lbl.lower(): code for code, lbl in Expense.Category.choices}
    cat_by_code = {code.upper(): code for code, _ in Expense.Category.choices}

    lines = []
    for r in rows[1:]:
        if c_dept >= len(r) or not r[c_dept] or not str(r[c_dept]).strip():
            continue
        dname = str(r[c_dept]).strip()
        try:
            amt = float(r[c_amt]) if c_amt < len(r) and r[c_amt] not in (None, "") else 0.0
        except (TypeError, ValueError):
            amt = 0.0
        if amt == 0:
            continue
        item = (str(r[c_item]).strip() if c_item is not None and c_item < len(r)
                and r[c_item] else "Budget line")
        raw_cat = (str(r[c_cat]).strip() if c_cat is not None and c_cat < len(r)
                   and r[c_cat] else "")
        cat_code = cat_by_code.get(raw_cat.upper()) or cat_by_label.get(raw_cat.lower()) or ""
        raw_src = (str(r[c_src]).strip() if c_src is not None and c_src < len(r)
                   and r[c_src] else "")
        dept, dscore = _match_department(dname, depts)
        if raw_src:
            src, sscore = _match_department(raw_src, depts)
        else:
            src, sscore = None, 1.0   # blank = own funds
        lines.append({
            "dept_name": dname, "dept_id": dept.id if dept else None,
            "dept_match": dept.name if dept else None,
            "item": item, "amount": amt, "category": cat_code,
            "src_name": raw_src, "src_id": src.id if src else None,
            "src_match": src.name if src else None,
            "src_blank": not raw_src,
        })

    if not lines:
        messages.error(request, "No budget lines with an amount were found.")
        return redirect("bulk_fund_import")

    # group by department for a clean review, and gather any unmatched names
    unmatched_depts = sorted({l["dept_name"] for l in lines if not l["dept_id"]})
    unmatched_srcs = sorted({l["src_name"] for l in lines
                             if l["src_name"] and not l["src_id"]})
    request.session["bulk_line_plan"] = {"year": year, "lines": lines}
    candidates = [{"id": d.id, "name": str(d)} for d in
                  sorted(depts, key=lambda x: str(x))]
    total = sum(l["amount"] for l in lines)
    return render(request, "departments/bulk_import.html", {
        "stage": "review_lines", "year": year, "lines": lines,
        "candidates": candidates, "line_total": total,
        "unmatched_depts": unmatched_depts, "unmatched_srcs": unmatched_srcs,
        "fund_types": Department.FundType.choices,
        "categories": Department.Category.choices,
    })


@db_tx.atomic
def _bulk_apply_line_items(view, request):
    data = request.session.get("bulk_line_plan")
    if not data:
        messages.error(request, "Your import session expired — please upload again.")
        return redirect("bulk_fund_import")
    year = data["year"]
    lines = data["lines"]

    # resolve any user-chosen department/source mappings from the review form
    def resolve(prefix, i, fallback_id, raw_name):
        choice = request.POST.get(f"{prefix}_{i}", "")
        if choice.startswith("dept:"):
            return Department.objects.filter(pk=choice.split(":", 1)[1]).first()
        if choice == "create" and raw_name:
            d, _ = Department.objects.get_or_create(
                name=raw_name.upper().strip(),
                defaults={"fund_type": Department.FundType.LOCAL,
                          "category": Department.Category.MINISTRY})
            return d
        if fallback_id:
            return Department.objects.filter(pk=fallback_id).first()
        return None

    from .models import Budget, BudgetLine
    written = skipped = created = 0
    touched_budgets = set()
    for i, l in enumerate(lines):
        dept = resolve("dept", i, l["dept_id"], l["dept_name"])
        if not dept:
            skipped += 1
            continue
        if l["src_blank"]:
            src = None
        else:
            src = resolve("src", i, l["src_id"], l["src_name"])
        budget, _ = Budget.objects.get_or_create(
            year=year, department=dept,
            defaults={"note": "From budget template"})
        if budget.pk not in touched_budgets:
            # clear previous template lines for a clean re-import per department
            budget.lines.all().delete()
            touched_budgets.add(budget.pk)
        BudgetLine.objects.create(budget=budget, name=l["item"][:120],
            category=l["category"], amount=Decimal(str(l["amount"])),
            source_fund=src)
        written += 1

    # recompute each touched budget's headline to the sum of its lines
    for bid in touched_budgets:
        b = Budget.objects.get(pk=bid)
        b.amount = b.lines_total
        b.save(update_fields=["amount"])

    request.session.pop("bulk_line_plan", None)
    messages.success(request,
        f"{written} budget line(s) imported across {len(touched_budgets)} "
        f"department(s) for {year}." + (f" {skipped} skipped." if skipped else ""))
    return redirect("budget")


# ===========================================================================
# Dedicated fund / department structure importer (with sub-groups)
# ===========================================================================
class FundStructureImportView(TreasurerRequiredMixin, View):
    """Build the chart of accounts in bulk: funds AND their sub-accounts, from a
    simple template — separate from budgeting (no amounts beyond an optional
    opening balance). Two-step wizard: upload → review (new vs existing) → apply.

    The template has one row per fund:
        Fund name | Parent fund | Fund type | Category | Opening balance
    A blank Parent makes a top-level fund; naming an existing (or also-in-sheet)
    fund as Parent makes the row a sub-account of it. Parents are created before
    children so order in the sheet doesn't matter.
    """
    template_name = "departments/fund_import.html"

    def get(self, request):
        if request.GET.get("template"):
            return self._template()
        return render(request, self.template_name, {"stage": "upload"})

    def post(self, request):
        if request.POST.get("apply"):
            return self._apply(request)
        return self._parse(request)

    # ---- template (pre-listing existing funds for the parent dropdown) ----
    def _template(self):
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation
        from django.http import HttpResponse

        existing = list(Department.objects.order_by("name"))
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Funds"
        head = ["Fund name", "Parent fund (blank = top level)", "Fund type",
                "Category", "Opening balance", "Show in expenses (Yes/No)"]
        ws.append(head)
        for c in range(1, len(head) + 1):
            cell = ws.cell(1, c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F5F4F")
            cell.alignment = Alignment(horizontal="center")
        # worked examples showing a parent and two sub-accounts
        ws.append(["YOUTH", "", "Local", "Ministry", 0, "Yes"])
        ws.append(["YOUTH POTLUCK", "YOUTH", "Local", "Ministry", 0, "Yes"])
        ws.append(["YOUTH MISSION", "YOUTH", "Local", "Development", 0, "No"])
        ws.append(["COMBINED OFFERING TRUST", "", "Trust", "Trust", 0, "Yes"])

        # reference lists
        ref = wb.create_sheet("Lists")
        ref["A1"] = "Existing funds (use as Parent)"; ref["A1"].font = Font(bold=True)
        for i, d in enumerate(existing, start=2):
            ref.cell(i, 1, d.name)
        ftypes = [lbl for _, lbl in Department.FundType.choices]
        cats = [lbl for _, lbl in Department.Category.choices]
        ref["B1"] = "Fund type"; ref["B1"].font = Font(bold=True)
        for i, v in enumerate(ftypes, start=2):
            ref.cell(i, 2, v)
        ref["C1"] = "Category"; ref["C1"].font = Font(bold=True)
        for i, v in enumerate(cats, start=2):
            ref.cell(i, 3, v)

        nrows = max(len(existing) + 20, 200)
        last_parent = len(existing) + 1
        if existing:
            dv_parent = DataValidation(type="list",
                formula1=f"=Lists!$A$2:$A${last_parent}", allow_blank=True)
            ws.add_data_validation(dv_parent); dv_parent.add(f"B2:B{nrows}")
        dv_ft = DataValidation(type="list",
            formula1=f"=Lists!$B$2:$B${len(ftypes) + 1}", allow_blank=True)
        dv_cat = DataValidation(type="list",
            formula1=f"=Lists!$C$2:$C${len(cats) + 1}", allow_blank=True)
        ws.add_data_validation(dv_ft); ws.add_data_validation(dv_cat)
        dv_ft.add(f"C2:C{nrows}"); dv_cat.add(f"D2:D{nrows}")
        dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        ws.add_data_validation(dv_yn); dv_yn.add(f"F2:F{nrows}")

        for col, w in zip("ABCDEF", (28, 30, 14, 16, 16, 22)):
            ws.column_dimensions[col].width = w

        info = wb.create_sheet("How to fill this in")
        for i, line in enumerate([
            "Fund / department structure import",
            "",
            "One row per fund or sub-account.",
            "  • Fund name — the fund's name (must be unique).",
            "  • Parent fund — leave BLANK for a top-level fund. To make a",
            "      sub-account, put the parent's exact name here (it can be an",
            "      existing fund or another row in this sheet).",
            "  • Fund type — Trust (remitted to the field) or Local (kept by the",
            "      church). A sub-account inherits its parent's type automatically.",
            "  • Category — Offering, Ministry, Development, Holding or Trust.",
            "  • Opening balance — optional brought-forward balance (usually 0 for",
            "      a new fund).",
            "  • Show in expenses — Yes (default) lists the fund in the expense",
            "      picker; No hides it (collection-only funds never spent directly).",
            "",
            "Existing funds are NOT changed; a row whose name already exists is",
            "skipped on import. Parents are created before their sub-accounts, so",
            "the order of rows does not matter.",
        ], start=1):
            info.cell(i, 1, line)
        info.column_dimensions["A"].width = 74

        buf = io.BytesIO(); wb.save(buf)
        resp = HttpResponse(buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = 'attachment; filename="fund_structure_template.xlsx"'
        return resp

    # ---- step 1: parse + build plan ----
    def _parse(self, request):
        import openpyxl
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "Choose a spreadsheet to upload.")
            return redirect("fund_structure_import")
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception:
            messages.error(request, "Could not read that file — please upload the .xlsx template.")
            return redirect("fund_structure_import")
        ws = wb["Funds"] if "Funds" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            messages.error(request, "The sheet is empty.")
            return redirect("fund_structure_import")
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

        def col(*names):
            for n in names:
                if n in header:
                    return header.index(n)
            return None
        c_name = col("fund name", "name")
        c_parent = col("parent fund (blank = top level)", "parent fund", "parent")
        c_type = col("fund type", "type")
        c_cat = col("category")
        c_open = col("opening balance", "opening", "b/f")
        c_show = col("show in expenses (yes/no)", "show in expenses", "show in expense", "show")
        if c_name is None:
            messages.error(request, "Couldn't find a 'Fund name' column — please use the template.")
            return redirect("fund_structure_import")

        ft_map = {lbl.lower(): val for val, lbl in Department.FundType.choices}
        ft_map.update({val.lower(): val for val, _ in Department.FundType.choices})
        cat_map = {lbl.lower(): val for val, lbl in Department.Category.choices}
        cat_map.update({val.lower(): val for val, _ in Department.Category.choices})
        existing_names = {d.name.upper() for d in Department.objects.all()}

        def cell(r, idx):
            if idx is None or idx >= len(r) or r[idx] in (None, ""):
                return ""
            return str(r[idx]).strip()

        plan = []
        for r in rows[1:]:
            name = cell(r, c_name)
            if not name:
                continue
            name_u = name.upper()
            parent = cell(r, c_parent)
            ftype_raw = cell(r, c_type).lower()
            ftype = ft_map.get(ftype_raw, Department.FundType.LOCAL)
            cat_raw = cell(r, c_cat).lower()
            cat = cat_map.get(cat_raw, Department.Category.MINISTRY)
            try:
                opening = float(r[c_open]) if c_open is not None and c_open < len(r) \
                    and r[c_open] not in (None, "") else 0.0
            except (TypeError, ValueError):
                opening = 0.0
            show_raw = cell(r, c_show).strip().lower()
            show = False if show_raw in ("no", "n", "false", "0", "hide") else True
            plan.append({
                "name": name_u, "parent": parent.upper() if parent else "",
                "ftype": ftype, "cat": cat, "opening": opening, "show": show,
                "exists": name_u in existing_names,
            })

        if not plan:
            messages.error(request, "No fund rows with a name were found.")
            return redirect("fund_structure_import")

        # validate parents: a parent must be an existing fund or another row here
        sheet_names = {p["name"] for p in plan}
        for p in plan:
            if p["parent"]:
                p["parent_ok"] = (p["parent"] in existing_names
                                  or p["parent"] in sheet_names)
            else:
                p["parent_ok"] = True

        request.session["fund_structure_plan"] = plan
        new_count = sum(1 for p in plan if not p["exists"])
        sub_count = sum(1 for p in plan if p["parent"])
        bad_parents = [p for p in plan if not p["parent_ok"]]
        return render(request, self.template_name, {
            "stage": "review", "plan": plan,
            "new_count": new_count, "existing_count": len(plan) - new_count,
            "sub_count": sub_count, "bad_parents": bad_parents,
        })

    # ---- step 2: apply (two-pass: parents first) ----
    @db_tx.atomic
    def _apply(self, request):
        plan = request.session.get("fund_structure_plan")
        if not plan:
            messages.error(request, "Your import session expired — please upload again.")
            return redirect("fund_structure_import")

        created = skipped = subs = 0
        # pass 1: create all top-level / parent funds (rows with no parent)
        for p in plan:
            if p["parent"]:
                continue
            if p["exists"]:
                skipped += 1
                continue
            Department.objects.get_or_create(
                name=p["name"],
                defaults={"fund_type": p["ftype"], "category": p["cat"],
                          "show_in_expenses": p.get("show", True),
                          "opening_balance": Decimal(str(p["opening"]))})
            created += 1
        # pass 2: create sub-accounts (parent now exists)
        for p in plan:
            if not p["parent"]:
                continue
            if p["exists"]:
                skipped += 1
                continue
            parent = Department.objects.filter(name=p["parent"]).first()
            if not parent:
                skipped += 1
                continue
            # a sub-account inherits the parent's fund type
            Department.objects.get_or_create(
                name=p["name"],
                defaults={"fund_type": parent.fund_type, "category": p["cat"],
                          "parent": parent, "show_in_expenses": p.get("show", True),
                          "opening_balance": Decimal(str(p["opening"]))})
            created += 1; subs += 1

        request.session.pop("fund_structure_plan", None)
        parts = [f"{created} fund(s) created"]
        if subs:
            parts.append(f"{subs} as sub-account(s)")
        if skipped:
            parts.append(f"{skipped} skipped (already existed)")
        messages.success(request, ", ".join(parts) + ".")
        return redirect("department_list")


# --- Account lifecycle + collection-account consolidation (#1, #2) ------------
import datetime as _dt
from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect, render
from django.views import View
from django.db import transaction as _db_tx
from .models import DepartmentStatusLog


def _log_status(dept, from_status, to_status, user, note=""):
    DepartmentStatusLog.objects.create(
        department=dept, from_status=from_status or "", to_status=to_status,
        changed_by=user, note=note[:200])


def _balance(dept, as_of=None):
    from reports.services.balances import fund_balance
    return fund_balance(dept, as_of) or Decimal(0)


class ConsolidateView(TreasurerRequiredMixin, View):
    """Transfer every non-zero child (collection account) balance under a parent
    into the parent in one operation, with proper transfer records and audit.
    Children end at zero; their history stays intact."""
    template_name = "departments/consolidate.html"

    def get(self, request, pk):
        parent = get_object_or_404(Department, pk=pk)
        rows = []
        for child in parent.subgroups.all():
            bal = _balance(child)
            rows.append({"child": child, "balance": bal})
        ctx = {"parent": parent, "rows": rows,
               "total": sum((r["balance"] for r in rows), Decimal(0)),
               "movable": [r for r in rows if r["balance"] != 0]}
        return render(request, self.template_name, ctx)

    def post(self, request, pk):
        from cashbook.models import FundTransfer
        parent = get_object_or_404(Department, pk=pk)
        today = _dt.date.today()
        moved = 0
        total = Decimal(0)
        with _db_tx.atomic():
            for child in parent.subgroups.all():
                bal = _balance(child)
                if bal == 0:
                    continue
                FundTransfer.objects.create(
                    date=today, source=child, destination=parent, amount=bal,
                    reason=f"Consolidation of {child.name} into {parent.name}",
                    recorded_by=request.user)
                _log_status(child, child.status, child.status, request.user,
                            note=f"Consolidated balance {bal:,.2f} into {parent.name}")
                moved += 1
                total += bal
        if moved:
            messages.success(request, f"Consolidated {moved} account(s) totalling "
                                      f"{total:,.2f} into {parent.name}. Their balances are now zero "
                                      f"and their history is preserved.")
        else:
            messages.info(request, "Nothing to consolidate — all sub-accounts are already at zero.")
        return redirect("department_list")


class CloseAccountView(TreasurerRequiredMixin, View):
    """Close an account once its purpose is complete. Only allowed at a zero
    balance (transfer or consolidate any remainder first)."""
    def post(self, request, pk):
        dept = get_object_or_404(Department, pk=pk)
        bal = _balance(dept)
        if bal != 0:
            messages.error(request, f"{dept.name} can't be closed — its balance is "
                                    f"{bal:,.2f}. Transfer or consolidate the remaining "
                                    f"balance to zero first.")
            return redirect("department_list")
        if dept.status != Department.Status.ACTIVE:
            messages.info(request, f"{dept.name} is already {dept.get_status_display().lower()}.")
            return redirect("department_list")
        prev = dept.status
        dept.status = Department.Status.CLOSED
        dept.save()
        _log_status(dept, prev, dept.status, request.user,
                    note=request.POST.get("note", "")[:200])
        messages.success(request, f"{dept.name} closed. It stays in historical reports "
                                  f"but won't accept new transactions.")
        return redirect("department_list")


class ArchiveAccountView(TreasurerRequiredMixin, View):
    """Archive a closed account (tidies it out of the main lists; still in reports)."""
    def post(self, request, pk):
        dept = get_object_or_404(Department, pk=pk)
        prev = dept.status
        dept.status = Department.Status.ARCHIVED
        dept.save()
        _log_status(dept, prev, dept.status, request.user)
        messages.success(request, f"{dept.name} archived.")
        return redirect("historical_accounts")


class ReopenAccountView(TreasurerRequiredMixin, View):
    """Reopen a closed/archived account back to active."""
    def post(self, request, pk):
        dept = get_object_or_404(Department, pk=pk)
        prev = dept.status
        dept.status = Department.Status.ACTIVE
        dept.save()
        _log_status(dept, prev, dept.status, request.user, note="Reopened")
        messages.success(request, f"{dept.name} reopened and active again.")
        return redirect("department_list")


class HistoricalAccountsView(ReadAccessMixin, View):
    """Closed and archived accounts, with their balances and status history."""
    template_name = "departments/historical_accounts.html"

    def get(self, request):
        depts = (Department.objects.filter(
            status__in=[Department.Status.CLOSED, Department.Status.ARCHIVED])
            .prefetch_related("status_logs").order_by("status", "name"))
        rows = [{"dept": d, "balance": _balance(d),
                 "last": d.status_logs.first()} for d in depts]
        from core.roles import is_treasurer
        return render(request, self.template_name,
                      {"rows": rows, "is_treasurer": is_treasurer(request.user)})
