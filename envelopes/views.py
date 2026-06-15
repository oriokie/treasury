import io
import datetime as dt
import json
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.db import transaction as db_tx
from django.shortcuts import redirect, render, get_object_or_404
from django.urls import reverse
from django.views import View
from django.views.generic import ListView, DetailView

from core.permissions import ReadAccessMixin, DataEntryRequiredMixin
from core.models import SiteConfig
from core.services.sms import send_receipt_sms
from core.utils import sabbath_week_of
from departments.models import Department
from giving.models import Transaction
from members.models import Member
from .models import Envelope, EnvelopeLine


from core.utils import (sabbath_bucket, sabbath_of, last_saturday as _last_saturday,
                        saturdays_of_month as _saturdays_of_month)


class EnvelopeListView(ReadAccessMixin, View):
    """Envelopes grouped by Sabbath within a chosen month."""
    template_name = "envelopes/list.html"

    def get(self, request):
        today = dt.date.today()
        raw = request.GET.get("month")
        try:
            year, month = (int(x) for x in raw.split("-")) if raw else (today.year, today.month)
        except (ValueError, AttributeError):
            year, month = today.year, today.month

        saturdays = _saturdays_of_month(year, month)
        envs = (Envelope.objects.select_related("member")
                .prefetch_related("lines__department"))
        groups = {s: [] for s in saturdays}
        other = []
        for e in envs:
            b = sabbath_bucket(e.date)
            if b in groups:
                groups[b].append(e)
        sections = []
        for s in saturdays:
            items = sorted(groups[s], key=lambda e: e.receipt_no)
            total = sum((e.total for e in items), Decimal(0))
            trust = sum((l.amount for e in items for l in e.lines.all()
                         if l.department.is_trust), Decimal(0))
            sections.append({"sabbath": s, "items": items, "total": total,
                             "trust": trust, "local": total - trust,
                             "count": len(items)})
        from core.models import SabbathClose, period_locked
        closed_map = {c.sabbath: c for c in SabbathClose.objects.filter(
            sabbath__in=[s["sabbath"] for s in sections])}
        for s in sections:
            s["closed"] = closed_map.get(s["sabbath"])
        # optional single-Sabbath filter (?sabbath=YYYY-MM-DD) so a busy month
        # can be narrowed to one Sabbath's receipts
        sel_sabbath = None
        raw_sab = request.GET.get("sabbath")
        if raw_sab:
            try:
                sel_sabbath = dt.datetime.strptime(raw_sab, "%Y-%m-%d").date()
            except ValueError:
                sel_sabbath = None
        visible = ([s for s in sections if s["sabbath"] == sel_sabbath]
                   if sel_sabbath else sections)
        first = dt.date(year, month, 1)
        prev = (first - dt.timedelta(days=1)).replace(day=1)
        nxt = (first + dt.timedelta(days=32)).replace(day=1)
        return render(request, self.template_name, {
            "sections": visible,
            "all_sections": sections,
            "sel_sabbath": sel_sabbath,
            "month_label": first.strftime("%B %Y"),
            "month_value": f"{year}-{month:02d}",
            "prev_month": f"{prev.year}-{prev.month:02d}",
            "next_month": f"{nxt.year}-{nxt.month:02d}",
            "all_saturdays": saturdays,
            "grand_total": sum((s["total"] for s in visible), Decimal(0)),
            "month_locked": period_locked(first),
            "sms_enabled": SiteConfig.get().sms_enabled,
            "whatsapp_enabled": SiteConfig.get().whatsapp_enabled,
        })


class EnvelopeReassignView(DataEntryRequiredMixin, View):
    """Move an envelope to a different Sabbath."""

    def post(self, request, pk):
        from django.shortcuts import get_object_or_404
        env = get_object_or_404(Envelope, pk=pk)
        try:
            new_date = dt.date.fromisoformat(request.POST.get("sabbath"))
        except (TypeError, ValueError):
            messages.error(request, "Pick a valid Sabbath date.")
            return redirect("envelope_list")
        env.date = new_date
        env.sabbath_week = sabbath_week_of(new_date)
        env.save(update_fields=["date", "sabbath_week"])
        # keep any linked cash transactions in step
        Transaction.objects.filter(envelope_lines__envelope=env).update(
            date=new_date, sabbath_week=env.sabbath_week)
        messages.success(request, f"Moved receipt {env.receipt_no} to {new_date:%d %b %Y}.")
        return redirect(f"{reverse('envelope_list')}?month={new_date.year}-{new_date.month:02d}")


class EnvelopeDetailView(ReadAccessMixin, DetailView):
    model = Envelope
    template_name = "envelopes/detail.html"
    context_object_name = "envelope"


# ---- Column catalogue for the ledger / template / import ----
PREFERRED = ["Tithe", "Combined Offering", "Camp Meeting", "Development",
             "Sabbath School", "Loose Offering", "LCB – Local Church Budget",
             "Thanksgiving Offering"]


def _is_building(name):
    return "building" in (name or "").lower()


def column_catalog(for_import=False):
    """Candidate ledger columns: active funds (excluding Building) + split funds,
    preferred ones first, with sensible defaults pre-selected. When for_import is
    set, sub-accounts (Trust Fund and LCB children) are excluded — imports use the
    standalone funds and split offerings only."""
    from giving.models import SplitFund
    from departments.models import split_component_dept_ids
    skip_ids = split_component_dept_ids() if for_import else set()
    cols = []
    for d in Department.objects.filter(active=True):
        if _is_building(d.name):
            continue
        if d.id in skip_ids:        # the 50% split halves — shown as one split column
            continue
        cols.append({"key": str(d.id), "label": d.name, "name": d.name,
                     "kind": "dept", "trust": d.is_trust})
    for s in SplitFund.objects.filter(active=True):
        cols.append({"key": f"split:{s.id}", "label": f"{s.name} (split)",
                     "name": s.name, "kind": "split", "trust": False})
    pref = [p.lower() for p in PREFERRED]

    def rank(c):
        n = c["name"].lower()
        return (0, pref.index(n)) if n in pref else (1, c["label"].lower())
    cols.sort(key=rank)
    for c in cols:
        c["default"] = c["name"].lower() in set(pref)
    return cols


def _amount(raw):
    try:
        v = Decimal(str(raw).replace(",", "").strip())
        return v if v else None
    except (InvalidOperation, TypeError, AttributeError):
        return None


def _expand_lines(amounts, funds, splits, dev_group=None):
    """amounts: {key: raw}. Returns list of (Department, Decimal[, DevelopmentGroup]),
    expanding splits. If `dev_group` is given it is attached to the Development line."""
    lines = []
    for key, raw in amounts.items():
        amt = _amount(raw)
        if not amt:
            continue
        if str(key).startswith("split:"):
            sf = splits.get(int(str(key).split(":", 1)[1]))
            if sf:
                for pdept, pamt in sf.split(amt):
                    if pamt:
                        lines.append((pdept, pamt))
        else:
            try:
                fid = int(key)
            except (ValueError, TypeError):
                continue
            if fid in funds:
                dept = funds[fid]
                if dev_group is not None and dept.category == "DEVELOPMENT":
                    lines.append((dept, amt, dev_group))
                else:
                    lines.append((dept, amt))
    return lines


def _save_envelope(*, date, name, receipt, channel, lines, member, user, cfg):
    env = Envelope.objects.create(
        date=date, sabbath_week=sabbath_week_of(date), receipt_no=receipt,
        member=member, contributor_name=name,
        channel=(Envelope.Channel.BANK if channel == "BANK" else Envelope.Channel.CASH),
        recorded_by=user)
    svc = sabbath_of(date)   # the Sabbath this gift is counted under
    for line in lines:
        dept, amt = line[0], line[1]
        dev_group = line[2] if len(line) > 2 else None
        # Every envelope line must create exactly one ledger transaction so the
        # money reaches the cash book / collections — for BANK envelopes too.
        # (To receipt money ALREADY imported from the bank statement, use the
        # "receipt as envelope" action on that transaction, which links instead
        # of creating, so nothing is double-counted.)
        txn = Transaction.objects.create(
            date=date, sabbath_week=env.sabbath_week, service_sabbath=svc,
            channel=Transaction.Channel.ENVELOPE,
            direction=Transaction.Direction.CREDIT, amount=amt,
            department=dept, dev_group=dev_group, member=member, payer_name=name,
            reference=f"envelope {receipt}",
            allocation_status=Transaction.Status.MANUAL,
            raw_narration=f"ENVELOPE {receipt}")
        EnvelopeLine.objects.create(envelope=env, department=dept, amount=amt,
                                    dev_group=dev_group, transaction=txn)
    env.recompute_total()
    env.save(update_fields=["total"])
    send_receipt_sms(env, cfg)
    return env


class EnvelopeLedgerCreate(DataEntryRequiredMixin, View):
    """Excel-like ledger entry: choose the columns for the Sabbath, then key a
    row per contributor (autocomplete + auto-incrementing receipt numbers)."""
    template_name = "envelopes/ledger.html"

    def get(self, request):
        from departments.models import DevelopmentGroup, Department
        dev = Department.objects.filter(category="DEVELOPMENT",
                                        parent__isnull=True).first()
        return render(request, self.template_name, {
            "columns": column_catalog(),
            "default_date": request.GET.get("date") or _last_saturday().isoformat(),
            "dev_groups": DevelopmentGroup.objects.filter(active=True).order_by("number"),
            "dev_fund_key": str(dev.id) if dev else "",
        })

    @db_tx.atomic
    def post(self, request):
        from giving.models import SplitFund
        cfg = SiteConfig.get()
        funds = {d.id: d for d in Department.objects.filter(active=True)}
        splits = {s.id: s for s in SplitFund.objects.filter(active=True)}
        try:
            sab = dt.date.fromisoformat(request.POST.get("date"))
        except (TypeError, ValueError):
            messages.error(request, "Please choose a valid Sabbath date.")
            return redirect("envelope_ledger")
        from core.models import entry_blocked
        _why = entry_blocked(sab)
        if _why:
            messages.error(request, _why)
            return redirect("envelope_ledger")
        try:
            rows = json.loads(request.POST.get("rows") or "[]")
        except json.JSONDecodeError:
            rows = []

        from departments.models import DevelopmentGroup
        created = 0
        for row in rows:
            name = (row.get("name") or "").strip()
            receipt = (row.get("receipt") or "").strip()
            channel = (row.get("channel") or "CASH").upper()
            dev_group = None
            if row.get("dev_group_id"):
                dev_group = DevelopmentGroup.objects.filter(
                    pk=row["dev_group_id"]).first()
            lines = _expand_lines(row.get("amounts") or {}, funds, splits,
                                  dev_group=dev_group)
            if not name or not receipt or not lines:
                continue
            if Envelope.objects.filter(receipt_no=receipt).exists():
                messages.warning(request, f"Receipt {receipt} already used — skipped {name}.")
                continue
            member = None
            if row.get("member_id"):
                member = Member.objects.filter(pk=row["member_id"]).first()
            if member is None:
                member = Member.objects.filter(name__iexact=name).first()
            _save_envelope(date=sab, name=name, receipt=receipt, channel=channel,
                           lines=lines, member=member, user=request.user, cfg=cfg)
            created += 1

        if created:
            messages.success(request, f"Recorded {created} envelope(s) for {sab:%d %b %Y}.")
        else:
            messages.warning(request, "Nothing recorded — add at least one named row with an amount.")
        return redirect(f"{reverse('envelope_list')}?date={sab.isoformat()}")


class EnvelopeTemplateView(DataEntryRequiredMixin, View):
    """Download an .xlsx template for the chosen columns, to fill in offline."""

    def get(self, request):
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        keys = [k for k in (request.GET.get("cols") or "").split(",") if k]
        catalog = {c["key"]: c for c in column_catalog(for_import=True)}
        chosen = [catalog[k] for k in keys if k in catalog] or \
                 [c for c in column_catalog(for_import=True) if c["default"]]
        headers = ["No", "Contributor Name", "Phone", "Receipt No", "Channel", "Group"] + \
                  [c["label"] for c in chosen]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Envelopes"
        ws.append(headers)
        bold = Font(bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor="1F5F4F")
        for col, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.font = bold
            cell.fill = fill
        # sample row + blank rows
        ws.append([1, "EXAMPLE: Jane Doe", "0712345678", "", "CASH", ""] +
                  [100 if i == 0 else "" for i in range(len(chosen))])
        for n in range(2, 30):
            ws.append([n] + [""] * (len(headers) - 1))
        ws.column_dimensions["B"].width = 26
        for i in range(len(headers)):
            ws.column_dimensions[chr(65 + i)].width = max(ws.column_dimensions[chr(65 + i)].width or 12, 14)
        ws.freeze_panes = "A2"
        buf = io.BytesIO()
        wb.save(buf)
        from django.http import HttpResponse
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = 'attachment; filename="envelope_template.xlsx"'
        return resp


class EnvelopeImportView(DataEntryRequiredMixin, View):
    """Upload a filled envelope template and create the envelopes."""
    template_name = "envelopes/import.html"

    def get(self, request):
        return render(request, self.template_name, {
            "columns": column_catalog(for_import=True),
            "default_date": _last_saturday().isoformat(),
        })

    @db_tx.atomic
    def post(self, request):
        if request.POST.get("resolve"):
            return self._resolve(request)
        try:
            sab = dt.date.fromisoformat(request.POST.get("date"))
        except (TypeError, ValueError):
            messages.error(request, "Choose the Sabbath date for this sheet.")
            return redirect("envelope_import")
        from core.models import entry_blocked
        _why = entry_blocked(sab)
        if _why:
            messages.error(request, f"Import rejected: {_why}")
            return redirect("envelope_import")
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "Choose an .xlsx file to import.")
            return redirect("envelope_import")
        content = f.read()
        # Item 7: detect fund columns in the sheet that don't match a known fund.
        # Rather than silently dropping them, ask the treasurer to map or create.
        unknown, err = self._scan_unknown(content)
        if err:
            messages.error(request, err)
            return redirect("envelope_import")
        if unknown:
            import base64
            request.session["env_import_b64"] = base64.b64encode(content).decode("ascii")
            request.session["env_import_date"] = sab.isoformat()
            return render(request, self.template_name, {
                "stage": "resolve", "unknown_cols": unknown, "sab": sab,
                "funds": Department.objects.filter(active=True, selectable=True).order_by("name"),
            })
        return self._import(request, sab, content)

    def _scan_unknown(self, content):
        """Return (unknown_columns, error). An unknown column is a header that is
        neither a recognised meta column nor a known fund, but does carry numbers
        (so it's a fund the sheet expects but we don't have)."""
        import openpyxl
        label_to_key = self._fund_label_map()
        meta = {"contributor name", "name", "phone", "receipt no", "receipt",
                "channel", "group", "group number", "dev group",
                "no", "no.", "s/no", "sno", "s/n", "#", "serial", "index", "row"}
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception:
            return None, "Could not read that file — is it a valid .xlsx?"
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None, "The sheet is empty."
        header = [str(h).strip() if h is not None else "" for h in rows[0]]
        unknown = []
        for i, h in enumerate(header):
            hl = h.lower()
            if not hl or hl in meta or hl in label_to_key:
                continue
            has_num = False
            for r in rows[1:]:
                if i < len(r) and r[i] not in (None, ""):
                    try:
                        float(str(r[i]).replace(",", ""))
                        has_num = True
                        break
                    except (ValueError, TypeError):
                        continue
            if has_num:
                unknown.append({"index": i, "name": h})
        return unknown, None

    def _fund_label_map(self):
        label_to_key = {c["label"].lower(): c["key"] for c in column_catalog(for_import=True)}
        for c in column_catalog(for_import=True):
            label_to_key.setdefault(c["name"].lower(), c["key"])
        return label_to_key

    def _resolve(self, request):
        """Apply the treasurer's decisions for unknown columns, then import."""
        import base64
        b64 = request.session.get("env_import_b64")
        sab_raw = request.session.get("env_import_date")
        if not b64 or not sab_raw:
            messages.error(request, "Your import session expired — please upload again.")
            return redirect("envelope_import")
        content = base64.b64decode(b64)
        sab = dt.date.fromisoformat(sab_raw)
        extra_cols = {}      # {col_index: department_id}
        created_funds = 0
        for k in list(request.POST.keys()):
            if not k.startswith("col_"):
                continue
            try:
                i = int(k[4:])
            except ValueError:
                continue
            choice = request.POST.get(k, "ignore")
            if choice.startswith("existing:"):
                try:
                    extra_cols[i] = int(choice.split(":", 1)[1])
                except ValueError:
                    pass
            elif choice == "create":
                name = (request.POST.get(f"name_{i}") or "").strip()
                if name:
                    dept = Department.objects.filter(name__iexact=name).first()
                    if not dept:
                        dept = Department.objects.create(
                            name=name, fund_type=Department.FundType.LOCAL,
                            category=Department.Category.OFFERING, selectable=True)
                        created_funds += 1
                    extra_cols[i] = dept.id
            # "ignore" -> leave out
        request.session.pop("env_import_b64", None)
        request.session.pop("env_import_date", None)
        if created_funds:
            messages.success(request, f"Created {created_funds} new fund(s).")
        return self._import(request, sab, content, extra_cols=extra_cols)

    def _import(self, request, sab, content, extra_cols=None):
        import openpyxl
        from giving.models import SplitFund
        extra_cols = extra_cols or {}
        cfg = SiteConfig.get()
        funds = {d.id: d for d in Department.objects.filter(active=True)}
        splits = {s.id: s for s in SplitFund.objects.filter(active=True)}
        label_to_key = self._fund_label_map()

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception:
            messages.error(request, "Could not read that file — is it a valid .xlsx?")
            return redirect("envelope_import")
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            messages.warning(request, "The sheet is empty.")
            return redirect("envelope_import")

        header = [str(h).strip() if h is not None else "" for h in rows[0]]
        idx = {h.lower(): i for i, h in enumerate(header)}
        name_i = idx.get("contributor name", idx.get("name"))
        phone_i = idx.get("phone")
        rcpt_i = idx.get("receipt no", idx.get("receipt"))
        chan_i = idx.get("channel")
        group_i = idx.get("group", idx.get("group number", idx.get("dev group")))
        fund_cols = []  # (col_index, key)
        for h, i in idx.items():
            if h in label_to_key:
                fund_cols.append((i, label_to_key[h]))
        # resolved unknown columns map straight to a department id (its own key)
        for i, dept_id in extra_cols.items():
            if dept_id in funds:
                fund_cols.append((i, str(dept_id)))

        nums = []
        for r in Envelope.objects.values_list("receipt_no", flat=True):
            d = "".join(ch for ch in str(r) if ch.isdigit())
            if d:
                nums.append(int(d))
        next_no = (max(nums) + 1) if nums else 1

        created = skipped = 0
        for r in rows[1:]:
            def cell(i):
                return r[i] if (i is not None and i < len(r)) else None
            name = str(cell(name_i) or "").strip() if name_i is not None else ""
            if not name or name.upper().startswith("EXAMPLE"):
                continue
            amounts = {}
            for i, key in fund_cols:
                v = cell(i)
                if v not in (None, ""):
                    amounts[key] = v
            dev_group = None
            if group_i is not None:
                gv = cell(group_i)
                if gv not in (None, ""):
                    from departments.models import DevelopmentGroup
                    digits = "".join(ch for ch in str(gv) if ch.isdigit())
                    if digits:
                        dev_group = DevelopmentGroup.objects.filter(
                            number=int(digits)).first()
            lines = _expand_lines(amounts, funds, splits, dev_group=dev_group)
            if not lines:
                continue
            receipt = str(cell(rcpt_i) or "").strip() if rcpt_i is not None else ""
            if not receipt:
                receipt = str(next_no)
                next_no += 1
            if Envelope.objects.filter(receipt_no=receipt).exists():
                skipped += 1
                continue
            channel = (str(cell(chan_i) or "CASH").strip().upper()
                       if chan_i is not None else "CASH")
            phone = str(cell(phone_i) or "").strip() if phone_i is not None else ""
            member = Member.objects.filter(name__iexact=name).first()
            if member is None and phone:
                from members.models import normalize_phone
                member = Member.objects.filter(phone=normalize_phone(phone)).first()
            _save_envelope(date=sab, name=name, receipt=receipt, channel=channel,
                           lines=lines, member=member, user=request.user, cfg=cfg)
            created += 1

        msg = f"Imported {created} envelope(s) for {sab:%d %b %Y}."
        if skipped:
            msg += f" Skipped {skipped} with duplicate receipts."
        messages.success(request, msg)
        return redirect(f"{reverse('envelope_list')}?date={sab.isoformat()}")


class EnvelopeSabbathExcelView(ReadAccessMixin, View):
    """Download one Sabbath's envelopes as .xlsx with a trust/local summary."""

    def get(self, request):
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.page import PageMargins
        from core.models import SiteConfig
        try:
            sab = dt.date.fromisoformat(request.GET.get("date"))
        except (TypeError, ValueError):
            sab = _last_saturday()
        church = SiteConfig.get().church_name or "Church Treasury"
        envs = [e for e in Envelope.objects.select_related("member")
                .prefetch_related("lines__department").order_by("receipt_no")
                if sabbath_bucket(e.date) == sab]

        # Map each split-half department to the single concept it belongs to
        # (e.g. "Combined Offering (Trust 50%)" -> "Combined Offering"), so the
        # per-contributor entries table shows ONE block for the full amount given,
        # while the summary table can still split it into trust/local halves.
        from giving.models import SplitFund
        split_parent = {}          # department_id -> split fund name
        split_name_is_trust = {}   # split fund name -> True if it has a trust half
        for sf in SplitFund.objects.filter(active=True).prefetch_related(
                "components__department"):
            for comp in sf.components.all():
                split_parent[comp.department_id] = sf.name
                if comp.department.is_trust:
                    split_name_is_trust[sf.name] = True

        def _strip_suffix(receipt):
            # receipts are stored globally-unique as "<MON><sidx>-<original>"
            # (e.g. "JUN1-0421"); show only the original number to the reader.
            s = str(receipt or "")
            if "-" in s:
                head, _, tail = s.partition("-")
                # only strip a leading month/sabbath tag like JUN1, MAY2…
                if head[:3].isalpha() and any(ch.isdigit() for ch in head):
                    return tail or s
            return s

        # ---- entries-table columns: real funds, but split halves collapsed to
        # their single concept name ----
        present = {}   # key -> ("dept", dept) or ("split", name); preserves order info
        for e in envs:
            for l in e.lines.all():
                if l.department_id in split_parent:
                    name = split_parent[l.department_id]
                    present.setdefault(("split", name),
                                       split_name_is_trust.get(name, False))
                else:
                    present.setdefault(("dept", l.department_id), l.department)

        # build ordered display columns (trust first), each with a label + a test
        entry_cols = []   # list of dicts: {label, is_trust, match(line)->bool}
        for key, val in present.items():
            if key[0] == "split":
                nm = key[1]
                entry_cols.append({
                    "label": nm, "is_trust": val,
                    "ids": {d for d, n in split_parent.items() if n == nm}})
            else:
                d = val
                entry_cols.append({
                    "label": d.name, "is_trust": d.is_trust, "ids": {d.id}})
        entry_cols.sort(key=lambda c: (not c["is_trust"], c["label"]))

        # ---- summary funds: the REAL funds (halves stay separate), trust first ----
        present_real = {}
        for e in envs:
            for l in e.lines.all():
                present_real[l.department_id] = l.department
        funds = sorted(present_real.values(), key=lambda d: (not d.is_trust, d.name))

        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Sabbath"
        bold = Font(bold=True); white = Font(bold=True, color="FFFFFF")
        head_fill = PatternFill("solid", fgColor="1F5F4F")
        amber = PatternFill("solid", fgColor="B07D2C")
        grey = PatternFill("solid", fgColor="E8E2D4")
        thin = Side(style="thin", color="999999")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ncols = 4 + len(entry_cols) + 1

        # church name + report title
        ws.append([church])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.cell(1, 1).font = Font(bold=True, size=15, color="1F5F4F")
        ws.cell(1, 1).alignment = Alignment(horizontal="center")
        ws.append([f"Treasurer's Cash Statement — Sabbath {sab:%d %B %Y}"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        ws.cell(2, 1).font = Font(bold=True, size=12)
        ws.cell(2, 1).alignment = Alignment(horizontal="center")

        header = (["No", "Contributor", "Receipt", "Channel"]
                  + [c["label"] for c in entry_cols] + ["Total"])
        ws.append(header)
        hr = ws.max_row
        for c in range(1, len(header) + 1):
            cell = ws.cell(hr, c); cell.font = white; cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border

        first_data = ws.max_row + 1
        for i, e in enumerate(envs, start=1):
            # sum each contributor's lines into the collapsed display columns
            amt = {}
            for l in e.lines.all():
                amt[l.department_id] = amt.get(l.department_id, Decimal(0)) + l.amount
            row_cells = []
            for col in entry_cols:
                v = sum((amt.get(did, Decimal(0)) for did in col["ids"]), Decimal(0))
                row_cells.append(float(v) if v else "")
            ws.append([i, e.contributor_name, _strip_suffix(e.receipt_no),
                       e.get_channel_display()] + row_cells + [float(e.total)])

        # column totals row (over the collapsed display columns)
        totals = {f.id: Decimal(0) for f in funds}   # per REAL fund (for summary)
        grand = Decimal(0)
        for e in envs:
            for l in e.lines.all():
                if l.department_id in totals:
                    totals[l.department_id] += l.amount
            grand += e.total
        col_totals = []
        for col in entry_cols:
            col_totals.append(float(sum((totals.get(did, Decimal(0))
                                         for did in col["ids"]), Decimal(0))))
        ws.append(["", "TOTAL", "", ""] + col_totals + [float(grand)])
        last_data = ws.max_row
        for c in range(1, len(header) + 1):
            ws.cell(last_data, c).font = bold
            ws.cell(last_data, c).fill = grey

        # borders + number format across the whole data grid (incl totals row)
        for r in range(hr, last_data + 1):
            for c in range(1, len(header) + 1):
                cell = ws.cell(r, c)
                cell.border = border
                if c >= 5 and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"

        # ---- summary block ----
        ws.append([])
        summary_start = ws.max_row + 1
        ws.append(["SUMMARY"]); ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
        ws.append(["Trust fund items (remitted)"]); ws.cell(ws.max_row, 1).font = bold
        ws.cell(ws.max_row, 1).fill = amber
        ws.cell(ws.max_row, 2).fill = amber
        trust_total = Decimal(0)
        for f in funds:
            if f.is_trust:
                ws.append([f.name, float(totals[f.id])]); trust_total += totals[f.id]
        ws.append(["Total trust funds", float(trust_total)]); ws.cell(ws.max_row, 1).font = bold
        ws.cell(ws.max_row, 2).font = bold; ws.cell(ws.max_row, 1).fill = grey
        ws.cell(ws.max_row, 2).fill = grey
        ws.append(["Local fund items (retained)"]); ws.cell(ws.max_row, 1).font = bold
        ws.cell(ws.max_row, 1).fill = amber; ws.cell(ws.max_row, 2).fill = amber
        local_total = Decimal(0)
        for f in funds:
            if not f.is_trust:
                ws.append([f.name, float(totals[f.id])]); local_total += totals[f.id]
        ws.append(["Total local funds", float(local_total)]); ws.cell(ws.max_row, 1).font = bold
        ws.cell(ws.max_row, 2).font = bold; ws.cell(ws.max_row, 1).fill = grey
        ws.cell(ws.max_row, 2).fill = grey
        ws.append(["GRAND TOTAL GIVEN", float(grand)])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
        ws.cell(ws.max_row, 2).font = Font(bold=True, size=12)
        summary_end = ws.max_row

        # borders + number format on the summary table (both columns)
        for r in range(summary_start, summary_end + 1):
            for c in (1, 2):
                cell = ws.cell(r, c)
                cell.border = border
                if c == 2 and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 12
        for i in range(4, len(header)):
            col = openpyxl.utils.get_column_letter(i + 1)
            ws.column_dimensions[col].width = max(ws.column_dimensions[col].width or 10, 13)
        ws.freeze_panes = "A4"

        # ---- print setup: landscape, fit to one page wide, repeat headers ----
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_options.horizontalCentered = True
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.5,
                                      header=0.3, footer=0.3)
        ws.print_title_rows = "1:3"        # church + title + column header on every page
        ws.oddFooter.center.text = f"{church} — Sabbath {sab:%d %b %Y} — Page &P of &N"

        buf = io.BytesIO(); wb.save(buf)
        from django.http import HttpResponse
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="sabbath_{sab}.xlsx"'
        return resp


class EnvelopeReceiptOneBankView(DataEntryRequiredMixin, View):
    """Receipt a SINGLE bank transaction (and any split siblings of the same contribution)
    as an envelope, on demand — the per-entry counterpart to the bulk monthly
    pull. Supports a user-supplied receipt number for hybrid manual receipting:
    the treasurer can enter the number written on the physical receipt/envelope so
    the system record and the hand-written one match. Leaving it blank auto-assigns
    the next 'B' number.

    Like the bulk flow, this never creates a second ledger posting — the bank
    transaction IS the money; the envelope just receipts it and the transaction is
    marked processed_via_envelope so it is not double-counted.
    """

    def get(self, request, pk):
        """Show a small confirmation form to receipt this bank contribution, with an
        optional manual receipt number."""
        txn = get_object_or_404(Transaction, pk=pk)
        eligible = (txn.channel == Transaction.Channel.BANK
                    and txn.direction == Transaction.Direction.CREDIT
                    and txn.department_id and not txn.processed_via_envelope
                    and not txn.manual_receipt
                    and not txn.sabbath_confirm_pending
                    and not txn.is_reversed and not txn.is_reversal)
        nums = []
        for r in Envelope.objects.values_list("receipt_no", flat=True):
            d = "".join(ch for ch in str(r) if ch.isdigit())
            if d:
                nums.append(int(d))
        next_no = f"B{(max(nums) + 1) if nums else 1}"
        return render(request, "envelopes/receipt_bank.html", {
            "txn": txn, "eligible": eligible, "next_no": next_no})

    @db_tx.atomic
    def post(self, request, pk):
        from core.models import period_locked
        from core.utils import sabbath_of as _sof
        txn = get_object_or_404(Transaction, pk=pk)

        if txn.channel != Transaction.Channel.BANK or \
                txn.direction != Transaction.Direction.CREDIT:
            messages.error(request, "Only bank/M-Pesa credits can be receipted as envelopes.")
            return redirect("transaction_list")
        if txn.processed_via_envelope or hasattr(txn, "envelope"):
            messages.info(request, "That contribution has already been receipted as an envelope.")
            return redirect("transaction_list")
        if txn.manual_receipt:
            messages.info(request, "That contribution is marked as a manual (paper) receipt. "
                                   "Untick 'manual receipt' on the entry first if you "
                                   "want to issue a system receipt instead.")
            return redirect("transaction_list")
        if txn.department_id is None:
            messages.error(request, "Allocate this contribution to a fund before receipting it.")
            return redirect("transaction_list")
        if txn.sabbath_confirm_pending:
            messages.error(request, "Confirm this contribution's Sabbath before receipting it.")
            return redirect("transaction_list")
        lk = period_locked(txn.service_sabbath or txn.date)
        if lk:
            messages.error(request, f"{lk} is locked — reopen the period first.")
            return redirect("transaction_list")

        # gather split siblings of the same gift so the receipt covers the full
        # amount — but never pull in a sibling that is already receipted (flagged)
        # or already has an envelope record, so we can't double-receipt a part.
        base_ref = ((txn.core_ref or "").split("-S")[0] or txn.mpesa_ref or str(txn.id))
        siblings = list(Transaction.objects.filter(
            channel=Transaction.Channel.BANK,
            direction=Transaction.Direction.CREDIT,
            processed_via_envelope=False, manual_receipt=False,
            department__isnull=False,
            date=txn.date, payer_name=txn.payer_name)
            .filter(envelope__isnull=True, envelope_lines__isnull=True)
            .select_related("department", "member"))
        txns = [t for t in siblings
                if ((t.core_ref or "").split("-S")[0] or t.mpesa_ref or str(t.id)) == base_ref] or [txn]

        # "mark only" — the envelope was already written/typed by hand, so just flag
        # "mark only" — the envelope was already written/typed by hand, so flag
        # the bank entry as a MANUAL RECEIPT (paper): keep it out of the review
        # queue and the receipt-bank-giving pull WITHOUT creating a system
        # envelope. Income is unaffected — the bank transaction is the income.
        if request.POST.get("mark_only"):
            for t in txns:
                t.mark_manual_receipt(value=True, cascade_split=False)
            messages.success(request, f"Marked as a manual (paper) receipt "
                                      f"(KSh {sum(t.amount for t in txns):,.2f}). No "
                                      f"system envelope was created.")
            return redirect("transaction_list")

        # receipt number: user-supplied (hybrid manual) or auto-assigned
        manual_no = (request.POST.get("receipt_no") or "").strip()
        if manual_no:
            if Envelope.objects.filter(receipt_no=manual_no).exists():
                messages.error(request, f"Receipt number {manual_no} is already used. "
                                        "Choose a different number.")
                return redirect("transaction_list")
            receipt_no = manual_no
        else:
            nums = []
            for r in Envelope.objects.values_list("receipt_no", flat=True):
                d = "".join(ch for ch in str(r) if ch.isdigit())
                if d:
                    nums.append(int(d))
            receipt_no = f"B{(max(nums) + 1) if nums else 1}"

        member = next((t.member for t in txns if t.member), None)
        sab_date = next((t.service_sabbath for t in txns if t.service_sabbath),
                        _sof(txn.date))
        env = Envelope.objects.create(
            date=sab_date, sabbath_week=sabbath_week_of(sab_date),
            receipt_no=receipt_no, member=member,
            contributor_name=txn.payer_name or (member.name if member else "(bank)"),
            channel=Envelope.Channel.BANK,
            bank_transaction=txns[0], recorded_by=request.user)
        for t in txns:
            EnvelopeLine.objects.create(envelope=env, department=t.department,
                                        amount=t.amount, transaction=t)
            t.processed_via_envelope = True
            t.save(update_fields=["processed_via_envelope"])
        env.recompute_total()
        env.save(update_fields=["total"])
        send_receipt_sms(env, SiteConfig.get())
        messages.success(request, f"Receipted as envelope #{receipt_no} "
                                  f"(KSh {env.total:,.2f}). You can now print or write it.")
        return redirect("transaction_list")


class EnvelopePullBankView(DataEntryRequiredMixin, View):
    """Create envelope records from auto/learned-allocated BANK giving so that
    tithe, combined offering, etc. also appear as envelopes — without creating a
    second ledger posting (the bank transaction is the money; the envelope just
    receipts it and is marked accounted-for)."""

    @db_tx.atomic
    def post(self, request):
        from collections import defaultdict
        today = dt.date.today()
        from core.utils import sabbath_of as _sof
        # Item 4: an optional specific Sabbath to receipt. If given, we receipt
        # only that Sabbath's bank giving; if not, the original whole-month logic
        # is used.
        raw_sab = (request.POST.get("sabbath") or "").strip()
        one_sabbath = None
        if raw_sab:
            try:
                one_sabbath = _sof(dt.date.fromisoformat(raw_sab))
            except ValueError:
                one_sabbath = None

        raw = request.POST.get("month")
        try:
            year, month = (int(x) for x in raw.split("-")) if raw else (today.year, today.month)
        except (ValueError, AttributeError):
            year, month = today.year, today.month
        if one_sabbath:
            year, month = one_sabbath.year, one_sabbath.month

        from core.models import period_locked
        _lk = period_locked(one_sabbath or dt.date(year, month, 1))
        if _lk:
            messages.error(request, f"{_lk} is locked — reopen the period before "
                                    "receipting bank giving into it.")
            return redirect(f"/envelopes/?month={year}-{month:02d}")

        # Select by the SERVICE SABBATH month, not the transaction date: a gift
        # dated late in a month (or rolled to the next Sabbath) is receipted under
        # the Sabbath it is counted on, which may fall in a different month.
        # Fall back to the transaction date for gifts with no service Sabbath set.
        from django.db.models import Q as _Q
        if one_sabbath:
            # a single Sabbath: match its exact service_sabbath, or (for gifts with
            # no Sabbath set) a transaction date within that Sabbath's week
            week_start = one_sabbath - dt.timedelta(days=6)
            period_q = (_Q(service_sabbath=one_sabbath) |
                        _Q(service_sabbath__isnull=True,
                           date__range=(week_start, one_sabbath)))
        else:
            period_q = (_Q(service_sabbath__year=year, service_sabbath__month=month) |
                        _Q(service_sabbath__isnull=True, date__year=year, date__month=month))
        eligible = (Transaction.objects.filter(
            channel=Transaction.Channel.BANK,
            direction=Transaction.Direction.CREDIT,
            allocation_status__in=[Transaction.Status.AUTO, Transaction.Status.LEARNED,
                                   Transaction.Status.MANUAL],
            processed_via_envelope=False, manual_receipt=False,
            department__isnull=False,
            # gifts still awaiting a Sabbath decision are not receipted yet
            sabbath_confirm_pending=False)
            .filter(period_q)
            # belt-and-braces: never re-receipt a gift that already has an envelope
            # record, even if its processed flag somehow wasn't set (older data,
            # a manual envelope, or a partially-receipted split). Matching on the
            # flag alone misses these and would create a duplicate receipt.
            .filter(envelope__isnull=True, envelope_lines__isnull=True)
            .distinct())
        cfg = SiteConfig.get()


        # group transactions that belong to one gift (split offerings share these).
        # NOTE: a sibling created by split_into shares the core_ref base, but may
        # not share mpesa_ref-vs-core_ref keying — normalise to the core_ref base
        # first so every part of one gift lands in the same envelope.
        groups = defaultdict(list)
        for t in eligible.select_related("department", "member"):
            base_ref = ((t.core_ref or "").split("-S")[0] or t.mpesa_ref or str(t.id))
            groups[(t.date, t.payer_name, base_ref)].append(t)

        # The trust-only scope applies per GIFT, not per line: if any part of a
        # gift is a trust fund, the WHOLE gift (all its funds) is receipted as one
        # envelope so the receipt accounts for the full amount given.
        if cfg.receipt_bank_scope == SiteConfig.ReceiptBankScope.TRUST_ONLY:
            groups = {k: ts for k, ts in groups.items()
                      if any(t.department and t.department.is_trust for t in ts)}

        # next receipt number — honour a user-supplied starting number if given
        # (optional), otherwise continue from the highest existing receipt.
        start_raw = (request.POST.get("start_receipt") or "").strip()
        start_digits = "".join(ch for ch in start_raw if ch.isdigit())
        prefix = "".join(ch for ch in start_raw if not ch.isdigit()) or "B"
        if start_digits:
            nxt = int(start_digits)
        else:
            nums = []
            for r in Envelope.objects.values_list("receipt_no", flat=True):
                d = "".join(ch for ch in str(r) if ch.isdigit())
                if d:
                    nums.append(int(d))
            nxt = (max(nums) + 1) if nums else 1

        created = 0
        skipped_dupe = 0
        for (gdate, payer, _), txns in groups.items():
            member = next((t.member for t in txns if t.member), None)
            # the envelope (and its receipt) belongs to the gift's service Sabbath
            sab_date = next((t.service_sabbath for t in txns if t.service_sabbath),
                            _sof(gdate))
            # find the next free receipt number from `nxt` (skip any already used)
            while Envelope.objects.filter(receipt_no=f"{prefix}{nxt}").exists():
                nxt += 1
            env = Envelope.objects.create(
                date=sab_date, sabbath_week=sabbath_week_of(sab_date),
                receipt_no=f"{prefix}{nxt}", member=member,
                contributor_name=payer or (member.name if member else "(bank)"),
                channel=Envelope.Channel.BANK,
                bank_transaction=txns[0], recorded_by=request.user)
            nxt += 1
            for t in txns:
                EnvelopeLine.objects.create(envelope=env, department=t.department,
                                            amount=t.amount, transaction=t)
                t.processed_via_envelope = True
                t.save(update_fields=["processed_via_envelope"])
            env.recompute_total()
            env.save(update_fields=["total"])
            send_receipt_sms(env, cfg)
            created += 1

        if created:
            messages.success(
                request, f"Receipted {created} bank contribution(s) as envelopes "
                         f"(accounted once — not double-counted).")
        else:
            messages.info(request, "No new bank giving to receipt for this month.")
        return redirect(f"{reverse('envelope_list')}?month={year}-{month:02d}")


DEFAULT_RECEIPT_MSG = ("Thank you for your faithful giving. \u201cBring ye all the tithes "
                       "into the storehouse\u2026\u201d \u2014 Malachi 3:10. May God bless you.")


class EnvelopeReceiptView(ReadAccessMixin, DetailView):
    """A printable contribution receipt for one envelope — standard or compact ETR
    (thermal) format via ?format=etr."""
    model = Envelope
    template_name = "envelopes/receipt.html"
    context_object_name = "envelope"

    def get_context_data(self, **kwargs):
        from core.models import SiteConfig
        ctx = super().get_context_data(**kwargs)
        cfg = SiteConfig.get()
        ctx["fmt"] = "etr" if self.request.GET.get("format") == "etr" else "standard"
        ctx["receipt_message"] = (cfg.receipt_message or "").strip() or DEFAULT_RECEIPT_MSG
        ctx["auto_print"] = self.request.GET.get("print") == "1"
        return ctx


class EnvelopeBulkReceiptsView(ReadAccessMixin, View):
    """Print every envelope receipt for one Sabbath at once. ?format=etr gives a
    continuous run of 72mm thermal receipts; the default A4 layout tiles several
    compact receipts per page with cut lines (for churches with only a plain
    printer)."""
    template_name = "envelopes/receipts_bulk.html"

    def get(self, request):
        from core.models import SiteConfig
        try:
            sabbath = dt.date.fromisoformat(request.GET.get("date", ""))
        except ValueError:
            sabbath = _last_saturday()
        sabbath = sabbath_of(sabbath)
        window_start = sabbath - dt.timedelta(days=6)
        envelopes = [e for e in Envelope.objects.filter(
                         date__range=(window_start, sabbath))
                     .select_related("member").prefetch_related("lines__department")
                     .order_by("receipt_no")
                     if sabbath_of(e.date) == sabbath]
        cfg = SiteConfig.get()
        fmt = "etr" if request.GET.get("format") == "etr" else "a4"
        return render(request, self.template_name, {
            "envelopes": envelopes, "sabbath": sabbath, "fmt": fmt,
            "receipt_message": (cfg.receipt_message or "").strip() or DEFAULT_RECEIPT_MSG,
            "count": len(envelopes),
            "total": sum((e.total for e in envelopes), Decimal(0)),
            "auto_print": request.GET.get("print") == "1",
        })


class CountSessionListView(ReadAccessMixin, ListView):
    template_name = "envelopes/count_list.html"
    context_object_name = "sessions"
    paginate_by = 30

    def get_queryset(self):
        from envelopes.models import CountSession
        return CountSession.objects.prefetch_related("witnesses").all()


class CountSessionCreate(DataEntryRequiredMixin, View):
    """Record a Sabbath cash count: denomination breakdown + witnesses, compared
    against the system's expected receipts for the Sabbath."""
    template_name = "envelopes/count_form.html"
    DENOMS = [1000, 500, 200, 100, 50, 40, 20, 10, 5, 1]

    def _expected(self, sabbath):
        b = self._breakdown(sabbath)
        return b["net"]

    def _breakdown(self, sabbath):
        """Transparent components of expected cash on hand for the Sabbath. Groups
        by each contribution's *service Sabbath*, so a contribution sent after the count closed is
        credited to the next Sabbath and never reopens this one.

        Crucially, this is a count of PHYSICAL CASH only. Bank/M-Pesa giving lives
        on the statement (channel=BANK) and is never included. We also subtract any
        'cash envelope' (an ENVELOPE-channel row) that turns out to duplicate a
        bank contribution for the same contributor this Sabbath — i.e. money that arrived
        in the bank but was also keyed on the cash envelope sheet. Counting it as
        cash would overstate the float and make the count impossible to balance."""
        from django.db.models import Sum, Case, When, DecimalField
        from giving.models import Transaction
        from cashbook.models import Expense
        from members.services.matching import name_key
        zero = DecimalField(max_digits=14, decimal_places=2)
        agg = (Transaction.objects.filter(
                    channel__in=[Transaction.Channel.CASH, Transaction.Channel.ENVELOPE],
                    direction=Transaction.Direction.CREDIT, confirmed=True,
                    is_reversed=False, is_reversal=False,
                    excluded_from_income=False,
                    service_sabbath=sabbath)
               .aggregate(
                    cash=Sum(Case(When(channel=Transaction.Channel.CASH, then="amount"),
                                  default=Decimal(0), output_field=zero)),
                    envelope=Sum(Case(When(channel=Transaction.Channel.ENVELOPE, then="amount"),
                                      default=Decimal(0), output_field=zero))))
        cash = agg["cash"] or Decimal(0)
        envelope = agg["envelope"] or Decimal(0)

        # --- exclude bank giving that was also keyed as a cash envelope ---------
        # Build a multiset of this Sabbath's BANK gifts by (contributor, amount);
        # any ENVELOPE-channel row that matches one is the same money entered twice
        # and must not count toward physical cash.
        bank_sig = {}
        for b in Transaction.objects.filter(
                channel=Transaction.Channel.BANK,
                direction=Transaction.Direction.CREDIT, confirmed=True,
                is_reversed=False, is_reversal=False,
                service_sabbath=sabbath).values("member_id", "payer_name", "amount"):
            who = b["member_id"] or name_key(b["payer_name"]) or ""
            bank_sig[(who, b["amount"])] = bank_sig.get((who, b["amount"]), 0) + 1
        bank_as_cash = Decimal(0)
        if bank_sig:
            for e in Transaction.objects.filter(
                    channel=Transaction.Channel.ENVELOPE,
                    direction=Transaction.Direction.CREDIT, confirmed=True,
                    is_reversed=False, is_reversal=False,
                    excluded_from_income=False,
                    service_sabbath=sabbath).values("member_id", "payer_name", "amount"):
                who = e["member_id"] or name_key(e["payer_name"]) or ""
                key = (who, e["amount"])
                if bank_sig.get(key, 0) > 0:
                    bank_sig[key] -= 1            # consume one match
                    bank_as_cash += e["amount"]
        envelope_cash = envelope - bank_as_cash

        window_start = sabbath - dt.timedelta(days=6)
        disbursed = (Expense.objects.filter(
                        method=Expense.Method.CASH,
                        status__in=[Expense.Status.APPROVED, Expense.Status.PAID],
                        date__range=(window_start, sabbath))
                     .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        return {"cash": cash, "envelope": envelope_cash, "envelope_raw": envelope,
                "bank_as_cash": bank_as_cash, "disbursed": disbursed,
                "net": cash + envelope_cash - disbursed}

    def get(self, request):
        try:
            sabbath = sabbath_of(dt.date.fromisoformat(request.GET["date"]))
        except (KeyError, ValueError):
            sabbath = _last_saturday()
        return render(request, self.template_name, {
            "denoms": self.DENOMS, "sabbath": sabbath,
            "expected": self._expected(sabbath),
            "breakdown": self._breakdown(sabbath)})

    def post(self, request):
        from envelopes.models import CountSession, CountDenomination, CountWitness
        try:
            sabbath = sabbath_of(dt.date.fromisoformat(request.POST["date"]))
        except (KeyError, ValueError):
            messages.error(request, "Choose the Sabbath being counted.")
            return redirect("count_new")
        counted = Decimal(0)
        denom_rows = []
        for d in self.DENOMS:
            try:
                qty = int(request.POST.get(f"d_{d}") or 0)
            except ValueError:
                qty = 0
            if qty:
                denom_rows.append((Decimal(d), qty))
                counted += Decimal(d) * qty
        expected = self._expected(sabbath)
        with db_tx.atomic():
            cs = CountSession.objects.create(
                date=sabbath, counted_total=counted, expected_total=expected,
                note=(request.POST.get("note") or "")[:200], recorded_by=request.user)
            for denom, qty in denom_rows:
                CountDenomination.objects.create(session=cs, denomination=denom, count=qty)
            names = request.POST.getlist("w_name")
            roles = request.POST.getlist("w_role")
            for i, nm in enumerate(names):
                nm = (nm or "").strip()
                if nm:
                    CountWitness.objects.create(session=cs, name=nm,
                        role=(roles[i] if i < len(roles) else "")[:60],
                        signed=bool(request.POST.get(f"w_signed_{i}")))
        if cs.has_discrepancy:
            messages.warning(request, f"Count saved with a discrepancy of "
                                      f"{cs.discrepancy:,.2f} vs expected {expected:,.2f}.")
        else:
            messages.success(request, "Count saved — it matches the expected receipts.")
        return redirect("count_detail", pk=cs.pk)


class CountSessionDetail(ReadAccessMixin, DetailView):
    from envelopes.models import CountSession as _CS
    model = _CS
    template_name = "envelopes/count_detail.html"
    context_object_name = "cs"

    def get_context_data(self, **kwargs):
        from core.models import SabbathClose
        ctx = super().get_context_data(**kwargs)
        ctx["sabbath_closed"] = SabbathClose.objects.filter(
            sabbath=self.object.date).first()
        return ctx


class EnvelopeUpdateView(DataEntryRequiredMixin, View):
    """Edit an existing envelope: contributor, receipt, and each fund line's amount
    and development group. Keeps the ledger in step with the changes."""
    template_name = "envelopes/edit.html"

    def get(self, request, pk):
        from envelopes.models import Envelope
        from departments.models import DevelopmentGroup
        env = get_object_or_404(Envelope, pk=pk)
        return render(request, self.template_name, {
            "env": env,
            "lines": env.lines.select_related("department", "dev_group").all(),
            "dev_groups": DevelopmentGroup.objects.filter(active=True).order_by("number"),
        })

    @db_tx.atomic
    def post(self, request, pk):
        from envelopes.models import Envelope, EnvelopeLine
        from departments.models import DevelopmentGroup
        env = get_object_or_404(Envelope, pk=pk)
        from core.models import entry_blocked
        _why = entry_blocked(env.date)
        if _why:
            messages.error(request, _why)
            return redirect("envelope_detail", pk=pk)
        name = (request.POST.get("contributor_name") or "").strip()
        receipt = (request.POST.get("receipt_no") or "").strip()
        if name:
            env.contributor_name = name
        if receipt and receipt != env.receipt_no:
            if Envelope.objects.filter(receipt_no=receipt).exclude(pk=env.pk).exists():
                messages.error(request, f"Receipt {receipt} is already in use.")
                return redirect("envelope_edit", pk=pk)
            env.receipt_no = receipt
        env.save(update_fields=["contributor_name", "receipt_no"])
        for line in env.lines.select_related("transaction").all():
            raw = request.POST.get(f"amount_{line.id}")
            if raw is None:
                continue
            amt = _amount(raw)
            gid = request.POST.get(f"group_{line.id}") or ""
            grp = DevelopmentGroup.objects.filter(pk=gid).first() if gid else None
            if not amt or amt == 0:
                # zeroed out: remove the line and its ledger entry
                if line.transaction_id and line.transaction.channel == "ENVELOPE":
                    line.transaction.delete()
                line.delete()
                continue
            line.amount = amt
            line.dev_group = grp
            line.save(update_fields=["amount", "dev_group"])
            t = line.transaction
            if t is not None and t.channel == "ENVELOPE":
                t.amount = amt
                t.dev_group = grp
                t.payer_name = env.contributor_name
                t.save(update_fields=["amount", "dev_group", "payer_name"])
        env.recompute_total()
        env.save(update_fields=["total"])
        messages.success(request, f"Envelope #{env.receipt_no} updated.")
        return redirect("envelope_detail", pk=pk)


class EnvelopeDeleteView(DataEntryRequiredMixin, View):
    """Delete an envelope entered in error. Removes the cash ledger entries it
    created; for a bank envelope it unlinks (but keeps) the bank deposit row."""
    def post(self, request, pk):
        from envelopes.models import Envelope
        env = get_object_or_404(Envelope, pk=pk)
        from core.models import entry_blocked
        _why = entry_blocked(env.date)
        if _why:
            messages.error(request, _why)
            return redirect(f"{reverse('envelope_list')}?date={env.date.isoformat()}")
        removed = 0
        for line in env.lines.select_related("transaction").all():
            t = line.transaction
            if t is not None:
                # detach then delete the ENVELOPE entry this line created
                line.transaction = None
                line.save(update_fields=["transaction"])
                if t.channel == t.Channel.ENVELOPE:
                    t.delete()
                    removed += 1
        # a bank envelope's deposit row is real money — keep it, just unlink
        if env.bank_transaction_id:
            env.bank_transaction = None
            env.save(update_fields=["bank_transaction"])
        sab = env.date.isoformat()
        env.delete()
        messages.success(request, f"Envelope deleted (removed {removed} ledger entr"
                                  f"{'y' if removed == 1 else 'ies'}).")
        return redirect(f"{reverse('envelope_list')}?date={sab}")


class EnvelopeSendReceiptView(DataEntryRequiredMixin, View):
    """Send one envelope's receipt by SMS or WhatsApp to the contributor's phone,
    as an alternative to printing."""
    def post(self, request, pk):
        from envelopes.models import Envelope
        from core.services.sms import build_receipt_text, send_sms
        env = get_object_or_404(Envelope, pk=pk)
        cfg = SiteConfig.get()
        channel = request.POST.get("channel", "sms")
        phone = env.member.receipt_phone if env.member_id else None
        back = f"{reverse('envelope_list')}?date={env.date.isoformat()}"
        if not phone:
            messages.error(request, "No phone number on file for this contributor "
                                    "(link a member with a phone to send a receipt).")
            return redirect(back)
        msg = build_receipt_text(env, cfg)
        if channel == "whatsapp":
            from core.services.whatsapp import send_whatsapp
            ok, detail = send_whatsapp(phone, msg, cfg)
        else:
            log = send_sms(phone, msg, cfg)
            ok = bool(log and log.status == "SENT")
            detail = (log.response[:120] if log else "SMS is not configured.")
        if ok:
            env.sms_sent = True
            env.save(update_fields=["sms_sent"])
            messages.success(request, f"Receipt sent by {channel.upper()} to {phone}.")
        else:
            messages.error(request, f"Could not send by {channel.upper()}: {detail}")
        return redirect(back)


class EnvelopeBulkSendView(DataEntryRequiredMixin, View):
    """Send all receipts for a Sabbath by SMS or WhatsApp (background, spaced out)."""
    def post(self, request):
        import threading
        from envelopes.models import Envelope
        channel = request.POST.get("channel", "sms")
        try:
            sabbath = sabbath_of(dt.date.fromisoformat(request.POST["date"]))
        except (KeyError, ValueError):
            sabbath = _last_saturday()
        cfg = SiteConfig.get()
        if channel == "whatsapp" and not cfg.whatsapp_enabled:
            messages.error(request, "WhatsApp isn't enabled in Settings → Channels.")
            return redirect(f"{reverse('envelope_list')}?date={sabbath.isoformat()}")
        if channel == "sms" and not cfg.sms_enabled:
            messages.error(request, "SMS isn't enabled in Settings → SMS.")
            return redirect(f"{reverse('envelope_list')}?date={sabbath.isoformat()}")
        window_start = sabbath - dt.timedelta(days=6)
        ids = [e.id for e in Envelope.objects.filter(
                   date__range=(window_start, sabbath), member__isnull=False)
               if sabbath_of(e.date) == sabbath and e.member.receipt_phone and not e.is_voided]
        if not ids:
            messages.info(request, "No contributors with a phone number for this Sabbath.")
            return redirect(f"{reverse('envelope_list')}?date={sabbath.isoformat()}")
        threading.Thread(target=self._send_all, args=(ids, channel), daemon=True).start()
        messages.success(request, f"Sending {len(ids)} receipt(s) by {channel.upper()} "
                                  f"in the background.")
        return redirect(f"{reverse('envelope_list')}?date={sabbath.isoformat()}")

    @staticmethod
    def _send_all(ids, channel):
        import time
        from django.db import connection
        from envelopes.models import Envelope
        from core.services.sms import build_receipt_text, send_sms
        from core.services.whatsapp import send_whatsapp
        cfg = SiteConfig.get()
        try:
            for i, eid in enumerate(ids):
                env = Envelope.objects.filter(pk=eid).select_related("member").first()
                if not env or not (env.member and env.member.receipt_phone):
                    continue
                msg = build_receipt_text(env, cfg)
                if channel == "whatsapp":
                    send_whatsapp(env.member.receipt_phone, msg, cfg)
                else:
                    send_sms(env.member.receipt_phone, msg, cfg)
                if i < len(ids) - 1:
                    time.sleep(2)
        finally:
            connection.close()


class SabbathCloseView(DataEntryRequiredMixin, View):
    """Close (or reopen) a Sabbath once counting is done. Closing fixes its
    offering/count figures; later contributions for it roll to the next open Sabbath."""
    def post(self, request):
        from core.models import SabbathClose
        try:
            sab = sabbath_of(dt.date.fromisoformat(request.POST["date"]))
        except (KeyError, ValueError):
            messages.error(request, "Choose a Sabbath to close.")
            return redirect("count_list")
        action = request.POST.get("action", "close")
        from django.utils.http import url_has_allowed_host_and_scheme
        back = request.POST.get("next") or reverse("count_list")
        if not url_has_allowed_host_and_scheme(back, allowed_hosts={request.get_host()}):
            back = reverse("count_list")
        if action == "reopen":
            from core.roles import is_treasurer
            if not is_treasurer(request.user):
                messages.error(request, "Only a treasurer can reopen a Sabbath.")
                return redirect(back)
            SabbathClose.objects.filter(sabbath=sab).delete()
            messages.success(request, f"Sabbath {sab:%d %b %Y} reopened.")
        else:
            SabbathClose.objects.get_or_create(
                sabbath=sab, defaults={"closed_by": request.user,
                                       "note": (request.POST.get("note") or "")[:200]})
            messages.success(request, f"Sabbath {sab:%d %b %Y} closed. Later contributions for "
                                      f"it will be credited to the next open Sabbath.")
        return redirect(back)


class SabbathReconciliationView(ReadAccessMixin, View):
    """Item 1: reconcile a Sabbath's bank giving (receipted + manual) against the
    envelopes counted for it, with fuzzy name matching and a balance check."""
    template_name = "envelopes/reconcile_sabbath.html"

    def get(self, request):
        from envelopes.reconcile import reconcile_sabbath, unsabbathed_bank_count
        try:
            sab = sabbath_of(dt.date.fromisoformat(request.GET["date"]))
        except (KeyError, ValueError):
            sab = _last_saturday()
        rec = reconcile_sabbath(sab)
        rec["unsabbathed"] = unsabbathed_bank_count()
        return render(request, self.template_name, rec)


class ReconcileApplyView(DataEntryRequiredMixin, View):
    """Item 1: apply selected reconciliation matches. Each selected pair is
    'env:<envelope_id>:bank:<txn_id>'. Applying marks the matched envelope as a
    BANK item (it was bank money), links it to the bank contribution when free, and
    neutralises the duplicate cash income the envelope created so the money is
    counted once — via the bank transaction."""

    @db_tx.atomic
    def post(self, request):
        from giving.models import Transaction
        pairs = request.POST.getlist("pair")
        sab = request.POST.get("date", "")
        applied = 0
        for token in pairs:
            try:
                _, env_id, _, txn_id = token.split(":")
                env_id, txn_id = int(env_id), int(txn_id)
            except (ValueError, AttributeError):
                continue
            env = Envelope.objects.filter(pk=env_id).first()
            txn = Transaction.objects.filter(pk=txn_id).first()
            if not env or not txn:
                continue
            changed = False
            # 1) mark the receipted entry (envelope) as a BANK item
            if env.channel != Envelope.Channel.BANK:
                env.channel = Envelope.Channel.BANK
                changed = True
            # 2) link it to the bank gift if neither side is already linked
            if env.bank_transaction_id is None and not hasattr(txn, "envelope"):
                env.bank_transaction = txn
                changed = True
            if changed:
                env.save(update_fields=["channel", "bank_transaction"])
            # 3) neutralise duplicate income: any ENVELOPE-channel transaction this
            #    envelope created is the same money as the bank gift — exclude it
            #    from income (the bank transaction is the income).
            for line in env.lines.select_related("transaction").all():
                lt = line.transaction
                if lt and lt.channel == Transaction.Channel.ENVELOPE \
                        and not lt.excluded_from_income:
                    lt.excluded_from_income = True
                    lt.save(update_fields=["excluded_from_income"])
            applied += 1
        if applied:
            messages.success(request, f"Applied {applied} match(es): marked as bank "
                                      "giving and removed the duplicate cash entry.")
        else:
            messages.info(request, "No matches were selected.")
        return redirect(f"{reverse('sabbath_reconcile')}?date={sab}")
