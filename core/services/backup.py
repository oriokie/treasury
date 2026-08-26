"""Backup and full-data export.

* database_backup(): the raw SQLite file, for a complete restorable snapshot.
* full_excel_export(): every operational table in one multi-sheet workbook,
  for analysis or off-system archival. Both are admin/treasurer-only and stream
  to the browser; nothing is written to disk on the server.
"""
import datetime as _dt
from decimal import Decimal

from django.conf import settings
from django.http import FileResponse, HttpResponse


def _mysql_host(db):
    """The app (PyMySQL) connects over TCP even for 'localhost'; the command-line
    dump/restore tools treat 'localhost' as a Unix socket, which can match a
    different grant and be denied even though the app authenticates fine. Force
    the same TCP host the app uses so authentication is consistent."""
    host = db.get("HOST") or "127.0.0.1"
    return "127.0.0.1" if host == "localhost" else host


def _write_mysql_defaults_file(db):
    """Write a temporary my.cnf-style [client] file carrying the app's exact
    credentials over TCP. Passing them via --defaults-extra-file is the reliable,
    documented way to authenticate the dump tool and it overrides any stray
    ~/.my.cnf that could otherwise supply the wrong user/password. Caller deletes
    the file. Returns its path."""
    import os
    import tempfile
    pw = str(db.get("PASSWORD", "")).replace("\\", "\\\\").replace('"', '\\"')
    fd = tempfile.NamedTemporaryFile("w", suffix=".cnf", delete=False)
    fd.write("[client]\n")
    fd.write(f"host={_mysql_host(db)}\n")
    fd.write(f"port={db.get('PORT') or 3306}\n")
    fd.write(f"user={db['USER']}\n")
    fd.write(f'password="{pw}"\n')
    fd.write("protocol=TCP\n")
    fd.close()
    os.chmod(fd.name, 0o600)
    return fd.name


def _mysql_tool(*names):
    """Resolve a MySQL/MariaDB command-line tool, preferring the modern MariaDB
    name (so we don't trip the 'mysqldump: Deprecated program name' notice) and
    falling back to the classic name."""
    import shutil
    for n in names:
        found = shutil.which(n)
        if found:
            return found
    return names[-1]


def database_backup_bytes():
    """Produce the raw backup as (filename, bytes), independent of any HTTP
    response. Reused by the download view and the automated backup command.
    Returns (filename, data) on success or raises RuntimeError with a message."""
    import os
    import subprocess
    db = settings.DATABASES["default"]
    engine = db["ENGINE"]
    stamp = _dt.datetime.now().strftime("%Y%m%d-%H%M%S")

    if engine.endswith("sqlite3"):
        db_path = db["NAME"]
        if db_path is None or str(db_path).startswith("file:") \
                or not os.path.exists(db_path):
            raise RuntimeError("No on-disk database to back up.")
        with open(db_path, "rb") as fh:
            return f"treasury-backup-{stamp}.sqlite3", fh.read()

    defaults_file = None
    if engine.endswith("mysql"):
        defaults_file = _write_mysql_defaults_file(db)
        # --no-tablespaces avoids needing the PROCESS privilege (shared hosting
        # users rarely have it); we drop --routines for the same reason — a church
        # database has no stored procedures.
        cmd = [_mysql_tool("mariadb-dump", "mysqldump"),
               f"--defaults-extra-file={defaults_file}",
               "--single-transaction", "--no-tablespaces", db["NAME"]]
        env = dict(os.environ)
        ext = "sql"
    elif engine.endswith("postgresql"):
        cmd = ["pg_dump", "-h", db.get("HOST") or "localhost",
               "-p", str(db.get("PORT") or "5432"),
               "-U", db["USER"], db["NAME"]]
        env = dict(os.environ, PGPASSWORD=db.get("PASSWORD", ""))
        ext = "sql"
    else:
        raise RuntimeError("Backup isn't supported for this database engine.")

    try:
        proc = subprocess.run(cmd, capture_output=True, env=env, timeout=300)
    except (OSError, subprocess.SubprocessError) as e:
        raise RuntimeError(f"Could not run the database backup tool ({e}).")
    finally:
        if defaults_file:
            try:
                os.unlink(defaults_file)
            except OSError:
                pass
    if proc.returncode != 0:
        raise RuntimeError("The database backup tool reported an error: "
                           + (proc.stderr.decode("utf-8", "replace")[:300] or "unknown"))
    return f"treasury-backup-{stamp}.{ext}", proc.stdout


def database_backup_response():
    """Download a database backup, appropriate to the engine in use:

    * SQLite   → the raw .sqlite3 file (a complete, restorable snapshot).
    * MySQL    → a .sql dump via mysqldump.
    * Postgres → a .sql dump via pg_dump.

    For MySQL/Postgres the dump tool must be installed and on PATH (it is on a
    standard cPanel/WHM server). The Excel data export is always available as a
    fallback regardless of engine.
    """
    try:
        filename, data = database_backup_bytes()
    except RuntimeError as e:
        return HttpResponse(f"{e} The Excel data export is available as an "
                            f"alternative.", content_type="text/plain")
    if filename.endswith(".sqlite3"):
        import io
        return FileResponse(io.BytesIO(data), as_attachment=True, filename=filename)
    resp = HttpResponse(data, content_type="application/sql")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def full_excel_export_response(year=None):
    """A complete, meaningful financial workbook as at the download date.

    Sheets:
      • Summary       — period, opening cash position, grand totals
      • Fund Balances — every fund: opening (B/F), receipts, operating expenses,
                        remittances, transfers, closing balance + grand totals
      • Trust Funds   — per trust fund: collected, remitted, still to remit
      • Income by Channel — bank / cash / envelope with counts
      • Cash Book     — every receipt and payment in date order with running balance
      • Members, Transactions, Expenses, Departments, Reconciliations — raw tables
    Figures cover the current year to the download date unless `year` is given.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from giving.models import Transaction
    from cashbook.models import Expense
    from members.models import Member
    from departments.models import Department
    from reports.services import balances
    from core.models import SiteConfig

    today = _dt.date.today()
    y = year or today.year
    start = _dt.date(y, 1, 1)
    end = today if today.year == y else _dt.date(y, 12, 31)
    cfg = SiteConfig.get()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HEAD_FILL = PatternFill("solid", fgColor="1F5F4F")
    HEAD_FONT = Font(bold=True, color="FFFFFF")
    TITLE_FONT = Font(bold=True, size=14, color="1F5F4F")
    TOTAL_FONT = Font(bold=True)
    money_cols_cache = {}

    def _creators(model):
        """Audit-only map pk -> username of whoever first created the row, read
        from simple-history's create record. Deliberately NOT shown in the UI or
        any on-screen report — it exists only in this backup workbook so an
        auditor can see who entered each record. One query per table."""
        try:
            return {h["id"]: (h["history_user__username"] or "")
                    for h in model.history.filter(history_type="+")
                    .values("id", "history_user__username")}
        except Exception:
            return {}

    def _sheet(name, header, rows, title=None, money_cols=(), total_row=None):
        ws = wb.create_sheet(name[:31])
        r = 1
        if title:
            ws.cell(row=r, column=1, value=title).font = TITLE_FONT
            r += 1
            ws.cell(row=r, column=1,
                    value=f"{cfg.church_name} · as at {today:%d %b %Y}").font = \
                Font(italic=True, size=9, color="666666")
            r += 2
        hr = r
        for ci, h in enumerate(header, 1):
            c = ws.cell(row=hr, column=ci, value=h)
            c.fill = HEAD_FILL
            c.font = HEAD_FONT
        r += 1
        for row in rows:
            for ci, val in enumerate(row, 1):
                ws.cell(row=r, column=ci, value=val)
            r += 1
        if total_row:
            for ci, val in enumerate(total_row, 1):
                c = ws.cell(row=r, column=ci, value=val)
                c.font = TOTAL_FONT
            r += 1
        # number format on money columns
        for ci in money_cols:
            for rr in range(hr + 1, r):
                cell = ws.cell(row=rr, column=ci)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
        # widths + frozen header
        for ci, h in enumerate(header, 1):
            width = max(len(str(h)) + 2,
                        *(len(str(ws.cell(row=rr, column=ci).value or "")) + 2
                          for rr in range(hr + 1, min(r, hr + 60))), 10)
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(width, 48)
        ws.freeze_panes = ws.cell(row=hr + 1, column=1)
        return ws

    # ---- Fund Balances (the master accounting picture) --------------------
    rows = balances.department_summary(start, end)
    t = balances.totals(rows)
    fb = []
    for row in rows:
        d = row["department"]
        fb.append([
            d.name, "Trust" if row["is_trust"] else "Local",
            float(row["opening"]), float(row["receipts"]),
            float(row.get("expenses_operating", row["expenses"])),
            float(row.get("remittances", 0)),
            float(row.get("net_transfer", 0)), float(row["closing"]),
        ])
    fb.sort(key=lambda x: (x[1], x[0]))
    _sheet("Fund Balances",
           ["Fund", "Type", "Opening (B/F)", "Receipts", "Operating expenses",
            "Remittances", "Net transfers", "Closing balance"],
           fb,
           title=f"Fund balances — {start:%d %b %Y} to {end:%d %b %Y}",
           money_cols=(3, 4, 5, 6, 7, 8),
           total_row=["TOTAL", "", float(t["opening"]), float(t["receipts"]),
                      float(t["expenses_operating"]), float(t["remittances"]),
                      0.0, float(t["closing"])])

    # ---- Summary ----------------------------------------------------------
    opening_cash = (cfg.opening_bank_balance + cfg.opening_cash_on_hand
                    - cfg.opening_unremitted_trust)
    trust_rows = balances.trust_summary(start, end)
    to_remit = sum((tr["to_remit"] for tr in trust_rows), Decimal(0))
    trust_unreceipted = sum((tr["unreceipted"] for tr in trust_rows), Decimal(0))
    _sheet("Summary",
           ["Item", "Amount"],
           [["Reporting period", f"{start:%d %b %Y} – {end:%d %b %Y}"],
            ["Opening bank balance", float(cfg.opening_bank_balance)],
            ["Opening cash on hand", float(cfg.opening_cash_on_hand)],
            ["Opening unremitted trust", float(cfg.opening_unremitted_trust)],
            ["Net opening position", float(opening_cash)],
            ["Total receipts (period)", float(t["receipts"])],
            ["Total operating expenses (period)", float(t["expenses_operating"])],
            ["Total trust remittances (period)", float(t["remittances"])],
            ["Trust outstanding to remit (receipted)", float(to_remit)],
            ["Trust unreceipted (pending receipting)", float(trust_unreceipted)],
            ["Closing fund balances (sum)", float(t["closing"])]],
           title="Financial summary", money_cols=(2,))

    # ---- Trust Funds ------------------------------------------------------
    _sheet("Trust Funds",
           ["Trust fund", "Collected", "Remitted", "Outstanding to remit (receipted)",
            "Unreceipted (pending)", "Total liability"],
           [[tr["department"].name, float(tr["collected"]),
             float(tr["remitted"]), float(tr["to_remit"]),
             float(tr["unreceipted"]), float(tr["total_liability"])]
            for tr in trust_rows],
           title="Trust fund remittance schedule",
           money_cols=(2, 3, 4, 5, 6),
           total_row=["TOTAL",
                      float(sum((tr["collected"] for tr in trust_rows), Decimal(0))),
                      float(sum((tr["remitted"] for tr in trust_rows), Decimal(0))),
                      float(to_remit), float(trust_unreceipted),
                      float(sum((tr["total_liability"] for tr in trust_rows), Decimal(0)))])

    # ---- Income by Channel ------------------------------------------------
    chan = balances.income_by_channel(start, end)
    _sheet("Income by Channel",
           ["Channel", "Amount", "Count"],
           [[c["channel"], float(c["total"] or 0), c["count"]] for c in chan],
           title="Income by channel", money_cols=(2,))

    # ---- Cash Book (receipts & payments with running balance) -------------
    entries = []
    for tx in (Transaction.objects.confirmed_credits()
               .filter(date__gte=start, date__lte=end, excluded_from_income=False)
               .select_related("department", "member").order_by("date", "id")):
        entries.append((tx.date, tx.member.name if tx.member_id else (tx.payer_name or "Receipt"),
                        tx.department.name if tx.department_id else "",
                        float(tx.amount), 0.0))
    for x in (Expense.objects.filter(date__gte=start, date__lte=end,
              status__in=[Expense.Status.APPROVED, Expense.Status.PAID])
              .select_related("department").order_by("date", "id")):
        entries.append((x.date, x.description,
                        x.department.name if x.department_id else "",
                        0.0, float(x.amount)))
    entries.sort(key=lambda e: e[0])
    cashbook = []
    running = float(opening_cash)
    cashbook.append(["", "Opening position", "", 0.0, 0.0, running])
    for d, desc, fund, cr, dr in entries:
        running += cr - dr
        cashbook.append([d.isoformat(), desc, fund, cr, dr, running])
    _sheet("Cash Book",
           ["Date", "Description", "Fund", "Receipt", "Payment", "Balance"],
           cashbook, title="Cash book", money_cols=(4, 5, 6))

    # ---- Raw tables -------------------------------------------------------
    _dcre = _creators(Department)
    _sheet("Departments",
           ["ID", "Name", "Type", "Category", "Trust", "Opening balance", "Created by"],
           [[d.id, d.name, d.fund_type, d.category, "Y" if d.is_trust else "",
             float(d.opening_balance or 0), _dcre.get(d.id, "")]
            for d in Department.objects.all().order_by("name")],
           money_cols=(6,))

    _mcre = _creators(Member)
    _sheet("Members",
           ["ID", "Name", "Primary phone", "All phones", "Group", "Source", "Active",
            "Created by"],
           [[m.id, m.name, m.receipt_phone or "",
             ", ".join(m.phones.values_list("number", flat=True)) or (m.phone or ""),
             m.group or "", m.get_source_display(), "Y" if m.active else "",
             _mcre.get(m.id, "")]
            for m in Member.objects.all().order_by("name")])

    _tcre = _creators(Transaction)
    _sheet("Transactions",
           ["ID", "Date", "Channel", "Direction", "Amount", "Fund", "Dev group",
            "Member / payer", "Reference", "M-Pesa ref", "Status", "Confirmed",
            "Excluded from income", "Created by"],
           [[t2.id, t2.date.isoformat(), t2.channel, t2.direction, float(t2.amount),
             t2.department.name if t2.department_id else "",
             t2.dev_group.number if t2.dev_group_id else "",
             t2.member.name if t2.member_id else (t2.payer_name or ""),
             t2.reference or "", t2.mpesa_ref or "", t2.allocation_status,
             "Y" if t2.confirmed else "", "Y" if t2.excluded_from_income else "",
             _tcre.get(t2.id, "")]
            for t2 in Transaction.objects.select_related(
                "department", "dev_group", "member").order_by("date")],
           money_cols=(5,))

    _xcre = _creators(Expense)
    _sheet("Expenses",
           ["ID", "Date", "Fund", "Description", "Amount", "Category", "Method",
            "Status", "Claimant", "Voucher", "Recorded by", "Created by"],
           [[x.id, x.date.isoformat(), x.department.name if x.department_id else "",
             x.description, float(x.amount), x.get_category_display(),
             x.method, x.get_status_display(), x.claimant or "",
             x.voucher_no or "", x.recorded_by.username if x.recorded_by_id else "",
             _xcre.get(x.id, "")]
            for x in Expense.objects.select_related(
                "department", "recorded_by").order_by("date")],
           money_cols=(5,))

    try:
        from statements.models import BankReconciliation
        _rcre = _creators(BankReconciliation)
        _sheet("Reconciliations",
               ["ID", "Statement date", "Bank balance", "Book balance", "Difference",
                "Reconciled", "Created by"],
               [[r.id, r.statement_date.isoformat(), float(r.bank_balance),
                 float(r.book_balance or 0), float(r.difference or 0),
                 "Y" if r.is_reconciled else "", _rcre.get(r.id, "")]
                for r in BankReconciliation.objects.order_by("statement_date")],
               money_cols=(3, 4, 5))
    except Exception:
        pass

    # ---- Payment instruments (cheques / EFT / RTGS / M-Pesa) --------------
    try:
        from cashbook.models import PaymentInstrument
        _sheet("Payments",
               ["ID", "Method", "Number / ref", "Payee", "Amount", "Bank account",
                "Date issued", "Date cleared", "Status", "Settles", "Source ID"],
               [[p.id, p.get_method_display(), p.instrument_number or "",
                 p.payee or "", float(p.amount),
                 str(p.bank_account) if p.bank_account_id else "",
                 p.date_issued.isoformat() if p.date_issued else "",
                 p.date_cleared.isoformat() if p.date_cleared else "",
                 p.get_status_display(), p.get_source_kind_display(),
                 p.expense_id or p.remittance_batch_id or p.refund_id
                 or p.transfer_id or ""]
                for p in PaymentInstrument.objects.select_related(
                    "bank_account").order_by("date_issued", "id")],
               title="Payment register", money_cols=(5,))
    except Exception:
        pass

    # ---- Staff advances ---------------------------------------------------
    try:
        from cashbook.models import StaffAdvance
        _sheet("Staff Advances",
               ["ID", "Staff", "Fund", "Amount", "Date issued", "Method",
                "From petty cash", "Purpose", "Status", "Reference", "Issued by"],
               [[a.id, a.staff_name,
                 a.department.name if a.department_id else "",
                 float(a.amount), a.date_issued.isoformat(), a.method,
                 "Y" if a.from_petty_cash else "", a.purpose or "",
                 a.get_status_display(), a.reference or "",
                 a.issued_by.username if a.issued_by_id else ""]
                for a in StaffAdvance.objects.select_related(
                    "department", "issued_by").order_by("date_issued", "id")],
               title="Staff advances", money_cols=(4,))
    except Exception:
        pass

    # ---- Remittance batches ----------------------------------------------
    try:
        from cashbook.models import RemittanceBatch
        _sheet("Remittances",
               ["Batch", "Date", "Period start", "Period end", "Amount",
                "Status", "Settlement", "Created by"],
               [[b.batch_number, b.date.isoformat() if b.date else "",
                 b.period_start.isoformat() if b.period_start else "",
                 b.period_end.isoformat() if b.period_end else "",
                 float(b.total_amount), b.get_status_display(),
                 b.settlement_label or "",
                 b.created_by.username if b.created_by_id else ""]
                for b in RemittanceBatch.objects.select_related(
                    "created_by", "payment").order_by("-created_at")],
               title="Conference remittance batches", money_cols=(5,))
    except Exception:
        pass

    # ---- Fund transfers ---------------------------------------------------
    try:
        from cashbook.models import FundTransfer
        _sheet("Fund Transfers",
               ["ID", "Date", "From fund", "To fund", "Amount", "Reason",
                "Reversed", "Recorded by"],
               [[t3.id, t3.date.isoformat(),
                 t3.source.name if t3.source_id else "",
                 t3.destination.name if t3.destination_id else "",
                 float(t3.amount), t3.reason or "",
                 "Y" if t3.is_reversed else "",
                 t3.recorded_by.username if t3.recorded_by_id else ""]
                for t3 in FundTransfer.objects.select_related(
                    "source", "destination", "recorded_by").order_by("date")],
               title="Inter-fund transfers", money_cols=(5,))
    except Exception:
        pass

    # ---- Payables & accruals ---------------------------------------------
    try:
        from cashbook.models import Payable, Accrual
        _sheet("Payables",
               ["ID", "Vendor", "Fund", "Description", "Amount", "Due date",
                "Settled"],
               [[p.id, getattr(p, "vendor", "") or "",
                 p.department.name if p.department_id else "", p.description or "",
                 float(p.amount), p.due_date.isoformat() if p.due_date else "",
                 "Y" if p.settled else ""]
                for p in Payable.objects.select_related("department").order_by("due_date")],
               title="Payables (money owed)", money_cols=(5,))
        _sheet("Accruals",
               ["ID", "Fund", "Description", "Amount", "Period", "Settled"],
               [[a.id, a.department.name if a.department_id else "",
                 a.description or "", float(a.amount),
                 a.period.isoformat() if getattr(a, "period", None) else "",
                 "Y" if a.settled else ""]
                for a in Accrual.objects.select_related("department").order_by("id")],
               title="Accruals", money_cols=(4,))
    except Exception:
        pass

    # ---- Pledges ----------------------------------------------------------
    try:
        from pledges.models import Pledge
        _sheet("Pledges",
               ["ID", "Member / pledger", "Campaign", "Amount pledged",
                "Fulfilled", "Balance", "Status"],
               [[pl.id, pl.member.name if pl.member_id else (pl.pledger_name or ""),
                 pl.campaign.name if getattr(pl, "campaign_id", None) else "",
                 float(pl.amount), float(getattr(pl, "fulfilled_amount", 0) or 0),
                 float(pl.amount - (getattr(pl, "fulfilled_amount", 0) or 0)),
                 getattr(pl, "status", "")]
                for pl in Pledge.objects.select_related("member").order_by("id")],
               title="Pledges", money_cols=(4, 5, 6))
    except Exception:
        pass

    # ---- Fixed assets -----------------------------------------------------
    try:
        from assets.models import FixedAsset
        _sheet("Fixed Assets",
               ["ID", "Name", "Category", "Acquired", "Cost",
                "Net book value", "Disposed"],
               [[a.id, a.name, getattr(a, "category", "") or "",
                 a.acquired_date.isoformat() if getattr(a, "acquired_date", None) else "",
                 float(getattr(a, "cost", 0) or 0),
                 float(a.net_book_value(today)), "Y" if a.disposed else ""]
                for a in FixedAsset.objects.order_by("name")],
               title="Fixed assets register", money_cols=(5, 6))
    except Exception:
        pass

    # ---- Petty cash top-ups ----------------------------------------------
    try:
        from cashbook.models import PettyCashTopUp
        _sheet("Petty Cash Top-ups",
               ["ID", "Date", "Amount", "Note", "Recorded by"],
               [[t4.id, t4.date.isoformat(), float(t4.amount),
                 getattr(t4, "note", "") or "",
                 t4.recorded_by.username if t4.recorded_by_id else ""]
                for t4 in PettyCashTopUp.objects.select_related(
                    "recorded_by").order_by("date")],
               title="Petty cash top-ups", money_cols=(3,))
    except Exception:
        pass

    # ---- Petty cash deposited to bank (float → bank) ---------------------
    try:
        from cashbook.models import PettyCashBankDeposit
        _sheet("Petty Cash Bank Deposits",
               ["ID", "Date", "Amount", "Note", "Bank txn", "Recorded by"],
               [[d.id, d.date.isoformat(), float(d.amount),
                 getattr(d, "note", "") or "",
                 d.bank_transaction_id or "",
                 d.recorded_by.username if d.recorded_by_id else ""]
                for d in PettyCashBankDeposit.objects.select_related(
                    "recorded_by").order_by("date")],
               title="Petty cash deposited to bank", money_cols=(3,))
    except Exception:
        pass

    # ---- Benevolent scheme -------------------------------------------------
    # The benevolent module holds the church's welfare obligations — who is
    # covered, who their household is, what has been paid in and what has been
    # paid out. A data export that omits it is not a copy of the church's
    # records, and it is the part least reconstructible from the bank statement:
    # a levy is just an amount until you know which case it settled.
    try:
        from benevolent.models import (BenevolentCase, BenevolentContribution,
                                       BenevolentScheme, SchemeDependant,
                                       SchemeMembership)
        _sheet("Benevolent Schemes",
               ["ID", "Code", "Name", "Status", "Fund"],
               [[sc.id, sc.code, sc.name, sc.get_status_display(),
                 sc.fund.name if sc.fund_id else ""]
                for sc in BenevolentScheme.objects.select_related("fund")
                .order_by("code")],
               title="Benevolent schemes")

        _sheet("Benevolent Members",
               ["ID", "Number", "Scheme", "Member", "Phone", "Joined", "Status",
                "Standing", "Type", "Household", "Registration fee paid",
                "Died on", "Left on"],
               [[m.id, m.number, m.scheme.code, m.member.name,
                 m.member.phone or "", m.joined_on.isoformat() if m.joined_on else "",
                 m.get_status_display(), m.get_standing_display(),
                 m.get_registration_type_display(), m.household_name or "",
                 "Yes" if m.registration_fee_paid else "No",
                 m.died_on.isoformat() if m.died_on else "",
                 m.left_on.isoformat() if m.left_on else ""]
                for m in SchemeMembership.objects
                .select_related("scheme", "member")
                .order_by("scheme__code", "member__name")],
               title="Benevolent memberships")

        # Living dependants and departed ones both, with the date: who was
        # covered when is the question a past claim is judged on.
        _sheet("Benevolent Dependants",
               ["Membership", "Member", "Dependant", "Relationship", "Phone",
                "Date of birth", "Registered", "Active", "Died on"],
               [[d.membership.number, d.membership.member.name, d.display_name,
                 d.get_relationship_display(), d.phone or "",
                 d.date_of_birth.isoformat() if d.date_of_birth else "",
                 d.registered_on.isoformat() if d.registered_on else "",
                 "Yes" if d.active else "No",
                 d.died_on.isoformat() if d.died_on else ""]
                for d in SchemeDependant.objects
                .select_related("membership__member", "member")
                .order_by("membership__number", "relationship")],
               title="Benevolent dependants")

        _sheet("Benevolent Cases",
               ["ID", "Number", "Scheme", "Member", "Event", "Event date",
                "Status", "Beneficiary"],
               [[c5.id, c5.number, c5.scheme.code,
                 c5.membership.member.name if c5.membership_id else "",
                 c5.event_type.name if c5.event_type_id else "",
                 c5.event_date.isoformat() if c5.event_date else "",
                 c5.get_status_display(), c5.beneficiary_display]
                for c5 in BenevolentCase.objects
                .select_related("scheme", "membership__member", "event_type")
                .order_by("-event_date")],
               title="Benevolent cases")

        _sheet("Benevolent Contributions",
               ["Date", "Scheme", "Member", "Kind", "Period", "Case", "Amount",
                "Automatic", "Reversed", "Note"],
               [[cb.date.isoformat() if cb.date else "", cb.scheme.code,
                 cb.membership.member.name if cb.membership_id else cb.payer_name,
                 cb.get_kind_display(), cb.period_label or "",
                 cb.case.number if cb.case_id else "", float(cb.amount or 0),
                 "Yes" if cb.allocated_automatically else "No",
                 cb.reversed_at.date().isoformat() if cb.reversed_at else "",
                 cb.note or ""]
                for cb in BenevolentContribution.objects
                .select_related("scheme", "membership__member", "case",
                                "transaction")
                .order_by("-id")[:20000]],
               title="Benevolent contributions", money_cols=(7,))
    except Exception:
        pass

    # ---- Envelopes, one sheet per month, as the counting schedule ----------
    # Laid out the way the envelope sheet is worked on and downloaded: a row per
    # contributor, a column per fund, a total down each. A flat list of lines
    # would hold the same numbers but could not be checked against the paper it
    # came from, which is the only reason anybody opens this.
    try:
        from envelopes.models import Envelope
        from collections import OrderedDict
        envs = list(Envelope.objects
                    .select_related("member")
                    .prefetch_related("lines__department")
                    .order_by("date", "receipt_no"))
        months = OrderedDict()
        for e in envs:
            months.setdefault(e.date.strftime("%Y-%m"), []).append(e)
        for month, rows_for_month in months.items():
            funds = []
            for e in rows_for_month:
                for ln in e.lines.all():
                    label = ln.department.name if ln.department_id else "Unallocated"
                    if label not in funds:
                        funds.append(label)
                        
            funds.sort()
            header = ["No", "Date", "Contributor Name", "Phone", "Receipt No",
                      "Channel"] + funds + ["Total"]
            body, totals = [], {f: 0.0 for f in funds}
            for n, e in enumerate(rows_for_month, start=1):
                by_fund = {f: 0.0 for f in funds}
                for ln in e.lines.all():
                    label = ln.department.name if ln.department_id else "Unallocated"
                    by_fund[label] = by_fund.get(label, 0.0) + float(ln.amount or 0)
                for f in funds:
                    totals[f] += by_fund[f]
                body.append([n, e.date.isoformat(),
                             (e.member.name if e.member_id else e.contributor_name) or "",
                             (e.member.phone if e.member_id else "") or "",
                             e.receipt_no or "", e.get_channel_display()]
                            + [by_fund[f] for f in funds]
                            + [float(e.total or 0)])
            grand = sum(totals.values())
            _sheet(f"Envelopes {month}", header, body,
                   title=f"Envelope schedule · {month}",
                   money_cols=tuple(range(7, 7 + len(funds) + 1)),
                   total_row=["", "", "Total", "", "", ""]
                             + [totals[f] for f in funds] + [grand])
    except Exception:
        pass

    stamp = today.strftime("%Y%m%d")
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="treasury-data-{stamp}.xlsx"'
    wb.save(resp)
    return resp


SQLITE_MAGIC = b"SQLite format 3\x00"


def _looks_like_sqlite(path):
    """Is this actually a SQLite database? Returns (ok, why_not).

    Restore copies the uploaded file straight over the live database. Without
    this, uploading the wrong file — a spreadsheet, a PDF, a SQL dump meant for
    a different engine — would destroy the church's data and only announce
    itself the next time somebody opened a page. The safety copy taken a moment
    later is no help if nobody realises anything happened.
    """
    import os
    try:
        with open(path, "rb") as fh:
            head = fh.read(16)
    except OSError:
        return False, "The uploaded file could not be read."
    if head != SQLITE_MAGIC:
        return False, (
            "That file is not a SQLite database, so it has not been restored "
            "and your data is untouched. A backup of this system downloads as "
            "a .sqlite3 file — check you are uploading that rather than a "
            "spreadsheet export or a dump from a different kind of database.")
    if os.path.getsize(path) < 4096:
        return False, ("That file is a SQLite database but far too small to be "
                       "this system's — nothing has been restored.")
    return True, ""


def _looks_like_sql_dump(path):
    """Is this a text SQL dump rather than a binary file? Returns (ok, why_not)."""
    try:
        with open(path, "rb") as fh:
            head = fh.read(4096)
    except OSError:
        return False, "The uploaded file could not be read."
    if head.startswith(SQLITE_MAGIC):
        return False, (
            "That is a SQLite backup, but this system runs on a different "
            "database, so it cannot be loaded directly. Nothing has been "
            "changed. Restore a dump taken from this system instead.")
    if b"\x00" in head:
        return False, ("That file is not a readable SQL dump, so nothing has "
                       "been restored and your data is untouched.")
    return True, ""


def database_restore(uploaded_file):
    """Restore the database from an uploaded backup. Returns (ok, message).

    SAFETY: takes a fresh backup of the CURRENT database first (so a bad restore
    is itself reversible), then loads the uploaded file. SQLite restores by
    replacing the file; MySQL/Postgres restore by piping the SQL dump through the
    client. This is a destructive, treasurer-only operation guarded by an
    explicit confirmation in the view.
    """
    import os
    import subprocess
    import tempfile
    import datetime as _dt
    from django.conf import settings
    from django.db import connection

    db = settings.DATABASES["default"]
    engine = db["ENGINE"]
    stamp = _dt.datetime.now().strftime("%Y%m%d-%H%M%S")

    # read upload to a temp file
    suffix = ".sqlite3" if engine.endswith("sqlite3") else ".sql"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        for chunk in uploaded_file.chunks():
            tmp.write(chunk)
        tmp.close()

        if engine.endswith("sqlite3"):
            # `NAME` is whatever settings.py put there. Django's own template
            # writes `BASE_DIR / "db.sqlite3"`, which is a PosixPath — and this
            # asked `isinstance(path, (str, bytes))`, which a Path fails. So a
            # perfectly ordinary installation was told there was no database to
            # restore into while the file sat right there. Coerced instead of
            # type-checked, which is what the backup side has always done.
            # What was uploaded is checked before where it would go. Somebody
            # who has just picked the wrong file needs to be told that, and it
            # is true regardless of how this installation stores its data.
            ok, why = _looks_like_sqlite(tmp.name)
            if not ok:
                return False, why
            path = os.fspath(db["NAME"]) if db["NAME"] is not None else ""
            if not path or path.startswith("file:") or not os.path.exists(path):
                return False, (
                    "This installation has no on-disk SQLite database to restore "
                    "into — it is running in memory or from a URI. Restore needs "
                    "a database file.")
            import shutil
            safety = f"{path}.pre-restore-{stamp}"
            shutil.copy2(path, safety)
            connection.close()
            shutil.copy2(tmp.name, path)
            return True, (f"Database restored from backup. The previous database "
                          f"was saved as {os.path.basename(safety)}.")

        if engine.endswith("mysql"):
            ok, why = _looks_like_sql_dump(tmp.name)
            if not ok:
                return False, why
            # safety dump of current state
            safety = f"/tmp/treasury-pre-restore-{stamp}.sql"
            defaults_file = _write_mysql_defaults_file(db)
            env = dict(os.environ)
            try:
                with open(safety, "wb") as out:
                    subprocess.run([_mysql_tool("mariadb-dump", "mysqldump"),
                                    f"--defaults-extra-file={defaults_file}",
                                    "--single-transaction", "--no-tablespaces",
                                    db["NAME"]],
                                   stdout=out, env=env, timeout=300, check=True)
                # load the uploaded dump
                with open(tmp.name, "rb") as src:
                    proc = subprocess.run([_mysql_tool("mariadb", "mysql"),
                                           f"--defaults-extra-file={defaults_file}",
                                           db["NAME"]],
                                          stdin=src, env=env, timeout=600,
                                          capture_output=True)
            finally:
                try:
                    os.unlink(defaults_file)
                except OSError:
                    pass
            if proc.returncode != 0:
                return False, ("Restore failed while loading the backup. Your data "
                               "was not changed beyond the safety dump at "
                               f"{safety}. Detail: {proc.stderr.decode()[:300]}")
            return True, (f"Database restored from backup. A safety copy of the "
                          f"previous data was saved to {safety} on the server.")

        if engine.endswith("postgresql"):
            ok, why = _looks_like_sql_dump(tmp.name)
            if not ok:
                return False, why
            safety = f"/tmp/treasury-pre-restore-{stamp}.sql"
            env = dict(os.environ, PGPASSWORD=db.get("PASSWORD", ""))
            host = db.get("HOST") or "localhost"
            port = str(db.get("PORT") or "5432")
            with open(safety, "wb") as out:
                subprocess.run(["pg_dump", "-h", host, "-p", port, "-U", db["USER"],
                                db["NAME"]], stdout=out, env=env, timeout=300, check=True)
            with open(tmp.name, "rb") as src:
                proc = subprocess.run(["psql", "-h", host, "-p", port, "-U", db["USER"],
                                       db["NAME"]], stdin=src, env=env, timeout=600,
                                      capture_output=True)
            if proc.returncode != 0:
                return False, ("Restore failed while loading the backup. A safety "
                               f"copy was saved to {safety}.")
            return True, (f"Database restored. Safety copy saved to {safety}.")

        return False, "Restore isn't supported for this database engine."
    except subprocess.TimeoutExpired:
        return False, "Restore timed out."
    except Exception as e:  # noqa: BLE001
        return False, f"Restore failed: {e}"
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass


def upload_offsite(filename, data, cfg=None):
    """Upload a backup file to off-site storage over HTTPS (dependency-free).

    Does an authenticated HTTP PUT to the configured destination URL with the
    file name appended — compatible with WebDAV (Nextcloud/ownCloud) and any
    object store / endpoint that accepts an authenticated PUT. Returns
    (ok, detail); never raises into the caller.
    """
    import base64
    import urllib.request
    import urllib.error
    from core.models import SiteConfig
    cfg = cfg or SiteConfig.get()
    if not getattr(cfg, "offsite_backup_enabled", False):
        return False, "Off-site backup is not enabled."
    base = (cfg.offsite_backup_url or "").strip()
    if not base:
        return False, "No off-site backup URL configured."
    url = base if base.endswith("/") else base + "/"
    url = url + filename
    try:
        req = urllib.request.Request(url, data=data, method="PUT")
        req.add_header("Content-Type", "application/octet-stream")
        user = (cfg.offsite_backup_user or "").strip()
        pwd = cfg.offsite_backup_password or ""
        if user:
            token = base64.b64encode(f"{user}:{pwd}".encode()).decode("ascii")
            req.add_header("Authorization", f"Basic {token}")
        with urllib.request.urlopen(req, timeout=30) as r:
            code = r.getcode()
        if 200 <= code < 300:
            return True, f"Uploaded {filename} to off-site storage."
        return False, f"Upload returned HTTP {code}."
    except urllib.error.HTTPError as e:
        return False, f"HTTP {e.code}: {e.reason}"
    except Exception as e:  # noqa: BLE001
        return False, f"{type(e).__name__}: {e}"
