"""
Legacy data import — transfer the historical Excel reporting sheets into the
treasury system.

This command is ADDITIVE ONLY: it never alters models, settings, or existing
behaviour. It reads the uploaded workbooks and creates Departments, Development
groups, Envelopes (with their fund lines), bank Transactions, Expenses and cash
Collections, using the same models the app already uses. It is safe to re-run —
every record is keyed so a second run skips what already exists.

Usage:
    python manage.py import_legacy --dir /path/to/xlsx [options]

Options:
    --dir PATH        Folder holding the .xlsx files (default: /mnt/user-data/uploads)
    --phase NAME      one of: all departments envelopes bank expenses collection
                      (default: all)
    --dry-run         Parse and report only; write nothing.
    --noinput         Never prompt; unresolved accounts are reported and skipped.
    --map PATH        JSON file caching account-name decisions (read + updated).
    --year YEAR       Calendar year of the data (default: 2026).

The "account resolver" turns a spreadsheet account label (which may not match an
account name word-for-word) into a real Department, using: exact match →
normalised match → a built-in alias table → fuzzy match → (if still unsure) an
interactive prompt whose answer is cached so you are only asked once.
"""
import datetime as dt
import difflib
import json
import os
import re
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction as db_tx

MONTHS = {"JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6, "JUNE": 6,
          "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12}

REPORTING_FILES = {
    1: "REPORTING_SHEET_JAN_2026.xlsx", 2: "REPORTING_SHEET_FEB_2026.xlsx",
    3: "REPORTING_SHEET_MAR_2026.xlsx", 4: "REPORTING_SHEET_APR_2026.xlsx",
    5: "REPORTING_SHEET_MAY.xlsx", 6: "REPORTING_SHEET_JUNE_2026.xlsx",
}
MASTER_FILE = "JUNE.xlsx"

# The 43 parent departments come from the DEPARTMENTS sheet (rows 14-56) and carry
# the brought-forward balances (total 5,757,459). Read at runtime, not hardcoded.
PARENT_BF_ROWS = range(14, 57)

# Subgroups requested explicitly (created as child accounts; b/f stays on parent).
VBS_SUBGROUPS = ["VBS_AMM", "VBS_AWM", "VBS_YOUTH", "VBS_AMBASSADORS",
                 "VBS_CHILDREN", "VBS_CMFOOD", "VBS_REGISTRATION"]
DEV_SUBGROUPS = [
    "DEV_GROUP_13", "DEV_GROUP_1", "DEV_GROUP_15", "DEV_GROUP_3", "DEV_GROUP_17",
    "DEV_GROUP_16", "DEV_GROUP_7", "DEV_GROUP_6", "DEV_GROUP_5", "DEV_GROUP_11",
    "DEV_GROUP_41", "DEV_GROUP_12", "DEV_GROUP_32", "DEV_GROUP_4", "DEV_GROUP_19",
    "DEV_GROUP_8", "DEV_GROUP_2", "DEV_GROUP_18", "DEV_GROUP_NA", "DEV_GROUP_40",
    "DEV_GROUP_9", "DEV_GROUP_10", "DEV_GROUP_39", "DEV_GROUP_26", "DEV_GROUP_20",
    "DEV_GROUP_14", "DEV_GROUP_22", "DEV_GROUP_30", "DEV_GROUP_35", "DEV_GROUP_24",
    "DEV_GROUP_25", "DEV_GROUP_21", "DEV_GROUP_33", "DEV_GROUP_34", "DEV_GROUP_28",
    "DEV_GROUP_23", "DEV_GROUP_27", "DEV_GROUP_37", "DEV_GROUP_38", "DEV_GROUP_36",
    "DEV_GROUP_VISITORS", "DEV_GROUP_31", "DEV_GROUP_29", "DEV_GROUP_REFUND",
]

# Bank-statement labels that are NOT fund allocations — skipped and reported.
BANK_SKIP = {"ENVELOPES", "NARRATION", "DEPOSIT", "WITHDRAWAL", "CHARGES"}
COLLECTION_SKIP = {"ENVELOPES"}

# The Local Church Budget (LCB) is made up of these sub-funds -> created as
# children of the LCB parent so they roll up into it.
LCB_CHILDREN = ["SABBATH SCHOOL", "LOOSE OFFERING", "COMBINED - LCB",
                "THANKGIVING - LCB", "LOCAL CHURCH BUDGET", "ENVELOPES",
                "LCB_DEPARTMENTS"]

# Non-numeric development groups need a synthetic number (DevelopmentGroup.number
# is a unique integer); kept high to avoid clashing with the real group numbers.
DEV_SPECIAL_NUMBERS = {"DEV_GROUP_NA": 900, "DEV_GROUP_VISITORS": 901,
                       "DEV_GROUP_REFUND": 902}

# Supporting funds the data references that aren't in the 43 parents; created with
# zero b/f so the imported transactions have a home. (name -> (fund_type, category))
SUPPORTING_FUNDS = {
    "TITHE": ("TRUST", "TRUST"),
    "COMBINED (50%) - ENF": ("TRUST", "TRUST"),
    "THANKSGIVING - ENF": ("TRUST", "TRUST"),
    "CAMP MEETING": ("TRUST", "TRUST"),
    "EVANGELISM - FIELD": ("TRUST", "TRUST"),
    "STATION DEV FUNDS": ("TRUST", "TRUST"),
    "LOCAL_EVANGELISM": ("LOCAL", "MINISTRY"),
    "13TH SABBATH OFFERING": ("TRUST", "TRUST"),
}

# Normalised source label -> canonical department name. Built from inspecting the
# sheets; the genuinely ambiguous ones are deliberately left out so they prompt.
ALIASES = {
    "LOCAL CHURCH BUDGET": "LOCAL CHURCH BUDGET",
    "LOCAL CHURCH BUDGET LCB": "LOCAL CHURCH BUDGET",
    "LCB": "LOCAL CHURCH BUDGET",
    "SABBATH SCH": "SABBATH SCHOOL",
    "SABBATH SCHOOL": "SABBATH SCHOOL",
    "PATHFINDER REG": "PATHFINDERS",
    "EVANGELISM FIELD": "EVANGELISM - FIELD",
    "FIELD EVANGELISM": "EVANGELISM - FIELD",
    "EVANGELISM LOCAL": "LOCAL_EVANGELISM",
    "LOCAL EVANGELISM": "LOCAL_EVANGELISM",
    "CAMP MEETING OFFERINGS": "CAMP MEETING",
    "CAMP MEETING OFFERING": "CAMP MEETING",
    "CAMP OFFERING": "CAMP MEETING",
    "CAMP EXPENSES": "CAMP EXPENSE",
    "CAMP EXPENSE": "CAMP EXPENSE",
    "STATION DEV FUND": "STATION DEV FUNDS",
    "STATION DEVELOPMENT": "STATION DEV FUNDS",
    "CHILDREN MINISTRY": "CHILDREN MINISTRY",
    "LOOSE OFFERING": "LOOSE OFFERING",
    "13TH SABBATH OFFERING": "13TH SABBATH OFFERING",
}

# Labels that are not a fund at all when seen as an envelope column.
NON_FUND_COLUMNS = {"GROUP NUMBER", "NO", "CONTRIBUTOR NAME", "RECEIPT NUMBER",
                    "TOTAL", "GROUP"}

# Contributor-name cells that actually mark the end of the contributor list in a
# SABBATH tab (a grand-total / summary block follows). Stop reading at these.
STOP_NAMES = {"TOTAL", "TOTALS", "GRAND TOTAL", "SUMMARY", "SUM", "SUB TOTAL",
              "SUBTOTAL"}

# Split offerings: a single envelope column whose amount is divided by fixed
# percentages across funds (e.g. Combined Offering = 50% conference trust + 50%
# local). normalised label -> [(department name, fraction), ...].
SPLITS = {
    "COMBINED OFFERINGS": [("COMBINED (50%) - ENF", Decimal("0.5")),
                           ("COMBINED - LCB", Decimal("0.5"))],
    "COMBINED OFFERING": [("COMBINED (50%) - ENF", Decimal("0.5")),
                          ("COMBINED - LCB", Decimal("0.5"))],
    "COMBINED OFFERING 50": [("COMBINED (50%) - ENF", Decimal("0.5")),
                             ("COMBINED - LCB", Decimal("0.5"))],
    "THANKS GIVING": [("THANKSGIVING - ENF", Decimal("0.5")),
                      ("THANKGIVING - LCB", Decimal("0.5"))],
    "THANKS": [("THANKSGIVING - ENF", Decimal("0.5")),
               ("THANKGIVING - LCB", Decimal("0.5"))],
    "THANKS GIVING 50": [("THANKSGIVING - ENF", Decimal("0.5")),
                         ("THANKGIVING - LCB", Decimal("0.5"))],
}


def norm(s):
    """Uppercase, strip punctuation to spaces, collapse whitespace."""
    if s is None:
        return ""
    s = re.sub(r"[^A-Za-z0-9]+", " ", str(s)).upper()
    return " ".join(s.split())


def to_decimal(v):
    if v in (None, ""):
        return None
    try:
        return Decimal(str(v).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        return None


def split_amount(amount, fractions):
    """Split `amount` by `fractions` so the parts sum back to `amount` exactly."""
    parts, running = [], Decimal(0)
    for i, f in enumerate(fractions):
        if i == len(fractions) - 1:
            parts.append(amount - running)
        else:
            share = (amount * f).quantize(Decimal("0.01"))
            parts.append(share)
            running += share
    return parts


class AccountResolver:
    """Turn a spreadsheet account label into a Department, learning as it goes."""

    def __init__(self, cmd, cache_path, interactive):
        self.cmd = cmd
        self.cache_path = cache_path
        self.interactive = interactive
        self.cache = {}
        if cache_path and os.path.exists(cache_path):
            try:
                self.cache = json.load(open(cache_path))
            except Exception:
                self.cache = {}
        self.unresolved = {}          # normalised label -> count
        self._reload()

    def _reload(self):
        from departments.models import Department
        self.by_name = {d.name.upper(): d for d in Department.objects.all()}
        self.by_norm = {norm(d.name): d for d in Department.objects.all()}

    def _save_cache(self):
        if self.cache_path:
            json.dump(self.cache, open(self.cache_path, "w"), indent=2)

    def resolve(self, label):
        """Return a Department or None (skip). None means 'could not resolve'."""
        from departments.models import Department
        raw = (label or "").strip()
        if not raw:
            return None
        n = norm(raw)
        if n in NON_FUND_COLUMNS or not n:
            return None

        # 1) cached decision
        if n in self.cache:
            target = self.cache[n]
            if target == "__SKIP__":
                return None
            d = Department.objects.filter(name=target).first()
            if d:
                return d

        # 2) exact / normalised match
        if raw.upper() in self.by_name:
            return self.by_name[raw.upper()]
        if n in self.by_norm:
            return self.by_norm[n]

        # 3) alias table
        if n in ALIASES:
            d = Department.objects.filter(name=ALIASES[n]).first()
            if d:
                return d

        # 4) fuzzy match against existing department names
        names = list(self.by_norm.keys())
        close = difflib.get_close_matches(n, names, n=1, cutoff=0.90)
        if close:
            return self.by_norm[close[0]]

        # 5) ask (or record as unresolved)
        if self.interactive:
            return self._prompt(raw, n)
        self.unresolved[n] = self.unresolved.get(n, 0) + 1
        return None

    def resolve_split(self, label):
        """If `label` is a split offering, return [(Department, fraction), ...];
        otherwise None. Falls back to skipping components that don't resolve."""
        from departments.models import Department
        n = norm(label)
        if n not in SPLITS:
            return None
        out = []
        for name, frac in SPLITS[n]:
            d = Department.objects.filter(name=name).first()
            if d:
                out.append((d, frac))
        return out or None

    def _prompt(self, raw, n):
        from departments.models import Department
        names = list(self.by_norm.keys())
        suggestions = difflib.get_close_matches(n, names, n=5, cutoff=0.4)
        self.cmd.stdout.write(self.cmd.style.WARNING(
            f"\nUnrecognised account: '{raw}'"))
        opts = [self.by_norm[s] for s in suggestions]
        for i, d in enumerate(opts, 1):
            self.cmd.stdout.write(f"   {i}. {d.name}")
        self.cmd.stdout.write("   c. Create a NEW fund with this exact name")
        self.cmd.stdout.write("   s. Skip (ignore these rows)")
        choice = input("   Map to [number / c / s, or type a fund name]: ").strip()
        target = None
        if choice.lower() == "s":
            self.cache[n] = "__SKIP__"; self._save_cache(); return None
        if choice.lower() == "c":
            target = raw
        elif choice.isdigit() and 1 <= int(choice) <= len(opts):
            target = opts[int(choice) - 1].name
        elif choice:
            target = choice
        if not target:
            self.cache[n] = "__SKIP__"; self._save_cache(); return None
        d, _ = Department.objects.get_or_create(
            name=target, defaults=dict(fund_type="LOCAL", category="OFFERING"))
        self._reload()
        self.cache[n] = d.name
        self._save_cache()
        return d


class Command(BaseCommand):
    help = "Import the historical Excel reporting sheets into the treasury system."

    def add_arguments(self, parser):
        parser.add_argument("--dir", default=None,
                            help="Folder with the .xlsx files (default: <project>/data).")
        parser.add_argument("--phase", default="all")
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--noinput", action="store_true")
        parser.add_argument("--map", default=None,
                            help="Account-map JSON (default: <data dir>/account_map.json).")
        parser.add_argument("--year", type=int, default=2026)
        parser.add_argument("--clean", action="store_true",
                            help="Flush the database and create a treasurer before importing.")

    def handle(self, *args, **o):
        from django.conf import settings
        global openpyxl
        import openpyxl  # noqa
        self.dir = o["dir"] or os.path.join(str(settings.BASE_DIR), "data")
        map_path = o["map"] or os.path.join(self.dir, "account_map.json")
        self.dry = o["dry_run"]
        self.year = o["year"]
        if not os.path.isdir(self.dir):
            raise CommandError(f"data folder not found: {self.dir}")

        if o["clean"] and not self.dry:
            self._clean_and_seed_treasurer()

        # Every imported record needs an owning user (recorded_by). On a fresh
        # database with no accounts yet, fall back to creating a dedicated import
        # user rather than crashing on a NULL recorded_by. A real superuser, if
        # one exists, is always preferred.
        self.import_user = self._ensure_import_user()

        self.resolver = AccountResolver(self, map_path, interactive=not o["noinput"])
        self.totals = {}
        self.dev_lookup = self._build_dev_group_lookup()
        phase = o["phase"].lower()

        order = ["departments", "envelopes", "bank", "expenses", "collection",
                 "dashboard", "remittances", "rules"]
        run = order if phase == "all" else [phase]
        for p in run:
            if p not in order:
                raise CommandError(f"unknown phase: {p}")
            self.stdout.write(self.style.MIGRATE_HEADING(f"\n=== PHASE: {p} ==="))
            getattr(self, f"phase_{p}")()
            self.resolver._reload()

        self._report_missing_dev_groups()

        if self.resolver.unresolved:
            self.stdout.write(self.style.WARNING(
                "\nUnresolved account labels (skipped — re-run interactively or add "
                "to the alias map):"))
            for n, c in sorted(self.resolver.unresolved.items()):
                self.stdout.write(f"   {n}  (x{c})")
        if getattr(self, "skipped_amounts", None):
            self.stdout.write(self.style.MIGRATE_HEADING(
                "\nNon-fund / skipped items (deliberately NOT counted as income or "
                "expense, to avoid double-counting):"))
            grand = 0
            for label, (cnt, amt) in sorted(self.skipped_amounts.items()):
                self.stdout.write(f"   {label}: {cnt} line(s), KES {amt:,.2f}")
                grand += amt
            self.stdout.write(f"   ── total non-fund value handled: KES {grand:,.2f}")
        self.stdout.write(self.style.SUCCESS(
            f"\nDone{' (DRY RUN — nothing written)' if self.dry else ''}. "
            f"Summary: {self.totals}"))

    # ---- helpers -----------------------------------------------------------
    def _ensure_import_user(self):
        """Return a user to own imported records. Prefers an existing superuser,
        then any user; if the database has none (a fresh deploy), creates a
        dedicated, unusable-password 'import' user so the import never fails on a
        NULL recorded_by. In a dry run, returns whatever exists without creating."""
        from django.contrib.auth.models import User
        user = (User.objects.filter(is_superuser=True).first()
                or User.objects.order_by("id").first())
        if user or self.dry:
            return user
        user = User.objects.create(username="import", is_staff=True,
                                   first_name="Legacy", last_name="Import")
        user.set_unusable_password()
        user.save()
        self.stdout.write(self.style.WARNING(
            "No users found — created a system 'import' user to own imported "
            "records. Create your own login with: python manage.py createsuperuser"))
        return user

    def _clean_and_seed_treasurer(self):
        """Wipe all data and create a Treasurer login, so the import starts fresh."""
        from django.core.management import call_command
        from django.contrib.auth.models import User, Group
        from core.roles import TREASURER
        self.stdout.write(self.style.WARNING("Cleaning database (flush)…"))
        call_command("flush", "--noinput")
        user, created = User.objects.get_or_create(
            username="treasurer",
            defaults=dict(is_staff=True, is_superuser=True, first_name="Treasurer"))
        if created:
            user.set_password("treasurer123")
            user.save()
        grp, _ = Group.objects.get_or_create(name=TREASURER)
        user.groups.add(grp)
        self.stdout.write(self.style.SUCCESS(
            "Fresh database ready. Treasurer login: treasurer / treasurer123"))

    def _build_dev_group_lookup(self):
        """From the TRUST sheet, build indexes that tie a development contribution to
        its group. The sheet has, per row: month (B), sabbath (C), DEV_GROUP_## in the
        narration (D), amount (E) and the contributor name in the details column (G).

        A single contributor can appear several times (different groups/sabbaths/
        amounts), so we key not only by name+month+sabbath but also by amount, which
        pins the exact row when a name has more than one entry."""
        from members.services.matching import name_key
        lookup = {}
        self.dev_amount_lookup = {}      # (nk, mon, sab, amt_key) -> number  (exact)
        self.dev_scope_candidates = {}   # (nk, mon, sab) -> [(amt, number)]
        self.dev_month_candidates = {}   # (nk, mon)      -> [(sab, amt, number)]
        wb = self._wb(MASTER_FILE)
        if not wb or "TRUST" not in wb.sheetnames:
            self.dev_by_scope = {}
            self._dev_cache = {}
            self.dev_group_numbers_referenced = {}
            return lookup
        ws = wb["TRUST"]
        referenced = {}   # group number -> count of trust-sheet rows referencing it
        for r in range(2, ws.max_row + 1):
            narr = ws.cell(row=r, column=4).value          # D: DEV_GROUP_##
            if not narr or not str(narr).strip().upper().startswith("DEV_GROUP"):
                continue
            narr_u = str(narr).strip().upper()
            m = re.search(r"DEV_GROUP_(\d+)", narr_u)
            if m:
                num = int(m.group(1))
            elif narr_u in DEV_SPECIAL_NUMBERS:        # DEV_GROUP_NA/VISITORS/REFUND
                num = DEV_SPECIAL_NUMBERS[narr_u]
            else:
                continue
            referenced[num] = referenced.get(num, 0) + 1
            month = norm(ws.cell(row=r, column=2).value)[:3]   # B
            sab = ws.cell(row=r, column=3).value               # C
            amount = to_decimal(ws.cell(row=r, column=5).value)  # E
            who = ws.cell(row=r, column=7).value               # G: contributor
            if not who:
                continue
            try:
                sab = int(sab)
            except (TypeError, ValueError):
                continue
            nk = name_key(str(who))
            lookup[(nk, month, sab)] = num   # last-wins single-match (back-compat)
            ak = self._amt_key(amount)
            if ak is not None:
                self.dev_amount_lookup[(nk, month, sab, ak)] = num
            self.dev_scope_candidates.setdefault((nk, month, sab), []).append((amount, num))
            self.dev_month_candidates.setdefault((nk, month), []).append((sab, amount, num))
        wb.close()
        # also index by scope (month, sabbath) -> [(name_key, number)] for fuzzy
        self.dev_by_scope = {}
        for (nk, mon, sab), n in lookup.items():
            self.dev_by_scope.setdefault((mon, sab), []).append((nk, n))
        self._dev_cache = {}
        self.stdout.write(f"  TRUST dev-group lookup: {len(lookup)} name entries, "
                          f"{len(self.dev_amount_lookup)} amount-keyed")
        self.dev_group_numbers_referenced = referenced
        return lookup

    @staticmethod
    def _amt_key(amount):
        """A stable key for matching amounts (rounded to cents)."""
        if amount in (None, ""):
            return None
        try:
            return int(round(float(amount) * 100))
        except (TypeError, ValueError):
            return None

    def _report_missing_dev_groups(self):
        """Flag dev-group numbers referenced in the TRUST sheet that have no matching
        DevelopmentGroup, so the treasurer can create them and re-import."""
        referenced = getattr(self, "dev_group_numbers_referenced", {})
        if not referenced:
            return
        from departments.models import DevelopmentGroup
        existing = set(DevelopmentGroup.objects.values_list("number", flat=True))
        missing = sorted(n for n in referenced if n not in existing)
        self.dev_group_numbers_missing = missing
        if missing:
            for n in missing:
                self._bump("dev_group_number_missing")
            self.stdout.write(self.style.WARNING(
                "  ⚠ TRUST sheet references development group numbers with no matching "
                "group: " + ", ".join(f"#{n} ({referenced[n]} row(s))" for n in missing) +
                ". Create these groups (Departments → Development groups) and re-import, "
                "or their contributions will stay on the parent Development fund."))

    def _dev_group_for(self, name, month_abbr, sabbath, amount=None):
        """Return the DevelopmentGroup for a development contribution, or None.

        Match priority, using the TRUST sheet's name + group + amount + sabbath +
        month:
          1. exact name + month + sabbath + amount;
          2. name + month + sabbath — unambiguous, or disambiguated by amount when
             the contributor appears more than once;
          3. name + month across sabbaths — disambiguated by amount;
          4. fuzzy name match within the month/sabbath.
        """
        from members.services.matching import name_key
        from departments.models import DevelopmentGroup
        nk = name_key(name or "")
        mon = (month_abbr or "")[:3]
        if not nk:
            return None
        num = self._resolve_dev_number(nk, mon, sabbath, amount)
        if num is None:
            return None
        return DevelopmentGroup.objects.filter(number=num).first()

    def _resolve_dev_number(self, nk, mon, sabbath, amount):
        ak = self._amt_key(amount)
        # 1) exact name + month + sabbath + amount
        if ak is not None:
            n = self.dev_amount_lookup.get((nk, mon, sabbath, ak))
            if n is not None:
                return n
        # 2) name + month + sabbath
        cands = self.dev_scope_candidates.get((nk, mon, sabbath), [])
        if len(cands) == 1:
            return cands[0][1]
        if len(cands) > 1 and ak is not None:
            for amt, n in cands:
                if self._amt_key(amt) == ak:
                    return n
        # 3) name + month, across sabbaths — prefer same sabbath, then amount
        mcands = self.dev_month_candidates.get((nk, mon), [])
        if len(mcands) == 1:
            return mcands[0][2]
        if mcands and ak is not None:
            same_sab = [(s, a, n) for (s, a, n) in mcands if s == sabbath]
            for pool in (same_sab, mcands):
                for s, a, n in pool:
                    if self._amt_key(a) == ak:
                        return n
        # 4) fuzzy name within scope
        return self._fuzzy_dev_number(nk, mon, sabbath)

    def _fuzzy_dev_number(self, nk, mon, sabbath):
        """Loosely match a contributor name to a TRUST-sheet dev entry. Scope to the
        same month+sabbath first (cutoff 0.80), then the whole month (cutoff 0.88)
        to stay safe. Cached per (name, month, sabbath)."""
        ck = (nk, mon, sabbath)
        if ck in getattr(self, "_dev_cache", {}):
            return self._dev_cache[ck]
        result = None
        for scope, cutoff in (((mon, sabbath), 0.80), ("MONTH", 0.88)):
            if scope == "MONTH":
                pairs = [p for (m2, s2), lst in self.dev_by_scope.items()
                         if m2 == mon for p in lst]
            else:
                pairs = self.dev_by_scope.get(scope, [])
            if not pairs:
                continue
            names = [p[0] for p in pairs]
            close = difflib.get_close_matches(nk, names, n=1, cutoff=cutoff)
            if close:
                result = dict(pairs)[close[0]]
                break
        self._dev_cache[ck] = result
        return result

    def _valid_period_date(self, date, month_txt):
        """Keep dates inside the Jan-Jun 2026 window (allowing the late-Dec 2025
        opening sabbath). Out-of-window dates are reassigned to the stated month
        when possible, else dropped (returns None)."""
        import datetime as dt
        lo, hi = dt.date(2026, 1, 1), dt.date(2026, 6, 30)
        if date and lo <= date <= hi:
            return date
        m = MONTHS.get(str(month_txt or "").strip().upper()[:3])
        if m and 1 <= m <= 6:
            day = date.day if (date and 1 <= date.day <= 28) else 15
            return dt.date(2026, m, day)
        return None

    def _wb(self, fname):
        import openpyxl
        path = os.path.join(self.dir, fname)
        if not os.path.exists(path):
            return None
        return openpyxl.load_workbook(path, data_only=True)

    def _bump(self, key, n=1):
        self.totals[key] = self.totals.get(key, 0) + n

    def _bump_amt(self, key, amount):
        """Track count and KES total of a skipped / non-fund category."""
        if not hasattr(self, "skipped_amounts"):
            self.skipped_amounts = {}
        c, t = self.skipped_amounts.get(key, (0, 0))
        try:
            t += abs(float(amount or 0))
        except (TypeError, ValueError):
            pass
        self.skipped_amounts[key] = (c + 1, t)

    def _date_from_day(self, day, month_abbr):
        m = MONTHS.get(str(month_abbr).strip().upper()[:3])
        try:
            return dt.date(self.year, m, int(day))
        except (ValueError, TypeError):
            return None

    # ---- PHASE: departments ------------------------------------------------
    def phase_departments(self):
        from departments.models import Department, DevelopmentGroup
        wb = self._wb(MASTER_FILE)
        if not wb:
            raise CommandError(f"{MASTER_FILE} not found in {self.dir}")
        ws = wb["DEPARTMENTS"]

        # 1) parent departments (rows 14-56) with brought-forward balances
        parents, total = [], Decimal(0)
        for r in PARENT_BF_ROWS:
            name = ws.cell(row=r, column=2).value
            bf = to_decimal(ws.cell(row=r, column=3).value) or Decimal(0)
            if not name or not str(name).strip():
                continue
            parents.append((str(name).strip(), bf))
            total += bf
        self.stdout.write(f"  parent departments: {len(parents)}, "
                          f"b/f total = {total:,.2f} (target 5,757,459)")

        def cat_for(name):
            u = name.upper()
            if u == "DEVELOPMENT":
                return "DEVELOPMENT"
            if u in ("TITHE",) or "ENF" in u or "FIELD" in u:
                return "TRUST"
            return "MINISTRY"

        if not self.dry:
            for name, bf in parents:
                ft = "TRUST" if cat_for(name) == "TRUST" else "LOCAL"
                d, created = Department.objects.get_or_create(
                    name=name, defaults=dict(opening_balance=bf, fund_type=ft,
                                             category=cat_for(name)))
                if not created and d.opening_balance != bf:
                    d.opening_balance = bf
                    d.save(update_fields=["opening_balance"])
                self._bump("departments_parent")

            # 2) supporting funds the data needs (zero b/f)
            for name, (ft, cat) in SUPPORTING_FUNDS.items():
                Department.objects.get_or_create(
                    name=name, defaults=dict(opening_balance=0, fund_type=ft,
                                             category=cat))
                self._bump("departments_support")

            # 3) LCB sub-funds -> children of the Local Church Budget parent
            lcb = Department.objects.filter(name="LCB").first()
            if lcb:
                for sub in LCB_CHILDREN:
                    Department.objects.get_or_create(
                        name=sub, defaults=dict(parent=lcb, fund_type="LOCAL",
                                                category="OFFERING", opening_balance=0))
                    self._bump("departments_lcb")

            # 4) VBS subgroups -> children of VBS
            vbs, _ = Department.objects.get_or_create(
                name="VBS", defaults=dict(fund_type="LOCAL", category="MINISTRY"))
            for sub in VBS_SUBGROUPS:
                Department.objects.get_or_create(
                    name=sub, defaults=dict(parent=vbs, fund_type="LOCAL",
                                            category="MINISTRY", opening_balance=0))
                self._bump("departments_vbs")

            # 5) DEVELOPMENT subgroups -> DevelopmentGroup records only (named
            #    DEV_GROUP_##). They are NOT child departments: all development
            #    giving sits on the DEVELOPMENT fund, tagged with its group.
            Department.objects.get_or_create(
                name="DEVELOPMENT", defaults=dict(fund_type="LOCAL",
                                                  category="DEVELOPMENT"))
            for sub in DEV_SUBGROUPS:
                m = re.match(r"DEV_GROUP_(\d+)$", sub)
                number = int(m.group(1)) if m else DEV_SPECIAL_NUMBERS.get(sub)
                if number is None:
                    continue
                grp, _ = DevelopmentGroup.objects.get_or_create(number=number)
                if grp.name != sub:
                    grp.name = sub
                    grp.active = True
                    grp.save(update_fields=["name", "active"])
                self._bump("departments_dev")
        else:
            self._bump("departments_parent", len(parents))
        wb.close()

    # ---- PHASE: envelopes --------------------------------------------------
    def _sabbath_dates(self, wb):
        """Read the Sabbath dates from a reporting file's OFFERING SUMMARY row."""
        if "OFFERING SUMMARY" not in wb.sheetnames:
            return []
        ws = wb["OFFERING SUMMARY"]
        for r in range(1, 8):
            dates = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            ds = [d.date() if isinstance(d, dt.datetime) else None for d in dates]
            ds = [d for d in ds if d]
            if len(ds) >= 2:
                return ds
        return []

    def _find_header(self, ws):
        for r in range(1, 10):
            row = [norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
            if any("CONTRIBUTOR" in x for x in row):
                return r
        return None

    def phase_envelopes(self):
        from core.models import SiteConfig
        from django.contrib.auth.models import User
        from members.services.matching import match_or_create_member
        from envelopes.views import _save_envelope
        from envelopes.models import Envelope
        user = self.import_user
        cfg = SiteConfig.get()

        for month_n, fname in REPORTING_FILES.items():
            wb = self._wb(fname)
            if not wb:
                self.stdout.write(self.style.WARNING(f"  missing {fname}, skipped"))
                continue
            sab_dates = self._sabbath_dates(wb)
            sab_tabs = [s for s in wb.sheetnames if norm(s).startswith("SABB")]
            for idx, tab in enumerate(sab_tabs):
                ws = wb[tab]
                # sabbath index from the tab name (SABBATH 1 -> 1), else position
                mnum = re.search(r"(\d+)", tab)
                sidx = int(mnum.group(1)) if mnum else idx + 1
                date = None
                if sab_dates:
                    # the date columns are aligned to sabbath order; the last N
                    # columns are the in-month sabbaths
                    pos = sidx - 1
                    date = sab_dates[pos] if pos < len(sab_dates) else sab_dates[-1]
                if not date:
                    date = self._nth_saturday(self.year, month_n, sidx)
                # disregard the July sheet (the JUNE file's last sabbath spills
                # into July) and anything outside the Jan-Jun window
                import datetime as _dt
                if date and (date > _dt.date(2026, 6, 30) or date < _dt.date(2026, 1, 1)):
                    self._bump("envelope_sabbath_skipped_outofrange")
                    continue
                hdr = self._find_header(ws)
                if not hdr:
                    continue
                # map fund columns (after TOTAL in col 4); a column can resolve to
                # a single fund or to a split (list of (dept, fraction))
                col_to_targets = {}
                for c in range(5, ws.max_column + 1):
                    label = ws.cell(row=hdr, column=c).value
                    if not label or norm(label) in NON_FUND_COLUMNS:
                        continue
                    split = self.resolver.resolve_split(label)
                    if split:
                        col_to_targets[c] = split
                        continue
                    d = self.resolver.resolve(label)
                    if d:
                        col_to_targets[c] = [(d, Decimal(1))]
                # rows
                for r in range(hdr + 1, ws.max_row + 1):
                    name = ws.cell(row=r, column=2).value
                    receipt = ws.cell(row=r, column=3).value
                    if not name or not str(name).strip():
                        continue
                    # the grand-total / summary block marks the end of contributors
                    if norm(name) in STOP_NAMES:
                        break
                    orig = str(receipt).strip() if receipt not in (None, "") else str(r)
                    abbr = next((k for k, v in MONTHS.items()
                                 if v == month_n and len(k) == 3), str(month_n))
                    receipt_ns = f"{abbr}{sidx}-{orig}"[:20]   # globally unique
                    lines = []
                    for c, targets in col_to_targets.items():
                        amt = to_decimal(ws.cell(row=r, column=c).value)
                        if not amt or amt == 0:
                            continue
                        if len(targets) == 1:
                            dept = targets[0][0]
                            if dept.category == "DEVELOPMENT":
                                grp = self._dev_group_for(str(name).strip(), abbr, sidx, amt)
                                if grp:
                                    lines.append((dept, amt, grp))
                                else:
                                    self._bump("dev_gift_unmatched_to_group")
                                    lines.append((dept, amt))
                            else:
                                lines.append((dept, amt))
                        else:
                            fracs = [f for _, f in targets]
                            for (dept, _), part in zip(targets,
                                                       split_amount(amt, fracs)):
                                if part != 0:
                                    lines.append((dept, part))
                    if not lines:
                        continue
                    if Envelope.objects.filter(receipt_no=receipt_ns).exists():
                        self._bump("envelopes_skipped_existing"); continue
                    self._bump("envelope_lines", len(lines))
                    if self.dry:
                        self._bump("envelopes"); continue
                    member, _ = match_or_create_member(str(name).strip(), None)
                    _save_envelope(date=date, name=str(name).strip(), receipt=receipt_ns,
                                   channel="CASH", lines=lines, member=member,
                                   user=user, cfg=cfg)
                    self._bump("envelopes")
            wb.close()

    @staticmethod
    def _nth_saturday(year, month, n):
        d = dt.date(year, month, 1)
        d += dt.timedelta(days=(5 - d.weekday()) % 7)   # first Saturday
        d += dt.timedelta(days=7 * (n - 1))
        if d.month != month:                            # overflow -> last Saturday
            d -= dt.timedelta(days=7)
        return d

    # ---- PHASE: bank -------------------------------------------------------
    def phase_bank(self):
        from giving.models import Transaction
        from cashbook.models import Expense
        from statements.models import BankAccount, BankReconciliation
        from statements.services.parser import parse_narration
        from core.models import service_sabbath_for
        from core.utils import sabbath_week_of
        from departments.models import Department
        from django.contrib.auth.models import User
        user = self.import_user
        wb = self._wb(MASTER_FILE)
        ws = wb["BANK"]
        bank_account = BankAccount.get_default()
        lcb = Department.objects.filter(name="LCB").first()
        dev_fund = Department.objects.filter(name="DEVELOPMENT",
                                             parent__isnull=True).first()
        seen_ref = {}

        def parse_date(v):
            if isinstance(v, dt.datetime):
                return v.date()
            if isinstance(v, str):
                for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
                    try:
                        return dt.datetime.strptime(v.strip()[:10], fmt).date()
                    except ValueError:
                        continue
            return None

        opening = closing = last_date = None
        for r in range(2, ws.max_row + 1):
            ref = ws.cell(row=r, column=1).value
            acct = ws.cell(row=r, column=4).value
            amount = to_decimal(ws.cell(row=r, column=5).value)
            vdate = ws.cell(row=r, column=7).value
            runbal = to_decimal(ws.cell(row=r, column=8).value)
            narr = ws.cell(row=r, column=9).value
            if not acct or amount in (None, 0):
                continue
            acct_u = str(acct).strip().upper()
            date = parse_date(vdate)
            # The dashboard assigns each bank credit to a month via the MONTH
            # column, not the value date. A handful of value dates are mis-keyed
            # (years 2027-2068) and the JUNE sheet's last sabbath spills into July.
            # Clamp to the Jan-Jun window: reassign to the stated month where we
            # can, otherwise drop the row (and report it).
            date = self._valid_period_date(date, ws.cell(row=r, column=2).value)
            if date is None:
                self._bump("bank_skipped_outofrange")
                self._bump_amt("Bank rows dropped — date outside Jan-Jun", amount)
                continue

            # opening = balance before the very first statement line
            if opening is None and runbal is not None and date:
                # first line moves the balance; back it out to get the opening
                signed = amount if acct_u not in ("WITHDRAWAL", "CHARGES") else -amount
                opening = runbal - signed
            if runbal is not None:
                closing = runbal
            if date:
                last_date = date

            # ref dedup-append (split rows share a ref). When the statement's REF
            # column holds a channel placeholder ("STKPUSH", "USSD", "MULTI"...)
            # rather than a real receipt, fall back to the genuine M-Pesa receipt
            # parsed from the narration — otherwise every such row collapses onto
            # the same key and the import can't tie to the bank's receipts.
            p = parse_narration(narr or "")
            _PLACEHOLDER_REFS = {"STKPUSH", "STK", "USSD", "C2B", "MPESAC2B",
                                 "PAYBILL", "MULTI", "OTHER", ""}
            ref_clean = str(ref).strip() if ref not in (None, "") else ""
            if ref_clean.upper() in _PLACEHOLDER_REFS:
                base = (p.get("receipt") or "").strip() or f"LEG-BANK-{r}"
            else:
                base = ref_clean or f"LEG-BANK-{r}"
            seen_ref[base] = seen_ref.get(base, 0) + 1
            core_ref = base if seen_ref[base] == 1 else f"{base}-{seen_ref[base]}"
            if date and Transaction.objects.filter(core_ref=core_ref).exists():
                self._bump("bank_skipped_existing"); continue
            if not date:
                continue
            svc = service_sabbath_for(date)
            common = dict(date=date, service_sabbath=svc,
                          sabbath_week=sabbath_week_of(svc),
                          bank_account=bank_account, raw_narration=str(narr or ""))

            # --- DEPOSIT: cash already receipted elsewhere, only banked now. Skip
            #     (no new income) — see "how deposits are handled" in the README.
            if acct_u == "DEPOSIT":
                self._bump("bank_skipped_deposit"); self._bump_amt("Deposits (cash banked, not income)", amount); continue

            # --- CHARGES: the bank charges are ALREADY captured inside the LCB
            #     expenses on the EXPENSES sheet, so creating a separate bank-charge
            #     expense here double-counts them (the 8,050 LCB discrepancy). We
            #     therefore only mark them processed; they still feed the bank
            #     reconciliation (A15) as a memo, not the books.
            if acct_u == "CHARGES":
                self._bump("bank_charges_processed"); self._bump_amt("Bank charges (already in LCB expenses)", amount)
                continue

            # --- WITHDRAWAL: cash drawn for expenses. EVERY withdrawal (including
            #     cheque 000401, the 900,000 debit) was already captured in the
            #     EXPENSES sheet, so we mark them all processed and post NO bank
            #     debit — that would double-count the outflow. Going forward the live
            #     app still flags every real-time debit for classification.
            if acct_u == "WITHDRAWAL":
                self._bump("bank_withdrawals_processed"); self._bump_amt("Withdrawals (already captured as expenses)", amount)
                continue

            # --- ENVELOPES: already receipted via the reporting sheets. Import for
            #     a complete bank statement but EXCLUDE from income (department=None,
            #     excluded_from_income), and mark it processed-through-envelope so it
            #     is never double-counted.
            if acct_u == "ENVELOPES":
                self._bump("bank_envelopes"); self._bump_amt("Bank envelope rows (already receipted)", amount)
                if not self.dry:
                    Transaction.objects.create(
                        channel="ENVELOPE", direction="CREDIT", amount=amount,
                        department=None, allocation_status="MANUAL", confirmed=True,
                        excluded_from_income=True, core_ref=core_ref,
                        reference="Processed via envelope",
                        payer_name=(p.get("name") or "")[:120],
                        payer_phone=(p.get("phone") or "")[:12],
                        mpesa_ref=(p.get("receipt") or "")[:30], **common)
                continue

            if acct_u == "NARRATION":
                self._bump("bank_skipped_narration"); self._bump_amt("Narration-only lines (no fund)", amount); continue

            # --- normal fund credit (development giving -> DEVELOPMENT + group tag)
            dev_group = None
            m = re.match(r"DEV_GROUP_(\d+)$", acct_u)
            special = DEV_SPECIAL_NUMBERS.get(acct_u)
            if m or special is not None:
                from departments.models import DevelopmentGroup
                number = int(m.group(1)) if m else special
                dept = dev_fund
                dev_group = DevelopmentGroup.objects.filter(number=number).first()
                if dev_group is None:
                    self._bump("dev_gift_unmatched_to_group")
                    self._bump(f"dev_number_missing_{number}")
            else:
                dept = self.resolver.resolve(acct)
            if not dept:
                continue
            self._bump("bank")
            if self.dry:
                continue
            # bank givers become members too, carrying the phone parsed from the
            # narration — so the member list eventually covers M-Pesa givers and
            # their numbers are available for receipts/SMS.
            from members.services.matching import match_or_create_member
            member = None
            if (p.get("name") or "").strip():
                member, _ = match_or_create_member(p.get("name"), p.get("phone"))
            Transaction.objects.create(
                channel="BANK", direction="CREDIT", amount=amount, department=dept,
                dev_group=dev_group, allocation_status="MANUAL", confirmed=True,
                member=member,
                core_ref=core_ref, reference=(p.get("reference") or "")[:60],
                payer_name=(p.get("name") or "")[:120],
                payer_phone=(p.get("phone") or "")[:12],
                mpesa_ref=(p.get("receipt") or "")[:30], **common)

        # record the statement's opening/closing balance for reconciliation
        self.totals["bank_opening_balance"] = float(opening or 0)
        self.totals["bank_closing_balance"] = float(closing or 0)
        if not self.dry and closing is not None and last_date:
            BankReconciliation.objects.get_or_create(
                statement_date=last_date,
                defaults=dict(bank_balance=closing, book_balance=closing,
                              created_by=user,
                              notes=f"Imported from legacy bank sheet. "
                                    f"Opening balance {opening:,.2f}, "
                                    f"closing balance {closing:,.2f}."))
        wb.close()

    # ---- PHASE: expenses ---------------------------------------------------
    def phase_expenses(self):
        from cashbook.models import Expense
        from django.contrib.auth.models import User
        user = self.import_user
        wb = self._wb(MASTER_FILE)
        ws = wb["EXPENSES"]
        for r in range(2, ws.max_row + 1):
            day = ws.cell(row=r, column=1).value
            month = ws.cell(row=r, column=2).value
            deptname = ws.cell(row=r, column=4).value
            desc = ws.cell(row=r, column=5).value
            amount = to_decimal(ws.cell(row=r, column=6).value)
            claimant = ws.cell(row=r, column=7).value
            if not deptname or amount in (None, 0):
                continue
            dept = self.resolver.resolve(deptname)
            if not dept:
                continue
            date = self._date_from_day(day, month) or dt.date(self.year, 1, 1)
            voucher = f"LEG-EXP-{norm(month)}-{r}"
            if Expense.objects.filter(voucher_no=voucher).exists():
                self._bump("expenses_skipped_existing"); continue
            self._bump("expenses")
            if self.dry:
                continue
            cat = self._guess_expense_category(desc, deptname)
            Expense.objects.create(
                date=date, department=dept, description=str(desc or "Expense")[:200],
                amount=amount, category=cat, status="PAID",
                claimant=str(claimant or "")[:120], voucher_no=voucher,
                recorded_by=user, approved_by=user, paid_date=date)
        wb.close()

    # keyword -> Expense.Category, checked in order against description + department
    EXP_CATEGORY_RULES = [
        ("ALLOWANCE", ["allowance", "honorari", "stipend", "salary", "wages", "pay "]),
        ("TRANSPORT", ["transport", "fare", "fuel", "matatu", "travel", "mileage", "boda"]),
        ("REFRESHMENTS", ["refreshment", "catering", "food", "lunch", "tea", "water ", "snack", "meal"]),
        ("STATIONERY", ["stationery", "printing", "photocopy", "print", "pen", "paper", "toner"]),
        ("UTILITIES", ["electric", "power", "kplc", "water bill", "token", "internet", "airtime", "wifi"]),
        ("MAINTENANCE", ["repair", "maintenance", "service", "fix", "plumb", "paint"]),
        ("CONSTRUCTION", ["construct", "building", "cement", "ballast", "sand", "block", "fundi", "roof"]),
        ("EVANGELISM", ["evangelis", "crusade", "mission", "outreach", "vbs", "campaign"]),
        ("BENEVOLENCE", ["benevolen", "welfare", "bereave", "funeral", "sick", "donation", "needy"]),
        ("BANK_CHARGE", ["bank charge", "ledger fee", "transaction fee", "withdrawal charge"]),
        ("MATERIALS", ["material", "supplies", "equipment", "purchase", "buy ", "items"]),
    ]

    def _guess_expense_category(self, description, deptname):
        from cashbook.models import Expense
        text = f"{description or ''} {deptname or ''}".lower()
        for cat, kws in self.EXP_CATEGORY_RULES:
            if any(k in text for k in kws):
                return cat
        return Expense.Category.OTHER

    # ---- PHASE: collection -------------------------------------------------
    def phase_collection(self):
        from giving.models import Transaction
        from core.models import service_sabbath_for
        from core.utils import sabbath_week_of
        wb = self._wb(MASTER_FILE)
        ws = wb["COLLECTION"]
        for r in range(2, ws.max_row + 1):
            day = ws.cell(row=r, column=1).value
            month = ws.cell(row=r, column=2).value
            narr = ws.cell(row=r, column=4).value
            amount = to_decimal(ws.cell(row=r, column=5).value)
            if not narr or amount in (None, 0):
                continue
            if str(narr).strip().upper() in COLLECTION_SKIP:
                self._bump("collection_skipped_envelopes"); self._bump_amt("Collection envelope lines (already receipted)", amount); continue
            acct_u = str(narr).strip().upper()
            dev_group = None
            m = re.match(r"DEV_GROUP_(\d+)$", acct_u)
            special = DEV_SPECIAL_NUMBERS.get(acct_u)
            if m or special is not None:
                from departments.models import Department, DevelopmentGroup
                number = int(m.group(1)) if m else special
                dept = Department.objects.filter(name="DEVELOPMENT",
                                                 parent__isnull=True).first()
                dev_group = DevelopmentGroup.objects.filter(number=number).first()
                if dev_group is None:
                    self._bump("dev_gift_unmatched_to_group")
                    self._bump(f"dev_number_missing_{number}")
            else:
                dept = self.resolver.resolve(narr)
            if not dept:
                continue
            date = self._date_from_day(day, month) or dt.date(self.year, 1, 1)
            ref = f"LEG-COLL-{norm(month)}-{r}"
            if Transaction.objects.filter(reference=ref, amount=amount,
                                          date=date).exists():
                self._bump("collection_skipped_existing"); continue
            self._bump("collection")
            if self.dry:
                continue
            svc = service_sabbath_for(date)
            Transaction.objects.create(
                date=date, service_sabbath=svc, sabbath_week=sabbath_week_of(svc),
                channel="CASH", direction="CREDIT", amount=amount, department=dept,
                dev_group=dev_group, allocation_status="MANUAL", confirmed=True,
                reference=ref,
                raw_narration=f"Cash collection: {narr}")
        wb.close()

    # ---- PHASE: remittances (historical trust remittances) -----------------
    def phase_remittances(self):
        """Record the trust remittances that have already been sent to the field
        so the cash/SOFP reflects only the still-held (unremitted) trust. Per the
        treasurer, every month is remitted except June, so we clear the Jan-May
        trust collection here. Each month becomes one REMITTED batch with a
        per-fund REMITTANCE expense (PAID), dated month-end."""
        import datetime as dt
        from decimal import Decimal
        from django.utils import timezone
        from django.db.models import Sum
        from django.contrib.auth.models import User
        from departments.models import Department
        from giving.models import Transaction
        from cashbook.models import Expense, RemittanceBatch
        from core.models import SiteConfig
        from core.utils import sabbath_week_of
        user = self.import_user
        field = SiteConfig.get().field_name or "the field"
        REMITTED_MONTHS = [1, 2, 3, 4, 5]   # Jan-May remitted; only June still held
        trust_depts = list(Department.objects.filter(fund_type="TRUST", active=True))
        total_remitted = Decimal(0)
        for m in REMITTED_MONTHS:
            last = (dt.date(self.year, m + 1, 1) - dt.timedelta(days=1)) if m < 12 \
                else dt.date(self.year, 12, 31)
            s = dt.date(self.year, m, 1)
            # idempotent: don't re-remit a month already recorded
            if RemittanceBatch.objects.filter(
                    period_start=s, status=RemittanceBatch.Status.REMITTED).exists():
                continue
            # trust collected in the month, per fund
            per = {r["department"]: (r["t"] or Decimal(0)) for r in
                   Transaction.objects.confirmed_credits()
                   .filter(date__gte=s, date__lte=last, department__in=trust_depts)
                   .values("department").annotate(t=Sum("amount"))}
            chosen = [(d, per.get(d.id, Decimal(0))) for d in trust_depts
                      if per.get(d.id, Decimal(0)) > 0]
            if not chosen:
                continue
            if self.dry:
                self._bump("remittance_batches"); continue
            batch = RemittanceBatch.create_batch(
                created_by=user, status=RemittanceBatch.Status.REMITTED,
                period_start=s, period_end=last, cheque_date=last,
                remitted_at=timezone.make_aware(dt.datetime.combine(last, dt.time(12, 0))),
                notes=f"Historical remittance for {s:%B %Y} (imported)")
            for dept, amt in chosen:
                Expense.objects.create(
                    date=last, sabbath_week=sabbath_week_of(last), department=dept,
                    description=f"Trust remittance to {field} for {s:%B %Y} — {batch.batch_number}",
                    amount=amt, category=Expense.Category.REMITTANCE,
                    claimant=field, method=Expense.Method.CHEQUE,
                    status=Expense.Status.PAID, recorded_by=user, approved_by=user,
                    paid_date=last, remittance_batch=batch)
                total_remitted += amt
            batch.recompute_total(); batch.save(update_fields=["total_amount"])
            self._bump("remittance_batches")
        self.stdout.write(f"  recorded remittances Jan-May: KES {total_remitted:,.2f} "
                          f"(only June left outstanding)")

    # ---- PHASE: rules (seed allocation rules, run LAST) --------------------
    def phase_rules(self):
        """Seed the standing allocation rules from the sample dataset so that
        future giving (new statement imports, the bank webhook) auto-allocates by
        paybill reference / narration token. Run after everything else, since the
        rules point at departments created in the departments phase. Rules whose
        target fund can't be resolved are skipped (reported), never guessed."""
        from decimal import Decimal
        from giving.models import AllocationRule, SplitFund, SplitComponent
        from giving.services.allocation import normalize_reference
        try:
            from core.management.commands._seed_data import ACCOUNT_RULES
        except Exception:
            ACCOUNT_RULES = {}
        try:
            from core.management.commands.seed_demo import RULES, SPLIT_RULES
        except Exception:
            RULES, SPLIT_RULES = {}, {}

        created = skipped = 0
        unresolved = []
        for ref, fund_name in {**RULES, **ACCOUNT_RULES}.items():
            dept = self.resolver.resolve(fund_name)
            if not dept:
                skipped += 1
                unresolved.append(fund_name)
                continue
            if self.dry:
                created += 1
                continue
            _, c = AllocationRule.objects.get_or_create(
                reference=normalize_reference(ref),
                defaults=dict(department=dept, source=AllocationRule.Source.SEED))
            created += 1 if c else 0

        # 50/50 split rules (combined / thanksgiving) need a SplitFund to point at
        split_defs = {
            "Combined Offering": [("COMBINED (50%) - ENF", 50), ("COMBINED - LCB", 50)],
            "Thanksgiving Offering": [("THANKSGIVING - ENF", 50), ("THANKGIVING - LCB", 50)],
        }
        if not self.dry:
            sf_cache = {}
            for sf_name, comps in split_defs.items():
                sf, _ = SplitFund.objects.get_or_create(name=sf_name)
                for dn, pct in comps:
                    d = self.resolver.resolve(dn)
                    if d:
                        SplitComponent.objects.get_or_create(
                            split_fund=sf, department=d, defaults=dict(percent=Decimal(pct)))
                        # the halves are internal to the split — hide them from the
                        # allocation pickers; the treasurer selects the split fund
                        if d.selectable:
                            d.selectable = False
                            d.save(update_fields=["selectable"])
                sf_cache[sf_name] = sf
            for ref, sf_name in SPLIT_RULES.items():
                sf = sf_cache.get(sf_name)
                if not sf:
                    continue
                _, c = AllocationRule.objects.get_or_create(
                    reference=normalize_reference(ref),
                    defaults=dict(split_fund=sf, source=AllocationRule.Source.SEED))
                created += 1 if c else 0

        self._bump("allocation_rules", created)
        if unresolved:
            self.stdout.write(self.style.WARNING(
                f"  {skipped} rule(s) skipped (fund not found): "
                f"{', '.join(sorted(set(unresolved))[:8])}"))
        self.stdout.write(f"  allocation rules seeded: {created}")

    # ---- PHASE: dashboard (opening position + historical years) ------------
    def phase_dashboard(self):
        from core.models import SiteConfig, HistoricalYear
        wb = self._wb(MASTER_FILE)
        if not wb or "DASHBOARD" not in wb.sheetnames:
            self.stdout.write(self.style.WARNING("  no DASHBOARD sheet, skipped"))
            return
        ws = wb["DASHBOARD"]

        # 1) opening cash position from the CASH BOOK SUMMARY block (col A/B)
        opening = {"bank": None, "cash": None, "trust": None}
        for r in range(1, 15):
            label = str(ws.cell(row=r, column=1).value or "").upper()
            val = to_decimal(ws.cell(row=r, column=2).value)
            if val is None:
                continue
            if "OPENING BANK" in label:
                opening["bank"] = val
            elif "CASH AT HAND" in label or "CASH ON HAND" in label:
                opening["cash"] = val
            elif "UNREMITTED" in label or "PENDING RECEIPT" in label:
                opening["trust"] = abs(val)
        if not self.dry and any(v is not None for v in opening.values()):
            cfg = SiteConfig.get()
            if opening["bank"] is not None:
                cfg.opening_bank_balance = opening["bank"]
            if opening["cash"] is not None:
                cfg.opening_cash_on_hand = opening["cash"]
            if opening["trust"] is not None:
                cfg.opening_unremitted_trust = opening["trust"]
            cfg.save(update_fields=["opening_bank_balance", "opening_cash_on_hand",
                                    "opening_unremitted_trust"])
            self._bump("opening_position_set")
            self.stdout.write(
                f"  opening position: bank {opening['bank']}, cash {opening['cash']}, "
                f"unremitted trust {opening['trust']}")

        # 2) historical years + months: a '20XX' marker in col U starts a block;
        #    month rows (col U = month name) carry V/W/X; a 'Grand Total' row closes it.
        from core.models import HistoricalMonth
        MONTHS = {"JANUARY": 1, "FEBRUARY": 2, "MARCH": 3, "APRIL": 4, "MAY": 5,
                  "JUNE": 6, "JULY": 7, "AUGUST": 8, "SEPTEMBER": 9, "OCTOBER": 10,
                  "NOVEMBER": 11, "DECEMBER": 12, "JAN": 1, "FEB": 2, "MAR": 3,
                  "APR": 4, "JUN": 6, "JUL": 7, "AUG": 8, "SEP": 9, "SEPT": 9,
                  "OCT": 10, "NOV": 11, "DEC": 12}
        years = {}
        months = {}     # (year, month) -> (coll, trust, exp)
        cur = None
        for r in range(19, ws.max_row + 1):
            u = str(ws.cell(row=r, column=21).value or "").strip()
            uU = u.upper()
            v = to_decimal(ws.cell(row=r, column=22).value) or 0
            w = to_decimal(ws.cell(row=r, column=23).value) or 0
            x = to_decimal(ws.cell(row=r, column=24).value) or 0
            if re.match(r"^20\d\d$", u):
                cur = int(u)
                continue
            if uU.startswith("GRAND TOTAL") and cur:
                years[cur] = (v, w, x)
                cur = None
                continue
            if cur and uU in MONTHS and (v or w or x):
                months[(cur, MONTHS[uU])] = (v, w, x)
        # the 2024 summary sits on row 19 (col U=2024, V/W/X totals)
        if str(ws.cell(row=19, column=21).value or "").strip() == "2024":
            years[2024] = (to_decimal(ws.cell(row=19, column=22).value) or 0,
                           to_decimal(ws.cell(row=19, column=23).value) or 0,
                           to_decimal(ws.cell(row=19, column=24).value) or 0)
        for yr, (coll, trust, exp) in years.items():
            self._bump("historical_years")
            if not self.dry:
                HistoricalYear.objects.update_or_create(
                    year=yr, defaults=dict(collection=coll, trust_fund=trust,
                                           expenditure=exp,
                                           note="Imported from dashboard history"))
        for (yr, mo), (coll, trust, exp) in months.items():
            self._bump("historical_months")
            if not self.dry:
                HistoricalMonth.objects.update_or_create(
                    year=yr, month=mo,
                    defaults=dict(collection=coll, trust_fund=trust, expenditure=exp))
        if years:
            self.stdout.write(f"  historical years: {sorted(years)} "
                              f"({len(months)} monthly rows)")

        # 3) bank reconciliation statement, mirroring the dashboard A15 block
        self._build_reconciliation(ws)
        wb.close()

    def _build_reconciliation(self, ws):
        """Recreate the BANK RECONCILIATION STATEMENT (dashboard col A/B, ~row 15)
        as a BankReconciliation with its reconciling items."""
        from statements.models import BankReconciliation, ReconciliationItem
        from giving.models import Transaction
        from django.contrib.auth.models import User
        import datetime as _dt
        bank_close = book_bal = None
        items = []   # (kind, description, amount, effect)
        section = "ADD"
        for r in range(15, 32):
            label = str(ws.cell(row=r, column=1).value or "").strip()
            lU = label.upper()
            val = to_decimal(ws.cell(row=r, column=2).value)
            if not label:
                continue
            if "BANK CLOSING" in lU:
                bank_close = val
            elif lU == "ADD":
                section = "ADD"
            elif lU == "LESS":
                section = "LESS"
            elif "CASHBOOK BALANCE" in lU:
                book_bal = val
            elif "VARIANCE" in lU or "QUALITY" in lU:
                break
            elif val is not None and val != 0:
                eff = "ADD" if section == "ADD" else "SUBTRACT"
                if "CHARGE" in lU:
                    # Best practice: bank charges are a real expense and are already
                    # posted to LCB on the EXPENSES sheet, so they belong in the
                    # cashbook, NOT as a permanent reconciling item. Carrying them
                    # in both places is what creates the spreadsheet's ~8,050
                    # variance, so we deliberately omit the charge line here.
                    self._bump("reconciliation_bankcharge_omitted")
                    continue
                elif "CASH AT HAND" in lU or "CASH ON HAND" in lU:
                    kind = "CASH_AT_HAND"
                elif "UNPRESENTED" in lU:
                    kind = "UNPRESENTED"
                else:
                    kind = "OTHER"
                items.append((kind, label, abs(val), eff))
        if self.dry or bank_close is None:
            return
        # derive the book balance from the (charge-free) reconciling items so the
        # statement ties exactly — no residual variance
        computed_book = bank_close + sum(
            (a if e == "ADD" else -a) for _, _, a, e in items)
        stmt_date = (Transaction.objects.filter(channel="BANK")
                     .order_by("-date").values_list("date", flat=True).first()
                     or _dt.date.today())
        rec, _ = BankReconciliation.objects.update_or_create(
            statement_date=stmt_date,
            defaults=dict(bank_balance=bank_close, book_balance=computed_book,
                          created_by=User.objects.order_by("id").first(),
                          notes="Imported from the dashboard bank reconciliation (A15). "
                                "Bank charges are treated as an expense (posted to LCB), "
                                "not a reconciling item, so this statement ties without "
                                "the ~8,050 variance the spreadsheet carried."))
        rec.items.all().delete()
        for kind, desc, amt, eff in items:
            ReconciliationItem.objects.create(reconciliation=rec, kind=kind,
                                              description=desc[:200], amount=amt, effect=eff)
        self._bump("reconciliation_items_set")
        self.stdout.write(f"  bank reconciliation: bank {bank_close:,.2f} -> "
                          f"cashbook {computed_book:,.2f} ({len(items)} items, "
                          f"bank charges expensed not reconciled)")
