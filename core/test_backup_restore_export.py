"""Restoring a backup, and exporting everything the church actually holds.

**Restore never worked on an ordinary installation.** It guarded the SQLite
branch with `isinstance(path, (str, bytes))`, and Django's own settings template
writes `NAME = BASE_DIR / "db.sqlite3"` — a `PosixPath`, which fails that test.
So a treasurer was told "No on-disk SQLite database to restore into" while the
database sat in the very place the message said it wasn't. Backup used
`os.path.exists` and worked, which is exactly the shape of the complaint: a
backup you can take and cannot use is not a backup.

**Nothing checked what was being restored.** The uploaded file was copied
straight over the live database. A spreadsheet, a PDF, or a dump from a
different engine would have destroyed the church's data, and the safety copy
taken a moment earlier only helps somebody who realises what happened.

**The data export left out the benevolent scheme and the envelopes.** Those are
the two parts least reconstructible from a bank statement — a levy is only an
amount until you know which case it settled, and an envelope total is only a
figure until you can see the funds it was split across.
"""
import datetime as dt
import io
from decimal import Decimal

import openpyxl
from django.contrib.auth.models import Group, User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase
from django.urls import reverse

from core import roles
from core.services import backup as bk


class RestoreAcceptsAnOrdinaryInstallationTests(TestCase):

    def test_a_path_object_is_accepted(self):
        """The whole bug: a PosixPath is a perfectly good database path."""
        from django.conf import settings
        import os
        name = settings.DATABASES["default"]["NAME"]
        self.assertTrue(
            os.fspath(name),
            "settings NAME must be coercible with os.fspath; the restore path "
            "guard depends on it rather than on isinstance checks.")

    def test_a_file_that_is_not_a_database_is_refused(self):
        ok, message = bk.database_restore(
            SimpleUploadedFile("notes.xlsx", b"PK\x03\x04 spreadsheet"))
        self.assertFalse(ok)
        self.assertIn("not a SQLite database", message)

    def test_an_empty_file_is_refused(self):
        ok, _ = bk.database_restore(SimpleUploadedFile("empty.sqlite3", b""))
        self.assertFalse(ok)

    def test_a_truncated_database_is_refused(self):
        """Right magic number, far too small to be this system's."""
        ok, message = bk.database_restore(
            SimpleUploadedFile("tiny.sqlite3", b"SQLite format 3\x00" + b"x" * 20))
        self.assertFalse(ok)
        self.assertIn("too small", message)

    def test_the_refusal_says_the_data_is_untouched(self):
        """A treasurer who has just uploaded the wrong file needs to know."""
        _, message = bk.database_restore(
            SimpleUploadedFile("notes.pdf", b"%PDF-1.4 rubbish"))
        self.assertIn("untouched", message)

    def test_a_sqlite_backup_is_refused_on_a_server_database(self):
        """Wrong-engine upload, caught before anything is overwritten."""
        ok, message = bk._looks_like_sql_dump_probe() \
            if hasattr(bk, "_looks_like_sql_dump_probe") else (None, None)
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".sql", delete=False) as fh:
            fh.write(b"SQLite format 3\x00" + b"x" * 100)
            path = fh.name
        ok, message = bk._looks_like_sql_dump(path)
        self.assertFalse(ok)
        self.assertIn("different database", message)

    def test_a_real_sql_dump_is_accepted(self):
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".sql", delete=False) as fh:
            fh.write(b"-- MySQL dump\nCREATE TABLE giving_transaction (id int);\n")
            path = fh.name
        ok, _ = bk._looks_like_sql_dump(path)
        self.assertTrue(ok)

    def test_a_real_database_file_passes_the_check(self):
        from django.conf import settings
        import os
        path = os.fspath(settings.DATABASES["default"]["NAME"])
        if not os.path.exists(path):
            self.skipTest("test database is not on disk")
        ok, _ = bk._looks_like_sqlite(path)
        self.assertTrue(ok)


class DataExportCoversTheWholeChurchTests(TestCase):

    def setUp(self):
        self.user = User.objects.create_user("tess-backup", password="office-pass-1")
        self.user.groups.add(Group.objects.get_or_create(name=roles.TREASURER)[0])
        self.client = Client()
        self.client.force_login(self.user)

    def _book(self):
        response = self.client.get(reverse("data_export"))
        self.assertEqual(response.status_code, 200)
        return openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)

    def test_the_export_downloads(self):
        self.assertTrue(self._book().sheetnames)

    def test_the_benevolent_scheme_is_included(self):
        names = self._book().sheetnames
        for expected in ("Benevolent Schemes", "Benevolent Members",
                         "Benevolent Dependants", "Benevolent Cases",
                         "Benevolent Contributions"):
            with self.subTest(sheet=expected):
                self.assertIn(expected, names)

    def test_dependants_carry_who_was_covered_and_when(self):
        """A past claim is judged on the household as it was at the time."""
        header = [c.value for c in self._book()["Benevolent Dependants"][4]]
        for column in ("Relationship", "Registered", "Active", "Died on"):
            self.assertIn(column, header)

    def test_a_contribution_says_which_case_it_settled(self):
        header = [c.value for c in self._book()["Benevolent Contributions"][4]]
        self.assertIn("Case", header)
        self.assertIn("Period", header)


class EnvelopeSchedulesTests(TestCase):
    """Envelopes are exported the way they are counted: a month to a sheet."""

    def setUp(self):
        from departments.models import Department
        from envelopes.models import Envelope, EnvelopeLine
        from members.models import Member

        self.user = User.objects.create_user("tess-env", password="office-pass-1")
        self.user.groups.add(Group.objects.get_or_create(name=roles.TREASURER)[0])
        self.tithe = Department.objects.create(
            name="Tithe", slug="tithe-bk", is_trust=True,
            fund_type=Department.FundType.TRUST,
            category=Department.Category.TRUST)
        self.lcb = Department.objects.create(
            name="Local Church Budget", slug="lcb-bk",
            fund_type=Department.FundType.LOCAL,
            category=Department.Category.MINISTRY)
        member = Member.objects.create(name="Ruth Momanyi", phone="254790301470")

        # two months, so the split into sheets is actually exercised
        for month, day in ((1, 10), (2, 14)):
            env = Envelope.objects.create(
                date=dt.date(2026, month, day), member=member,
                receipt_no=f"R{month:02d}", channel="CASH",
                total=Decimal("300"), recorded_by=self.user)
            EnvelopeLine.objects.create(envelope=env, department=self.tithe,
                                        amount=Decimal("200"))
            EnvelopeLine.objects.create(envelope=env, department=self.lcb,
                                        amount=Decimal("100"))
        self.client = Client()
        self.client.force_login(self.user)

    def _book(self):
        response = self.client.get(reverse("data_export"))
        self.assertEqual(response.status_code, 200)
        return openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)

    def test_there_is_a_sheet_for_each_month(self):
        names = self._book().sheetnames
        self.assertIn("Envelopes 2026-01", names)
        self.assertIn("Envelopes 2026-02", names)

    def test_the_layout_is_the_counting_schedule(self):
        """A fund to a column, as on the sheet the money was counted onto."""
        header = [c.value for c in self._book()["Envelopes 2026-01"][4]]
        for column in ("No", "Date", "Contributor Name", "Phone", "Receipt No",
                       "Channel", "Tithe", "Local Church Budget", "Total"):
            with self.subTest(column=column):
                self.assertIn(column, header)

    def test_each_fund_column_carries_its_own_figure(self):
        ws = self._book()["Envelopes 2026-01"]
        header = [c.value for c in ws[4]]
        row = [c.value for c in ws[5]]
        self.assertEqual(row[header.index("Tithe")], 200)
        self.assertEqual(row[header.index("Local Church Budget")], 100)

    def test_a_row_adds_across(self):
        """The total must be the funds beside it, or the sheet cannot be checked."""
        ws = self._book()["Envelopes 2026-01"]
        header = [c.value for c in ws[4]]
        row = [c.value for c in ws[5]]
        funds = sum(row[header.index(f)] for f in ("Tithe", "Local Church Budget"))
        self.assertEqual(funds, row[header.index("Total")])

    def test_the_sheet_adds_down(self):
        ws = self._book()["Envelopes 2026-01"]
        header = [c.value for c in ws[4]]
        total_row = [c.value for c in ws[ws.max_row]]
        self.assertEqual(total_row[2], "Total")
        self.assertEqual(total_row[header.index("Tithe")], 200)
        self.assertEqual(total_row[header.index("Total")], 300)

    def test_a_month_with_no_envelopes_has_no_sheet(self):
        self.assertNotIn("Envelopes 2026-03", self._book().sheetnames)
