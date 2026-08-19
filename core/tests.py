from django.test import TestCase, override_settings

# Create your tests here.


class AssistantTests(TestCase):
    def setUp(self):
        import datetime as dt
        from decimal import Decimal
        from departments.models import Department
        from giving.models import Transaction
        self.dept = Department.objects.create(name="Tithe", fund_type=Department.FundType.TRUST)
        Transaction.objects.create(date=dt.date.today(), channel="CASH", direction="CREDIT",
                                   amount=Decimal("1000"), department=self.dept, allocation_status="AUTO")

    def test_collections_intent(self):
        from core.services import assistant
        d = assistant.answer("total collections this month")
        self.assertIn("Total received", d["text"])

    def test_balance_intent(self):
        from core.services import assistant
        d = assistant.answer("balance of Tithe")
        self.assertIn("Tithe", d["text"])
        self.assertTrue(d.get("rows"))

    def test_spending_not_confused_with_pending(self):
        from core.services import assistant
        d = assistant.answer("spending this month")
        self.assertIn("Expenses", d["text"])


class SmsScopeTests(TestCase):
    def test_scope_off_blocks_send(self):
        import datetime as dt
        from decimal import Decimal
        from core.models import SiteConfig
        from members.models import Member
        from envelopes.models import Envelope
        from core.services.sms import send_receipt_sms
        from django.contrib.auth.models import User
        u = User.objects.create_user("sms1", password="x")
        cfg = SiteConfig.get(); cfg.sms_enabled = True
        cfg.sms_receipt_scope = SiteConfig.SmsReceiptScope.OFF; cfg.save()
        m = Member.objects.create(name="A B", phone="254700000000")
        env = Envelope.objects.create(date=dt.date.today(), sabbath_week=1,
                                      receipt_no="Z1", member=m, contributor_name="A B",
                                      channel="BANK", total=Decimal("100"), recorded_by=u)
        self.assertIsNone(send_receipt_sms(env, cfg))

    def test_bank_scope_skips_cash(self):
        import datetime as dt
        from decimal import Decimal
        from core.models import SiteConfig
        from members.models import Member
        from envelopes.models import Envelope
        from core.services.sms import send_receipt_sms
        from django.contrib.auth.models import User
        u = User.objects.create_user("sms2", password="x")
        cfg = SiteConfig.get(); cfg.sms_enabled = True
        cfg.sms_receipt_scope = SiteConfig.SmsReceiptScope.BANK; cfg.save()
        m = Member.objects.create(name="C D", phone="254700000001")
        env = Envelope.objects.create(date=dt.date.today(), sabbath_week=1,
                                      receipt_no="Z2", member=m, contributor_name="C D",
                                      channel="CASH", total=Decimal("100"), recorded_by=u)
        self.assertIsNone(send_receipt_sms(env, cfg))  # cash skipped under BANK scope


class ExecutiveDashboardTests(TestCase):
    def setUp(self):
        import datetime as dt
        from decimal import Decimal
        from django.contrib.auth.models import User
        from departments.models import Department, Budget
        from giving.models import Transaction
        from cashbook.models import Expense
        self.u = User.objects.create_superuser("ex2", password="x")
        self.local = Department.objects.create(name="Youth", fund_type=Department.FundType.LOCAL)
        self.trust = Department.objects.create(name="Tithe", fund_type=Department.FundType.TRUST)
        Budget.objects.create(year=dt.date.today().year, department=self.local, amount=Decimal("1000"))
        # over-budget spend this year
        Expense.objects.create(date=dt.date(dt.date.today().year, 1, 5), department=self.local,
            description="big", amount=Decimal("5000"), category="OTHER",
            status=Expense.Status.PAID, recorded_by=self.u)
        # old unremitted trust receipt
        Transaction.objects.create(date=dt.date.today() - dt.timedelta(days=90), channel="ENVELOPE",
            direction="CREDIT", amount=Decimal("2000"), department=self.trust, allocation_status="AUTO")

    def test_cards_and_charts(self):
        from core.services import dashboard
        self.assertEqual(len(dashboard.cards()), 6)
        ch = dashboard.charts()
        self.assertIn("giving_trend", ch)
        self.assertEqual(len(ch["giving_trend"]["labels"]), 12)

    def test_anomalies_flag_overbudget_and_remittance(self):
        from core.services import health
        a = health.anomalies()
        titles = " ".join(x["title"] for x in a).lower()
        self.assertIn("over budget", titles)
        self.assertTrue(any("trust" in x["title"].lower() or "remit" in x["title"].lower()
                            for x in a))

    def test_page_renders(self):
        from django.urls import reverse
        self.client.login(username="ex2", password="x")
        r = self.client.get(reverse("executive"))
        self.assertEqual(r.status_code, 200)
        self.assertIn(b"chartData", r.content)


class YearEndCarryForwardTests(TestCase):
    def setUp(self):
        import datetime as dt
        from decimal import Decimal
        from django.contrib.auth.models import User
        from departments.models import Department
        from giving.models import Transaction
        from cashbook.models import Expense
        self.u = User.objects.create_superuser("ye", password="x")
        self.fund = Department.objects.create(name="General", fund_type=Department.FundType.LOCAL,
                                              opening_balance=Decimal("1000"))
        # 2025 activity: +4000 receipts, -500 expense  => closing 4500
        Transaction.objects.create(date=dt.date(2025, 3, 1), channel="CASH", direction="CREDIT",
            amount=Decimal("4000"), department=self.fund, allocation_status="MANUAL")
        Expense.objects.create(date=dt.date(2025, 6, 1), department=self.fund, description="x",
            amount=Decimal("500"), category="OTHER", status=Expense.Status.PAID, recorded_by=self.u)
        self.client.login(username="ye", password="x")

    def _val(self, rows, key):
        return {r["department"].id: r[key] for r in rows}[self.fund.id]

    def test_balance_carries_forward_automatically(self):
        import datetime as dt
        from decimal import Decimal
        from reports.services import balances
        c25 = self._val(balances.department_summary(dt.date(2025, 1, 1), dt.date(2025, 12, 31)), "closing")
        o26 = self._val(balances.department_summary(dt.date(2026, 1, 1), dt.date(2026, 12, 31)), "opening")
        self.assertEqual(c25, Decimal("4500"))
        self.assertEqual(o26, Decimal("4500"))   # next year's opening == this year's closing

    def test_close_year_snapshots_and_locks(self):
        from decimal import Decimal
        from django.urls import reverse
        from core.models import YearEndClose, FundCarryForward, PeriodLock
        self.client.post(reverse("controls"), {"action": "close_year", "close_year": "2025"})
        close = YearEndClose.objects.get(year=2025)
        self.assertEqual(close.total_carried, Decimal("4500"))
        self.assertEqual(FundCarryForward.objects.get(close=close, department=self.fund).closing_balance,
                         Decimal("4500"))
        self.assertEqual(PeriodLock.objects.filter(year=2025).count(), 12)   # year locked

    def test_double_close_blocked_and_reopen(self):
        from django.urls import reverse
        from core.models import YearEndClose, PeriodLock
        self.client.post(reverse("controls"), {"action": "close_year", "close_year": "2025"})
        self.client.post(reverse("controls"), {"action": "close_year", "close_year": "2025"})
        self.assertEqual(YearEndClose.objects.filter(year=2025).count(), 1)
        # reopen (superuser) clears the close and the locks
        self.client.post(reverse("controls"), {"action": "reopen_year", "close_year": "2025"})
        self.assertEqual(YearEndClose.objects.filter(year=2025).count(), 0)
        self.assertEqual(PeriodLock.objects.filter(year=2025).count(), 0)


class AssistantRevampTests(TestCase):
    def setUp(self):
        import datetime as dt
        from decimal import Decimal
        from django.contrib.auth.models import User
        from departments.models import Department
        from giving.models import Transaction
        from cashbook.models import Expense
        from ledger.services import posting
        self.u = User.objects.create_superuser("as", password="x")
        self.local = Department.objects.create(name="Church Budget", fund_type=Department.FundType.LOCAL,
                                               opening_balance=Decimal("1000"))
        self.trust = Department.objects.create(name="Tithe", fund_type=Department.FundType.TRUST)
        Transaction.objects.create(date=dt.date(2026, 5, 4), channel="CASH", direction="CREDIT",
            amount=Decimal("9000"), department=self.local, allocation_status="MANUAL")
        Transaction.objects.create(date=dt.date(2026, 5, 5), channel="BANK", direction="CREDIT",
            amount=Decimal("4000"), department=self.trust, allocation_status="AUTO")
        Expense.objects.create(date=dt.date(2026, 5, 6), department=self.local, description="chairs",
            amount=Decimal("70000"), category="CONSTRUCTION", expenditure_type="CAPITAL",
            status=Expense.Status.PAID, recorded_by=self.u)
        posting.rebuild()
        self.client.login(username="as", password="x")

    def _ask(self, q):
        from core.services import assistant
        return assistant.answer(q, self.u)

    def test_books_balanced_intent(self):
        d = self._ask("Are the books balanced?")
        self.assertIn("balance", d["text"].lower())

    def test_cash_intent_matches_pool(self):
        # cash = 1000 opening + 9000 + 4000 - 70000 capital = -56000
        from decimal import Decimal
        d = self._ask("How much cash do we have?")
        self.assertIn("56,000", d["text"])  # negative pool from the big capital outlay

    def test_capital_expenditure_intent(self):
        d = self._ask("capital expenditure this year")
        self.assertIn("70,000", d["text"])

    def test_fund_balance_uses_engine(self):
        d = self._ask("balance of Church Budget")
        self.assertTrue(any("Balance" in r[0] for r in d["rows"]))

    def test_fallback_offers_suggestions(self):
        d = self._ask("zxcv qwer nonsense")
        self.assertIn("suggestions", d)

    def test_page_and_endpoint(self):
        import json
        from django.urls import reverse
        self.assertEqual(self.client.get(reverse("assistant")).status_code, 200)
        r = self.client.post(reverse("assistant_ask"),
            data=json.dumps({"q": "trust funds to remit"}), content_type="application/json")
        self.assertEqual(r.status_code, 200)
        self.assertIn("text", r.json())


class SecurityAndControlsTests(TestCase):
    def test_credentials_encrypted_at_rest(self):
        from core.models import SiteConfig
        from django.db import connection
        cfg = SiteConfig.get()
        cfg.sms_api_key = "SUPER-SECRET-KEY"
        cfg.save()
        with connection.cursor() as c:
            c.execute("SELECT sms_api_key FROM core_siteconfig WHERE id=1")
            raw = c.fetchone()[0]
        self.assertTrue(raw.startswith("enc1:"))
        self.assertNotIn("SUPER-SECRET", raw)
        cfg.refresh_from_db()
        self.assertEqual(cfg.sms_api_key, "SUPER-SECRET-KEY")  # decrypts on load

    def test_siteconfig_is_audited(self):
        from core.models import SiteConfig
        cfg = SiteConfig.get()
        cfg.require_expense_approval = False
        cfg.save()
        self.assertTrue(cfg.history.exists())

    def test_dual_yearend_requires_second_treasurer(self):
        import datetime as dt
        from decimal import Decimal
        from django.contrib.auth.models import User, Group
        from django.urls import reverse
        from departments.models import Department
        from core.models import SiteConfig, YearEndClose
        g, _ = Group.objects.get_or_create(name="Treasurer")
        t1 = User.objects.create_user("yt1", password="x"); t1.groups.add(g)
        t2 = User.objects.create_user("yt2", password="x"); t2.groups.add(g)
        Department.objects.create(name="LCB", fund_type=Department.FundType.LOCAL,
                                  opening_balance=Decimal("1000"))
        cfg = SiteConfig.get(); cfg.require_dual_yearend = True; cfg.save()
        y = dt.date.today().year - 1
        # first treasurer initiates -> pending, not effective
        self.client.login(username="yt1", password="x")
        self.client.post(reverse("controls"), {"action": "close_year", "close_year": str(y)})
        close = YearEndClose.objects.get(year=y)
        self.assertFalse(close.is_effective)
        # same treasurer cannot confirm
        self.client.post(reverse("controls"), {"action": "confirm_close", "close_year": str(y)})
        close.refresh_from_db(); self.assertFalse(close.is_effective)
        # a second treasurer confirms -> effective
        self.client.login(username="yt2", password="x")
        self.client.post(reverse("controls"), {"action": "confirm_close", "close_year": str(y)})
        close.refresh_from_db()
        self.assertTrue(close.is_effective)
        self.assertEqual(close.confirmed_by, t2)


class OutboundAndLlmTests(TestCase):
    def test_post_json_returns_http_error_status(self):
        # 401 from a bad key should be returned, not raised (so callers see the reason)
        from core.services.net import post_json
        try:
            status, body = post_json("https://api.anthropic.com/v1/messages", {},
                headers={"x-api-key": "bad", "anthropic-version": "2023-06-01"}, timeout=15)
        except Exception:
            self.skipTest("network unavailable in this environment")
        self.assertGreaterEqual(status, 400)

    def test_test_llm_reports_when_disabled(self):
        from core.models import SiteConfig
        from core.services.assistant import test_llm
        cfg = SiteConfig.get(); cfg.llm_enabled = False; cfg.save()
        ok, detail = test_llm(cfg)
        self.assertFalse(ok)
        self.assertIn("off", detail.lower())

    def test_test_llm_reports_missing_key(self):
        from core.models import SiteConfig
        from core.services.assistant import test_llm
        cfg = SiteConfig.get(); cfg.llm_enabled = True; cfg.llm_api_key = ""; cfg.save()
        ok, detail = test_llm(cfg)
        self.assertFalse(ok)
        self.assertIn("key", detail.lower())

    def test_groq_default_model_used(self):
        # blank model + Groq provider -> a Groq model name, never gpt-4o-mini
        from core.services.assistant import _DEFAULT_MODEL
        self.assertIn("llama", _DEFAULT_MODEL["GROQ"])
        self.assertNotEqual(_DEFAULT_MODEL["GROQ"], _DEFAULT_MODEL["OPENAI"])


class SettingsSaveTests(TestCase):
    def setUp(self):
        from django.contrib.auth.models import User, Group
        self.u = User.objects.create_user("st", password="x")
        self.u.groups.add(Group.objects.get_or_create(name="Treasurer")[0])
        self.client.login(username="st", password="x")

    def _post_data(self, **overrides):
        from core.models import SiteConfig
        from core.forms import SiteConfigForm
        form = SiteConfigForm(instance=SiteConfig.get())
        data = {}
        for name, field in form.fields.items():
            val = form[name].value()
            if getattr(field.widget, "input_type", None) == "checkbox":
                if val:
                    data[name] = "on"
            else:
                data[name] = "" if val is None else val
        data.update(overrides)
        return data

    def test_settings_form_has_no_nested_forms(self):
        from django.urls import reverse
        h = self.client.get(reverse("settings")).content.decode()
        start = h.find('<form method="post">')
        end = h.find("</form>", start)
        # no second <form> opens between the main form's open and close tags
        self.assertEqual(h.find("<form", start + 6, end), -1)

    def test_save_button_persists_settings(self):
        from django.urls import reverse
        from core.models import SiteConfig
        self.client.post(reverse("settings"), self._post_data(church_name="Berea SDA"))
        self.assertEqual(SiteConfig.get().church_name, "Berea SDA")

    def test_pledge_modes_persist_and_are_not_duplicated_on_the_page(self):
        """Regression: f.name == fields used to also render under Other, so
        save posted the unchanged duplicate and the user's choice bounced back."""
        from django.urls import reverse
        from core.models import SiteConfig
        html = self.client.get(reverse("settings")).content.decode()
        self.assertEqual(
            html.count('name="pledge_match_mode"'), 1,
            "pledge_match_mode rendered more than once — save would revert it")
        self.assertEqual(
            html.count('name="pledge_public_submit_mode"'), 1,
            "pledge_public_submit_mode rendered more than once — save would "
            "revert it")
        self.client.post(reverse("settings"), self._post_data(
            pledge_match_mode="AUTO",
            pledge_public_submit_mode="ACTIVE"))
        cfg = SiteConfig.get()
        self.assertEqual(cfg.pledge_match_mode, "AUTO")
        self.assertEqual(cfg.pledge_public_submit_mode, "ACTIVE")

    def test_save_and_test_persists_before_testing(self):
        from django.urls import reverse
        from core.models import SiteConfig
        self.client.post(reverse("settings"), self._post_data(
            llm_enabled="on", llm_provider="GROQ", llm_model="llama-3.3-70b-versatile",
            llm_api_key="gsk_fake", test_llm="1"))
        cfg = SiteConfig.get()
        self.assertTrue(cfg.llm_enabled)
        self.assertEqual(cfg.llm_model, "llama-3.3-70b-versatile")


class TelegramEnhancementsTests(TestCase):
    """Per-user PIN identifies the user; the expense flow matches a member,
    asks the payment method, and records a bank charge."""

    def setUp(self):
        from django.contrib.auth.models import User
        from core.models import SiteConfig, TelegramProfile
        from departments.models import Department
        self.cfg = SiteConfig.get()
        self.cfg.telegram_enabled = True
        self.cfg.telegram_pin = "1234"
        self.cfg.save()
        self.user = User.objects.create_user("tguser", password="x")
        TelegramProfile.objects.create(user=self.user, pin="4827")
        Department.objects.get_or_create(name="LCB Departments",
            defaults=dict(fund_type="LOCAL", category="OFFERING"))

    def _reply(self, text, cid=55001):
        from core.services.telegram_bot import handle_update
        out = handle_update({"message": {"chat": {"id": cid}, "text": text}})
        return out[0]["text"] if out else ""

    def test_personal_pin_identifies_user(self):
        from core.models import TelegramSession
        r = self._reply("4827")
        self.assertIn("Signed in", r)
        self.assertEqual(TelegramSession.objects.get(chat_id="55001").user_id,
                         self.user.id)

    def test_expense_matches_member_and_records_charge(self):
        from members.models import Member
        from cashbook.models import Expense
        from departments.models import Department
        Member.objects.create(name="Grace Wanjiru", active=True)
        d = Department.objects.create(name="Evangelism", fund_type="LOCAL",
                                      category="MINISTRY")
        self._reply("4827")
        self._reply("/expense")
        self._reply("2000")
        # choose the Evangelism fund by name
        self._reply("Evangelism")
        self._reply("Transport refund")   # description
        r = self._reply("Grace Wanjiru")  # claimant
        self.assertIn("Matched member", r)
        r = self._reply("2")              # M-Pesa
        self.assertIn("charge", r.lower())
        self._reply("50")                 # charge amount
        before = Expense.objects.count()
        self._reply("yes")
        self.assertEqual(Expense.objects.count() - before, 2)  # expense + charge
        exp = Expense.objects.filter(claimant__iexact="Grace Wanjiru").latest("id")
        self.assertEqual(exp.recorded_by_id, self.user.id)
        self.assertEqual(exp.method, "MPESA")
        self.assertTrue(Expense.objects.filter(category="BANK_CHARGE").exists())


class DuplicateDetectionTests(TestCase):
    """Offerings flag same-reference duplicates; expenses exclude bank charges."""

    def test_offerings_shared_reference_not_flagged(self):
        # distinct givers sharing a paybill reference (each with a unique bank
        # receipt) are NOT duplicates under the corrected logic.
        from giving.models import Transaction
        from core.views import _duplicate_offerings
        import datetime as dt
        from decimal import Decimal
        for i in range(2):
            Transaction.objects.create(
                date=dt.date(2026, 6, 6), channel="BANK", direction="CREDIT",
                amount=Decimal("500"), allocation_status="MANUAL", confirmed=True,
                payer_name=f"Person {i}", reference="SAMEREF123",
                core_ref=f"DR{i}")
        out = _duplicate_offerings()
        ref_dups = [o for o in out if o["reference"] == "SAMEREF123"]
        self.assertFalse(ref_dups)

    def test_expenses_exclude_bank_charges(self):
        from django.contrib.auth.models import User
        from cashbook.models import Expense
        from departments.models import Department
        from core.views import _duplicate_expenses
        import datetime as dt
        from decimal import Decimal
        u = User.objects.create_user("e", password="x")
        d = Department.objects.create(name="LCB", fund_type="LOCAL", category="OFFERING")
        # two identical bank charges (legitimately repeat) must NOT be flagged
        for i in range(2):
            Expense.objects.create(
                date=dt.date(2026, 6, 6), department=d, description="M-Pesa charge",
                amount=Decimal("30"), category="BANK_CHARGE", status="PAID",
                claimant="System", recorded_by=u)
        flagged = [g for g in _duplicate_expenses()
                   if g["description"] == "M-Pesa charge"]
        self.assertEqual(flagged, [])


class ProductionReadinessTests(TestCase):
    """Health check, update endpoints, and the rich Excel export."""

    def setUp(self):
        from django.contrib.auth.models import User, Group
        self.u = User.objects.create_user("pr", password="x")
        g, _ = Group.objects.get_or_create(name="Treasurer")
        self.u.groups.add(g)

    def test_healthz_no_auth(self):
        from django.test import Client
        r = Client().get("/healthz/")
        self.assertEqual(r.status_code, 200)
        body = r.json()
        self.assertEqual(body["status"], "ok")
        self.assertTrue(body["database"])
        self.assertIn("version", body)

    def test_update_page_and_status(self):
        from django.test import Client
        c = Client(); c.force_login(self.u)
        self.assertEqual(c.get("/update/").status_code, 200)
        s = c.get("/update/status/")
        self.assertEqual(s.status_code, 200)
        self.assertIn("running", s.json())

    def test_update_status_requires_treasurer(self):
        from django.test import Client
        from django.contrib.auth.models import User
        viewer = User.objects.create_user("v", password="x")
        c = Client(); c.force_login(viewer)
        # a non-treasurer should not reach the update controls
        self.assertNotEqual(c.get("/update/").status_code, 200)

    def test_full_excel_export_has_accounting_sheets(self):
        import io
        import openpyxl
        from core.services.backup import full_excel_export_response
        r = full_excel_export_response()
        self.assertEqual(r.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        for sheet in ("Summary", "Fund Balances", "Trust Funds",
                      "Income by Channel", "Cash Book", "Members",
                      "Transactions", "Expenses"):
            self.assertIn(sheet, wb.sheetnames)
        # Fund Balances must carry a TOTAL row
        fb = wb["Fund Balances"]
        totals = [row[0] for row in fb.iter_rows(values_only=True)]
        self.assertIn("TOTAL", totals)


class GithubTokenAuthTests(TestCase):
    """The release checker sends an Authorization header when GITHUB_TOKEN is
    set (needed for private repos), and omits it when not."""

    def _capture_headers(self):
        import json
        from unittest.mock import patch
        from core.services import updates
        captured = {}

        class FakeResp:
            def __enter__(self): return self
            def __exit__(self, *a): pass
            def read(self):
                return json.dumps({"tag_name": "v1.0.3", "html_url": "u",
                                   "body": "n"}).encode()

        def fake_urlopen(req, timeout=4):
            captured["headers"] = dict(req.header_items())
            return FakeResp()

        updates._release_cache["value"] = None   # bypass the time cache in tests
        with patch.object(updates.urllib.request, "urlopen", fake_urlopen):
            rel = updates.latest_release(force=True)
        return captured.get("headers", {}), rel

    @override_settings(GITHUB_REPO="o/r", GITHUB_TOKEN="ghp_secret")
    def test_token_adds_auth_header(self):
        headers, rel = self._capture_headers()
        self.assertEqual(rel["tag"], "v1.0.3")
        self.assertEqual(headers.get("Authorization"), "Bearer ghp_secret")

    @override_settings(GITHUB_REPO="o/r", GITHUB_TOKEN="")
    def test_no_token_no_auth_header(self):
        headers, rel = self._capture_headers()
        self.assertNotIn("Authorization", headers)


class DashboardAttentionTests(TestCase):
    """The dashboard 'needs attention' list surfaces actionable items with
    correct counts and resolvable links, and stays empty when all is clear."""

    def setUp(self):
        from django.contrib.auth.models import User, Group
        self.u = User.objects.create_user("att", password="x")
        g, _ = Group.objects.get_or_create(name="Treasurer")
        self.u.groups.add(g)

    def _ctx(self):
        from django.test import RequestFactory
        from core.views import DashboardView
        req = RequestFactory().get("/")
        req.user = self.u
        v = DashboardView(); v.request = req
        return v.get_context_data()

    def test_empty_when_nothing_pending(self):
        ctx = self._ctx()
        # a clean demo-less DB should have no attention items
        self.assertEqual(ctx["attention"], [])

    def test_pending_expense_surfaces(self):
        from cashbook.models import Expense
        from departments.models import Department
        import datetime as dt
        from decimal import Decimal
        d = Department.objects.create(name="X", fund_type="LOCAL", category="OFFERING")
        Expense.objects.create(date=dt.date(2026, 6, 1), department=d,
            description="t", amount=Decimal("10"), status="PENDING",
            recorded_by=self.u)
        labels = [a["label"] for a in self._ctx()["attention"]]
        self.assertTrue(any("awaiting approval" in l for l in labels))

    def test_pledge_draft_surfaces(self):
        from pledges.models import PledgeCampaign, Pledge
        from members.models import Member
        from decimal import Decimal
        m = Member.objects.create(name="D Giver")
        camp = PledgeCampaign.objects.create(name="C", status="ACTIVE")
        Pledge.objects.create(campaign=camp, member=m, amount=Decimal("100"),
                              status="DRAFT")
        ctx = self._ctx()
        self.assertEqual(ctx["pledge_draft_count"], 1)
        self.assertTrue(any("pledges awaiting approval" in a["label"]
                            for a in ctx["attention"]))

    def test_all_attention_links_resolve(self):
        # every link we might emit must be a real, resolvable path
        from django.urls import resolve
        from cashbook.models import Expense
        from departments.models import Department
        from pledges.models import PledgeCampaign, Pledge
        from members.models import Member
        import datetime as dt
        from decimal import Decimal
        d = Department.objects.create(name="Y", fund_type="LOCAL", category="OFFERING")
        Expense.objects.create(date=dt.date(2026, 6, 1), department=d, description="t",
            amount=Decimal("10"), status="PENDING", recorded_by=self.u)
        m = Member.objects.create(name="E Giver")
        camp = PledgeCampaign.objects.create(name="C2", status="ACTIVE")
        Pledge.objects.create(campaign=camp, member=m, amount=Decimal("100"),
                              status="DRAFT")
        for a in self._ctx()["attention"]:
            path = a["url"].split("?")[0]
            self.assertTrue(resolve(path))  # raises if unresolvable


class MultiYearTrendThroughMonthTests(TestCase):
    """Item 4: the multi-year trend compares January–current-month of each year,
    not full prior years vs a part-year-in-progress."""

    def setUp(self):
        import datetime as dt
        from django.contrib.auth.models import User
        from core.models import SiteConfig, HistoricalMonth
        u = User.objects.create_superuser("trend", password="x")
        cfg = SiteConfig.get(); cfg.require_2fa_for_treasurers = False; cfg.save()
        self.client.force_login(u)
        self.cur_month = dt.date.today().month
        self.prev_year = dt.date.today().year - 1
        # prior year: 1,000 in every month -> through current month = cur_month*1000
        for m in range(1, 13):
            HistoricalMonth.objects.create(year=self.prev_year, month=m,
                collection=1000, trust_fund=100, expenditure=200)

    def test_prior_year_truncated_to_current_month(self):
        r = self.client.get("/")
        self.assertEqual(r.status_code, 200)
        td = {t["year"]: t for t in __import__("json").loads(r.context["trend_json"])}
        self.assertIn(self.prev_year, td)
        # only months <= current month are counted (not the full 12,000)
        self.assertEqual(td[self.prev_year]["collection"], self.cur_month * 1000)
        self.assertEqual(td[self.prev_year]["expenditure"], self.cur_month * 200)

    def test_current_year_is_year_to_date(self):
        import datetime as dt
        from decimal import Decimal
        from giving.models import Transaction
        from departments.models import Department
        d = Department.objects.create(name="TY Fund", fund_type="LOCAL",
                                      category="OFFERING", selectable=True)
        today = dt.date.today()
        Transaction.objects.create(date=dt.date(today.year, 1, 15), channel="CASH",
            direction="CREDIT", amount=Decimal("5000"), department=d,
            confirmed=True, allocation_status="MANUAL")
        r = self.client.get("/")
        td = {t["year"]: t for t in __import__("json").loads(r.context["trend_json"])}
        self.assertIn(today.year, td)
        self.assertEqual(td[today.year]["collection"], 5000.0)
        # label reflects the cut-off month
        self.assertEqual(r.context["trend_through"],
            ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"][self.cur_month-1])
