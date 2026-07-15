"""Round 9, item 5 — every benevolent table can be downloaded as a spreadsheet.

The tables a treasurer works from — the register, the memberships, the
contributions, the cases — had no export at all, while the rest of the app has
had xlsx/csv downloads for a long time. These add them, reusing the one styled
workbook helper (reports.exports) so the format matches everything else, and
respecting whatever filters the page already has applied.
"""
import datetime as dt
import io
from decimal import Decimal

from django.contrib.auth.models import Group, User
from django.test import TestCase
from django.urls import reverse

from core.roles import TREASURER
from departments.models import Department
from members.models import Member

from benevolent.models import (BenevolentEventType, BenevolentScheme,
                               SchemeMembership, SchemePolicy)
from benevolent.services import contributions as contrib_svc
from benevolent.services import registry as reg_svc
from benevolent.services import schemes as scheme_svc

TODAY = dt.date.today()


class ExportFixture(TestCase):
    def setUp(self):
        self.treasurer = User.objects.create_user("t_ex", password="x")
        self.treasurer.groups.add(Group.objects.get_or_create(name=TREASURER)[0])
        self.client.force_login(self.treasurer)

        fund = Department.objects.create(name="EX Fund", slug="ex-fund",
                                         fund_type=Department.FundType.LOCAL)
        self.scheme = BenevolentScheme.objects.create(
            name="Benevolent", code="EXB", fund=fund, created_by=self.treasurer)
        BenevolentEventType.objects.create(
            scheme=self.scheme, name="Bereavement", code="BER",
            triggers_on_death=True)
        policy = SchemePolicy.objects.create(
            scheme=self.scheme, effective_from=TODAY - dt.timedelta(days=900),
            membership_required=True, waiting_period_days=0,
            contribution_mode=SchemePolicy.ContributionMode.PER_CASE_LEVY,
            levy_amount=Decimal("500"), registration_required=True,
            registration_fee=Decimal("500"),
            benefit_mode=SchemePolicy.BenefitMode.FIXED,
            benefit_amount=Decimal("50000"),
            arrears_treatment=SchemePolicy.ArrearsTreatment.IGNORE,
            created_by=self.treasurer)
        scheme_svc.publish_policy(policy, user=self.treasurer)
        scheme_svc.activate_scheme(self.scheme, user=self.treasurer)

        self.m1 = reg_svc.register(
            self.scheme, Member.objects.create(name="AARON EXPORT"),
            joined_on=TODAY - dt.timedelta(days=300), user=self.treasurer)
        self.m2 = reg_svc.register(
            self.scheme, Member.objects.create(name="BETH EXPORT"),
            joined_on=TODAY - dt.timedelta(days=200), user=self.treasurer)
        contrib_svc.record_contribution(
            self.scheme, date=TODAY, amount=Decimal("500"),
            user=self.treasurer, membership=self.m1, channel="CASH")

    def _open_xlsx(self, content):
        import openpyxl
        return openpyxl.load_workbook(io.BytesIO(content)).active


class MembershipExportTests(ExportFixture):
    def test_xlsx_download(self):
        r = self.client.get(reverse("benevolent_membership_list"),
                            {"export": "xlsx"})
        self.assertEqual(r.status_code, 200)
        self.assertIn("spreadsheetml", r["Content-Type"])
        self.assertIn("benevolent-members", r["Content-Disposition"])
        ws = self._open_xlsx(r.content)
        text = " ".join(str(c.value) for row in ws.iter_rows() for c in row
                        if c.value)
        self.assertIn("AARON EXPORT", text)
        self.assertIn("BETH EXPORT", text)

    def test_csv_download(self):
        r = self.client.get(reverse("benevolent_membership_list"),
                            {"export": "csv"})
        self.assertEqual(r.status_code, 200)
        self.assertIn("text/csv", r["Content-Type"])
        body = r.content.decode()
        self.assertIn("AARON EXPORT", body)
        self.assertIn("Number", body)     # header row

    def test_export_respects_the_status_filter(self):
        # lapse one member so a status filter has something to exclude
        self.m2.status = SchemeMembership.Status.SUSPENDED
        self.m2.save(update_fields=["status"])
        r = self.client.get(reverse("benevolent_membership_list"),
                            {"export": "csv", "status": "ACTIVE"})
        body = r.content.decode()
        self.assertIn("AARON EXPORT", body)       # active, included
        self.assertNotIn("BETH EXPORT", body)     # lapsed, excluded


class ContributionExportTests(ExportFixture):
    def test_xlsx_has_the_contribution_amount(self):
        r = self.client.get(reverse("benevolent_contribution_list"),
                            {"export": "xlsx"})
        self.assertEqual(r.status_code, 200)
        ws = self._open_xlsx(r.content)
        values = [c.value for row in ws.iter_rows() for c in row]
        self.assertIn(500.0, values)     # the contribution's amount
        self.assertIn("AARON EXPORT", " ".join(str(v) for v in values if v))


class CaseExportTests(ExportFixture):
    def test_case_export_downloads(self):
        # a death opens a case we can export
        reg_svc.record_death(self.m1, died_on=TODAY, user=self.treasurer)
        r = self.client.get(reverse("benevolent_case_list"), {"export": "csv"})
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertIn("Number", body)
        self.assertIn("AARON EXPORT", body)


class RegistryExportTests(ExportFixture):
    def test_registry_export_downloads(self):
        r = self.client.get(reverse("benevolent_registry"), {"export": "xlsx"})
        self.assertEqual(r.status_code, 200)
        self.assertIn("spreadsheetml", r["Content-Type"])
        ws = self._open_xlsx(r.content)
        text = " ".join(str(c.value) for row in ws.iter_rows() for c in row
                        if c.value)
        self.assertIn("AARON EXPORT", text)


class ExportButtonTests(ExportFixture):
    def test_pages_show_export_links_that_carry_filters(self):
        r = self.client.get(reverse("benevolent_membership_list"),
                            {"status": "ACTIVE"})
        body = r.content.decode()
        self.assertIn("export=xlsx", body)
        self.assertIn("export=csv", body)
        # the active filter must survive into the export link
        self.assertIn("status=ACTIVE", body)
