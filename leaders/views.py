import datetime as dt
from decimal import Decimal

from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import TemplateView

from core.permissions import RoleRequiredMixin  # noqa (kept for symmetry)
from core.utils import parse_period
from reports.services import balances
from departments.models import Department, DevelopmentGroup
from members.models import mask_phone
from core.rights import display_phone, display_giver
from .permissions import LeaderRequiredMixin, allowed_departments, assert_department_allowed


def _parse_date(raw):
    if not raw:
        return None
    try:
        return dt.date.fromisoformat(raw.strip())
    except (ValueError, AttributeError):
        return None


def _scoped_rows(user, start, end):
    """Per-department balance rows (un-consolidated) limited to the leader's
    departments. Filtering happens here, server-side, on the id set."""
    allowed_ids = set(allowed_departments(user).values_list("id", flat=True))
    rows = balances.department_summary(start, end, consolidated=False)
    return [r for r in rows if r["department"].id in allowed_ids]


class LeaderDashboardView(LeaderRequiredMixin, TemplateView):
    template_name = "leaders/dashboard.html"

    def get(self, request, *args, **kwargs):
        # a leader who leads exactly one department goes straight to that
        # department's page — no need for an intermediate overview
        if not request.GET.get("stay"):
            led = list(allowed_departments(request.user))
            led_ids = {d.id for d in led}
            roots = [d for d in led if d.parent_id not in led_ids]
            if len(roots) == 1:
                return redirect("leader_department_detail", pk=roots[0].pk)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        start, end = parse_period(self.request)
        ctx["start"], ctx["end"] = start, end
        rows = _scoped_rows(self.request.user, start, end)
        # A row heads the list (is a "top") when its parent is NOT also in the
        # leader's set — i.e. it's a root of the leader's allowed tree. So a leader
        # assigned the parent fund sees it with its sub-accounts nested; a leader
        # assigned a single subgroup directly still sees that subgroup at the top
        # (it isn't hidden under a parent they don't lead).
        allowed_ids = set(allowed_departments(self.request.user)
                          .values_list("id", flat=True))
        by_parent = {}
        tops = []
        for r in rows:
            d = r["department"]
            if d.parent_id and d.parent_id in allowed_ids:
                by_parent.setdefault(d.parent_id, []).append(r)
            else:
                tops.append(r)
        for r in tops:
            r["subrows"] = by_parent.get(r["department"].id, [])
        ctx["rows"] = tops
        ctx["total_receipts"] = sum((r["receipts"] for r in rows), Decimal(0))
        ctx["total_expenses"] = sum((r["expenses_operating"] for r in rows), Decimal(0))
        ctx["dept_count"] = len({r["department"].id for r in rows})
        return ctx


class LeaderDepartmentDetailView(LeaderRequiredMixin, TemplateView):
    template_name = "leaders/department_detail.html"

    def get(self, request, *args, **kwargs):
        dept = assert_department_allowed(request.user, kwargs["pk"])
        if not dept:
            # not their department — refuse, send back to their dashboard
            return redirect("leader_dashboard")
        self.dept = dept
        export = request.GET.get("export")
        if export in ("groups_csv", "groups_xlsx"):
            return self._export_groups(request, export)
        return super().get(request, *args, **kwargs)

    def _dev_group_rows(self, start, end):
        from decimal import Decimal
        from django.db.models import Sum
        from giving.models import Transaction
        base = Transaction.objects.filter(
            direction=Transaction.Direction.CREDIT, confirmed=True,
            is_reversal=False, is_reversed=False, dev_group__isnull=False)
        # collected within the period (receipts) and before it (opening)
        collected_map = {r["dev_group"]: (r["s"] or Decimal(0)) for r in
            base.filter(date__gte=start, date__lte=end)
            .values("dev_group").annotate(s=Sum("amount"))}
        opening_map = {r["dev_group"]: (r["s"] or Decimal(0)) for r in
            base.filter(date__lt=start)
            .values("dev_group").annotate(s=Sum("amount"))}
        rows = []
        for g in DevelopmentGroup.objects.filter(active=True):
            collected = collected_map.get(g.id, Decimal(0))
            opening = opening_map.get(g.id, Decimal(0))
            closing = opening + collected
            tgt = g.target or Decimal(0)
            pct = int(min(closing / tgt * 100, 100)) if tgt else None
            rows.append({"group": g, "opening": opening, "collected": collected,
                         "closing": closing, "target": tgt, "pct": pct})
        rows.sort(key=lambda r: r["closing"], reverse=True)
        return rows

    def _export_groups(self, request, export):
        from reports.exports import csv_response, xlsx_response
        from core.models import SiteConfig
        start, end = parse_period(request)
        rows = self._dev_group_rows(start, end)
        header = ["Group", "Opening", "Receipts", "Closing", "Target", "% of target"]
        data = [[str(r["group"]), float(r["opening"]), float(r["collected"]),
                 float(r["closing"]), float(r["target"]),
                 (r["pct"] if r["pct"] is not None else "")] for r in rows]
        title = (f"{self.dept.name} — development groups, "
                 f"{start:%d %b %Y} to {end:%d %b %Y}")
        fn = f"dev_groups_{self.dept.slug or self.dept.id}_{start:%Y%m%d}_{end:%Y%m%d}"
        if export == "groups_xlsx":
            return xlsx_response(f"{fn}.xlsx", header, data, title=title,
                                 church=SiteConfig.get().church_name)
        return csv_response(f"{fn}.csv", header, data)

    def get_context_data(self, **kwargs):
        import json
        from decimal import Decimal
        from django.db.models import Sum, Count
        from django.db.models.functions import ExtractMonth
        from giving.models import Transaction
        from cashbook.models import Expense
        ctx = super().get_context_data(**kwargs)
        start, end = parse_period(self.request)
        ctx["start"], ctx["end"] = start, end
        dept = self.dept
        ctx["dept"] = dept

        # this department's row + its sub-accounts (the leader may see those)
        allowed_ids = set(allowed_departments(self.request.user).values_list("id", flat=True))
        rows = balances.department_summary(start, end, consolidated=False)
        ctx["row"] = next((r for r in rows if r["department"].id == dept.id), None)
        ctx["subrows"] = [r for r in rows
                          if r["department"].parent_id == dept.id
                          and r["department"].id in allowed_ids]
        # the whole area this leader manages here: the department + visible subs
        dept_ids = {dept.id} | {r["department"].id for r in ctx["subrows"]}

        MN = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        credit_q = dict(department_id__in=dept_ids,
                        direction=Transaction.Direction.CREDIT, confirmed=True,
                        is_reversal=False, is_reversed=False, excluded_from_income=False)

        # --- headline KPIs ---------------------------------------------------
        receipts = (Transaction.objects.filter(date__gte=start, date__lte=end, **credit_q)
                    .aggregate(s=Sum("amount"))["s"] or Decimal(0))
        spend = (Expense.objects.filter(department_id__in=dept_ids,
                    status__in=[Expense.Status.APPROVED, Expense.Status.PAID],
                    date__gte=start, date__lte=end)
                 .exclude(category=Expense.Category.REMITTANCE)
                 .aggregate(s=Sum("amount"))["s"] or Decimal(0))
        gift_count = (Transaction.objects.filter(date__gte=start, date__lte=end, **credit_q)
                      .count())
        ctx["kpi"] = {
            "closing": ctx["row"]["closing"] if ctx["row"] else Decimal(0),
            "receipts": receipts, "spend": spend, "gifts": gift_count,
            "net": receipts - spend,
        }

        # --- monthly trend (receipts vs spend) -------------------------------
        mrec = {r["m"]: float(r["t"] or 0) for r in
                Transaction.objects.filter(date__gte=start, date__lte=end, **credit_q)
                .annotate(m=ExtractMonth("date")).values("m").annotate(t=Sum("amount"))}
        mexp = {r["m"]: float(r["t"] or 0) for r in
                Expense.objects.filter(department_id__in=dept_ids,
                    status__in=[Expense.Status.APPROVED, Expense.Status.PAID],
                    date__gte=start, date__lte=end)
                .exclude(category=Expense.Category.REMITTANCE)
                .annotate(m=ExtractMonth("date")).values("m").annotate(t=Sum("amount"))}
        active_m = sorted(set(mrec) | set(mexp))
        ctx["monthly_json"] = json.dumps({
            "labels": [MN[m - 1] for m in active_m],
            "receipts": [mrec.get(m, 0) for m in active_m],
            "expenses": [mexp.get(m, 0) for m in active_m]})
        ctx["has_monthly"] = len(active_m) >= 2

        # --- income by channel ----------------------------------------------
        ch_labels = {"BANK": "Bank / M-Pesa", "CASH": "Cash", "ENVELOPE": "Envelopes"}
        ch = (Transaction.objects.filter(date__gte=start, date__lte=end, **credit_q)
              .values("channel").annotate(t=Sum("amount")).order_by("-t"))
        ctx["channel_json"] = json.dumps([
            {"label": ch_labels.get(c["channel"], c["channel"] or "Other"),
             "value": float(c["t"] or 0)} for c in ch if c["t"]])
        ctx["has_channel"] = bool(ch)

        # --- top contributors ------------------------------------------------
        top = (Transaction.objects.filter(date__gte=start, date__lte=end, **credit_q)
               .values("member__name", "payer_name")
               .annotate(t=Sum("amount"), n=Count("id")).order_by("-t")[:6])
        ctx["top_givers"] = [{
            "who": display_giver(self.request.user, r["member__name"] or r["payer_name"]) or "—",
            "total": r["t"] or Decimal(0), "n": r["n"]} for r in top]

        # --- budget vs actual ------------------------------------------------
        if dept.annual_budget:
            pct = int(min(spend / dept.annual_budget * 100, 100)) if dept.annual_budget else 0
            ctx["budget"] = {"budget": dept.annual_budget, "spent": spend, "pct": pct,
                             "over": spend > dept.annual_budget}
        else:
            ctx["budget"] = None

        # --- recent collections / expenses previews --------------------------
        txns = (Transaction.objects.filter(department=dept,
                    direction=Transaction.Direction.CREDIT, confirmed=True,
                    is_reversal=False, is_reversed=False,
                    date__gte=start, date__lte=end)
                .select_related("member").order_by("-date")[:8])
        ctx["collections"] = [{
            "date": t.date, "amount": t.amount, "channel": t.get_channel_display(),
            "who": display_giver(self.request.user, t.member.name if t.member_id else t.payer_name) or "—",
            "phone": display_phone(self.request.user, t.member.phone if t.member_id else t.payer_phone),
            "reference": t.reference,
        } for t in txns]
        ctx["expenses"] = (Expense.objects.filter(department=dept,
                    date__gte=start, date__lte=end).order_by("-date")[:8])

        # which of these funds may actually carry expenses — used to hide the
        # expenses preview and columns for collection-only subgroups.
        from departments.models import expense_departments
        elig_ids = {d.id for d in expense_departments()}
        ctx["expenses_eligible"] = dept.id in elig_ids
        for r in ctx["subrows"]:
            r["expenses_eligible"] = r["department"].id in elig_ids
        ctx["any_sub_expenses"] = any(r.get("expenses_eligible") for r in ctx["subrows"])

        # --- development groups (for a development leader) -------------------
        if dept.category == Department.Category.DEVELOPMENT:
            ctx["dev_groups"] = self._dev_group_rows(start, end)
            ctx["dev_collected"] = sum((r["collected"] for r in ctx["dev_groups"]),
                                       Decimal(0))
        else:
            ctx["dev_groups"] = None

        # --- pledges summary -------------------------------------------------
        try:
            from pledges.models import Pledge
            pledges = (Pledge.objects.filter(campaign__target_department=dept)
                       .exclude(status=Pledge.Status.CANCELLED)
                       .select_related("member", "campaign"))
            p_total = sum((p.amount for p in pledges), Decimal(0))
            p_paid = sum((p.paid for p in pledges), Decimal(0))
            ctx["pledge_summary"] = {
                "count": pledges.count(), "pledged": p_total, "paid": p_paid,
                "outstanding": p_total - p_paid,
                "pct": int(min(p_paid / p_total * 100, 100)) if p_total else 0,
            } if pledges.exists() else None
        except Exception:
            ctx["pledge_summary"] = None

        # staff advances on this fund (and its sub-accounts the leader sees)
        try:
            from cashbook.models import StaffAdvance
            advs = list(StaffAdvance.objects.filter(department_id__in=dept_ids)
                        .order_by("-date_issued")[:50])
            open_advs = [a for a in advs if a.status != StaffAdvance.Status.CLOSED]
            ctx["advance_summary"] = {
                "open_count": len(open_advs),
                "outstanding": sum((a.balance for a in open_advs if a.balance > 0),
                                   Decimal(0)),
                "recent": advs[:5],
            } if advs else None
        except Exception:  # noqa: BLE001
            ctx["advance_summary"] = None

        return ctx


# ===========================================================================
# Detailed, downloadable leader pages (item 2)
# ===========================================================================
def _leads_a_development_dept(user):
    return allowed_departments(user).filter(
        category=Department.Category.DEVELOPMENT).exists()


def _collection_rows(dept, start, end, user):
    """Full collections (credits) for a department in a period, newest first."""
    from giving.models import Transaction
    txns = (Transaction.objects.filter(
                department=dept, direction=Transaction.Direction.CREDIT,
                confirmed=True, is_reversal=False, is_reversed=False,
                date__gte=start, date__lte=end)
            .select_related("member", "dev_group").order_by("-date", "-id"))
    out = []
    for t in txns:
        out.append({
            "date": t.date, "amount": t.amount, "channel": t.get_channel_display(),
            "who": display_giver(user, t.member.name if t.member_id else t.payer_name) or "—",
            "phone": display_phone(user, t.member.phone if t.member_id else t.payer_phone),
            "reference": t.reference or "",
            "group": t.dev_group.label if t.dev_group_id else "",
        })
    return out


class LeaderCollectionsView(LeaderRequiredMixin, TemplateView):
    """Full, downloadable collections list for one of the leader's departments."""
    template_name = "leaders/collections.html"

    def get(self, request, *args, **kwargs):
        self.dept = assert_department_allowed(request.user, kwargs["pk"])
        if not self.dept:
            return redirect("leader_dashboard")
        start, end = parse_period(request)
        export = request.GET.get("export")
        if export in ("csv", "xlsx"):
            from reports.exports import csv_response, xlsx_response
            from core.models import SiteConfig
            rows = _collection_rows(self.dept, start, end, self.request.user)
            header = ["Date", "Contributor", "Phone", "Reference", "Channel", "Group", "Amount"]
            data = [[r["date"].isoformat(), r["who"], r["phone"], r["reference"],
                     r["channel"], r["group"], float(r["amount"])] for r in rows]
            title = f"{self.dept.name} collections {start:%d %b %Y}–{end:%d %b %Y}"
            fn = f"collections_{self.dept.slug or self.dept.id}"
            if export == "xlsx":
                return xlsx_response(f"{fn}.xlsx", header, data, title=title,
                                     church=SiteConfig.get().church_name)
            return csv_response(f"{fn}.csv", header, data)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        from django.core.paginator import Paginator
        ctx = super().get_context_data(**kwargs)
        start, end = parse_period(self.request)
        ctx["dept"] = self.dept
        ctx["start"], ctx["end"] = start, end
        rows = _collection_rows(self.dept, start, end, self.request.user)
        q = (self.request.GET.get("q") or "").strip().lower()
        if q:
            rows = [r for r in rows if q in (r.get("who") or "").lower()
                    or q in (r.get("reference") or "").lower()
                    or q in (r.get("phone") or "").lower()]
        ctx["q"] = self.request.GET.get("q", "")
        ctx["total"] = sum((r["amount"] for r in rows), Decimal(0))
        ctx["count"] = len(rows)
        page = Paginator(rows, 50).get_page(self.request.GET.get("page"))
        ctx["page_obj"] = page
        ctx["is_paginated"] = page.has_other_pages()
        ctx["rows"] = page.object_list
        return ctx


class LeaderExpensesView(LeaderRequiredMixin, TemplateView):
    """Full, downloadable expense list for one of the leader's departments."""
    template_name = "leaders/expenses.html"

    def get(self, request, *args, **kwargs):
        self.dept = assert_department_allowed(request.user, kwargs["pk"])
        if not self.dept:
            return redirect("leader_dashboard")
        start, end = parse_period(request)
        export = request.GET.get("export")
        if export in ("csv", "xlsx"):
            from reports.exports import csv_response, xlsx_response
            from core.models import SiteConfig
            from cashbook.models import Expense
            exps = (Expense.objects.filter(department=self.dept,
                        date__gte=start, date__lte=end).order_by("-date", "-id"))
            header = ["Date", "Description", "Category", "Claimant", "Method",
                      "Status", "Amount"]
            data = [[e.date.isoformat(), e.description, e.get_category_display(),
                     e.claimant or "", e.get_method_display(),
                     e.get_status_display(), float(e.amount)] for e in exps]
            title = f"{self.dept.name} expenses {start:%d %b %Y}–{end:%d %b %Y}"
            fn = f"expenses_{self.dept.slug or self.dept.id}"
            if export == "xlsx":
                return xlsx_response(f"{fn}.xlsx", header, data, title=title,
                                     church=SiteConfig.get().church_name)
            return csv_response(f"{fn}.csv", header, data)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        from cashbook.models import Expense
        from django.db.models import Sum, Q
        from django.core.paginator import Paginator
        ctx = super().get_context_data(**kwargs)
        start, end = parse_period(self.request)
        ctx["dept"] = self.dept
        ctx["start"], ctx["end"] = start, end
        q = (self.request.GET.get("q") or "").strip()
        status = (self.request.GET.get("status") or "").strip()
        exps = (Expense.objects.filter(department=self.dept,
                    date__gte=start, date__lte=end).order_by("-date", "-id"))
        if q:
            exps = exps.filter(Q(description__icontains=q) | Q(claimant__icontains=q)
                               | Q(voucher_no__icontains=q))
        if status:
            exps = exps.filter(status=status)
        ctx["q"] = q
        ctx["status"] = status
        ctx["status_choices"] = Expense.Status.choices
        ctx["total"] = exps.aggregate(s=Sum("amount"))["s"] or Decimal(0)
        ctx["count"] = exps.count()
        page = Paginator(exps, 50).get_page(self.request.GET.get("page"))
        ctx["page_obj"] = page
        ctx["is_paginated"] = page.has_other_pages()
        ctx["expenses"] = page.object_list
        return ctx


class LeaderGroupDetailView(LeaderRequiredMixin, TemplateView):
    """Drill-down for a single development group: its performance vs target and
    the full per-contributor list, downloadable. Visible only to a leader who
    leads a development department."""
    template_name = "leaders/group_detail.html"

    def get(self, request, *args, **kwargs):
        if not _leads_a_development_dept(request.user):
            return redirect("leader_dashboard")
        self.group = get_object_or_404(DevelopmentGroup, pk=kwargs["pk"])
        start, end = parse_period(request)
        export = request.GET.get("export")
        if export in ("csv", "xlsx"):
            from reports.exports import csv_response, xlsx_response
            from core.models import SiteConfig
            data_obj = balances.dev_group_members(self.group, start, end)
            header = ["Contributor", "Phone", "Contributions", "Total"]
            data = [[display_giver(request.user, r["name"]), display_phone(request.user, r["phone"]), r["count"], float(r["total"])]
                    for r in data_obj["rows"]]
            title = f"{self.group.label} contributions {start:%d %b %Y}–{end:%d %b %Y}"
            fn = f"group_{self.group.number}_contributions"
            if export == "xlsx":
                return xlsx_response(f"{fn}.xlsx", header, data, title=title,
                                     church=SiteConfig.get().church_name)
            return csv_response(f"{fn}.csv", header, data)
        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        start, end = parse_period(self.request)
        ctx["start"], ctx["end"] = start, end
        g = self.group
        ctx["group"] = g
        data_obj = balances.dev_group_members(g, start, end)
        # mask phones for display
        ctx["rows"] = [{"name": display_giver(self.request.user, r["name"]), "phone": display_phone(self.request.user, r["phone"]),
                        "count": r["count"], "total": r["total"]}
                       for r in data_obj["rows"]]
        ctx["total"] = data_obj["total"]
        ctx["target"] = g.target or Decimal(0)
        ctx["pct"] = (int(min(data_obj["total"] / g.target * 100, 100))
                      if g.target else None)
        return ctx


class LeaderPledgesView(LeaderRequiredMixin, TemplateView):
    """Full, downloadable pledge list for one of the leader's departments."""
    template_name = "leaders/pledges.html"

    def get(self, request, *args, **kwargs):
        self.dept = assert_department_allowed(request.user, kwargs["pk"])
        if not self.dept:
            return redirect("leader_dashboard")
        export = request.GET.get("export")
        if export in ("csv", "xlsx"):
            from reports.exports import csv_response, xlsx_response
            from core.models import SiteConfig
            rows = self._rows()
            header = ["Member", "Phone", "Campaign", "Pledged", "Paid",
                      "Outstanding", "Status"]
            data = [[r["member"], r["phone"], r["campaign"], float(r["amount"]),
                     float(r["paid"]), float(r["outstanding"]), r["status"]]
                    for r in rows]
            title = f"{self.dept.name} pledges"
            fn = f"pledges_{self.dept.slug or self.dept.id}"
            if export == "xlsx":
                return xlsx_response(f"{fn}.xlsx", header, data, title=title,
                                     church=SiteConfig.get().church_name)
            return csv_response(f"{fn}.csv", header, data)
        return super().get(request, *args, **kwargs)

    def _rows(self):
        from pledges.models import Pledge
        out = []
        try:
            pledges = (Pledge.objects.filter(campaign__target_department=self.dept)
                       .exclude(status=Pledge.Status.CANCELLED)
                       .select_related("member", "campaign").order_by("-start_date"))
            for p in pledges:
                out.append({
                    "member": display_giver(self.request.user, p.member.name), "phone": display_phone(self.request.user, p.member.phone),
                    "campaign": p.campaign.name, "amount": p.amount,
                    "paid": p.paid, "outstanding": p.outstanding,
                    "status": p.get_status_display(), "pct": p.percent_paid
                    if hasattr(p, "percent_paid") else
                    (int(min(p.paid / p.amount * 100, 100)) if p.amount else 0)})
        except Exception:
            pass
        return out

    def get_context_data(self, **kwargs):
        from decimal import Decimal
        ctx = super().get_context_data(**kwargs)
        ctx["dept"] = self.dept
        rows = self._rows()
        ctx["rows"] = rows
        ctx["pledged"] = sum((r["amount"] for r in rows), Decimal(0))
        ctx["paid"] = sum((r["paid"] for r in rows), Decimal(0))
        ctx["outstanding"] = sum((r["outstanding"] for r in rows), Decimal(0))
        ctx["pct"] = (int(min(ctx["paid"] / ctx["pledged"] * 100, 100))
                      if ctx["pledged"] else 0)
        return ctx


# ---------------------------------------------------------------------------
# Leader: staff advances issued to their departments
# A leader may view advances on their funds and record the expenses that
# account for them. Each settling expense is APPROVED + PAID, with the leader
# named as the claimant. Scope is enforced server-side via allowed_departments.
# ---------------------------------------------------------------------------
from django.contrib import messages  # noqa: E402
from django.views import View  # noqa: E402
from cashbook.models import StaffAdvance, Expense  # noqa: E402
from .permissions import LeaderRequiredMixin  # noqa: E402


def _leader_advances(user):
    ids = set(allowed_departments(user).values_list("id", flat=True))
    return (StaffAdvance.objects.filter(department_id__in=ids)
            .select_related("department").order_by("-date_issued", "-id"))


class LeaderAdvancesView(LeaderRequiredMixin, TemplateView):
    template_name = "leaders/advances.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        advs = list(_leader_advances(self.request.user))
        ctx["advances"] = advs
        ctx["outstanding"] = sum((a.balance for a in advs
                                  if a.status != StaffAdvance.Status.CLOSED
                                  and a.balance > 0), Decimal(0))
        return ctx


class LeaderAdvanceDetailView(LeaderRequiredMixin, View):
    template_name = "leaders/advance_detail.html"

    def _get_advance(self, request, pk):
        adv = get_object_or_404(StaffAdvance, pk=pk)
        if not assert_department_allowed(request.user, adv.department_id):
            return None
        return adv

    def get(self, request, pk):
        import datetime as _dt
        from django.core.paginator import Paginator
        from cashbook.views import _advance_detail_ctx
        adv = self._get_advance(request, pk)
        if not adv:
            return redirect("leader_advances")
        ctx = _advance_detail_ctx(adv, leader_mode=True, user=request.user)
        # Excel download of the full statement
        if request.GET.get("export") == "xlsx":
            return self._export_xlsx(adv, ctx.get("statement", []))
        # optional date-range + search filters over the statement timeline; the
        # running balance on each row is already computed, so filtering only hides
        # rows without distorting the figures.
        rows = ctx.get("statement", [])
        q = (request.GET.get("q") or "").strip().lower()
        start = _parse_date(request.GET.get("start"))
        end = _parse_date(request.GET.get("end"))
        if start:
            rows = [r for r in rows if r["date"] >= start]
        if end:
            rows = [r for r in rows if r["date"] <= end]
        if q:
            rows = [r for r in rows if q in (r.get("label") or "").lower()]
        ctx["q"] = request.GET.get("q", "")
        ctx["f_start"] = request.GET.get("start", "")
        ctx["f_end"] = request.GET.get("end", "")
        ctx["filtered"] = bool(q or start or end)
        page = Paginator(rows, 25).get_page(request.GET.get("page"))
        ctx["page_obj"] = page
        ctx["is_paginated"] = page.has_other_pages()
        ctx["statement"] = page.object_list
        return render(request, self.template_name, ctx)

    def _export_xlsx(self, adv, statement):
        import io, openpyxl
        from openpyxl.styles import Font, PatternFill
        from django.http import HttpResponse
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Advance statement"
        ws.append(["Staff advance statement"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append(["Staff", adv.staff_name])
        ws.append(["Fund", adv.department.name])
        ws.append(["Purpose", adv.purpose])
        ws.append(["Issued", adv.date_issued.isoformat() if adv.date_issued else ""])
        ws.append(["Total advanced", float(adv.amount or 0)])
        ws.append(["Balance to account for", float(getattr(adv, "balance", 0) or 0)])
        ws.append(["Status", adv.get_status_display()])
        ws.append([])
        hdr_row = ws.max_row + 1
        head = ["Date", "Detail", "Out to holder", "Accounted", "Still to account"]
        ws.append(head)
        for c in range(1, len(head) + 1):
            cell = ws.cell(hdr_row, c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F5F4F")
        for r in statement:
            ws.append([
                r["date"].isoformat() if r.get("date") else "",
                r.get("label", ""),
                float(r["out"]) if r.get("out") is not None else "",
                float(r["back"]) if r.get("back") is not None else "",
                float(r["running"]) if r.get("running") is not None else "",
            ])
        ws.column_dimensions["A"].width = 13
        ws.column_dimensions["B"].width = 46
        for col in ("C", "D", "E"):
            ws.column_dimensions[col].width = 16
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f"advance_{adv.id}_{adv.staff_name}".replace(" ", "_")[:60] + ".xlsx"
        resp = HttpResponse(buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp

    def post(self, request, pk):
        from decimal import Decimal, InvalidOperation
        from cashbook.views import _record_advance_expense
        from cashbook.models import Expense, ExpenseAttachment
        from core.utils import block_if_locked
        adv = self._get_advance(request, pk)
        if not adv:
            return redirect("leader_advances")
        if adv.status == StaffAdvance.Status.CLOSED:
            messages.error(request, "This advance is closed; ask the treasurer to amend it.")
            return redirect("leader_advance_detail", pk=pk)
        action = request.POST.get("action", "add_expense")

        # --- attach a receipt / M-Pesa confirmation to one of the advance's lines ---
        if action == "add_attachment":
            exp = Expense.objects.filter(pk=request.POST.get("expense_id"),
                                         advance=adv).first()
            if not exp:
                messages.error(request, "That expense is not on this advance.")
                return redirect("leader_advance_detail", pk=pk)
            f = request.FILES.get("file")
            text = (request.POST.get("text") or "").strip()
            link = (request.POST.get("link") or "").strip()
            ALLOWED = (".pdf", ".jpg", ".jpeg", ".png", ".heic", ".webp", ".gif")
            if f and (not f.name.lower().endswith(ALLOWED) or f.size > 10 * 1024 * 1024):
                messages.error(request, "Receipts must be a PDF or image up to 10 MB.")
                return redirect("leader_advance_detail", pk=pk)
            if f or text or link:
                ExpenseAttachment.objects.create(expense=exp, file=f or None,
                    text=text, link=link, label=(request.POST.get("label") or "")[:120],
                    uploaded_by=request.user)
                messages.success(request, "Receipt attached.")
            else:
                messages.error(request, "Add a file, paste an M-Pesa message, or enter a link.")
            return redirect("leader_advance_detail", pk=pk)

        # --- edit one of the leader's own expense lines (before closure) ---
        if action == "edit_expense":
            exp = Expense.objects.filter(pk=request.POST.get("expense_id"),
                                         advance=adv).first()
            mine = (request.user.get_full_name() or request.user.username)
            if not exp or exp.recorded_by_id != request.user.id:
                messages.error(request, "You can only edit expenses you entered.")
                return redirect("leader_advance_detail", pk=pk)
            if exp.category == Expense.Category.BANK_CHARGE:
                messages.error(request, "Transaction-charge lines can't be edited directly.")
                return redirect("leader_advance_detail", pk=pk)
            try:
                new_amt = Decimal(request.POST.get("amount") or exp.amount)
            except InvalidOperation:
                new_amt = exp.amount
            if new_amt <= 0:
                messages.error(request, "Amount must be positive.")
                return redirect("leader_advance_detail", pk=pk)
            # re-check the advance limit with the proposed amount swapped in
            if (adv.balance + exp.amount - new_amt) < 0:
                messages.error(request, "That amount would exceed the advance balance.")
                return redirect("leader_advance_detail", pk=pk)
            try:
                exp.date = dt.date.fromisoformat(request.POST.get("date"))
            except (TypeError, ValueError):
                pass
            exp.amount = new_amt
            exp.description = (request.POST.get("description") or exp.description)[:200]
            exp.save(update_fields=["date", "amount", "description"])
            messages.success(request, "Expense updated.")
            return redirect("leader_advance_detail", pk=pk)

        # --- delete one of the leader's own lines (an expense or a transaction
        #     charge) while the period is open and the advance is still pending ---
        if action == "delete_expense":
            exp = Expense.objects.filter(pk=request.POST.get("expense_id"),
                                         advance=adv).first()
            if not exp or exp.recorded_by_id != request.user.id:
                messages.error(request, "You can only delete lines you entered.")
                return redirect("leader_advance_detail", pk=pk)
            if adv.status in (StaffAdvance.Status.SETTLED, StaffAdvance.Status.CLOSED):
                messages.error(request, "This advance is settled or closed; ask the "
                    "treasurer to make changes.")
                return redirect("leader_advance_detail", pk=pk)
            if block_if_locked(request, exp.date):
                return redirect("leader_advance_detail", pk=pk)
            is_charge = exp.category == Expense.Category.BANK_CHARGE
            # deleting an expense also removes any transaction charge attached to it
            exp.charges.all().delete()
            exp.delete()
            messages.success(request,
                "Charge deleted." if is_charge else "Expense deleted.")
            return redirect("leader_advance_detail", pk=pk)

        # --- default: record a new expense (with optional transaction charge) ---
        try:
            amount = Decimal(request.POST.get("amount") or "0")
        except InvalidOperation:
            amount = Decimal(0)
        try:
            charge = Decimal(request.POST.get("charge") or "0")
        except InvalidOperation:
            charge = Decimal(0)
        if charge < 0:
            charge = Decimal(0)
        desc = (request.POST.get("description") or "").strip()
        if not (desc and amount > 0):
            messages.error(request, "A description and a positive amount are required.")
            return redirect("leader_advance_detail", pk=pk)
        try:
            d = dt.date.fromisoformat(request.POST.get("date"))
        except (TypeError, ValueError):
            d = dt.date.today()
        if block_if_locked(request, d):
            return redirect("leader_advance_detail", pk=pk)
        claimant = (request.user.get_full_name() or request.user.username)
        _exp, err = _record_advance_expense(adv, date=d, desc=desc, amount=amount,
            category=request.POST.get("category"), user=request.user,
            claimant=claimant, charge=charge)
        if err:
            messages.error(request, err)
        else:
            messages.success(request, "Expense recorded against the advance."
                + (f" Transaction charge of KSh {charge:,.2f} added." if charge else ""))
        return redirect("leader_advance_detail", pk=pk)
