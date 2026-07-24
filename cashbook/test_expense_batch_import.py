"""Batch entry, and the spreadsheet import's field coverage.

Both exist for the same reason: the expense form is the reference, and every
other way of getting an expense into the system must produce the same record.
When they diverge, the divergence is silent — a spreadsheet row missing a
supplier looks fine until someone asks what the church owes that supplier.
"""
import datetime as dt
from decimal import Decimal

from django.contrib.auth.models import Group, User
from django.test import Client, TestCase
from django.urls import reverse

from core.roles import TREASURER
from departments.models import Department
from vendors.models import Vendor

from .models import Expense
from .services import expenses as expense_svc

TODAY = dt.date.today()


class Base(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("bt", password="batch-pass-1")
        self.user.groups.add(Group.objects.get_or_create(name=TREASURER)[0])
        self.fund = Department.objects.create(
            name="Youth Fund", slug="youth-fund",
            fund_type=Department.FundType.LOCAL,
            category=Department.Category.MINISTRY,
            active=True, show_in_expenses=True, selectable=True)
        self.supplier = Vendor.objects.create(name="Mwangi Hardware Ltd")
        self.client = Client()
        self.client.get("/accounts/login/")
        self.client.post("/accounts/login/",
                         {"username": "bt", "password": "batch-pass-1"}, follow=True)


class ExpenseBatchEntryTests(Base):
    """Several receipts, one date, one fund, one claimant."""

    def _post(self, **over):
        data = {
            "date": TODAY.isoformat(), "department": self.fund.pk,
            "claimant": "J. Mwangi", "method": "MPESA", "category": "TRANSPORT",
            "vendor": self.supplier.pk,
            "line_description": ["Fuel to Kisii", "Fare for choir", ""],
            "line_amount": ["1200", "800", ""],
            "line_category": ["", "REFRESHMENTS", ""],
            "line_charge": ["30", "", ""],
        }
        data.update(over)
        return self.client.post(reverse("expense_batch"), data)

    def test_the_page_renders(self):
        self.assertEqual(self.client.get(reverse("expense_batch")).status_code, 200)

    def test_each_line_becomes_its_own_expense(self):
        self._post()
        self.assertEqual(
            Expense.objects.filter(charge_for__isnull=True).count(), 2)

    def test_the_shared_header_is_applied_to_every_line(self):
        self._post()
        for expense in Expense.objects.filter(charge_for__isnull=True):
            self.assertEqual(expense.department, self.fund)
            self.assertEqual(expense.date, TODAY)
            self.assertEqual(expense.claimant, "J. Mwangi")
            self.assertEqual(expense.method, "MPESA")
            self.assertEqual(expense.vendor, self.supplier)

    def test_a_line_may_override_the_default_category(self):
        self._post()
        fuel = Expense.objects.get(description="Fuel to Kisii")
        choir = Expense.objects.get(description="Fare for choir")
        self.assertEqual(fuel.category, "TRANSPORT")     # header default
        self.assertEqual(choir.category, "REFRESHMENTS") # line override

    def test_a_charge_becomes_its_own_bank_charge_expense(self):
        """Not added to the amount: 1,200 of transport plus 30 of bank charge,
        because merging them would overstate transport and hide the fee."""
        self._post()
        fuel = Expense.objects.get(description="Fuel to Kisii")
        charge = Expense.objects.get(charge_for=fuel)
        self.assertEqual(fuel.amount, Decimal("1200"))
        self.assertEqual(charge.amount, Decimal("30"))
        self.assertEqual(charge.category, Expense.Category.BANK_CHARGE)
        self.assertEqual(charge.department, self.fund)

    def test_blank_lines_are_ignored(self):
        self._post()
        self.assertFalse(Expense.objects.filter(description="").exists())

    def test_nothing_is_saved_without_a_fund_or_date(self):
        self._post(department="")
        self.assertEqual(Expense.objects.count(), 0)

    def test_the_batch_is_all_or_nothing(self):
        """A half-entered stack of receipts is worse than none — the treasurer
        cannot tell which ones went in."""
        before = Expense.objects.count()
        self._post(line_amount=["1200", "not-a-number", ""])
        self.assertEqual(Expense.objects.count(), before,
                         "Part of the batch was saved after a bad line.")


class SharedRecordingRulesTests(Base):
    """The rules that used to exist in three places."""

    def test_petty_cash_expenses_are_recorded_as_paid_not_pending(self):
        """Money already out of the tin is paid. Recording it as awaiting
        approval leaves the float disagreeing with the drawer."""
        expense, _ = expense_svc.record(
            date=TODAY, department=self.fund, description="Tea", amount=200,
            user=self.user, paid_from_petty_cash=True)
        self.assertEqual(expense.status, Expense.Status.PAID)
        self.assertEqual(expense.paid_date, TODAY)

    def test_the_charge_row_inherits_the_payee_and_status(self):
        """It has to: a charge left pending while its expense is approved sits
        in the queue forever, and one with no payee cannot be matched on the
        bank statement."""
        expense, charge = expense_svc.record(
            date=TODAY, department=self.fund, description="Transfer",
            amount=5000, user=self.user, payee="Mwangi Hardware",
            claimant="J. Mwangi", auto_approve=True, charge=Decimal("55"))
        self.assertEqual(charge.status, expense.status)
        self.assertEqual(charge.payee, "Mwangi Hardware")
        self.assertEqual(charge.claimant, "J. Mwangi")
        self.assertEqual(charge.charge_for, expense)


class ImportTemplateCoverageTests(Base):
    """The spreadsheet must offer what the form offers."""

    def test_the_template_has_a_column_for_every_form_field(self):
        import io
        import openpyxl
        from cashbook.forms import ExpenseForm

        response = self.client.get(reverse("expense_import") + "?download=1")
        self.assertEqual(response.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        header = {str(c.value).strip().lower()
                  for c in next(wb["Expenses"].iter_rows(max_row=1))}

        # Field name -> the column that carries it.
        expected = {
            "date": "date", "department": "fund", "description": "description",
            "amount": "amount", "category": "category", "method": "method",
            "claimant": "claimant", "vendor": "supplier", "payee": "payee",
            "voucher_no": "voucher no", "charge": "m-pesa charge",
            "paid_from_petty_cash": "paid from petty cash",
            "expenditure_type": "expenditure type", "budget_line": "budget item",
        }
        missing = []
        for field in ExpenseForm().fields:
            if field == "capitalized_asset":
                continue     # chosen per asset, not sensible in a flat sheet
            column = expected.get(field)
            if column is None or column not in header:
                missing.append(field)
        self.assertFalse(
            missing,
            f"The import template has no column for: {missing}. A spreadsheet "
            f"that cannot carry a field silently produces the incomplete rows "
            f"the form no longer allows.")

    def test_the_supplier_register_is_offered_as_a_list(self):
        import io
        import openpyxl
        response = self.client.get(reverse("expense_import") + "?download=1")
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        names = {str(c.value) for row in wb["Lists"].iter_rows() for c in row}
        self.assertIn("Mwangi Hardware Ltd", names,
                      "Registered suppliers are not listed for lookup.")


class BatchSharedChargeTests(Base):
    """A stack settled with one transfer attracts one fee, not one per line."""

    def _post(self, **over):
        data = {
            "date": TODAY.isoformat(), "department": self.fund.pk,
            "claimant": "J. Mwangi", "method": "MPESA", "category": "TRANSPORT",
            "voucher_no": "BATCH-7",
            "line_description": ["Fuel", "Fare", "Water"],
            "line_amount": ["1000", "2000", "3000"],
            "line_category": ["", "", ""],
            "line_charge": ["", "", ""],
            "shared_charge": "55",
        }
        data.update(over)
        return self.client.post(reverse("expense_batch"), data)

    def test_one_charge_is_recorded_for_the_whole_batch(self):
        self._post()
        charges = Expense.objects.filter(category=Expense.Category.BANK_CHARGE)
        self.assertEqual(charges.count(), 1,
                         "The batch fee was multiplied across the lines.")
        self.assertEqual(charges.get().amount, Decimal("55"))

    def test_the_fee_is_not_split_across_the_lines(self):
        """Splitting would invent charges the bank never levied, and would
        misstate every line's cost."""
        self._post()
        for expense in Expense.objects.exclude(
                category=Expense.Category.BANK_CHARGE):
            self.assertIn(expense.amount,
                          [Decimal("1000"), Decimal("2000"), Decimal("3000")])

    def test_the_batch_charge_is_not_attached_to_any_single_line(self):
        """`charge_for` means "the fee levied on THIS expense". The batch fee
        belongs to no single line, and attaching it to the first would expose
        it to deletion when that line's own charge is edited."""
        self._post()
        charge = Expense.objects.get(category=Expense.Category.BANK_CHARGE)
        self.assertIsNone(charge.charge_for_id)
        self.assertIn("BATCH-7", charge.description)
        self.assertIn("3 item(s)", charge.description)

    def test_it_lands_on_the_same_fund_and_carries_the_header(self):
        self._post()
        charge = Expense.objects.get(category=Expense.Category.BANK_CHARGE)
        self.assertEqual(charge.department, self.fund)
        self.assertEqual(charge.date, TODAY)
        self.assertEqual(charge.voucher_no, "BATCH-7")
        self.assertEqual(charge.method, "MPESA")

    def test_per_line_and_shared_charges_can_both_apply(self):
        """Most of a stack paid in one transfer, one item paid separately."""
        self._post(line_charge=["", "", "20"])
        charges = Expense.objects.filter(category=Expense.Category.BANK_CHARGE)
        self.assertEqual(charges.count(), 2)
        self.assertEqual(sum(c.amount for c in charges), Decimal("75"))
        # the per-line one stays linked to its own expense; the batch one does not
        self.assertEqual(charges.filter(charge_for__isnull=False).count(), 1)

    def test_no_shared_charge_means_no_extra_row(self):
        self._post(shared_charge="")
        self.assertEqual(
            Expense.objects.filter(category=Expense.Category.BANK_CHARGE).count(), 0)
