"""Enhancement: the Expenses Excel export now includes a Payment Method
column, using the actual recorded payment source (Expense.method plus the
separate paid_from_petty_cash flag) rather than inferring anything."""
import datetime as dt
from decimal import Decimal
from django.test import TestCase, Client
from django.contrib.auth.models import User, Group
from departments.models import Department
from cashbook.models import Expense


def _tr():
    u = User.objects.create_user("tr_pmexport", password="x", is_superuser=True)
    u.groups.add(Group.objects.get_or_create(name="Treasurer")[0])
    return u


class PaymentMethodExportTests(TestCase):
    def setUp(self):
        self.tr = _tr()
        self.d = Department.objects.create(name="PMExportFund", fund_type="LOCAL",
            category="MINISTRY")
        self.c = Client(); self.c.force_login(self.tr)

    def _rows(self, response):
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content))
        return list(wb.active.iter_rows(values_only=True))

    def test_header_includes_payment_method(self):
        rows = self._rows(self.c.get("/expenses/?export=xlsx"))
        header = next(r for r in rows if r and r[0] == "ID")
        self.assertIn("Payment Method", header)

    def test_cash_method_shown_as_cash(self):
        Expense.objects.create(date=dt.date(2026, 6, 1), department=self.d,
            description="PM Cash", amount=Decimal("100"), category="OTHER",
            method="CASH", status="PAID", recorded_by=self.tr, approved_by=self.tr,
            paid_from_petty_cash=False)
        rows = self._rows(self.c.get("/expenses/?export=xlsx"))
        row = next(r for r in rows if r and r[2] == "PM Cash")
        self.assertEqual(row[9], "Cash")

    def test_petty_cash_flag_overrides_method_label(self):
        Expense.objects.create(date=dt.date(2026, 6, 2), department=self.d,
            description="PM Petty", amount=Decimal("100"), category="OTHER",
            method="CASH", status="PAID", recorded_by=self.tr, approved_by=self.tr,
            paid_from_petty_cash=True)
        rows = self._rows(self.c.get("/expenses/?export=xlsx"))
        row = next(r for r in rows if r and r[2] == "PM Petty")
        self.assertEqual(row[9], "Petty Cash")

    def test_bank_method(self):
        Expense.objects.create(date=dt.date(2026, 6, 3), department=self.d,
            description="PM Bank", amount=Decimal("100"), category="OTHER",
            method="BANK", status="PAID", recorded_by=self.tr, approved_by=self.tr)
        rows = self._rows(self.c.get("/expenses/?export=xlsx"))
        row = next(r for r in rows if r and r[2] == "PM Bank")
        self.assertEqual(row[9], "Bank")

    def test_cheque_method(self):
        Expense.objects.create(date=dt.date(2026, 6, 4), department=self.d,
            description="PM Cheque", amount=Decimal("100"), category="OTHER",
            method="CHEQUE", status="PAID", recorded_by=self.tr, approved_by=self.tr)
        rows = self._rows(self.c.get("/expenses/?export=xlsx"))
        row = next(r for r in rows if r and r[2] == "PM Cheque")
        self.assertEqual(row[9], "Cheque")

    def test_mpesa_method_shown_as_mobile_money(self):
        Expense.objects.create(date=dt.date(2026, 6, 5), department=self.d,
            description="PM Mpesa", amount=Decimal("100"), category="OTHER",
            method="MPESA", status="PAID", recorded_by=self.tr, approved_by=self.tr)
        rows = self._rows(self.c.get("/expenses/?export=xlsx"))
        row = next(r for r in rows if r and r[2] == "PM Mpesa")
        self.assertEqual(row[9], "Mobile Money")

    def test_csv_export_also_includes_column(self):
        Expense.objects.create(date=dt.date(2026, 6, 6), department=self.d,
            description="PM Csv", amount=Decimal("100"), category="OTHER",
            method="BANK", status="PAID", recorded_by=self.tr, approved_by=self.tr)
        b = self.c.get("/expenses/?export=csv").content.decode()
        self.assertIn("Payment Method", b)
        self.assertIn("Bank", b)
