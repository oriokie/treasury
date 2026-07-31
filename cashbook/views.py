import datetime as dt

from django.contrib import messages
from django.db import transaction as db_tx
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, View

from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin

from core.utils import block_if_locked as _block_if_locked, PrefPaginationMixin
from core.permissions import DataEntryRequiredMixin, ReadAccessMixin, TreasurerRequiredMixin, AdvanceAccessMixin, PaymentViewMixin
from core.utils import sabbath_week_of
from decimal import InvalidOperation
from django.core.exceptions import ValidationError


#: Prefix used in settlement messages. Read from site config so the wording
#: matches the rest of the app rather than hard-coding a currency here.
def _currency():
    try:
        from core.models import SiteConfig
        return (SiteConfig.get().currency or "") + " "
    except Exception:
        return ""


def _parse_date(value):
    """A date the user typed, or None for 'today'."""
    if not value:
        return None
    try:
        return dt.date.fromisoformat(str(value)[:10])
    except ValueError:
        return None
from departments.models import Department
from .forms import ExpenseForm, FundTransferForm, RecurringExpenseForm
from .models import Expense

# Financial totals live in the treasury-position service (see that module's
# docstring). Re-exported here under their original names — including the
# private `_petty_balance_asof` — so every existing `from cashbook.views import
# ...` keeps working unchanged: assistant, dashboards, period close, statements,
# liabilities, and this module's own views all depend on these names.
from cashbook.services.treasury_position import (
    petty_balance_asof as _petty_balance_asof,
    outstanding_advances_total, outstanding_bank_advances_total,
    outstanding_petty_advances_total, open_payables_total, open_accruals_total,
    unexpired_prepayments_total, unpresented_cheques_total, unpresented_payments_qs)

# Advance helpers shared with leaders/views.py now live in the advances service.
# Re-exported under their original private names so every call site (this
# module's advance views, leaders/views.py) keeps working unchanged.
from cashbook.services.advances import (
    advance_detail_ctx as _advance_detail_ctx,
    record_advance_expense as _record_advance_expense)



def _fund_available(dept_id, as_of):
    """A fund's available (closing) balance as at a date, or None if unknown.
    Targeted single-fund aggregation (no full-portfolio loop)."""
    if not dept_id:
        return None
    from reports.services.balances import fund_balance
    return fund_balance(dept_id, as_of)


class ExpenseListView(PrefPaginationMixin, ReadAccessMixin, ListView):
    model = Expense
    template_name = "cashbook/list.html"
    context_object_name = "expenses"
    paginate_by = 50

    def get_queryset(self):
        import datetime as dt
        qs = (Expense.objects.select_related("department", "recorded_by")
              .prefetch_related("attachments").order_by("-id"))
        from django.db.models import Count
        qs = qs.annotate(n_attachments=Count("attachments"))
        # The Expense Register is OPERATIONAL expenditure only. Liability
        # settlements (trust releases, loan repayments, and any category
        # flagged as liability) live in the Liability Register instead —
        # historical rows included. The posting engine is unaffected.
        qs = qs.filter(doc_class=Expense.DocClass.EXPENSE)
        g = self.request.GET
        if g.get("status"):
            qs = qs.filter(status=g["status"])
        if g.get("department"):
            qs = qs.filter(department_id=g["department"])
        if g.get("category"):
            qs = qs.filter(category=g["category"])
        if g.get("q"):
            from django.db.models import Q
            term = g["q"].strip()
            qs = qs.filter(Q(description__icontains=term)
                           | Q(claimant__icontains=term)
                           | Q(voucher_no__icontains=term))
        from core.utils import default_to_current_month
        start, end = default_to_current_month(self.request, from_param="start", to_param="end")
        if start:
            qs = qs.filter(date__gte=start)
        if end:
            qs = qs.filter(date__lte=end)
        return qs

    def get(self, request, *args, **kwargs):
        export = request.GET.get("export")
        if export == "support-pdf":
            return self._supporting_docs_pdf(request)
        if export in ("csv", "xlsx"):
            from reports.exports import csv_response, xlsx_response
            from core.models import SiteConfig
            qs = self.get_queryset()
            header = ["ID", "Date", "Description", "Fund", "Category", "Type", "Status",
                      "Claimant", "Voucher", "Payment Method", "Amount"]

            def _payment_method(x):
                # the actual recorded payment source — never inferred: petty
                # cash is its own recorded flag (paid_from_petty_cash), distinct
                # from the payment method (which is otherwise always CASH for
                # a petty-cash disbursement); everything else comes straight
                # from Expense.method.
                if x.paid_from_petty_cash:
                    return "Petty Cash"
                return {"CASH": "Cash", "BANK": "Bank", "CHEQUE": "Cheque",
                        "MPESA": "Mobile Money"}.get(x.method) or (x.method or "Other")

            rows = [[x.id, x.date.isoformat(), x.description,
                     x.department.name if x.department_id else "",
                     x.get_category_display(), x.get_expenditure_type_display(),
                     x.get_status_display(), x.claimant or "", x.voucher_no or "",
                     _payment_method(x),
                     float(x.amount)] for x in qs]
            if export == "xlsx":
                return xlsx_response("expenses.xlsx", header, rows, title="Expenses",
                                     church=SiteConfig.get().church_name)
            return csv_response("expenses.csv", header, rows)
        return super().get(request, *args, **kwargs)

    def _supporting_docs_pdf(self, request):
        from django.http import HttpResponse
        from core.models import SiteConfig
        from .services.supporting_pdf import (build_supporting_docs_pdf,
                                              HAVE_PDF_LIBS)
        if not HAVE_PDF_LIBS:
            messages.error(request, "PDF generation isn't available on the server yet. "
                "Ask the administrator to install the reportlab and pypdf packages.")
            return redirect("expense_list")
        # only expenses with an actual FILE attachment (a scanned receipt or
        # photo). Text (M-Pesa message) and link-only attachments are excluded —
        # the Receipts view already presents those, and a summary page with no
        # document adds nothing to a printed supporting-documents bundle.
        # NOTE: .exclude(attachments__file="") on a to-many relation excludes
        # the PARENT if it has *any* related row matching — so an expense with
        # one real file attachment AND one text-only attachment (file="") gets
        # wrongly excluded entirely (a well-known Django ORM gotcha with
        # exclude() across multi-valued relations). Both conditions must be in
        # the SAME .filter() call so they're required of the SAME related row.
        file_ids = (Expense.objects
                    .filter(attachments__file__isnull=False, attachments__file__gt="")
                    .values("id"))
        qs = (self.get_queryset().filter(id__in=file_ids)
              .distinct().prefetch_related("attachments"))
        if not qs.exists():
            messages.info(request, "None of the expenses matching the current "
                          "filters have a supporting document (file) attached.")
            params = request.GET.copy()
            params.pop("export", None)
            suffix = f"?{params.urlencode()}" if params else ""
            return redirect(f"{request.path}{suffix}")
        data, stats = build_supporting_docs_pdf(
            qs, church=SiteConfig.get().church_name or "")
        resp = HttpResponse(data, content_type="application/pdf")
        resp["Content-Disposition"] = 'attachment; filename="supporting_documents.pdf"'
        return resp

    def get_context_data(self, **kwargs):
        from django.db.models import Sum, Count
        ctx = super().get_context_data(**kwargs)
        ctx["statuses"] = Expense.Status.choices
        ctx["exp_types"] = Expense.ExpenditureType.choices
        ctx["categories"] = Expense.Category.choices
        ctx["departments"] = Department.objects.filter(active=True)
        # As with the ledger (giving.TransactionListView): show what's
        # ACTUALLY being filtered, including the current-month default
        # applied on a bare visit, rather than leave the date inputs blank
        # while secretly narrowing the results.
        from core.utils import default_to_current_month
        filters = self.request.GET.copy()
        start, end = default_to_current_month(self.request, from_param="start", to_param="end")
        filters["start"] = start.isoformat() if start else ""
        filters["end"] = end.isoformat() if end else ""
        ctx["filters"] = filters
        ctx["date_default_applied"] = not self.request.GET
        ctx["filtered_total"] = self.get_queryset().aggregate(t=Sum("amount"))["t"] or 0
        # status money-bar: totals within the other active filters, ignoring the
        # status filter itself, so the chips always show the whole picture
        g = self.request.GET.copy()
        g.pop("status", None)
        base = Expense.objects.all()
        if g.get("department"):
            base = base.filter(department_id=g["department"])
        if g.get("category"):
            base = base.filter(category=g["category"])
        if g.get("q"):
            from django.db.models import Q
            term = g["q"].strip()
            base = base.filter(Q(description__icontains=term)
                               | Q(claimant__icontains=term)
                               | Q(voucher_no__icontains=term))
        import datetime as _dt
        for key, lookup in (("start", "date__gte"), ("end", "date__lte")):
            if g.get(key):
                try:
                    base = base.filter(**{lookup: _dt.date.fromisoformat(g[key])})
                except ValueError:
                    pass
        agg = {r["status"]: r for r in base.values("status")
               .annotate(n=Count("id"), t=Sum("amount"))}
        ctx["status_bar"] = [
            {"code": code, "label": label,
             "n": agg.get(code, {}).get("n", 0),
             "t": agg.get(code, {}).get("t", 0) or 0,
             "active": self.request.GET.get("status") == code}
            for code, label in Expense.Status.choices]
        ctx["status_qs"] = g.urlencode()
        # split the filtered total into operating vs trust-remittance, so the
        # page can show a true operating-expense figure (remittances are a
        # liability settlement, not expenditure)
        qs = self.get_queryset()
        remit_total = (qs.filter(category=Expense.Category.REMITTANCE)
                       .aggregate(t=Sum("amount"))["t"] or 0)
        ctx["filtered_remittances"] = remit_total
        ctx["filtered_operating"] = (ctx["filtered_total"] or 0) - remit_total
        from core.models import SiteConfig
        ctx["dual_threshold"] = SiteConfig.get().dual_approval_threshold or 0
        return ctx


class ExpenseCreate(DataEntryRequiredMixin, CreateView):
    model = Expense
    form_class = ExpenseForm
    template_name = "cashbook/form.html"
    success_url = reverse_lazy("expense_list")

    def _settle_target(self):
        """Resolve a ?settle=payable:5 / accrual:3 marker to the obligation being
        paid, so settling opens this form pre-filled and editable (method,
        claimant, charges) instead of silently posting a fixed expense."""
        from .models import Payable, Accrual
        raw = self.request.GET.get("settle") or self.request.POST.get("settle") or ""
        kind, _, sid = raw.partition(":")
        if not sid.isdigit():
            return None, None
        if kind == "payable":
            return "payable", Payable.objects.filter(pk=sid, settled=False).first()
        if kind == "accrual":
            return "accrual", Accrual.objects.filter(pk=sid, settled=False).first()
        return None, None

    def get_initial(self):
        initial = super().get_initial()
        # `?petty=1` — how the petty cash page now sends someone here. The old
        # separate "record a disbursement" form wrote exactly this Expense with
        # paid_from_petty_cash=True; retiring it left this form to do the job, so
        # the link that used to open that one opens this, ready.
        if self.request.GET.get("petty"):
            initial["paid_from_petty_cash"] = True
            initial["method"] = Expense.Method.CASH
        kind, obj = self._settle_target()
        if obj:
            initial.update({
                "date": dt.date.today(), "department": obj.department_id,
                "description": getattr(obj, "description", "")[:200],
                # The balance, not the invoice total: a payable already part
                # paid should open at what is left, or the treasurer has to
                # work it out and retype it every instalment.
                "amount": getattr(obj, "balance", None) or obj.amount,
                "category": obj.category,
                "method": Expense.Method.BANK})
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        kind, obj = self._settle_target()
        if obj:
            ctx["settle"] = f"{kind}:{obj.pk}"
            ctx["settle_label"] = (f"Settling {kind}: {getattr(obj, 'vendor', '') or ''} "
                                   f"{obj.description}".strip())
        from core.models import SiteConfig
        from core.roles import is_treasurer
        from statements.models import BankAccount
        from .models import PaymentInstrument
        ctx["bank_accounts"] = BankAccount.objects.all()
        ctx["payment_methods"] = PaymentInstrument.Method.choices
        # issuing a payment at entry only makes sense once the expense is approved;
        # when approval is required, this expense will start PENDING, so offer it
        # to a treasurer only (who can approve immediately after) — otherwise hide it.
        cfg = SiteConfig.get()
        ctx["can_issue_payment_now"] = (not cfg.require_expense_approval) or is_treasurer(self.request.user)
        ctx["require_expense_approval"] = cfg.require_expense_approval
        return ctx

    def form_valid(self, form):
        from core.models import SiteConfig
        from core.roles import is_treasurer
        exp = form.save(commit=False)
        if _block_if_locked(self.request, exp.date):
            return redirect("expense_create")
        # H4: block an expense that would overdraw the fund. Treasurers may override
        # with an explicit, logged confirmation; assistants are always blocked.
        cfg = SiteConfig.get()
        if cfg.enforce_fund_balance:
            available = _fund_available(exp.department_id, exp.date)
            if available is not None and exp.amount > available:
                override = is_treasurer(self.request.user) and \
                    self.request.POST.get("override_balance")
                if not override:
                    messages.error(self.request,
                        f"{exp.department.name} has {available:,.2f} available — not enough for "
                        f"this expense of {exp.amount:,.2f}." +
                        (" Tick \u201coverride\u201d to record it anyway." if is_treasurer(self.request.user)
                         else " A treasurer must record an over-budget expense."))
                    ctx = self.get_context_data(form=form)
                    ctx["overspend_warning"] = is_treasurer(self.request.user)
                    return self.render_to_response(ctx)
                messages.warning(self.request,
                    f"Override: {exp.department.name} taken below zero by "
                    f"{exp.recorded_by if False else self.request.user.username}.")
        exp.recorded_by = self.request.user
        exp.sabbath_week = sabbath_week_of(exp.date)
        issue_now = bool(self.request.POST.get("issue_payment"))
        auto = not SiteConfig.get().require_expense_approval
        # a treasurer issuing a payment at entry is implicitly approving it too —
        # matches the existing detail-page pattern where "Mark paid" also approves.
        if auto or (issue_now and is_treasurer(self.request.user)):
            exp.status = Expense.Status.APPROVED
            exp.approved_by = self.request.user
        exp.save()
        # optional M-Pesa / bank transaction charge -> separate bank-charge expense
        charge = form.cleaned_data.get("charge")
        if charge and charge > 0:
            # One implementation of the charge row, shared with the import and
            # the batch screen — see services.expenses. The copies here had
            # already drifted on `payee` and `claimant`, which is invisible
            # until a bank reconciliation cannot match the charge.
            from .services.expenses import _record_charge
            _record_charge(exp, charge, self.request.user)
        payment_note = ""
        if issue_now and exp.status == Expense.Status.APPROVED:
            from .models import PaymentInstrument
            method = self.request.POST.get("payment_method") or "CHEQUE"
            if method not in dict(PaymentInstrument.Method.choices):
                method = "CHEQUE"
            ref = (self.request.POST.get("payment_reference") or "").strip()[:40]
            try:
                issued = (dt.date.fromisoformat(self.request.POST.get("payment_date"))
                          if self.request.POST.get("payment_date") else exp.date)
            except ValueError:
                issued = exp.date
            bank_id = self.request.POST.get("payment_bank_account") or ""
            inst = PaymentInstrument(
                method=method, instrument_number=ref,
                # the payee is who the money goes TO; the claimant is who asked
                # for it. They are often the same person, and often not.
                payee=(exp.payee or exp.claimant or exp.department.name)[:160],
                amount=exp.amount, date_issued=issued,
                status=PaymentInstrument.Status.ISSUED,
                source_kind=PaymentInstrument.SourceKind.EXPENSE,
                expense=exp, recorded_by=self.request.user)
            if bank_id.isdigit():
                from statements.models import BankAccount
                inst.bank_account = BankAccount.objects.filter(pk=bank_id).first()
            inst.save()
            payment_note = (f" {inst.get_method_display()} "
                            f"{ref or '(no reference)'} issued to settle it.")
        if exp.status == Expense.Status.PENDING:
            from core.services.notifications import notify
            from django.urls import reverse
            notify("APPROVAL",
                   f"Expense awaiting approval: {exp.description} — "
                   f"{exp.amount:,.2f} ({exp.department.name})",
                   link=reverse("expense_list"))
        # if this expense settles a payable/accrual, link and close it
        kind, obj = self._settle_target()
        if obj and not obj.settled:
            if kind in ("payable", "accrual"):
                # A payable may be paid over several instalments, so the expense
                # is LINKED and the service decides whether that clears it. The
                # flags are never set here — services.payables.refresh_settlement
                # is the only writer, which is what keeps them agreeing with the
                # payments.
                from .services import obligations as obligation_svc
                field = "accrual" if kind == "accrual" else "payable"
                setattr(exp, field, obj)
                exp.save(update_fields=[field])
                obligation_svc.refresh_settlement(obj)
                obj.refresh_from_db()
                if obj.settled:
                    messages.success(
                        self.request,
                        f"{kind.title()} settled in full and recorded as an expense.")
                else:
                    who = getattr(obj, "vendor", "") or obj.description
                    messages.success(
                        self.request,
                        f"Part payment recorded. {obj.balance:,.2f} still owing "
                        f"on {who}.")
            else:
                obj.settled = True
                obj.settled_on = exp.date
                obj.settled_expense = exp
                obj.save()
                messages.success(self.request,
                    f"{kind.title()} settled and recorded as an expense.")
            return redirect("accruals")
        messages.success(self.request, f"Expense recorded.{payment_note}")
        return redirect(self.success_url)


class ExpenseUpdate(DataEntryRequiredMixin, UpdateView):
    model = Expense
    form_class = ExpenseForm
    template_name = "cashbook/form.html"
    success_url = reverse_lazy("expense_list")

    def get_initial(self):
        initial = super().get_initial()
        charge = self.object.charges.first() if self.object else None
        if charge:
            initial["charge"] = charge.amount
        return initial

    def form_valid(self, form):
        from core.roles import is_treasurer
        exp = form.save(commit=False)
        # #1 — respect period locks on edits, just like creation
        if _block_if_locked(self.request, exp.date):
            return redirect(self.success_url)
        original = Expense.objects.get(pk=exp.pk) if exp.pk else None
        # also block if the expense is being moved OUT of a locked period
        if original and _block_if_locked(self.request, original.date):
            return redirect(self.success_url)
        if original and original.status != Expense.Status.PENDING:
            # #2a — only a treasurer may edit an expense that is past PENDING
            if not is_treasurer(self.request.user):
                messages.error(self.request,
                    "This expense has been approved/paid — only a treasurer can edit "
                    "it. Ask a treasurer, or it must be reversed first.")
                return redirect(self.success_url)
            # #2b — a material change (amount or fund) voids the prior approval
            if (original.amount != exp.amount
                    or original.department_id != exp.department_id):
                exp.status = Expense.Status.PENDING
                exp.approved_by = None
                exp.second_approved_by = None
                exp.paid_date = None
                messages.warning(self.request,
                    "Amount or fund changed — the expense was returned to pending "
                    "approval and must be re-approved.")
        exp.sabbath_week = sabbath_week_of(exp.date)
        exp.save()
        from core.models import reconciled_period_warning
        warn = reconciled_period_warning(exp.date)
        if warn:
            messages.warning(self.request, warn)
        # Sync the linked transaction-charge entry (M-Pesa/bank charge). Editing an
        # expense must create the charge if newly added, update it in place if it
        # already exists (never duplicate), and remove it if the charge is cleared.
        charge = form.cleaned_data.get("charge")
        existing = exp.charges.first()
        if charge and charge > 0:
            ref = exp.voucher_no or f"exp #{exp.id}"
            desc = f"Transaction charge — {exp.description} [for {ref}]"[:200]
            if existing:
                existing.date = exp.date
                existing.sabbath_week = exp.sabbath_week
                existing.department = exp.department
                existing.description = desc
                existing.amount = charge
                existing.method = exp.method
                existing.voucher_no = exp.voucher_no
                existing.paid_from_petty_cash = exp.paid_from_petty_cash
                existing.status = exp.status
                existing.save()
                # remove any accidental extra charge rows to prevent double-counting
                exp.charges.exclude(pk=existing.pk).delete()
            else:
                from .services.expenses import _record_charge
                created_charge = _record_charge(exp, charge, self.request.user)
                # The edit screen composes its own description (it may mention
                # the change), so keep that wording rather than the default.
                if desc and created_charge.description != desc:
                    created_charge.description = desc
                    created_charge.save(update_fields=["description"])
        elif existing:
            # charge removed on edit — delete the linked charge entry
            exp.charges.all().delete()
        messages.success(self.request, "Expense updated.")
        return redirect(self.success_url)


class ExpenseApprove(TreasurerRequiredMixin, View):
    def post(self, request, pk):
        from core.models import SiteConfig
        exp = get_object_or_404(Expense, pk=pk)
        action = request.POST.get("action")
        cfg = SiteConfig.get()
        threshold = cfg.dual_approval_threshold or 0
        needs_two = threshold and exp.amount >= threshold
        if action == "approve":
            if (cfg.require_different_approver
                    and exp.recorded_by_id == request.user.id):
                messages.error(request,
                    "You recorded this expense yourself — a different treasurer "
                    "must approve it (Settings → require a different approver).")
                return redirect("expense_list")
            exp.status = Expense.Status.APPROVED
            exp.approved_by = request.user
        elif action == "second_approve":
            if exp.status != Expense.Status.APPROVED or not exp.approved_by_id:
                messages.error(request, "This expense must receive its first approval "
                                        "before a second treasurer can co-approve it.")
                return redirect("expense_list")
            if exp.approved_by_id == request.user.id:
                messages.error(request, "A second, different treasurer must co-approve this "
                                        "high-value expense.")
                return redirect("expense_list")
            exp.second_approved_by = request.user
            messages.success(request, "Second approval recorded.")
            exp.save()
            return redirect("expense_list")
        elif action == "reject":
            exp.status = Expense.Status.REJECTED
            exp.rejected_by = request.user
            # do NOT set approved_by on a rejection — it corrupts the audit trail
            # (an auditor would otherwise read "approved by X" on a rejected claim).
            exp.save()
            from core.services.notifications import notify
            reason = (request.POST.get("note") or "").strip()
            if exp.recorded_by_id and exp.recorded_by_id != request.user.id:
                notify("REJECTION",
                       f"Your expense “{exp.description}” ({exp.amount:,.2f}) was rejected"
                       + (f": {reason}" if reason else "."),
                       link="/expenses/", recipients=[exp.recorded_by])
            messages.success(request, "Expense rejected and the submitter notified.")
            return redirect("expense_list")
        elif action == "pay":
            # M1: a high-value expense needs two distinct treasurer approvals before pay
            if needs_two and (not exp.approved_by_id or not exp.second_approved_by_id):
                messages.error(request,
                    f"Expenses of {threshold:,.2f} or more need two treasurers' approval "
                    f"before payment. Have a second treasurer co-approve first.")
                return redirect("expense_list")
            exp.status = Expense.Status.PAID
            exp.paid_date = dt.date.today()
            if not exp.approved_by:
                exp.approved_by = request.user
        exp.save()
        messages.success(request, f"Expense marked {exp.get_status_display()}.")
        return redirect("expense_list")


class ExpenseDeleteView(TreasurerRequiredMixin, View):
    """Delete an expense (treasurer only; kept in the audit log).

    Only a PENDING expense may be hard-deleted — it has no ledger effect yet
    (post_expense() only posts APPROVED/PAID), so removing it is a genuine
    correction with nothing to reverse. Once an expense is APPROVED or PAID it
    is a posted transaction: deleting it would silently erase it from the
    general ledger and every report that has already been generated from it,
    with no trace of what happened (a hard delete rewrites history, unlike a
    reversal which preserves it). Use the Refund/reversal workflow instead,
    which posts an offsetting entry and keeps both sides of the story."""
    def post(self, request, pk):
        exp = get_object_or_404(Expense, pk=pk)
        if _block_if_locked(request, exp.date):
            return redirect("expense_list")
        if exp.status != Expense.Status.PENDING:
            messages.error(request,
                f"This expense is {exp.get_status_display().lower()} — it has "
                "already been posted to the ledger. Record a refund/reversal "
                "instead of deleting it, so the audit trail shows what "
                "happened rather than making it disappear.")
            return redirect(request.META.get("HTTP_REFERER") or "expense_list")
        exp.delete()
        messages.success(request, "Expense deleted.")
        return redirect(request.META.get("HTTP_REFERER") or "expense_list")


# ===================== Inter-fund transfers =====================
class TransferListView(ReadAccessMixin, ListView):
    template_name = "cashbook/transfer_list.html"
    context_object_name = "transfers"
    paginate_by = 50

    def get_queryset(self):
        from django.db.models import Q
        from core.utils import default_to_current_month
        from .models import FundTransfer

        qs = FundTransfer.objects.select_related(
            "source", "destination", "recorded_by").order_by("-date", "-id")

        # This page had no filters at all, and no date bound — it loaded every
        # transfer the church had ever made, on every visit, with no way to
        # narrow it. Same treatment as the ledger and expense lists: default to
        # the current month on a bare visit, and honour anything explicitly
        # asked for (including a deliberate "everything").
        start, end = default_to_current_month(self.request, from_param="start",
                                              to_param="end")
        if start:
            qs = qs.filter(date__gte=start)
        if end:
            qs = qs.filter(date__lte=end)

        fund = self.request.GET.get("fund")
        if fund:
            # either side of the transfer — a treasurer asking "what moved in or
            # out of the Building Fund" does not care which direction it was
            qs = qs.filter(Q(source_id=fund) | Q(destination_id=fund))
        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(Q(reason__icontains=q)
                          | Q(source__name__icontains=q)
                          | Q(destination__name__icontains=q))
        return qs

    def get_context_data(self, **kwargs):
        from django.db.models import Sum
        from core.utils import default_to_current_month
        from departments.models import Department

        ctx = super().get_context_data(**kwargs)
        start, end = default_to_current_month(self.request, from_param="start",
                                              to_param="end")
        filters = self.request.GET.copy()
        filters["start"] = start.isoformat() if start else ""
        filters["end"] = end.isoformat() if end else ""
        ctx["filters"] = filters
        ctx["date_default_applied"] = not self.request.GET
        ctx["funds"] = Department.objects.filter(active=True).order_by("name")
        ctx["filtered_total"] = (
            self.get_queryset().aggregate(t=Sum("amount"))["t"] or 0)
        return ctx


class TransferCreate(DataEntryRequiredMixin, CreateView):
    from .models import FundTransfer
    model = FundTransfer
    form_class = FundTransferForm
    template_name = "cashbook/transfer_form.html"
    success_url = reverse_lazy("transfer_list")

    def form_valid(self, form):
        if _block_if_locked(self.request, form.cleaned_data.get("date")):
            return redirect("transfer_create")
        form.instance.recorded_by = self.request.user
        messages.success(self.request,
                         "Transfer recorded — the balance has moved between the funds.")
        return super().form_valid(form)


class TransferEdit(DataEntryRequiredMixin, UpdateView):
    from .models import FundTransfer
    model = FundTransfer
    form_class = FundTransferForm
    template_name = "cashbook/transfer_form.html"
    success_url = reverse_lazy("transfer_list")

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if self.object.is_locked:
            messages.error(request, "This transfer can't be edited — it has been "
                "reversed, is a reversal entry, or falls in a locked period.")
            return redirect("transfer_list")
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        # block if either the original date or the new date is locked
        if _block_if_locked(self.request, self.object.date) or \
           _block_if_locked(self.request, form.cleaned_data.get("date")):
            return redirect("transfer_edit", pk=self.object.pk)
        messages.success(self.request, "Transfer updated — balances, the journal and "
            "the audit trail were adjusted to match.")
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["editing"] = True
        return ctx


class TransferReverseView(TreasurerRequiredMixin, View):
    def post(self, request, pk):
        from .models import FundTransfer
        tr = get_object_or_404(FundTransfer, pk=pk)
        if _block_if_locked(request, tr.date):
            return redirect("transfer_list")
        try:
            tr.reverse(request.user)
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect("transfer_list")
        messages.success(request, "Transfer reversed — a mirror entry was posted and "
                                  "both remain on record.")
        return redirect("transfer_list")


# ===================== Recurring (scheduled) expenses =====================
class RecurringListView(ReadAccessMixin, ListView):
    template_name = "cashbook/recurring_list.html"
    context_object_name = "schedules"

    def get_queryset(self):
        from .models import RecurringExpense
        return RecurringExpense.objects.select_related("department").all()

    def get_context_data(self, **kwargs):
        from .services.recurring import next_due, upcoming_instalments
        ctx = super().get_context_data(**kwargs)
        # `upcoming` excludes instalments already settled — including any paid
        # ahead of time — so the same period cannot be offered for payment twice.
        ctx["rows"] = [{"s": s,
                        "next": next_due(s) if s.active else None,
                        "upcoming": upcoming_instalments(s, 3) if s.active else []}
                       for s in ctx["schedules"]]
        return ctx


class RecurringCreate(DataEntryRequiredMixin, CreateView):
    from .models import RecurringExpense
    model = RecurringExpense
    form_class = RecurringExpenseForm
    template_name = "cashbook/recurring_form.html"
    success_url = reverse_lazy("recurring_list")

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, "Recurring expense saved. Use “Generate due” to "
                                       "create the entries up to today.")
        return super().form_valid(form)


class RecurringUpdate(DataEntryRequiredMixin, UpdateView):
    from .models import RecurringExpense
    model = RecurringExpense
    form_class = RecurringExpenseForm
    template_name = "cashbook/recurring_form.html"
    success_url = reverse_lazy("recurring_list")


class RecurringToggle(DataEntryRequiredMixin, View):
    def post(self, request, pk):
        from .models import RecurringExpense
        s = get_object_or_404(RecurringExpense, pk=pk)
        s.active = not s.active
        s.save(update_fields=["active"])
        messages.success(request, f"“{s.description}” is now {'active' if s.active else 'paused'}.")
        return redirect("recurring_list")


class RecurringDelete(DataEntryRequiredMixin, View):
    """Delete a recurring schedule. Expenses it already generated are kept (they
    are real ledger entries); only the future schedule is removed."""
    def post(self, request, pk):
        from .models import RecurringExpense
        s = get_object_or_404(RecurringExpense, pk=pk)
        desc = s.description
        s.delete()
        messages.success(request, f"Recurring schedule “{desc}” deleted. "
                                  f"Any expenses it already created remain in the ledger.")
        return redirect("recurring_list")


class RecurringPayEarly(DataEntryRequiredMixin, View):
    """Settle a scheduled instalment before its due date.

    The expense is dated today, because that is when the money leaves and fund
    balances are kept on a cash basis; it records the instalment it settles
    separately, so the schedule will not raise the same charge again when the
    due date comes round.
    """
    def post(self, request, pk):
        from .services.recurring import pay_early
        from .models import RecurringExpense
        import datetime as _dt
        sched = get_object_or_404(RecurringExpense, pk=pk)
        raw = request.POST.get("due_date") or ""
        try:
            due = _dt.date.fromisoformat(raw)
        except ValueError:
            messages.error(request, "Pick which instalment is being paid.")
            return redirect("recurring_list")
        try:
            exp = pay_early(sched, due, on=_dt.date.today(), user=request.user)
        except ValueError as err:
            messages.error(request, str(err))
            return redirect("recurring_list")
        messages.success(
            request,
            f"Paid early: {sched.description} for {due:%d %b %Y}, recorded today. "
            + ("Awaiting approval." if exp.status == Expense.Status.PENDING
               else "Approved."))
        return redirect("recurring_list")


class RecurringGenerate(DataEntryRequiredMixin, View):
    def post(self, request):
        from .services.recurring import generate_due, generate_schedule
        from .models import RecurringExpense
        pk = request.POST.get("schedule")
        if pk:
            s = get_object_or_404(RecurringExpense, pk=pk)
            n = generate_schedule(s, user=request.user)
        else:
            n = generate_due(user=request.user)
        if n:
            messages.success(request, f"Generated {n} expense entr{'y' if n == 1 else 'ies'} "
                                      f"from the recurring schedule(s), up to today.")
        else:
            messages.info(request, "Nothing due — all recurring expenses are already up to date.")
        return redirect("recurring_list")


# ============================ Petty cash ============================
# Petty cash is a CASH LOCATION (a physical float), not a fund. Top-ups put cash
# into the float; disbursements are real expenses charged to the relevant ministry
# fund and flagged as paid from petty cash. The float balance is a control total
# (top-ups less petty disbursements) reconciled against the cash in the box; the
# ministry funds carry the actual cost, so fund balances stay correct.
from django.views.generic import TemplateView
from .forms import PettyCashTopUpForm
from .models import PettyCashTopUp as PettyTopUp


# The canonical implementation now lives in the treasury-position service
# (cashbook/services/treasury_position.py) — see that module's docstring.
# `_petty_balance_asof` is now re-exported from the top of this file.


class PettyCashView(ReadAccessMixin, TemplateView):
    template_name = "cashbook/petty_cash.html"

    def get(self, request, *args, **kwargs):
        if request.GET.get("export") in ("csv", "xlsx"):
            return self._export(request)
        return super().get(request, *args, **kwargs)

    def _export(self, request):
        from core.utils import parse_period
        from reports.exports import csv_response, xlsx_response
        from core.models import SiteConfig
        ctx = self.get_context_data()
        header = ["Date", "Description", "Fund", "In", "Out", "Balance"]
        rows = [["", "Opening balance", "", "", "", ctx["opening"]]]
        for m in ctx["movements"]:
            rows.append([m["date"].isoformat(), m["desc"], m.get("fund") or "",
                         m["in"] or "", m["out"] or "", m["balance"]])
        rows.append(["", "Closing balance", "", "", "", ctx["closing"]])
        start, end = parse_period(request)
        fn = f"petty_cash_{start}_{end}"
        if request.GET.get("export") == "xlsx":
            return xlsx_response(fn + ".xlsx", header, rows,
                title=f"Petty cash register {start} to {end}",
                church=SiteConfig.get().church_name)
        return csv_response(fn + ".csv", header, rows)

    def get_context_data(self, **kwargs):
        from decimal import Decimal
        from core.models import SiteConfig
        from core.utils import parse_period
        ctx = super().get_context_data(**kwargs)
        start, end = parse_period(self.request)
        opening = _petty_balance_asof(start - dt.timedelta(days=1))
        movements = []
        for t in PettyTopUp.objects.filter(date__gte=start, date__lte=end):
            movements.append({"date": t.date, "desc": "Top-up" + (f" — {t.note}" if t.note else ""),
                              "in": t.amount, "out": None, "fund": None})
        disb = (Expense.objects.filter(paid_from_petty_cash=True, date__gte=start, date__lte=end,
                status__in=[Expense.Status.APPROVED, Expense.Status.PAID])
                .select_related("department"))
        for x in disb:
            movements.append({"date": x.date, "desc": x.description
                              + (f" · {x.claimant}" if x.claimant else ""), "in": None,
                              "out": x.amount, "fund": x.department.name, "cat": x.get_category_display()})
        # Cash physically handed back into the tin when an expense is refunded.
        # `petty_balance_asof` — which the "Float on hand (today)" card reads —
        # has always counted these, but the register did not list them, so the
        # two disagreed by exactly the refunds falling inside the period. The
        # opening balance comes from the same helper, so a refund before the
        # period was counted and one during it silently vanished. It is a real
        # movement of cash: whoever counts the box has to see it.
        from .models import ExpenseRefund
        for ref in (ExpenseRefund.objects.filter(to_petty_cash=True, date__gte=start,
                    date__lte=end).select_related("expense", "expense__department")):
            movements.append({"date": ref.date,
                "desc": f"Refund returned to petty cash — {ref.expense.description}"
                        + (f" · {ref.note}" if getattr(ref, "note", "") else ""),
                "in": ref.amount, "out": None,
                "fund": ref.expense.department.name, "cat": "Refund"})
        # petty-cash-funded staff advances: the float drops when issued and rises
        # again only if unspent cash is returned
        from .models import StaffAdvance, AdvanceTopUp
        for adv in StaffAdvance.objects.filter(from_petty_cash=True,
                date_issued__gte=start, date_issued__lte=end).select_related("department"):
            movements.append({"date": adv.date_issued,
                "desc": f"Advance to {adv.staff_name} — {adv.purpose}",
                "in": None, "out": adv.base_amount, "fund": adv.department.name,
                "cat": "Staff advance"})
        # each top-up onto a petty-funded advance is a further outflow on its own date
        for tu in AdvanceTopUp.objects.filter(advance__from_petty_cash=True,
                date__gte=start, date__lte=end).select_related("advance", "advance__department"):
            movements.append({"date": tu.date,
                "desc": f"Advance top-up — {tu.advance.staff_name}"
                        + (f" · {tu.note}" if tu.note else ""),
                "in": None, "out": tu.amount,
                "fund": tu.advance.department.name, "cat": "Advance top-up"})
        for adv in StaffAdvance.objects.filter(from_petty_cash=True,
                returned_to_petty__gt=0, settled_on__gte=start,
                settled_on__lte=end).select_related("department"):
            movements.append({"date": adv.settled_on,
                "desc": f"Unspent advance returned — {adv.staff_name}",
                "in": adv.returned_to_petty, "out": None,
                "fund": adv.department.name, "cat": "Advance return"})
        movements.sort(key=lambda m: m["date"])
        running = opening
        for m in movements:
            running += (m["in"] or Decimal(0)) - (m["out"] or Decimal(0))
            m["balance"] = running
        cfg = SiteConfig.get()
        balance_now = _petty_balance_asof(dt.date.today())

        # The register is computed forward — a running balance has to be — and
        # then shown backward, because the row a treasurer wants is almost
        # always the one they just entered. Reversing AFTER the balances are
        # worked out means each row still carries the balance as it stood on
        # its own date; reversing before would have produced a column that
        # counted down to the opening balance, which is not what it says.
        from django.core.paginator import Paginator
        newest_first = list(reversed(movements))
        paginator = Paginator(newest_first, 50)
        page = paginator.get_page(self.request.GET.get("page"))

        ctx.update({
            "start": start, "end": end, "opening": opening,
            # `movements` stays chronological for the export and for anything
            # that needs to read the register forward.
            "movements": movements,
            "page": page, "rows": page.object_list,
            "closing": running, "balance_now": balance_now,
            "float_target": cfg.petty_cash_float,
            "to_topup": max(cfg.petty_cash_float - balance_now, Decimal(0)) if cfg.petty_cash_float else Decimal(0),
            "topup_form": PettyCashTopUpForm(),
        })
        return ctx


class PettyCashTopUp(DataEntryRequiredMixin, View):
    def post(self, request):
        form = PettyCashTopUpForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            if _block_if_locked(request, cd["date"]):
                return redirect("petty_cash")
            PettyTopUp.objects.create(date=cd["date"], amount=cd["amount"],
                                      note=cd["note"], recorded_by=request.user)
            messages.success(request, f"Petty cash float increased by {cd['amount']}.")
        else:
            messages.error(request, "Could not record the top-up: " + form.errors.as_text())
        return redirect("petty_cash")


class PettyCashDisburse(DataEntryRequiredMixin, View):
    """RETIRED — redirects to the expense form.

    This wrote an ordinary `Expense` with `paid_from_petty_cash=True` — exactly
    what the expense form writes when that box is ticked. Two forms for one row,
    and the lesser one at that: it could not attach a receipt, could not set an
    expenditure type or a budget line, and had its own approval shortcut. A
    treasurer who used it ended up with a voucher that looked different from
    every other voucher in the book.

    Kept as a redirect rather than deleted so an old bookmark still lands
    somewhere useful. The form and its template are gone.
    """

    def get(self, request):
        return redirect(f"{reverse_lazy('expense_create')}?petty=1")

    def post(self, request):
        return redirect(f"{reverse_lazy('expense_create')}?petty=1")


# ==================== Expense detail + receipt attachments ====================
from django.views.generic import DetailView
from .models import ExpenseAttachment


class ExpenseDetailView(ReadAccessMixin, DetailView):
    model = Expense
    template_name = "cashbook/expense_detail.html"
    context_object_name = "expense"

    def get_context_data(self, **kwargs):
        from core.models import SiteConfig
        ctx = super().get_context_data(**kwargs)
        # same source as the list, so the dual-approval gate behaves identically
        ctx["dual_threshold"] = SiteConfig.get().dual_approval_threshold or 0
        ctx["refunds"] = self.object.refunds.select_related("recorded_by")
        ctx["refundable"] = self.object.refundable_balance
        from core import roles
        # gates the "put on the asset register" action (treasurer-only, and the
        # capitalise view enforces it again server-side)
        ctx["can_approve"] = roles.can_approve(self.request.user)
        import datetime as _dt
        ctx["today_iso"] = _dt.date.today().isoformat()
        return ctx


class ExpenseRefundCreate(DataEntryRequiredMixin, View):
    """Record cash returned to the fund against an expense (e.g. unspent change
    from an over-issued purchase). The original expense is never altered."""
    def post(self, request, pk):
        from .models import ExpenseRefund
        from django.core.exceptions import ValidationError
        from decimal import Decimal
        exp = get_object_or_404(Expense, pk=pk)
        if _block_if_locked(request, exp.date):
            return redirect("expense_detail", pk=pk)
        import datetime as _dt
        raw_date = request.POST.get("date") or ""
        try:
            d = _dt.date.fromisoformat(raw_date) if raw_date else _dt.date.today()
        except ValueError:
            d = _dt.date.today()
        try:
            amount = Decimal(request.POST.get("amount") or "0")
        except Exception:  # noqa: BLE001
            amount = Decimal(0)
        ref = ExpenseRefund(expense=exp, date=d, amount=amount,
            method=request.POST.get("method") or Expense.Method.CASH,
            to_petty_cash=bool(request.POST.get("to_petty_cash")),
            reference=request.POST.get("reference", "")[:40],
            note=request.POST.get("note", "")[:200], recorded_by=request.user)
        try:
            ref.full_clean(exclude=["recorded_by"])
        except ValidationError as exc:
            messages.error(request, "; ".join(
                m for msgs in exc.message_dict.values() for m in msgs))
            return redirect("expense_detail", pk=pk)
        ref.save()
        messages.success(request, f"Refund of {amount:,.2f} recorded — the fund "
            f"balance has been restored by this amount.")
        return redirect("expense_detail", pk=pk)


class ExpenseRefundDelete(TreasurerRequiredMixin, View):
    def post(self, request, pk, ref):
        from .models import ExpenseRefund
        r = get_object_or_404(ExpenseRefund, pk=ref, expense_id=pk)
        if _block_if_locked(request, r.date):
            return redirect("expense_detail", pk=pk)
        r.delete()
        messages.success(request, "Refund removed.")
        return redirect("expense_detail", pk=pk)


# The canonical implementations now live in cashbook/services/receipts.py.
# Re-exported here (constants and functions) so every existing call site and
# `from cashbook.views import ...` keeps working unchanged.
from cashbook.services.receipts import (                    # noqa: E402
    RECEIPT_ALLOWED_EXT, RECEIPT_MAX_BYTES, validate_receipt_upload,
    missing_receipts_queryset)


class ExpenseAttachmentUpload(LoginRequiredMixin, View):
    """Attach a supporting document to an expense. Staff may attach to any fund;
    a department leader only to funds they lead. Accepts a file (pdf/image), a
    pasted text receipt (also read from `mpesa_ref`), or a link."""
    ALLOWED_EXT = RECEIPT_ALLOWED_EXT
    MAX_BYTES = RECEIPT_MAX_BYTES

    def _may_attach(self, user, exp):
        from core import roles
        if roles.is_staff_role(user):
            return True
        if roles.is_leader(user):
            from departments.models import departments_led_by
            return exp.department_id in {d.id for d in departments_led_by(user)}
        return False

    def _back(self, request, pk):
        nxt = request.POST.get("next") or request.GET.get("next")
        return redirect(nxt) if nxt else redirect("expense_detail", pk=pk)

    def post(self, request, pk):
        exp = get_object_or_404(Expense, pk=pk)
        if not self._may_attach(request.user, exp):
            messages.error(request, "You can't attach to that expense.")
            return redirect("dashboard")
        f = request.FILES.get("file")
        text = (request.POST.get("text") or request.POST.get("mpesa_ref") or "").strip()
        link = (request.POST.get("link") or "").strip()
        err = validate_receipt_upload(f)
        if err:
            messages.error(request, err)
            return self._back(request, pk)
        if f or text or link:
            ExpenseAttachment.objects.create(expense=exp, file=f or None,
                text=text, link=link,
                label=(request.POST.get("label", "")[:120]
                       or ("M-Pesa message" if text and not f else "")),
                uploaded_by=request.user)
            messages.success(request, f"Supporting document added to expense #{exp.id}.")
        else:
            messages.error(request, "Add a file, paste a text receipt, or enter a link.")
        return self._back(request, pk)


class ExpenseAttachmentDelete(DataEntryRequiredMixin, View):
    def post(self, request, pk, att):
        a = get_object_or_404(ExpenseAttachment, pk=att, expense_id=pk)
        a.file.delete(save=False)
        a.delete()
        messages.success(request, "Attachment removed.")
        return redirect("expense_detail", pk=pk)


# ============== Payables, accruals & prepayments (accrual overlay) ==============
# This app keeps spendable fund balances on a cash basis. Payables, accruals and
# prepayments are an accrual overlay: they are tracked here and shown as memoranda
# on the Statement of Financial Position so the treasurer can see obligations and
# prepaid amounts. Settling a payable/accrual records the actual payment (an
# Expense in the fund), so the cash books recognise it at payment date.
from .models import Payable, Accrual, Prepayment
from .forms import PayableForm, AccrualForm, PrepaymentForm

# (outstanding_advances_total, open_payables_total, open_accruals_total,
# unexpired_prepayments_total et al. are re-exported from the top of this file.)


class AccrualsView(ReadAccessMixin, TemplateView):
    template_name = "cashbook/accruals.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        today = dt.date.today()
        # `supplier` and `payments` are both read for every row — the supplier
        # to link to their account, the payments to show what is still owed —
        # so both are fetched up front rather than one query per bill.
        payables = list(Payable.objects
                        .select_related("department", "supplier")
                        .prefetch_related("payments"))
        accruals = list(Accrual.objects.select_related("department")
                        .prefetch_related("payments"))
        prepayments = list(Prepayment.objects.select_related("department"))
        for p in prepayments:
            p.unexpired_now = p.unexpired(today)
        ctx.update({
            "payables": payables, "accruals": accruals, "prepayments": prepayments,
            # Bills with no supplier record cannot appear on any supplier
            # account, so the page says how many there are rather than leaving
            # the register quietly incomplete.
            "unlinked_payables": sum(1 for p in payables
                                     if not p.supplier_id and not p.settled),
            "open_payables": open_payables_total(), "open_accruals": open_accruals_total(),
            "unexpired_prepay": unexpired_prepayments_total(),
            "payable_form": PayableForm(), "accrual_form": AccrualForm(),
            "prepayment_form": PrepaymentForm(), "today": today,
        })
        return ctx


class PayableCreate(DataEntryRequiredMixin, View):
    def post(self, request):
        form = PayableForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            if _block_if_locked(request, cd["date"]):
                return redirect("accruals")
            Payable.objects.create(date=cd["date"], vendor=cd["vendor"],
                supplier=cd.get("supplier"),
                description=cd["description"], amount=cd["amount"], department=cd["department"],
                category=cd["category"], due_date=cd.get("due_date"), recorded_by=request.user)
            messages.success(request, f"Payable to {cd['vendor']} recorded.")
        else:
            messages.error(request, "Could not record payable: " + form.errors.as_text())
        return redirect("accruals")


class AccrualCreate(DataEntryRequiredMixin, View):
    def post(self, request):
        form = AccrualForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            if _block_if_locked(request, cd["date"]):
                return redirect("accruals")
            Accrual.objects.create(date=cd["date"], description=cd["description"],
                amount=cd["amount"], department=cd["department"], category=cd["category"],
                recorded_by=request.user)
            messages.success(request, "Accrual recorded.")
        else:
            messages.error(request, "Could not record accrual: " + form.errors.as_text())
        return redirect("accruals")


class PrepaymentCreate(DataEntryRequiredMixin, View):
    def post(self, request):
        form = PrepaymentForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            # the cash leaves now: record the payment as a PAID expense in the fund
            exp = Expense.objects.create(date=cd["date"], sabbath_week=sabbath_week_of(cd["date"]),
                department=cd["department"], description=f"Prepayment: {cd['description']}",
                amount=cd["amount"], category=cd["category"], method=Expense.Method.BANK,
                status=Expense.Status.PAID, paid_date=cd["date"], recorded_by=request.user,
                approved_by=request.user)
            Prepayment.objects.create(date=cd["date"], description=cd["description"],
                amount=cd["amount"], department=cd["department"], category=cd["category"],
                months=cd["months"], start_date=cd["start_date"], source_expense=exp,
                recorded_by=request.user)
            messages.success(request, "Prepayment recorded; the unexpired portion is shown as a "
                                      "prepaid asset.")
        else:
            messages.error(request, "Could not record prepayment: " + form.errors.as_text())
        return redirect("accruals")


@db_tx.atomic
def _settle(obj, request):
    """Kept only for callers outside this module. Both payables and accruals
    now settle through services.obligations, which supports instalments; this
    delegates rather than keeping a second, all-or-nothing implementation."""
    from .services import obligations as obligation_svc
    if obj.settled:
        return
    obligation_svc.settle(obj, user=request.user)


class PayableSettle(DataEntryRequiredMixin, View):
    """Pay a payable — all of it, or an instalment.

    An `amount` on the form pays that much and leaves the rest outstanding; no
    amount pays the balance, which is what the single button used to do. The
    work itself is in `services.payables.settle`, so this view and any other
    caller settle a payable exactly one way.
    """

    def post(self, request, pk):
        from .services import payables as payable_svc

        obj = get_object_or_404(Payable, pk=pk)
        raw = (request.POST.get("amount") or "").strip()
        try:
            expense = payable_svc.settle(
                obj, amount=raw or None, user=request.user,
                on=_parse_date(request.POST.get("paid_on")),
                method=request.POST.get("method") or Expense.Method.BANK,
                reference=request.POST.get("reference", ""),   # -> voucher_no
                paid_from_petty_cash=bool(request.POST.get("paid_from_petty_cash")))
        except (ValidationError, InvalidOperation) as exc:
            msg = "; ".join(exc.messages) if hasattr(exc, "messages") \
                else "That is not a valid amount."
            messages.error(request, msg)
            return redirect("accruals")

        obj.refresh_from_db()
        if obj.settled:
            messages.success(
                request,
                f"Payable to {obj.vendor} is now settled in full "
                f"({_currency()}{expense.amount:,.2f} paid, charged to "
                f"{obj.department.name}).")
        else:
            messages.success(
                request,
                f"{_currency()}{expense.amount:,.2f} paid to {obj.vendor}. "
                f"{_currency()}{obj.balance:,.2f} still owing.")
        return redirect("accruals")


class PayableUnlinkPayment(TreasurerRequiredMixin, View):
    """Detach a payment linked to the wrong payable.

    Treasurer-only, and it removes the link rather than the expense: the money
    did leave the bank, and undoing a mis-linking by deleting a real payment
    would fix the paperwork by losing the cash.
    """

    def post(self, request, pk):
        from .services import payables as payable_svc

        expense = get_object_or_404(Expense, pk=pk)
        payable = payable_svc.unlink_payment(expense, user=request.user)
        if payable is None:
            messages.error(request, "That payment is not linked to a payable.")
        else:
            messages.success(
                request,
                f"Payment detached. {payable.vendor} now shows "
                f"{_currency()}{payable.balance:,.2f} outstanding.")
        return redirect("accruals")


class AccrualSettle(DataEntryRequiredMixin, View):
    """Pay an accrual — all of it, or an instalment. Same shape as
    PayableSettle, same service underneath."""

    def post(self, request, pk):
        from .services import obligations as obligation_svc

        obj = get_object_or_404(Accrual, pk=pk)
        try:
            expense = obligation_svc.settle(
                obj, amount=(request.POST.get("amount") or "").strip() or None,
                user=request.user, on=_parse_date(request.POST.get("paid_on")),
                method=request.POST.get("method") or Expense.Method.BANK,
                reference=request.POST.get("reference", ""))
        except (ValidationError, InvalidOperation) as exc:
            messages.error(request, "; ".join(exc.messages)
                           if hasattr(exc, "messages") else "That is not a valid amount.")
            return redirect("accruals")

        obj.refresh_from_db()
        if obj.settled:
            messages.success(request, "Accrual settled in full and recorded as "
                                      "an expense.")
        else:
            messages.success(
                request,
                f"{_currency()}{expense.amount:,.2f} paid. "
                f"{_currency()}{obj.balance:,.2f} still owing.")
        return redirect("accruals")


class AccrualUnlinkPayment(TreasurerRequiredMixin, View):
    """Detach a payment linked to the wrong accrual, keeping the expense."""

    def post(self, request, pk):
        from .services import obligations as obligation_svc

        expense = get_object_or_404(Expense, pk=pk)
        obligation = obligation_svc.unlink_payment(expense, user=request.user)
        if obligation is None:
            messages.error(request, "That payment is not linked to an accrual.")
        else:
            messages.success(request, f"Payment detached. "
                                      f"{_currency()}{obligation.balance:,.2f} outstanding.")
        return redirect("accruals")


# ============================ Staff advances / imprest ============================
class AdvanceListView(ReadAccessMixin, ListView):
    template_name = "cashbook/advance_list.html"
    context_object_name = "advances"
    paginate_by = 40

    def get_queryset(self):
        from .models import StaffAdvance
        return StaffAdvance.objects.select_related("department").all()

    def get_context_data(self, **kwargs):
        from decimal import Decimal
        ctx = super().get_context_data(**kwargs)
        today = dt.date.today()
        bank_out = outstanding_bank_advances_total(today)
        all_out = outstanding_advances_total(today)
        petty_out = max(all_out - bank_out, Decimal(0))
        ctx.update({
            "total_outstanding": all_out,
            "bank_outstanding": bank_out,
            "petty_outstanding": petty_out,
        })
        return ctx


class AdvanceCreate(AdvanceAccessMixin, View):
    template_name = "cashbook/advance_form.html"

    def get(self, request):
        return render(request, self.template_name, {
            "departments": Department.objects.filter(active=True).order_by("name"),
            "petty_balance": _petty_balance_asof(dt.date.today()),
            "today": dt.date.today().isoformat()})

    def post(self, request):
        from decimal import Decimal, InvalidOperation
        from .models import StaffAdvance
        dept = Department.objects.filter(pk=request.POST.get("department")).first()
        try:
            amount = Decimal(request.POST.get("amount") or "0")
        except InvalidOperation:
            amount = Decimal(0)
        name = (request.POST.get("staff_name") or "").strip()
        if not (dept and name and amount > 0):
            messages.error(request, "Staff name, fund and a positive amount are required.")
            return redirect("advance_new")
        try:
            issued = dt.date.fromisoformat(request.POST.get("date_issued"))
        except (TypeError, ValueError):
            issued = dt.date.today()
        if _block_if_locked(request, issued):
            return redirect("advance_new")
        from_petty = request.POST.get("from_petty_cash") in ("1", "on", "true")
        method = request.POST.get("method") or "CASH"
        try:
            charge = Decimal(request.POST.get("bank_charge") or "0")
        except InvalidOperation:
            charge = Decimal(0)
        if charge < 0:
            charge = Decimal(0)
        if from_petty:
            method = "CASH"   # petty cash is, by definition, cash
            avail = _petty_balance_asof(issued)
            if amount > avail:
                messages.error(request,
                    f"The petty-cash float is only KSh {avail:,.2f} on {issued:%d %b %Y}; "
                    f"top it up before issuing an advance of KSh {amount:,.2f}.")
                return redirect("advance_new")
        adv = StaffAdvance.objects.create(
            staff_name=name, department=dept, amount=amount, date_issued=issued,
            purpose=(request.POST.get("purpose") or "")[:200],
            method=method, from_petty_cash=from_petty, bank_charge=charge,
            reference=(request.POST.get("reference") or "")[:40],
            issued_by=request.user)
        _sync_advance_charge(adv, request.user)
        messages.success(request, "Advance recorded." + (
            " Petty-cash float reduced accordingly." if from_petty else "")
            + (f" Bank/M-Pesa charge of KSh {charge:,.2f} recorded." if charge else ""))
        return redirect("advance_detail", pk=adv.pk)


def apply_advance_edit(adv, post, user, *, allow_petty_toggle=True):
    """Apply an edit to an advance from POST data and keep everything in step
    (petty float recomputes from the live fields; the bank-charge expense is
    re-synced). Returns (ok, error_message)."""
    from decimal import Decimal, InvalidOperation
    from .models import StaffAdvance
    name = (post.get("staff_name") or adv.staff_name).strip()
    dept = Department.objects.filter(pk=post.get("department")).first() or adv.department
    try:
        amount = Decimal(post.get("amount") or adv.amount)
    except InvalidOperation:
        amount = adv.amount
    try:
        charge = Decimal(post.get("bank_charge") or "0")
    except InvalidOperation:
        charge = Decimal(0)
    if amount <= 0:
        return False, "Amount must be positive."
    if charge < 0:
        charge = Decimal(0)
    try:
        issued = dt.date.fromisoformat(post.get("date_issued"))
    except (TypeError, ValueError):
        issued = adv.date_issued
    from_petty = adv.from_petty_cash
    if allow_petty_toggle:
        from_petty = post.get("from_petty_cash") in ("1", "on", "true")
    method = "CASH" if from_petty else (post.get("method") or adv.method)
    # validate petty float can still cover a petty advance (excluding this advance's
    # own current contribution so an unchanged edit doesn't false-trip)
    if from_petty:
        avail = _petty_balance_asof(issued) + adv.petty_outstanding_asof(issued)
        if amount - adv.settled_asof(issued) - (adv.returned_to_petty or Decimal(0)) > avail:
            return False, (f"The petty-cash float can't cover that amount on "
                           f"{issued:%d %b %Y}.")
    adv.staff_name = name[:120]
    adv.department = dept
    adv.amount = amount
    adv.date_issued = issued
    adv.method = method
    adv.from_petty_cash = from_petty
    adv.bank_charge = charge
    adv.purpose = (post.get("purpose") or adv.purpose)[:200]
    adv.reference = (post.get("reference") or adv.reference)[:40]
    adv.save()
    _sync_advance_charge(adv, user)
    # refresh status against the new amount
    bal = adv.balance
    if bal == 0 and adv.settled_total > 0:
        adv.status = StaffAdvance.Status.SETTLED
    elif adv.settled_total > 0:
        adv.status = StaffAdvance.Status.PARTLY
    elif adv.status in (StaffAdvance.Status.SETTLED, StaffAdvance.Status.PARTLY):
        adv.status = StaffAdvance.Status.ISSUED
    adv.save(update_fields=["status"])
    return True, None


class AdvanceEdit(AdvanceAccessMixin, View):
    """Edit an advance. Treasurers/assistants may edit any advance; a closed
    advance can only be amended by a treasurer (enforced below)."""
    template_name = "cashbook/advance_form.html"

    def _get(self, pk):
        from .models import StaffAdvance
        return get_object_or_404(StaffAdvance, pk=pk)

    def get(self, request, pk):
        adv = self._get(pk)
        return render(request, self.template_name, {
            "adv": adv, "editing": True,
            "departments": Department.objects.filter(active=True).order_by("name"),
            "petty_balance": _petty_balance_asof(dt.date.today()),
            "today": adv.date_issued.isoformat()})

    def post(self, request, pk):
        from core import roles
        adv = self._get(pk)
        if adv.status == "CLOSED" and not roles.is_treasurer(request.user):
            messages.error(request, "A closed advance can only be amended by a treasurer.")
            return redirect("advance_detail", pk=pk)
        if _block_if_locked(request, adv.date_issued):
            return redirect("advance_detail", pk=pk)
        ok, err = apply_advance_edit(adv, request.POST, request.user)
        if not ok:
            messages.error(request, err)
            return redirect("advance_edit", pk=pk)
        messages.success(request, "Advance updated.")
        return redirect("advance_detail", pk=pk)


class AdvanceDelete(TreasurerRequiredMixin, View):
    """Delete an advance end-to-end: its settling expenses and bank-charge expense
    go with it, and the petty float recomputes automatically."""
    def post(self, request, pk):
        from .models import StaffAdvance
        adv = get_object_or_404(StaffAdvance, pk=pk)
        if _block_if_locked(request, adv.date_issued):
            return redirect("advance_detail", pk=pk)
        # an advance can only be deleted if nothing has been accounted against it
        if adv.expenses.exists():
            messages.error(request, "This advance has expenses recorded against it. "
                "Remove those expenses first, then delete the advance.")
            return redirect("advance_detail", pk=pk)
        # detach and remove the church's sending-charge expense (not an accounting
        # line against the advance), then delete the advance
        if adv.charge_expense_id:
            ce = adv.charge_expense
            adv.charge_expense = None
            adv.save(update_fields=["charge_expense"])
            ce.delete()
        adv.delete()
        messages.success(request, "Advance deleted.")
        return redirect("advance_list")


class AdvanceDetail(ReadAccessMixin, View):
    template_name = "cashbook/advance_detail.html"

    def get(self, request, pk):
        from .models import StaffAdvance
        adv = get_object_or_404(StaffAdvance, pk=pk)
        return render(request, self.template_name, _advance_detail_ctx(adv, user=request.user))


def _sync_advance_charge(adv, user):
    """Create / update / remove the BANK_CHARGE expense for the charge the CHURCH
    incurs when sending an advance to the holder (adv.bank_charge). This is the
    church's cost, so it is booked against the fund but does NOT reduce the advance
    the holder must account for — it is not linked via the `advance` FK. The
    charge_expense one-to-one keeps a handle on it for edits/removal."""
    from decimal import Decimal
    from .models import Expense
    charge = adv.bank_charge or Decimal(0)
    exp = adv.charge_expense
    if charge and charge > 0:
        if exp:
            exp.date = adv.date_issued
            exp.amount = charge
            exp.department = adv.department
            exp.method = adv.method
            exp.claimant = adv.staff_name
            exp.paid_from_petty_cash = adv.from_petty_cash
            exp.description = f"Bank/M-Pesa charge — sending advance to {adv.staff_name}"
            exp.save(update_fields=["date", "amount", "department", "method",
                                    "claimant", "paid_from_petty_cash", "description"])
        else:
            exp = Expense.objects.create(
                date=adv.date_issued, sabbath_week=sabbath_week_of(adv.date_issued),
                department=adv.department,
                description=f"Bank/M-Pesa charge — sending advance to {adv.staff_name}",
                amount=charge, category=Expense.Category.BANK_CHARGE,
                claimant=adv.staff_name, method=adv.method,
                paid_from_petty_cash=adv.from_petty_cash,
                status=Expense.Status.PAID, paid_date=adv.date_issued,
                recorded_by=user, approved_by=user)
            adv.charge_expense = exp
            adv.save(update_fields=["charge_expense"])
    elif exp:
        # charge removed — delete the linked expense
        adv.charge_expense = None
        adv.save(update_fields=["charge_expense"])
        exp.delete()


def _advance_import_sample():
    """A sample .xlsx for importing expenses against a staff advance."""
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Advance expenses"
    head = ["Date", "Description", "Category", "Amount", "Charge"]
    ws.append(head)
    for c in range(1, len(head) + 1):
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, c).fill = PatternFill("solid", fgColor="1F5F4F")
    ws.append(["2026-06-06", "Bus fare to venue", "Transport", 800, 0])
    ws.append(["2026-06-06", "Refreshments", "Refreshments / catering", 1500, 30])
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 24
    ref = wb.create_sheet("Lists")
    ref["A1"] = "Category"; ref["A1"].font = Font(bold=True)
    cats = [c.label for c in Expense.Category]
    for i, c in enumerate(cats, start=2):
        ref.cell(i, 1, c)
    dv = DataValidation(type="list", formula1=f"=Lists!$A$2:$A${len(cats)+1}",
                        allow_blank=True)
    ws.add_data_validation(dv); dv.add("C2:C500")
    info = wb.create_sheet("How to fill this in")
    for i, line in enumerate([
        "Import expenses onto a staff advance",
        "",
        "One row per expense the advance holder spent.",
        "  - Date  — YYYY-MM-DD (required).",
        "  - Description — what it was for (required).",
        "  - Category — pick from the list (optional; defaults to Other).",
        "  - Amount — the amount spent (required).",
        "  - Charge — any M-Pesa/bank fee paid on that payment (optional).",
        "",
        "The total of Amount + Charge across all rows cannot exceed the",
        "advance's remaining balance. If it does, nothing is imported.",
    ], start=1):
        info.cell(i, 1, line)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


class AdvanceExpenseImportView(View):
    """Import a spreadsheet of expenses onto one staff advance. Available to the
    treasurer/assistant and to the department leader who owns the advance. The
    combined total (amount + charge) may not exceed the advance's remaining
    balance — an over-budget file is rejected in full."""
    template_name = "cashbook/advance_import.html"

    def _get_advance(self, request, pk):
        from .models import StaffAdvance
        from core import roles
        adv = get_object_or_404(StaffAdvance, pk=pk)
        user = request.user
        if roles.can_enter_data(user):
            return adv, False
        # a leader may import onto advances in their departments
        if roles.is_leader(user):
            from leaders.permissions import assert_department_allowed
            if assert_department_allowed(user, adv.department_id):
                return adv, True
        return None, False

    def get(self, request, pk):
        from django.http import HttpResponse
        if request.GET.get("download"):
            buf = _advance_import_sample()
            resp = HttpResponse(buf.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            resp["Content-Disposition"] = 'attachment; filename="advance_expenses_template.xlsx"'
            return resp
        adv, is_leader = self._get_advance(request, pk)
        if not adv:
            return redirect("leader_dashboard" if request.user else "dashboard")
        return render(request, self.template_name,
                      {"adv": adv, "is_leader_view": is_leader})

    def post(self, request, pk):
        import openpyxl, datetime as _dt
        from decimal import Decimal, InvalidOperation
        from core.utils import block_if_locked
        adv, is_leader = self._get_advance(request, pk)
        if not adv:
            return redirect("dashboard")
        back = ("leader_advance_detail" if is_leader else "advance_detail")
        if adv.status == adv.Status.CLOSED:
            messages.error(request, "This advance is closed.")
            return redirect(back, pk=pk)
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "Choose a spreadsheet to upload.")
            return redirect("advance_import", pk=pk)
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception:  # noqa: BLE001
            messages.error(request, "Could not read that file — please upload a .xlsx.")
            return redirect("advance_import", pk=pk)
        ws = (wb["Advance expenses"] if "Advance expenses" in wb.sheetnames
              else wb.active)
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            messages.error(request, "The sheet is empty.")
            return redirect("advance_import", pk=pk)
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

        def col(*names):
            for n in names:
                if n in header:
                    return header.index(n)
            return None
        c_date, c_desc = col("date"), col("description", "details", "expense")
        c_amt, c_cat = col("amount", "value"), col("category")
        c_charge = col("charge", "m-pesa charge", "transaction charge")
        if c_date is None or c_desc is None or c_amt is None:
            messages.error(request, "Need at least Date, Description and Amount columns "
                                    "— please use the template.")
            return redirect("advance_import", pk=pk)

        cat_labels = {c.label.upper(): c.value for c in Expense.Category}
        cat_labels.update({c.value: c.value for c in Expense.Category})

        def pdate(v):
            if isinstance(v, _dt.datetime):
                return v.date()
            if isinstance(v, _dt.date):
                return v
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
                try:
                    return _dt.datetime.strptime(str(v).strip(), fmt).date()
                except (ValueError, TypeError):
                    continue
            return None

        def num(v):
            try:
                return Decimal(str(v)) if v not in (None, "") else Decimal(0)
            except (InvalidOperation, TypeError):
                return Decimal(0)

        plan, errors, total = [], [], Decimal(0)
        for i, r in enumerate(rows[1:], start=2):
            desc = (str(r[c_desc]).strip() if c_desc < len(r) and r[c_desc] else "")
            d = pdate(r[c_date]) if c_date < len(r) else None
            amt = num(r[c_amt]) if c_amt < len(r) else Decimal(0)
            charge = num(r[c_charge]) if c_charge is not None and c_charge < len(r) else Decimal(0)
            if not desc and not d and amt <= 0:
                continue
            if not (desc and d and amt > 0):
                errors.append(f"Row {i}: needs a date, description and positive amount.")
                continue
            cat_raw = (str(r[c_cat]).strip().upper() if c_cat is not None
                       and c_cat < len(r) and r[c_cat] else "")
            category = cat_labels.get(cat_raw, Expense.Category.OTHER)
            plan.append({"date": d, "desc": desc, "amount": amt,
                         "charge": charge, "category": category})
            total += amt + charge

        if not plan:
            messages.error(request, "No valid rows found. " + " ".join(errors[:3]))
            return redirect("advance_import", pk=pk)

        # the whole-batch balance cap: reject the import entirely if it exceeds
        if total > adv.balance:
            messages.error(request,
                f"This file accounts for KSh {total:,.2f}, but only "
                f"KSh {adv.balance:,.2f} is left on the advance. Nothing was "
                "imported — reduce the entries and try again.")
            return redirect("advance_import", pk=pk)

        claimant = adv.staff_name
        if is_leader:
            claimant = request.user.get_full_name() or request.user.username
        created = 0
        for p in plan:
            if block_if_locked(request, p["date"]):
                continue
            exp, err = _record_advance_expense(
                adv, date=p["date"], desc=p["desc"], amount=p["amount"],
                category=p["category"], user=request.user, claimant=claimant,
                charge=p["charge"])
            if exp:
                created += 1
        msg = f"{created} expense(s) imported onto the advance."
        if errors:
            msg += f" {len(errors)} row(s) skipped."
        messages.success(request, msg)
        return redirect(back, pk=pk)


class AdvanceAddExpense(AdvanceAccessMixin, View):
    """Record a receipt/expense that settles part of an advance."""
    def post(self, request, pk):
        from decimal import Decimal, InvalidOperation
        from .models import StaffAdvance
        adv = get_object_or_404(StaffAdvance, pk=pk)
        try:
            amount = Decimal(request.POST.get("amount") or "0")
        except InvalidOperation:
            amount = Decimal(0)
        try:
            charge = Decimal(request.POST.get("charge") or "0")
        except InvalidOperation:
            charge = Decimal(0)
        if charge < 0:
            charge = Decimal(0)
        desc = (request.POST.get("description") or "").strip()
        if not (desc and amount > 0):
            messages.error(request, "A description and positive amount are required.")
            return redirect("advance_detail", pk=pk)
        try:
            d = dt.date.fromisoformat(request.POST.get("date"))
        except (TypeError, ValueError):
            d = dt.date.today()
        _exp, err = _record_advance_expense(adv, date=d, desc=desc, amount=amount,
            category=request.POST.get("category"), user=request.user, charge=charge)
        if err:
            messages.error(request, err)
        else:
            messages.success(request, "Expense recorded against the advance."
                + (f" Transaction charge of KSh {charge:,.2f} added." if charge else ""))
        return redirect("advance_detail", pk=pk)


class AdvanceTopUpView(AdvanceAccessMixin, View):
    """Issue more cash onto an open advance (carry a small leftover forward into a
    larger working advance instead of retiring and re-issuing). An optional
    bank/M-Pesa charge for sending the top-up is the church's own cost — booked
    as an expense against the fund but not added to what the holder must
    account for (mirrors how the charge on the original issue is handled)."""
    def post(self, request, pk):
        from decimal import Decimal, InvalidOperation
        from .models import StaffAdvance, AdvanceTopUp
        adv = get_object_or_404(StaffAdvance, pk=pk)
        if adv.status == StaffAdvance.Status.CLOSED:
            messages.error(request, "This advance is closed; issue a new one instead.")
            return redirect("advance_detail", pk=pk)
        try:
            amount = Decimal(request.POST.get("amount") or "0")
        except InvalidOperation:
            amount = Decimal(0)
        try:
            charge = Decimal(request.POST.get("charge") or "0")
        except InvalidOperation:
            charge = Decimal(0)
        if amount <= 0:
            messages.error(request, "Enter a positive top-up amount.")
            return redirect("advance_detail", pk=pk)
        try:
            d = dt.date.fromisoformat(request.POST.get("date"))
        except (TypeError, ValueError):
            d = dt.date.today()
        if _block_if_locked(request, d):
            return redirect("advance_detail", pk=pk)
        if adv.from_petty_cash:
            avail = _petty_balance_asof(d)
            if amount + charge > avail:
                messages.error(request, f"The petty-cash float is only KSh {avail:,.2f} "
                    f"on {d:%d %b %Y}; top it up before issuing more.")
                return redirect("advance_detail", pk=pk)
        tu = AdvanceTopUp.objects.create(advance=adv, date=d, amount=amount,
            charge=charge, note=(request.POST.get("note") or "")[:200],
            issued_by=request.user)
        adv.amount = (adv.amount or Decimal(0)) + amount
        if adv.status == StaffAdvance.Status.SETTLED:
            adv.status = StaffAdvance.Status.PARTLY   # reopened by fresh cash
        adv.save(update_fields=["amount", "status"])
        if charge > 0:
            _sync_topup_charge(tu, adv, request.user)
        msg = f"Topped up by KSh {amount:,.2f}. New advance total KSh {adv.amount:,.2f}."
        if charge > 0:
            msg += f" A KSh {charge:,.2f} sending charge was booked against the fund."
        messages.success(request, msg)
        return redirect("advance_detail", pk=pk)


def _sync_topup_charge(tu, adv, user):
    """Create the BANK_CHARGE expense for the charge incurred sending a top-up —
    the church's cost, booked against the fund but not linked via `advance` (so
    it never reduces what the holder must account for)."""
    from .models import Expense
    exp = Expense.objects.create(
        date=tu.date, sabbath_week=sabbath_week_of(tu.date), department=adv.department,
        description=f"Bank/M-Pesa charge — topping up advance for {adv.staff_name}",
        amount=tu.charge, category=Expense.Category.BANK_CHARGE,
        claimant=adv.staff_name, method=adv.method,
        paid_from_petty_cash=adv.from_petty_cash,
        status=Expense.Status.PAID, paid_date=tu.date,
        recorded_by=user, approved_by=user)
    tu.charge_expense = exp
    tu.save(update_fields=["charge_expense"])


class AdvanceTopUpReverseView(TreasurerRequiredMixin, View):
    """Reverse a top-up on an advance: remove the top-up, decrement the advance
    total, and restore the source (the petty-cash float recovers automatically
    once the top-up no longer counts as outstanding)."""
    def post(self, request, pk, topup_id):
        from decimal import Decimal
        from .models import StaffAdvance, AdvanceTopUp
        adv = get_object_or_404(StaffAdvance, pk=pk)
        tu = get_object_or_404(AdvanceTopUp, pk=topup_id, advance=adv)
        if _block_if_locked(request, tu.date):
            return redirect("advance_detail", pk=pk)
        amount = tu.amount
        adv.amount = max((adv.amount or Decimal(0)) - amount, Decimal(0))
        adv.save(update_fields=["amount"])
        charge_note = ""
        if tu.charge_expense_id:
            tu.charge_expense.delete()
            charge_note = " Its sending charge was removed too."
        tu.delete()
        messages.success(request, f"Top-up of KSh {amount:,.2f} reversed. "
            f"New advance total KSh {adv.amount:,.2f}"
            + (" — the petty-cash float has been restored." if adv.from_petty_cash
               else ".") + charge_note)
        return redirect("advance_detail", pk=pk)


class AdvanceClose(TreasurerRequiredMixin, View):
    def post(self, request, pk):
        from decimal import Decimal, InvalidOperation
        from .models import StaffAdvance
        adv = get_object_or_404(StaffAdvance, pk=pk)
        if adv.from_petty_cash:
            try:
                ret = Decimal(request.POST.get("returned_to_petty") or "0")
            except InvalidOperation:
                ret = Decimal(0)
            if ret > 0:
                adv.returned_to_petty = ret
        adv.status = StaffAdvance.Status.CLOSED
        adv.settled_on = dt.date.today()
        adv.note = (request.POST.get("note") or adv.note)[:1000]
        adv.save(update_fields=["status", "settled_on", "note", "returned_to_petty"])
        messages.success(request, "Advance closed." + (
            " Returned cash credited back to petty cash." if adv.from_petty_cash
            and adv.returned_to_petty else ""))
        return redirect("advance_detail", pk=pk)


class ExpenseCategoryList(TreasurerRequiredMixin, View):
    template_name = "cashbook/categories.html"

    def get(self, request):
        from .models import ExpenseCategory, Expense
        return render(request, self.template_name, {
            "builtin": Expense.Category.choices,
            "custom": ExpenseCategory.objects.all()})

    def post(self, request):
        from .models import ExpenseCategory
        action = request.POST.get("action")
        if action == "add":
            label = (request.POST.get("label") or "").strip()
            code = (request.POST.get("code") or "").strip().upper().replace(" ", "_")[:20]
            is_liab = bool(request.POST.get("is_liability"))
            if label and code:
                ExpenseCategory.objects.get_or_create(
                    code=code, defaults={"label": label, "is_liability": is_liab})
                messages.success(request, f"Added category “{label}”.")
            else:
                messages.error(request, "Both a code and a label are required.")
        elif action == "toggle":
            ec = ExpenseCategory.objects.filter(pk=request.POST.get("id")).first()
            if ec:
                ec.active = not ec.active
                ec.save(update_fields=["active"])
                messages.success(request, f"“{ec.label}” is now {'active' if ec.active else 'inactive'}.")
        elif action == "toggle_liability":
            ec = ExpenseCategory.objects.filter(pk=request.POST.get("id")).first()
            if ec:
                ec.is_liability = not ec.is_liability
                ec.save(update_fields=["is_liability"])
                # refile existing vouchers in this category (classification
                # only — accounting entries and audit history untouched)
                from .models import Expense as _E
                new_class = (_E.DocClass.LIABILITY if ec.is_liability
                             else _E.DocClass.EXPENSE)
                n = _E.objects.filter(category=ec.code).update(doc_class=new_class)
                messages.success(request,
                    f"“{ec.label}” is now a {'liability' if ec.is_liability else 'normal expense'} "
                    f"category ({n} existing voucher{'s' if n != 1 else ''} refiled).")
        return redirect("expense_categories")


class ExpenseRecategorizeView(TreasurerRequiredMixin, View):
    """Download all expenses as a spreadsheet for offline editing, then re-import
    to update the category and/or the expenditure type (capital vs recurrent).
    The round-trip is keyed on the expense ID; amounts, dates, descriptions and
    other fields in the file are ignored, so a treasurer can fix many
    mis-categorised expenses quickly without risking the rest of the ledger."""
    template_name = "cashbook/recategorize.html"

    def get(self, request):
        if request.GET.get("download"):
            return self._download(request)
        # show the upload form + the valid category list
        return render(request, self.template_name, {
            "categories": Expense.Category.choices})

    def _download(self, request):
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from django.http import HttpResponse
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Expenses"
        head = ["ID", "Date", "Description", "Fund", "Amount",
                "Current category", "New category (edit this)",
                "Current type", "New type (capital/recurrent)"]
        ws.append(head)
        for c in range(1, len(head) + 1):
            ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
            ws.cell(1, c).fill = PatternFill("solid", fgColor="1F5F4F")
        # a reference sheet of valid category codes
        ref = wb.create_sheet("Valid categories")
        ref.append(["Code", "Label"])
        ref.cell(1, 1).font = Font(bold=True); ref.cell(1, 2).font = Font(bold=True)
        for code, label in Expense.Category.choices:
            ref.append([code, label])
        ref.append([])
        ref.append(["Expenditure type", ""])
        for code, label in Expense.ExpenditureType.choices:
            ref.append([code, label])
        for x in (Expense.objects.select_related("department")
                  .order_by("-date", "-id")):
            ws.append([x.id, x.date.isoformat(), x.description,
                       x.department.name if x.department_id else "",
                       float(x.amount), x.get_category_display(),
                       x.get_category_display(),
                       x.get_expenditure_type_display() if x.expenditure_type else "",
                       x.get_expenditure_type_display() if x.expenditure_type else ""])
        ws.column_dimensions["C"].width = 34
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 22
        ws.column_dimensions["I"].width = 24
        buf = io.BytesIO(); wb.save(buf)
        resp = HttpResponse(buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = 'attachment; filename="expenses_to_recategorize.xlsx"'
        return resp

    @db_tx.atomic
    def post(self, request):
        import openpyxl
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "Choose the edited spreadsheet to upload.")
            return redirect("expense_recategorize")
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception:
            from core.utils import log_exception as _lx; _lx('cashbook/views.py')
            messages.error(request, "Could not read that file — is it the .xlsx you downloaded?")
            return redirect("expense_recategorize")
        ws = wb["Expenses"] if "Expenses" in wb.sheetnames else wb.active

        # build label -> code and code -> code lookups so the file can carry
        # either the human label ("Transport") or the raw code ("TRANSPORT")
        by_label = {lbl.lower(): code for code, lbl in Expense.Category.choices}
        by_code = {code.upper(): code for code, _ in Expense.Category.choices}
        type_by_label = {lbl.lower(): code for code, lbl in Expense.ExpenditureType.choices}
        type_by_code = {code.upper(): code for code, _ in Expense.ExpenditureType.choices}

        header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        try:
            id_col = header.index("id")
            new_col = header.index("new category (edit this)")
        except ValueError:
            messages.error(request, "That doesn't look like the recategorise template "
                                    "(missing the ID or New category column).")
            return redirect("expense_recategorize")
        # optional type column (older templates won't have it)
        type_col = next((i for i, h in enumerate(header)
                         if h.startswith("new type")), None)

        updated = unchanged = bad = missing = typed = 0
        bad_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if id_col >= len(row) or row[id_col] in (None, ""):
                continue
            try:
                eid = int(row[id_col])
            except (TypeError, ValueError):
                continue
            exp = Expense.objects.filter(pk=eid).first()
            if not exp:
                missing += 1
                continue
            dirty = []
            # category
            raw = row[new_col] if new_col < len(row) else None
            if raw not in (None, ""):
                key = str(raw).strip()
                code = by_code.get(key.upper()) or by_label.get(key.lower())
                if not code:
                    bad += 1
                    if len(bad_rows) < 10:
                        bad_rows.append(f"#{eid}: “{key}”")
                elif exp.category != code:
                    exp.category = code
                    dirty.append("category")
            # expenditure type (capital / recurrent)
            if type_col is not None and type_col < len(row):
                traw = row[type_col]
                if traw not in (None, ""):
                    tkey = str(traw).strip()
                    tcode = type_by_code.get(tkey.upper()) or type_by_label.get(tkey.lower())
                    if not tcode:
                        bad += 1
                        if len(bad_rows) < 10:
                            bad_rows.append(f"#{eid}: type “{tkey}”")
                    elif exp.expenditure_type != tcode:
                        exp.expenditure_type = tcode
                        dirty.append("expenditure_type")
            if dirty:
                exp.save(update_fields=dirty)
                updated += 1 if "category" in dirty else 0
                typed += 1 if "expenditure_type" in dirty else 0
            else:
                unchanged += 1

        parts = []
        if updated:
            parts.append(f"{updated} re-categorised")
        if typed:
            parts.append(f"{typed} type changed")
        if not parts:
            parts.append("0 changed")
        if unchanged:
            parts.append(f"{unchanged} unchanged")
        if missing:
            parts.append(f"{missing} not found")
        if bad:
            parts.append(f"{bad} with an unrecognised value")
        msg = ", ".join(parts) + "."
        if bad_rows:
            msg += " Unrecognised: " + "; ".join(bad_rows)
        (messages.success if (updated or typed) else messages.warning)(request, msg)
        return redirect("expense_list")


# ===========================================================================
# Bulk expense import (item 5)
# ===========================================================================
class ExpenseImportView(DataEntryRequiredMixin, View):
    """Bulk-load expenses from a spreadsheet: date, fund, description, amount,
    category, method, claimant, voucher. Honours the approval setting (auto-approve
    when approval isn't required, otherwise lands as pending)."""
    template_name = "cashbook/import.html"

    def _scoped_departments(self, user):
        from departments.models import Department
        return Department.objects.all()

    CATEGORY_LABELS = {c.label.upper(): c.value for c in Expense.Category}
    CATEGORY_LABELS.update({c.value: c.value for c in Expense.Category})
    # friendly aliases
    CATEGORY_LABELS.update({
        "ALLOWANCE": "ALLOWANCE", "HONORARIA": "ALLOWANCE", "TRANSPORT": "TRANSPORT",
        "FARE": "TRANSPORT", "REFRESHMENTS": "REFRESHMENTS", "CATERING": "REFRESHMENTS",
        "FOOD": "REFRESHMENTS", "MATERIALS": "MATERIALS", "SUPPLIES": "MATERIALS",
        "STATIONERY": "STATIONERY", "PRINTING": "STATIONERY", "UTILITIES": "UTILITIES",
        "POWER": "UTILITIES", "WATER": "UTILITIES", "MAINTENANCE": "MAINTENANCE",
        "REPAIRS": "MAINTENANCE", "CONSTRUCTION": "CONSTRUCTION", "EVANGELISM": "EVANGELISM",
        "MISSION": "EVANGELISM", "BENEVOLENCE": "BENEVOLENCE", "WELFARE": "BENEVOLENCE",
        "BANK CHARGE": "BANK_CHARGE", "BANK CHARGES": "BANK_CHARGE", "CHARGES": "BANK_CHARGE",
        "REMITTANCE": "REMITTANCE", "OTHER": "OTHER",
    })
    METHOD_LABELS = {"CASH": "CASH", "BANK": "BANK", "CHEQUE": "CHEQUE", "CHECK": "CHEQUE",
                     "MPESA": "MPESA", "M-PESA": "MPESA", "MOBILE": "MPESA"}

    def get(self, request):
        if request.GET.get("download"):
            return self._download()
        return render(request, self.template_name, {"stage": "upload"})

    def post(self, request):
        if request.POST.get("apply"):
            return self._apply(request)
        return self._parse(request)

    def _download(self):
        import io, openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
        from django.http import HttpResponse
        from departments.models import Department
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Expenses"
        # Every field the expense form offers, so the two entry paths produce
        # the same record. Before v3.18.0 the upload was missing Supplier,
        # Payee, Expenditure type and Budget item, which meant a spreadsheet
        # silently produced the incomplete rows the form no longer allows.
        head = ["Date", "Fund", "Description", "Amount", "Category", "Method",
                "Claimant", "Supplier", "Payee", "Voucher no", "Expenditure type",
                "Budget item", "M-Pesa charge", "Paid from petty cash"]
        ws.append(head)
        for c in range(1, len(head) + 1):
            ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
            ws.cell(1, c).fill = PatternFill("solid", fgColor="1F5F4F")
        ws.append(["2026-06-06", "LCB", "Pulpit microphone", 4500, "Materials",
                   "Cash", "J. Mwangi", "", "", "V-001", "Recurrent", "", "", ""])
        ws.append(["2026-06-07", "YOUTH", "Bus fare for rally", 2000, "Transport",
                   "M-Pesa", "S. Achieng", "", "", "V-002", "Recurrent", "", 30, "Yes"])
        ref = wb.create_sheet("Lists")
        ref["A1"] = "Funds"; ref["A1"].font = Font(bold=True)
        funds = list(Department.objects.filter(active=True, selectable=True).order_by("name"))
        for i, d in enumerate(funds, start=2):
            ref.cell(i, 1, d.name)
        ref["B1"] = "Category"; ref["B1"].font = Font(bold=True)
        cats = [c.label for c in Expense.Category]
        for i, c in enumerate(cats, start=2):
            ref.cell(i, 2, c)
        ref["C1"] = "Method"; ref["C1"].font = Font(bold=True)
        for i, m in enumerate([mm.label for mm in Expense.Method], start=2):
            ref.cell(i, 3, m)
        # The supplier register, so whoever fills this in picks a name that
        # already exists instead of typing a fourth spelling of it.
        from vendors.models import Vendor
        ref["D1"] = "Supplier"; ref["D1"].font = Font(bold=True)
        suppliers = list(Vendor.objects.exclude(status=Vendor.Status.ARCHIVED)
                         .order_by("name").values_list("name", flat=True))
        for i, name in enumerate(suppliers, start=2):
            ref.cell(i, 4, name)
        ref["E1"] = "Expenditure type"; ref["E1"].font = Font(bold=True)
        etypes = [t.label for t in Expense.ExpenditureType]
        for i, t in enumerate(etypes, start=2):
            ref.cell(i, 5, t)
        nrows = 500
        if funds:
            dv = DataValidation(type="list", formula1=f"=Lists!$A$2:$A${len(funds)+1}", allow_blank=True)
            ws.add_data_validation(dv); dv.add(f"B2:B{nrows}")
        dvc = DataValidation(type="list", formula1=f"=Lists!$B$2:$B${len(cats)+1}", allow_blank=True)
        ws.add_data_validation(dvc); dvc.add(f"E2:E{nrows}")
        dvm = DataValidation(type="list", formula1="=Lists!$C$2:$C$5", allow_blank=True)
        ws.add_data_validation(dvm); dvm.add(f"F2:F{nrows}")
        if suppliers:
            dvs = DataValidation(type="list",
                                 formula1=f"=Lists!$D$2:$D${len(suppliers)+1}",
                                 allow_blank=True)
            ws.add_data_validation(dvs); dvs.add(f"H2:H{nrows}")
        dvt = DataValidation(type="list",
                             formula1=f"=Lists!$E$2:$E${len(etypes)+1}",
                             allow_blank=True)
        ws.add_data_validation(dvt); dvt.add(f"K2:K{nrows}")
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["H"].width = 24
        ws.column_dimensions["I"].width = 20
        ws.column_dimensions["L"].width = 20
        info = wb.create_sheet("How to fill this in")
        for i, line in enumerate([
            "Expense import",
            "",
            "One row per expense.",
            "  - Date — YYYY-MM-DD (required).",
            "  - Fund — the fund charged (pick from the list, required).",
            "  - Description — what it was for (required).",
            "  - Amount — required, > 0.",
            "  - Category — Allowance / Transport / Materials / ... (defaults to Other).",
            "  - Method — Cash / Bank / Cheque / M-Pesa (defaults to Cash).",
            "  - Claimant — who incurred or requested it (optional).",
            "  - Supplier — pick from the register on the Lists sheet. Optional, but",
            "    choosing one puts the payment on that supplier's account.",
            "  - Payee — who the money actually goes to, if not the claimant.",
            "  - Voucher no — optional reference.",
            "  - Expenditure type — Recurrent (running cost) or Capital (creates or",
            "    improves an asset). Defaults to Recurrent.",
            "  - Budget item — optional. Matched by name within the fund on the same",
            "    row, so the same item name may be used in different funds.",
            "  - M-Pesa charge — optional. The transaction/withdrawal charge for this",
            "    payment; it's recorded as a separate bank-charge expense on the same",
            "    fund and linked back to this expense.",
            "  - Paid from petty cash — optional Yes/No. If Yes, the expense (and any",
            "    charge) is paid from the petty cash float and reduces it.",
            "",
            "If approval is required, imported expenses arrive as Pending for you to",
            "approve. If it isn't, they are approved automatically.",
        ], start=1):
            info.cell(i, 1, line)
        info.column_dimensions["A"].width = 74
        buf = io.BytesIO(); wb.save(buf)
        resp = HttpResponse(buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = 'attachment; filename="expense_import_template.xlsx"'
        return resp

    def _parse(self, request):
        import openpyxl, datetime as _dt
        from departments.models import Department
        f = request.FILES.get("file")
        if not f:
            messages.error(request, "Choose a spreadsheet to upload.")
            return redirect("expense_import")
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
        except Exception:
            from core.utils import log_exception as _lx; _lx('cashbook/views.py')
            messages.error(request, "Could not read that file — please upload a .xlsx.")
            return redirect("expense_import")
        ws = wb["Expenses"] if "Expenses" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            messages.error(request, "The sheet is empty.")
            return redirect("expense_import")
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

        def col(*names):
            for n in names:
                if n in header:
                    return header.index(n)
            return None
        c_date = col("date")
        c_fund = col("fund", "department")
        c_desc = col("description", "details", "expense")
        c_amt = col("amount", "value")
        c_cat = col("category")
        c_method = col("method", "paid by", "payment")
        c_claim = col("claimant", "paid to")
        c_supplier = col("supplier", "vendor")
        c_payee = col("payee")
        c_etype = col("expenditure type", "type")
        c_budget = col("budget item", "budget line", "budget")
        c_vouch = col("voucher no", "voucher", "ref")
        c_charge = col("m-pesa charge", "mpesa charge", "charge", "transaction charge")
        c_petty = col("paid from petty cash", "petty cash", "petty")
        if c_date is None or c_amt is None or c_desc is None:
            messages.error(request, "Need at least Date, Description and Amount columns "
                                    "— please use the template.")
            return redirect("expense_import")

        funds = {d.name.strip().lower(): d for d in Department.objects.all()}
        from vendors.models import Vendor, name_key as _vkey
        # Matched on the register's own normalised key, so "Mwangi Hardware Ltd"
        # in the sheet finds "MWANGI HARDWARE" in the register rather than
        # creating a fourth spelling. Never creates a supplier: a typo in a
        # spreadsheet must not silently add to the register.
        suppliers = {}
        for v in Vendor.objects.exclude(status=Vendor.Status.ARCHIVED):
            suppliers.setdefault(v.name_key, v)
            suppliers.setdefault(v.name.strip().lower(), v)
        etypes = {t.label.upper(): t.value for t in Expense.ExpenditureType}
        etypes.update({t.value: t.value for t in Expense.ExpenditureType})

        def cell(r, idx):
            if idx is None or idx >= len(r) or r[idx] in (None, ""):
                return ""
            return str(r[idx]).strip()

        def pdate(v):
            if isinstance(v, _dt.datetime):
                return v.date().isoformat()
            if isinstance(v, _dt.date):
                return v.isoformat()
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y"):
                try:
                    return _dt.datetime.strptime(str(v).strip(), fmt).date().isoformat()
                except (ValueError, TypeError):
                    continue
            return None

        plan = []
        for r in rows[1:]:
            desc = cell(r, c_desc)
            d = pdate(r[c_date]) if c_date < len(r) else None
            try:
                amt = float(r[c_amt]) if c_amt < len(r) and r[c_amt] not in (None, "") else 0.0
            except (TypeError, ValueError):
                amt = 0.0
            if not desc and not d and amt <= 0:
                continue
            fund_raw = cell(r, c_fund)
            fund = funds.get(fund_raw.lower()) if fund_raw else None
            cat = self.CATEGORY_LABELS.get(cell(r, c_cat).upper(), "OTHER")
            method = self.METHOD_LABELS.get(cell(r, c_method).upper(), "CASH")
            try:
                charge = (float(r[c_charge]) if c_charge is not None and c_charge < len(r)
                          and r[c_charge] not in (None, "") else 0.0)
            except (TypeError, ValueError):
                charge = 0.0
            petty_raw = cell(r, c_petty) if c_petty is not None else ""
            petty = str(petty_raw).strip().lower() in ("yes", "y", "true", "1", "x")
            supplier_raw = cell(r, c_supplier) if c_supplier is not None else ""
            supplier = None
            if supplier_raw:
                supplier = (suppliers.get(_vkey(supplier_raw))
                            or suppliers.get(supplier_raw.strip().lower()))
            etype_raw = cell(r, c_etype) if c_etype is not None else ""
            etype = etypes.get(etype_raw.upper(), Expense.ExpenditureType.RECURRENT)
            budget_raw = cell(r, c_budget) if c_budget is not None else ""
            budget = None
            if budget_raw and fund:
                # Scoped to the row's own fund: the same item name may exist in
                # several budgets, and charging spend to another fund's line
                # would corrupt both.
                from .models import BudgetLine
                budget = BudgetLine.objects.filter(
                    name__iexact=budget_raw, budget__department=fund).first()
            plan.append({
                "date": d, "fund_raw": fund_raw,
                "fund_id": fund.id if fund else None,
                "fund_name": fund.name if fund else None,
                "description": desc[:200], "amount": amt,
                "category": cat, "method": method, "charge": round(charge, 2),
                "petty": petty,
                "claimant": cell(r, c_claim)[:120], "voucher": cell(r, c_vouch)[:30],
                "supplier_raw": supplier_raw,
                "supplier_id": supplier.id if supplier else None,
                "supplier_name": supplier.name if supplier else None,
                "payee": (cell(r, c_payee)[:160] if c_payee is not None else ""),
                "expenditure_type": etype,
                "budget_line_id": budget.id if budget else None,
                # An unrecognised supplier is a warning, not a rejection: the
                # expense is still real and the register can be tidied later.
                "warn": ("Supplier not on the register — will be recorded "
                         "without one." if supplier_raw and not supplier else ""),
                "ok": bool(d and desc and amt > 0 and fund),
            })
        if not plan:
            messages.error(request, "No expense rows were found.")
            return redirect("expense_import")
        request.session["expense_import_plan"] = plan
        from core.models import SiteConfig
        return render(request, self.template_name, {
            "stage": "review", "plan": plan,
            "ready": sum(1 for p in plan if p["ok"]),
            "problems": sum(1 for p in plan if not p["ok"]),
            "total": sum(p["amount"] for p in plan if p["ok"]),
            "auto_approve": not SiteConfig.get().require_expense_approval,
        })

    @db_tx.atomic
    def _apply(self, request):
        from departments.models import Department
        from core.models import SiteConfig
        plan = request.session.get("expense_import_plan")
        if not plan:
            messages.error(request, "Your import session expired — please upload again.")
            return redirect("expense_import")
        auto = not SiteConfig.get().require_expense_approval
        created = skipped = 0
        for p in plan:
            if not p["ok"]:
                skipped += 1
                continue
            fund = Department.objects.filter(pk=p["fund_id"]).first()
            if not fund:
                skipped += 1
                continue
            import datetime as _dt
            from .models import BudgetLine
            from .services import expenses as expense_svc
            from vendors.models import Vendor

            d = _dt.date.fromisoformat(p["date"])
            # Through the same function the form and the batch screen use, so
            # an imported expense is indistinguishable from a typed one — and
            # so the status and transaction-charge rules cannot drift between
            # the three entry paths again.
            expense, charge_expense = expense_svc.record(
                date=d, department=fund, description=p["description"],
                amount=p["amount"], user=request.user,
                category=p["category"], method=p["method"],
                claimant=p["claimant"], payee=p.get("payee", ""),
                voucher_no=p["voucher"],
                vendor=(Vendor.objects.filter(pk=p["supplier_id"]).first()
                        if p.get("supplier_id") else None),
                expenditure_type=p.get("expenditure_type") or None,
                budget_line=(BudgetLine.objects.filter(pk=p["budget_line_id"]).first()
                             if p.get("budget_line_id") else None),
                paid_from_petty_cash=bool(p.get("petty")),
                auto_approve=auto, charge=p.get("charge"))
            created += 1
            if charge_expense is not None:
                created += 1
        request.session.pop("expense_import_plan", None)
        state = "approved" if auto else "pending approval"
        parts = [f"{created} expense(s) imported ({state})"]
        if skipped:
            parts.append(f"{skipped} row(s) skipped (missing date, fund, description or amount)")
        messages.success(request, ", ".join(parts) + ".")
        return redirect("expense_list")


class ExpenseBulkActionView(TreasurerRequiredMixin, View):
    """Apply one action (approve / reject / pay / delete) to several selected
    expenses at once. Each item is checked against the same guards as the single
    action, and ones that don't qualify are skipped (and counted) rather than
    erroring the whole batch."""
    def post(self, request):
        import datetime as dt
        from django.shortcuts import redirect
        from core.models import SiteConfig
        ids = request.POST.getlist("ids")
        action = request.POST.get("action")
        if not ids:
            messages.info(request, "No expenses were selected.")
            return redirect("expense_list")
        threshold = SiteConfig.get().dual_approval_threshold or 0
        require_diff = SiteConfig.get().require_different_approver
        from core.models import period_locked
        done = skipped = 0
        S = Expense.Status
        for exp in Expense.objects.filter(pk__in=ids):
            if period_locked(exp.date):
                skipped += 1
                continue
            if action == "approve" and exp.status == S.PENDING:
                if require_diff and exp.recorded_by_id == request.user.id:
                    skipped += 1
                    continue
                exp.status = S.APPROVED
                exp.approved_by = request.user
                exp.save()
                done += 1
            elif action == "reject" and exp.status in (S.PENDING, S.APPROVED):
                exp.status = S.REJECTED
                exp.approved_by = request.user
                exp.save()
                done += 1
            elif action == "pay" and exp.status == S.APPROVED:
                needs_two = threshold and exp.amount >= threshold
                if needs_two and not (exp.approved_by_id and exp.second_approved_by_id):
                    skipped += 1
                    continue
                exp.status = S.PAID
                exp.paid_date = dt.date.today()
                exp.save()
                done += 1
            elif action == "delete":
                exp.delete()
                done += 1
            else:
                skipped += 1
        verb = {"approve": "approved", "reject": "rejected", "pay": "marked paid",
                "delete": "deleted"}.get(action, "updated")
        msg = f"{done} expense(s) {verb}."
        if skipped:
            msg += f" {skipped} skipped (locked period or not eligible)."
        (messages.success if done else messages.info)(request, msg)
        return redirect("expense_list")


def _post_decimal(request, name):
    """A POST field as a Decimal, or None if blank/invalid. Shared by both of
    FundBudgetView's independent goal-editing forms."""
    from decimal import Decimal, InvalidOperation
    try:
        v = request.POST.get(name, "").strip()
        return Decimal(v) if v else None
    except InvalidOperation:
        return None


class FundBudgetView(LoginRequiredMixin, View):
    """Budget & goals for a fund (e.g. Camp Meeting): per-category budget vs actual
    spend for a year, plus the contribution goal (Department.target) and the
    yearly goal (Department.annual_budget) tracked against what's been collected.

    Viewing (GET) is available to treasurers/assistants for any fund, and to a
    leader for a fund they lead once granted the view_fund_budget right — see
    core.roles.can_view_fund_budget. Editing (POST) always stays
    treasurer/assistant only, regardless of that right."""
    template_name = "cashbook/fund_budget.html"

    def _ctx(self, request, dept):
        import datetime as dt
        from decimal import Decimal
        from django.db.models import Sum
        from giving.models import Transaction
        from .models import BudgetLine, Expense
        year = int(request.GET.get("year") or dt.date.today().year)

        lines = list(BudgetLine.objects.filter(department=dept, year=year))
        # actual spend per budget item, from expenses tagged to that item
        spend = {r["budget_line"]: r["t"] for r in (Expense.objects.filter(
            budget_line__in=lines,
            status__in=[Expense.Status.APPROVED, Expense.Status.PAID])
            .values("budget_line").annotate(t=Sum("amount")))}
        # spend on this fund not tagged to any item (so nothing is hidden)
        untagged = (Expense.objects.filter(department=dept, date__year=year,
            budget_line__isnull=True,
            status__in=[Expense.Status.APPROVED, Expense.Status.PAID])
            .aggregate(t=Sum("amount"))["t"] or Decimal(0))
        rows, tot_budget, tot_actual = [], Decimal(0), Decimal(0)
        for b in lines:
            actual = spend.get(b.id, Decimal(0))
            tot_budget += b.amount; tot_actual += actual
            rows.append({"id": b.id, "name": b.name,
                         "category": b.get_category_display() if b.category else "",
                         "budget": b.amount, "actual": actual,
                         "variance": b.amount - actual,
                         "pct": int(min(actual / b.amount * 100, 999)) if b.amount else 0,
                         "note": b.note})

        # --- collections aggregated across this fund AND its sub-accounts ---
        def _fund_ids(d):
            ids = [d.id]
            for sub in d.subgroups.all():
                ids.extend(_fund_ids(sub))
            return ids

        def _collected(fund, year_):
            if fund is None:
                return Decimal(0)
            return (Transaction.objects.filter(
                department_id__in=_fund_ids(fund),
                direction=Transaction.Direction.CREDIT, confirmed=True,
                is_reversal=False, is_reversed=False, excluded_from_income=False,
                date__year=year_).aggregate(t=Sum("amount"))["t"] or Decimal(0))

        # Camp Meeting EXPENSE goal: collected across the expense fund + subgroups
        expense_collected = _collected(dept, year)

        def _goal(g, collected):
            g = g or Decimal(0)
            return {"goal": g, "collected": collected,
                    "pct": int(min(collected / g * 100, 100)) if g else 0,
                    "short": max(g - collected, Decimal(0)),
                    "variance": collected - g}

        # Group Contribution goal: one goal set on the parent, progress from each
        # sub-account's own collection (shown per group, plus the aggregate)
        group_rows = []
        for sub in dept.subgroups.all():
            c = _collected(sub, year)
            g = sub.contribution_goal or Decimal(0)
            group_rows.append({"name": sub.name, "id": sub.id, "collected": c,
                "goal": g, "pct": int(min(c / g * 100, 100)) if g else 0,
                "short": max(g - c, Decimal(0))})
        group_rows.sort(key=lambda r: r["collected"], reverse=True)
        contribution_collected = sum((r["collected"] for r in group_rows), Decimal(0))
        contribution_target = sum((r["goal"] for r in group_rows), Decimal(0))

        # Camp Meeting OFFERING goal: a church-wide Trust-fund target, now
        # configured in Settings → Goals rather than on this fund — shown here
        # read-only for context, never editable/saved from this page.
        from core.models import SiteConfig
        cfg = SiteConfig.get()
        offering = cfg.camp_offering_fund if dept.goal_type == "CAMP_EXPENSE" else None
        offering_collected = _collected(offering, year)

        # loan financing block (only shown when the fund has loans): loans
        # raise the fund's available cash but never its income, so they are
        # presented as their own line alongside contributions and expenses
        from loans.services.loans import fund_loan_summary
        loan_summary = fund_loan_summary(dept)

        return {
            "dept": dept, "year": year,
            "loan_summary": loan_summary,
            "years": range(dt.date.today().year + 1, dt.date.today().year - 4, -1),
            "rows": rows, "tot_budget": tot_budget, "tot_actual": tot_actual,
            "tot_variance": tot_budget - tot_actual, "untagged": untagged,
            "expense_goal": _goal(dept.year_goal, expense_collected),
            "contribution_goal": _goal(contribution_target, contribution_collected),
            "group_rows": group_rows,
            "offering": offering,
            "offering_goal": _goal(cfg.camp_offering_goal if dept.goal_type == "CAMP_EXPENSE"
                                   else None, offering_collected),
            "categories": Expense.Category.choices,
            "is_camp_expense": dept.goal_type == "CAMP_EXPENSE",
            "goal_type": dept.goal_type,
            "all_funds": Department.objects.filter(active=True, is_trust=True)
                         .exclude(pk=dept.pk).order_by("name"),
        }

    def get(self, request, pk):
        from departments.models import Department
        from core import roles
        dept = get_object_or_404(Department, pk=pk)
        if not roles.can_view_fund_budget(request.user, dept):
            if roles.is_leader(request.user):
                messages.error(request, "You don't lead that fund, or don't "
                               "have the budget-viewing right for it.")
                return redirect("leader_dashboard")
            messages.error(request, "This page is restricted to Treasurers.")
            return redirect("dashboard")
        return render(request, self.template_name, self._ctx(request, dept))

    def post(self, request, pk):
        import datetime as dt
        from decimal import Decimal, InvalidOperation
        from departments.models import Department
        from core import roles
        from .models import BudgetLine
        dept = get_object_or_404(Department, pk=pk)
        # editing/saving is always treasurer/assistant only, regardless of the
        # view_fund_budget right a leader might hold for this fund
        if not (roles.is_treasurer(request.user) or roles.is_assistant(request.user)):
            messages.error(request, "This page is restricted to Treasurers.")
            if roles.is_leader(request.user):
                return redirect("leader_dashboard")
            return redirect("dashboard")
        year = int(request.POST.get("year") or dt.date.today().year)
        # update the fund's goals — two independent forms/buttons on the page
        # (expense goal, and per-group contribution goals) must not clobber
        # each other: submitting one used to blank the other's fields because
        # both posted to the same "save_goals" flag and the shared handler
        # unconditionally rewrote every field regardless of which form was
        # actually submitted.
        if "save_expense_goal" in request.POST:
            dept.year_goal = _post_decimal(request, "expense_goal")          # Camp Meeting Expense goal
            gt = request.POST.get("goal_type") or "NONE"
            dept.goal_type = gt if gt in ("NONE", "CAMP_EXPENSE") else "NONE"
            # The Camp Meeting OFFERING goal (a church-wide Trust-fund target) is
            # configured in Settings → Goals, not here — this page only tracks
            # this fund's own expense goal and its subgroups' contribution goals.
            dept.offering_goal = None
            dept.offering_fund = None
            dept.save(update_fields=["year_goal", "offering_goal",
                                     "offering_fund", "goal_type"])
            messages.success(request, "Goals updated.")
            return redirect(f"{request.path}?year={year}")
        if "save_group_goals" in request.POST:
            # each development group has its own contribution goal — this form
            # only ever touches subgroup rows, never the fund's own year_goal
            for sub in dept.subgroups.all():
                val = _post_decimal(request, f"group_goal_{sub.id}")
                if sub.contribution_goal != val:
                    sub.contribution_goal = val
                    sub.save(update_fields=["contribution_goal"])
            messages.success(request, "Group goals updated.")
            return redirect(f"{request.path}?year={year}")
        # add / update a named budget item
        name = (request.POST.get("name") or "").strip()
        try:
            amount = Decimal(request.POST.get("amount") or "0")
        except InvalidOperation:
            amount = Decimal(0)
        if name:
            BudgetLine.objects.update_or_create(
                department=dept, year=year, name=name,
                defaults={"amount": amount, "category": request.POST.get("category", ""),
                          "note": request.POST.get("note", "")[:120]})
            messages.success(request, "Budget item saved.")
        return redirect(f"{request.path}?year={year}")


class GroupGoalsPngView(LoginRequiredMixin, View):
    """Server-rendered PNG of a fund's Group Contribution Goals — a proper
    per-group progress bar chart, generated with Pillow so it looks identical
    everywhere and needs no client-side rendering (canvas/screenshot). PNG
    rather than JPEG: this is a table of sharp text and flat fills, and PNG's
    lossless compression keeps it crisp instead of JPEG-blurred.

    Permission (v2.44 fix): must match FundBudgetView's own
    can_view_fund_budget check exactly, not a narrower Treasurer-only check —
    this page is embedded/linked FROM the budget page, so anyone who can see
    that page (assistants, and leaders granted view_fund_budget for their own
    fund) must also be able to load the image on it, or the img tag simply
    shows broken/blank for them with no obvious reason why."""
    def get(self, request, pk):
        from departments.models import Department
        from core.models import SiteConfig
        from core import roles
        from django.http import HttpResponse
        from .services.goal_chart import build_group_goals_png
        dept = get_object_or_404(Department, pk=pk)
        if not roles.can_view_fund_budget(request.user, dept):
            return HttpResponse(status=403)
        ctx = FundBudgetView()._ctx(request, dept)
        data = build_group_goals_png(
            dept_name=dept.name, year=ctx["year"], group_rows=ctx["group_rows"],
            contribution_goal=ctx["contribution_goal"],
            church_name=SiteConfig.get().church_name or "")
        resp = HttpResponse(data, content_type="image/png")
        fname = f"group-contribution-goals-{dept.slug or dept.id}-{ctx['year']}.png"
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp


class BudgetItemsPngView(LoginRequiredMixin, View):
    """Server-rendered PNG of a fund's 'Budget vs actual by item' table,
    generated with Pillow — same approach as GroupGoalsPngView, so it looks
    identical wherever it's downloaded, matching the on-screen table exactly
    (Budget item / Budget / Actual / Variance / Used), including the totals
    row.

    Permission: see GroupGoalsPngView's docstring — same fix, same reason."""
    def get(self, request, pk):
        from departments.models import Department
        from core.models import SiteConfig
        from core import roles
        from django.http import HttpResponse
        from .services.goal_chart import build_budget_items_png
        dept = get_object_or_404(Department, pk=pk)
        if not roles.can_view_fund_budget(request.user, dept):
            return HttpResponse(status=403)
        ctx = FundBudgetView()._ctx(request, dept)
        data = build_budget_items_png(
            dept_name=dept.name, year=ctx["year"], rows=ctx["rows"],
            tot_budget=ctx["tot_budget"], tot_actual=ctx["tot_actual"],
            tot_variance=ctx["tot_variance"],
            church_name=SiteConfig.get().church_name or "")
        resp = HttpResponse(data, content_type="image/png")
        fname = f"budget-vs-actual-{dept.slug or dept.id}-{ctx['year']}.png"
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp


class BudgetItemsJSONView(DataEntryRequiredMixin, View):
    """Budget items for a fund + year, for the expense form's 'Budget item'
    picker. Returns [] for funds that have no budget set."""
    def get(self, request):
        import datetime as dt
        from django.http import JsonResponse
        from .models import BudgetLine
        try:
            dept_id = int(request.GET.get("dept") or 0)
        except ValueError:
            dept_id = 0
        year = request.GET.get("year")
        try:
            year = int(year) if year else dt.date.today().year
        except ValueError:
            year = dt.date.today().year
        items = (BudgetLine.objects.filter(department_id=dept_id, year=year)
                 .order_by("name"))
        return JsonResponse({"items": [
            {"id": b.id, "name": b.name, "category": b.category,
             "amount": float(b.amount)} for b in items]})


# --- Payables / accruals / prepayments: edit, delete, settle-against-expense ---

class _ObligationEditView(DataEntryRequiredMixin, View):
    """Edit a payable/accrual/prepayment. Settled items are read-only (safe)."""
    model = None
    form_class = None
    title = ""
    template_name = "cashbook/obligation_edit.html"

    def _get(self, pk):
        return get_object_or_404(self.model, pk=pk)

    def get(self, request, pk):
        obj = self._get(pk)
        if getattr(obj, "settled", False):
            messages.error(request, "This item is settled and can no longer be edited.")
            return redirect("accruals")
        fields = self.form_class().fields
        initial = {f: getattr(obj, f) for f in fields if hasattr(obj, f)}
        return render(request, self.template_name,
                      {"form": self.form_class(initial=initial), "title": self.title, "obj": obj})

    def post(self, request, pk):
        obj = self._get(pk)
        if getattr(obj, "settled", False):
            messages.error(request, "This item is settled and can no longer be edited.")
            return redirect("accruals")
        form = self.form_class(request.POST)
        if form.is_valid():
            if _block_if_locked(request, form.cleaned_data.get("date", dt.date.today())):
                return redirect("accruals")
            for f, v in form.cleaned_data.items():
                if hasattr(obj, f):
                    setattr(obj, f, v)
            obj.save()
            messages.success(request, "Saved.")
            return redirect("accruals")
        return render(request, self.template_name,
                      {"form": form, "title": self.title, "obj": obj})


class PayableEdit(_ObligationEditView):
    model = Payable; form_class = PayableForm; title = "Edit payable"

class AccrualEdit(_ObligationEditView):
    model = Accrual; form_class = AccrualForm; title = "Edit accrual"

class PrepaymentEdit(_ObligationEditView):
    model = Prepayment; form_class = PrepaymentForm; title = "Edit prepayment"


class _ObligationDeleteView(DataEntryRequiredMixin, View):
    model = None
    label = "item"
    def post(self, request, pk):
        obj = get_object_or_404(self.model, pk=pk)
        if getattr(obj, "settled", False):
            messages.error(request, f"A settled {self.label} can't be deleted. "
                                    f"Reverse the settlement first if it was a mistake.")
        else:
            obj.delete()
            messages.success(request, f"{self.label.capitalize()} deleted.")
        return redirect("accruals")


class PayableDelete(_ObligationDeleteView):
    model = Payable; label = "payable"

class AccrualDelete(_ObligationDeleteView):
    model = Accrual; label = "accrual"

class PrepaymentDelete(_ObligationDeleteView):
    model = Prepayment; label = "prepayment"


class SettleAgainstExpenseView(DataEntryRequiredMixin, View):
    """Settle a payable/accrual by linking an expense that was already entered
    (e.g. the treasurer keyed the payment as an expense by mistake), instead of
    creating a second one. Lists unlinked expenses on the same fund to choose from."""
    template_name = "cashbook/settle_existing.html"

    def _obj(self, kind, pk):
        model = {"payable": Payable, "accrual": Accrual}.get(kind)
        return model.objects.filter(pk=pk).first() if model else None

    def get(self, request, kind, pk):
        obj = self._obj(kind, pk)
        if not obj or obj.settled:
            messages.error(request, "That item can't be settled.")
            return redirect("accruals")
        cands = (Expense.objects.filter(department=obj.department,
                    payable__isnull=True, accrual__isnull=True)
                 .exclude(category=Expense.Category.BANK_CHARGE)
                 .order_by("-date")[:100])
        return render(request, self.template_name,
                      {"obj": obj, "kind": kind, "candidates": cands})

    def post(self, request, kind, pk):
        obj = self._obj(kind, pk)
        if not obj or obj.settled:
            messages.error(request, "That item can't be settled.")
            return redirect("accruals")
        exp = get_object_or_404(Expense, pk=request.POST.get("expense"))
        if kind in ("payable", "accrual"):
            from .services import obligations as obligation_svc
            field = "accrual" if kind == "accrual" else "payable"
            setattr(exp, field, obj)
            exp.save(update_fields=[field])
            obligation_svc.refresh_settlement(obj)
            obj.refresh_from_db()
            if obj.settled:
                messages.success(
                    request, f"{kind.title()} settled against the existing expense "
                             f"“{exp.description}”.")
            else:
                messages.success(
                    request, f"“{exp.description}” linked as a part payment. "
                             f"{obj.balance:,.2f} still owing.")
            return redirect("accruals")
        obj.settled = True
        obj.settled_on = exp.date
        obj.settled_expense = exp
        obj.save()
        messages.success(request, f"{kind.title()} settled against the existing expense "
                                  f"“{exp.description}”.")
        return redirect("accruals")


# (unpresented_cheques_total, unpresented_payments_qs are re-exported from the
# top of this file.)


class ChequeRegisterView(PaymentViewMixin, View):
    """Payment register — cheques today, extensible to EFT/RTGS/M-Pesa. Lists
    payment instruments, filterable by method and status, and drives the full
    lifecycle (draft -> approved -> issued -> cleared, plus void/stop)."""
    template_name = "cashbook/payment_register.html"

    def _scoped(self, request, qs):
        """Department leaders only see instruments on funds they lead."""
        from core import roles
        from .models import PaymentInstrument
        if roles.can_view_payments(request.user):
            return qs
        from leaders.permissions import allowed_departments
        ids = list(allowed_departments(request.user).values_list("id", flat=True))
        from django.db.models import Q
        return qs.filter(Q(expense__department_id__in=ids)
                         | Q(extra_expenses__department_id__in=ids)).distinct()

    def get(self, request):
        import datetime as dtt
        from decimal import Decimal
        from django.core.paginator import Paginator
        from django.db.models import Count, Q, Sum
        from core import roles
        from .models import PaymentInstrument
        g = request.GET
        status = g.get("status", "")
        method = g.get("method", "")
        q = (g.get("q") or "").strip()
        qs = PaymentInstrument.objects.select_related(
            "expense__department", "remittance_batch", "refund", "transfer",
            "bank_account", "recorded_by", "approved_by", "bank_transaction")
        qs = self._scoped(request, qs)
        if status == "_outstanding":
            qs = qs.filter(status__in=PaymentInstrument.OUTSTANDING_STATES)
        elif status:
            qs = qs.filter(status=status)
        if method:
            qs = qs.filter(method=method)
        if g.get("source_kind"):
            qs = qs.filter(source_kind=g["source_kind"])
        if g.get("bank"):
            qs = qs.filter(bank_account_id=g["bank"])
        if g.get("fund"):
            qs = qs.filter(Q(expense__department_id=g["fund"])
                           | Q(extra_expenses__department_id=g["fund"])).distinct()
        for key, lookup in (("start", "date_issued__gte"),
                            ("end", "date_issued__lte")):
            if g.get(key):
                try:
                    qs = qs.filter(**{lookup: dtt.date.fromisoformat(g[key])})
                except ValueError:
                    pass
        if q:
            amount_q = Q()
            try:
                from decimal import Decimal as _D
                amount_q = Q(amount=_D(q.replace(",", "")))
            except Exception:  # noqa: BLE001
                pass
            qs = qs.filter(
                Q(instrument_number__icontains=q) | Q(payee__icontains=q)
                | Q(note__icontains=q) | Q(expense__voucher_no__icontains=q)
                | Q(expense__description__icontains=q)
                | Q(bank_transaction__core_ref__icontains=q) | amount_q).distinct()
        sort = g.get("sort") or "-date_issued"
        allowed_sorts = {"date_issued", "-date_issued", "amount", "-amount",
                         "date_cleared", "-date_cleared", "payee", "-payee",
                         "status", "-status", "instrument_number"}
        if sort in allowed_sorts:
            qs = qs.order_by(sort, "-id")

        export = g.get("export")
        if export in ("csv", "xlsx"):
            return self._export_register(request, qs)

        # ---- dashboard metrics (Part 10) ----
        today = dtt.date.today()
        base = PaymentInstrument.objects.all()
        base = self._scoped(request, base)
        outq = base.filter(status__in=PaymentInstrument.OUTSTANDING_STATES)
        by_method = {r["method"]: r for r in outq.values("method").annotate(
            n=Count("id"), t=Sum("amount"))}
        cleared_pairs = list(base.filter(
            status=PaymentInstrument.Status.CLEARED,
            date_issued__isnull=False, date_cleared__isnull=False)
            .values_list("date_issued", "date_cleared")[:2000])
        avg_days = (sum((c - i).days for i, c in cleared_pairs)
                    / len(cleared_pairs)) if cleared_pairs else None
        oldest = outq.filter(date_issued__isnull=False)                      .order_by("date_issued").first()
        stats = {
            "out_cheques": by_method.get("CHEQUE", {}),
            "out_eft": by_method.get("EFT", {}),
            "out_rtgs": by_method.get("RTGS", {}),
            "awaiting_n": outq.count(),
            "awaiting_t": outq.aggregate(t=Sum("amount"))["t"] or Decimal(0),
            "cleared_today": base.filter(date_cleared=today).count(),
            "cancelled_n": base.filter(
                status__in=PaymentInstrument.TERMINAL_STATES).count(),
            "avg_days": round(avg_days, 1) if avg_days is not None else None,
            "oldest": oldest,
        }

        page = Paginator(qs, 40).get_page(g.get("page"))
        from departments.models import Department
        from statements.models import BankAccount
        ctx = {
            "cheques": page.object_list, "page_obj": page,
            "is_paginated": page.has_other_pages(),
            "status": status, "method": method, "q": q, "f": g,
            "sort": sort, "stats": stats,
            "sort_options": [("-date_issued", "Newest issued first"),
                             ("date_issued", "Oldest issued first"),
                             ("-amount", "Largest amount"), ("amount", "Smallest amount"),
                             ("-date_cleared", "Recently cleared"),
                             ("payee", "Payee A–Z"), ("status", "Status"),
                             ("instrument_number", "Instrument no.")],
            "statuses": PaymentInstrument.Status.choices,
            "methods": PaymentInstrument.Method.choices,
            "source_kinds": PaymentInstrument.SourceKind.choices,
            "funds": Department.objects.filter(active=True).order_by("name"),
            "banks": BankAccount.objects.all(),
            "unpresented_total": unpresented_cheques_total(),
            "can_enter_data": roles.can_manage_payments(request.user),
            "can_approve": roles.can_approve_payments(request.user),
            "can_clear": roles.can_clear_payments(request.user),
            "can_void": roles.can_void_payments(request.user),
            "is_treasurer": roles.is_treasurer(request.user)
                            or getattr(request.user, "is_superuser", False),
        }
        # deep-link prefill: e.g. from an expense's "Issue a payment" link
        for_kind = request.GET.get("for_kind", "")
        for_id = request.GET.get("for_id", "")
        if for_kind == "EXPENSE" and for_id.isdigit():
            exp = Expense.objects.filter(pk=for_id).first()
            if exp:
                ctx["prefill_kind"] = "EXPENSE"
                ctx["prefill_id"] = exp.id
                ctx["prefill_amount"] = exp.amount
                ctx["prefill_payee"] = exp.claimant or exp.department.name
        return render(request, self.template_name, ctx)

    def _export_register(self, request, qs):
        from reports.exports import csv_response, xlsx_response
        from core.models import SiteConfig
        header = ["Instrument no", "Type", "Source type", "Source ref", "Payee",
                  "Fund(s)", "Bank account", "Amount", "Issue date",
                  "Cleared date", "Status", "Outstanding?", "Days to clear",
                  "Created by", "Approved by", "Bank ref"]
        rows = []
        for p in qs[:5000]:
            src = p.source
            rows.append([
                p.instrument_number or f"#{p.pk}", p.get_method_display(),
                p.get_source_kind_display(),
                (p.expense.voucher_no or f"EXP-{p.expense_id}") if p.expense_id
                else (str(src) if src else ""),
                p.payee, p.fund_names,
                str(p.bank_account) if p.bank_account_id else "",
                float(p.amount),
                p.date_issued.isoformat() if p.date_issued else "",
                p.date_cleared.isoformat() if p.date_cleared else "",
                p.get_status_display(),
                "Yes" if p.is_outstanding else "",
                p.clearance_days if p.clearance_days is not None else "",
                getattr(p.recorded_by, "username", ""),
                getattr(p.approved_by, "username", ""),
                p.bank_transaction.core_ref if p.bank_transaction_id else ""])
        fn = "payment_register"
        if request.GET.get("export") == "xlsx":
            return xlsx_response(fn + ".xlsx", header, rows,
                                 title="Payment Register",
                                 church=SiteConfig.get().church_name)
        return csv_response(fn + ".csv", header, rows)

    def post(self, request):
        from decimal import Decimal, InvalidOperation
        import datetime as dt
        from django.core.exceptions import ValidationError
        from core import roles
        from .models import (PaymentInstrument, Expense, RemittanceBatch,
                             ExpenseRefund, FundTransfer)
        from .services.payments import apply_event, reissue
        action = request.POST.get("action")
        treas = roles.is_treasurer(request.user) or getattr(request.user, "is_superuser", False)
        # granular gates per action (Part 13); everything else needs manage
        _need = {"add": roles.can_manage_payments,
                 "prepare": roles.can_manage_payments,
                 "issue": roles.can_manage_payments,
                 "present": roles.can_manage_payments,
                 "approve": roles.can_approve_payments,
                 "clear": roles.can_clear_payments,
                 "void": roles.can_void_payments,
                 "cancel": roles.can_void_payments,
                 "reject": roles.can_void_payments,
                 "reverse": roles.can_void_payments,
                 "expire": roles.can_void_payments,
                 "stop": roles.can_void_payments,
                 "reissue": roles.can_void_payments,
                 "delete": roles.can_void_payments,
                 "sync": roles.can_manage_payments}
        gate = _need.get(action, roles.can_manage_payments)
        if not gate(request.user):
            messages.error(request, "You do not have the right for that payment action.")
            return redirect("payment_register")

        if action == "add":
            try:
                amount = Decimal(request.POST.get("amount") or "0")
            except InvalidOperation:
                amount = Decimal(0)
            try:
                issued = dt.date.fromisoformat(request.POST.get("date_issued"))
            except (TypeError, ValueError):
                issued = None
            kind = request.POST.get("source_kind") or "EXPENSE"
            src_id = (request.POST.get("source_id") or "").strip()
            inst = PaymentInstrument(
                method=request.POST.get("method") or "CHEQUE",
                instrument_number=(request.POST.get("instrument_number") or "")[:40],
                payee=(request.POST.get("payee") or "")[:160],
                amount=amount, date_issued=issued, source_kind=kind,
                signatory_1=(request.POST.get("signatory_1") or "")[:120],
                signatory_2=(request.POST.get("signatory_2") or "")[:120],
                note=(request.POST.get("note") or "")[:200],
                recorded_by=request.user, status="DRAFT")
            # attach the referenced source. For EXPENSE, a comma-separated
            # id list makes one instrument settle several vouchers (one EFT
            # covering multiple expenses); the total must match exactly.
            extra_expenses = []
            if kind == "EXPENSE" and src_id:
                ids = [i.strip() for i in src_id.split(",") if i.strip().isdigit()]
                # preserve the order the ids were given in — the first is the
                # primary expense, the rest become extra_expenses
                by_id = Expense.objects.in_bulk([int(i) for i in ids])
                exps = [by_id[int(i)] for i in ids if int(i) in by_id]
                if exps:
                    inst.expense = exps[0]
                    extra_expenses = exps[1:]
                if len(exps) > 1:
                    total = sum((e.amount for e in exps), Decimal(0))
                    if abs(total - amount) > Decimal("0.01"):
                        messages.error(request,
                            f"The {len(exps)} expenses total {total:,.2f} but the "
                            f"payment is {amount:,.2f} — they must match.")
                        return redirect("payment_register")
                    # tell clean() this payment legitimately covers several
                    # vouchers, so the single-expense ceiling guard is skipped
                    inst._covers_multiple = True
            elif kind == "REMITTANCE" and src_id.isdigit():
                inst.remittance_batch = RemittanceBatch.objects.filter(pk=src_id).first()
            elif kind == "REFUND" and src_id.isdigit():
                inst.refund = ExpenseRefund.objects.filter(pk=src_id).first()
            elif kind == "TRANSFER" and src_id.isdigit():
                inst.transfer = FundTransfer.objects.filter(pk=src_id).first()
            # manual / supplier payments need the standalone permission
            if kind in ("MANUAL", "SUPPLIER") and not treas:
                messages.error(request, "Standalone payments require treasurer rights.")
                return redirect("payment_register")
            try:
                inst.full_clean(exclude=["amount"] if amount <= 0 else None)
            except ValidationError as exc:
                msg = "; ".join(m for v in getattr(exc, "message_dict", {}).values()
                                for m in v) or "; ".join(exc.messages)
                messages.error(request, msg)
                return redirect("payment_register")
            inst.save()
            if extra_expenses:
                inst.extra_expenses.set(extra_expenses)
            from .models import PaymentEvent
            PaymentEvent.objects.create(
                payment=inst, event=PaymentEvent.Event.CREATE,
                from_status="", to_status=inst.status,
                on=inst.date_issued or dt.date.today(), user=request.user,
                comment=(inst.note or "")[:200])
            messages.success(request, f"Payment {inst.instrument_number or '(draft)'} recorded"
                + (f" covering {1 + len(extra_expenses)} expenses." if extra_expenses else "."))

        elif action in ("approve", "prepare", "issue", "present", "clear",
                        "void", "cancel", "reject", "reverse", "expire", "stop"):
            inst = get_object_or_404(PaymentInstrument, pk=request.POST.get("pk"))
            verb = {"stop": "CANCEL"}.get(action, action.upper())
            on = None
            if request.POST.get("on"):
                try:
                    on = dt.date.fromisoformat(request.POST["on"])
                except ValueError:
                    on = None
            try:
                apply_event(inst, verb, request.user, on=on,
                            comment=(request.POST.get("comment") or "")[:200])
            except ValidationError as exc:
                messages.error(request, "; ".join(exc.messages))
                return redirect("payment_register")
            when = f" ({on:%d %b %Y})" if on else ""
            messages.success(request, f"Payment marked {inst.get_status_display()}{when}.")

        elif action == "reissue":
            inst = get_object_or_404(PaymentInstrument, pk=request.POST.get("pk"))
            try:
                copy = reissue(inst, request.user,
                               number=(request.POST.get("new_number") or "")[:40],
                               comment=(request.POST.get("comment") or "")[:200])
            except ValidationError as exc:
                messages.error(request, "; ".join(exc.messages))
                return redirect("payment_register")
            messages.success(request,
                f"{inst.get_method_display()} {inst.instrument_number or inst.pk} "
                f"cancelled; replacement draft #{copy.pk} created — set its "
                f"number and issue it when ready.")

        elif action == "delete":
            inst = get_object_or_404(PaymentInstrument, pk=request.POST.get("pk"))
            if inst.is_locked:
                messages.error(request, "A cleared payment cannot be deleted — void it instead.")
            else:
                inst.delete()
                messages.success(request, "Payment removed.")

        elif action == "sync":
            created = self._sync_from_records(request.user)
            messages.success(request, f"Imported {created} payment(s) from existing "
                                      f"cheque expenses and remittances.")
        return redirect("payment_register")

    def _sync_from_records(self, user):
        """Create instrument entries for cheque-method expenses and cheque
        remittance batches that aren't in the register yet."""
        from .models import PaymentInstrument, Expense, RemittanceBatch
        have = set(PaymentInstrument.objects.exclude(instrument_number="")
                   .values_list("instrument_number", flat=True))
        created = 0
        for e in Expense.objects.filter(method=Expense.Method.CHEQUE).exclude(voucher_no=""):
            if e.voucher_no in have:
                continue
            PaymentInstrument.objects.create(
                method="CHEQUE", instrument_number=e.voucher_no[:40],
                payee=(e.payee or e.claimant or e.description)[:160], amount=e.amount,
                date_issued=e.paid_date or e.date, status="OUTSTANDING",
                source_kind="EXPENSE", expense=e, recorded_by=user)
            have.add(e.voucher_no)
            created += 1
        for b in RemittanceBatch.objects.filter(payment__isnull=True).exclude(cheque_no=""):
            if b.cheque_no in have:
                continue
            inst = PaymentInstrument.objects.create(
                method="CHEQUE", instrument_number=b.cheque_no[:40],
                payee="Conference remittance", amount=b.total_amount,
                date_issued=b.cheque_date or b.date, status="OUTSTANDING",
                source_kind="REMITTANCE", remittance_batch=b, recorded_by=user)
            b.payment = inst
            b.save(update_fields=["payment"])
            have.add(b.cheque_no)
            created += 1
        return created



class ReceiptArchiveView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Print-friendly archive of expense receipts for a period, grouped by month,
    plus a one-click ZIP download — so a year's supporting documents can be
    printed or filed together for audit. Staff see every department; a department
    leader sees only the funds they lead (the data is scoped per user)."""
    template_name = "cashbook/receipt_archive.html"

    def test_func(self):
        from core import roles
        u = self.request.user
        return roles.is_staff_role(u) or roles.is_leader(u)

    def handle_no_permission(self):
        if self.request.user.is_authenticated:
            messages.error(self.request, "You don't have permission to view that.")
            return redirect("dashboard")
        return super().handle_no_permission()

    def get(self, request, *args, **kwargs):
        if request.GET.get("download") == "zip":
            return self._zip(request)
        if request.GET.get("export") == "pdf":
            return self._pdf(request)
        return super().get(request, *args, **kwargs)

    def _pdf(self, request):
        from decimal import Decimal
        from django.http import HttpResponse
        from core.models import SiteConfig
        from .services.supporting_pdf import build_receipt_grid_pdf, HAVE_PDF_LIBS
        if not HAVE_PDF_LIBS:
            messages.error(request, "PDF generation isn't available on the server yet. "
                "Ask the administrator to install the reportlab and pypdf packages.")
            return redirect("receipt_archive")
        start, end, qs = self._attachments(request)
        from collections import OrderedDict
        groups = OrderedDict()
        for a in qs:
            groups.setdefault(a.expense.date.strftime("%B %Y"), []).append(a)
        if not groups:
            messages.info(request, "No receipts to print in that period.")
            return redirect("receipt_archive")
        cfg = SiteConfig.get()
        # one expense can carry several attachments — sum each expense once,
        # or the header total silently overstates the bundle
        total = sum({a.expense_id: a.expense.amount for a in qs}.values(), Decimal(0))
        data, stats = build_receipt_grid_pdf(
            groups, church=cfg.church_name or "",
            currency=cfg.currency_symbol or "KES",
            period_label=f"{start:%d %b %Y} – {end:%d %b %Y}",
            filters_label=self._filters_label(request),
            total=total)
        resp = HttpResponse(data, content_type="application/pdf")
        resp["Content-Disposition"] = (
            f'attachment; filename="receipts_{start}_{end}.pdf"')
        return resp

    def _attachments(self, request):
        from core.utils import parse_period
        from .models import ExpenseAttachment
        import datetime as dt
        # this page's own default: "this month" (parse_period's normal
        # fallback) is often empty, which made the page — and the PDF/ZIP
        # downloads that depend on the same range — look broken on a fresh
        # visit. Default to "this year so far" instead, whenever the request
        # has no period/date params of its own; an explicit choice (preset
        # or custom dates) always takes precedence.
        has_explicit_period = any(request.GET.get(k) for k in
                                  ("period", "start", "end", "year", "month"))
        if not has_explicit_period:
            today = dt.date.today()
            start, end = dt.date(today.year, 1, 1), today
        else:
            start, end = parse_period(request)
        qs = (ExpenseAttachment.objects.filter(
                  expense__date__gte=start, expense__date__lte=end)
              .select_related("expense", "expense__department", "expense__vendor")
              .order_by("expense__date"))
        depts = self._scoped_department_ids(request)
        if depts is not None:
            qs = qs.filter(expense__department_id__in=depts)
        qs = self._apply_filters(request, qs)
        return start, end, qs

    # Applied AFTER the leader scoping above, never instead of it: a leader
    # picking a department they don't lead must still see nothing, so this
    # can only ever narrow what the scope already allowed.
    FILTERS = {
        "department": "expense__department_id",
        "category": "expense__category",
        "status": "expense__status",
        "method": "expense__method",
    }

    def _apply_filters(self, request, qs):
        for param, field in self.FILTERS.items():
            raw = (request.GET.get(param) or "").strip()
            if not raw:
                continue
            if field.endswith("_id"):
                # a hand-edited ?department=abc would otherwise raise
                # ValueError when the queryset is evaluated — a 500 on a
                # URL a user can type. Ignore it instead.
                try:
                    raw = int(raw)
                except ValueError:
                    continue
            qs = qs.filter(**{field: raw})
        q = (request.GET.get("q") or "").strip()
        if q:
            from django.db.models import Q
            qs = qs.filter(Q(expense__description__icontains=q)
                           | Q(expense__payee__icontains=q)
                           | Q(expense__voucher_no__icontains=q))
        return qs

    def _filters_label(self, request):
        """Human-readable summary of the filters, printed in the PDF header —
        without it a filtered bundle is indistinguishable from a complete one,
        which is exactly the wrong thing to hand an auditor."""
        from departments.models import Department
        from .models import Expense
        bits = []
        dept = (request.GET.get("department") or "").strip()
        if dept.isdigit():
            d = Department.objects.filter(pk=dept).first()
            if d:
                bits.append(f"Department: {d.name}")
        for param, choices, label in (
                ("category", Expense.Category.choices, "Category"),
                ("status", Expense.Status.choices, "Status"),
                ("method", Expense.Method.choices, "Method")):
            raw = (request.GET.get(param) or "").strip()
            if raw:
                bits.append(f"{label}: {dict(choices).get(raw, raw)}")
        q = (request.GET.get("q") or "").strip()
        if q:
            bits.append(f'Search: "{q}"')
        return "  ·  ".join(bits)

    def _active_filters(self, request):
        """The filters in force, as {param: value} — used to keep them on the
        period-preset links and the PDF/ZIP download links, which would
        otherwise silently drop them and export a different set of receipts
        than the one on screen."""
        keep = {}
        for param in list(self.FILTERS) + ["q"]:
            raw = (request.GET.get(param) or "").strip()
            if raw:
                keep[param] = raw
        return keep

    def _scoped_department_ids(self, request):
        """None for treasurers/auditors (all departments); a list for leaders."""
        from core import roles
        u = request.user
        if roles.is_treasurer(u) or roles.can_enter_data(u) or roles.is_auditor(u):
            return None
        if roles.is_leader(u):
            from departments.models import departments_led_by
            return [d.id for d in departments_led_by(u)]
        return []

    def get_context_data(self, **kwargs):
        from collections import OrderedDict
        ctx = super().get_context_data(**kwargs)
        start, end, qs = self._attachments(self.request)
        groups = OrderedDict()
        n_files = 0
        for a in qs:
            key = a.expense.date.strftime("%B %Y")
            groups.setdefault(key, []).append(a)
            if a.file:
                n_files += 1
        from urllib.parse import urlencode
        from decimal import Decimal
        from .models import Expense
        req = self.request

        # Departments offered in the picker are the ones this user may
        # actually see — a leader is never shown funds they don't lead.
        from departments.models import Department
        dept_qs = Department.objects.order_by("name")
        scoped = self._scoped_department_ids(req)
        if scoped is not None:
            dept_qs = dept_qs.filter(id__in=scoped)

        active = self._active_filters(req)
        # the period is carried on the filter links, and the filters on the
        # period-preset links, so neither ever silently drops the other
        period_qs = urlencode({"start": start.isoformat(), "end": end.isoformat()})
        ctx.update({
            "start": start, "end": end, "groups": groups,
            "count": qs.count(), "file_count": n_files,
            "total_amount": sum({a.expense_id: a.expense.amount
                                 for a in qs}.values(), Decimal(0)),
            "departments": dept_qs,
            "categories": Expense.Category.choices,
            "statuses": Expense.Status.choices,
            "methods": Expense.Method.choices,
            "f": {k: req.GET.get(k, "") for k in list(self.FILTERS) + ["q"]},
            "has_filters": bool(active),
            "filter_qs": urlencode(active),
            "period_qs": period_qs,
            "export_qs": urlencode({**active, "start": start.isoformat(),
                                    "end": end.isoformat()}),
        })
        return ctx

    def _zip(self, request):
        import io
        import zipfile
        from django.http import HttpResponse
        start, end, qs = self._attachments(request)
        buf = io.BytesIO()
        added = 0
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for a in qs:
                if not a.file:
                    continue
                try:
                    d = a.expense.date
                    base = a.file.name.split("/")[-1]
                    arc = f"{d:%Y}/{d:%m}/exp{a.expense_id}_{base}"
                    a.file.open("rb")
                    zf.writestr(arc, a.file.read())
                    a.file.close()
                    added += 1
                except Exception:  # noqa: BLE001
                    from core.utils import log_exception as _lx; _lx("receipt zip")
            # an index of text/links too
            lines = ["Expense receipts index", f"Period: {start} to {end}", ""]
            for a in qs:
                lines.append(f"{a.expense.date}  {a.expense.description}  "
                             f"{a.expense.amount}  "
                             + (a.file.name if a.file else (a.link or a.text or "")))
            zf.writestr("index.txt", "\n".join(lines))
        if not added:
            messages.info(request, "No uploaded receipt files in that period to download.")
            return redirect("receipt_archive")
        resp = HttpResponse(buf.getvalue(), content_type="application/zip")
        resp["Content-Disposition"] = (
            f'attachment; filename="receipts_{start}_{end}.zip"')
        return resp


# Canonical implementation now in cashbook/services/receipts.py; re-exported
# above alongside validate_receipt_upload. (core.views, leaders.views and this
# module's MissingReceiptsView all call it via that name.)


class MissingReceiptsView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Queue of expenses still awaiting a receipt / supporting document. Staff see
    all funds; a leader sees only funds they lead. An expense drops off the queue
    as soon as any document is attached to it."""
    template_name = "cashbook/missing_receipts.html"

    def test_func(self):
        from core import roles
        u = self.request.user
        return roles.is_staff_role(u) or roles.is_leader(u)

    def handle_no_permission(self):
        if self.request.user.is_authenticated:
            messages.error(self.request, "You don't have permission to view that.")
            return redirect("dashboard")
        return super().handle_no_permission()

    def _dept_ids(self):
        from core import roles
        u = self.request.user
        if roles.is_staff_role(u):
            return None
        from departments.models import departments_led_by
        return [d.id for d in departments_led_by(u)]

    def get_context_data(self, **kwargs):
        from core.utils import parse_period
        from django.db.models import Sum
        ctx = super().get_context_data(**kwargs)
        start, end = parse_period(self.request)
        qs = missing_receipts_queryset(start, end, self._dept_ids())
        ctx.update({
            "start": start, "end": end, "expenses": qs,
            "count": qs.count(),
            "total": qs.aggregate(s=Sum("amount"))["s"] or 0,
            "is_leader_view": self._dept_ids() is not None,
        })
        return ctx


# Canonical implementation now in cashbook/services/cheque_words.py — a
# number-to-words renderer is not a view. Re-exported under its original name so
# the cheque-print views (and the payment-instrument test) keep calling it.
from cashbook.services.cheque_words import amount_in_words as _amount_in_words  # noqa: E402


class ChequePrintView(ReadAccessMixin, View):
    """Two quite different things, and it matters which you want.

    **The advice/voucher** (default): a facsimile on plain paper. A record of the
    payment, for the file and for the payee. Prints a whole document.

    **The leaf** (`?mode=leaf`): ink ONLY where the values go, positioned in
    millimetres, to be fed through the printer on a real, pre-printed bank cheque.
    No borders, no labels, no headings — the leaf already has all of that, and
    printing our own on top of it would ruin it.

    Cheque leaves differ between banks by a few millimetres, and a numbered leaf
    spoiled by a bad guess is not free. So the layout is not guessed: it is
    configured, and `?mode=calibrate` prints a measuring sheet onto one spoiled
    leaf so a treasurer can read off where the marks actually land and correct it.
    Once.
    """

    def get(self, request, pk):
        from core.models import SiteConfig
        from .models import PaymentInstrument
        inst = get_object_or_404(PaymentInstrument, pk=pk)
        mode = request.GET.get("mode") or "advice"
        cfg = SiteConfig.get()

        if mode in ("leaf", "calibrate"):
            words = _amount_in_words(inst.amount)
            # the words often need two lines on a narrow leaf; split on a word
            # boundary near the middle rather than letting it run off the edge
            w1, w2 = words, ""
            if len(words) > 46:
                cut = words.rfind(" ", 0, 46)
                if cut > 0:
                    w1, w2 = words[:cut], words[cut + 1:]
            return render(request, "cashbook/cheque_leaf.html", {
                "c": inst, "cfg": cfg, "calibrate": (mode == "calibrate"),
                "words1": w1, "words2": w2,
                "date_str": (inst.date_issued or dt.date.today()).strftime("%d %m %Y"),
            })

        return render(request, "cashbook/payment_print.html", {
            "c": inst, "amount_words": _amount_in_words(inst.amount)})


class ChequeOutstandingView(ReadAccessMixin, View):
    """Outstanding (unpresented) payments AS AT a date, per method, with
    exports — judged on issue/cleared DATES via outstanding_asof, so the
    report is correct for any historical date (the reconciliation view of
    the world), not just today."""
    def get(self, request):
        import datetime as dtt
        from .models import PaymentInstrument
        from reports.exports import csv_response, xlsx_response
        as_of = dtt.date.today()
        if request.GET.get("as_of"):
            try:
                as_of = dtt.date.fromisoformat(request.GET["as_of"])
            except ValueError:
                pass
        qs = (PaymentInstrument.outstanding_asof(as_of)
              .select_related("expense__department", "remittance_batch")
              .order_by("date_issued"))
        method = request.GET.get("method", "")
        if method:
            qs = qs.filter(method=method)
        header = ["Method", "Number", "Payee", "Source", "Fund(s)", "Amount",
                  "Date issued", "Days outstanding"]
        rows = [[c.get_method_display(), c.instrument_number, c.payee,
                 c.source_label, c.fund_names, float(c.amount), c.date_issued,
                 (as_of - c.date_issued).days if c.date_issued else ""]
                for c in qs]
        export = request.GET.get("export")
        if export == "csv":
            return csv_response(f"outstanding_payments_{as_of}", header, rows)
        if export == "xlsx":
            return xlsx_response(f"outstanding_payments_{as_of}", header, rows,
                                 title=f"Outstanding payments as at {as_of:%d %b %Y}")
        total = sum((c.amount for c in qs), __import__("decimal").Decimal(0))
        return render(request, "cashbook/payment_outstanding.html", {
            "rows": qs, "total": total, "method": method, "as_of": as_of,
            "methods": PaymentInstrument.Method.choices})


class PaymentAnalysisView(ReadAccessMixin, View):
    """Payments grouped by fund / bank account / method / source type over a
    period, plus the clearance-performance figures (cleared count, average
    and slowest days-to-clear) and cancelled/voided listing — the reporting
    side of the payment lifecycle."""

    def get(self, request):
        import datetime as dtt
        from decimal import Decimal
        from core.utils import parse_period
        from .models import PaymentInstrument
        from reports.exports import csv_response, xlsx_response
        start, end = parse_period(request)
        group = request.GET.get("group") or "fund"
        qs = (PaymentInstrument.objects.filter(
                date_issued__gte=start, date_issued__lte=end)
              .select_related("expense__department", "bank_account"))

        def _key(p):
            if group == "bank":
                return str(p.bank_account) if p.bank_account_id else "(no bank account)"
            if group == "method":
                return p.get_method_display()
            if group == "source":
                return p.get_source_kind_display()
            return p.fund_names or "(no fund)"

        agg = {}
        for pmt in qs:
            row = agg.setdefault(_key(pmt), {
                "count": 0, "total": Decimal(0),
                "cleared": 0, "cleared_total": Decimal(0),
                "outstanding": 0, "outstanding_total": Decimal(0),
                "cancelled": 0, "days": []})
            row["count"] += 1
            row["total"] += pmt.amount
            if pmt.status == PaymentInstrument.Status.CLEARED:
                row["cleared"] += 1
                row["cleared_total"] += pmt.amount
                if pmt.clearance_days is not None:
                    row["days"].append(pmt.clearance_days)
            elif pmt.is_outstanding:
                row["outstanding"] += 1
                row["outstanding_total"] += pmt.amount
            elif pmt.status in PaymentInstrument.TERMINAL_STATES:
                row["cancelled"] += 1
        rows = []
        for key in sorted(agg):
            r = agg[key]
            rows.append({
                "group": key, **r,
                "avg_days": (round(sum(r["days"]) / len(r["days"]), 1)
                             if r["days"] else None),
                "max_days": max(r["days"]) if r["days"] else None})

        export = request.GET.get("export")
        if export in ("csv", "xlsx"):
            from core.models import SiteConfig
            header = [group.title(), "Payments", "Total", "Cleared",
                      "Cleared total", "Outstanding", "Outstanding total",
                      "Cancelled", "Avg days to clear", "Slowest (days)"]
            data = [[r["group"], r["count"], float(r["total"]), r["cleared"],
                     float(r["cleared_total"]), r["outstanding"],
                     float(r["outstanding_total"]), r["cancelled"],
                     r["avg_days"] if r["avg_days"] is not None else "",
                     r["max_days"] if r["max_days"] is not None else ""]
                    for r in rows]
            fn = f"payments_by_{group}_{start}_{end}"
            if export == "xlsx":
                return xlsx_response(fn + ".xlsx", header, data,
                    title=f"Payments by {group} — {start:%d %b %Y} to {end:%d %b %Y}",
                    church=SiteConfig.get().church_name)
            return csv_response(fn + ".csv", header, data)
        return render(request, "cashbook/payment_analysis.html", {
            "rows": rows, "group": group, "start": start, "end": end,
            "f": request.GET})


class ExpenseBatchCreate(DataEntryRequiredMixin, TemplateView):
    """Several expenses that share a date, fund, claimant and method.

    A treasurer settling a stack of receipts from one person, on one day, from
    one fund, was re-typing those four fields for every line — which is the slow
    part, and the part where the fund gets mistyped on line seven and nobody
    notices until the fund balances are wrong.

    So the shared facts are entered once at the top, and each line carries only
    what actually differs: the narration, the amount, and the transaction charge
    where there was one. The work itself goes through
    `services.expenses.record_batch`, the same function the single form and the
    spreadsheet import use, so a batch-entered expense is indistinguishable from
    one entered any other way.
    """
    template_name = "cashbook/expense_batch.html"

    def get_context_data(self, **kwargs):
        from core.models import SiteConfig
        from vendors.models import Vendor
        ctx = super().get_context_data(**kwargs)
        ctx.update({
            "categories": Expense.Category.choices,
            "methods": Expense.Method.choices,
            "expenditure_types": Expense.ExpenditureType.choices,
            "vendors": Vendor.objects.exclude(
                status=Vendor.Status.ARCHIVED).order_by("name"),
            "today": dt.date.today(),
            "auto_approve": not SiteConfig.get().require_expense_approval,
        })
        return ctx

    def post(self, request, *args, **kwargs):
        from core.models import SiteConfig
        from departments.models import Department
        from vendors.models import Vendor
        from .services import expenses as expense_svc

        post = request.POST
        fund = Department.objects.filter(pk=post.get("department") or 0).first()
        when = _parse_date(post.get("date"))
        if fund is None or when is None:
            messages.error(request, "Choose the fund and the date first.")
            return self.render_to_response(self.get_context_data(**kwargs))

        vendor = None
        if (post.get("vendor") or "").isdigit():
            vendor = Vendor.objects.filter(pk=int(post["vendor"])).first()

        header = {
            "date": when, "department": fund,
            "claimant": post.get("claimant", ""), "payee": post.get("payee", ""),
            "vendor": vendor,
            "method": post.get("method") or Expense.Method.CASH,
            "category": post.get("category") or Expense.Category.OTHER,
            "expenditure_type": post.get("expenditure_type") or None,
            "voucher_no": post.get("voucher_no", ""),
            "paid_from_petty_cash": bool(post.get("paid_from_petty_cash")),
        }

        # Lines arrive as parallel lists, which is what a table of inputs
        # produces; zip rather than index arithmetic so a ragged post cannot
        # silently pair the wrong amount with the wrong narration.
        descriptions = post.getlist("line_description")
        amounts = post.getlist("line_amount")
        categories = post.getlist("line_category")
        charges = post.getlist("line_charge")
        lines = []
        for i, description in enumerate(descriptions):
            lines.append({
                "description": description,
                "amount": amounts[i] if i < len(amounts) else None,
                "category": categories[i] if i < len(categories) else None,
                "charge": charges[i] if i < len(charges) else None,
            })

        try:
            created = expense_svc.record_batch(
                header=header, lines=lines, user=request.user,
                auto_approve=not SiteConfig.get().require_expense_approval,
                shared_charge=post.get("shared_charge"))
        except Exception as exc:
            from core.utils import log_exception as _lx
            _lx("cashbook/views.py")
            messages.error(request, f"Nothing was saved — {exc}")
            return self.render_to_response(self.get_context_data(**kwargs))

        if not created:
            messages.error(request, "No lines had both a description and an amount.")
            return self.render_to_response(self.get_context_data(**kwargs))

        messages.success(
            request,
            f"{len(created)} expense(s) recorded against {fund.name} "
            f"for {when:%d %b %Y}.")
        return redirect("expense_list")
