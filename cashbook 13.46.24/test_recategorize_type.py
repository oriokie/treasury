"""Recategorize round-trip also edits expenditure type (capital/recurrent) (#1)."""
import io
import datetime as dt

from django.test import TestCase, Client
from django.contrib.auth.models import User, Group
from django.core.files.uploadedfile import SimpleUploadedFile

from departments.models import Department
from cashbook.models import Expense


class RecategorizeTypeTests(TestCase):
    def setUp(self):
        u = User.objects.create_user("rc", password="x", is_superuser=True)
        u.groups.add(Group.objects.get_or_create(name="Treasurer")[0])
        self.u = u
        self.c = Client(); self.c.force_login(u)
        self.d = Department.objects.create(name="LCB", fund_type="LOCAL",
                                           category="OFFERING", show_in_expenses=True)
        self.e = Expense.objects.create(date=dt.date(2026, 6, 1), department=self.d,
            description="Chairs", amount=5000, category="OTHER",
            expenditure_type="RECURRENT", status="PAID", recorded_by=u)

    def test_download_has_type_column(self):
        import openpyxl
        r = self.c.get("/expenses/recategorize/?download=1")
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        hdr = [c.value for c in wb["Expenses"][1]]
        self.assertIn("New type (capital/recurrent)", hdr)

    def test_reimport_updates_type_and_category(self):
        import openpyxl
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Expenses"
        ws.append(["ID", "Date", "Description", "Fund", "Amount", "Current category",
                   "New category (edit this)", "Current type", "New type (capital/recurrent)"])
        ws.append([self.e.id, "2026-06-01", "Chairs", self.d.name, 5000,
                   "Other", "Materials", "Recurrent", "Capital"])
        buf = io.BytesIO(); wb.save(buf)
        up = SimpleUploadedFile("e.xlsx", buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.c.post("/expenses/recategorize/", {"file": up})
        self.e.refresh_from_db()
        self.assertEqual(self.e.category, "MATERIALS")
        self.assertEqual(self.e.expenditure_type, "CAPITAL")

    def test_type_only_change(self):
        import openpyxl
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Expenses"
        ws.append(["ID", "New category (edit this)", "New type (capital/recurrent)"])
        ws.append([self.e.id, "", "Capital"])
        buf = io.BytesIO(); wb.save(buf)
        up = SimpleUploadedFile("e.xlsx", buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.c.post("/expenses/recategorize/", {"file": up})
        self.e.refresh_from_db()
        self.assertEqual(self.e.expenditure_type, "CAPITAL")
        self.assertEqual(self.e.category, "OTHER")   # untouched
