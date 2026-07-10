"""v1.90 batch: richer backup, assistant context, remittance notification,
recon advance auto-populate + section removal."""
import datetime as dt, io
from decimal import Decimal
from django.test import TestCase, Client
from django.contrib.auth.models import User, Group
from django.core.cache import cache
from openpyxl import load_workbook
from departments.models import Department
from giving.models import Transaction
from cashbook.models import Expense, StaffAdvance
from statements.models import BankReconciliation, ReconciliationItem
from ledger.services.posting import ensure_chart


def _tr():
    u = User.objects.create_user("tr_190", password="x", is_superuser=True)
    u.groups.add(Group.objects.get_or_create(name="Treasurer")[0])
    return u


class BackupExportTests(TestCase):
    def test_backup_has_all_key_sheets(self):
        from core.services.backup import full_excel_export_response
        resp = full_excel_export_response()
        data = resp.content
        wb = load_workbook(io.BytesIO(data))
        for sheet in ["Fund Balances", "Trust Funds", "Cash Book", "Payments",
                      "Staff Advances", "Remittances", "Fund Transfers",
                      "Pledges", "Fixed Assets", "Petty Cash Top-ups",
                      "Transactions", "Expenses"]:
            self.assertIn(sheet, wb.sheetnames)


class AssistantContextTests(TestCase):
    def setUp(self):
        u = User.objects.create_user("m_ctx", password="x", is_superuser=True)
        d = Department.objects.create(name="Tithe", fund_type="LOCAL",
            category="OFFERING", show_in_expenses=True)
        Transaction.objects.create(date=dt.date.today(), amount=Decimal("1000"),
            department=d, direction="CREDIT", confirmed=True, channel="BANK",
            allocation_status="MANUAL")
        Expense.objects.create(date=dt.date.today(), department=d,
            description="x", amount=Decimal("50"), category="UTILITIES",
            status="PAID", recorded_by=u, approved_by=u)
        from members.models import Member
        Member.objects.create(name="Test Member", active=True)
        BankReconciliation.objects.create(statement_date=dt.date.today(),
            bank_balance=Decimal("0"), created_by=u)

    def test_context_covers_key_dimensions(self):
        from core.services.assistant import _data_context
        ctx = _data_context()
        for token in ["Collections this month", "Channel", "Tithe received YTD",
                      "remittance compliance", "bank reconciliation",
                      "Active members", "expense categories"]:
            self.assertIn(token, ctx)


class RemittanceNotificationTests(TestCase):
    def setUp(self):
        cache.clear()
        self.tr = _tr()
        self.trust = Department.objects.create(name="ENF N", fund_type="TRUST",
            category="TRUST", is_trust=True)

    def _to_remit(self):
        from reports.services import balances
        cache.clear()
        row = [r for r in balances.trust_summary()
               if r["department"].id == self.trust.id][0]
        return row["to_remit"], row["unreceipted"]

    def test_unreceipted_does_not_show(self):
        Transaction.objects.create(date=dt.date(2026, 6, 1), amount=Decimal("5000"),
            department=self.trust, direction="CREDIT", confirmed=True,
            channel="BANK", allocation_status="MANUAL", manual_receipt=False)
        to_remit, unrec = self._to_remit()
        self.assertEqual(to_remit, 0)
        self.assertEqual(unrec, Decimal("5000"))

    def test_clears_after_remit(self):
        Transaction.objects.create(date=dt.date(2026, 6, 2), amount=Decimal("8000"),
            department=self.trust, direction="CREDIT", confirmed=True,
            channel="ENVELOPE", allocation_status="MANUAL", manual_receipt=True)
        self.assertEqual(self._to_remit()[0], Decimal("8000"))
        Expense.objects.create(date=dt.date(2026, 6, 30), department=self.trust,
            description="Remit", amount=Decimal("8000"), category="REMITTANCE",
            status="PAID", recorded_by=self.tr, approved_by=self.tr)
        self.assertEqual(self._to_remit()[0], 0)


class ReconAdvanceTests(TestCase):
    def setUp(self):
        ensure_chart()
        self.tr = _tr()
        self.c = Client(); self.c.force_login(self.tr)
        self.d = Department.objects.create(name="LCB R", fund_type="LOCAL",
            category="OFFERING", show_in_expenses=True)

    def test_advance_auto_populated_at_creation(self):
        StaffAdvance.objects.create(staff_name="Bob", department=self.d,
            amount=Decimal("7000"), date_issued=dt.date(2026, 6, 1), purpose="x",
            method="BANK", from_petty_cash=False, issued_by=self.tr)
        self.c.post("/reconciliations/new/",
            {"statement_date": "2026-06-30", "bank_balance": "0"})
        rec = BankReconciliation.objects.order_by("-id").first()
        item = ReconciliationItem.objects.filter(reconciliation=rec,
            description__icontains="Staff advances").first()
        self.assertIsNotNone(item)
        self.assertEqual(item.amount, Decimal("7000"))
        self.assertEqual(item.effect, "ADD")

    def test_info_cards_removed(self):
        rec = BankReconciliation.objects.create(statement_date=dt.date(2026, 6, 30),
            bank_balance=Decimal("0"), created_by=self.tr)
        body = self.c.get(f"/reconciliations/{rec.id}/").content.decode()
        self.assertNotIn("Staff advances (bank-funded)", body)
        self.assertNotIn("Unpresented cheques (from the register)", body)
        self.assertIn("added to this statement", body)  # the auto note
